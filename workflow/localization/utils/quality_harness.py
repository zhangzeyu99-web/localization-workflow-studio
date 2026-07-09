"""Quality regression harness for localization outputs.

The harness is intentionally independent from one project workbook. It can run
small string fixtures and scan Excel language tables so quality regressions are
caught before final delivery.

Implementation detail: row-level rule checks live in
utils/quality_harness_rules.py and term-base/sheet detection lives in
utils/quality_harness_terms.py. This module re-exports those symbols so
existing `from utils.quality_harness import ...` imports keep working.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

from utils.language_config import normalize_language_code
from utils.quality_harness_rules import (  # noqa: F401  (re-exported)
    BROKEN_BULLET_PATTERN,
    CJK_PATTERN,
    COLOR_TAG_PATTERN,
    DEFAULT_HARD_ISSUES,
    FULLWIDTH_PUNCTUATION_PATTERN,
    HASH_CODE_PATTERN,
    HASHED_INTERNAL_CODE_PATTERN,
    HTML_ENTITY_PATTERN,
    HarnessResult,
    INTERNAL_TOKEN_PATTERN,
    LETTER_PLACEHOLDER_COMPACTION_PATTERN,
    MIN_NUMBERED_TERM_GROUP_SIZE,
    NUMBERED_SOURCE_PATTERN,
    NUMBERED_TARGET_PATTERN,
    ORPHAN_LEADING_CLITIC_PATTERN,
    PLACEHOLDER_WORD_GLUE_PATTERN,
    SANDWICHED_QUESTION_PATTERN,
    SOURCE_SEPARATOR_PATTERN,
    WORD_START_PATTERN,
    _append_numbered_term_consistency_issues,
    _check_person_name_terms,
    _check_surface_regressions,
    _check_terms,
    _check_ui_length,
    _choose_numbered_canonical,
    _coerce_check_result,
    _contains_expected_person_name,
    _format_numbered_target_signature,
    _has_leading_lowercase,
    _has_punctuation_corruption,
    _issue,
    _looks_like_allowed_runtime_code,
    _matched_person_name_terms,
    _numbered_item_is_readable,
    _numbered_signature_from_term,
    _parse_numbered_source,
    _parse_numbered_target,
    _primary_term,
    _starts_with_runtime_payload,
    _visible_start,
    check_row,
)
from utils.quality_harness_terms import (  # noqa: F401  (re-exported)
    AUTO_SOFT_SOURCE_TERMS,
    AUTO_SOFT_TARGET_WORDS,
    GENERIC_ROLE_TARGETS,
    GLOSSARY_SHEET_NAMES,
    LANGUAGE_TARGET_HEADERS,
    LANGUAGE_VARIANT_HEADERS,
    LEGAL_TERM_CHECK_MIN_LENGTH,
    LEGAL_TERM_CHECK_SKIP_MARKERS,
    OUTPUT_DIR_NAME_HINTS,
    PERSON_NAME_CATEGORY_MARKERS,
    SOFT_TERM_CATEGORY_MARKERS,
    SUPPORT_SHEET_NAME_KEYWORDS,
    TERM_BASE_FILENAME_KEYWORDS,
    _add_term_lookup_entry,
    _all_target_header_candidates,
    _all_variant_header_candidates,
    _clean_term_cell,
    _collect_term_context,
    _collect_terms_from_json,
    _collect_terms_from_workbook,
    _detect_columns,
    _discover_term_base_paths,
    _fallback_index,
    _find_header,
    _has_explicit_language_header,
    _is_auto_soft_term,
    _is_generic_role_target,
    _is_glossary_sheet,
    _is_person_name_category,
    _is_soft_term_category,
    _is_support_sheet,
    _iter_term_base_paths,
    _looks_like_output_dir,
    _resolve_term_base_paths,
    _should_skip_term_checks,
    _split_term_variants,
    _target_header_candidates,
    _truthy_cell,
    _variant_header_candidates,
)


def load_fixture(path: str | Path) -> dict:
    return json.loads(Path(path).read_text(encoding='utf-8'))


def run_fixture(fixture: dict, lang: str = 'en') -> HarnessResult:
    cases = fixture.get('cases', [])
    result = HarnessResult(passed=True, total_cases=len(cases))

    for case in cases:
        row_id = case.get('id', '')
        case_lang = case.get('lang', lang)
        issues = check_row(
            row_id=row_id,
            source=case.get('source', ''),
            translation=case.get('translation', ''),
            lang=case_lang,
        )
        actual = sorted({issue.check_type for issue in issues})
        expected = sorted(case.get('expected_issues', []))

        for issue_type in actual:
            result.issue_counts[issue_type] += 1

        if actual != expected:
            result.passed = False
            result.failures.append({
                'id': row_id,
                'source': case.get('source', ''),
                'translation': case.get('translation', ''),
                'expected_issues': expected,
                'actual_issues': actual,
            })

    return result


def scan_workbook(
    path: str | Path,
    lang: str = 'en',
    fail_on: Iterable[str] | None = None,
    term_base: str | Path | Sequence[str | Path] | None = None,
    auto_discover_terms: bool = True,
) -> HarnessResult:
    """Scan a workbook language table.

    The scanner expects either headers containing ID/CN/EN-like names or a
    simple first-three-column layout: ID, source, target.
    """
    lang = normalize_language_code(lang)
    fail_set = set(fail_on or DEFAULT_HARD_ISSUES)
    result = HarnessResult(passed=True)
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, read_only=False, data_only=False)

    try:
        term_sources = _resolve_term_base_paths(workbook_path, term_base, auto_discover_terms)
        term_context = _collect_term_context(wb, term_sources, lang=lang)
        strong_term_lookup = term_context['strong']
        soft_term_lookup = term_context['soft']
        person_name_terms = term_context['person_names']
        for ws in wb.worksheets:
            if _is_glossary_sheet(ws) or _is_support_sheet(ws):
                continue
            id_col, src_col, tgt_col = _detect_columns(ws, lang=lang)
            if src_col is None or tgt_col is None:
                continue
            max_col = max(c for c in (id_col, src_col, tgt_col) if c is not None) + 1
            numbered_rows: list[dict] = []
            for row_index, row in enumerate(
                ws.iter_rows(min_row=2, max_col=max_col, values_only=True),
                start=2,
            ):
                row_id = row[id_col] if id_col is not None else row_index
                source = row[src_col]
                target = row[tgt_col]
                if not isinstance(source, str) or not isinstance(target, str):
                    continue

                result.rows_scanned += 1
                numbered_rows.append({
                    'file': str(workbook_path),
                    'sheet': ws.title,
                    'row': row_index,
                    'id': row_id,
                    'source': source,
                    'translation': target,
                })
                row_issues = check_row(row_id, source, target, lang=lang)
                row_issues.extend(_check_ui_length(row_id, source, target, lang=lang))
                if not _should_skip_term_checks(source):
                    row_issues.extend(_check_terms(row_id, source, target, strong_term_lookup, lang=lang))
                    row_issues.extend(_check_terms(row_id, source, target, soft_term_lookup, soft=True, lang=lang))
                    row_issues.extend(_check_person_name_terms(row_id, source, target, person_name_terms))
                for issue in row_issues:
                    result.issue_counts[issue.check_type] += 1
                    if issue.check_type in fail_set:
                        result.passed = False
                        result.issues.append({
                            'file': str(workbook_path),
                            'sheet': ws.title,
                            'row': row_index,
                            'id': row_id,
                            'check_type': issue.check_type,
                            'severity': issue.severity,
                            'message': issue.message,
                            'source': source,
                            'translation': target,
                            'auto_fix': issue.auto_fix,
                        })

            _append_numbered_term_consistency_issues(numbered_rows, result, fail_set, strong_term_lookup, lang=lang)

        if result.rows_scanned == 0:
            result.passed = False
            result.issue_counts['workbook_scan_empty'] += 1
            result.issues.append({
                'file': str(workbook_path),
                'sheet': '',
                'row': 0,
                'id': '',
                'check_type': 'workbook_scan_empty',
                'severity': 'error',
                'message': 'No workbook rows were scanned; check sheet names and headers before treating QA as passed',
                'source': '',
                'translation': '',
                'auto_fix': '',
            })
    finally:
        wb.close()

    return result


def merge_results(results: list[HarnessResult]) -> HarnessResult:
    merged = HarnessResult(passed=all(r.passed for r in results))
    for result in results:
        merged.total_cases += result.total_cases
        merged.rows_scanned += result.rows_scanned
        merged.issue_counts.update(result.issue_counts)
        merged.issues.extend(result.issues)
        merged.failures.extend(result.failures)
    return merged
