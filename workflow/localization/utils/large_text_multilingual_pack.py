"""Fast workbook extraction for large multilingual localization packs."""
from __future__ import annotations

import json
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from utils.language_config import SOURCE_HEADERS, normalize_language_code, target_header_candidates
from utils.large_text_multilingual_gate import protected_tokens
from utils.source_reference import (
    EnglishReferenceStatus,
    classify_english_reference,
    normalize_source_mode,
)


TERM_CATEGORY_HEADERS = {"分类", "类别", "术语分类", "category", "type"}
STRICT_TERM_CATEGORIES = {
    "主角",
    "角色名",
    "人物名",
    "英雄名",
    "怪物名",
    "boss",
    "npc",
}


@dataclass(frozen=True)
class PackArtifacts:
    items_jsonl: Path
    source_rows_jsonl: Path
    prepare_stats: Path
    source_rows: int
    unique_items: int
    estimated_target_cells: int
    elapsed_seconds: float


def stable_row_key(source_file: str, sheet: str, row: int, row_id: object) -> str:
    return f"{source_file}::{sheet}::{row_id}::{row}"


def _write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n",
        encoding="utf-8",
    )


def _normalize_headers(values: tuple[Any, ...]) -> list[str]:
    return [str(value or "").strip().lower() for value in values]


def _reset_stale_dimensions(sheet: Any) -> None:
    if sheet.max_row == 1 and sheet.max_column == 1 and hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()


def _find_header(headers: list[str], candidates: set[str]) -> int | None:
    return next((index for index, header in enumerate(headers) if header in candidates), None)


def _language_columns(
    headers: list[str],
    target_langs: list[str],
    *,
    excluded_columns: set[int] | None = None,
) -> dict[str, int]:
    excluded_columns = excluded_columns or set()
    columns: dict[str, int] = {}
    for raw_lang in target_langs:
        lang = normalize_language_code(raw_lang)
        candidates = target_header_candidates(lang)
        column = next(
            (index for index, header in enumerate(headers) if index not in excluded_columns and header in candidates),
            None,
        )
        if column is None:
            raise ValueError(f"missing target language column: {raw_lang}")
        columns[str(raw_lang).upper()] = column
    return columns


def _is_strict_term_category(value: object) -> bool:
    category = str(value or "").strip().casefold()
    return category in {item.casefold() for item in STRICT_TERM_CATEGORIES}


def _load_terms(path: Path | None, target_langs: list[str]) -> list[dict[str, Any]]:
    if path is None:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            _reset_stale_dimensions(sheet)
            iterator = sheet.iter_rows(values_only=True)
            first = next(iterator, ())
            headers = _normalize_headers(first)
            id_col = _find_header(headers, {"id", "key", "索引id"})
            source_col = _find_header(headers, {header.lower() for header in SOURCE_HEADERS})
            category_col = _find_header(headers, TERM_CATEGORY_HEADERS)
            if source_col is None:
                continue
            try:
                language_columns = _language_columns(
                    headers,
                    target_langs,
                    excluded_columns={id_col} if id_col is not None else set(),
                )
            except ValueError:
                continue
            terms: dict[str, dict[str, Any]] = {}
            for values in iterator:
                source = values[source_col] if source_col < len(values) else None
                if source in (None, ""):
                    continue
                translations = {
                    lang: str((values[column] if column < len(values) else "") or "").strip()
                    for lang, column in language_columns.items()
                }
                if all(translations.values()):
                    category = (
                        str((values[category_col] if category_col < len(values) else "") or "").strip()
                        if category_col is not None
                        else ""
                    )
                    strict = _is_strict_term_category(category)
                    terms[str(source).strip()] = {
                        "translations": translations,
                        "category": category,
                        "required": strict,
                        "strict": strict,
                    }
            return [
                {"source": source, **metadata}
                for source, metadata in sorted(terms.items(), key=lambda item: len(item[0]), reverse=True)
            ]
    finally:
        workbook.close()
    return []


def _load_history(history_dirs: list[Path], target_langs: list[str]) -> dict[str, dict[str, str]]:
    memory: dict[str, dict[str, str]] = {}
    conflicts: set[str] = set()
    for directory in history_dirs:
        for path in sorted(directory.glob("*.xlsx")):
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    _reset_stale_dimensions(sheet)
                    iterator = sheet.iter_rows(values_only=True)
                    first = next(iterator, ())
                    headers = _normalize_headers(first)
                    id_col = _find_header(headers, {"id", "key", "索引id"})
                    source_col = _find_header(headers, {header.lower() for header in SOURCE_HEADERS})
                    if source_col is None:
                        continue
                    try:
                        language_columns = _language_columns(
                            headers,
                            target_langs,
                            excluded_columns={id_col} if id_col is not None else set(),
                        )
                    except ValueError:
                        continue
                    for values in iterator:
                        source = values[source_col] if source_col < len(values) else None
                        if source in (None, ""):
                            continue
                        translations = {
                            lang: str((values[column] if column < len(values) else "") or "").strip()
                            for lang, column in language_columns.items()
                        }
                        if not all(translations.values()):
                            continue
                        source_text = str(source)
                        if source_text in memory and memory[source_text] != translations:
                            conflicts.add(source_text)
                        else:
                            memory[source_text] = translations
            finally:
                workbook.close()
    for source in conflicts:
        memory.pop(source, None)
    return memory


def _term_hits(text: str, terms: list[dict[str, Any]], limit: int = 12) -> list[dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    occupied: list[tuple[int, int]] = []
    for term in terms:
        source = term["source"]
        if len(source) < 2 and source != text:
            continue
        start = text.find(source)
        if start < 0:
            continue
        end = start + len(source)
        if any(start < used_end and end > used_start for used_start, used_end in occupied):
            continue
        hits.append(term)
        occupied.append((start, end))
        if len(hits) >= limit:
            break
    return hits


def prepare_pack(
    *,
    inputs: list[Path],
    term_base: Path | None,
    history_dirs: list[Path],
    target_langs: list[str],
    work_dir: Path,
    source_mode: str = "cn",
) -> PackArtifacts:
    started = time.perf_counter()
    target_langs = [str(lang).upper() for lang in target_langs]
    source_mode = normalize_source_mode(source_mode)
    if source_mode != "cn" and "EN" in target_langs:
        raise ValueError("source_mode cn+en/en requires EN to be a reference, not a target language")
    terms = _load_terms(term_base, target_langs)
    history = _load_history(history_dirs, target_langs)
    exact_terms = {term["source"]: term["translations"] for term in terms}
    rows: list[dict[str, Any]] = []
    seed_memory: dict[str, dict[str, str]] = {}
    english_column_name = ""
    english_usable = 0
    english_empty = 0
    english_chinese = 0

    for input_path in inputs:
        workbook = load_workbook(input_path, read_only=True, data_only=False)
        try:
            for sheet in workbook.worksheets:
                _reset_stale_dimensions(sheet)
                iterator = sheet.iter_rows(values_only=True)
                first = next(iterator, ())
                headers = _normalize_headers(first)
                id_col = _find_header(headers, {"id", "key", "索引id"})
                source_col = _find_header(headers, {header.lower() for header in SOURCE_HEADERS})
                if source_col is None:
                    continue
                english_col = _find_header(headers, target_header_candidates("en"))
                if source_mode != "cn" and english_col is None:
                    raise ValueError(
                        f"source_mode {source_mode} requires an English reference column: "
                        f"{input_path.name}:{sheet.title}"
                    )
                if english_col is not None:
                    english_column_name = str(first[english_col] or "")
                language_columns = _language_columns(
                    headers,
                    target_langs,
                    excluded_columns={id_col} if id_col is not None else set(),
                )
                context = "ui" if "ui" in f"{input_path.stem} {sheet.title}".lower() else "language"
                for row_index, values in enumerate(iterator, 2):
                    source = values[source_col] if source_col < len(values) else None
                    row_id = values[id_col] if id_col is not None and id_col < len(values) else row_index
                    if source in (None, "") and row_id in (None, ""):
                        continue
                    if source in (None, ""):
                        raise ValueError(f"blank source text: {input_path.name}:{sheet.title}!R{row_index}")
                    for lang, column in language_columns.items():
                        target = values[column] if column < len(values) else None
                        if target not in (None, ""):
                            raise ValueError(
                                f"target column is not empty: {input_path.name}:{sheet.title}!R{row_index}:{lang}"
                            )
                    source_text = str(source)
                    if source_mode == "cn":
                        reference_en = ""
                        reference_status = "not_requested"
                    else:
                        reference_raw = values[english_col] if english_col is not None and english_col < len(values) else None
                        reference = classify_english_reference(reference_raw)
                        reference_en = reference.text
                        reference_status = reference.status
                        if reference_status == "usable":
                            english_usable += 1
                        elif reference_status == "missing":
                            english_empty += 1
                        else:
                            english_chinese += 1
                        if source_mode == "en" and reference_status != "usable":
                            raise ValueError(
                                "source_mode en requires complete usable English for every row: "
                                f"{input_path.name}:{sheet.title}!R{row_index} status={reference_status}"
                            )
                    translation_source = reference_en if source_mode == "en" else source_text
                    if source_text in history:
                        seed_memory[source_text] = history[source_text]
                        seed_origin = "history"
                    elif source_text in exact_terms:
                        seed_memory[source_text] = exact_terms[source_text]
                        seed_origin = "glossary_exact"
                    else:
                        seed_origin = "api"
                    row = {
                        "key": stable_row_key(input_path.name, sheet.title, row_index, row_id),
                        "id": str(row_id),
                        "source_file": input_path.name,
                        "sheet": sheet.title,
                        "row": row_index,
                        "context": context,
                        "cn": source_text,
                        "translation_source": translation_source,
                        "source_mode": source_mode,
                        "reference_en": reference_en,
                        "reference_en_status": reference_status,
                        "tokens": protected_tokens({"cn": source_text}),
                        "term_hits": _term_hits(source_text, terms),
                        "seed_origin": seed_origin,
                        "translations": {},
                    }
                    rows.append(row)
        finally:
            workbook.close()

    work_dir.mkdir(parents=True, exist_ok=True)
    items_path = work_dir / "items.jsonl"
    source_rows_path = work_dir / "source_rows.jsonl"
    stats_path = work_dir / "prepare_stats.json"
    _write_jsonl(items_path, rows)
    _write_jsonl(
        source_rows_path,
        [
            {
                key: row[key]
                for key in (
                    "key",
                    "id",
                    "source_file",
                    "sheet",
                    "row",
                    "cn",
                    "translation_source",
                    "source_mode",
                    "reference_en",
                    "reference_en_status",
                )
            }
            for row in rows
        ],
    )
    (work_dir / "seed_memory.json").write_text(
        json.dumps(seed_memory, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    elapsed = time.perf_counter() - started
    result = PackArtifacts(
        items_jsonl=items_path,
        source_rows_jsonl=source_rows_path,
        prepare_stats=stats_path,
        source_rows=len(rows),
        unique_items=len({
            (row["cn"], row["translation_source"], row["source_mode"])
            for row in rows
        }),
        estimated_target_cells=len(rows) * len(target_langs),
        elapsed_seconds=elapsed,
    )
    stats_path.write_text(
        json.dumps(
            {
                **asdict(result),
                "items_jsonl": str(items_path),
                "source_rows_jsonl": str(source_rows_path),
                "prepare_stats": str(stats_path),
                "source_mode": source_mode,
                "english_reference_status": (
                    EnglishReferenceStatus(
                        column=english_column_name,
                        total_rows=english_usable + english_empty + english_chinese,
                        usable_rows=english_usable,
                        empty_rows=english_empty,
                        chinese_rows=english_chinese,
                    ).to_dict()
                    if source_mode != "cn"
                    else None
                ),
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )
    return result
