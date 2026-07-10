"""Term-base collection and workbook sheet/column detection for the quality harness.

Implementation module split out of utils/quality_harness.py. Import these
symbols through utils.quality_harness to keep the public surface stable.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from utils.language_config import (
    LANGUAGE_NAMES,
    SOURCE_HEADERS,
    all_language_target_headers,
    normalize_language_code,
    target_header_candidates,
    variant_header_candidates,
)

GLOSSARY_SHEET_NAMES = {'术语表', 'glossary', 'terms', 'term base', 'termbase'}
TERM_BASE_FILENAME_KEYWORDS = {'术语', 'glossary', 'term', 'termbase'}
OUTPUT_DIR_NAME_HINTS = {'output', 'out', 'final', 'result'}
SUPPORT_SHEET_NAME_KEYWORDS = {
    '裁决',
    '审计',
    '返修',
    '需确认',
    '总览',
    '错误模式',
    '学习笔记',
    '详细记录',
    '术语行筛选',
    'review',
    'summary',
    'overview',
    'details',
    'audit',
    'decision',
    'review log',
    'fix log',
}
LEGAL_TERM_CHECK_SKIP_MARKERS = {'隐私政策', '用户协议'}
LEGAL_TERM_CHECK_MIN_LENGTH = 1000
PERSON_NAME_CATEGORY_MARKERS = {'人名', '角色', 'person', 'name', 'character'}
SOFT_TERM_CATEGORY_MARKERS = {'soft', 'generic', 'common', '参考', '泛词', '通用词'}
CONTEXTUAL_HARD_TERM_CATEGORY_MARKERS = {'强术语', '强制', '固定译名', '专有名词', '专名'}
AUTO_SOFT_SOURCE_TERMS = {
    '使用',
    '激活',
    '领取',
    '解锁',
    '购买',
    '重置',
    '刷新',
    '失败',
    '可领取',
    '普通',
    '增加',
    '随机',
    '提升',
    '额外',
    '唯一',
    '时间',
    '开启',
    '自动',
    '注意',
    '选择',
    '完成',
    '通过',
    '通关',
    '成功',
    '基础',
    '击杀',
    '发送',
    '可获',
    '数量',
    '同时',
    '需要',
    '已领取',
    '获得',
    '获得了',
    '任务',
    '当前',
    '操作',
    '申请',
    '达到',
    '玩家',
    '未解锁',
    '创建',
    '最大',
    '匹配',
    '分解',
    '已有',
    '加入',
    '未开启',
    '进阶',
    '输出',
    '提交',
    '设置',
    '跟随',
    '穿戴',
    '星级',
    '属性',
    '接取',
    '冷却',
    '全服',
    '分钟',
    '道具',
    '邮件',
    '主角',
    '排名',
    '每日',
    '赠送',
    '获取',
    '兑换',
    '取消',
    '提示',
    '其他',
    '直接',
    '团队',
    '奖励',
    '补偿',
}
AUTO_SOFT_TARGET_WORDS = {
    'additional',
    'amount',
    'at',
    'auto',
    'available',
    'base',
    'basic',
    'buy',
    'claim',
    'claimable',
    'claimed',
    'clear',
    'complete',
    'completed',
    'death',
    'extra',
    'failed',
    'kill',
    'notice',
    'obtain',
    'open',
    'random',
    'refresh',
    'required',
    'reset',
    'select',
    'send',
    'stage',
    'success',
    'time',
    'unique',
    'unlock',
    'use',
}
GENERIC_ROLE_TARGETS = {
    'ally',
    'base',
    'boss',
    'captain',
    'character',
    'commander',
    'enemy',
    'hero',
    'member',
    'miner',
    'player',
    'protagonist',
    'role',
    'soldier',
    'survivor',
}

_LANGUAGE_CODES = tuple(LANGUAGE_NAMES.keys())

LANGUAGE_TARGET_HEADERS = {
    lang: target_header_candidates(lang, include_generic=True)
    for lang in _LANGUAGE_CODES
}

LANGUAGE_VARIANT_HEADERS = {
    lang: variant_header_candidates(lang)
    for lang in _LANGUAGE_CODES
}


def _target_header_candidates(lang: str) -> set[str]:
    return target_header_candidates(lang, include_generic=True)


def _variant_header_candidates(lang: str) -> set[str]:
    return variant_header_candidates(lang)


def _all_target_header_candidates() -> set[str]:
    return all_language_target_headers()


def _all_variant_header_candidates() -> set[str]:
    values: set[str] = set()
    for lang in _LANGUAGE_CODES:
        values.update(variant_header_candidates(lang))
    return values


def _has_explicit_language_header(headers: Sequence[str], language_headers: set[str]) -> bool:
    return any(
        header in language_headers and not (idx == 0 and header == 'id')
        for idx, header in enumerate(headers)
    )


def _should_skip_term_checks(source: str) -> bool:
    text = str(source or '')
    return (
        len(text) >= LEGAL_TERM_CHECK_MIN_LENGTH
        and any(marker in text for marker in LEGAL_TERM_CHECK_SKIP_MARKERS)
    )


def _is_auto_soft_term(cn: str, target: str, category: str = '') -> bool:
    """Downgrade obvious generic words when the glossary has no category.

    Project glossaries often mix real domain terms with rows like
    "获得 -> Obtain" and "成功 -> Success". When no category is present,
    those rows should guide wording, not block delivery.
    """
    if _clean_term_cell(category):
        return False
    source_term = _clean_term_cell(cn)
    if source_term in AUTO_SOFT_SOURCE_TERMS:
        return True
    if len(source_term) > 3:
        return False
    target_words = {
        word
        for word in re.findall(r'[A-Za-z]+', _clean_term_cell(target).lower())
        if word not in {'a', 'an', 'the', 'to', 'of', 'for', 'in', 'on'}
    }
    return bool(target_words) and target_words.issubset(AUTO_SOFT_TARGET_WORDS)


def _collect_term_context(
    workbook,
    term_base: str | Path | Sequence[str | Path] | None,
    lang: str = 'en',
) -> dict[str, object]:
    lang = normalize_language_code(lang)
    strong_terms: dict[str, dict] = {}
    soft_terms: dict[str, dict] = {}
    person_name_terms: list[tuple[str, str]] = []
    seen_person_names: set[tuple[str, str]] = set()

    def add(
        cn: str,
        target: str,
        category: str = '',
        variants: Sequence[str] | None = None,
        enforce_case: bool = False,
    ) -> None:
        cn = _clean_term_cell(cn)
        target = _clean_term_cell(target)
        category = _clean_term_cell(category)
        variants = [_clean_term_cell(v) for v in (variants or [])]
        variants = [v for v in variants if v and v.lower() != target.lower()]
        if not cn or not target:
            return
        if _is_person_name_category(category):
            if _is_generic_role_target(target):
                return
            key = (cn, target)
            if key not in seen_person_names:
                seen_person_names.add(key)
                person_name_terms.append(key)
            return
        bucket = soft_terms if _is_soft_term_category(category) or _is_auto_soft_term(cn, target, category) else strong_terms
        _add_term_lookup_entry(
            bucket,
            cn,
            target,
            variants,
            enforce_case,
            enforce_in_context=_enforces_contextual_term(category),
        )

    _collect_terms_from_workbook(workbook, add, lang=lang)

    for path in _iter_term_base_paths(term_base):
        term_path = Path(path)
        if not term_path.exists():
            continue
        if term_path.suffix.lower() == '.json':
            _collect_terms_from_json(term_path, add, lang=lang)
            continue
        term_wb = load_workbook(term_path, read_only=False, data_only=True)
        try:
            _collect_terms_from_workbook(term_wb, add, all_sheets=True, lang=lang)
        finally:
            term_wb.close()

    # Strong terms win when a later soft row repeats the same source term.
    for cn in list(soft_terms):
        if cn in strong_terms:
            del soft_terms[cn]

    person_name_terms.sort(key=lambda item: len(item[0]), reverse=True)
    return {
        'strong': strong_terms,
        'soft': soft_terms,
        'person_names': person_name_terms,
    }


def _collect_terms_from_workbook(workbook, add, all_sheets: bool = False, lang: str = 'en') -> None:
    lang = normalize_language_code(lang)
    target_candidates = target_header_candidates(lang, include_generic=True)
    variant_candidates = variant_header_candidates(lang)
    language_headers = all_language_target_headers()
    for ws in workbook.worksheets:
        is_glossary_sheet = _is_glossary_sheet(ws)
        header = [str(value or '').strip().lower() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
        cn_idx = _find_header(header, set(SOURCE_HEADERS))
        target_idx = _find_header(header, target_candidates)
        variant_idx = _find_header(header, variant_candidates)
        category_idx = _find_header(header, {'分类', '类别', 'category', 'type', 'tag', 'tags'})
        enforce_idx = _find_header(header, {'enforce_case', '大小写', '大小写约束'})
        has_language_header = _has_explicit_language_header(header, language_headers)

        if is_glossary_sheet:
            cn_idx = cn_idx if cn_idx is not None else _fallback_index(header, 0)
            if target_idx is None:
                if has_language_header:
                    continue
                target_idx = _fallback_index(header, 1)
            if variant_idx is None and not has_language_header:
                variant_idx = _fallback_index(header, 2)
            category_idx = category_idx if category_idx is not None else _fallback_index(header, 3)
        elif not all_sheets:
            continue
        elif cn_idx is None or target_idx is None:
            continue
        elif category_idx is None and len(workbook.worksheets) > 1:
            # Avoid treating a full delivery workbook passed as --term-base as a glossary.
            continue

        if cn_idx is None or target_idx is None:
            continue
        max_idx = max(index for index in (cn_idx, target_idx, variant_idx, category_idx, enforce_idx) if index is not None)
        for row in ws.iter_rows(min_row=2, max_col=max_idx + 1, values_only=True):
            cn = row[cn_idx] if cn_idx < len(row) else ''
            target = row[target_idx] if target_idx < len(row) else ''
            raw_variants = row[variant_idx] if variant_idx is not None and variant_idx < len(row) else ''
            category = row[category_idx] if category_idx is not None and category_idx < len(row) else ''
            enforce_raw = row[enforce_idx] if enforce_idx is not None and enforce_idx < len(row) else ''
            add(cn, target, category, _split_term_variants(raw_variants), _truthy_cell(enforce_raw))


def _collect_terms_from_json(path: Path, add, lang: str = 'en') -> None:
    data = json.loads(path.read_text(encoding='utf-8'))
    lookup = data.get('lookup', data) if isinstance(data, dict) else {}
    if not isinstance(lookup, dict):
        return
    for cn, value in lookup.items():
        if isinstance(value, str):
            add(cn, value)
        elif isinstance(value, list):
            values = [_clean_term_cell(v) for v in value]
            values = [v for v in values if v]
            if values:
                add(cn, values[0], variants=values[1:])
        elif isinstance(value, dict):
            target = value.get('primary') or value.get(lang) or value.get('target') or value.get('en') or ''
            raw_variants = value.get('variants', [])
            if isinstance(raw_variants, str):
                variants = _split_term_variants(raw_variants)
            else:
                variants = [_clean_term_cell(v) for v in raw_variants if _clean_term_cell(v)]
            category = value.get('category') or value.get('type') or value.get('constraint') or ''
            add(cn, target, category, variants, bool(value.get('enforce_case', False)))


def _add_term_lookup_entry(
    lookup: dict[str, dict],
    cn: str,
    target: str,
    variants: Sequence[str],
    enforce_case: bool,
    enforce_in_context: bool = False,
) -> None:
    entry = lookup.setdefault(cn, {'primary': target, 'variants': [], 'enforce_case': False, 'enforce_in_context': False})
    if not entry.get('primary'):
        entry['primary'] = target
    for variant in variants:
        if variant.lower() == str(entry.get('primary', '')).lower():
            continue
        if all(variant.lower() != str(existing).lower() for existing in entry.get('variants', [])):
            entry.setdefault('variants', []).append(variant)
    entry['enforce_case'] = bool(entry.get('enforce_case')) or enforce_case
    entry['enforce_in_context'] = bool(entry.get('enforce_in_context')) or enforce_in_context


def _split_term_variants(value) -> list[str]:
    text = _clean_term_cell(value)
    if not text:
        return []
    return [_clean_term_cell(part) for part in re.split(r'[;,|/、]+', text) if _clean_term_cell(part)]


def _clean_term_cell(value) -> str:
    text = str(value or '').strip()
    return '' if text.lower() == 'nan' else text


def _truthy_cell(value) -> bool:
    text = _clean_term_cell(value).lower()
    return text in {'1', 'true', 'yes', 'y', '是', '强制'}


def _is_soft_term_category(category: str) -> bool:
    text = str(category or '').strip().lower()
    return any(marker in text for marker in SOFT_TERM_CATEGORY_MARKERS)


def _enforces_contextual_term(category: str) -> bool:
    text = str(category or '').strip().lower()
    return any(marker in text for marker in CONTEXTUAL_HARD_TERM_CATEGORY_MARKERS)


def _is_person_name_category(category: str) -> bool:
    text = str(category or '').strip().lower()
    return any(marker in text for marker in PERSON_NAME_CATEGORY_MARKERS)


def _is_generic_role_target(target: str) -> bool:
    normalized = re.sub(r"[^A-Za-z\s]", '', str(target or '')).strip().lower()
    return normalized in GENERIC_ROLE_TARGETS


def _resolve_term_base_paths(
    workbook_path: Path,
    term_base: str | Path | Sequence[str | Path] | None,
    auto_discover_terms: bool,
) -> list[Path]:
    paths: list[Path] = []
    seen: set[str] = set()

    def add(path: str | Path) -> None:
        candidate = Path(path)
        try:
            key = str(candidate.resolve()).lower()
        except OSError:
            key = str(candidate.absolute()).lower()
        if key in seen:
            return
        seen.add(key)
        paths.append(candidate)

    explicit_paths = _iter_term_base_paths(term_base)
    for path in explicit_paths:
        add(path)

    if auto_discover_terms and not explicit_paths:
        for path in _discover_term_base_paths(workbook_path):
            add(path)

    return paths


def _discover_term_base_paths(workbook_path: Path) -> list[Path]:
    workbook_path = Path(workbook_path)
    search_dirs = [workbook_path.parent]
    if _looks_like_output_dir(workbook_path.parent):
        search_dirs.append(workbook_path.parent.parent)

    discovered: list[Path] = []
    seen: set[str] = set()
    for directory in search_dirs:
        if not directory.exists() or not directory.is_dir():
            continue
        for candidate in directory.iterdir():
            if not candidate.is_file() or candidate.name.startswith('~$'):
                continue
            if candidate.resolve() == workbook_path.resolve():
                continue
            if candidate.suffix.lower() not in {'.xlsx', '.xlsm', '.json'}:
                continue
            stem = candidate.stem.lower().replace(' ', '')
            if not any(keyword in stem for keyword in TERM_BASE_FILENAME_KEYWORDS):
                continue
            key = str(candidate.resolve()).lower()
            if key in seen:
                continue
            seen.add(key)
            discovered.append(candidate)
    return discovered


def _looks_like_output_dir(path: Path) -> bool:
    name = path.name.lower().replace('-', '_')
    return any(name == hint or name.startswith(f'{hint}_') or name.endswith(f'_{hint}') for hint in OUTPUT_DIR_NAME_HINTS)


def _iter_term_base_paths(term_base: str | Path | Sequence[str | Path] | None) -> list[str | Path]:
    if term_base is None:
        return []
    if isinstance(term_base, (str, Path)):
        return [term_base]
    return list(term_base)


def _find_header(headers: list[str], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in candidates:
            return index
    return None


def _fallback_index(headers: list[str], fallback: int) -> int | None:
    return fallback if fallback < len(headers) else None


def _detect_columns(ws, lang: str = 'en') -> tuple[int | None, int | None, int | None]:
    lang = normalize_language_code(lang)
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(v or '').strip().lower() for v in first]

    def pick(candidates: set[str], fallback: int | None = None) -> int | None:
        for idx, header in enumerate(headers):
            if header in candidates:
                return idx
        return fallback if fallback is None or fallback < len(headers) else None

    id_col = pick({'id', 'key'}, 0)
    src_col = pick(set(SOURCE_HEADERS), 1)
    tgt_col = pick(target_header_candidates(lang, include_generic=True))
    if tgt_col is None and _has_explicit_language_header(headers, all_language_target_headers()):
        return id_col, src_col, None
    if tgt_col is None:
        tgt_col = 2 if len(headers) > 2 else None
    return id_col, src_col, tgt_col


def _is_glossary_sheet(ws) -> bool:
    title = str(ws.title or '').strip().lower()
    compact_title = re.sub(r'\s+', '', title)
    if title in GLOSSARY_SHEET_NAMES or compact_title in {'termbase', 'terms'} or '术语' in title:
        return True

    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(v or '').strip().lower() for v in first[:4]]
    if headers == ['cn', 'en', 'en2', '分类']:
        return True
    return (
        len(headers) >= 4
        and headers[0] in {'cn', 'zh', 'source', 'original', '中文', '简体中文', '原文'}
        and headers[1] in _all_target_header_candidates()
        and headers[2] in _all_variant_header_candidates()
        and headers[3] in {'分类', '类别', 'category', 'type', 'tag', 'tags'}
    )


def _is_support_sheet(ws) -> bool:
    title = str(ws.title or '').strip().lower()
    compact_title = re.sub(r'[\s_-]+', '', title)
    return any(keyword in title or keyword.replace(' ', '') in compact_title for keyword in SUPPORT_SHEET_NAME_KEYWORDS)
