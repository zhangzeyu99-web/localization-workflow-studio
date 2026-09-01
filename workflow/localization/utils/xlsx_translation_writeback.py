"""Precise XLSX target-cell writeback without rebuilding workbook objects."""
from __future__ import annotations

import json
import posixpath
import re
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils.language_config import SOURCE_HEADERS, normalize_language_code, target_header_candidates


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"m": MAIN_NS, "r": REL_NS, "pr": PKG_REL_NS}
ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)


@dataclass(frozen=True)
class WritebackResult:
    output_dir: Path
    output_files: int
    source_rows: int
    written_cells: int
    invalid_style_refs: int


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8-sig").splitlines()
        if line.strip()
    ]


def _headers(sheet: Any) -> list[str]:
    return [str(cell.value or "").strip().lower() for cell in sheet[1]]


def _find_source_column(headers: list[str]) -> int:
    candidates = {str(value).strip().lower() for value in SOURCE_HEADERS}
    column = next((index for index, header in enumerate(headers) if header in candidates), None)
    if column is None:
        raise ValueError("source language column is missing")
    return column + 1


def _target_columns(headers: list[str], target_langs: list[str]) -> dict[str, int]:
    id_column = next((index for index, header in enumerate(headers) if header in {"id", "key"}), None)
    result: dict[str, int] = {}
    for raw_lang in target_langs:
        lang = str(raw_lang).upper()
        candidates = target_header_candidates(normalize_language_code(lang))
        index = next(
            (
                index
                for index, header in enumerate(headers)
                if index != id_column and header in candidates
            ),
            None,
        )
        if index is None:
            raise ValueError(f"missing target language column: {lang}")
        result[lang] = index + 1
    return result


def _sheet_xml_paths(entries: dict[str, bytes]) -> dict[str, str]:
    workbook = ET.fromstring(entries["xl/workbook.xml"])
    relationships = ET.fromstring(entries["xl/_rels/workbook.xml.rels"])
    targets = {
        relation.attrib["Id"]: relation.attrib["Target"]
        for relation in relationships.findall("pr:Relationship", NS)
    }
    result: dict[str, str] = {}
    for sheet in workbook.findall("m:sheets/m:sheet", NS):
        relation_id = sheet.attrib[f"{{{REL_NS}}}id"]
        target = targets[relation_id].replace("\\", "/")
        if target.startswith("/"):
            archive_path = target.lstrip("/")
        else:
            archive_path = posixpath.normpath(posixpath.join("xl", target))
        result[sheet.attrib["name"]] = archive_path
    return result


def _patch_sheet(xml: bytes, changes: list[tuple[int, int, str]]) -> bytes:
    text = xml.decode("utf-8")
    for row_number, column, value in changes:
        row_pattern = re.compile(
            rf'(<row\b(?=[^>]*\br="{row_number}")[^>]*>)(.*?)(</row>)',
            re.DOTALL,
        )
        row_match = row_pattern.search(text)
        if row_match is None:
            raise ValueError(f"worksheet row is missing: {row_number}")
        reference = f"{get_column_letter(column)}{row_number}"
        preserve = ' xml:space="preserve"' if value[:1].isspace() or value[-1:].isspace() else ""
        encoded = escape(value)
        row_body = row_match.group(2)
        cell_pattern = re.compile(
            rf'<c\b(?=[^>]*\br="{re.escape(reference)}")(?P<attrs>[^>]*?)(?:\s*/>|>(?P<body>.*?)</c>)',
            re.DOTALL,
        )
        cell_match = cell_pattern.search(row_body)
        if cell_match is not None:
            attributes = re.sub(r'\s+t="[^"]*"', "", cell_match.group("attrs"))
            replacement = f'<c{attributes} t="inlineStr"><is><t{preserve}>{encoded}</t></is></c>'
            row_body = row_body[: cell_match.start()] + replacement + row_body[cell_match.end() :]
        else:
            replacement = f'<c r="{reference}" t="inlineStr"><is><t{preserve}>{encoded}</t></is></c>'
            insert_at = len(row_body)
            for existing in re.finditer(r'<c\b[^>]*\br="([A-Z]+)\d+"', row_body):
                if _column_index(existing.group(1)) > column:
                    insert_at = existing.start()
                    break
            row_body = row_body[:insert_at] + replacement + row_body[insert_at:]
        text = text[: row_match.start(2)] + row_body + text[row_match.end(2) :]
    return text.encode("utf-8")


def _column_index(letters: str) -> int:
    value = 0
    for char in letters.upper():
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _style_validation(entries: dict[str, bytes]) -> int:
    styles = ET.fromstring(entries["xl/styles.xml"])
    cell_xfs = styles.find("m:cellXfs", NS)
    style_count = len(cell_xfs) if cell_xfs is not None else 0
    invalid = 0
    for name, raw in entries.items():
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        root = ET.fromstring(raw)
        for cell in root.findall(".//m:c", NS):
            style = cell.attrib.get("s")
            if style is not None and (not style.isdigit() or int(style) >= style_count):
                invalid += 1
    return invalid


def validate_xlsx(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path, "r") as archive:
        entries = {name: archive.read(name) for name in archive.namelist()}
    invalid = _style_validation(entries)
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet_count = len(workbook.sheetnames)
    finally:
        workbook.close()
    return {
        "path": str(path),
        "normal_open": True,
        "sheet_count": sheet_count,
        "invalid_style_refs": invalid,
        "ok": invalid == 0,
    }


def _patch_workbook(
    source: Path,
    output: Path,
    changes_by_sheet: dict[str, list[tuple[int, int, str]]],
) -> int:
    with zipfile.ZipFile(source, "r") as archive:
        infos = archive.infolist()
        entries = {info.filename: archive.read(info.filename) for info in infos}
    paths = _sheet_xml_paths(entries)
    for sheet_name, changes in changes_by_sheet.items():
        if sheet_name not in paths:
            raise ValueError(f"worksheet is missing: {sheet_name}")
        archive_path = paths[sheet_name]
        entries[archive_path] = _patch_sheet(entries[archive_path], changes)
    invalid = _style_validation(entries)
    if invalid:
        raise ValueError(f"workbook contains {invalid} invalid style references")
    output.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(dir=output.parent, suffix=".xlsx", delete=False) as handle:
        temporary = Path(handle.name)
    try:
        with zipfile.ZipFile(temporary, "w") as archive:
            for info in infos:
                archive.writestr(info, entries[info.filename])
        temporary.replace(output)
    finally:
        if temporary.exists():
            temporary.unlink()
    return invalid


def write_translation_workbooks(
    *,
    inputs: list[Path],
    cache_jsonl: Path,
    target_langs: list[str],
    output_dir: Path,
) -> WritebackResult:
    target_langs = [str(lang).upper() for lang in target_langs]
    cache_rows = _read_jsonl(cache_jsonl)
    inputs_by_name = {path.name: path for path in inputs}
    if len(inputs_by_name) != len(inputs):
        raise ValueError("input workbooks must have unique file names")
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cache_rows:
        grouped[str(row.get("source_file") or "")].append(row)
    unknown = sorted(set(grouped) - set(inputs_by_name))
    if unknown:
        raise ValueError(f"cache references unknown source workbook(s): {unknown}")

    written_cells = 0
    invalid_style_refs = 0
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, source in inputs_by_name.items():
        workbook = load_workbook(source, read_only=False, data_only=False)
        changes_by_sheet: dict[str, list[tuple[int, int, str]]] = defaultdict(list)
        try:
            columns_by_sheet: dict[str, tuple[int | None, int, dict[str, int]]] = {}
            for row in grouped.get(name, []):
                sheet_name = str(row.get("sheet") or "")
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"worksheet is missing: {name}:{sheet_name}")
                sheet = workbook[sheet_name]
                if sheet_name not in columns_by_sheet:
                    headers = _headers(sheet)
                    id_column = next(
                        (index + 1 for index, header in enumerate(headers) if header in {"id", "key"}),
                        None,
                    )
                    columns_by_sheet[sheet_name] = (
                        id_column,
                        _find_source_column(headers),
                        _target_columns(headers, target_langs),
                    )
                id_column, source_column, target_columns = columns_by_sheet[sheet_name]
                row_number = int(row["row"])
                current_id = sheet.cell(row=row_number, column=id_column).value if id_column else row_number
                if str(current_id) != str(row.get("id") or row_number):
                    raise ValueError(
                        f"source row id drift: {name}:{sheet_name}!R{row_number}"
                    )
                expected_key = f"{name}::{sheet_name}::{current_id}::{row_number}"
                if row.get("key") and str(row["key"]) != expected_key:
                    raise ValueError(
                        f"source row key drift: {name}:{sheet_name}!R{row_number}"
                    )
                current_source = sheet.cell(row=row_number, column=source_column).value
                current_source_text = "" if current_source is None else str(current_source)
                expected_source_text = "" if row.get("cn") is None else str(row.get("cn"))
                if current_source_text != expected_source_text:
                    raise ValueError(
                        f"source text drift: {name}:{sheet_name}!R{row_number}"
                    )
                translations = row.get("translations") or {}
                for lang in target_langs:
                    target = str(translations.get(lang) or "").strip()
                    if not target:
                        raise ValueError(
                            f"blank target translation: {name}:{sheet_name}!R{row_number}:{lang}"
                        )
                    changes_by_sheet[sheet_name].append(
                        (row_number, target_columns[lang], target)
                    )
                    written_cells += 1
        finally:
            workbook.close()
        output = output_dir / name
        invalid_style_refs += _patch_workbook(source, output, changes_by_sheet)
        validation = validate_xlsx(output)
        if not validation["ok"]:
            raise ValueError(f"writeback validation failed: {validation}")

    return WritebackResult(
        output_dir=output_dir,
        output_files=len(inputs),
        source_rows=len(cache_rows),
        written_cells=written_cells,
        invalid_style_refs=invalid_style_refs,
    )


def verify_translation_cache(
    delivery_dir: Path,
    cache_jsonl: Path,
    target_langs: list[str],
) -> dict[str, Any]:
    target_langs = [str(lang).upper() for lang in target_langs]
    rows = _read_jsonl(cache_jsonl)
    workbooks: dict[str, Any] = {}
    columns: dict[tuple[str, str], tuple[int | None, int, dict[str, int]]] = {}
    issues: list[dict[str, str]] = []
    try:
        for row in rows:
            name = str(row.get("source_file") or "")
            sheet_name = str(row.get("sheet") or "")
            row_number = int(row["row"])
            if name not in workbooks:
                path = delivery_dir / name
                if not path.exists():
                    issues.append({"type": "output_file_missing", "key": name, "lang": ""})
                    continue
                workbooks[name] = load_workbook(path, read_only=False, data_only=False)
            workbook = workbooks.get(name)
            if workbook is None:
                continue
            if sheet_name not in workbook.sheetnames:
                issues.append({"type": "output_sheet_missing", "key": f"{name}:{sheet_name}", "lang": ""})
                continue
            sheet = workbook[sheet_name]
            column_key = (name, sheet_name)
            if column_key not in columns:
                headers = _headers(sheet)
                id_column = next(
                    (index + 1 for index, header in enumerate(headers) if header in {"id", "key"}),
                    None,
                )
                columns[column_key] = (
                    id_column,
                    _find_source_column(headers),
                    _target_columns(headers, target_langs),
                )
            id_column, source_column, target_columns = columns[column_key]
            current_id = sheet.cell(row=row_number, column=id_column).value if id_column else row_number
            current_source = sheet.cell(row=row_number, column=source_column).value
            expected_key = f"{name}::{sheet_name}::{current_id}::{row_number}"
            if str(current_id) != str(row.get("id") or row_number) or (
                row.get("key") and str(row["key"]) != expected_key
            ):
                issues.append({"type": "source_row_identity_mismatch", "key": str(row.get("key") or expected_key), "lang": ""})
                continue
            current_source_text = "" if current_source is None else str(current_source)
            expected_source_text = "" if row.get("cn") is None else str(row.get("cn"))
            if current_source_text != expected_source_text:
                issues.append({"type": "source_value_mismatch", "key": expected_key, "lang": ""})
            translations = row.get("translations") or {}
            for lang in target_langs:
                current = sheet.cell(row=row_number, column=target_columns[lang]).value
                expected = str(translations.get(lang) or "")
                if str(current or "") != expected:
                    issues.append(
                        {"type": "translation_value_mismatch", "key": expected_key, "lang": lang}
                    )
    finally:
        for workbook in workbooks.values():
            workbook.close()
    return {
        "cache_jsonl": str(cache_jsonl),
        "checked_rows": len(rows),
        "target_languages": target_langs,
        "hard_blockers": len(issues),
        "issues": issues,
        "cache_matches_delivery": len(issues) == 0,
    }
