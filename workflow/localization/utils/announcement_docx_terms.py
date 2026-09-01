"""Announcement term workbook loading and term-hit extraction.

Implementation module split out of utils/announcement_docx_harness.py. Import
these symbols through utils.announcement_docx_harness to keep the public
surface stable.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils.announcement_docx_common import (
    CANONICAL_LANGUAGE_HEADER,
    _CJK_RE,
    _clean_cell,
    _language_code_for_header,
)

_ANNOUNCEMENT_PLACEHOLDER_RE = re.compile(r"<@\d+>|\$\{\d+\}|%[sdif]|\{[^{}]+\}|<[^>]+>|\\n")
_ANNOUNCEMENT_SENTENCE_PUNCTUATION = set("，。！？；：、,.?!;:")
_LOW_VALUE_ANNOUNCEMENT_TERMS = {
    "好的",
    "游戏",
    "时间",
    "小时",
    "完成",
    "发放",
    "领取",
    "获得",
    "获取",
    "更新",
    "进入",
    "点击",
    "下载",
    "界面",
    "弹窗",
    "邮件",
    "活动",
    "系统",
    "玩法",
    "服务器",
    "维护",
    "补偿",
    "内容",
    "任务",
    "奖励",
    "开启",
    "开放",
    "修复",
    "新增",
    "部分",
    "所有",
    "普通",
    "使用",
    "问题",
    "确认",
    "取消",
    "返回",
    "查看",
    "选择",
    "购买",
    "前往",
    "本次",
    "当前",
    "公告",
    "进入游戏",
    "开始游戏",
}
_LOW_VALUE_ANNOUNCEMENT_PREFIXES = (
    "使用后",
    "用于",
    "可",
    "将",
    "已",
    "未",
    "请",
    "点击",
    "前往",
    "打开",
    "关闭",
    "获得",
    "完成",
    "通关",
    "达到",
    "活动期间",
    "当前",
    "本期",
    "每日",
    "成功",
    "失败",
    "进入",
    "开始",
    "返回",
    "查看",
    "选择",
)
_LOW_VALUE_ANNOUNCEMENT_SUBSTRINGS = (
    "是否",
    "不足",
    "尚未",
    "暂无",
    "已达",
    "未领取",
    "已领取",
    "补发",
    "异常",
    "排行奖励",
    "奖励邮件",
    "正在",
    "倒计时",
)
_TERM_NOUN_SUFFIXES = (
    "活动",
    "系统",
    "玩法",
    "副本",
    "战场",
    "商店",
    "商城",
    "公会",
    "英雄",
    "角色",
    "皮肤",
    "碎片",
    "宝箱",
    "礼包",
    "月卡",
    "通行证",
    "圣器",
    "遗物",
    "机关",
    "装备",
    "材料",
    "水晶",
    "金币",
    "硬币",
    "积分",
    "图鉴",
    "技能",
)


@dataclass(frozen=True)
class LanguageSpec:
    header: str
    code: str
    column_index: int


@dataclass(frozen=True)
class TermEntry:
    source: str
    target: str
    term_id: str = ""


@dataclass(frozen=True)
class SentenceAdaptationEntry:
    priority: int
    match_type: str
    entry_id: str
    announcement_cn: str
    official_cn_template: str
    targets: dict[str, str]
    source_row: int


@dataclass(frozen=True)
class AnnouncementTerms:
    path: Path
    languages: list[LanguageSpec]
    by_language: dict[str, dict[str, TermEntry]]
    sentence_adaptations: list[SentenceAdaptationEntry]


def load_announcement_terms(path: str | Path) -> AnnouncementTerms:
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Glossary"] if "Glossary" in wb.sheetnames else wb.active
        languages = _language_specs_from_headers(
            [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        )
        by_language: dict[str, dict[str, TermEntry]] = {}
        for spec in languages:
            by_language[spec.header] = {}

        for row in range(2, ws.max_row + 1):
            term_id = _clean_cell(ws.cell(row, 1).value)
            source = _clean_cell(ws.cell(row, 2).value)
            if not source:
                continue
            for spec in languages:
                target = _clean_cell(ws.cell(row, spec.column_index).value)
                if target:
                    by_language[spec.header][source] = TermEntry(source=source, target=target, term_id=term_id)
        sentence_adaptations = _load_sentence_adaptations(wb, languages)
    finally:
        wb.close()
    return AnnouncementTerms(
        path=path,
        languages=languages,
        by_language=by_language,
        sentence_adaptations=sentence_adaptations,
    )


def _read_announcement_language_specs(path: Path) -> list[LanguageSpec]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Glossary"] if "Glossary" in wb.sheetnames else wb.active
        headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
        return _language_specs_from_headers(headers)
    finally:
        wb.close()


def _language_specs_from_headers(headers: list[str]) -> list[LanguageSpec]:
    languages: list[LanguageSpec] = []
    seen_codes: set[str] = set()
    for col in range(3, len(headers) + 1):
        input_header = headers[col - 1]
        code = _language_code_for_header(input_header)
        if not code or code in seen_codes:
            continue
        seen_codes.add(code)
        header = CANONICAL_LANGUAGE_HEADER[code]
        languages.append(LanguageSpec(header=header, code=code, column_index=col))
    return languages


def _find_term_hits(source: str, terms: AnnouncementTerms) -> list[dict[str, Any]]:
    all_sources = {
        term
        for lang_terms in terms.by_language.values()
        for term in lang_terms
        if len(term) >= 2 and not _is_low_value_announcement_term(term) and _term_occurs(source, term)
    }
    selected: list[str] = []
    for term in sorted(all_sources, key=len, reverse=True):
        if any(term in longer for longer in selected):
            continue
        selected.append(term)

    hits: list[dict[str, Any]] = []
    for source_term in selected:
        targets = {}
        for spec in terms.languages:
            entry = terms.by_language.get(spec.header, {}).get(source_term)
            if entry:
                targets[spec.header] = entry.target
        if targets:
            hits.append({"source": source_term, "targets": targets})
    return hits


def _find_sentence_adaptations(source: str, terms: AnnouncementTerms) -> list[dict[str, Any]]:
    hits: list[SentenceAdaptationEntry] = []
    for entry in terms.sentence_adaptations:
        if entry.match_type == "official_exact":
            matched = _sentence_exact_match(source, entry)
        else:
            matched = _normalized_sentence_contains(source, entry.announcement_cn)
        if matched:
            hits.append(entry)

    hits.sort(
        key=lambda entry: (
            entry.priority,
            0 if entry.match_type == "official_exact" else 1,
            -len(_normalize_sentence_text(entry.announcement_cn)),
            entry.source_row,
        )
    )
    return [
        {
            "priority": entry.priority,
            "match_type": entry.match_type,
            "id": entry.entry_id,
            "announcement_cn": entry.announcement_cn,
            "official_cn_template": entry.official_cn_template,
            "targets": entry.targets,
        }
        for entry in hits
    ]


def _load_sentence_adaptations(wb, languages: list[LanguageSpec]) -> list[SentenceAdaptationEntry]:
    if "SentenceTemplates" not in wb.sheetnames:
        return []

    ws = wb["SentenceTemplates"]
    headers = [_clean_cell(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    required_headers = ("Priority", "MatchType", "ID", "AnnouncementCN", "OfficialCNTemplate")
    header_index: dict[str, int] = {}
    for col, header in enumerate(headers, start=1):
        if header and header not in header_index:
            header_index[header] = col
    missing = [header for header in required_headers if header not in header_index]
    if missing:
        raise ValueError(f"SentenceTemplates missing columns: {missing}")

    required_columns = {header_index[header] for header in required_headers}
    language_columns: dict[str, int] = {}
    for col, header in enumerate(headers, start=1):
        if col in required_columns:
            continue
        code = _language_code_for_header(header)
        if code:
            language_columns[CANONICAL_LANGUAGE_HEADER[code]] = col
    missing_languages = [spec.header for spec in languages if spec.header not in language_columns]
    if missing_languages:
        raise ValueError(f"SentenceTemplates missing language columns: {missing_languages}")

    entries: list[SentenceAdaptationEntry] = []
    for row in range(2, ws.max_row + 1):
        values = [_clean_cell(ws.cell(row, col).value) for col in required_columns]
        if not any(values):
            continue
        try:
            priority = int(float(_clean_cell(ws.cell(row, header_index["Priority"]).value)))
        except ValueError as exc:
            raise ValueError(f"SentenceTemplates row {row} has invalid Priority") from exc
        if priority < 1:
            raise ValueError(f"SentenceTemplates row {row} has invalid Priority")
        match_type = _clean_cell(ws.cell(row, header_index["MatchType"]).value).lower()
        if match_type not in {"official_exact", "official_similar"}:
            raise ValueError(f"SentenceTemplates row {row} has unsupported MatchType: {match_type}")
        entry_id = _clean_cell(ws.cell(row, header_index["ID"]).value)
        announcement_cn = _clean_cell(ws.cell(row, header_index["AnnouncementCN"]).value)
        official_cn_template = _clean_cell(ws.cell(row, header_index["OfficialCNTemplate"]).value)
        if not entry_id or not announcement_cn or not official_cn_template:
            raise ValueError(f"SentenceTemplates row {row} has empty required fields")
        targets = {
            spec.header: _clean_cell(ws.cell(row, language_columns[spec.header]).value)
            for spec in languages
        }
        missing_targets = [header for header, target in targets.items() if not target]
        if missing_targets:
            raise ValueError(f"SentenceTemplates row {row} has empty targets: {missing_targets}")
        entries.append(
            SentenceAdaptationEntry(
                priority=priority,
                match_type=match_type,
                entry_id=entry_id,
                announcement_cn=announcement_cn,
                official_cn_template=official_cn_template,
                targets=targets,
                source_row=row,
            )
        )
    return entries


def _sentence_exact_match(source: str, entry: SentenceAdaptationEntry) -> bool:
    if _normalized_sentence_contains(source, entry.announcement_cn):
        return True
    normalized_source = _normalize_sentence_text(source)
    normalized_template = _normalize_sentence_text(entry.official_cn_template)
    parts = re.split(r"(<@\d+>)", normalized_template)
    pattern = "".join(r".+?" if re.fullmatch(r"<@\d+>", part) else re.escape(part) for part in parts)
    return bool(pattern and re.search(pattern, normalized_source))


def _normalized_sentence_contains(source: str, cue: str) -> bool:
    normalized_cue = _normalize_sentence_text(cue)
    return bool(normalized_cue and normalized_cue in _normalize_sentence_text(source))


def _normalize_sentence_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(text or ""))
    normalized = normalized.translate(
        str.maketrans(
            {
                "，": ",",
                "。": ".",
                "：": ":",
                "；": ";",
                "！": "!",
                "？": "?",
                "【": "[",
                "】": "]",
                "（": "(",
                "）": ")",
                "～": "~",
            }
        )
    )
    return re.sub(r"\s+", "", normalized).casefold()


def _is_low_value_announcement_term(source: str) -> bool:
    text = str(source or "").strip()
    if not text:
        return True
    if text in _LOW_VALUE_ANNOUNCEMENT_TERMS:
        return True
    if not _CJK_RE.search(text):
        return True
    if len(text) <= 1:
        return True
    if _ANNOUNCEMENT_PLACEHOLDER_RE.search(text):
        return True
    if any(char.isdigit() for char in text) and not any(text.endswith(suffix) for suffix in _TERM_NOUN_SUFFIXES):
        return True
    cjk_len = sum(1 for char in text if "\u3400" <= char <= "\u9fff")
    if cjk_len <= 2 and text.endswith(("级", "次", "个", "点", "日", "月")):
        return True
    if any(mark in text for mark in _ANNOUNCEMENT_SENTENCE_PUNCTUATION) and cjk_len > 4:
        return True
    if any(text.startswith(prefix) for prefix in _LOW_VALUE_ANNOUNCEMENT_PREFIXES) and cjk_len > 4:
        return True
    if any(part in text for part in _LOW_VALUE_ANNOUNCEMENT_SUBSTRINGS) and cjk_len > 5:
        return True
    if cjk_len > 10 and not any(text.endswith(suffix) for suffix in _TERM_NOUN_SUFFIXES):
        return True
    return False


def _term_occurs(source: str, term: str) -> bool:
    start = 0
    while True:
        index = source.find(term, start)
        if index < 0:
            return False
        before = source[index - 1] if index > 0 else ""
        after_index = index + len(term)
        after = source[after_index] if after_index < len(source) else ""
        if term[0].isdigit() and before.isdigit():
            start = index + 1
            continue
        if term[-1].isdigit() and after.isdigit():
            start = index + 1
            continue
        return True
