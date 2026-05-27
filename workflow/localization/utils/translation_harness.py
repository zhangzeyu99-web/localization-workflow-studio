"""Agent-operated full translation harness for English language tables.

The harness does not call any model API. It prepares row-level workpacks for the
main agent, validates the agent-written response, and writes translations back
by ID so the existing QA pipeline can run unchanged.
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from process_language import _load_term_base
from utils.excel_reader import get_text_pairs, read_language_file
from utils.text_normalize import extract_vars, strip_tags_and_vars
from utils.ui_detector import is_ui_text
from utils.ui_length_checker import assess_ui_length, is_short_text_candidate


_CJK_RE = re.compile(r"[\u3400-\u9fff]")
_SENTENCE_PUNCT_RE = re.compile(r"[.!?\u3002\uff01\uff1f\uff1b;]")
_GENERIC_BBCODE_RE = re.compile(
    r"\[/?(?:size|color|b|i|u|s)(?:=[^\]]+)?\]",
    re.IGNORECASE,
)

_SOFT_TERMS = {
    "\u83b7\u5f97",
    "\u6392\u540d",
    "\u664b\u7ea7",
    "\u6982\u7387",
    "\u9009\u62e9",
    "\u4e0d\u8db3",
    "\u7ee7\u7eed",
    "\u52a0\u901f",
    "\u5f3a\u5316",
    "\u7ae0",
}

_SYSTEM_PROMPT_HINTS = (
    "\u8bf7",
    "\u5931\u8d25",
    "\u6210\u529f",
    "\u65e0\u6cd5",
    "\u4e0d\u80fd",
    "\u4e0d\u8db3",
    "\u672a",
    "\u5df2",
    "\u7a0d\u540e",
    "\u9891\u7e41",
    "\u606d\u559c",
    "\u9057\u61be",
)

_RULE_HINTS = (
    "\u53ef",
    "\u6bcf",
    "\u5f53",
    "\u82e5",
    "\u5982\u679c",
    "\u9700\u8981",
    "\u7528\u4e8e",
    "\u5f71\u54cd",
    "\u901a\u8fc7",
    "\u53c2\u4e0e",
    "\u5b8c\u6210",
)

RowId = int | str
SUPPORTED_HARNESS_LANGS = {"en", "ko", "ja"}


@dataclass(frozen=True)
class TargetColumnStatus:
    total_rows: int
    empty_rows: int
    chinese_rows: int
    requires_full_translation: bool
    reason: str


@dataclass(frozen=True)
class PreparedTranslationHarness:
    workpack_path: Path
    manifest_path: Path
    response_template_path: Path
    target_status: TargetColumnStatus
    manifest: dict[str, Any]


@dataclass(frozen=True)
class AppliedTranslationHarness:
    final_workbook_path: Path
    cache_path: Path
    summary_path: Path
    row_count: int


def prepare_translation_harness(
    input_path: str | Path,
    term_base_path: str | Path | None = None,
    lang: str = "en",
    output_dir: str | Path | None = None,
    lang_index: int = 0,
    style_hint: str = "",
) -> PreparedTranslationHarness:
    """Prepare a translation workpack and strict manifest."""
    lang = _require_supported_lang(lang)

    input_path = Path(input_path)
    output_dir = Path(output_dir) if output_dir else input_path.parent / "translation_harness"
    output_dir.mkdir(parents=True, exist_ok=True)
    style_hint = _normalize_style_hint(style_hint)

    df, col_map = read_language_file(str(input_path))
    pairs = get_text_pairs(df, col_map, lang_index=lang_index)
    term_lookup = _load_term_base(str(term_base_path) if term_base_path else None, lang=lang)
    cache = _load_translation_cache(input_path.parent, lang, style_hint=style_hint)

    target_status = analyze_target_column(pairs, lang=lang)
    rows = []
    for _, pair in pairs.iterrows():
        row_id = _coerce_row_id(pair["id"])
        if row_id is None:
            continue
        source = str(pair["original"])
        current_target = _seed_target(source, str(pair["translation"]), target_status, lang=lang)
        cached_translation = cache.get(source, "")
        text_type = classify_text(source, current_target)
        ui_meta = _build_ui_length_meta(row_id, source, current_target, lang=lang)
        rows.append(
            {
                "id": row_id,
                "source": source,
                "current_target": current_target,
                "text_type": text_type,
                "placeholders": extract_vars(source),
                "tags": _extract_tags(source),
                "newline_shape": _newline_shape(source),
                "term_hits": _term_hits(source, term_lookup),
                "ui_length_meta": ui_meta,
                "style_hint": style_hint,
                "cache_hit": bool(cached_translation),
                "cached_translation": cached_translation,
            }
        )

    manifest = {
        "version": 1,
        "language": lang,
        "input_path": str(input_path),
        "input_sha256": _sha256_file(input_path),
        "lang_index": lang_index,
        "row_ids": [row["id"] for row in rows],
        "target_status": target_status.__dict__,
        "style_profile": _build_style_profile(rows, style_hint=style_hint),
        "response_protocol": "jsonl:{id:int|str,translation:str}",
    }

    workpack_path = output_dir / "translation_workpack.jsonl"
    manifest_path = output_dir / "translation_manifest.json"
    response_template_path = output_dir / "translation_response.jsonl"

    _write_jsonl(workpack_path, rows)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_jsonl(
        response_template_path,
        [{"id": row["id"], "translation": row.get("cached_translation", "")} for row in rows],
    )

    return PreparedTranslationHarness(
        workpack_path=workpack_path,
        manifest_path=manifest_path,
        response_template_path=response_template_path,
        target_status=target_status,
        manifest=manifest,
    )


def apply_translation_response(
    input_path: str | Path,
    manifest_path: str | Path,
    response_path: str | Path,
    output_dir: str | Path | None = None,
    lang: str = "en",
) -> AppliedTranslationHarness:
    """Validate a translation response and write a final workbook by ID."""
    input_path = Path(input_path)
    manifest_path = Path(manifest_path)
    response_path = Path(response_path)
    output_dir = Path(output_dir) if output_dir else input_path.parent / "translation_harness"
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest_lang = str(manifest.get("language") or "").strip().lower()
    if lang == "en" and manifest_lang and manifest_lang != "en":
        lang = manifest_lang
    lang = _require_supported_lang(lang)
    if manifest_lang and manifest_lang != lang:
        raise ValueError(f"manifest language {manifest_lang!r} does not match requested language {lang!r}")
    if _sha256_file(input_path) != manifest.get("input_sha256"):
        raise ValueError("input drift detected before applying translation response")

    workpack_path = manifest_path.parent / "translation_workpack.jsonl"
    workpack = _read_jsonl(workpack_path)
    expected_ids = [_coerce_row_id(row["id"]) for row in workpack]
    responses = parse_translation_response(response_path)
    _validate_response_ids(expected_ids, responses)
    _validate_response_surface(workpack, responses)

    final_path = output_dir / f"{input_path.stem}_\u6700\u7ec8\u7248.xlsx"
    shutil.copy2(input_path, final_path)
    _write_translations_to_workbook(
        workbook_path=final_path,
        responses=responses,
        lang_index=int(manifest.get("lang_index", 0)),
    )

    style_profile = manifest.get("style_profile", {})
    style_hint = ""
    if isinstance(style_profile, dict):
        style_hint = _normalize_style_hint(style_profile.get("project_hint", ""))
    cache_path = _update_translation_cache(input_path.parent, lang, workpack, responses, style_hint=style_hint)
    summary_path = output_dir / "translation_apply_summary.json"
    summary = {
        "final_workbook": str(final_path),
        "row_count": len(expected_ids),
        "cache_path": str(cache_path),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return AppliedTranslationHarness(
        final_workbook_path=final_path,
        cache_path=cache_path,
        summary_path=summary_path,
        row_count=len(expected_ids),
    )


def analyze_target_column(pairs, lang: str = "en") -> TargetColumnStatus:
    total = 0
    empty = 0
    chinese = 0
    for _, row in pairs.iterrows():
        source = str(row["original"]).strip()
        target = str(row["translation"]).strip()
        if not source:
            continue
        total += 1
        if not target or target.lower() == "nan":
            empty += 1
        elif _looks_like_source_seed(target, lang):
            chinese += 1

    combined = empty + chinese
    requires = total > 0 and (empty / total >= 0.8 or chinese / total >= 0.5 or combined / total >= 0.5)
    if empty / max(total, 1) >= 0.8:
        reason = "empty_target_column"
    elif chinese / max(total, 1) >= 0.5:
        reason = "chinese_seed_target_column"
    elif combined / max(total, 1) >= 0.5:
        reason = "empty_or_chinese_seed_target_column"
    else:
        reason = "existing_target_translation"

    return TargetColumnStatus(
        total_rows=total,
        empty_rows=empty,
        chinese_rows=chinese,
        requires_full_translation=requires,
        reason=reason,
    )


def classify_text(source: str, current_target: str = "") -> str:
    visible = strip_tags_and_vars(source)
    is_ui, _, _ = is_ui_text(source, current_target or visible)
    if is_short_text_candidate(source, current_target or source) and is_ui:
        return "ui_short"
    if any(hint in source for hint in _SYSTEM_PROMPT_HINTS):
        return "system_prompt"
    if any(hint in source for hint in _RULE_HINTS):
        return "rule_text"
    if len(visible) <= 12 and not _SENTENCE_PUNCT_RE.search(visible):
        return "name_or_item"
    return "description"


def parse_translation_response(response_path: str | Path) -> dict[RowId, str]:
    responses: dict[RowId, str] = {}
    duplicate_ids: set[RowId] = set()
    for raw_line in Path(response_path).read_text(encoding="utf-8").splitlines():
        line = raw_line.strip().lstrip("\ufeff")
        if not line:
            continue
        row_id: RowId | None
        translation: str
        try:
            item = json.loads(line)
            row_id = _coerce_row_id(item["id"])
            translation = str(item["translation"])
        except json.JSONDecodeError:
            match = re.match(r"^(\d+)\s*\|\s*(.+)$", line)
            if not match:
                raise ValueError(f"invalid response line: {line}") from None
            row_id = _coerce_row_id(match.group(1))
            translation = match.group(2)
        if row_id is None:
            raise ValueError(f"invalid response ID: {line}")
        if row_id in responses:
            duplicate_ids.add(row_id)
        responses[row_id] = translation
    if duplicate_ids:
        raise ValueError(f"duplicate response IDs: {sorted(duplicate_ids)}")
    return responses


def _seed_target(source: str, target: str, status: TargetColumnStatus, lang: str = "en") -> str:
    target = str(target or "").strip()
    if not status.requires_full_translation:
        return target
    if not target or target.lower() == "nan" or _looks_like_source_seed(target, lang):
        return source
    return target


def _looks_like_source_seed(target: str, lang: str) -> bool:
    text = str(target or "")
    if not _CJK_RE.search(text):
        return False
    if lang == "ja" and re.search(r"[\u3040-\u30ff]", text):
        return False
    return lang != "ja"


def _build_ui_length_meta(row_id: RowId, source: str, target: str, lang: str = "en") -> dict[str, Any] | None:
    is_ui, _, _ = is_ui_text(source, target)
    assessment = assess_ui_length(row_id, source, target, is_ui=is_ui, lang=lang)
    if not assessment:
        return None
    return {
        "policy": assessment.policy,
        "source_length": assessment.source_length,
        "target_length": assessment.target_length,
        "budget": assessment.budget,
        "reason": assessment.reason,
    }


def _term_hits(source: str, term_lookup: dict[str, dict]) -> list[dict[str, str]]:
    hits = []
    for cn_term, term in sorted(term_lookup.items(), key=lambda item: len(str(item[0])), reverse=True):
        cn = str(cn_term)
        if len(cn) < 2 or cn not in source:
            continue
        primary = str(term.get("primary", "")) if isinstance(term, dict) else str(term)
        variants = term.get("variants", []) if isinstance(term, dict) else []
        strength = "soft" if cn in _SOFT_TERMS else "strong"
        hits.append(
            {
                "source": cn,
                "target": primary,
                "variants": [str(v) for v in variants],
                "strength": strength,
            }
        )
        if len(hits) >= 12:
            break
    return hits


def _build_style_profile(rows: list[dict[str, Any]], style_hint: str = "") -> dict[str, Any]:
    buckets: dict[str, list[RowId]] = {}
    for row in rows:
        buckets.setdefault(str(row["text_type"]), []).append(row["id"])
    samples = {key: value[:5] for key, value in buckets.items()}
    return {
        "project_hint": _normalize_style_hint(style_hint),
        "quality_target": "production_readable",
        "case_policy": "sentence_case_for_status_prompt_text",
        "term_policy": "strong_terms_required_soft_terms_guidance",
        "calibration_samples": samples,
    }


def _extract_tags(text: str) -> list[str]:
    return _GENERIC_BBCODE_RE.findall(str(text))


def _newline_shape(text: str) -> dict[str, int]:
    raw = str(text)
    return {
        "actual": raw.count("\n"),
        "escaped": raw.count("\\n"),
    }


def _validate_response_ids(expected_ids: list[RowId | None], responses: dict[RowId, str]) -> None:
    if any(row_id is None for row_id in expected_ids):
        raise ValueError("manifest contains invalid row IDs")
    actual_ids = list(responses.keys())
    expected_set = set(row_id for row_id in expected_ids if row_id is not None)
    actual_set = set(actual_ids)
    missing = sorted(expected_set - actual_set, key=lambda value: str(value))
    extra = sorted(actual_set - expected_set, key=lambda value: str(value))
    if missing:
        raise ValueError(f"missing response IDs: {missing}")
    if extra:
        raise ValueError(f"extra response IDs: {extra}")
    if actual_ids != expected_ids:
        raise ValueError("response ID order differs from manifest")


def _validate_response_surface(workpack: list[dict[str, Any]], responses: dict[RowId, str]) -> None:
    for row in workpack:
        row_id = _coerce_row_id(row["id"])
        if row_id is None:
            raise ValueError(f"invalid workpack ID: {row.get('id')}")
        translation = responses[row_id]
        if not translation.strip():
            raise ValueError(f"empty translation for ID {row_id}")
        expected_vars = Counter(row.get("placeholders", []))
        actual_vars = Counter(extract_vars(translation))
        if actual_vars != expected_vars:
            raise ValueError(f"placeholder mismatch for ID {row_id}: expected {dict(expected_vars)}, got {dict(actual_vars)}")
        expected_tags = Counter(row.get("tags", []))
        actual_tags = Counter(_extract_tags(translation))
        if actual_tags != expected_tags:
            raise ValueError(f"tag mismatch for ID {row_id}: expected {dict(expected_tags)}, got {dict(actual_tags)}")
        expected_newlines = row.get("newline_shape", {})
        if _newline_shape(translation) != expected_newlines:
            raise ValueError(f"newline mismatch for ID {row_id}")


def _write_translations_to_workbook(workbook_path: Path, responses: dict[RowId, str], lang_index: int) -> None:
    df, col_map = read_language_file(str(workbook_path))
    id_col_name = col_map["id_col"]
    trans_col_name = col_map["languages"][lang_index]["translation_col"]

    wb = load_workbook(workbook_path)
    ws = wb.worksheets[0]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    id_col = headers.index(id_col_name) + 1
    trans_col = headers.index(trans_col_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        row_id = _coerce_row_id(ws.cell(row_idx, id_col).value)
        if row_id in responses:
            ws.cell(row_idx, trans_col).value = responses[row_id]
    wb.save(workbook_path)
    wb.close()


def _load_translation_cache(project_dir: Path, lang: str, style_hint: str = "") -> dict[str, str]:
    cache_path = project_dir / ".translation_cache" / f"{lang}.jsonl"
    if not cache_path.exists():
        return {}
    requested_hint = _normalize_style_hint(style_hint)
    cache: dict[str, str] = {}
    for row in _read_jsonl(cache_path):
        cached_hint = _normalize_style_hint(row.get("style_hint", ""))
        if cached_hint != requested_hint:
            continue
        source = str(row.get("source", ""))
        translation = str(row.get("translation", ""))
        if source and translation:
            cache[source] = translation
    return cache


def _update_translation_cache(
    project_dir: Path,
    lang: str,
    workpack: list[dict[str, Any]],
    responses: dict[RowId, str],
    style_hint: str = "",
) -> Path:
    cache_dir = project_dir / ".translation_cache"
    cache_dir.mkdir(exist_ok=True)
    cache_path = cache_dir / f"{lang}.jsonl"
    style_hint = _normalize_style_hint(style_hint)
    current = {}
    if cache_path.exists():
        for row in _read_jsonl(cache_path):
            cache_key = (str(row.get("source", "")), _normalize_style_hint(row.get("style_hint", "")))
            current[cache_key] = row
    for row in workpack:
        row_id = _coerce_row_id(row["id"])
        if row_id is None:
            continue
        source = str(row["source"])
        current[(source, style_hint)] = {
            "source": source,
            "translation": responses[row_id],
            "text_type": row.get("text_type", ""),
            "lang": lang,
            "style_hint": style_hint,
        }
    _write_jsonl(cache_path, [row for _, row in sorted(current.items())])
    return cache_path


def _normalize_style_hint(style_hint: Any) -> str:
    text = str(style_hint or "").strip()
    return re.sub(r"\s+", " ", text)


def _require_supported_lang(lang: Any) -> str:
    code = str(lang or "en").strip().lower()
    if code not in SUPPORTED_HARNESS_LANGS:
        raise ValueError(f"translation harness supports only {sorted(SUPPORTED_HARNESS_LANGS)}, got {code!r}")
    return code


def _coerce_row_id(value: Any) -> RowId | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else str(value).strip()
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"-?(0|[1-9]\d*)", text):
        return int(text)
    return text


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + ("\n" if rows else ""),
        encoding="utf-8",
    )


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
