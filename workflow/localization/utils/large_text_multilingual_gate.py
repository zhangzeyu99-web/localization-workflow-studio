"""Agent-only large-text multilingual gate (preflight/cache-lint/apply-dry-run/readback-gate).

`backend/app/workflow/large_text.py` is the product-side port of this module's
readback-gate logic. Lint rule changes here (token/number/CJK detection, etc.)
must be mirrored there, and `backend/tests/test_large_text_productization.py`
parity tests must be extended to cover the change. Where the two sides
disagree, this module is authoritative.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TOKEN_RE = re.compile(r"\\n|\{[^{}\s]+\}|%[sdif]|##\d+|</?[A-Za-z][^>\s]*[^>]*>|\[[A-Za-z0-9_:/#=.,-]+\]")
CJK_RE = re.compile(r"[\u3400-\u9fff]")
NUMBER_RE = re.compile(
    r"\d+(?:[,.]\d+)?(?:\s*(?:千|万|萬|亿|億|(?i:thousand|million|billion|ribu|rb|juta|miliar|millones|millón|milhao|milhão|milhões|mil)\b)|[KkMBWw](?![A-Za-z]))%?"
    r"|\d{1,3}(?:[,\s.]\d{3})+(?:[,.]\d+)?%?"
    r"|\d+(?:[,.]\d+)?%?"
)
WORD_MULTIPLIERS = {
    "thousand": Decimal("1000"),
    "ribu": Decimal("1000"),
    "rb": Decimal("1000"),
    "mil": Decimal("1000"),
    "million": Decimal("1000000"),
    "juta": Decimal("1000000"),
    "millones": Decimal("1000000"),
    "millón": Decimal("1000000"),
    "milhao": Decimal("1000000"),
    "milhão": Decimal("1000000"),
    "milhões": Decimal("1000000"),
    "billion": Decimal("1000000000"),
    "miliar": Decimal("1000000000"),
}
NUMBER_WORDS = {
    "zero": Decimal("0"),
    "one": Decimal("1"),
    "once": Decimal("1"),
    "single": Decimal("1"),
    "two": Decimal("2"),
    "three": Decimal("3"),
    "four": Decimal("4"),
    "five": Decimal("5"),
    "six": Decimal("6"),
    "seven": Decimal("7"),
    "eight": Decimal("8"),
    "nine": Decimal("9"),
    "ten": Decimal("10"),
    "satu": Decimal("1"),
    "sekali": Decimal("1"),
    "dua": Decimal("2"),
    "tiga": Decimal("3"),
    "empat": Decimal("4"),
    "lima": Decimal("5"),
    "seis": Decimal("6"),
    "tujuh": Decimal("7"),
    "delapan": Decimal("8"),
    "sembilan": Decimal("9"),
    "sepuluh": Decimal("10"),
    "uno": Decimal("1"),
    "una": Decimal("1"),
    "un": Decimal("1"),
    "dos": Decimal("2"),
    "tres": Decimal("3"),
    "cuatro": Decimal("4"),
    "cinco": Decimal("5"),
    "seis": Decimal("6"),
    "siete": Decimal("7"),
    "ocho": Decimal("8"),
    "nueve": Decimal("9"),
    "diez": Decimal("10"),
    "um": Decimal("1"),
    "uma": Decimal("1"),
    "dois": Decimal("2"),
    "duas": Decimal("2"),
    "quatro": Decimal("4"),
    "sete": Decimal("7"),
    "oito": Decimal("8"),
    "nove": Decimal("9"),
    "dez": Decimal("10"),
}
CJK_ALLOWED_LANGS = {"CN", "ZH", "ZH-CN", "JA", "JP"}
PROCESS_SUFFIXES = {".jsonl", ".log", ".tmp", ".bak", ".manifest"}
SOURCE_HEADERS = {"CN", "ZH", "SOURCE", "TEXT", "原文"}


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line_no, line in enumerate(path.read_text(encoding="utf-8-sig", errors="replace").splitlines(), 1):
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{path}:{line_no} is not valid JSONL: {exc}") from exc
        if not isinstance(value, dict):
            raise ValueError(f"{path}:{line_no} must be a JSON object")
        rows.append(value)
    return rows


def parse_langs(value: str | list[str]) -> list[str]:
    if isinstance(value, list):
        langs = value
    else:
        langs = [part.strip() for part in value.split(",")]
    return [lang.upper() for lang in langs if str(lang).strip()]


def row_key(row: dict[str, Any], fallback: int) -> str:
    return str(row.get("key") or row.get("id") or row.get("para_id") or fallback)


def source_text(row: dict[str, Any]) -> str:
    return str(row.get("cn") or row.get("CN") or row.get("source") or "")


def row_translation(row: dict[str, Any], lang: str) -> str:
    translations = row.get("translations")
    if isinstance(translations, dict):
        return "" if translations.get(lang) is None else str(translations.get(lang))
    return "" if row.get(lang) is None else str(row.get(lang))


def explicit_translation_langs(row: dict[str, Any]) -> set[str]:
    translations = row.get("translations")
    if isinstance(translations, dict):
        return {str(key).upper() for key in translations}
    return set()


def protected_tokens(row: dict[str, Any]) -> list[str]:
    tokens = []
    raw_tokens = row.get("tokens") or row.get("protected_tokens") or []
    if isinstance(raw_tokens, str):
        try:
            parsed = json.loads(raw_tokens)
            raw_tokens = parsed if isinstance(parsed, list) else [raw_tokens]
        except json.JSONDecodeError:
            raw_tokens = [raw_tokens]
    if isinstance(raw_tokens, list):
        tokens.extend(str(token) for token in raw_tokens if str(token))
    tokens.extend(token for token in TOKEN_RE.findall(source_text(row)) if is_auto_protected_token(token))
    return sorted(set(tokens), key=len, reverse=True)


def is_auto_protected_token(token: str) -> bool:
    if token.startswith("[") and token.endswith("]"):
        inner = token[1:-1]
        return bool(re.search(r"[\d_:/#=.,-]", inner) or (inner.isupper() and len(inner) <= 12))
    return True


def parse_number_token(token: str) -> Decimal | None:
    raw_with_spaces = token.strip()
    raw = raw_with_spaces.replace(" ", "")
    if not raw:
        return None

    if raw.endswith("%"):
        raw = raw[:-1]
        raw_with_spaces = raw_with_spaces[:-1].strip()

    suffix = ""
    word_multiplier: Decimal | None = None
    lowered = raw_with_spaces.lower()
    for word, multiplier in sorted(WORD_MULTIPLIERS.items(), key=lambda item: len(item[0]), reverse=True):
        match = re.search(rf"\s+{re.escape(word)}\.?$", lowered)
        if match:
            word_multiplier = multiplier
            raw = raw_with_spaces[: match.start()].strip().replace(" ", "")
            break
    if raw and raw[-1] in "KkMBWw":
        suffix = raw[-1].upper()
        raw = raw[:-1]
    elif raw.endswith(("千", "万", "萬", "亿", "億")):
        suffix = raw[-1]
        raw = raw[:-1]
    if not raw:
        return None

    has_comma = "," in raw
    has_dot = "." in raw
    has_unit = bool(suffix or word_multiplier)
    if has_unit and (has_comma or has_dot):
        if has_comma and has_dot:
            last_comma = raw.rfind(",")
            last_dot = raw.rfind(".")
            last_sep = max(last_comma, last_dot)
            after = raw[last_sep + 1 :]
            before = re.sub(r"[,.]", "", raw[:last_sep])
            raw = before + "." + after
        elif has_comma:
            raw = raw.replace(",", ".")
    elif has_comma or has_dot:
        last_comma = raw.rfind(",")
        last_dot = raw.rfind(".")
        last_sep = max(last_comma, last_dot)
        sep = raw[last_sep]
        after = raw[last_sep + 1 :]
        before = raw[:last_sep]
        thousands_like = len(after) == 3 and all(group.isdigit() for group in re.split(r"[,.]", before) if group)
        if has_comma and has_dot:
            if thousands_like:
                raw = re.sub(r"[,.]", "", raw)
            else:
                raw = re.sub(r"[,.]", "", before) + "." + after
        elif sep == "," and not thousands_like:
            raw = before.replace(",", "") + "." + after
        elif thousands_like:
            raw = re.sub(r"[,.]", "", raw)
        elif sep == ".":
            raw = before.replace(".", "") + "." + after

    try:
        multiplier = word_multiplier or {
            "千": Decimal("1000"),
            "万": Decimal("10000"),
            "萬": Decimal("10000"),
            "亿": Decimal("100000000"),
            "億": Decimal("100000000"),
            "K": Decimal("1000"),
            "M": Decimal("1000000"),
            "B": Decimal("1000000000"),
            "W": Decimal("10000"),
        }.get(suffix, Decimal(1))
        return (Decimal(raw) * multiplier).normalize()
    except InvalidOperation:
        return None


def numeric_values(text: str) -> set[Decimal]:
    text = text or ""
    text = re.sub(r"(\d(?:[\d,.]*))(?:<[^>]+>)+\s*([千万萬亿億KkMBWw])", r"\1\2", text)
    text = re.sub(r"(?<=\d)\uff0c(?=\d{3}(?!\d))", ",", text)
    text = text.replace("\uff0c", " ")
    values = set()
    for token in NUMBER_RE.findall(text):
        parsed = parse_number_token(token)
        if parsed is not None:
            values.add(parsed)
    lowered = text.lower()
    for word, value in NUMBER_WORDS.items():
        if re.search(rf"\b{re.escape(word)}\b", lowered):
            values.add(value)
    return values


def source_numeric_values(row: dict[str, Any]) -> set[Decimal]:
    src = source_text(row)
    src = re.sub(r"\d+(?:[,.]\d+)?\s*月", "", src)
    src = re.sub(r"(?<=\d)\.(?=\d点)", " ", src)
    values = numeric_values(src)
    if CJK_RE.search(src):
        values = {
            value
            for value in values
            if not (
                value == value.to_integral_value()
                and Decimal("0") <= value <= Decimal("10")
                and not re.search(rf"(?<!\d){int(value)}\s*%", src)
            )
        }
    return values


def numeric_value_present(value: Decimal, targets: set[Decimal]) -> bool:
    if value in targets:
        return True
    if abs(value) >= Decimal("1000"):
        for target in targets:
            if target == 0:
                continue
            if abs(target - value) / abs(value) <= Decimal("0.005"):
                return True
    if Decimal("1") <= value <= Decimal("99"):
        if (Decimal("1900") + value) in targets or (Decimal("2000") + value) in targets:
            return True
    return False


def preflight(
    items_jsonl: Path,
    *,
    target_langs: list[str],
    source_rows_jsonl: Path | None = None,
    workbook_count: int = 1,
    full_proofread: bool = False,
) -> dict[str, Any]:
    rows = read_jsonl(items_jsonl)
    unique_keys = {row_key(row, index) for index, row in enumerate(rows, 1)}
    long_text = [row for row in rows if int(row.get("char_len") or len(source_text(row))) > 300]
    source_rows = len(read_jsonl(source_rows_jsonl)) if source_rows_jsonl and source_rows_jsonl.exists() else None
    reasons = []
    if len(unique_keys) > 5000:
        reasons.append("unique_items>5000")
    if len(target_langs) > 4:
        reasons.append("target_languages>4")
    if workbook_count > 1:
        reasons.append("workbook_count>1")
    if full_proofread:
        reasons.append("full_proofread_requested")
    estimated_cells = len(unique_keys) * len(target_langs)
    recommended_shards = 1
    if reasons:
        recommended_shards = max(2, min(8, (estimated_cells // 25000) + 2))
    return {
        "items_jsonl": str(items_jsonl),
        "source_rows_jsonl": str(source_rows_jsonl) if source_rows_jsonl else "",
        "unique_items": len(unique_keys),
        "source_rows": source_rows,
        "target_languages": target_langs,
        "target_language_count": len(target_langs),
        "estimated_target_cells": estimated_cells,
        "long_text_items": len(long_text),
        "workbook_count": workbook_count,
        "large_pack": bool(reasons),
        "large_pack_reasons": reasons,
        "recommended_translation_shards": recommended_shards,
        "recommended_deep_proofread_shards": max(recommended_shards, 4) if full_proofread else recommended_shards,
    }


def add_issue(issues: list[dict[str, Any]], issue_type: str, key: str, lang: str, detail: str) -> None:
    issues.append({"severity": "hard", "type": issue_type, "key": key, "lang": lang, "detail": detail})


def _term_hits(row: dict[str, Any]) -> list[dict[str, Any]]:
    raw = row.get("term_hits") or row.get("term_hits_json") or []
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            return []
        raw = parsed
    if not isinstance(raw, list):
        return []
    return [hit for hit in raw if isinstance(hit, dict)]


def _accepted_term_variants(hit: dict[str, Any], lang: str) -> list[str]:
    values: list[str] = []
    translations = hit.get("translations")
    if isinstance(translations, dict) and translations.get(lang):
        values.append(str(translations[lang]))
    if hit.get(lang):
        values.append(str(hit[lang]))

    variants = hit.get("accepted_variants") or hit.get("variants")
    if isinstance(variants, dict):
        raw = variants.get(lang) or []
    else:
        raw = variants or []
    if isinstance(raw, str):
        values.append(raw)
    elif isinstance(raw, list):
        values.extend(str(value) for value in raw if str(value))

    return sorted({value for value in values if value}, key=len, reverse=True)


def _check_required_terms(issues: list[dict[str, Any]], row: dict[str, Any], key: str, lang: str, target: str) -> None:
    for hit in _term_hits(row):
        if not (hit.get("required") or hit.get("strict")):
            continue
        variants = _accepted_term_variants(hit, lang)
        if variants and not any(variant in target for variant in variants):
            source = hit.get("source") or hit.get("CN") or hit.get("term") or ""
            add_issue(issues, "term_missing", key, lang, f"required term not used: {source}")


def cache_lint(cache_jsonl: Path, *, target_langs: list[str]) -> dict[str, Any]:
    rows = read_jsonl(cache_jsonl)
    issues: list[dict[str, Any]] = []
    seen: set[str] = set()
    unauthorized = Counter()
    requested = set(target_langs)
    for index, row in enumerate(rows, 1):
        key = row_key(row, index)
        if key in seen:
            add_issue(issues, "duplicate_key", key, "", "cache contains duplicate source key")
        seen.add(key)
        extras = explicit_translation_langs(row) - requested
        for lang in sorted(extras):
            unauthorized[lang] += 1
            add_issue(issues, "unauthorized_language", key, lang, "translation cache contains a language that was not requested")

        src = source_text(row)
        src_numbers = source_numeric_values(row)
        tokens = protected_tokens(row)
        for lang in target_langs:
            target = row_translation(row, lang).strip()
            if not target:
                add_issue(issues, "empty_translation", key, lang, "target translation is empty")
                continue
            if lang.upper() not in CJK_ALLOWED_LANGS and CJK_RE.search(target):
                add_issue(issues, "cjk_residue", key, lang, "target translation still contains Chinese/Japanese ideographs")
            for token in tokens:
                if token and token not in target:
                    add_issue(issues, "protected_token_missing", key, lang, f"missing protected token {token}")
            target_numbers = numeric_values(target)
            missing_numbers = {number for number in src_numbers if not numeric_value_present(number, target_numbers)}
            if missing_numbers:
                add_issue(issues, "number_missing", key, lang, f"missing numeric value(s): {sorted(str(value) for value in missing_numbers)}")
            _check_required_terms(issues, row, key, lang, target)

    by_type = Counter(issue["type"] for issue in issues)
    return {
        "cache_jsonl": str(cache_jsonl),
        "checked_items": len(rows),
        "target_languages": target_langs,
        "hard_blockers": len(issues),
        "hard_by_type": dict(sorted(by_type.items())),
        "unauthorized_languages": dict(sorted(unauthorized.items())),
        "issues": issues,
        "ok_to_apply": len(issues) == 0,
    }


def copy_cell_style(src: Any, dst: Any) -> None:
    if not src.has_style:
        return
    dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def apply_dry_run(template_xlsx: Path, output_xlsx: Path) -> dict[str, Any]:
    workbook = load_workbook(template_xlsx)
    try:
        sheet = workbook.active
        source = sheet.cell(row=1, column=min(2, sheet.max_column))
        target = sheet.cell(row=2, column=source.column)
        target.value = "dry-run"
        copy_cell_style(source, target)
        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_xlsx)
    finally:
        workbook.close()
    return {"ok": True, "template": str(template_xlsx), "output": str(output_xlsx), "style_copy": "safe_copy"}


def is_process_file(path: Path) -> bool:
    lowered = path.name.lower()
    return (
        path.name.startswith("_")
        or path.suffix.lower() in PROCESS_SUFFIXES
        or "manifest" in lowered
        or "workpack" in lowered
        or "response" in lowered
    )


def _looks_like_translation_sheet(headers: list[str], target_langs: list[str]) -> bool:
    header_set = {header.upper() for header in headers if header}
    if header_set.intersection(set(target_langs)):
        return True
    return bool(header_set.intersection(SOURCE_HEADERS))


def readback_gate(delivery_dir: Path, *, target_langs: list[str]) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    files = []
    if not delivery_dir.exists():
        add_issue(issues, "delivery_dir_missing", str(delivery_dir), "", "delivery directory does not exist")
        return {"delivery_dir": str(delivery_dir), "exists": False, "hard_blockers": len(issues), "issues": issues, "files": files}

    for path in sorted(p for p in delivery_dir.iterdir() if p.is_file()):
        files.append({"name": path.name, "bytes": path.stat().st_size})
        if is_process_file(path):
            add_issue(issues, "process_file_in_delivery", path.name, "", "delivery directory contains process/log/work files")
        if path.suffix.lower() != ".xlsx":
            continue

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [str(value).strip().upper() if value is not None else "" for value in first_row]
                if not _looks_like_translation_sheet(headers, target_langs):
                    continue
                col_by_lang = {header: index for index, header in enumerate(headers) if header}
                target_columns: list[tuple[str, int]] = []
                for lang in target_langs:
                    col_index = col_by_lang.get(lang.upper())
                    if col_index is None:
                        add_issue(issues, "target_column_missing", f"{path.name}:{sheet.title}", lang, "target language column is missing")
                        continue
                    target_columns.append((lang, col_index))
                if not target_columns:
                    continue
                for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    for lang, col_index in target_columns:
                        value = row_values[col_index] if col_index < len(row_values) else None
                        if value is None or str(value).strip() == "":
                            add_issue(issues, "blank_target_cell", f"{path.name}:{sheet.title}!R{row_index}C{col_index + 1}", lang, "target cell is blank")
        finally:
            workbook.close()

    by_type = Counter(issue["type"] for issue in issues)
    return {
        "delivery_dir": str(delivery_dir),
        "exists": True,
        "file_count": len(files),
        "files": files,
        "target_languages": target_langs,
        "hard_blockers": len(issues),
        "hard_by_type": dict(sorted(by_type.items())),
        "issues": issues,
        "readback_verified": len(issues) == 0,
    }


def write_or_print(result: dict[str, Any], out_path: Path | None, *, quiet: bool = False) -> None:
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    if quiet:
        summary = {key: result.get(key) for key in ["hard_blockers", "checked_items", "unique_items", "estimated_target_cells", "file_count", "ok_to_apply", "readback_verified"] if key in result}
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Preflight and deterministic gates for large multilingual localization packs.")
    sub = parser.add_subparsers(dest="command", required=True)

    pre = sub.add_parser("preflight", help="Estimate task size and required large-pack gates before translation.")
    pre.add_argument("--items-jsonl", required=True, type=Path)
    pre.add_argument("--source-rows-jsonl", type=Path)
    pre.add_argument("--target-langs", required=True)
    pre.add_argument("--workbook-count", type=int, default=1)
    pre.add_argument("--full-proofread", action="store_true")
    pre.add_argument("--out", type=Path)
    pre.add_argument("--quiet", action="store_true")

    lint = sub.add_parser("cache-lint", help="Block workbook/docx writes until translation cache has no hard blockers.")
    lint.add_argument("--cache-jsonl", required=True, type=Path)
    lint.add_argument("--target-langs", required=True)
    lint.add_argument("--out", type=Path)
    lint.add_argument("--quiet", action="store_true")

    dry = sub.add_parser("apply-dry-run", help="Verify workbook style copy/write/save before a large apply.")
    dry.add_argument("--template-xlsx", required=True, type=Path)
    dry.add_argument("--output-xlsx", required=True, type=Path)
    dry.add_argument("--out", type=Path)
    dry.add_argument("--quiet", action="store_true")

    readback = sub.add_parser("readback-gate", help="Verify final delivery directory is clean and target columns are filled.")
    readback.add_argument("--delivery-dir", required=True, type=Path)
    readback.add_argument("--target-langs", required=True)
    readback.add_argument("--out", type=Path)
    readback.add_argument("--quiet", action="store_true")

    args = parser.parse_args(argv)
    if args.command == "preflight":
        result = preflight(
            args.items_jsonl,
            target_langs=parse_langs(args.target_langs),
            source_rows_jsonl=args.source_rows_jsonl,
            workbook_count=args.workbook_count,
            full_proofread=args.full_proofread,
        )
        write_or_print(result, args.out, quiet=args.quiet)
        return 0
    if args.command == "cache-lint":
        result = cache_lint(args.cache_jsonl, target_langs=parse_langs(args.target_langs))
        write_or_print(result, args.out, quiet=args.quiet)
        return 0 if result["hard_blockers"] == 0 else 1
    if args.command == "apply-dry-run":
        result = apply_dry_run(args.template_xlsx, args.output_xlsx)
        write_or_print(result, args.out, quiet=args.quiet)
        return 0
    if args.command == "readback-gate":
        result = readback_gate(args.delivery_dir, target_langs=parse_langs(args.target_langs))
        write_or_print(result, args.out, quiet=args.quiet)
        return 0 if result["hard_blockers"] == 0 else 1
    return 2


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
