from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils.large_text_multilingual_gate import (
    apply_dry_run,
    cache_lint,
    preflight,
    readback_gate,
)


def write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    path.write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n",
        encoding="utf-8",
    )


class LargeTextMultilingualGateTests(unittest.TestCase):
    def test_cache_lint_recomputes_term_hits_and_blocks_missing_strict_term(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            term_base = root / "terms.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN", "分类"])
            sheet.append([2179, "双生魔偶", "Clockwork Twins", "主角"])
            workbook.save(term_base)
            workbook.close()

            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "row-25",
                        "cn": "战斗开始时，为双生魔偶恢复生命",
                        "term_hits": [],
                        "translations": {"EN": "At the start of battle, restores HP to Twin Doll."},
                    }
                ],
            )

            result = cache_lint(cache, target_langs=["EN"], term_base=term_base)

            self.assertEqual(result["hard_by_type"], {"term_hit_snapshot_mismatch": 1, "term_missing": 1})
            self.assertEqual(result["hard_blockers"], 2)

    def test_preflight_flags_large_pack_and_long_text(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            items = Path(tmp) / "items.jsonl"
            write_jsonl(
                items,
                [
                    {"key": "1", "cn": "短文本"},
                    {"key": "2", "cn": "长" * 301},
                ],
            )

            result = preflight(items, target_langs=["DE", "FR", "ES", "PT", "RU"])

            self.assertEqual(result["unique_items"], 2)
            self.assertEqual(result["target_languages"], ["DE", "FR", "ES", "PT", "RU"])
            self.assertEqual(result["estimated_target_cells"], 10)
            self.assertEqual(result["long_text_items"], 1)
            self.assertIs(result["large_pack"], True)
            self.assertIn("target_languages>4", result["large_pack_reasons"])

    def test_preflight_accepts_windows_utf8_bom_jsonl(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            items = Path(tmp) / "items.jsonl"
            items.write_text('\ufeff{"key":"1","cn":"开始"}\n', encoding="utf-8")

            result = preflight(items, target_langs=["EN"])

            self.assertEqual(result["unique_items"], 1)
            self.assertEqual(result["target_languages"], ["EN"])

    def test_cache_lint_blocks_missing_translation_cjk_and_token_loss(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "ok",
                        "cn": "获得{num}金币",
                        "tokens": ["{num}"],
                        "translations": {"DE": "{num} Gold erhalten"},
                    },
                    {
                        "key": "bad",
                        "cn": "消耗{num}钻石和20点体力",
                        "tokens": ["{num}"],
                        "translations": {"DE": "消耗 Diamanten"},
                    },
                    {"key": "missing", "cn": "开始战斗", "translations": {}},
                ],
            )

            result = cache_lint(cache, target_langs=["DE"])

            self.assertEqual(result["hard_blockers"], 4)
            issue_types = {issue["type"] for issue in result["issues"]}
            self.assertLessEqual({"cjk_residue", "protected_token_missing", "number_missing", "empty_translation"}, issue_types)

    def test_cache_lint_blocks_missing_and_extra_angle_at_placeholders(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "missing",
                        "cn": "造成<@1>%攻击力伤害",
                        "translations": {"EN": "Deals ATK DMG"},
                    },
                    {
                        "key": "extra",
                        "cn": "造成<@1>%攻击力伤害",
                        "translations": {"EN": "Deals <@1>% ATK DMG <@2>"},
                    },
                ],
            )

            result = cache_lint(cache, target_langs=["EN"])

            self.assertEqual(result["hard_blockers"], 2, result["issues"])
            issue_types = {issue["type"] for issue in result["issues"]}
            self.assertEqual(issue_types, {"protected_token_missing", "protected_token_extra"})

    def test_cache_lint_accepts_equivalent_wan_number(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "number",
                        "cn": "战力达到66.9万",
                        "translations": {"ES": "Alcanza 669,000 de poder"},
                    }
                ],
            )

            result = cache_lint(cache, target_langs=["ES"])

            self.assertEqual(result["hard_blockers"], 0)

    def test_cache_lint_accepts_russian_compact_number_units(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {"key": "thousand", "cn": "获得32万金币", "translations": {"RU": "Получено 320 тыс. монет"}},
                    {"key": "million", "cn": "战力8138万", "translations": {"RU": "Боевая мощь 81,38 млн"}},
                    {"key": "billion", "cn": "累计10亿", "translations": {"RU": "Всего 1 млрд"}},
                ],
            )

            result = cache_lint(cache, target_langs=["RU"])

            self.assertEqual(result["hard_blockers"], 0)

    def test_cache_lint_uses_hyphenated_reference_number_equivalence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "shop",
                        "cn": "811便利店",
                        "reference_en": "8-11 Mart",
                        "translations": {"RU": "Магазин 8-11"},
                    }
                ],
            )

            result = cache_lint(cache, target_langs=["RU"])

            self.assertEqual(result["hard_blockers"], 0)

    def test_cache_lint_preserves_explicit_opaque_payloads(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "opaque-ok",
                        "cn": "_x0001_损坏载荷???",
                        "opaque_payload_preserved": True,
                        "translations": {"RU": "_x0001_损坏载荷???"},
                    },
                    {
                        "key": "opaque-bad",
                        "cn": "_x0002_损坏载荷???",
                        "opaque_payload_preserved": True,
                        "translations": {"RU": "изменено"},
                    },
                ],
            )

            result = cache_lint(cache, target_langs=["RU"])

            self.assertEqual(result["hard_blockers"], 1)
            self.assertEqual(result["issues"][0]["type"], "opaque_payload_changed")

    def test_cache_lint_blocks_mojibake_outside_opaque_payloads(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "mojibake",
                        "cn": "成功占领城市",
                        "translations": {"RU": "успеш��о занимает город"},
                    }
                ],
            )

            result = cache_lint(cache, target_langs=["RU"])

            self.assertEqual(result["hard_blockers"], 1)
            self.assertEqual(result["issues"][0]["type"], "mojibake")

    def test_cache_lint_accepts_game_number_formats_without_false_positives(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "rich-unit",
                        "cn": "123.1<font=GameFont_SDF>\u4e07",
                        "translations": {
                            "EN": "1.231<font=GameFont_SDF>M</font>",
                            "IDN": "1.231M<font=GameFont_SDF>",
                            "ES": "1.231M<font=GameFont_SDF>",
                            "PT": "1.231M<font=GameFont_SDF>",
                            "RU": "1,231<font=GameFont_SDF>млн",
                        },
                    },
                    {
                        "key": "w-unit",
                        "cn": "10w\u8d44\u6e90\u968f\u673a\u7bb1",
                        "translations": {
                            "EN": "100K Resource Random Box",
                            "IDN": "Kotak Acak 100K Sumber Daya",
                            "ES": "Caja aleatoria de 100K Recursos",
                            "PT": "Caixa aleatoria de 100K Recursos",
                            "RU": "Случайный ящик ресурсов, 100 тыс.",
                        },
                    },
                    {
                        "key": "time-units",
                        "cn": "\u4e0b\u6b21\u5237\u65b0\uff1a24\u65f645\u520616\u79d2",
                        "translations": {
                            "EN": "Next Refresh: 24h 45m 16s",
                            "IDN": "Segarkan berikutnya: 24 j 45 m 16 dtk",
                            "ES": "Proxima actualizacion: 24 h 45 min 16 s",
                            "PT": "Proxima atualizacao: 24 h 45 min 16 s",
                            "RU": "Следующее обновление: 24 ч 45 мин 16 с",
                        },
                    },
                    {
                        "key": "month-name",
                        "cn": "\u6839\u636e2019\u5e7410\u670815\u65e5\u53d1\u5e03",
                        "translations": {
                            "EN": "issued on October 15, 2019",
                            "IDN": "diterbitkan pada 15 Oktober 2019",
                            "ES": "emitida el 15 de octubre de 2019",
                            "PT": "emitida em 15 de outubro de 2019",
                            "RU": "опубликовано 15 октября 2019 г.",
                        },
                    },
                ],
            )

            result = cache_lint(cache, target_langs=["EN", "IDN", "ES", "PT", "RU"])

            self.assertEqual(result["hard_blockers"], 0)

    def test_cache_lint_only_auto_protects_machine_like_bracket_tokens(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            cache = Path(tmp) / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "translatable-label",
                        "source": "Defeat [Monster] for [Name]",
                        "translations": {"ES": "Derrota [Monstruo] para [Nombre]"},
                    },
                    {
                        "key": "machine-token",
                        "source": "Share [SDT] {num}",
                        "translations": {"ES": "Comparte {num}"},
                    },
                ],
            )

            result = cache_lint(cache, target_langs=["ES"])

            self.assertEqual(result["hard_blockers"], 1)
            self.assertEqual(result["issues"][0]["type"], "protected_token_missing")
            self.assertIn("[SDT]", result["issues"][0]["detail"])

    def test_apply_dry_run_uses_safe_style_copy(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            template = root / "template.xlsx"
            output = root / "dry_run.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "source"
            sheet["B1"] = "target"
            sheet["B1"].style = "Headline 1"
            workbook.save(template)

            result = apply_dry_run(template, output)

            self.assertIs(result["ok"], True)
            self.assertTrue(output.exists())
            copied = load_workbook(output)
            try:
                self.assertEqual(copied.active["B2"].value, "dry-run")
            finally:
                copied.close()

    def test_readback_gate_rejects_process_files_and_blank_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            delivery = Path(tmp) / "delivery"
            delivery.mkdir()
            (delivery / "_work.jsonl").write_text("{}", encoding="utf-8")
            workbook_path = delivery / "final.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "DE"])
            sheet.append([1, "开始", "Starten"])
            sheet.append([2, "结束", None])
            workbook.save(workbook_path)

            result = readback_gate(delivery, target_langs=["DE"])

            self.assertEqual(result["hard_blockers"], 2)
            issue_types = {issue["type"] for issue in result["issues"]}
            self.assertLessEqual({"process_file_in_delivery", "blank_target_cell"}, issue_types)

    def test_readback_gate_skips_qa_summary_support_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            delivery = Path(tmp) / "delivery"
            delivery.mkdir()
            final_path = delivery / "final.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN", "IDN"])
            sheet.append([1, "领取", "Claim", "Klaim"])
            workbook.save(final_path)

            qa_path = delivery / "QA摘要.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "LineReview"
            sheet.append(["source_file", "row", "id", "CN", "status"])
            sheet.append(["final.xlsx", 2, 1, "领取", "KEEP"])
            workbook.save(qa_path)

            result = readback_gate(delivery, target_langs=["EN", "IDN"])

            self.assertEqual(result["hard_blockers"], 0, result["issues"])
            self.assertTrue(result["readback_verified"])

    def test_readback_gate_skips_trailing_styled_rows_without_id_or_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            delivery = Path(tmp) / "delivery"
            delivery.mkdir()
            final_path = delivery / "final.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN", "FR"])
            sheet.append([1, "领取", "Claim", "Récupérer"])
            sheet["C100"].style = "Headline 1"
            sheet["D100"].style = "Headline 1"
            workbook.save(final_path)

            result = readback_gate(delivery, target_langs=["EN", "FR"])

            self.assertEqual(result["hard_blockers"], 0, result["issues"])
            self.assertTrue(result["readback_verified"])


if __name__ == "__main__":
    unittest.main()
