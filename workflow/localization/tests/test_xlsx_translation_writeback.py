from __future__ import annotations

import json
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from utils.xlsx_translation_writeback import verify_translation_cache, write_translation_workbooks


def write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    path.write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n",
        encoding="utf-8",
    )


class XlsxTranslationWritebackTests(unittest.TestCase):
    def test_writeback_handles_underreported_worksheet_dimension(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append([1, "Attack", None])
            workbook.save(source)
            workbook.close()
            with zipfile.ZipFile(source, "r") as archive:
                entries = {name: archive.read(name) for name in archive.namelist()}
                infos = archive.infolist()
            xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
            xml = xml.replace('<dimension ref="A1:C2"/>', '<dimension ref="A1"/>')
            entries["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
            with zipfile.ZipFile(source, "w") as archive:
                for info in infos:
                    archive.writestr(info, entries[info.filename])
            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "source.xlsx::Sheet::1::2",
                        "id": "1",
                        "source_file": "source.xlsx",
                        "sheet": "Sheet",
                        "row": 2,
                        "cn": "Attack",
                        "translations": {"EN": "Attack"},
                    }
                ],
            )

            write_translation_workbooks(
                inputs=[source],
                cache_jsonl=cache,
                target_langs=["EN"],
                output_dir=root / "delivery",
            )

            saved = load_workbook(root / "delivery" / "source.xlsx", read_only=True)
            try:
                self.assertEqual(saved["Sheet"]["C2"].value, "Attack")
            finally:
                saved.close()

    def test_writeback_preserves_worksheet_namespace_extensions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append([1, "Attack", None])
            workbook.save(source)
            workbook.close()
            with zipfile.ZipFile(source, "r") as archive:
                entries = {name: archive.read(name) for name in archive.namelist()}
                infos = archive.infolist()
            xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
            xml = xml.replace(
                "<worksheet ",
                '<worksheet xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
                'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
                'mc:Ignorable="x14" ',
                1,
            ).replace(
                "</worksheet>",
                '<extLst><ext uri="test"><x14:test/></ext></extLst></worksheet>',
            )
            entries["xl/worksheets/sheet1.xml"] = xml.encode("utf-8")
            with zipfile.ZipFile(source, "w") as archive:
                for info in infos:
                    archive.writestr(info, entries[info.filename])
            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [{"key": "source.xlsx::Sheet::1::2", "id": "1", "source_file": "source.xlsx", "sheet": "Sheet", "row": 2, "cn": "Attack", "translations": {"EN": "Attack"}}],
            )

            write_translation_workbooks(
                inputs=[source],
                cache_jsonl=cache,
                target_langs=["EN"],
                output_dir=root / "delivery",
            )

            with zipfile.ZipFile(root / "delivery" / "source.xlsx", "r") as archive:
                saved_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
            self.assertIn("xmlns:x14=", saved_xml)
            self.assertIn('mc:Ignorable="x14"', saved_xml)
            self.assertIn("<x14:test", saved_xml)

    def test_writeback_only_changes_requested_target_cells_and_opens_normally(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Language"
            sheet.append(["ID", "CN", "Note", "EN", "DE"])
            sheet.append([1001, "Attack", "keep", None, None])
            sheet["C2"].fill = PatternFill("solid", fgColor="00FF00")
            sheet["D2"].fill = PatternFill("solid", fgColor="FFFF00")
            workbook.save(source)
            workbook.close()

            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "source.xlsx::Language::1001::2",
                        "id": "1001",
                        "source_file": "source.xlsx",
                        "sheet": "Language",
                        "row": 2,
                        "cn": "Attack",
                        "translations": {"EN": "Attack", "DE": "Angriff"},
                    }
                ],
            )

            result = write_translation_workbooks(
                inputs=[source],
                cache_jsonl=cache,
                target_langs=["EN", "DE"],
                output_dir=root / "delivery",
            )

            self.assertEqual(result.written_cells, 2)
            self.assertEqual(result.invalid_style_refs, 0)
            self.assertEqual(result.output_files, 1)
            output = root / "delivery" / "source.xlsx"
            saved = load_workbook(output)
            try:
                sheet = saved["Language"]
                self.assertEqual(sheet["A2"].value, 1001)
                self.assertEqual(sheet["B2"].value, "Attack")
                self.assertEqual(sheet["C2"].value, "keep")
                self.assertEqual(sheet["D2"].value, "Attack")
                self.assertEqual(sheet["E2"].value, "Angriff")
                self.assertEqual(sheet["C2"].fill.fgColor.rgb, "0000FF00")
                self.assertEqual(sheet["D2"].fill.fgColor.rgb, "00FFFF00")
            finally:
                saved.close()

    def test_writeback_rejects_source_text_drift(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append([1, "Attack", None])
            workbook.save(source)
            workbook.close()
            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "source.xlsx::Sheet::1::2",
                        "id": "1",
                        "source_file": "source.xlsx",
                        "sheet": "Sheet",
                        "row": 2,
                        "cn": "Changed source",
                        "translations": {"EN": "Attack"},
                    }
                ],
            )

            with self.assertRaisesRegex(ValueError, "source text drift"):
                write_translation_workbooks(
                    inputs=[source],
                    cache_jsonl=cache,
                    target_langs=["EN"],
                    output_dir=root / "delivery",
                )

    def test_writeback_preserves_numeric_zero_source_text(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append([1, 0, None])
            workbook.save(source)
            workbook.close()
            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [
                    {
                        "key": "source.xlsx::Sheet::1::2",
                        "id": "1",
                        "source_file": "source.xlsx",
                        "sheet": "Sheet",
                        "row": 2,
                        "cn": "0",
                        "translations": {"EN": "0"},
                    }
                ],
            )

            write_translation_workbooks(
                inputs=[source],
                cache_jsonl=cache,
                target_langs=["EN"],
                output_dir=root / "delivery",
            )

            saved = load_workbook(root / "delivery" / "source.xlsx", read_only=True)
            try:
                self.assertEqual(saved["Sheet"]["C2"].value, "0")
            finally:
                saved.close()
            verification = verify_translation_cache(root / "delivery", cache, ["EN"])
            self.assertEqual(verification["hard_blockers"], 0)

    def test_writeback_rejects_repeated_source_row_with_wrong_id(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append([1, "Attack", None])
            sheet.append([2, "Attack", None])
            workbook.save(source)
            workbook.close()
            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [{"key": "source.xlsx::Sheet::1::2", "id": "1", "source_file": "source.xlsx", "sheet": "Sheet", "row": 3, "cn": "Attack", "translations": {"EN": "First"}}],
            )

            with self.assertRaisesRegex(ValueError, "source row id drift"):
                write_translation_workbooks(
                    inputs=[source],
                    cache_jsonl=cache,
                    target_langs=["EN"],
                    output_dir=root / "delivery",
                )

    def test_cache_readback_detects_nonblank_but_wrong_translation(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            delivery = root / "delivery"
            delivery.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append([1, "Attack", "Wrong but nonblank"])
            workbook.save(delivery / "source.xlsx")
            workbook.close()
            cache = root / "cache.jsonl"
            write_jsonl(
                cache,
                [{"key": "source.xlsx::Sheet::1::2", "id": "1", "source_file": "source.xlsx", "sheet": "Sheet", "row": 2, "cn": "Attack", "translations": {"EN": "Attack"}}],
            )

            result = verify_translation_cache(delivery, cache, ["EN"])

            self.assertEqual(result["hard_blockers"], 1)
            self.assertEqual(result["issues"][0]["type"], "translation_value_mismatch")


if __name__ == "__main__":
    unittest.main()
