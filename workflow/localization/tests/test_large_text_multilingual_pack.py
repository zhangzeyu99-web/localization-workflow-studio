from __future__ import annotations

import json
import re
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook

from utils.large_text_multilingual_pack import _load_terms, prepare_pack, stable_row_key


LANGS = ["EN", "IDN", "DE", "FR", "ES", "PT", "RU", "IT", "TR", "TH"]


def create_workbook(path: Path, start_id: int, rows: int, unique_count: int = 145) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "UI" if "UI" in path.stem else "Language"
    sheet.append(["ID", "CN", *LANGS])
    for offset in range(rows):
        sheet.append([start_id + offset, f"文本{offset % unique_count}", *([None] * len(LANGS))])
    workbook.save(path)
    workbook.close()


def create_reference_workbook(path: Path, *, missing_reference: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Language"
    sheet.append(["ID", "CN", "EN", "FR", "DE"])
    sheet.append([1, "烈焰斩", "Flame Strike", None, None])
    sheet.append([2, "太阳神殿", None if missing_reference else "Temple of the Sun", None, None])
    workbook.save(path)
    workbook.close()


class LargeTextMultilingualPackTests(unittest.TestCase):
    def test_load_terms_ignores_stale_xlsx_dimension_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "terms.xlsx"
            rewritten = Path(tmp) / "rewritten.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN", "分类"])
            sheet.append([2179, "双生魔偶", "Clockwork Twins", "主角"])
            workbook.save(path)
            workbook.close()
            with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(rewritten, "w") as target:
                for info in source.infolist():
                    payload = source.read(info.filename)
                    if info.filename == "xl/worksheets/sheet1.xml":
                        payload = re.sub(rb'<dimension ref="[^"]+"', b'<dimension ref="A1"', payload, count=1)
                    target.writestr(info, payload)
            terms = _load_terms(rewritten, ["EN"])

            self.assertEqual([term["source"] for term in terms], ["双生魔偶"])

    def test_load_terms_marks_main_character_names_as_required(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "terms.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN", "FR", "DE", "PT", "ES", "TR", "RU", "备注", "分类"])
            sheet.append(
                [
                    2179,
                    "双生魔偶",
                    "Clockwork Twins",
                    "Jumeaux Mécaniques",
                    "Uhrwerk-Zwillinge",
                    "Gêmeas Mecânicas",
                    "Gemelos Mecánicos",
                    "Kurmalı İkizler",
                    "Заводные близнецы",
                    None,
                    "主角",
                ]
            )
            workbook.save(path)
            workbook.close()

            terms = _load_terms(path, ["EN", "FR", "DE", "PT", "ES", "TR", "RU"])

            self.assertEqual(len(terms), 1)
            self.assertEqual(terms[0]["source"], "双生魔偶")
            self.assertEqual(terms[0]["category"], "主角")
            self.assertIs(terms[0]["required"], True)
            self.assertIs(terms[0]["strict"], True)

    def test_stable_row_key_includes_file_sheet_id_and_row(self) -> None:
        self.assertEqual(
            stable_row_key("a.xlsx", "Sheet1", 7, 1001),
            "a.xlsx::Sheet1::1001::7",
        )

    def test_prepare_pack_extracts_rows_and_deduplicates_unique_texts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first = root / "7.13新增.xlsx"
            second = root / "7.13UI新增.xlsx"
            create_workbook(first, 1000, 245)
            create_workbook(second, 2000, 22)

            result = prepare_pack(
                inputs=[first, second],
                term_base=None,
                history_dirs=[],
                target_langs=LANGS,
                work_dir=root / "work",
            )

            self.assertEqual(result.source_rows, 267)
            self.assertEqual(result.unique_items, 145)
            self.assertEqual(result.estimated_target_cells, 2670)
            self.assertLess(result.elapsed_seconds, 5.0)
            self.assertTrue(result.items_jsonl.exists())
            self.assertTrue(result.source_rows_jsonl.exists())
            rows = [json.loads(line) for line in result.items_jsonl.read_text(encoding="utf-8").splitlines()]
            self.assertEqual(len(rows), 267)
            self.assertEqual(rows[0]["context"], "language")
            self.assertEqual(rows[-1]["context"], "ui")
            self.assertEqual(set(rows[0]["translations"]), set())

    def test_prepare_pack_rejects_nonempty_target_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            path = root / "input.xlsx"
            create_workbook(path, 1, 1)
            from openpyxl import load_workbook

            workbook = load_workbook(path)
            workbook.active["C2"] = "Existing"
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "target column is not empty"):
                prepare_pack(
                    inputs=[path],
                    term_base=None,
                    history_dirs=[],
                    target_langs=LANGS,
                    work_dir=root / "work",
                )

    def test_prepare_pack_adds_english_reference_for_non_english_targets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            path = root / "input.xlsx"
            create_reference_workbook(path)

            result = prepare_pack(
                inputs=[path],
                term_base=None,
                history_dirs=[],
                target_langs=["FR", "DE"],
                work_dir=root / "work",
                source_mode="cn+en",
            )
            rows = [
                json.loads(line)
                for line in result.items_jsonl.read_text(encoding="utf-8").splitlines()
            ]
            stats = json.loads(result.prepare_stats.read_text(encoding="utf-8"))

            self.assertEqual(rows[0]["source_mode"], "cn+en")
            self.assertEqual(rows[0]["translation_source"], "烈焰斩")
            self.assertEqual(rows[0]["reference_en"], "Flame Strike")
            self.assertEqual(rows[0]["reference_en_status"], "usable")
            self.assertEqual(stats["english_reference_status"]["usable_rows"], 2)

    def test_prepare_pack_en_mode_rejects_incomplete_reference(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            path = root / "input.xlsx"
            create_reference_workbook(path, missing_reference=True)

            with self.assertRaisesRegex(ValueError, "complete usable English"):
                prepare_pack(
                    inputs=[path],
                    term_base=None,
                    history_dirs=[],
                    target_langs=["FR", "DE"],
                    work_dir=root / "work",
                    source_mode="en",
                )


if __name__ == "__main__":
    unittest.main()
