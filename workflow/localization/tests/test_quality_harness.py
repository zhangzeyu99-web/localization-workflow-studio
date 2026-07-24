import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from utils.quality_harness import load_fixture, run_fixture, scan_workbook


class QualityHarnessTests(unittest.TestCase):
    def test_fixture_detects_known_bad_and_good_cases(self):
        fixture = {
            "cases": [
                {
                    "id": "bad-title-case",
                    "source": "该帐号创角过多",
                    "translation": "Too Many Roles",
                    "expected_issues": ["title_case_overuse"],
                },
                {
                    "id": "bad-internal-token",
                    "source": "调试残留",
                    "translation": "ZXN37Q",
                    "expected_issues": ["internal_token_leak"],
                },
                {
                    "id": "good-feature-title",
                    "source": "七日登录",
                    "translation": "7-Day Login",
                    "expected_issues": [],
                },
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)
        self.assertEqual(result.total_cases, 3)
        self.assertEqual(result.issue_counts["title_case_overuse"], 1)
        self.assertEqual(result.issue_counts["internal_token_leak"], 1)

    def test_fixture_reports_expectation_mismatch(self):
        fixture = {
            "cases": [
                {
                    "id": "bad-clipped-word",
                    "source": "登录超时",
                    "translation": "Logi time",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertFalse(result.passed)
        self.assertEqual(result.failures[0]["id"], "bad-clipped-word")
        self.assertIn("clipped_word", result.failures[0]["actual_issues"])

    def test_load_fixture_reads_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "fixture.json"
            path.write_text(json.dumps({"cases": []}), encoding="utf-8")

            self.assertEqual(load_fixture(path), {"cases": []})

    def test_scan_workbook_reports_hard_issues(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "系统错误", "System Error"])
            ws.append([2, "战令", "Battle Pass"])
            wb.save(path)

            result = scan_workbook(path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["title_case_overuse"], 1)
            self.assertEqual(result.rows_scanned, 2)
            self.assertEqual(result.issues[0]["row"], 2)

    def test_scan_workbook_flags_numbered_term_inconsistency(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "numbered_terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "消灭怪物-1", "Kill Monsters-1"])
            ws.append([2, "消灭怪物-2", "Destroy Monsters-2"])
            ws.append([3, "消灭怪物-3", "Kill Monsters -3"])
            ws.append([4, "搜集物资-1", "Collect Supplies-1"])
            ws.append([5, "搜集物资-2", "Collect Supplies-2"])
            ws.append([6, "搜集物资-3", "Collect Supplies-3"])
            wb.save(path)

            result = scan_workbook(path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["numbered_term_inconsistency"], 2)
            self.assertEqual(
                [issue["id"] for issue in result.issues if issue["check_type"] == "numbered_term_inconsistency"],
                [2, 3],
            )

    def test_scan_workbook_uses_first_good_numbered_term_not_majority(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "numbered_terms_first_good.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "消灭怪物-1", "Kill Monsters-1"])
            ws.append([2, "消灭怪物-2", "Destroy Monsters-2"])
            ws.append([3, "消灭怪物-3", "Destroy Monsters-3"])
            wb.save(path)

            result = scan_workbook(path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["numbered_term_inconsistency"], 2)
            self.assertEqual(
                [issue["id"] for issue in result.issues if issue["check_type"] == "numbered_term_inconsistency"],
                [2, 3],
            )

    def test_scan_workbook_prefers_term_base_for_numbered_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "numbered_terms_with_base.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "消灭怪物-1", "Destroy Monsters-1"])
            ws.append([2, "消灭怪物-2", "Kill Monsters-2"])
            ws.append([3, "消灭怪物-3", "Kill Monsters-3"])
            wb.save(path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["消灭怪物", "Kill Monsters", None, "任务"])
            wb.save(term_path)

            result = scan_workbook(path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["numbered_term_inconsistency"], 1)
            self.assertEqual(
                [issue["id"] for issue in result.issues if issue["check_type"] == "numbered_term_inconsistency"],
                [1],
            )

    def test_scan_workbook_enforces_glossary_term_for_standalone_ui_label(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "战机", "Fighter"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["战机", "Warplane", None, "系统名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_missing"], 1)
            self.assertEqual(result.issues[0]["check_type"], "term_missing")

    def test_scan_workbook_hard_blocks_partial_multiword_term_hit(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "菇勇者传说", "Mushroom Legend"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["菇勇者传说", "Legend of Mushroom", None, "游戏名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_partial_hit"], 1)
            self.assertEqual(result.issues[0]["check_type"], "term_partial_hit")
            self.assertEqual(result.issues[0]["severity"], "error")

    def test_scan_workbook_treats_game_name_category_as_exact(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "菇勇者传说", "Legend of Mushrooms"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["菇勇者传说", "Legend of Mushroom", None, "游戏名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_partial_hit"], 1)
            self.assertEqual(result.issues[0]["severity"], "error")

    def test_scan_workbook_preserves_json_exact_match_metadata(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "菇勇者传说", "Legend of Mushrooms"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.json"
            term_path.write_text(
                json.dumps(
                    {
                        "lookup": {
                            "菇勇者传说": {
                                "primary": "Legend of Mushroom",
                                "variants": [],
                                "exact_match": True,
                            }
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_partial_hit"], 1)
            self.assertEqual(result.issues[0]["severity"], "error")

    def test_scan_workbook_accepts_pluralized_common_multiword_term(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "联盟成员", "Alliance Members"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["联盟成员", "Alliance Member", None, "系统/组织"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_missing"], 0)
            self.assertEqual(result.issue_counts["term_partial_hit"], 0)

    def test_scan_workbook_keeps_explicit_soft_terms_non_blocking(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "领取奖励", "Get rewards"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["领取", "Claim", None, "泛词"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_soft_missing"], 1)
            self.assertEqual(result.issues, [])

    def test_explicit_term_base_suppresses_nearby_auto_discovered_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "ABC message", "Message"])
            wb.save(workbook_path)

            explicit_path = Path(tmp) / "selected_terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "glossary"
            ws.append(["CN", "EN", "EN2", "category"])
            ws.append(["ABC", "Canonical", None, "soft"])
            wb.save(explicit_path)

            stale_path = Path(tmp) / "stale_glossary.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "glossary"
            ws.append(["CN", "EN", "EN2", "category"])
            ws.append(["ABC", "Canonical", None, "hard"])
            wb.save(stale_path)

            result = scan_workbook(workbook_path, term_base=explicit_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_soft_missing"], 1)
            self.assertEqual(result.issue_counts["term_missing"], 0)

    def test_scan_workbook_skips_game_terms_for_long_legal_text(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "legal.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            source = "四三九九游戏隐私政策" + (
                "我们的合作伙伴将按照用户协议处理信息。" * 80
            )
            ws.append([1, source, "Our partners process information under the user agreement."])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["伙伴", "Companion", None, "强术语"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_missing"], 0)

    def test_scan_workbook_auto_softens_uncategorized_generic_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "获得奖励成功", "Rewards delivered"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "获得", "Obtain"])
            ws.append([2, "成功", "Success"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_soft_missing"], 2)
            self.assertEqual(result.issues, [])

    def test_scan_workbook_does_not_force_subterm_into_natural_phrase(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "战机升级", "Fighter upgrade"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "战机", "Warplane"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_missing"], 0)

    def test_scan_workbook_enforces_explicit_contextual_strong_term(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "战机升级", "Fighter upgrade"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["战机", "Warplane", None, "强术语"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_missing"], 1)

    def test_scan_workbook_applies_ui_length_gate(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "ui_length.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "领取奖励", "Claim all rewards now immediately"])
            ws.append([2, "一段提示文本", "This is a readable short note that keeps expanding"])
            wb.save(path)

            result = scan_workbook(path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["ui_length_overflow"], 1)
            self.assertEqual(result.issue_counts["short_text_length_watch"], 1)
            self.assertEqual(
                [issue["id"] for issue in result.issues if issue["check_type"] == "ui_length_overflow"],
                [1],
            )

    def test_scan_workbook_uses_non_read_only_and_fails_empty_scan(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "unrecognized.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Only", "Notes"])
            ws.append(["x", "not a language table"])
            wb.save(path)

            with patch("utils.quality_harness.load_workbook", wraps=load_workbook) as mocked:
                result = scan_workbook(path)

            self.assertFalse(mocked.call_args.kwargs["read_only"])
            self.assertFalse(result.passed)
            self.assertEqual(result.rows_scanned, 0)
            self.assertEqual(result.issue_counts["workbook_scan_empty"], 1)
            self.assertEqual(result.issues[0]["check_type"], "workbook_scan_empty")

    def test_scan_workbook_uses_japanese_column_and_allows_kanji(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "ja_language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "JA"])
            ws.append([1, "领取奖励", "報酬を受け取る"])
            wb.save(path)

            result = scan_workbook(path, lang="ja")

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)
            self.assertEqual(result.issue_counts["chinese_residue"], 0)

    def test_scan_workbook_uses_idn_column_instead_of_id_primary_key(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "idn_language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "IDN"])
            ws.append([1, "领取", "Klaim"])
            wb.save(path)

            result = scan_workbook(path, lang="idn", auto_discover_terms=False)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)
            self.assertEqual(result.issue_counts["workbook_scan_empty"], 0)

    def test_scan_workbook_uses_korean_term_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "ko_language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "KO"])
            ws.append([1, "战机", "Fighter"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "KO", "KO2", "分类"])
            ws.append(["战机", "전투기", "", "系统名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, lang="ko", term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_missing"], 1)

    def test_scan_workbook_accepts_jp_and_kr_header_aliases(self):
        with tempfile.TemporaryDirectory() as tmp:
            ja_path = Path(tmp) / "ja_language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "Notes", "JP"])
            ws.append([1, "领取奖励", 123, "報酬を受け取る"])
            wb.save(ja_path)

            ja_result = scan_workbook(ja_path, lang="ja")

            self.assertTrue(ja_result.passed, ja_result.issues)
            self.assertEqual(ja_result.rows_scanned, 1)

            ko_path = Path(tmp) / "ko_language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "Notes", "KR"])
            ws.append([1, "战机", 123, "Fighter"])
            wb.save(ko_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Lookup"
            ws.append(["CN", "KR", "KR2", "分类"])
            ws.append(["战机", "전투기", "", "系统名"])
            wb.save(term_path)

            ko_result = scan_workbook(ko_path, lang="ko", term_base=term_path)

            self.assertEqual(ko_result.rows_scanned, 1)
            self.assertFalse(ko_result.passed)
            self.assertEqual(ko_result.issue_counts["term_missing"], 1)

    def test_scan_workbook_skips_glossary_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "with_glossary.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["未加入军团", "Not in a Legion", None, "社交/军团"])
            ws.append(["次数不足", "Not Enough Attempts", None, "通用提示"])

            ws = wb.create_sheet("语言表5.8")
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "系统错误", "System error"])
            wb.save(path)

            result = scan_workbook(path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)
            self.assertEqual(result.issue_counts["title_case_overuse"], 0)

    def test_scan_workbook_enforces_embedded_person_name_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "with_person_names.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["艾莉娅", "Aria", None, "角色/人名"])
            ws.append(["里昂", "Leon", None, "角色/人名"])
            ws.append(["丹尼尔", "Daniel", None, "角色/人名"])
            ws.append(["尼尔", "Neil", None, "角色/人名"])
            ws.append(["布雷兹", "Blaze", None, "角色"])

            ws = wb.create_sheet("语言表")
            ws.append(["ID", "CN", "EN"])
            ws.append([1480, "艾莉娅的勋章", "Arya's Medal"])
            ws.append([1482, "里昂的勋章", "Lyon's Medal"])
            ws.append([1849, "丹尼尔", "Daniel"])
            ws.append([1481, "布雷兹的勋章", "Blaise's Medal"])
            wb.save(path)

            result = scan_workbook(path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["person_name_term_mismatch"], 3)
            self.assertEqual([issue["id"] for issue in result.issues], [1480, 1482, 1481])

    def test_scan_workbook_accepts_external_person_name_term_base(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "艾莉娅的勋章", "Aria's Medal"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["艾莉娅", "Aria", None, "角色/人名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)

    def test_scan_workbook_auto_discovers_same_dir_person_name_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = root / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "艾莉娅的勋章", "Arya's Medal"])
            wb.save(workbook_path)

            term_path = root / "项目术语表.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["艾莉娅", "Aria", None, "角色/人名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["person_name_term_mismatch"], 1)

    def test_scan_workbook_auto_discovers_parent_terms_from_output_dir(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "output_en_final"
            output_dir.mkdir()
            workbook_path = output_dir / "result_en.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "里昂的勋章", "Lyon's Medal"])
            wb.save(workbook_path)

            term_path = root / "项目术语表.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["里昂", "Leon", None, "角色/人名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["person_name_term_mismatch"], 1)

    def test_scan_workbook_skips_support_audit_sheets(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "with_audit_sheet.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["艾莉娅", "Aria", None, "角色/人名"])

            ws = wb.create_sheet("语言表")
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "艾莉娅的勋章", "Aria's Medal"])

            ws = wb.create_sheet("研发意见裁决")
            ws.append(["Sheet", "Row", "ID", "CN", "原EN", "研发意见", "最终EN", "裁决"])
            ws.append(["语言表", 2, 1, "艾莉娅的勋章", "Arya's Medal", "人名相关术语统一问题", "Aria's Medal", "fix"])
            wb.save(workbook_path)

            result = scan_workbook(workbook_path, term_base=workbook_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)

    def test_scan_workbook_skips_result_review_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "result_en.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "完整结果"
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "系统错误", "System error"])

            ws = wb.create_sheet("需确认")
            ws.append(["ID", "CN", "EN"])
            ws.append([2, "系统错误", "System Error"])
            wb.save(workbook_path)

            result = scan_workbook(workbook_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)

    def test_runtime_placeholder_sentence_does_not_trigger_leading_lowercase(self):
        fixture = {
            "cases": [
                {
                    "id": "runtime-message",
                    "source": "<color=#457B9F>##1</color>已同意<color=#457B9F>##2</color>加入军团！",
                    "translation": "<color=#457B9F>##1</color> approved <color=#457B9F>##2</color> to join the Legion!",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_placeholder_only_literal_newline_does_not_trigger_leading_lowercase(self):
        fixture = {
            "cases": [
                {
                    "id": "placeholder-only-newline",
                    "source": "[{0}]\\n{2}",
                    "translation": "[{0}]\\n{2}",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_rich_text_macro_display_text_does_not_trigger_variable_drift(self):
        fixture = {
            "cases": [
                {
                    "id": "rich-text-display",
                    "source": "#{item_blue,{0,PlayerName}}手气爆棚，在#{item_blue,{1,PlayerName}}的红包中豪夺#{luck_red,运气王}！",
                    "translation": "#{item_blue,{0,PlayerName}} is extremely lucky and won #{luck_red,Luck King} from #{item_blue,{1,PlayerName}}'s red packet!",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_link_macro_display_text_does_not_trigger_variable_drift(self):
        fixture = {
            "cases": [
                {
                    "id": "link-macro-display",
                    "source": "#L{linkType:Url;url:www.baidu.com,item_red,www.baidu.com 点击前往 }",
                    "translation": "#L{linkType:Url;url:www.baidu.com,item_red,www.baidu.com Click to go }",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_number_unit_prefix_does_not_trigger_leading_lowercase(self):
        fixture = {
            "cases": [
                {
                    "id": "voice-limit",
                    "source": "限30秒语音\\n松手发送\\n划开取消发送",
                    "translation": "30s Voice Limit\\nRelease to Send\\nSlide Away to Cancel",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_bbcode_size_prefix_does_not_trigger_leading_lowercase(self):
        fixture = {
            "cases": [
                {
                    "id": "bbcode-size-prefix",
                    "source": "[size=80][c0]\u81ea\u9009\u4f20\u8bf4\u6280\u80fd[s0][/size]",
                    "translation": "[size=80][c0]Select Legendary Skill[s0][/size]",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_color_hex_tag_does_not_trigger_hash_code_abbreviation(self):
        fixture = {
            "cases": [
                {
                    "id": "color-hex-tag",
                    "source": "<color=#FFFFFF>\u7b7e\u5230\u9001</color>",
                    "translation": "Day <color=#FFFFFF>Sign-in Gift</color>",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_bbcode_color_hex_tag_does_not_trigger_hash_code_abbreviation(self):
        fixture = {
            "cases": [
                {
                    "id": "bbcode-color-hex-tag",
                    "source": "[color=#FFFFFF]\u7b7e\u5230\u9001[/color]",
                    "translation": "Day [color=#FFFFFF]Sign-in Gift[/color]",
                    "expected_issues": [],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_html_color_codes_must_match(self):
        fixture = {
            "cases": [
                {
                    "id": "html-color-mismatch",
                    "source": "<color=#608F01>##1</color>\u5956\u52b1",
                    "translation": "<color=#FF0000>##1</color> rewards",
                    "expected_issues": ["bbcode_color_mismatch"],
                }
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_sandwiched_question_mark_separator_is_punctuation_corruption(self):
        fixture = {
            "cases": [
                {
                    "id": "bad-middle-dot-downgraded-to-question",
                    "source": "重装·普攻I",
                    "translation": "Tank ? Basic Attack I",
                    "expected_issues": ["punctuation_corruption"],
                },
                {
                    "id": "good-natural-question-in-question-source",
                    "source": "开飞机？试试看嘛！",
                    "translation": "Flying a Warplane? Let's try!",
                    "expected_issues": [],
                },
                {
                    "id": "good-question-key-prompt",
                    "source": "帮助键",
                    "translation": "Press ? for help",
                    "expected_issues": [],
                },
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_hash_code_and_placeholder_compaction_are_hard_issues(self):
        fixture = {
            "cases": [
                {
                    "id": "blacklist-limit-code",
                    "source": "黑名单数量已达上限",
                    "translation": "#BRUL",
                    "expected_issues": ["hash_code_abbreviation"],
                },
                {
                    "id": "spend-item-code",
                    "source": "花费##1个##2",
                    "translation": "S##1##2",
                    "expected_issues": ["placeholder_compaction"],
                },
                {
                    "id": "employed-glue-code",
                    "source": "##1上岗##2位居民",
                    "translation": "##1Employed##2",
                    "expected_issues": ["placeholder_word_glue"],
                },
            ]
        }

        result = run_fixture(fixture)

        self.assertTrue(result.passed, result.failures)

    def test_scan_workbook_uses_requested_target_language_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN", "TH", "VI", "ID"])
            ws.append([1, "领取奖励", "领取奖励", "รับรางวัล", "Nhận thưởng", "Klaim Hadiah"])
            wb.save(workbook_path)

            result = scan_workbook(workbook_path, lang="vi", auto_discover_terms=False)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.rows_scanned, 1)

    def test_scan_workbook_uses_requested_language_term_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "language.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "VI"])
            ws.append([1, "求生之路", "Con Đường Sinh Tồn"])
            wb.save(workbook_path)

            term_path = Path(tmp) / "terms.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "VI"])
            ws.append(["求生之路", "Survival Road", "Con Đường Sinh Tồn"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, lang="vi", term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_missing"], 0)


if __name__ == "__main__":
    unittest.main()
