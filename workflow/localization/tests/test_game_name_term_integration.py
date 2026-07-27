import json
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from process_language import _load_term_base
from utils.quality_harness import scan_workbook
from utils.term_checker import check_term_hit


class GameNameTermIntegrationTests(unittest.TestCase):
    def test_term_loader_preserves_game_name_category_with_case_constraint(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "terms.xlsx"
            pd.DataFrame(
                {
                    "CN": ["菇勇者传说"],
                    "EN": ["Legend of Mushroom"],
                    "大小写约束": ["是"],
                    "分类": ["游戏名"],
                }
            ).to_excel(path, index=False)

            term_lookup = _load_term_base(str(path), lang="en")
            results = check_term_hit(
                row_id=1,
                original="菇勇者传说",
                translation="Legend of Mushrooms",
                term_lookup=term_lookup,
            )

            self.assertEqual(term_lookup["菇勇者传说"]["category"], "游戏名")
            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].severity, "error")

    def test_quality_harness_hard_blocks_partial_game_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "language.xlsx"
            term_path = tmp_path / "terms.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "菇勇者传说", "Mushroom Legend"])
            wb.save(workbook_path)

            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["菇勇者传说", "Legend of Mushroom", None, "游戏名"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_partial_hit"], 1)
            self.assertEqual(result.issues[0]["severity"], "error")

    def test_quality_harness_preserves_json_exact_match_metadata(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "language.xlsx"
            term_path = tmp_path / "terms.json"

            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "菇勇者传说", "Legend of Mushrooms"])
            wb.save(workbook_path)

            term_path.write_text(
                json.dumps(
                    {
                        "lookup": {
                            "菇勇者传说": {
                                "primary": "Legend of Mushroom",
                                "variants": [],
                                "exact_match": True,
                            }
                        }
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertFalse(result.passed)
            self.assertEqual(result.issue_counts["term_partial_hit"], 1)
            self.assertEqual(result.issues[0]["severity"], "error")

    def test_quality_harness_accepts_pluralized_common_multiword_term(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "language.xlsx"
            term_path = tmp_path / "terms.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "CN", "EN"])
            ws.append([1, "联盟成员", "Alliance Members"])
            wb.save(workbook_path)

            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            ws.append(["CN", "EN", "EN2", "分类"])
            ws.append(["联盟成员", "Alliance Member", None, "系统/组织"])
            wb.save(term_path)

            result = scan_workbook(workbook_path, term_base=term_path)

            self.assertTrue(result.passed, result.issues)
            self.assertEqual(result.issue_counts["term_missing"], 0)
            self.assertEqual(result.issue_counts["term_partial_hit"], 0)


if __name__ == "__main__":
    unittest.main()
