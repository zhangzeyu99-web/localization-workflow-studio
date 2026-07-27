import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils.translation_harness import (
    apply_translation_response,
    prepare_translation_harness,
)


SRC_CLAIM = "\u9886\u53d6\u5956\u52b1"
SRC_SURVIVAL = "\u7d2f\u8ba1\u53c2\u4e0e{0}\u6b21\u6c42\u751f\u4e4b\u8def"
SRC_RICH = "[size=80][c0]\u81ea\u9009\u4f20\u8bf4\u6280\u80fd[s0][/size]"
SRC_SKILL = "\u7ec8\u7109\u4e4b\u5883"
SRC_LOCATION = "\u592a\u9633\u795e\u6bbf"


def _write_language_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Cn", "En"])
    ws.append([1, SRC_CLAIM, ""])
    ws.append([2, SRC_SURVIVAL, SRC_SURVIVAL])
    ws.append([3, SRC_RICH, ""])
    wb.save(path)


def _write_term_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "\u672f\u8bed\u8868"
    ws.append(["ID", "CN", "EN", "EN2", "\u5206\u7c7b"])
    ws.append([1, SRC_CLAIM, "Claim Reward", "", "UI\u64cd\u4f5c\u52a8\u8bcd"])
    ws.append([2, "\u6c42\u751f\u4e4b\u8def", "Survival Road", "", "\u4efb\u52a1/\u6d3b\u52a8/\u73a9\u6cd5"])
    ws.append([3, "\u4f20\u8bf4\u6280\u80fd", "Legendary Skill", "", "\u9053\u5177/\u88c5\u5907/\u793c\u5305"])
    wb.save(path)


def _write_name_policy_language_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Cn", "En"])
    ws.append([4, SRC_SKILL, ""])
    ws.append([5, SRC_LOCATION, ""])
    wb.save(path)


def _write_name_policy_term_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "\u672f\u8bed\u8868"
    ws.append(["ID", "CN", "EN", "EN2", "\u5206\u7c7b"])
    ws.append([4, SRC_SKILL, "Final Realm", "", "\u6280\u80fd\u540d"])
    ws.append([5, SRC_LOCATION, "Temple of the Sun", "", "\u5730\u70b9\u540d"])
    wb.save(path)


def _write_english_reference_language_workbook(path: Path, *, missing_second_reference: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "CN", "EN", "FR"])
    ws.append([10, "\u70c8\u7130\u65a9", "Flame Strike", ""])
    ws.append([11, "\u592a\u9633\u795e\u6bbf", "" if missing_second_reference else "Temple of the Sun", ""])
    wb.save(path)


def _write_english_reference_term_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "\u672f\u8bed\u8868"
    ws.append(["CN", "EN", "FR", "\u5206\u7c7b"])
    ws.append(["\u70c8\u7130\u65a9", "Flame Strike", "Frappe ardente", "\u6280\u80fd\u540d"])
    ws.append(["\u592a\u9633\u795e\u6bbf", "Temple of the Sun", "Temple du Soleil", "\u5730\u70b9\u540d"])
    wb.save(path)


def _write_korean_language_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "KO"])
    ws.append(["btn.claim", "\u9886\u53d6\u5956\u52b1", ""])
    ws.append(["msg.welcome", "\u6b22\u8fce\u56de\u6765\uff0c{playerName}", ""])
    wb.save(path)


def _write_multilang_language_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "CN", "EN", "TH", "VI", "ID"])
    ws.append([1, SRC_CLAIM, "", "", "", ""])
    ws.append([2, SRC_SURVIVAL, SRC_SURVIVAL, SRC_SURVIVAL, SRC_SURVIVAL, SRC_SURVIVAL])
    wb.save(path)


def _write_multilang_term_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "\u672f\u8bed\u8868"
    ws.append(["CN", "EN", "TH", "VI", "ID"])
    ws.append([SRC_CLAIM, "Claim Reward", "รับรางวัล", "Nhận thưởng", "Klaim Hadiah"])
    ws.append(["\u6c42\u751f\u4e4b\u8def", "Survival Road", "เส้นทางเอาชีวิตรอด", "Con Đường Sinh Tồn", "Jalan Bertahan Hidup"])
    wb.save(path)


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n",
        encoding="utf-8",
    )


class TranslationHarnessTests(unittest.TestCase):
    def test_translation_harness_has_no_external_mt_client(self):
        source_path = Path(__file__).resolve().parents[1] / "utils" / "translation_harness.py"
        source = source_path.read_text(encoding="utf-8").lower()
        forbidden_markers = [
            "deep_translator",
            "googletranslator",
            "googletrans",
            "translate.googleapis",
            "translate_a/single",
            "requests.",
            "httpx.",
            "aiohttp",
        ]

        for marker in forbidden_markers:
            with self.subTest(marker=marker):
                self.assertNotIn(marker, source)

    def test_prepare_builds_workpack_for_empty_or_chinese_target_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            term_path = tmp_path / "terms.xlsx"
            out_dir = tmp_path / "out"
            _write_language_workbook(lang_path)
            _write_term_workbook(term_path)

            result = prepare_translation_harness(
                input_path=lang_path,
                term_base_path=term_path,
                lang="en",
                output_dir=out_dir,
            )

            self.assertTrue(result.target_status.requires_full_translation)
            self.assertEqual(result.manifest["language"], "en")
            self.assertEqual(result.manifest["row_ids"], [1, 2, 3])
            self.assertTrue(result.workpack_path.exists())
            self.assertTrue(result.manifest_path.exists())

            rows = [
                json.loads(line)
                for line in result.workpack_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual(len(rows), 3)
            self.assertEqual(rows[0]["current_target"], SRC_CLAIM)
            self.assertEqual(rows[0]["text_type"], "ui_short")
            self.assertEqual(rows[1]["placeholders"], ["{0}"])
            self.assertIn("Survival Road", [term["target"] for term in rows[1]["term_hits"]])
            self.assertIn("[size=80]", rows[2]["tags"])

    def test_prepare_adds_explicit_skill_and_location_name_policy(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            term_path = tmp_path / "terms.xlsx"
            _write_name_policy_language_workbook(lang_path)
            _write_name_policy_term_workbook(term_path)

            result = prepare_translation_harness(
                input_path=lang_path,
                term_base_path=term_path,
                lang="en",
                output_dir=tmp_path / "out",
            )
            rows = [
                json.loads(line)
                for line in result.workpack_path.read_text(encoding="utf-8").splitlines()
            ]

            self.assertEqual(rows[0]["text_type"], "ui_skill_name")
            self.assertEqual(rows[0]["name_type"], "ui_skill_name")
            self.assertEqual(rows[0]["name_policy"]["preferred_words"], 2)
            self.assertEqual(rows[1]["text_type"], "ui_location_name")
            self.assertEqual(rows[1]["name_policy"]["preferred_content_words"], 2)
            self.assertEqual(rows[0]["term_hits"][0]["category"], "\u6280\u80fd\u540d")

    def test_prepare_cn_plus_en_keeps_chinese_primary_and_adds_english_reference(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            term_path = tmp_path / "terms.xlsx"
            _write_english_reference_language_workbook(lang_path)
            _write_english_reference_term_workbook(term_path)

            prepared = prepare_translation_harness(
                input_path=lang_path,
                term_base_path=term_path,
                lang="fr",
                output_dir=tmp_path / "out",
                source_mode="cn+en",
            )
            rows = [
                json.loads(line)
                for line in prepared.workpack_path.read_text(encoding="utf-8").splitlines()
            ]

            self.assertEqual(prepared.manifest["source_mode"], "cn+en")
            self.assertEqual(prepared.manifest["english_reference_status"]["usable_rows"], 2)
            self.assertEqual(rows[0]["source"], "\u70c8\u7130\u65a9")
            self.assertEqual(rows[0]["translation_source"], "\u70c8\u7130\u65a9")
            self.assertEqual(rows[0]["reference_en"], "Flame Strike")
            self.assertEqual(rows[0]["reference_en_status"], "usable")
            self.assertEqual(rows[0]["source_mode"], "cn+en")

    def test_prepare_en_mode_uses_english_as_translation_source_but_keeps_cn_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            term_path = tmp_path / "terms.xlsx"
            _write_english_reference_language_workbook(lang_path)
            _write_english_reference_term_workbook(term_path)

            prepared = prepare_translation_harness(
                input_path=lang_path,
                term_base_path=term_path,
                lang="fr",
                output_dir=tmp_path / "out",
                source_mode="en",
            )
            rows = [
                json.loads(line)
                for line in prepared.workpack_path.read_text(encoding="utf-8").splitlines()
            ]

            self.assertEqual(rows[0]["translation_source"], "Flame Strike")
            self.assertEqual(rows[0]["reference_en"], "Flame Strike")
            self.assertEqual(rows[0]["source_cn"], "\u70c8\u7130\u65a9")
            self.assertIn("Frappe ardente", [term["target"] for term in rows[0]["term_hits"]])

    def test_prepare_en_mode_rejects_incomplete_english_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            _write_english_reference_language_workbook(lang_path, missing_second_reference=True)

            with self.assertRaisesRegex(ValueError, "complete usable English"):
                prepare_translation_harness(
                    input_path=lang_path,
                    lang="fr",
                    output_dir=tmp_path / "out",
                    source_mode="en",
                )

    def test_prepare_cn_plus_en_marks_missing_reference_and_falls_back_to_chinese(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            _write_english_reference_language_workbook(lang_path, missing_second_reference=True)

            prepared = prepare_translation_harness(
                input_path=lang_path,
                lang="fr",
                output_dir=tmp_path / "out",
                source_mode="cn+en",
            )
            rows = [
                json.loads(line)
                for line in prepared.workpack_path.read_text(encoding="utf-8").splitlines()
            ]

            self.assertEqual(prepared.manifest["english_reference_status"]["usable_rows"], 1)
            self.assertEqual(prepared.manifest["english_reference_status"]["empty_rows"], 1)
            self.assertEqual(rows[1]["reference_en"], "")
            self.assertEqual(rows[1]["reference_en_status"], "missing")
            self.assertEqual(rows[1]["translation_source"], "\u592a\u9633\u795e\u6bbf")

    def test_translation_cache_is_isolated_by_source_mode(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            _write_english_reference_language_workbook(lang_path)
            out_dir = tmp_path / "out"

            prepared_cn = prepare_translation_harness(
                input_path=lang_path,
                lang="fr",
                output_dir=out_dir,
                source_mode="cn",
            )
            response_path = tmp_path / "translation_response.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": 10, "translation": "Frappe ardente"},
                    {"id": 11, "translation": "Temple du Soleil"},
                ],
            )
            apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared_cn.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
                lang="fr",
            )

            prepared_ref = prepare_translation_harness(
                input_path=lang_path,
                lang="fr",
                output_dir=out_dir,
                source_mode="cn+en",
            )
            rows = [
                json.loads(line)
                for line in prepared_ref.workpack_path.read_text(encoding="utf-8").splitlines()
            ]

            self.assertFalse(rows[0]["cache_hit"])
            self.assertFalse(rows[1]["cache_hit"])

    def test_translation_cache_is_invalidated_when_english_reference_changes(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            _write_english_reference_language_workbook(lang_path)
            out_dir = tmp_path / "out"

            prepared = prepare_translation_harness(
                input_path=lang_path,
                lang="fr",
                output_dir=out_dir,
                source_mode="cn+en",
            )
            response_path = tmp_path / "translation_response.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": 10, "translation": "Frappe ardente"},
                    {"id": 11, "translation": "Temple du Soleil"},
                ],
            )
            apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
                lang="fr",
            )

            wb = load_workbook(lang_path)
            wb.active["C2"] = "Blaze Slash"
            wb.save(lang_path)
            wb.close()

            changed_reference = prepare_translation_harness(
                input_path=lang_path,
                lang="fr",
                output_dir=out_dir,
                source_mode="cn+en",
            )
            rows = [
                json.loads(line)
                for line in changed_reference.workpack_path.read_text(encoding="utf-8").splitlines()
            ]

            self.assertFalse(rows[0]["cache_hit"])
            self.assertTrue(rows[1]["cache_hit"])

    def test_prepare_supports_thai_vietnamese_and_indonesian_targets(self):
        expected_terms = {
            "th": "เส้นทางเอาชีวิตรอด",
            "vi": "Con Đường Sinh Tồn",
            "idn": "Jalan Bertahan Hidup",
        }
        lang_indexes = {"th": 1, "vi": 2, "idn": 3}

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            term_path = tmp_path / "terms.xlsx"
            _write_multilang_language_workbook(lang_path)
            _write_multilang_term_workbook(term_path)

            for lang, expected_term in expected_terms.items():
                with self.subTest(lang=lang):
                    out_dir = tmp_path / f"out_{lang}"
                    prepared = prepare_translation_harness(
                        input_path=lang_path,
                        term_base_path=term_path,
                        lang=lang,
                        output_dir=out_dir,
                        lang_index=lang_indexes[lang],
                    )
                    rows = [
                        json.loads(line)
                        for line in prepared.workpack_path.read_text(encoding="utf-8").splitlines()
                    ]

                    self.assertEqual(prepared.manifest["language"], lang)
                    self.assertIn(expected_term, [term["target"] for term in rows[1]["term_hits"]])

    def test_apply_writes_non_english_target_column_and_language_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            term_path = tmp_path / "terms.xlsx"
            out_dir = tmp_path / "out_th"
            _write_multilang_language_workbook(lang_path)
            _write_multilang_term_workbook(term_path)
            prepared = prepare_translation_harness(
                input_path=lang_path,
                term_base_path=term_path,
                lang="th",
                output_dir=out_dir,
                lang_index=1,
            )
            response_path = tmp_path / "translation_response_th.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": 1, "translation": "รับรางวัล"},
                    {"id": 2, "translation": "เข้าร่วมเส้นทางเอาชีวิตรอด {0} ครั้ง"},
                ],
            )

            applied = apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
                lang="th",
            )

            wb = load_workbook(applied.final_workbook_path, read_only=True, data_only=False)
            try:
                ws = wb.active
                self.assertEqual(ws.cell(2, 4).value, "รับรางวัล")
                self.assertEqual(ws.cell(3, 4).value, "เข้าร่วมเส้นทางเอาชีวิตรอด {0} ครั้ง")
                self.assertIsNone(ws.cell(2, 3).value)
            finally:
                wb.close()
            self.assertTrue((tmp_path / ".translation_cache" / "th.jsonl").exists())

    def test_prepare_includes_project_style_hint_and_scopes_cache_by_hint(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            out_dir = tmp_path / "out"
            _write_language_workbook(lang_path)
            style_hint = "US mobile SLG; concise, idiomatic UI wording"

            prepared = prepare_translation_harness(
                input_path=lang_path,
                lang="en",
                output_dir=out_dir,
                style_hint=style_hint,
            )

            self.assertEqual(prepared.manifest["style_profile"]["project_hint"], style_hint)
            rows = [
                json.loads(line)
                for line in prepared.workpack_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual(rows[0]["style_hint"], style_hint)

            response_path = tmp_path / "translation_response.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": 1, "translation": "Claim"},
                    {"id": 2, "translation": "Do Survival Road {0} times"},
                    {"id": 3, "translation": "[size=80][c0]Pick Legendary Skill[s0][/size]"},
                ],
            )
            apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
            )

            same_hint = prepare_translation_harness(
                input_path=lang_path,
                lang="en",
                output_dir=out_dir,
                style_hint=style_hint,
            )
            same_rows = [
                json.loads(line)
                for line in same_hint.workpack_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertTrue(same_rows[0]["cache_hit"])
            self.assertEqual(same_rows[0]["cached_translation"], "Claim")

            different_hint = prepare_translation_harness(
                input_path=lang_path,
                lang="en",
                output_dir=out_dir,
                style_hint="UK PC strategy; formal wording",
            )
            different_rows = [
                json.loads(line)
                for line in different_hint.workpack_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertFalse(different_rows[0]["cache_hit"])
            self.assertEqual(different_rows[0]["cached_translation"], "")

    def test_apply_rejects_incomplete_duplicate_extra_or_placeholder_drift(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            out_dir = tmp_path / "out"
            _write_language_workbook(lang_path)
            prepared = prepare_translation_harness(lang_path, output_dir=out_dir)

            cases = [
                ([{"id": 1, "translation": "Claim Reward"}], "missing"),
                (
                    [
                        {"id": 1, "translation": "Claim Reward"},
                        {"id": 1, "translation": "Claim Reward Again"},
                        {"id": 2, "translation": "Join {0} Survival Road"},
                        {"id": 3, "translation": "[size=80][c0]Select Legendary Skill[s0][/size]"},
                    ],
                    "duplicate",
                ),
                (
                    [
                        {"id": 1, "translation": "Claim Reward"},
                        {"id": 2, "translation": "Join {0} Survival Road"},
                        {"id": 3, "translation": "[size=80][c0]Select Legendary Skill[s0][/size]"},
                        {"id": 99, "translation": "Extra"},
                    ],
                    "extra",
                ),
                (
                    [
                        {"id": 2, "translation": "Join {0} Survival Road"},
                        {"id": 1, "translation": "Claim Reward"},
                        {"id": 3, "translation": "[size=80][c0]Select Legendary Skill[s0][/size]"},
                    ],
                    "order",
                ),
                (
                    [
                        {"id": 1, "translation": "Claim Reward"},
                        {"id": 2, "translation": "Join Survival Road"},
                        {"id": 3, "translation": "[size=80][c0]Select Legendary Skill[s0][/size]"},
                    ],
                    "placeholder",
                ),
            ]

            for rows, expected in cases:
                response_path = tmp_path / f"{expected}.jsonl"
                _write_jsonl(response_path, rows)
                with self.assertRaisesRegex(ValueError, expected):
                    apply_translation_response(
                        input_path=lang_path,
                        manifest_path=prepared.manifest_path,
                        response_path=response_path,
                        output_dir=out_dir,
                    )

    def test_apply_rejects_input_drift_between_prepare_and_apply(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            out_dir = tmp_path / "out"
            _write_language_workbook(lang_path)
            prepared = prepare_translation_harness(lang_path, output_dir=out_dir)

            response_path = tmp_path / "translation_response.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": 1, "translation": "Claim Reward"},
                    {"id": 2, "translation": "Join Survival Road {0} times"},
                    {"id": 3, "translation": "[size=80][c0]Select Legendary Skill[s0][/size]"},
                ],
            )

            wb = load_workbook(lang_path)
            ws = wb.active
            ws.cell(2, 2).value = "\u9886\u53d6\u5956\u52b1!"
            wb.save(lang_path)
            wb.close()

            with self.assertRaisesRegex(ValueError, "input drift"):
                apply_translation_response(
                    input_path=lang_path,
                    manifest_path=prepared.manifest_path,
                    response_path=response_path,
                    output_dir=out_dir,
                )

    def test_apply_writes_final_workbook_and_same_project_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            out_dir = tmp_path / "out"
            _write_language_workbook(lang_path)
            prepared = prepare_translation_harness(lang_path, output_dir=out_dir)

            response_path = tmp_path / "translation_response.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": 1, "translation": "Claim Reward"},
                    {"id": 2, "translation": "Join Survival Road {0} times"},
                    {"id": 3, "translation": "[size=80][c0]Select Legendary Skill[s0][/size]"},
                ],
            )

            applied = apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
            )

            wb = load_workbook(applied.final_workbook_path, read_only=True, data_only=False)
            ws = wb.active
            self.assertEqual(ws.cell(2, 3).value, "Claim Reward")
            self.assertEqual(ws.cell(3, 3).value, "Join Survival Road {0} times")
            self.assertEqual(ws.cell(4, 3).value, "[size=80][c0]Select Legendary Skill[s0][/size]")
            wb.close()

            cache_path = tmp_path / ".translation_cache" / "en.jsonl"
            self.assertTrue(cache_path.exists())
            cache_rows = [json.loads(line) for line in cache_path.read_text(encoding="utf-8").splitlines()]
            self.assertIn("Claim Reward", [row["translation"] for row in cache_rows])

            prepared_again = prepare_translation_harness(lang_path, output_dir=out_dir)
            rows = [
                json.loads(line)
                for line in prepared_again.workpack_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual(rows[0]["cached_translation"], "Claim Reward")
            self.assertTrue(rows[0]["cache_hit"])

    def test_apply_accepts_utf8_bom_response_written_by_windows_tools(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            out_dir = tmp_path / "out"
            _write_language_workbook(lang_path)
            prepared = prepare_translation_harness(lang_path, output_dir=out_dir)

            response_path = tmp_path / "translation_response.jsonl"
            response_path.write_text(
                "\ufeff"
                + "\n".join(
                    [
                        json.dumps({"id": 1, "translation": "Claim Reward"}, ensure_ascii=False),
                        json.dumps({"id": 2, "translation": "Join Survival Road {0} times"}, ensure_ascii=False),
                        json.dumps(
                            {
                                "id": 3,
                                "translation": "[size=80][c0]Select Legendary Skill[s0][/size]",
                            },
                            ensure_ascii=False,
                        ),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            applied = apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
            )

            self.assertEqual(applied.row_count, 3)

    def test_prepare_and_apply_support_korean_target_column_and_string_ids(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "ko.xlsx"
            out_dir = tmp_path / "out"
            _write_korean_language_workbook(lang_path)

            prepared = prepare_translation_harness(lang_path, lang="ko", output_dir=out_dir)

            self.assertEqual(prepared.manifest["language"], "ko")
            self.assertEqual(prepared.manifest["row_ids"], ["btn.claim", "msg.welcome"])

            response_path = tmp_path / "translation_response.jsonl"
            _write_jsonl(
                response_path,
                [
                    {"id": "btn.claim", "translation": "보상 받기"},
                    {"id": "msg.welcome", "translation": "{playerName}님, 다시 오신 것을 환영합니다"},
                ],
            )
            applied = apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
                lang="ko",
            )

            wb = load_workbook(applied.final_workbook_path, read_only=True, data_only=False)
            try:
                ws = wb["Language"]
                self.assertEqual(ws.cell(2, 3).value, "보상 받기")
                self.assertIn("{playerName}", ws.cell(3, 3).value)
            finally:
                wb.close()
            self.assertEqual(applied.cache_path.name, "ko.jsonl")

    def test_prepare_and_apply_support_european_and_arabic_targets(self):
        cases = {
            "es": ("ES", "Reclamar recompensa", "西班牙语"),
            "pt": ("PT", "Resgatar recompensa", "葡萄牙语"),
            "ru": ("RU", "Получить награду", "俄语"),
            "tr": ("TR", "Ödülü al", "土耳其语"),
            "ar": ("AR", "استلام المكافأة", "阿拉伯语"),
        }
        for lang, (header, translation, cn_header) in cases.items():
            with self.subTest(lang=lang):
                with tempfile.TemporaryDirectory() as tmp:
                    tmp_path = Path(tmp)
                    lang_path = tmp_path / "lang.xlsx"
                    out_dir = tmp_path / "out"
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Sheet1"
                    ws.append(["ID", "CN", header])
                    ws.append([1, SRC_CLAIM, ""])
                    wb.save(lang_path)

                    term_path = tmp_path / "terms.xlsx"
                    term_wb = Workbook()
                    term_ws = term_wb.active
                    term_ws.title = "术语表"
                    term_ws.append(["CN", cn_header])
                    term_ws.append([SRC_CLAIM, translation])
                    term_wb.save(term_path)

                    prepared = prepare_translation_harness(
                        input_path=lang_path,
                        term_base_path=term_path,
                        lang=lang,
                        output_dir=out_dir,
                    )
                    self.assertEqual(prepared.manifest["language"], lang)
                    rows = [
                        json.loads(line)
                        for line in prepared.workpack_path.read_text(encoding="utf-8").splitlines()
                    ]
                    self.assertIn(translation, [term["target"] for term in rows[0]["term_hits"]])

                    response_path = tmp_path / f"translation_response_{lang}.jsonl"
                    _write_jsonl(response_path, [{"id": 1, "translation": translation}])
                    applied = apply_translation_response(
                        input_path=lang_path,
                        manifest_path=prepared.manifest_path,
                        response_path=response_path,
                        output_dir=out_dir,
                        lang=lang,
                    )
                    wb = load_workbook(applied.final_workbook_path, read_only=True, data_only=False)
                    try:
                        self.assertEqual(wb.active.cell(2, 3).value, translation)
                    finally:
                        wb.close()
                    self.assertEqual(applied.cache_path.name, f"{lang}.jsonl")

    def test_multilingual_workbook_writes_requested_language_column_by_header(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "multilingual.xlsx"
            out_dir = tmp_path / "out"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["ID", "CN", "EN", "IT"])
            ws.append([1, SRC_CLAIM, "Existing English", ""])
            wb.save(lang_path)
            wb.close()

            prepared = prepare_translation_harness(lang_path, lang="it", output_dir=out_dir)
            self.assertEqual(prepared.manifest["lang_index"], 1)

            response_path = tmp_path / "translation_response_it.jsonl"
            _write_jsonl(response_path, [{"id": 1, "translation": "Riscatta ricompensa"}])
            applied = apply_translation_response(
                input_path=lang_path,
                manifest_path=prepared.manifest_path,
                response_path=response_path,
                output_dir=out_dir,
                lang="it",
            )

            result = load_workbook(applied.final_workbook_path, read_only=True, data_only=False)
            try:
                self.assertEqual(result.active.cell(2, 3).value, "Existing English")
                self.assertEqual(result.active.cell(2, 4).value, "Riscatta ricompensa")
            finally:
                result.close()

    def test_prepare_rejects_unsupported_language_code(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            lang_path = tmp_path / "lang.xlsx"
            _write_language_workbook(lang_path)
            with self.assertRaisesRegex(ValueError, "supports only"):
                prepare_translation_harness(lang_path, lang="xx", output_dir=tmp_path / "out")


if __name__ == "__main__":
    unittest.main()
