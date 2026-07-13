from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils.large_text_multilingual_pipeline import run_pipeline


LANGS = ["EN", "IDN", "DE", "FR", "ES", "PT", "RU", "IT", "TR", "TH"]


def create_source(path: Path, start_id: int, rows: int) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "UI" if "UI" in path.stem else "Language"
    sheet.append(["ID", "CN", *LANGS])
    for offset in range(rows):
        sheet.append([start_id + offset, f"Text {offset % 7}", *([None] * len(LANGS))])
    workbook.save(path)
    workbook.close()


class Translator:
    def translate_batch(self, rows, target_langs):  # type: ignore[no-untyped-def]
        return [
            {
                "request_key": row["request_key"],
                "translations": {
                    lang: f"{lang} {row['cn']}" for lang in target_langs
                },
            }
            for row in rows
        ]


class Reviewer:
    def review_batch(self, rows, target_langs):  # type: ignore[no-untyped-def]
        return [
            {
                "review_key": row["review_key"],
                "lang": lang,
                "status": "KEEP",
                "suggested": row["translations"][lang],
                "reason": "ok",
            }
            for row in rows
            for lang in target_langs
        ]


class Auditor:
    def audit_batch(self, suggestions):  # type: ignore[no-untyped-def]
        self.assert_no_suggestions = suggestions
        return []


class LargeTextMultilingualPipelineTests(unittest.TestCase):
    def test_pipeline_rejects_input_named_like_qa_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "QA摘要.xlsx"
            create_source(source, 1, 1)

            with self.assertRaisesRegex(ValueError, "reserved delivery file name"):
                run_pipeline(
                    inputs=[source],
                    term_base=None,
                    history_dirs=[],
                    target_langs=LANGS,
                    task_dir=root,
                    relay_config=None,
                    proofread_mode="basic",
                    translation_client=Translator(),
                )

    def test_full_pipeline_delivers_clean_workbooks_with_all_gates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            first = root / "7.13.xlsx"
            second = root / "7.13UI.xlsx"
            create_source(first, 1000, 20)
            create_source(second, 2000, 5)

            result = run_pipeline(
                inputs=[first, second],
                term_base=None,
                history_dirs=[],
                target_langs=LANGS,
                task_dir=root,
                relay_config=None,
                proofread_mode="full",
                translation_client=Translator(),
                reviewer=Reviewer(),
                auditor=Auditor(),
            )

            self.assertEqual(result.source_rows, 25)
            self.assertEqual(result.unique_items, 7)
            self.assertEqual(result.hard_blockers, 0)
            self.assertLess(result.elapsed_seconds, 30)
            self.assertEqual(
                sorted(path.name for path in result.delivery_dir.iterdir()),
                ["7.13.xlsx", "7.13UI.xlsx", "QA摘要.xlsx"],
            )
            delivered = load_workbook(result.delivery_dir / "7.13.xlsx", read_only=True)
            try:
                self.assertEqual(delivered["Language"]["C2"].value, "EN Text 0")
                self.assertEqual(delivered["Language"]["L2"].value, "TH Text 0")
            finally:
                delivered.close()
            manifest = json.loads(result.manifest.read_text(encoding="utf-8"))
            self.assertEqual(manifest["status"], "complete")
            self.assertTrue(all(value in {"done", "skipped"} for value in manifest["phase_status"].values()))
            self.assertTrue(result.retro_metrics.exists())


if __name__ == "__main__":
    unittest.main()
