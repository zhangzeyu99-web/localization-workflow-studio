from __future__ import annotations

import importlib
import subprocess
import sys
import tempfile
from pathlib import Path
import unittest

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "extract_glossary.py"


def sentence_templates_module():
    try:
        return importlib.import_module("glossary_extraction.sentence_templates")
    except ModuleNotFoundError:
        raise AssertionError("official sentence-template module is missing") from None


SKILL_CN = (
    "萝莉炮手每进行<@1>次普攻,展开<@2>秒【狂鲨海岸】,期间每秒对领域内的敌人造成"
    "<@3>%攻击力的大范围伤害,并降低<@4>%生命回复,并提高领域内的敌人受到的<@5>%直接伤害"
)
SKILL_EN = (
    "For every <@1> BATKs launched by Cannon Girl, summons Shore of Sharks for <@2>s, "
    "dealing <@3>% ATK DMG per second to enemies in a large area within the field, reducing "
    "their HP Regen by <@4>%, and increasing their Direct DMG taken by <@5>%."
)
SKILL_NOTICE = (
    "萝莉炮手每进行5次普攻，展开10秒【狂鲨海岸】，期间每秒对领域内的敌人造成1600%攻击力的"
    "大范围伤害，并降低30%生命回复，并提高领域内的敌人受到的40%直接伤害。"
)
SKILL_EXPECTED_EN = (
    "For every 5 BATKs launched by Cannon Girl, summons Shore of Sharks for 10s, dealing "
    "1600% ATK DMG per second to enemies in a large area within the field, reducing their "
    "HP Regen by 30%, and increasing their Direct DMG taken by 40%."
)


class SentenceTemplateMatchingTests(unittest.TestCase):
    def test_historical_script_facade_exports_sentence_template_api(self):
        spec = importlib.util.spec_from_file_location("sentence_template_facade", SCRIPT_PATH)
        facade = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        sys.modules[spec.name] = facade
        spec.loader.exec_module(facade)

        self.assertTrue(hasattr(facade, "build_sentence_template_matches"))

    def test_exact_official_template_captures_values_and_renders_target(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[{"ID": "9233", "CN": SKILL_CN, "translations": {"EN": SKILL_EN}}],
            announcement_text=SKILL_NOTICE,
        )

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["Priority"], 1)
        self.assertEqual(matches[0]["MatchType"], "official_exact")
        self.assertEqual(matches[0]["translations"]["EN"], SKILL_EXPECTED_EN)
        self.assertEqual(matches[0]["_render_status"]["EN"], "rendered")

    def test_exact_context_phrase_uses_language_specific_official_wording(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[
                {
                    "ID": "6235",
                    "CN": "自身增伤+<@1>%",
                    "translations": {
                        "EN": "DMG Boost +<@1>%",
                        "TH": "DMG INC ตัวเอง +<@1>%",
                    },
                }
            ],
            announcement_text="皮肤属性：全队全属性+200，自身增伤+40%",
        )

        self.assertEqual(matches[0]["translations"]["EN"], "DMG Boost +40%")
        self.assertEqual(matches[0]["translations"]["TH"], "DMG INC ตัวเอง +40%")

    def test_structured_language_table_cells_expand_into_aligned_phrase_templates(self):
        module = sentence_templates_module()
        _languages, candidates = module.sentence_template_candidates_from_sheet_rows(
            rows=[
                ["ID", "CN", "EN"],
                [
                    "6235",
                    '[["全队全属性+<@1>"],["自身增伤+<@2>%"]]',
                    '[["Team All Stats +<@1>"],["DMG Boost +<@2>%"]]',
                ],
            ],
            sheet_title="Sheet0",
            id_column="ID",
            source_column="CN",
            target_column="EN",
        )
        matches = module.build_sentence_template_matches(
            candidate_rows=candidates,
            announcement_text="皮肤属性：全队全属性+200，自身增伤+40%",
        )

        exact = {row["OfficialCNTemplate"]: row for row in matches if row["Priority"] == 1}
        self.assertEqual(exact["全队全属性+<@1>"]["translations"]["EN"], "Team All Stats +200")
        self.assertEqual(exact["自身增伤+<@2>%"]["translations"]["EN"], "DMG Boost +40%")

    def test_multilingual_template_merge_can_require_the_primary_language(self):
        module = sentence_templates_module()
        candidates = [
            {"ID": "EN1", "CN": "用于提升奇幻英雄的异能强度。", "translations": {"EN": "EN context"}},
            {"ID": "TH1", "CN": "用于提升奇幻英雄的异能强度。", "translations": {"TH": "TH context"}},
            {"ID": "TH2", "CN": "仅次要语言存在的完整句式。", "translations": {"TH": "TH only"}},
        ]

        merged = module.merge_sentence_template_candidates(candidates, required_language="EN")

        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0]["ID"], "EN1")
        self.assertEqual(merged[0]["translations"], {"EN": "EN context", "TH": "TH context"})

    def test_similar_official_sentence_is_evidence_not_rendered_translation(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[
                {
                    "ID": "8894",
                    "CN": "异能凝粹，灵华内蕴，晶辉外显。用于提升奇幻阵营的英雄异能强度。",
                    "translations": {
                        "EN": "Arcane power and spiritual essence with glorious brilliance. Used to enhance Fantasy heroes' Ability."
                    },
                },
                {
                    "ID": "N1",
                    "CN": "这是与公告内容无关的完整系统说明。",
                    "translations": {"EN": "Unrelated system description."},
                },
                {
                    "ID": "G1",
                    "CN": "激活专属异能，解锁英雄隐藏潜能！",
                    "translations": {"EN": "Activate exclusive Ability."},
                },
            ],
            announcement_text="新增兑换道具【奇幻结晶】：用于提升奇幻阵营英雄的异能阶级。",
            matched_terms=["异能"],
        )

        self.assertEqual(matches[0]["Priority"], 2)
        self.assertEqual(matches[0]["MatchType"], "official_similar")
        self.assertEqual(matches[0]["ID"], "8894")
        self.assertEqual(
            matches[0]["translations"]["EN"],
            "Arcane power and spiritual essence with glorious brilliance. Used to enhance Fantasy heroes' Ability.",
        )

    def test_similar_evidence_uses_uncovered_terms_and_has_a_global_limit(self):
        module = sentence_templates_module()
        terms = [f"专属词甲{index}" for index in range(25)]
        candidates = [
            {
                "ID": f"C{index}",
                "CN": f"关于{term}的完整官方说明文本。",
                "translations": {"EN": f"Official context for {term}."},
            }
            for index, term in enumerate(terms)
        ]
        matches = module.build_sentence_template_matches(
            candidate_rows=candidates,
            announcement_text="、".join(terms),
            matched_terms=terms,
        )

        self.assertEqual(len(matches), 20)
        self.assertTrue(all(row["AnnouncementCN"] in terms for row in matches))

    def test_similar_evidence_is_omitted_when_exact_template_already_covers_term(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[
                {"ID": "9233", "CN": SKILL_CN, "translations": {"EN": SKILL_EN}},
                {
                    "ID": "8849",
                    "CN": "罗刹女每进行<@1>次普攻，展开沙瀑葬域并对领域内敌人造成伤害。",
                    "translations": {"EN": "A similar official skill sentence."},
                },
            ],
            announcement_text=SKILL_NOTICE,
            matched_terms=["普攻", "攻击", "狂鲨海岸"],
        )

        self.assertEqual([row["MatchType"] for row in matches], ["official_exact"])

    def test_unmatched_target_placeholders_are_not_silently_rendered(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[
                {
                    "ID": "P1",
                    "CN": "获得<@1>个奖励",
                    "translations": {"EN": "Get <@1> rewards from <@2>"},
                }
            ],
            announcement_text="获得5个奖励",
        )

        self.assertEqual(matches[0]["translations"]["EN"], "Get <@1> rewards from <@2>")
        self.assertEqual(matches[0]["_render_status"]["EN"], "unverifiable_placeholder")

    def test_repeated_unnumbered_placeholders_capture_values_by_position(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[
                {
                    "ID": "P2",
                    "CN": "玩家%s获得%s积分奖励",
                    "translations": {"EN": "Player %s received %s point rewards"},
                }
            ],
            announcement_text="玩家Alice获得500积分奖励",
        )

        self.assertEqual(matches[0]["translations"]["EN"], "Player Alice received 500 point rewards")

    def test_translation_validation_reports_warning_without_failing(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[{"ID": "9233", "CN": SKILL_CN, "translations": {"EN": SKILL_EN}}],
            announcement_text=SKILL_NOTICE,
        )
        result = module.validate_official_template_usage(
            template_matches=matches,
            translated_texts={
                "EN": "After Cannon Girl performs 5 BATKs, deploys Shore of Sharks for 10 seconds."
            },
        )

        self.assertEqual(result["status"], "warning")
        self.assertEqual(result["checked"], 1)
        self.assertEqual(result["mismatches"], 1)
        self.assertEqual(result["issues"][0]["ID"], "9233")

    def test_translation_validation_warns_when_language_has_no_official_target(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[{"ID": "9233", "CN": SKILL_CN, "translations": {"EN": SKILL_EN}}],
            announcement_text=SKILL_NOTICE,
        )
        result = module.validate_official_template_usage(
            template_matches=matches,
            translated_texts={"TH": "ข้อความแปลภาษาไทย"},
        )

        self.assertEqual(result["status"], "warning")
        self.assertEqual(result["checked"], 1)
        self.assertEqual(result["issues"][0]["reason"], "missing_official_target")

    def test_translation_validation_language_codes_are_case_insensitive(self):
        module = sentence_templates_module()
        matches = module.build_sentence_template_matches(
            candidate_rows=[{"ID": "9233", "CN": SKILL_CN, "translations": {"en": SKILL_EN}}],
            announcement_text=SKILL_NOTICE,
        )
        result = module.validate_official_template_usage(
            template_matches=matches,
            translated_texts={"EN": SKILL_EXPECTED_EN},
        )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["matches"], 1)


class SentenceTemplateCliTests(unittest.TestCase):
    def test_cli_writes_second_sheet_and_warning_report(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            language_path = temp_path / "language.xlsx"
            notice_path = temp_path / "notice.txt"
            translated_path = temp_path / "translated.txt"
            output_path = temp_path / "announcement_terms.xlsx"
            validation_path = temp_path / "announcement_validation.md"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet0"
            sheet.append(["ID", "CN", "EN"])
            sheet.append(["9233", SKILL_CN, SKILL_EN])
            sheet.append(["S1", "狂鲨海岸", "Shore of Sharks"])
            sheet.append(["S2", "普攻", "BATK"])
            sheet.append(["S3", "攻击", "ATK"])
            sheet.append(["S4", "生命回复", "HP Regen"])
            sheet.append(["S5", "直接伤害", "Direct DMG"])
            workbook.save(language_path)
            workbook.close()

            notice_path.write_text(SKILL_NOTICE, encoding="utf-8")
            translated_path.write_text(
                "After Cannon Girl performs 5 BATKs, deploys Shore of Sharks for 10 seconds.",
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(language_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(output_path),
                    "--announcement-validation-output",
                    str(validation_path),
                    "--translated-material",
                    f"EN={translated_path}",
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("ANNOUNCEMENT_TEMPLATES=1", result.stdout)
            self.assertIn("OFFICIAL_TEMPLATE_MISMATCHES=1", result.stdout)
            output_workbook = load_workbook(output_path, read_only=True, data_only=True)
            self.assertEqual(output_workbook.sheetnames, ["Glossary", "SentenceTemplates"])
            rows = list(output_workbook["SentenceTemplates"].iter_rows(values_only=True))
            output_workbook.close()
            self.assertEqual(
                rows[0],
                ("Priority", "MatchType", "ID", "AnnouncementCN", "OfficialCNTemplate", "EN"),
            )
            self.assertEqual(rows[1][0:3], (1, "official_exact", "9233"))
            self.assertEqual(rows[1][5], SKILL_EXPECTED_EN)
            report = validation_path.read_text(encoding="utf-8")
            self.assertIn("official_template_qa: warning", report)
            self.assertIn("official_template_mismatches: 1", report)

    def test_multilingual_cli_merges_language_specific_exact_templates(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            en_path = temp_path / "language_en.xlsx"
            th_path = temp_path / "language_th.xlsx"
            notice_path = temp_path / "notice.txt"
            output_path = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            for path, language, target in [
                (en_path, "EN", "DMG Boost +<@1>%"),
                (th_path, "TH", "DMG INC ตัวเอง +<@1>%"),
            ]:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["ID", "CN", language])
                sheet.append(["6235", "自身增伤+<@1>%", target])
                workbook.save(path)
                workbook.close()
            notice_path.write_text("皮肤属性：自身增伤+40%", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    "--language-table",
                    f"EN={en_path}",
                    "--language-table",
                    f"TH={th_path}",
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(output_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            rows = list(workbook["SentenceTemplates"].iter_rows(values_only=True))
            workbook.close()
            self.assertEqual(rows[0][-2:], ("EN", "TH"))
            self.assertEqual(rows[1][-2:], ("DMG Boost +40%", "DMG INC ตัวเอง +40%"))

    def test_multilingual_cli_reads_headerless_target_column_after_chinese_key(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            en_path = temp_path / "language_en.xlsx"
            th_path = temp_path / "language_th.xlsx"
            notice_path = temp_path / "notice.txt"
            output_path = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            en_book = Workbook()
            en_sheet = en_book.active
            en_sheet.append(["ID", "CN", "EN"])
            en_sheet.append(["6235", "自身增伤+<@1>%", "DMG Boost +<@1>%"])
            en_sheet.append(["6720", "自身增伤", "Self DMG Boost"])
            en_book.save(en_path)
            en_book.close()

            th_book = Workbook()
            th_sheet = th_book.active
            th_sheet.append(["中文key", None, None, None, None, "所属模块"])
            for index in range(60):
                th_sheet.append([f"待翻译文本{index}", None, None, None, None, "属性说明-TD"])
            th_sheet.append(
                ["自身增伤+<@1>%", "DMG INC ตัวเอง +<@1>%", None, None, None, "属性说明-TD"]
            )
            th_sheet.append(["自身增伤", "DMG INC ตัวเอง", None, None, None, "属性说明-TD"])
            th_book.save(th_path)
            th_book.close()
            notice_path.write_text("皮肤属性：自身增伤+40%", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    "--language-table",
                    f"EN={en_path}",
                    "--language-table",
                    f"TH={th_path}",
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(output_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            workbook = load_workbook(output_path, read_only=True, data_only=True)
            glossary_rows = list(workbook["Glossary"].iter_rows(values_only=True))
            template_rows = list(workbook["SentenceTemplates"].iter_rows(values_only=True))
            workbook.close()
            self.assertEqual(glossary_rows[0][-2:], ("EN", "TH"))
            self.assertEqual(glossary_rows[1][-2:], ("Self DMG Boost", "DMG INC ตัวเอง"))
            self.assertEqual(template_rows[0][-2:], ("EN", "TH"))
            self.assertEqual(template_rows[1][-2:], ("DMG Boost +40%", "DMG INC ตัวเอง +40%"))

    def test_translated_material_enables_default_validation_report(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            language_path = temp_path / "language.xlsx"
            notice_path = temp_path / "notice.txt"
            translated_path = temp_path / "translated.txt"
            output_path = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "CN", "EN"])
            sheet.append(["9233", SKILL_CN, SKILL_EN])
            workbook.save(language_path)
            workbook.close()
            notice_path.write_text(SKILL_NOTICE, encoding="utf-8")
            translated_path.write_text(SKILL_EXPECTED_EN, encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(language_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(output_path),
                    "--translated-material",
                    f"EN={translated_path}",
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("OFFICIAL_TEMPLATE_MISMATCHES=0", result.stdout)
            reports = list(temp_path.glob("notice_announcement_validation_*.md"))
            self.assertEqual(len(reports), 1)
            self.assertIn("official_template_qa: ok", reports[0].read_text(encoding="utf-8"))


class SentenceTemplateAiPacketTests(unittest.TestCase):
    def test_ai_packet_places_official_templates_before_atomic_terms(self):
        supplement = importlib.import_module("glossary_extraction.ai_supplement")
        exact = {
            "Priority": 1,
            "MatchType": "official_exact",
            "ID": "9233",
            "AnnouncementCN": SKILL_NOTICE,
            "OfficialCNTemplate": SKILL_CN,
            "translations": {"EN": SKILL_EXPECTED_EN},
            "_render_status": {"EN": "rendered"},
        }
        similar = {
            "Priority": 2,
            "MatchType": "official_similar",
            "ID": "8894",
            "AnnouncementCN": "用于提升奇幻阵营",
            "OfficialCNTemplate": "用于提升奇幻阵营的英雄异能强度。",
            "translations": {"EN": "Used to enhance Fantasy heroes' Ability."},
            "_render_status": {"EN": "evidence"},
        }
        packet = supplement.build_ai_supplement_packet(
            announcement_text=SKILL_NOTICE,
            matched_rows=[{"ID": "S1", "CN": "普攻", "EN": "BATK"}],
            candidate_rows=[],
            headers=["ID", "CN", "EN"],
            sentence_template_matches=[exact, similar],
        )

        self.assertEqual(packet["official_sentence_matches"][0]["ID"], "9233")
        self.assertEqual(packet["official_context_evidence"][0]["ID"], "8894")
        self.assertIn("official sentence", packet["instructions"][0].lower())


if __name__ == "__main__":
    unittest.main()
