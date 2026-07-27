from __future__ import annotations

from openpyxl import load_workbook

from glossary_extraction.excel_io import write_final_workbook
from glossary_extraction.quality import readback_delivery_workbook


def test_clean_delivery_keeps_one_target_and_category(tmp_path):
    output = tmp_path / "final.xlsx"
    rows = [
        {
            "ID": "SkillName_1001",
            "CN": "鲨潮护盾",
            "EN": "Sharkguard",
            "EN2": "Shark Shield",
            "Category": "技能名",
        }
    ]

    write_final_workbook(
        output_path=output,
        final_rows=rows,
        target_header="EN",
        include_en2=False,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Glossary"]
    sheet_rows = list(workbook["Glossary"].iter_rows(values_only=True))
    assert sheet_rows[0] == ("ID", "CN", "EN", "分类")
    assert sheet_rows[1] == ("SkillName_1001", "鲨潮护盾", "Sharkguard", "技能名")
    workbook.close()


def test_readback_blocks_duplicate_cn_and_name_collision(tmp_path):
    output = tmp_path / "invalid.xlsx"
    rows = [
        {"ID": "1", "CN": "鲨潮护盾", "EN": "Sharkguard", "Category": "技能名"},
        {"ID": "2", "CN": "鲨卫", "EN": "Sharkguard", "Category": "技能名"},
        {"ID": "3", "CN": "鲨卫", "EN": "Shark Ward", "Category": "技能名"},
    ]
    write_final_workbook(output, rows, "EN")

    report = readback_delivery_workbook(
        output,
        target_header="EN",
        require_target=True,
        include_en2=False,
    )

    assert report.duplicate_cn == 1
    assert report.name_collisions == 1
    assert report.hard_blockers == 2


def test_clean_delivery_can_include_en2_when_explicitly_requested(tmp_path):
    output = tmp_path / "with_en2.xlsx"
    rows = [
        {
            "ID": "1",
            "CN": "报名",
            "EN": "Registration",
            "EN2": "Sign Up",
            "Category": "动作",
        }
    ]

    write_final_workbook(output, rows, "EN", include_en2=True)

    workbook = load_workbook(output, read_only=True, data_only=True)
    sheet_rows = list(workbook["Glossary"].iter_rows(values_only=True))
    workbook.close()
    assert sheet_rows[0] == ("ID", "CN", "EN", "EN2", "分类")
    assert sheet_rows[1] == ("1", "报名", "Registration", "Sign Up", "动作")
