from __future__ import annotations

import importlib.util
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
import unittest
from unittest.mock import patch
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "extract_glossary.py"
SPEC = importlib.util.spec_from_file_location("extract_glossary", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


def write_minimal_docx(path: Path, paragraphs: list[str]) -> None:
    body = "".join(
        f"<w:p><w:r><w:t>{paragraph}</w:t></w:r></w:p>"
        for paragraph in paragraphs
    )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body}</w:body>"
        "</w:document>"
    )
    with ZipFile(path, "w") as archive:
        archive.writestr("word/document.xml", document_xml)


class UtilityTests(unittest.TestCase):
    def test_normalize_english_ignores_case_spacing_and_plural(self):
        self.assertEqual(
            MODULE.normalize_english_for_compare("Rewards"),
            MODULE.normalize_english_for_compare("reward"),
        )
        self.assertEqual(
            MODULE.normalize_english_for_compare("Dual Guns"),
            MODULE.normalize_english_for_compare("dual   guns"),
        )

    def test_is_valid_term_rejects_composed_effect_and_batch_item_names(self):
        self.assertFalse(MODULE.is_valid_term("\u51b0\u5c01\u6269\u6563\u4f24\u5bb3\u63d0\u9ad8"))
        self.assertFalse(MODULE.is_valid_term("\u6d3b\u52a8\u6295\u653e40\u7ea7\u6b66\u5668"))
        self.assertFalse(MODULE.is_valid_term("\u6d3b\u52a8\u6295\u653e1\u7ea7\u6b66\u5668"))
        self.assertFalse(MODULE.is_valid_term("10\u7ea7\u7ea2\u8272\u6124\u6012\u6b66\u5668"))
        self.assertFalse(MODULE.is_valid_term("70\u7ea7\u7cbe\u9009\u7279\u7ea7\u6b66\u5668"))
        self.assertFalse(MODULE.is_valid_term("\u968f\u673a\u7ea2\u8272\u88c5\u5907"))
        self.assertTrue(MODULE.is_valid_term("\u83b7\u5f97"))
        self.assertTrue(MODULE.is_valid_term("\u9644\u9b54"))
        self.assertTrue(MODULE.is_valid_term("\u51b0\u5c01\u6269\u6563"))

    def test_common_combat_power_term_remains_a_deliverable_attribute(self):
        records = [
            MODULE.Record(row_id="A-1", source="\u6218\u529b", target="S\u1ee9c m\u1ea1nh"),
            MODULE.Record(row_id="A-2", source="\u63d0\u5347\u6218\u529b", target="T\u0103ng s\u1ee9c m\u1ea1nh"),
            MODULE.Record(row_id="A-3", source="\u6218\u529b\u5956\u52b1", target="Ph\u1ea7n th\u01b0\u1edfng s\u1ee9c m\u1ea1nh"),
            MODULE.Record(row_id="A-4", source="\u6218\u529b\u6392\u884c", target="X\u1ebfp h\u1ea1ng s\u1ee9c m\u1ea1nh"),
            MODULE.Record(row_id="A-5", source="\u6218\u529b\u7cfb\u7edf", target="H\u1ec7 th\u1ed1ng s\u1ee9c m\u1ea1nh"),
            MODULE.Record(row_id="A-6", source="\u6218\u529b\u5c5e\u6027", target="Thu\u1ed9c t\u00ednh s\u1ee9c m\u1ea1nh"),
            MODULE.Record(row_id="A-7", source="\u6218\u529b\u6210\u957f", target="T\u0103ng tr\u01b0\u1edfng s\u1ee9c m\u1ea1nh"),
        ]

        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=5,
            glossary_hit_threshold=6,
            include_empty_final_terms=True,
            target_language="VN",
        )

        combat_power = next(row for row in final_rows if row["CN"] == "\u6218\u529b")
        self.assertEqual(combat_power["Category"], "\u5c5e\u6027")

    def test_records_from_rows_preserves_term_type_context(self):
        rows = [
            ["ID", "中文", "英文", "术语类型"],
            ["SkillName_1001", "鲨潮护盾", "Sharkguard", "技能名"],
            ["MapName_2001", "暮色海岸", "Dusk Coast", "地名"],
        ]

        records = MODULE.records_from_rows(
            rows=rows,
            sheet_title="技能与地图",
            id_column="ID",
            source_column="中文",
            target_column="英文",
        )

        self.assertEqual(records[0].sheet_name, "技能与地图")
        self.assertEqual(records[0].row_number, 2)
        self.assertEqual(records[0].source_field, "中文")
        self.assertEqual(records[0].term_type_hint, "技能名")
        self.assertEqual(records[1].term_type_hint, "地名")

    def test_singleton_proper_names_bypass_frequency_threshold(self):
        records = [
            MODULE.Record(
                "SkillName_1001",
                "鲨潮护盾",
                "Sharkguard",
                sheet_name="技能名称",
                row_number=2,
            ),
            MODULE.Record(
                "MapName_2001",
                "暮色海岸",
                "Dusk Coast",
                sheet_name="地图名称",
                row_number=3,
            ),
            MODULE.Record("Text_1", "普通文本", "Normal Text"),
        ]

        all_rows, glossary_rows, high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=5,
            glossary_hit_threshold=10,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="singleton-proper-names",
        )

        final = {row["CN"]: row for row in final_rows}
        self.assertEqual(final["鲨潮护盾"]["Category"], "技能名")
        self.assertEqual(final["鲨潮护盾"]["HitRows"], 1)
        self.assertEqual(final["暮色海岸"]["Category"], "地名")
        self.assertNotIn("普通文本", final)
        self.assertFalse(any(row["CN"] == "普通文本" for row in glossary_rows))
        self.assertFalse(any(row["CN"] == "普通文本" for row in high_risk_rows))
        self.assertTrue(any(row["CN"] == "鲨潮护盾" for row in all_rows))

    def test_conflicting_name_type_stays_in_review_only(self):
        records = [
            MODULE.Record(
                "MapName_2001",
                "暮色海岸",
                "Dusk Coast",
                sheet_name="地图",
                term_type_hint="技能名",
            )
        ]

        all_rows, glossary_rows, high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=5,
            glossary_hit_threshold=10,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="conflicting-name-type",
        )

        self.assertEqual(all_rows[0]["NeedsReview"], "Yes")
        self.assertEqual(all_rows[0]["Category"], "待确认")
        self.assertEqual(high_risk_rows[0]["CN"], "暮色海岸")
        self.assertEqual(glossary_rows, [])
        self.assertEqual(final_rows, [])

    def test_name_budget_warning_does_not_remove_proper_name(self):
        records = [
            MODULE.Record(
                "SkillName_1001",
                "巨齿鲨水盾",
                "Megalodon Water Shield",
                term_type_hint="技能名",
            )
        ]

        all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=5,
            glossary_hit_threshold=10,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="name-budget-warning",
            target_language="EN",
        )

        self.assertEqual(all_rows[0]["NameWordCount"], 3)
        self.assertIn("english_skill_word_budget", all_rows[0]["NamePolicyWarnings"])
        self.assertEqual(final_rows[0]["CN"], "巨齿鲨水盾")

    def test_name_collision_is_high_risk_and_excluded_from_delivery(self):
        records = [
            MODULE.Record("SkillName_1001", "鲨潮护盾", "Sharkguard", term_type_hint="技能名"),
            MODULE.Record("SkillName_1002", "鲨卫", "Sharkguard", term_type_hint="技能名"),
        ]

        all_rows, glossary_rows, high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=5,
            glossary_hit_threshold=10,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="name-collision",
            target_language="EN",
        )

        self.assertTrue(all(row["NameCollision"] == "Yes" for row in all_rows))
        self.assertTrue(all("鲨" in row["NameCollisionWith"] for row in all_rows))
        self.assertEqual({row["CN"] for row in high_risk_rows}, {"鲨潮护盾", "鲨卫"})
        self.assertEqual(glossary_rows, [])
        self.assertEqual(final_rows, [])

    def test_collect_translation_diff_marks_manual_adaptation(self):
        counter = MODULE.Counter(
            {
                "Sign Up": 3,
                "Registration": 2,
                "Registration Countdown": 1,
            }
        )
        diff = MODULE.collect_translation_diff("Sign Up", counter)
        self.assertEqual(diff["has_diff"], "Yes")
        self.assertEqual(diff["same_or_format_only_count"], 3)
        self.assertEqual(diff["diff_count"], 3)
        self.assertEqual(diff["diff_variants"], "Registration (2) | Registration Countdown (1)")

    def test_collect_translation_diff_ignores_context_extension(self):
        counter = MODULE.Counter(
            {
                "Registration": 2,
                "Registration Countdown": 2,
                "Registration Requirements": 1,
            }
        )
        diff = MODULE.collect_translation_diff("Registration", counter)
        self.assertEqual(diff["has_diff"], "No")
        self.assertEqual(diff["same_or_format_only_count"], 5)
        self.assertEqual(diff["diff_count"], 0)

    def test_split_usage_buckets_separates_example_and_manual_adaptation(self):
        counter = MODULE.Counter(
            {
                "Registration": 2,
                "Registration Countdown": 2,
                "Sign Up": 1,
                "Registered": 1,
            }
        )
        example_counter, manual_counter = MODULE.split_usage_buckets("Registration", counter)
        self.assertEqual(example_counter["Registration"], 2)
        self.assertEqual(example_counter["Registration Countdown"], 2)
        self.assertEqual(manual_counter["Sign Up"], 1)
        self.assertEqual(manual_counter["Registered"], 1)

    def test_choose_en2_prefers_exact_and_can_derive_compact_variant(self):
        self.assertEqual(
            MODULE.choose_en2_value(
                example_en="Registration",
                exact_diff_counter=MODULE.Counter({"Sign Up": 1}),
                manual_counter=MODULE.Counter({"Registered": 1}),
            ),
            "Sign Up",
        )
        self.assertEqual(
            MODULE.choose_en2_value(
                example_en="Level Up",
                exact_diff_counter=MODULE.Counter(),
                manual_counter=MODULE.Counter(
                    {
                        "Upgrade Module": 1,
                        "Upgrade Gold Mine": 1,
                        "Upgrade Defense Tower": 1,
                        "Upgrade Camp": 1,
                    }
                ),
            ),
            "Upgrade",
        )

    def test_build_project_brief_infers_project_signals_and_prompt(self):
        records = [
            MODULE.Record("1", "合成花束完成订单", "Merge bouquets to complete orders"),
            MODULE.Record("2", "修复花店装饰", "Restore the flower shop decor"),
            MODULE.Record("3", "领取奖励", "Claim Rewards"),
            MODULE.Record("4", "先生，我最后再问一次……您确定要这么做吗？", "Sir, I'll ask one last time... Are you sure?"),
            MODULE.Record("5", "这里唯一危险的东西是你和你的电锯！", "The only dangerous thing here is you and your chainsaw!"),
        ]
        all_rows, glossary_rows, _high_risk_rows, manual_rows, _final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="brief-fixture",
        )

        markdown, prompt = MODULE.build_project_brief(
            project_name="Fixture Game",
            sheet_name="Sheet0",
            records=records,
            all_rows=all_rows,
            glossary_rows=glossary_rows,
            manual_rows=manual_rows,
        )

        self.assertIn("Fixture Game", markdown)
        self.assertIn("AI 生成的专属翻译提示词", markdown)
        self.assertIn("项目元信息", markdown)
        self.assertIn("合成经营", markdown)
        self.assertNotIn("输入快照", markdown)
        self.assertIn("译文需符合以下要求", prompt)
        self.assertIn("美剧日常对白", prompt)
        self.assertIn("游戏内容/UI/玩法说明尽量精简", prompt)

    def test_load_records_falls_back_to_raw_xlsx_reader(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "styled_language_table.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Main"
            worksheet.append(["ID", "简体中文", "英文"])
            worksheet.append(["UIMail101", "领取", "Claim"])
            workbook.save(input_path)
            workbook.close()

            original_loader = MODULE.load_workbook

            def failing_loader(*_args, **_kwargs):
                raise TypeError("expected <class 'openpyxl.styles.fills.Fill'>")

            try:
                MODULE.load_workbook = failing_loader
                records, sheet_name = MODULE.load_records(
                    input_path=input_path,
                    sheet_name="Main",
                    id_column="ID",
                    source_column="简体中文",
                    target_column="英文",
                )
            finally:
                MODULE.load_workbook = original_loader

            self.assertEqual(sheet_name, "Main")
            self.assertEqual(records, [MODULE.Record("UIMail101", "领取", "Claim")])

    def test_project_brief_prioritizes_aircraft_combat_over_noise(self):
        records = [
            MODULE.Record("1", "战机攻击提升", "Aircraft ATK Up"),
            MODULE.Record("2", "导弹伤害增加", "Missile DMG Up"),
            MODULE.Record("3", "弹幕射击技能", "Barrage Skill"),
            MODULE.Record("4", "英雄装备强化", "Enhance Hero Gear"),
            MODULE.Record("5", "礼包奖励", "Pack Rewards"),
            MODULE.Record("6", "修复失败", "Repair failed"),
        ]
        all_rows, glossary_rows, _high_risk_rows, manual_rows, _final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="aircraft-brief-fixture",
        )

        markdown, prompt = MODULE.build_project_brief(
            project_name="Aircraft",
            sheet_name="Sheet0",
            records=records,
            all_rows=all_rows,
            glossary_rows=glossary_rows,
            manual_rows=manual_rows,
        )

        self.assertIn("科幻战机 / 飞行射击 / RPG养成", markdown)
        self.assertIn("偏科幻军事", prompt)
        self.assertIn("避免可爱化", prompt)
        self.assertNotIn("花店修复", markdown)

    def test_project_brief_uses_extra_materials_and_notes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            design_doc = temp_path / "setting.md"
            design_doc.write_text("项目设定：科幻战机、导弹、弹幕射击、英雄装备养成。", encoding="utf-8")
            screenshot = temp_path / "aircraft_missile_battle_ui.png"
            screenshot.write_bytes(b"fake-image")

            material_records, material_sources = MODULE.load_project_material_records(
                material_paths=[design_doc, screenshot],
                notes=["截图显示深色科幻机库和战机强化界面。"],
            )
            records = [MODULE.Record("1", "领取", "Claim")] + material_records
            markdown, prompt = MODULE.build_project_brief(
                project_name="Fixture Game",
                sheet_name="Sheet0",
                records=records,
                all_rows=[],
                glossary_rows=[],
                manual_rows=[],
                material_sources=material_sources,
            )

            self.assertIn("科幻战机 / 飞行射击 / RPG养成", markdown)
            self.assertIn("信息来源", markdown)
            self.assertIn("setting.md", markdown)
            self.assertIn("aircraft_missile_battle_ui.png", markdown)
            self.assertIn("偏科幻军事", prompt)

    def test_records_from_docx_material_reads_body_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            docx_path = Path(temp_dir) / "notice.docx"
            write_minimal_docx(docx_path, ["新增秘境玩法", "开放纹章系统"])

            records = MODULE.records_from_docx_material(docx_path)

            self.assertEqual([record.source for record in records], ["新增秘境玩法", "开放纹章系统"])

    def test_select_announcement_term_rows_matches_notice_terms_with_translations(self):
        records = [
            MODULE.Record("T1", "秘境", "Trial Realm"),
            MODULE.Record("T2", "纹章", "Emblem"),
            MODULE.Record("T3", "商城", "Shop"),
        ]
        all_rows, _glossary_rows, _high_risk_rows, _manual_rows, _final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="announcement-match-fixture",
        )

        matched = MODULE.select_announcement_term_rows(
            term_rows=all_rows,
            announcement_text="新增秘境和纹章系统",
            include_empty=False,
        )

        self.assertEqual([row["CN"] for row in matched], ["秘境", "纹章"])
        self.assertEqual([row["EN"] for row in matched], ["Trial Realm", "Emblem"])

    def test_select_announcement_term_rows_orders_by_first_position_and_suppresses_overlaps(self):
        records = [
            MODULE.Record("T1", "试炼", "Trial"),
            MODULE.Record("T2", "试炼秘境", "Trial Realm"),
            MODULE.Record("T3", "秘境", "Realm"),
        ]
        all_rows, _glossary_rows, _high_risk_rows, _manual_rows, _final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="announcement-order-fixture",
        )

        matched = MODULE.select_announcement_term_rows(
            term_rows=all_rows,
            announcement_text="新增试炼秘境玩法，秘境奖励提升",
            include_empty=False,
        )

        self.assertEqual([row["CN"] for row in matched], ["试炼秘境", "秘境"])

    def test_select_announcement_term_rows_suppresses_later_repeated_subterms(self):
        records = [
            MODULE.Record("T1", "新大陆", "Newland"),
            MODULE.Record("T2", "大陆", "Continent"),
        ]
        rows = MODULE.build_announcement_candidate_rows(
            records=records,
            curated_rules=MODULE.new_curated_rules(),
            min_hit=1,
        )

        matched = MODULE.select_announcement_term_rows(
            term_rows=rows,
            announcement_text="新大陆开放，新大陆主城显示优化",
            include_empty=False,
        )

        self.assertEqual([row["CN"] for row in matched], ["新大陆"])

    def test_load_announcement_texts_reads_xlsx_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "notice.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Notice"
            worksheet.append(["公告"])
            worksheet.append(["新增纹章系统"])
            workbook.save(workbook_path)
            workbook.close()

            text = MODULE.load_announcement_texts([workbook_path])

            self.assertIn("新增纹章系统", text)

    def test_announcement_candidate_prefers_current_language_table_translation(self):
        curated = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "",
                    "block_en2": True,
                    "ignore": False,
                    "note": "",
                }
            },
        }
        rows = MODULE.build_announcement_candidate_rows(
            records=[MODULE.Record("1", "报名", "Sign Up")],
            curated_rules=curated,
            min_hit=1,
        )

        self.assertEqual(rows[0]["EN"], "Sign Up")


class MemoryTests(unittest.TestCase):
    def test_preferences_can_block_en2_and_accumulate_observations(self):
        curated = {
            "version": 1,
            "terms": {
                "奖励": {
                    "approved_en": "Reward",
                    "approved_en2": "",
                    "block_en2": True,
                    "ignore": False,
                    "note": ""
                }
            },
        }
        observations = MODULE.new_observation_store()
        records = [
            MODULE.Record("1", "奖励", "Reward"),
            MODULE.Record("2", "奖励补发", "Promo"),
            MODULE.Record("3", "奖励", "Rewards"),
        ]
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=observations,
            input_digest="fixture-1",
        )
        row = {item["CN"]: item for item in final_rows}["奖励"]
        self.assertEqual(row["EN"], "Reward")
        self.assertEqual(row["EN2"], "")
        state = observations["terms"]["奖励"]
        self.assertEqual(state["seen_runs"], 1)
        self.assertIn("Reward", state["observed_exact_candidates"])

    def test_current_language_table_translation_wins_over_curated(self):
        curated = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "",
                    "block_en2": True,
                    "ignore": False,
                    "note": ""
                }
            },
        }
        records = [
            MODULE.Record("1", "报名", "Sign Up"),
            MODULE.Record("2", "报名倒计时", "Registration Countdown"),
        ]
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=MODULE.new_observation_store(),
            input_digest="fixture-2",
        )
        row = {item["CN"]: item for item in final_rows}["报名"]
        self.assertEqual(row["EN"], "Sign Up")
        self.assertEqual(row["TranslationSource"], "current_table")
        self.assertEqual(row["TranslationConflict"], "Yes")
        self.assertIn("Registration", row["TranslationConflictValues"])

    def test_curated_translation_only_fills_blank_current_translation(self):
        curated = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "",
                    "block_en2": True,
                    "ignore": False,
                    "note": "",
                }
            },
        }
        records = [MODULE.Record("1", "报名", "")]

        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=MODULE.new_observation_store(),
            input_digest="curated-fills-blank",
            include_empty_final_terms=True,
        )

        self.assertEqual(final_rows[0]["EN"], "Registration")
        self.assertEqual(final_rows[0]["TranslationSource"], "curated")
        self.assertEqual(final_rows[0]["TranslationConflict"], "No")

    def test_observation_history_cannot_override_current_translation(self):
        observations = MODULE.new_observation_store()
        observations["terms"]["报名"] = {
            "observed_exact_candidates": {"Registration": 8},
            "observed_example_usages": {},
            "observed_manual_adaptations": {},
            "seen_runs": 4,
            "last_seen_at": "2026-07-01T00:00:00+00:00",
            "last_input_digest": "old",
        }
        records = [MODULE.Record("1", "报名", "Sign Up")]

        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=observations,
            input_digest="current-run",
        )

        self.assertEqual(final_rows[0]["EN"], "Sign Up")
        self.assertEqual(final_rows[0]["TranslationSource"], "current_table")

    def test_curated_and_observation_stores_roundtrip(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            curated = MODULE.new_curated_rules()
            curated["terms"]["传说"] = {
                "approved_en": "Legend",
                "approved_en2": "Legendary",
                "block_en2": False,
                "ignore": False,
                "note": "fixture"
            }
            observations = MODULE.new_observation_store()
            observations["terms"]["传说"] = {
                "observed_exact_candidates": {"Legend": 2},
                "observed_example_usages": {"Legendary Hero": 3},
                "observed_manual_adaptations": {"Legendary": 1},
                "seen_runs": 2,
                "last_seen_at": "2026-04-24T00:00:00+00:00",
                "last_input_digest": "abc"
            }

            MODULE.save_curated_rules(curated_path, curated)
            MODULE.save_observation_store(observations_path, observations)

            loaded_curated = MODULE.load_curated_rules(curated_path)
            loaded_observations = MODULE.load_observation_store(observations_path)

            self.assertEqual(loaded_curated["terms"]["传说"]["approved_en"], "Legend")
            self.assertEqual(loaded_observations["terms"]["传说"]["seen_runs"], 2)

    def test_legacy_term_memory_can_split_into_two_layers(self):
        legacy = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "Sign Up",
                    "block_en2": False,
                    "ignore": False,
                    "note": "legacy",
                    "observed_exact_candidates": {"Registration": 2},
                    "observed_example_usages": {"Registration Countdown": 1},
                    "observed_manual_adaptations": {"Sign Up": 1},
                    "seen_runs": 2,
                    "last_seen_at": "2026-04-24T00:00:00+00:00",
                    "last_input_digest": "legacy-digest",
                }
            },
        }
        curated, observations = MODULE.split_legacy_term_memory(legacy)
        self.assertEqual(curated["terms"]["报名"]["approved_en"], "Registration")
        self.assertEqual(observations["terms"]["报名"]["seen_runs"], 2)
        self.assertIn("Sign Up", observations["terms"]["报名"]["observed_manual_adaptations"])

    def test_observations_can_backfill_en2_when_curated_en2_is_blank(self):
        curated = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "",
                    "block_en2": False,
                    "ignore": False,
                    "note": ""
                }
            },
        }
        observations = MODULE.new_observation_store()
        observations["terms"]["报名"] = {
            "observed_exact_candidates": {"Registration": 2},
            "observed_example_usages": {"Registration Countdown": 2},
            "observed_manual_adaptations": {"Sign Up": 4},
            "seen_runs": 2,
            "last_seen_at": "2026-04-24T00:00:00+00:00",
            "last_input_digest": "legacy-run",
        }
        records = [
            MODULE.Record("1", "报名", "Registration"),
            MODULE.Record("2", "报名条件", "Registration Requirements"),
        ]
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=observations,
            input_digest="fixture-3",
        )
        row = {item["CN"]: item for item in final_rows}["报名"]
        self.assertEqual(row["EN"], "Registration")
        self.assertEqual(row["EN2"], "Sign Up")

    def test_build_term_rows_does_not_add_empty_curated_rules(self):
        curated = MODULE.new_curated_rules()
        records = [
            MODULE.Record("1", "奖励", "Reward"),
            MODULE.Record("2", "升级", "Level Up"),
        ]
        MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=MODULE.new_observation_store(),
            input_digest="fixture-no-curated-pollution",
        )
        self.assertEqual(curated["terms"], {})


class CliIntegrationTests(unittest.TestCase):
    def test_cli_writes_compact_name_review_packet(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "proper_names.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            packet_path = temp_path / "name_review.json"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Names"
            worksheet.append(["ID", "cn", "en", "术语类型"])
            worksheet.append(["Name_1001", "巨齿鲨水盾", "Megalodon Water Shield", "技能名"])
            worksheet.append(["Text_1", "普通文本", "Normal Text", ""])
            workbook.save(input_path)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--min-hit",
                    "5",
                    "--glossary-hit-threshold",
                    "10",
                    "--target-language",
                    "EN",
                    "--name-review-packet-output",
                    str(packet_path),
                    "--no-project-brief",
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            packet = json.loads(packet_path.read_text(encoding="utf-8"))
            self.assertEqual([row["CN"] for row in packet["candidates"]], ["巨齿鲨水盾"])
            self.assertNotIn("普通文本", packet_path.read_text(encoding="utf-8"))
            self.assertIn("english_skill_word_budget", packet["candidates"][0]["warnings"])

    def test_cli_blocks_project_name_collision_before_final_delivery(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "colliding_names.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            packet_path = temp_path / "name_review.json"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Names"
            worksheet.append(["ID", "cn", "en", "术语类型"])
            worksheet.append(["Name_1001", "鲨潮护盾", "Sharkguard", "技能名"])
            worksheet.append(["Name_1002", "鲨卫", "Sharkguard", "技能名"])
            workbook.save(input_path)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--min-hit",
                    "5",
                    "--glossary-hit-threshold",
                    "10",
                    "--target-language",
                    "EN",
                    "--name-review-packet-output",
                    str(packet_path),
                    "--no-project-brief",
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )

            self.assertEqual(result.returncode, 2, msg=result.stderr or result.stdout)
            self.assertIn("delivery_hard_blockers: 1", result.stdout)
            self.assertTrue(detail_path.exists())
            self.assertTrue(packet_path.exists())
            self.assertFalse(final_path.exists())

    def test_cli_proper_name_end_to_end_blocks_collision_then_passes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "proper_names.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            packet_path = temp_path / "name_review.json"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "名称"
            worksheet.append(["ID", "cn", "en", "术语类型"])
            worksheet.append(["SkillName_1001", "鲨潮护盾", "Sharkguard", "技能名"])
            worksheet.append(
                [
                    "SkillDesc_1001",
                    "召唤鲨潮并获得护盾",
                    "Summons a shark tide and gains a shield.",
                    "",
                ]
            )
            worksheet.append(["SkillName_1002", "鲨卫", "Sharkguard", "技能名"])
            worksheet.append(["MapName_2001", "暮色海岸", "Dusk Coast", "地名"])
            worksheet.append(["Text_1", "终极挑战", "Final Challenge", ""])
            workbook.save(input_path)
            workbook.close()

            curated_path.write_text(
                json.dumps(
                    {
                        "version": 1,
                        "terms": {
                            "鲨潮护盾": {
                                "approved_en": "Megalodon Water Shield",
                                "approved_en2": "",
                                "block_en2": True,
                                "ignore": False,
                                "note": "",
                                "category_override": "",
                                "term_type_override": "ui_skill_name",
                            }
                        },
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            args = [
                sys.executable,
                str(SCRIPT_PATH),
                str(input_path),
                "--output",
                str(detail_path),
                "--final-output",
                str(final_path),
                "--curated-rules",
                str(curated_path),
                "--observations-store",
                str(observations_path),
                "--min-hit",
                "5",
                "--glossary-hit-threshold",
                "10",
                "--target-language",
                "EN",
                "--name-review-packet-output",
                str(packet_path),
                "--no-project-brief",
            ]

            blocked = subprocess.run(
                args,
                cwd=ROOT,
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )
            self.assertEqual(blocked.returncode, 2, msg=blocked.stderr or blocked.stdout)
            self.assertIn("delivery_hard_blockers: 1", blocked.stdout)

            workbook = load_workbook(input_path)
            workbook["名称"]["C4"] = "Shark Ward"
            workbook.save(input_path)
            workbook.close()

            passed = subprocess.run(
                args,
                cwd=ROOT,
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )
            self.assertEqual(passed.returncode, 0, msg=passed.stderr or passed.stdout)

            final_workbook = load_workbook(final_path, read_only=True, data_only=True)
            self.assertEqual(final_workbook.sheetnames, ["Glossary"])
            rows = list(final_workbook["Glossary"].iter_rows(values_only=True))
            self.assertEqual(rows[0], ("ID", "CN", "EN", "分类"))
            lookup = {row[1]: row for row in rows[1:]}
            self.assertEqual(lookup["鲨潮护盾"][2], "Sharkguard")
            self.assertEqual(lookup["鲨潮护盾"][3], "技能名")
            self.assertEqual(lookup["暮色海岸"][3], "地名")
            self.assertNotIn("终极挑战", lookup)
            final_workbook.close()

            packet = json.loads(packet_path.read_text(encoding="utf-8"))
            self.assertEqual(
                {item["CN"] for item in packet["candidates"]},
                {"鲨潮护盾", "鲨卫", "暮色海岸"},
            )
            self.assertNotIn("召唤鲨潮并获得护盾", packet_path.read_text(encoding="utf-8"))

    def test_cli_can_generate_source_only_final_terms(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "source_only_language_table.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "cn"])
            worksheet.append(["1", "奖励"])
            worksheet.append(["2", "奖励补发"])
            worksheet.append(["3", "升级"])
            workbook.save(input_path)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--source-only",
                    "--include-empty-final-terms",
                    "--min-hit",
                    "1",
                    "--glossary-hit-threshold",
                    "1",
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            final_workbook = load_workbook(final_path, read_only=True, data_only=True)
            rows = list(final_workbook["Glossary"].iter_rows(values_only=True))
            final_workbook.close()
            lookup = {row[1]: row for row in rows[1:]}
            self.assertEqual(rows[0], ("ID", "CN", "EN", "分类"))
            self.assertIn("奖励", lookup)
            self.assertEqual(lookup["奖励"][2], None)
            self.assertEqual(lookup["奖励"][3], "资源")

    def test_cli_generates_detail_final_and_store_outputs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "sample_language_table.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"
            project_brief_path = temp_path / "project_brief.md"
            prompt_path = temp_path / "translation_prompt.txt"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "cn", "en"])
            worksheet.append(["1", "报名", "Registration"])
            worksheet.append(["2", "报名条件", "Registration Requirements"])
            worksheet.append(["3", "报名", "Sign Up"])
            worksheet.append(["4", "升级", "Level Up"])
            worksheet.append(["5", "升级模块", "Upgrade Module"])
            workbook.save(input_path)
            workbook.close()

            curated_path.write_text(
                json.dumps(
                    {
                        "version": 1,
                        "terms": {
                            "报名": {
                                "approved_en": "Registration",
                                "approved_en2": "Sign Up",
                                "block_en2": False,
                                "ignore": False,
                                "note": ""
                            },
                            "升级": {
                                "approved_en": "Level Up",
                                "approved_en2": "Upgrade",
                                "block_en2": False,
                                "ignore": False,
                                "note": ""
                            }
                        },
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--project-name",
                    "Fixture Game",
                    "--project-brief-output",
                    str(project_brief_path),
                    "--translation-prompt-output",
                    str(prompt_path),
                    "--min-hit",
                    "1",
                    "--glossary-hit-threshold",
                    "1",
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertTrue(detail_path.exists())
            self.assertTrue(final_path.exists())
            self.assertTrue(curated_path.exists())
            self.assertTrue(observations_path.exists())
            self.assertTrue(project_brief_path.exists())
            self.assertTrue(prompt_path.exists())
            project_brief = project_brief_path.read_text(encoding="utf-8")
            prompt = prompt_path.read_text(encoding="utf-8")
            self.assertIn("Fixture Game", project_brief)
            self.assertIn("AI 生成的专属翻译提示词", project_brief)
            self.assertIn("项目元信息", project_brief)
            self.assertIn("译文需符合以下要求", prompt)
            self.assertIn("PROJECT_BRIEF_OUTPUT=", result.stdout)
            self.assertIn("TRANSLATION_PROMPT_OUTPUT=", result.stdout)
            self.assertIn("PROJECT_MATERIALS=0", result.stdout)

            final_workbook = load_workbook(final_path, read_only=True, data_only=True)
            glossary_sheet = final_workbook["Glossary"]
            rows = list(glossary_sheet.iter_rows(values_only=True))
            final_workbook.close()
            lookup = {row[1]: row for row in rows[1:]}
            self.assertEqual(rows[0], ("ID", "CN", "EN", "分类"))
            self.assertEqual(lookup["报名"][2], "Registration")
            self.assertEqual(lookup["报名"][3], "动作")
            self.assertEqual(lookup["升级"][2], "Level Up")
            self.assertEqual(lookup["升级"][3], "动作")
            self.assertIn("DELIVERY_HARD_BLOCKERS=0", result.stdout)

    def test_cli_generates_announcement_term_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "language_table.xlsx"
            notice_path = temp_path / "notice.txt"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            announcement_output = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "CN", "EN"])
            worksheet.append(["T1", "秘境", "Trial Realm"])
            worksheet.append(["T2", "纹章", "Emblem"])
            worksheet.append(["T3", "商城", "Shop"])
            workbook.save(input_path)
            workbook.close()
            notice_path.write_text("本次更新新增秘境和纹章系统。", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--no-project-brief",
                    "--min-hit",
                    "1",
                    "--glossary-hit-threshold",
                    "1",
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("ANNOUNCEMENT_OUTPUT=", result.stdout)
            self.assertIn("ANNOUNCEMENT_TERMS=2", result.stdout)

            final_workbook = load_workbook(announcement_output, read_only=True, data_only=True)
            rows = list(final_workbook["Glossary"].iter_rows(values_only=True))
            self.assertEqual(rows[0], ("ID", "CN", "EN"))
            self.assertEqual(rows[1:], [("T1", "秘境", "Trial Realm"), ("T2", "纹章", "Emblem")])
            final_workbook.close()

    def test_cli_announcement_term_workbook_preserves_all_language_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "language_table.xlsx"
            notice_path = temp_path / "notice.txt"
            announcement_output = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "CN", "EN", "FR", "DE", "ID", "TH"])
            worksheet.append(["T1", "秘境", "Trial Realm", "Royaume d'épreuve", "Prüfungsreich", "Alam Ujian", "แดนทดสอบ"])
            worksheet.append(["T2", "纹章", "Emblem", "Emblème", "Emblem", "Lambang", "ตราสัญลักษณ์"])
            worksheet.append(["T3", "商城", "Shop", "Boutique", "Shop", "Toko", "ร้านค้า"])
            workbook.save(input_path)
            workbook.close()
            notice_path.write_text("新增秘境和纹章系统。", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            output_workbook = load_workbook(announcement_output, read_only=True, data_only=True)
            rows = list(output_workbook["Glossary"].iter_rows(values_only=True))
            self.assertEqual(rows[0], ("ID", "CN", "EN", "FR", "DE", "ID", "TH"))
            self.assertEqual(rows[1], ("T1", "秘境", "Trial Realm", "Royaume d'épreuve", "Prüfungsreich", "Alam Ujian", "แดนทดสอบ"))
            self.assertEqual(rows[2], ("T2", "纹章", "Emblem", "Emblème", "Emblem", "Lambang", "ตราสัญลักษณ์"))
            output_workbook.close()

    def test_cli_announcement_term_workbook_prefers_clean_standalone_term_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "language_table.xlsx"
            notice_path = temp_path / "notice.txt"
            announcement_output = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "CN", "EN", "FR"])
            worksheet.append(["T1", "{0}结束", "{0} Ended", "Terminé : {0}"])
            worksheet.append(["T2", "结束", "Ended", "Terminé"])
            workbook.save(input_path)
            workbook.close()
            notice_path.write_text("维护结束后发放奖励。", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            output_workbook = load_workbook(announcement_output, read_only=True, data_only=True)
            rows = list(output_workbook["Glossary"].iter_rows(values_only=True))
            self.assertEqual(rows[1], ("T2", "结束", "Ended", "Terminé"))
            output_workbook.close()

    def test_cli_announcement_only_skips_full_glossary_outputs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "language_table.xlsx"
            notice_path = temp_path / "notice.txt"
            announcement_output = temp_path / "announcement_terms.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "CN", "EN"])
            worksheet.append(["T1", "秘境", "Trial Realm"])
            worksheet.append(["T2", "纹章", "Emblem"])
            workbook.save(input_path)
            workbook.close()
            notice_path.write_text("新增秘境玩法。", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("DETAIL_OUTPUT=disabled", result.stdout)
            self.assertIn("FINAL_OUTPUT=disabled", result.stdout)
            self.assertIn("ANNOUNCEMENT_TERMS=1", result.stdout)
            self.assertTrue(announcement_output.exists())
            self.assertEqual(list(temp_path.glob("*_glossary_details_*.xlsx")), [])
            self.assertEqual(list(temp_path.glob("*_ID_CN_EN_EN2_*.xlsx")), [])
            self.assertEqual(list(temp_path.glob("*_announcement_validation_*.md")), [])

    def test_announcement_candidate_rows_auto_detects_export_header_after_metadata(self):
        rows = [
            ["0", None, None],
            ["\u8bed\u8a00\u603b\u8868", None, None],
            ["\u7d22\u5f15ID", "\u5185\u5bb9", "\u4e2d\u6587\uff0c\u7528\u4e8e\u5bfc\u51fa\uff08\u4e0d\u8981\u5728\u539f\u8868\u7b80\u8f6c\u7e41\uff09"],
            ["1*1", "stack_error,n_list", ""],
            ["id", "text", ""],
            ["int", "string", ""],
            ["1001", "Trial Realm", "\u79d8\u5883"],
            ["1002", "Emblem", "\u7eb9\u7ae0"],
        ]

        headers, candidate_rows = MODULE.announcement_candidate_rows_from_sheet_rows(
            rows=rows,
            sheet_title="out_language_en",
            id_column="ID",
            source_column="CN",
            target_column="EN",
            curated_rules=MODULE.new_curated_rules(),
            min_hit=1,
        )

        self.assertEqual(headers, ["ID", "CN", "EN"])
        self.assertEqual([row["CN"] for row in candidate_rows], ["\u79d8\u5883", "\u7eb9\u7ae0"])
        self.assertEqual(candidate_rows[0]["_AnnouncementValues"], ["1001", "\u79d8\u5883", "Trial Realm"])

    def test_announcement_candidate_rows_extracts_json_wrapped_terms(self):
        rows = [
            ["ID", "CN", "KR"],
            ["A1", "[\"\u5e7b\u620f\u793c\u5305\"]", "[\"\ud658\uc0c1\uadf9 \ud328\ud0a4\uc9c0\"]"],
            ["A2", "[\"\u72c2\u6b22\u591c\u5178\",\"#FFFCDC\"]", "[\"\ubc24\uc758 \ucd95\uc81c\",\"#FFFCDC\"]"],
            [
                "A3",
                "[[\"\u4e94   \u8be1\u620f\u5386\u9669\",\"\u73a9\u6cd5\u8bf4\u660e\"]]",
                "[[\"5   \uc18d\uc784\uc218 \ubaa8\ud5d8\",\"\ucf58\ud150\uce20 \uc124\uba85\"]]",
            ],
        ]

        headers, candidate_rows = MODULE.announcement_candidate_rows_from_sheet_rows(
            rows=rows,
            sheet_title="config",
            id_column="ID",
            source_column="CN",
            target_column="KR",
            curated_rules=MODULE.new_curated_rules(),
            min_hit=1,
        )

        self.assertEqual(headers, ["ID", "CN", "KR"])
        lookup = {row["CN"]: row for row in candidate_rows}
        self.assertEqual(lookup["\u5e7b\u620f\u793c\u5305"]["_AnnouncementValues"], ["A1", "\u5e7b\u620f\u793c\u5305", "\ud658\uc0c1\uadf9 \ud328\ud0a4\uc9c0"])
        self.assertEqual(lookup["\u72c2\u6b22\u591c\u5178"]["_AnnouncementValues"], ["A2", "\u72c2\u6b22\u591c\u5178", "\ubc24\uc758 \ucd95\uc81c"])
        self.assertEqual(lookup["\u8be1\u620f\u5386\u9669"]["_AnnouncementValues"], ["A3", "\u8be1\u620f\u5386\u9669", "\uc18d\uc784\uc218 \ubaa8\ud5d8"])

    def test_select_announcement_term_rows_demotes_low_value_terms_without_dropping_them(self):
        term_rows = [
            {"ID": "T1", "CN": "\u73a9\u5bb6", "EN": "Player"},
            {"ID": "T2", "CN": "\u67e5\u770b", "EN": "View"},
            {"ID": "T3", "CN": "\u822a\u6d77\u8d5b\u5b63", "EN": "Sail Season"},
        ]

        matched = MODULE.select_announcement_term_rows(
            term_rows=term_rows,
            announcement_text="\u73a9\u5bb6\u53ef\u67e5\u770b\u822a\u6d77\u8d5b\u5b63",
            include_empty=False,
        )

        self.assertEqual([row["CN"] for row in matched], ["\u822a\u6d77\u8d5b\u5b63", "\u73a9\u5bb6", "\u67e5\u770b"])

    def test_build_ai_supplement_packet_limits_context_to_relevant_evidence(self):
        packet = MODULE.build_ai_supplement_packet(
            announcement_text="\u65b0\u589e\u79d8\u5883\u548c\u661f\u754c\u88c2\u9699\u73a9\u6cd5\u3002",
            matched_rows=[{"ID": "T1", "CN": "\u79d8\u5883", "EN": "Trial Realm"}],
            candidate_rows=[
                {"ID": "T1", "CN": "\u79d8\u5883", "EN": "Trial Realm"},
                {"ID": "S1", "CN": "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218", "EN": "Unlock Astral Rift Challenge"},
                {"ID": "N1", "CN": "\u5b8c\u5168\u65e0\u5173\u7cfb\u7edf", "EN": "Unrelated System"},
            ],
            headers=["ID", "CN", "EN"],
            project_name="",
        )

        evidence_sources = [item["source_text"] for item in packet["evidence_rows"]]
        self.assertIn("\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218", evidence_sources)
        self.assertNotIn("\u5b8c\u5168\u65e0\u5173\u7cfb\u7edf", evidence_sources)
        self.assertEqual(packet["matched_terms"], [{"ID": "T1", "CN": "\u79d8\u5883", "EN": "Trial Realm"}])

    def test_ai_evidence_candidates_keep_long_sentence_context(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            table_path = temp_path / "language.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Main"
            worksheet.append(["ID", "CN", "EN"])
            worksheet.append(["S1", "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218\uff0c\u8d62\u53d6\u4e30\u539a\u5956\u52b1\u3002", "Unlock Astral Rift Challenge to win rich rewards."])
            worksheet.append(["N1", "\u5b8c\u5168\u65e0\u5173\u7684\u957f\u53e5\u5b50\u4e0d\u5e94\u8fdb\u5165\u8bc1\u636e\u5305\u3002", "This unrelated long sentence should not be included."])
            workbook.save(table_path)
            workbook.close()

            candidate_rows = MODULE.build_ai_evidence_candidate_rows_from_workbook(
                input_path=table_path,
                sheet_name=None,
                id_column="ID",
                source_column="CN",
                target_column="EN",
                language="EN",
                source_only=False,
            )
            packet = MODULE.build_ai_supplement_packet(
                announcement_text="\u65b0\u589e\u661f\u754c\u88c2\u9699\u73a9\u6cd5\u3002",
                matched_rows=[],
                candidate_rows=candidate_rows,
                headers=["ID", "CN", "EN"],
                project_name="",
            )

            evidence_sources = [item["source_text"] for item in packet["evidence_rows"]]
            self.assertIn("\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218\uff0c\u8d62\u53d6\u4e30\u539a\u5956\u52b1\u3002", evidence_sources)
            self.assertNotIn("\u5b8c\u5168\u65e0\u5173\u7684\u957f\u53e5\u5b50\u4e0d\u5e94\u8fdb\u5165\u8bc1\u636e\u5305\u3002", evidence_sources)

    def test_apply_ai_supplement_response_adds_evidence_backed_term_to_main_rows(self):
        packet = {
            "evidence_rows": [
                {
                    "evidence_id": "S1",
                    "ID": "S1",
                    "source_text": "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218",
                    "target_text": "Unlock Astral Rift Challenge",
                    "language": "EN",
                    "reason": "announcement_overlap",
                }
            ]
        }
        response = {
            "supplement_terms": [
                {
                    "cn": "\u661f\u754c\u88c2\u9699",
                    "translations": {"EN": "Astral Rift"},
                    "source_ids": ["S1"],
                    "confidence": "medium",
                    "reason": "split from bilingual sentence",
                    "evidence_ids": ["S1"],
                    "action": "add_to_main",
                }
            ]
        }

        rows, report = MODULE.apply_ai_supplement_response(
            announcement_rows=[{"ID": "T1", "CN": "\u79d8\u5883", "EN": "Trial Realm"}],
            headers=["ID", "CN", "EN"],
            announcement_text="\u65b0\u589e\u661f\u754c\u88c2\u9699\u73a9\u6cd5\u3002",
            packet=packet,
            response=response,
            project_name="",
        )

        self.assertEqual([row["CN"] for row in rows], ["\u79d8\u5883", "\u661f\u754c\u88c2\u9699"])
        self.assertEqual(rows[1]["ID"], "S1")
        self.assertEqual(rows[1]["EN"], "Astral Rift")
        self.assertEqual(report["terms"][0]["status"], "added_to_main")

    def test_apply_ai_supplement_response_keeps_low_confidence_or_unbacked_terms_out_of_main_rows(self):
        packet = {
            "evidence_rows": [
                {
                    "evidence_id": "S1",
                    "ID": "S1",
                    "source_text": "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218",
                    "target_text": "Unlock Astral Rift Challenge",
                    "language": "EN",
                    "reason": "announcement_overlap",
                }
            ]
        }
        response = {
            "supplement_terms": [
                {
                    "cn": "\u661f\u754c\u88c2\u9699",
                    "translations": {"EN": "Astral Rift"},
                    "source_ids": ["S1"],
                    "confidence": "low",
                    "reason": "too uncertain",
                    "evidence_ids": ["S1"],
                    "action": "add_to_main",
                },
                {
                    "cn": "\u865a\u7a7a\u5546\u57ce",
                    "translations": {"EN": "Void Shop"},
                    "source_ids": [],
                    "confidence": "high",
                    "reason": "no language-table evidence",
                    "evidence_ids": [],
                    "action": "add_to_main",
                },
            ]
        }

        rows, report = MODULE.apply_ai_supplement_response(
            announcement_rows=[],
            headers=["ID", "CN", "EN"],
            announcement_text="\u65b0\u589e\u661f\u754c\u88c2\u9699\u548c\u865a\u7a7a\u5546\u57ce\u3002",
            packet=packet,
            response=response,
            project_name="",
        )

        self.assertEqual(rows, [])
        self.assertEqual([term["status"] for term in report["terms"]], ["report_only", "report_only"])

    def test_resolve_ai_supplement_provider_auto_falls_back_to_packet_without_api_key(self):
        with patch.dict(os.environ, {}, clear=True):
            provider = MODULE.resolve_ai_supplement_provider(
                provider_name="auto",
                response_path=None,
                model="test-model",
                api_url="https://example.test/v1/responses",
                timeout_seconds=1,
            )

        self.assertIsInstance(provider, MODULE.PacketOnlyAiSupplementProvider)

    def test_resolve_ai_supplement_provider_auto_prefers_response_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            response_path = Path(temp_dir) / "ai_response.json"
            response_path.write_text('{"supplement_terms":[]}', encoding="utf-8")
            with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
                provider = MODULE.resolve_ai_supplement_provider(
                    provider_name="auto",
                    response_path=response_path,
                    model="test-model",
                    api_url="https://example.test/v1/responses",
                    timeout_seconds=1,
                )

        self.assertIsInstance(provider, MODULE.FileAiSupplementProvider)

    def test_openai_ai_supplement_provider_parses_structured_response(self):
        captured = {}

        class FakeResponse:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, tb):
                return False

            def read(self):
                return json.dumps(
                    {
                        "output_text": json.dumps(
                            {
                                "supplement_terms": [
                                    {
                                        "cn": "\u661f\u754c\u88c2\u9699",
                                        "translations": {"EN": "Astral Rift"},
                                        "source_ids": ["S1"],
                                        "confidence": "high",
                                        "reason": "evidence-backed split",
                                        "evidence_ids": ["S1"],
                                        "action": "add_to_main",
                                    }
                                ]
                            },
                            ensure_ascii=False,
                        )
                    },
                    ensure_ascii=False,
                ).encode("utf-8")

        def fake_urlopen(request, timeout):
            captured["url"] = request.full_url
            captured["timeout"] = timeout
            captured["headers"] = dict(request.header_items())
            captured["payload"] = json.loads(request.data.decode("utf-8"))
            return FakeResponse()

        provider = MODULE.OpenAiSupplementProvider(
            api_key="test-key",
            model="test-model",
            api_url="https://example.test/v1/responses",
            timeout_seconds=7,
        )
        with patch.object(MODULE.urllib.request, "urlopen", fake_urlopen):
            response = provider.generate({"announcement_text": "\u65b0\u589e\u661f\u754c\u88c2\u9699"})

        self.assertEqual(captured["url"], "https://example.test/v1/responses")
        self.assertEqual(captured["timeout"], 7)
        self.assertEqual(captured["headers"]["Authorization"], "Bearer test-key")
        self.assertEqual(captured["payload"]["model"], "test-model")
        self.assertEqual(captured["payload"]["text"]["format"]["type"], "json_schema")
        self.assertEqual(response["supplement_terms"][0]["cn"], "\u661f\u754c\u88c2\u9699")

    def test_cli_ai_supplement_auto_uses_packet_fallback_without_api_key(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            table_path = temp_path / "language.xlsx"
            notice_path = temp_path / "notice.txt"
            announcement_output = temp_path / "announcement_terms.xlsx"
            packet_output = temp_path / "ai_packet.json"
            report_output = temp_path / "ai_report.md"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Main"
            worksheet.append(["ID", "CN", "EN"])
            worksheet.append(["T1", "\u79d8\u5883", "Trial Realm"])
            worksheet.append(["S1", "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218", "Unlock Astral Rift Challenge"])
            workbook.save(table_path)
            workbook.close()

            notice_path.write_text("\u65b0\u589e\u79d8\u5883\u548c\u661f\u754c\u88c2\u9699\u73a9\u6cd5\u3002", encoding="utf-8")
            env = os.environ.copy()
            env.pop("OPENAI_API_KEY", None)
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(table_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                    "--ai-supplement",
                    "--ai-supplement-packet-output",
                    str(packet_output),
                    "--ai-supplement-report-output",
                    str(report_output),
                    "--curated-rules",
                    str(temp_path / "curated.json"),
                    "--observations-store",
                    str(temp_path / "observations.json"),
                ],
                cwd=ROOT,
                env=env,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("AI_SUPPLEMENT_PROVIDER=packet", result.stdout)
            self.assertTrue(packet_output.exists())
            self.assertTrue(report_output.exists())
            output_workbook = load_workbook(announcement_output, read_only=True, data_only=True)
            rows = list(output_workbook["Glossary"].iter_rows(values_only=True))
            output_workbook.close()
            self.assertEqual(rows[0], ("ID", "CN", "EN"))
            self.assertEqual(rows[1:], [("T1", "\u79d8\u5883", "Trial Realm")])

    def test_cli_ai_supplement_reports_missing_project_name_translation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            table_path = temp_path / "language.xlsx"
            notice_path = temp_path / "notice.txt"
            response_path = temp_path / "ai_response.json"
            announcement_output = temp_path / "announcement_terms.xlsx"
            report_output = temp_path / "ai_report.md"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Main"
            worksheet.append(["ID", "CN", "EN"])
            worksheet.append(["T1", "\u79d8\u5883", "Trial Realm"])
            worksheet.append(["S1", "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218", "Unlock Astral Rift Challenge"])
            workbook.save(table_path)
            workbook.close()

            notice_path.write_text("\u65b0\u589e\u79d8\u5883\u548c\u661f\u754c\u88c2\u9699\u73a9\u6cd5\u3002", encoding="utf-8")
            response_path.write_text(
                json.dumps(
                    {
                        "supplement_terms": [
                            {
                                "cn": "\u661f\u754c\u88c2\u9699",
                                "translations": {"EN": "Astral Rift"},
                                "source_ids": ["S1"],
                                "confidence": "high",
                                "reason": "split from bilingual sentence",
                                "evidence_ids": ["S1"],
                                "action": "add_to_main",
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "cp1252"
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(table_path),
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                    "--ai-supplement",
                    "--ai-supplement-response",
                    str(response_path),
                    "--ai-supplement-report-output",
                    str(report_output),
                    "--project-name",
                    "\u52c7\u8005\u8054\u76df",
                    "--curated-rules",
                    str(temp_path / "curated.json"),
                    "--observations-store",
                    str(temp_path / "observations.json"),
                ],
                cwd=ROOT,
                env=env,
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("PROJECT_NAME_TRANSLATION_MISSING=\u52c7\u8005\u8054\u76df", result.stdout)
            self.assertTrue(report_output.exists())

            output_workbook = load_workbook(announcement_output, read_only=True, data_only=True)
            rows = list(output_workbook["Glossary"].iter_rows(values_only=True))
            output_workbook.close()
            self.assertEqual(rows[0], ("ID", "CN", "EN"))
            self.assertEqual(rows[1:], [("T1", "\u79d8\u5883", "Trial Realm"), ("S1", "\u661f\u754c\u88c2\u9699", "Astral Rift")])
            self.assertIn("\u8bf7\u8865\u5145\u9879\u76ee\u540d\u6807\u51c6\u8bd1\u6587", report_output.read_text(encoding="utf-8"))

    def test_cli_generates_multilingual_announcement_terms_and_validation_report(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            en_path = temp_path / "language_en.xlsx"
            fr_path = temp_path / "language_fr.xlsx"
            notice_path = temp_path / "notice.txt"
            announcement_output = temp_path / "announcement_terms.xlsx"
            validation_output = temp_path / "announcement_validation.md"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            for workbook_path, translations in [
                (en_path, {"\u79d8\u5883": "Trial Realm", "\u7eb9\u7ae0": "Emblem", "\u73a9\u5bb6": "Player"}),
                (
                    fr_path,
                    {
                        "\u79d8\u5883": "Royaume d'epreuve",
                        "\u7eb9\u7ae0": "Embleme",
                        "\u73a9\u5bb6": "Joueur",
                        "\u5956\u52b1": "Recompense",
                    },
                ),
            ]:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "out_language"
                worksheet.append(["0", None, None])
                worksheet.append(["\u8bed\u8a00\u603b\u8868", None, None])
                worksheet.append(["\u7d22\u5f15ID", "\u5185\u5bb9", "\u4e2d\u6587\uff0c\u7528\u4e8e\u5bfc\u51fa\uff08\u4e0d\u8981\u5728\u539f\u8868\u7b80\u8f6c\u7e41\uff09"])
                worksheet.append(["1*1", "stack_error,n_list", ""])
                worksheet.append(["id", "text", ""])
                worksheet.append(["int", "string", ""])
                for index, (cn, target) in enumerate(translations.items(), start=1001):
                    worksheet.append([str(index), target, cn])
                workbook.save(workbook_path)
                workbook.close()

            notice_path.write_text(
                "\u65b0\u589e\u79d8\u5883\u548c\u7eb9\u7ae0\u7cfb\u7edf\uff0c\u73a9\u5bb6\u53ef\u67e5\u770b\u79d8\u5883\u5956\u52b1\u3002",
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    "--language-table",
                    f"EN={en_path}",
                    "--language-table",
                    f"FR={fr_path}",
                    "--announcement-material",
                    str(notice_path),
                    "--announcement-output",
                    str(announcement_output),
                    "--announcement-validation-output",
                    str(validation_output),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertIn("ANNOUNCEMENT_TERMS=3", result.stdout)
            self.assertTrue(validation_output.exists())

            output_workbook = load_workbook(announcement_output, read_only=True, data_only=True)
            rows = list(output_workbook["Glossary"].iter_rows(values_only=True))
            self.assertEqual(rows[0], ("ID", "CN", "EN", "FR"))
            self.assertEqual(
                rows[1:],
                [
                    ("1001", "\u79d8\u5883", "Trial Realm", "Royaume d'epreuve"),
                    ("1002", "\u7eb9\u7ae0", "Emblem", "Embleme"),
                    ("1003", "\u73a9\u5bb6", "Player", "Joueur"),
                ],
            )
            output_workbook.close()

            validation_text = validation_output.read_text(encoding="utf-8")
            self.assertIn("term_count: 3", validation_text)
            self.assertIn("duplicate_cn: 0", validation_text)
            self.assertIn("empty_translation_cells: 0", validation_text)


if __name__ == "__main__":
    unittest.main()
