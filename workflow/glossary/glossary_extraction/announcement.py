"""Announcement lookup: candidate rows, multilingual rows, workbook and validation outputs."""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from glossary_extraction.constants import (
    DELIMITED_MATERIAL_EXTENSIONS,
    DOCX_MATERIAL_EXTENSIONS,
    LOW_VALUE_ANNOUNCEMENT_TERMS,
    TABLE_MATERIAL_EXTENSIONS,
    TEXT_MATERIAL_EXTENSIONS,
)
from glossary_extraction.excel_io import (
    display_header_name,
    iter_raw_xlsx_sheets,
    language_table_layout_from_rows,
    load_table_material_records,
    records_from_delimited_material,
    records_from_docx_material,
    records_from_text_material,
    style_sheet,
    value_at,
)
from glossary_extraction.experience import get_curated_term_state, new_curated_rules
from glossary_extraction.heuristics import clean_text, extract_structured_term_pairs, is_valid_term
from glossary_extraction.models import LanguageTableSpec, Record


def load_announcement_texts(material_paths: list[Path]) -> str:
    chunks: list[str] = []
    for material_path in material_paths:
        path = Path(material_path)
        if not path.exists():
            raise FileNotFoundError(f"Missing announcement material: {path}")

        suffix = path.suffix.lower()
        if suffix in DOCX_MATERIAL_EXTENSIONS:
            records = records_from_docx_material(path)
        elif suffix in TABLE_MATERIAL_EXTENSIONS:
            records = load_table_material_records(path)
        elif suffix in DELIMITED_MATERIAL_EXTENSIONS:
            records = records_from_delimited_material(path)
        elif suffix in TEXT_MATERIAL_EXTENSIONS:
            records = records_from_text_material(path)
        else:
            records = records_from_text_material(path)

        chunks.extend(record.source for record in records if record.source)
    return clean_text(" ".join(chunks))


def build_announcement_candidate_rows(
    records: list[Record],
    curated_rules: dict[str, Any] | None = None,
    min_hit: int = 1,
) -> list[dict[str, object]]:
    curated_rules = curated_rules if curated_rules is not None else new_curated_rules()
    records_by_term: dict[str, list[Record]] = defaultdict(list)
    translations_by_term: dict[str, Counter[str]] = defaultdict(Counter)
    for record in records:
        term = clean_text(record.source)
        if not is_valid_term(term):
            continue
        records_by_term[term].append(record)
        if record.target:
            translations_by_term[term][record.target] += 1

    rows: list[dict[str, object]] = []
    for term, term_records in records_by_term.items():
        if len(term_records) < min_hit:
            continue
        curated_state = get_curated_term_state(curated_rules, term, create=False)
        if curated_state.get("ignore"):
            continue

        approved_en = clean_text(curated_state.get("approved_en"))
        approved_en2 = "" if curated_state.get("block_en2") else clean_text(curated_state.get("approved_en2"))
        common_en = translations_by_term[term].most_common(1)[0][0] if translations_by_term[term] else ""
        en = approved_en or common_en
        example_record = next((record for record in term_records if record.target == en), term_records[0])
        rows.append(
            {
                "ID": example_record.row_id,
                "CN": term,
                "EN": en,
                "EN2": approved_en2,
            }
        )
    return rows


def announcement_candidate_rows_from_sheet_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    curated_rules: dict[str, Any] | None = None,
    min_hit: int = 1,
    source_only: bool = False,
) -> tuple[list[str], list[dict[str, object]]]:
    if not rows:
        return [], []
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        return [], []

    curated_rules = curated_rules if curated_rules is not None else new_curated_rules()
    records_by_term: dict[str, list[tuple[str, str, list[object], str]]] = defaultdict(list)
    for row_number, row in enumerate(rows[layout.header_row_index + 1 :], start=layout.header_row_index + 2):
        row_values = list(row)
        raw_source = value_at(row_values, layout.source_index)
        raw_id = value_at(row_values, layout.id_index)
        row_id = "" if raw_id is None else str(raw_id)
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        raw_target = "" if layout.target_index is None else value_at(row_values, layout.target_index)
        raw_source_text = "" if raw_source is None else str(raw_source).strip()
        for term, target in extract_structured_term_pairs(raw_source, raw_target):
            term = clean_text(term)
            if not is_valid_term(term):
                continue
            values = [value_at(row_values, index) for index in layout.output_indexes]
            if len(values) >= 2:
                values[1] = term
            if len(values) >= 3:
                values[2] = target
            records_by_term[term].append((row_id, clean_text(target), values, raw_source_text))

    candidate_rows: list[dict[str, object]] = []
    for term, entries in records_by_term.items():
        if len(entries) < min_hit:
            continue
        curated_state = get_curated_term_state(curated_rules, term, create=False)
        if curated_state.get("ignore"):
            continue
        row_id, target, values, _raw_source_text = next(
            (entry for entry in entries if entry[3] == term),
            entries[0],
        )
        candidate_rows.append(
            {
                "ID": row_id,
                "CN": term,
                "EN": target,
                "EN2": "" if curated_state.get("block_en2") else clean_text(curated_state.get("approved_en2")),
                "_AnnouncementValues": values,
            }
        )
    return layout.headers, candidate_rows


def build_announcement_candidate_rows_from_workbook(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    curated_rules: dict[str, Any] | None = None,
    min_hit: int = 1,
    source_only: bool = False,
) -> tuple[list[str], list[dict[str, object]]]:
    headers: list[str] = []
    candidate_rows: list[dict[str, object]] = []
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
        for worksheet in worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            sheet_headers, sheet_rows = announcement_candidate_rows_from_sheet_rows(
                rows=rows,
                sheet_title=worksheet.title,
                id_column=id_column,
                source_column=source_column,
                target_column=target_column,
                curated_rules=curated_rules,
                min_hit=min_hit,
                source_only=source_only,
            )
            if sheet_rows and not headers:
                headers = sheet_headers
            candidate_rows.extend(sheet_rows)
        workbook.close()
        return headers, candidate_rows
    except Exception:
        for raw_sheet_name, rows in iter_raw_xlsx_sheets(input_path):
            if sheet_name and raw_sheet_name != sheet_name:
                continue
            sheet_headers, sheet_rows = announcement_candidate_rows_from_sheet_rows(
                rows=rows,
                sheet_title=raw_sheet_name,
                id_column=id_column,
                source_column=source_column,
                target_column=target_column,
                curated_rules=curated_rules,
                min_hit=min_hit,
                source_only=source_only,
            )
            if sheet_rows and not headers:
                headers = sheet_headers
            candidate_rows.extend(sheet_rows)
        return headers, candidate_rows


def is_low_value_announcement_term(term: str) -> bool:
    return clean_text(term) in LOW_VALUE_ANNOUNCEMENT_TERMS


def select_announcement_term_rows(
    term_rows: list[dict[str, object]],
    announcement_text: str,
    include_empty: bool = False,
) -> list[dict[str, object]]:
    normalized_notice = clean_text(announcement_text)
    candidates: list[tuple[int, int, int, int, str, dict[str, object]]] = []
    spans_by_term: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for row in term_rows:
        cn = clean_text(row.get("CN"))
        if not cn:
            continue

        en = clean_text(row.get("EN")) or clean_text(row.get("EN2"))
        if not include_empty and not en:
            continue

        output_row = dict(row)
        output_row["CN"] = cn
        output_row["EN"] = en
        for match in re.finditer(re.escape(cn), normalized_notice):
            span = (match.start(), match.end())
            low_value_rank = 1 if is_low_value_announcement_term(cn) else 0
            candidates.append((low_value_rank, span[0], span[1], -len(cn), cn, output_row))
            spans_by_term[cn].append(span)

    candidates.sort(key=lambda item: (item[0], item[1], item[3], item[4]))
    selected_spans: list[tuple[int, int]] = []
    selected_terms: set[str] = set()
    selected_rows: list[dict[str, object]] = []
    for _low_value_rank, start, end, _negative_length, cn, row in candidates:
        if cn in selected_terms:
            continue
        if any(start < selected_end and end > selected_start for selected_start, selected_end in selected_spans):
            continue
        selected_spans.extend(spans_by_term.get(cn, [(start, end)]))
        selected_terms.add(cn)
        selected_rows.append(row)
    return selected_rows


def write_announcement_glossary_workbook(
    output_path: Path,
    matched_rows: list[dict[str, object]],
    id_header: str,
    source_header: str,
    target_header: str,
    headers: list[str] | None = None,
    sentence_template_matches: list[dict[str, object]] | None = None,
    template_languages: list[str] | None = None,
) -> None:
    workbook = Workbook()
    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    output_headers = headers or [
        display_header_name(id_header, "ID"),
        display_header_name(source_header, "CN"),
        display_header_name(target_header, "EN"),
    ]
    glossary_sheet.append(output_headers)
    for row in matched_rows:
        source_values = row.get("_AnnouncementValues")
        if isinstance(source_values, list):
            values = source_values[: len(output_headers)]
            if len(values) < len(output_headers):
                values.extend([""] * (len(output_headers) - len(values)))
            glossary_sheet.append(values)
        else:
            values = []
            for header in output_headers:
                if clean_text(header).lower() == "en":
                    values.append(clean_text(row.get("EN")) or clean_text(row.get("EN2")))
                else:
                    values.append(row.get(header, ""))
            glossary_sheet.append(values)
    style_sheet(glossary_sheet)

    template_sheet = workbook.create_sheet("SentenceTemplates")
    languages = template_languages or output_headers[2:]
    template_headers = [
        "Priority",
        "MatchType",
        "ID",
        "AnnouncementCN",
        "OfficialCNTemplate",
        *languages,
    ]
    template_sheet.append(template_headers)
    for row in sentence_template_matches or []:
        translations = row.get("translations", {})
        translations = translations if isinstance(translations, dict) else {}
        template_sheet.append(
            [
                row.get("Priority", ""),
                row.get("MatchType", ""),
                row.get("ID", ""),
                row.get("AnnouncementCN", ""),
                row.get("OfficialCNTemplate", ""),
                *[translations.get(language, "") for language in languages],
            ]
        )
    style_sheet(template_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def parse_language_table_spec(raw_spec: str) -> LanguageTableSpec:
    if "=" not in raw_spec:
        raise ValueError(f"Invalid --language-table value '{raw_spec}'. Expected LANG=path.")
    raw_language, raw_path = raw_spec.split("=", 1)
    language = clean_text(raw_language).upper()
    if not language or not re.match(r"^[A-Z0-9_-]+$", language):
        raise ValueError(f"Invalid language code in --language-table value '{raw_spec}'.")
    if not raw_path.strip():
        raise ValueError(f"Missing path in --language-table value '{raw_spec}'.")
    path = Path(raw_path.strip())
    return LanguageTableSpec(language=language, path=path)


def parse_language_table_specs(raw_specs: list[str]) -> list[LanguageTableSpec]:
    specs = [parse_language_table_spec(raw_spec) for raw_spec in raw_specs]
    seen_languages: set[str] = set()
    for spec in specs:
        if spec.language in seen_languages:
            raise ValueError(f"Duplicate --language-table language code: {spec.language}")
        seen_languages.add(spec.language)
    return specs


def build_multilingual_announcement_rows(
    language_table_specs: list[LanguageTableSpec],
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    curated_rules: dict[str, Any],
    announcement_min_hit: int,
    source_only: bool,
    announcement_text: str,
    include_empty: bool = False,
) -> tuple[list[dict[str, object]], dict[str, int]]:
    candidate_by_cn: dict[str, dict[str, object]] = {}
    translations_by_language: dict[str, dict[str, str]] = {spec.language: {} for spec in language_table_specs}
    primary_terms: set[str] = set()
    duplicate_source_terms = 0

    for spec_index, spec in enumerate(language_table_specs):
        _headers, candidate_rows = build_announcement_candidate_rows_from_workbook(
            input_path=spec.path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=spec.language,
            curated_rules=curated_rules,
            min_hit=announcement_min_hit,
            source_only=source_only,
        )
        for row in candidate_rows:
            cn = clean_text(row.get("CN"))
            if not cn:
                continue
            if spec_index == 0:
                primary_terms.add(cn)
            target = clean_text(row.get("EN")) or clean_text(row.get("EN2"))
            if target:
                translations_by_language[spec.language][cn] = target
            candidate = candidate_by_cn.get(cn)
            if candidate is None:
                candidate_by_cn[cn] = {
                    "ID": row.get("ID", ""),
                    "CN": cn,
                    "EN": target,
                }
            else:
                if clean_text(candidate.get("ID")) and clean_text(row.get("ID")) and clean_text(candidate.get("ID")) != clean_text(row.get("ID")):
                    duplicate_source_terms += 1
                if not clean_text(candidate.get("EN")) and target:
                    candidate["EN"] = target

    matched_terms = select_announcement_term_rows(
        term_rows=[candidate_by_cn[cn] for cn in candidate_by_cn if cn in primary_terms],
        announcement_text=announcement_text,
        include_empty=include_empty,
    )

    rows: list[dict[str, object]] = []
    for matched in matched_terms:
        cn = clean_text(matched.get("CN"))
        row = {
            "ID": matched.get("ID", ""),
            "CN": cn,
        }
        for spec in language_table_specs:
            row[spec.language] = translations_by_language[spec.language].get(cn, "")
        rows.append(row)

    stats = {
        "candidate_terms": len(primary_terms),
        "duplicate_source_terms": duplicate_source_terms,
    }
    return rows, stats


def announcement_output_values(row: dict[str, object], headers: list[str]) -> list[object]:
    source_values = row.get("_AnnouncementValues")
    if isinstance(source_values, list):
        values = source_values[: len(headers)]
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        return values
    values: list[object] = []
    for header in headers:
        if clean_text(header).lower() == "en":
            values.append(clean_text(row.get("EN")) or clean_text(row.get("EN2")))
        else:
            values.append(row.get(header, ""))
    return values


def build_announcement_validation_markdown(
    announcement_materials: list[Path],
    language_tables: list[str],
    glossary_output_path: Path,
    rows: list[dict[str, object]],
    headers: list[str],
    stats: dict[str, int] | None = None,
    sentence_template_matches: list[dict[str, object]] | None = None,
    template_qa: dict[str, object] | None = None,
) -> str:
    stats = stats or {}
    sentence_template_matches = sentence_template_matches or []
    template_qa = template_qa or {
        "status": "not_run",
        "checked": 0,
        "matches": 0,
        "mismatches": 0,
        "unverifiable_placeholders": 0,
        "issues": [],
    }
    cn_values = [clean_text(row.get("CN")) for row in rows if clean_text(row.get("CN"))]
    duplicate_cn = len(cn_values) - len(set(cn_values))
    language_headers = headers[2:]
    empty_translation_cells = 0
    for row in rows:
        values = announcement_output_values(row, headers)
        for index in range(2, len(headers)):
            if index >= len(values) or not clean_text(values[index]):
                empty_translation_cells += 1
    low_value_terms = sum(1 for row in rows if is_low_value_announcement_term(clean_text(row.get("CN"))))

    lines = [
        "# Announcement Glossary Validation",
        "",
        f"status: {'warning' if template_qa.get('status') == 'warning' else 'ok'}",
        f"term_count: {len(rows)}",
        f"languages: {', '.join(language_headers) if language_headers else 'none'}",
        f"duplicate_cn: {duplicate_cn}",
        f"duplicate_source_terms: {int(stats.get('duplicate_source_terms', 0))}",
        f"empty_translation_cells: {empty_translation_cells}",
        f"missing_language_values: {empty_translation_cells}",
        f"low_value_terms: {low_value_terms}",
        f"candidate_terms: {int(stats.get('candidate_terms', len(rows)))}",
        f"sentence_template_count: {len(sentence_template_matches)}",
        f"official_exact_templates: {sum(1 for row in sentence_template_matches if row.get('MatchType') == 'official_exact')}",
        f"official_similar_evidence: {sum(1 for row in sentence_template_matches if row.get('MatchType') == 'official_similar')}",
        f"official_template_qa: {template_qa.get('status', 'not_run')}",
        f"official_template_checked: {int(template_qa.get('checked', 0))}",
        f"official_template_matches: {int(template_qa.get('matches', 0))}",
        f"official_template_mismatches: {int(template_qa.get('mismatches', 0))}",
        f"unverifiable_placeholders: {int(template_qa.get('unverifiable_placeholders', 0))}",
        f"output: {glossary_output_path}",
        "",
        "## Announcement Materials",
    ]
    lines.extend(f"- {path}" for path in announcement_materials)
    lines.append("")
    lines.append("## Language Tables")
    lines.extend(f"- {source}" for source in language_tables)
    lines.append("")
    issues = template_qa.get("issues", [])
    if isinstance(issues, list) and issues:
        lines.extend(["## Official Template Warnings", ""])
        for issue in issues:
            if not isinstance(issue, dict):
                continue
            lines.append(
                f"- ID={issue.get('ID', '')} | language={issue.get('language', '')} | "
                f"reason={issue.get('reason', '')} | expected={issue.get('expected', '')}"
            )
        lines.append("")
    return "\n".join(lines)


def write_announcement_validation_report(
    output_path: Path,
    announcement_materials: list[Path],
    language_tables: list[str],
    glossary_output_path: Path,
    rows: list[dict[str, object]],
    headers: list[str],
    stats: dict[str, int] | None = None,
    sentence_template_matches: list[dict[str, object]] | None = None,
    template_qa: dict[str, object] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        build_announcement_validation_markdown(
            announcement_materials=announcement_materials,
            language_tables=language_tables,
            glossary_output_path=glossary_output_path,
            rows=rows,
            headers=headers,
            stats=stats,
            sentence_template_matches=sentence_template_matches,
            template_qa=template_qa,
        ),
        encoding="utf-8",
    )
