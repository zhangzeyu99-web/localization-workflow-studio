"""Workbook, CSV/TSV, DOCX, and text material readers plus workbook writers."""

from __future__ import annotations

import csv
import hashlib
import json
import posixpath
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from glossary_extraction.constants import (
    AUTO_ID_HEADERS,
    AUTO_SOURCE_HEADERS,
    AUTO_TARGET_HEADERS,
    DELIMITED_MATERIAL_EXTENSIONS,
    HEADER_SCAN_LIMIT,
    IMAGE_MATERIAL_EXTENSIONS,
    TABLE_MATERIAL_EXTENSIONS,
    TEXT_MATERIAL_EXTENSIONS,
)
from glossary_extraction.heuristics import clean_text
from glossary_extraction.models import Record, SheetColumnLayout


def file_digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def set_widths(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        worksheet.column_dimensions[letter].width = max(10, min(max_len + 2, 42))


def style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    set_widths(worksheet)


def resolve_column_index(headers: list[object], expected_name: str) -> int:
    key = clean_text(expected_name).lower()
    for index, name in enumerate(headers):
        if clean_text(name).lower() == key:
            return index
    available = ", ".join(str(name) for name in headers)
    raise ValueError(f"Missing column '{expected_name}'. Available headers: {available}")


XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XLSX_NS = {"a": XLSX_MAIN_NS, "rel": PACKAGE_REL_NS}


def cell_column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - 64
    return index - 1


def xml_text(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return "".join(node.itertext())


def workbook_sheet_targets(archive: ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("rel:Relationship", XLSX_NS)
    }
    sheets: list[tuple[str, str]] = []
    for sheet in workbook_root.find("a:sheets", XLSX_NS).findall("a:sheet", XLSX_NS):
        rel_id = sheet.attrib[f"{{{XLSX_REL_NS}}}id"]
        target = rel_map[rel_id]
        target_path = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
        sheets.append((sheet.attrib["name"], target_path))
    return sheets


def load_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return [xml_text(item) for item in root.findall("a:si", XLSX_NS)]


def raw_cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value = cell.find("a:v", XLSX_NS)
        if value is None or value.text is None:
            return ""
        try:
            return shared_strings[int(value.text)]
        except (IndexError, ValueError):
            return ""
    if cell_type == "inlineStr":
        return xml_text(cell.find("a:is", XLSX_NS))
    value = cell.find("a:v", XLSX_NS)
    return "" if value is None or value.text is None else value.text


def iter_raw_xlsx_sheets(input_path: Path) -> list[tuple[str, list[list[str]]]]:
    sheets: list[tuple[str, list[list[str]]]] = []
    with ZipFile(input_path) as archive:
        shared_strings = load_shared_strings(archive)
        for sheet_name, target_path in workbook_sheet_targets(archive):
            root = ET.fromstring(archive.read(target_path))
            rows: list[list[str]] = []
            for row in root.findall(".//a:sheetData/a:row", XLSX_NS):
                cells: list[tuple[int, str]] = []
                max_column = -1
                for cell in row.findall("a:c", XLSX_NS):
                    column_index = cell_column_index(cell.attrib.get("r", ""))
                    max_column = max(max_column, column_index)
                    cells.append((column_index, raw_cell_value(cell, shared_strings)))
                values = [""] * (max_column + 1)
                for column_index, value in cells:
                    values[column_index] = value
                rows.append(values)
            sheets.append((sheet_name, rows))
    return sheets


def records_from_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> list[Record]:
    if not rows:
        return []
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        headers = list(rows[0])
        id_index = resolve_column_index(headers, id_column)
        source_index = resolve_column_index(headers, source_column)
        target_index = None if source_only else resolve_column_index(headers, target_column)
        data_rows = rows[1:]
        first_data_row_number = 2
    else:
        id_index = layout.id_index
        source_index = layout.source_index
        target_index = layout.target_index
        data_rows = rows[layout.header_row_index + 1 :]
        first_data_row_number = layout.header_row_index + 2

    records: list[Record] = []
    for row_number, row in enumerate(data_rows, start=first_data_row_number):
        row_values = list(row)
        row_id = "" if id_index >= len(row_values) or row_values[id_index] is None else str(row_values[id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        source = "" if source_index >= len(row_values) else clean_text(row_values[source_index])
        target = "" if target_index is None or target_index >= len(row_values) else clean_text(row_values[target_index])
        if not source:
            continue
        records.append(Record(row_id=row_id, source=source, target=target))
    return records


def load_records_from_raw_xlsx(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> tuple[list[Record], str]:
    sheets = iter_raw_xlsx_sheets(input_path)
    if not sheets:
        return [], ""
    selected_sheet: tuple[str, list[list[str]]] | None = None
    for candidate in sheets:
        if sheet_name is None or candidate[0] == sheet_name:
            selected_sheet = candidate
            break
    if selected_sheet is None:
        available = ", ".join(name for name, _rows in sheets)
        raise ValueError(f"Missing worksheet '{sheet_name}'. Available worksheets: {available}")
    title, rows = selected_sheet
    return records_from_rows(
        rows=rows,
        sheet_title=title,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    ), title


def normalized_header_lookup(headers: list[object]) -> dict[str, int]:
    return {clean_text(name).lower(): index for index, name in enumerate(headers)}


def first_matching_header(headers: list[object], candidates: list[str]) -> int | None:
    lookup = normalized_header_lookup(headers)
    for candidate in candidates:
        key = clean_text(candidate).lower()
        if key in lookup:
            return lookup[key]
    return None


def first_matching_header_fuzzy(headers: list[object], candidates: list[str]) -> int | None:
    exact_index = first_matching_header(headers, candidates)
    if exact_index is not None:
        return exact_index
    normalized_candidates = [clean_text(candidate).lower() for candidate in candidates if clean_text(candidate)]
    for index, header in enumerate(headers):
        header_key = clean_text(header).lower()
        if not header_key:
            continue
        if any(candidate in header_key for candidate in normalized_candidates):
            return index
    return None


def canonical_output_header(requested_header: str, default_header: str) -> str:
    clean_header = clean_text(requested_header)
    return default_header if clean_header.lower() == default_header.lower() else clean_header


def value_at(values: list[object], index: int | None) -> object:
    if index is None or index >= len(values):
        return ""
    return values[index]


def exact_sheet_column_layout(
    headers: list[str],
    header_row_index: int,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> SheetColumnLayout | None:
    try:
        id_index = resolve_column_index(headers, id_column)
        source_index = resolve_column_index(headers, source_column)
        target_index = None if source_only else resolve_column_index(headers, target_column)
    except ValueError:
        return None
    return SheetColumnLayout(
        header_row_index=header_row_index,
        headers=headers,
        id_index=id_index,
        source_index=source_index,
        target_index=target_index,
        output_indexes=list(range(len(headers))),
    )


def auto_sheet_column_layout(
    headers: list[str],
    header_row_index: int,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> SheetColumnLayout | None:
    id_index = first_matching_header_fuzzy(headers, [id_column, *AUTO_ID_HEADERS])
    source_index = first_matching_header_fuzzy(headers, [source_column, *AUTO_SOURCE_HEADERS])
    target_index = None if source_only else first_matching_header_fuzzy(headers, [target_column, *AUTO_TARGET_HEADERS])
    if id_index is None or source_index is None or (not source_only and target_index is None):
        return None
    output_indexes = [id_index, source_index]
    output_headers = ["ID", "CN"]
    if not source_only:
        output_indexes.append(target_index if target_index is not None else -1)
        output_headers.append(canonical_output_header(target_column, "EN"))
    return SheetColumnLayout(
        header_row_index=header_row_index,
        headers=output_headers,
        id_index=id_index,
        source_index=source_index,
        target_index=target_index,
        output_indexes=output_indexes,
    )


def language_table_layout_from_rows(
    rows: list[list[object]],
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> SheetColumnLayout | None:
    for header_row_index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        headers, _empty_values = trim_trailing_empty_columns(list(row))
        if not headers:
            continue
        exact_layout = exact_sheet_column_layout(
            headers=headers,
            header_row_index=header_row_index,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        if exact_layout is not None:
            return exact_layout
        auto_layout = auto_sheet_column_layout(
            headers=headers,
            header_row_index=header_row_index,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        if auto_layout is not None:
            return auto_layout
    return None


def auto_records_from_sheet_rows(sheet_title: str, rows: list[list[object]]) -> list[Record]:
    if not rows:
        return []
    headers = list(rows[0])
    source_index = first_matching_header(
        headers,
        ["简体中文", "中文", "正常对话", "资料", "简介", "内容", "说明", "描述", "cn", "source", "zh", "Chinese", "note", "description"],
    )
    if source_index is None:
        return []
    target_index = first_matching_header(
        headers,
        ["英文", "英语", "en", "English", "优化翻译"],
    )
    id_index = first_matching_header(
        headers,
        ["唯一标识ID", "ID", "id", "章节", "关卡序号"],
    )

    records: list[Record] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = list(row)
        source = "" if source_index >= len(row_values) else clean_text(row_values[source_index])
        if not source:
            continue
        target = "" if target_index is None or target_index >= len(row_values) else clean_text(row_values[target_index])
        row_id = ""
        if id_index is not None and id_index < len(row_values):
            row_id = clean_text(row_values[id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        records.append(Record(row_id=row_id, source=source, target=target))
    return records


def generic_records_from_sheet_rows(sheet_title: str, rows: list[list[object]]) -> list[Record]:
    records: list[Record] = []
    for row_number, row in enumerate(rows[1:], start=2):
        parts = [clean_text(value) for value in row if clean_text(value)]
        if not parts:
            continue
        records.append(Record(row_id=f"{sheet_title}:{row_number}", source=" ".join(parts), target=""))
    return records


def load_project_records(input_path: Path) -> list[Record]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        records: list[Record] = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            records.extend(auto_records_from_sheet_rows(worksheet.title, rows))
        workbook.close()
        return records
    except Exception:
        records = []
        for sheet_title, rows in iter_raw_xlsx_sheets(input_path):
            records.extend(auto_records_from_sheet_rows(sheet_title, rows))
        return records


def load_table_material_records(input_path: Path) -> list[Record]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        records: list[Record] = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            sheet_records = auto_records_from_sheet_rows(worksheet.title, rows)
            records.extend(sheet_records or generic_records_from_sheet_rows(worksheet.title, rows))
        workbook.close()
        return records
    except Exception:
        records = []
        for sheet_title, rows in iter_raw_xlsx_sheets(input_path):
            sheet_records = auto_records_from_sheet_rows(sheet_title, rows)
            records.extend(sheet_records or generic_records_from_sheet_rows(sheet_title, rows))
        return records


def chunk_text_material(text: str, limit: int = 160) -> list[str]:
    cleaned = clean_text(text)
    if not cleaned:
        return []
    raw_chunks = re.split(r"[\r\n]+|(?<=[。！？.!?])\s*", cleaned)
    chunks: list[str] = []
    buffer = ""
    for raw_chunk in raw_chunks:
        chunk = clean_text(raw_chunk)
        if not chunk:
            continue
        if len(chunk) > limit:
            chunks.append(chunk[:limit])
            continue
        if not buffer:
            buffer = chunk
        elif len(buffer) + len(chunk) + 1 <= limit:
            buffer = f"{buffer} {chunk}"
        else:
            chunks.append(buffer)
            buffer = chunk
    if buffer:
        chunks.append(buffer)
    return chunks[:200]


def records_from_text_material(path: Path) -> list[Record]:
    try:
        content = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        content = path.read_text(encoding="gb18030", errors="ignore")
    records: list[Record] = []
    for index, chunk in enumerate(chunk_text_material(content), start=1):
        records.append(Record(row_id=f"{path.name}:{index}", source=chunk, target=""))
    return records


DOCX_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def records_from_docx_material(path: Path) -> list[Record]:
    with ZipFile(path) as archive:
        if "word/document.xml" not in archive.namelist():
            raise ValueError(f"Missing word/document.xml in DOCX file: {path}")
        root = ET.fromstring(archive.read("word/document.xml"))

    records: list[Record] = []
    for index, paragraph in enumerate(root.findall(".//w:p", DOCX_NS), start=1):
        text = clean_text("".join(node.text or "" for node in paragraph.findall(".//w:t", DOCX_NS)))
        if text:
            records.append(Record(row_id=f"{path.name}:{index}", source=text, target=""))
    return records


def records_from_delimited_material(path: Path) -> list[Record]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        content = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        content = path.read_text(encoding="gb18030", errors="ignore")
    rows = list(csv.reader(content.splitlines(), delimiter=delimiter))
    if not rows:
        return []
    return auto_records_from_sheet_rows(path.name, rows) or generic_records_from_sheet_rows(path.name, rows)


def records_from_image_material(path: Path) -> list[Record]:
    source = " ".join(
        part
        for part in [
            "图片资料",
            path.stem.replace("_", " ").replace("-", " "),
            path.parent.name if path.parent else "",
        ]
        if part
    )
    return [Record(row_id=f"{path.name}:image", source=source, target="")]


def load_project_material_records(
    material_paths: list[Path],
    notes: list[str] | None = None,
) -> tuple[list[Record], list[str]]:
    records: list[Record] = []
    sources: list[str] = []
    for note_index, note in enumerate(notes or [], start=1):
        note_text = clean_text(note)
        if note_text:
            records.append(Record(row_id=f"project-note:{note_index}", source=note_text, target=""))
            sources.append(f"备注: {note_text[:40]}")

    for material_path in material_paths:
        path = Path(material_path)
        suffix = path.suffix.lower()
        if not path.exists():
            sources.append(f"缺失资料: {path}")
            continue
        if suffix in TABLE_MATERIAL_EXTENSIONS:
            material_records = load_table_material_records(path)
        elif suffix in DELIMITED_MATERIAL_EXTENSIONS:
            material_records = records_from_delimited_material(path)
        elif suffix in TEXT_MATERIAL_EXTENSIONS:
            material_records = records_from_text_material(path)
        elif suffix in IMAGE_MATERIAL_EXTENSIONS:
            material_records = records_from_image_material(path)
        else:
            material_records = records_from_text_material(path)

        records.extend(material_records)
        sources.append(f"{path.name} ({len(material_records)} 条)")
    return records, sources


def trim_trailing_empty_columns(headers: list[object], values: list[object] | None = None) -> tuple[list[str], list[object]]:
    last_index = len(headers) - 1
    while last_index >= 0 and not clean_text(headers[last_index]):
        last_index -= 1
    trimmed_headers = [clean_text(header) for header in headers[: last_index + 1]]
    raw_values = list(values or [])
    trimmed_values = raw_values[: len(trimmed_headers)]
    if len(trimmed_values) < len(trimmed_headers):
        trimmed_values.extend([""] * (len(trimmed_headers) - len(trimmed_values)))
    return trimmed_headers, trimmed_values


def load_records(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> tuple[list[Record], str]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        records = records_from_rows(
            rows=rows,
            sheet_title=worksheet.title,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        workbook.close()
        return records, worksheet.title
    except Exception:
        return load_records_from_raw_xlsx(
            input_path=input_path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )


def write_text_output(output_path: Path, content: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding="utf-8")


def append_rows(worksheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_sheet(worksheet)


def write_detail_workbook(
    output_path: Path,
    sheet_name: str,
    records: list[Record],
    all_rows: list[dict[str, object]],
    glossary_rows: list[dict[str, object]],
    high_risk_rows: list[dict[str, object]],
    manual_rows: list[dict[str, object]],
    curated_rules_path: Path | None,
    observations_store_path: Path | None,
) -> None:
    workbook = Workbook()
    headers = [
        "ID",
        "CN",
        "EN",
        "EN2",
        "SuggestedEN",
        "ExactCandidates",
        "ExampleUsages",
        "ManualAdaptations",
        "ActualShortUsages",
        "HasActualDiff",
        "DiffType",
        "DiffVariants",
        "SameOrFormatOnlyCount",
        "DiffCount",
        "Category",
        "Risk",
        "Priority",
        "HitRows",
        "ExactRows",
        "ExampleID",
        "ExampleSource",
        "ExampleEN",
        "DiffExampleID",
        "DiffExampleSource",
        "DiffExampleEN",
        "Note",
    ]

    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    append_rows(glossary_sheet, headers, glossary_rows)

    high_risk_sheet = workbook.create_sheet("HighRisk")
    append_rows(high_risk_sheet, headers, high_risk_rows)

    manual_sheet = workbook.create_sheet("ManualAdaptation")
    append_rows(manual_sheet, headers, manual_rows)

    all_sheet = workbook.create_sheet("Candidates")
    append_rows(all_sheet, headers, all_rows)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["Item", "Value"])
    for item, value in [
        ("SourceRows", len(records)),
        ("Sheet", sheet_name),
        ("CandidateTerms", len(all_rows)),
        ("GlossaryRows", len(glossary_rows)),
        ("HighRiskRows", len(high_risk_rows)),
        ("ManualAdaptationRows", len(manual_rows)),
        ("CuratedRules", str(curated_rules_path) if curated_rules_path else ""),
        ("ObservationsStore", str(observations_store_path) if observations_store_path else ""),
        ("Rule", "Extract short source terms from the source column and use target column only for English alignment and drift checks."),
        ("ManualAdaptation", "A term is marked as manual adaptation when short target usages introduce a stable wording different from the example EN."),
        ("LearningModel", "Curated rules keep approved EN/EN2 decisions; observation store accumulates seen variants and usage drift."),
    ]:
        notes_sheet.append([item, value])
    style_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def write_final_workbook(output_path: Path, final_rows: list[dict[str, object]]) -> None:
    workbook = Workbook()

    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    final_headers = ["ID", "CN", "EN", "EN2"]
    glossary_sheet.append(final_headers)
    for row in final_rows:
        glossary_sheet.append([row.get(header, "") for header in final_headers])
    style_sheet(glossary_sheet)

    detail_sheet = workbook.create_sheet("Buckets")
    detail_headers = ["ID", "CN", "EN", "EN2", "ExampleUsages", "ManualAdaptations", "Note"]
    detail_sheet.append(detail_headers)
    for row in final_rows:
        detail_sheet.append([row.get(header, "") for header in detail_headers])
    style_sheet(detail_sheet)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["Item", "Value"])
    for item, value in [
        ("Columns", "ID = text id, CN = source term, EN = example English, EN2 = manual adaptation English"),
        ("Rule", "EN2 remains blank when the alternative wording is not stable enough or is explicitly blocked by curated rules."),
        ("RowCount", len(final_rows)),
    ]:
        notes_sheet.append([item, value])
    style_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def display_header_name(header: str, default_header: str) -> str:
    clean_header = clean_text(header)
    return default_header if clean_header.lower() == default_header.lower() else clean_header


def write_json_output(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
