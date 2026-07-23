"""Official sentence-template extraction, matching, rendering, and QA."""

from __future__ import annotations

import html
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from glossary_extraction.constants import LOW_VALUE_ANNOUNCEMENT_TERMS
from glossary_extraction.excel_io import (
    iter_raw_xlsx_sheets,
    language_table_layout_from_rows,
    load_table_material_records,
    records_from_delimited_material,
    records_from_docx_material,
    value_at,
)


DYNAMIC_PLACEHOLDER_RE = re.compile(r"<@\d+>|\{\d+(?:,[^{}]+)?\}|%[sd]")
CJK_RUN_RE = re.compile(r"[\u4e00-\u9fff]+")
SPACE_RE = re.compile(r"\s+")
PUNCTUATION_MAP = str.maketrans(
    {
        "，": ",",
        "。": ".",
        "：": ":",
        "；": ";",
        "！": "!",
        "？": "?",
        "（": "(",
        "）": ")",
    }
)
MAX_SIMILAR_ROWS_PER_OVERLAP = 2
MAX_SIMILAR_TEMPLATE_ROWS = 20


def raw_template_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = html.unescape(text)
    return SPACE_RE.sub(" ", text).strip()


def normalize_template_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", raw_template_text(value))
    text = text.translate(PUNCTUATION_MAP)
    return SPACE_RE.sub("", text)


def placeholder_tokens(value: object) -> list[str]:
    return DYNAMIC_PLACEHOLDER_RE.findall(raw_template_text(value))


def placeholder_occurrence_keys(tokens: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    keys: list[str] = []
    for token in tokens:
        if token not in {"%s", "%d"}:
            keys.append(token)
            continue
        occurrence = counts.get(token, 0)
        counts[token] = occurrence + 1
        keys.append(f"{token}#{occurrence}")
    return keys


def structured_text_items(value: object) -> list[tuple[tuple[object, ...], str]]:
    raw = raw_template_text(value)
    if not raw or raw[0] not in "[{":
        return [((), raw)] if raw else []
    try:
        parsed = json.loads(raw)
    except (TypeError, ValueError):
        return [((), raw)]

    items: list[tuple[tuple[object, ...], str]] = []

    def walk(node: object, path: tuple[object, ...]) -> None:
        if isinstance(node, str):
            text = raw_template_text(node)
            if text:
                items.append((path, text))
            return
        if isinstance(node, list):
            for index, child in enumerate(node):
                walk(child, (*path, index))
            return
        if isinstance(node, dict):
            for key, child in node.items():
                walk(child, (*path, str(key)))

    walk(parsed, ())
    return items or [((), raw)]


def fixed_template_length(source_template: str) -> int:
    fixed = DYNAMIC_PLACEHOLDER_RE.sub("", normalize_template_text(source_template))
    return len(fixed)


def fixed_cjk_length(source_template: str) -> int:
    fixed = DYNAMIC_PLACEHOLDER_RE.sub("", normalize_template_text(source_template))
    return sum(len(run) for run in CJK_RUN_RE.findall(fixed))


def is_sentence_template_candidate(source_template: str) -> bool:
    source = raw_template_text(source_template)
    if not source:
        return False
    cjk_length = fixed_cjk_length(source)
    if placeholder_tokens(source):
        return cjk_length >= 4
    normalized = normalize_template_text(source)
    return cjk_length >= 8 and (len(normalized) >= 12 or bool(re.search(r"[,.;!?，。；！？]", source)))


def template_pattern(source_template: str) -> tuple[re.Pattern[str], dict[str, str]]:
    normalized = normalize_template_text(source_template)
    tokens = DYNAMIC_PLACEHOLDER_RE.findall(normalized)
    token_keys = placeholder_occurrence_keys(tokens)
    parts = DYNAMIC_PLACEHOLDER_RE.split(normalized)
    token_groups: dict[str, str] = {}
    pattern_parts: list[str] = []
    for index, fixed in enumerate(parts):
        pattern_parts.append(re.escape(fixed))
        if index >= len(tokens):
            continue
        token_key = token_keys[index]
        group_name = token_groups.get(token_key)
        if group_name is None:
            group_name = f"value_{len(token_groups)}"
            token_groups[token_key] = group_name
            if parts[index + 1]:
                pattern_parts.append(f"(?P<{group_name}>.+?)")
            else:
                pattern_parts.append(f"(?P<{group_name}>[^,.;!?]{{1,80}})")
        else:
            pattern_parts.append(f"(?P={group_name})")
    return re.compile("".join(pattern_parts)), token_groups


def render_official_target(
    target_template: str,
    source_token_keys: list[str],
    captured_values: dict[str, str],
) -> tuple[str, str]:
    target = raw_template_text(target_template)
    target_tokens = placeholder_tokens(target)
    target_token_keys = placeholder_occurrence_keys(target_tokens)
    if not target_tokens:
        return target, "rendered"

    target_mapping: dict[str, str] = {}
    if all(token_key in captured_values for token_key in set(target_token_keys)):
        target_mapping = {token_key: captured_values[token_key] for token_key in set(target_token_keys)}
    else:
        if len(source_token_keys) != len(target_token_keys):
            return target, "unverifiable_placeholder"
        if any(token_key not in captured_values for token_key in source_token_keys):
            return target, "unverifiable_placeholder"
        target_mapping = {
            target_token_key: captured_values[source_token_key]
            for source_token_key, target_token_key in zip(source_token_keys, target_token_keys)
        }

    target_index = 0

    def replace_target(_match: re.Match[str]) -> str:
        nonlocal target_index
        token_key = target_token_keys[target_index]
        target_index += 1
        return target_mapping[token_key]

    rendered = DYNAMIC_PLACEHOLDER_RE.sub(replace_target, target)
    return rendered, "rendered"


def longest_cjk_overlap(source_template: str, announcement_text: str, minimum: int = 4, maximum: int = 16) -> str:
    notice = normalize_template_text(announcement_text)
    best = ""
    for run in CJK_RUN_RE.findall(normalize_template_text(source_template)):
        upper = min(maximum, len(run))
        for size in range(upper, minimum - 1, -1):
            if len(best) >= size:
                break
            for start in range(0, len(run) - size + 1):
                fragment = run[start : start + size]
                if fragment in notice:
                    best = fragment
                    break
            if len(best) >= size:
                break
    return best


def build_sentence_template_matches(
    candidate_rows: list[dict[str, object]],
    announcement_text: str,
    matched_terms: list[str] | None = None,
) -> list[dict[str, object]]:
    normalized_notice = normalize_template_text(announcement_text)
    exact_candidates: list[tuple[int, int, int, dict[str, object]]] = []

    for row in candidate_rows:
        source_template = raw_template_text(row.get("CN"))
        translations = row.get("translations", {})
        if not is_sentence_template_candidate(source_template) or not isinstance(translations, dict):
            continue
        if not any(raw_template_text(value) for value in translations.values()):
            continue
        pattern, token_groups = template_pattern(source_template)
        for match in pattern.finditer(normalized_notice):
            source_token_keys = placeholder_occurrence_keys(placeholder_tokens(source_template))
            captured_values = {
                token_key: match.group(group_name)
                for token_key, group_name in token_groups.items()
            }
            rendered_translations: dict[str, str] = {}
            render_status: dict[str, str] = {}
            for language, target_template in translations.items():
                rendered, status = render_official_target(
                    raw_template_text(target_template),
                    source_token_keys,
                    captured_values,
                )
                rendered_translations[str(language)] = rendered
                render_status[str(language)] = status
            output_row = {
                "Priority": 1,
                "MatchType": "official_exact",
                "ID": raw_template_text(row.get("ID")),
                "AnnouncementCN": match.group(0),
                "OfficialCNTemplate": source_template,
                "translations": rendered_translations,
                "_render_status": render_status,
            }
            exact_candidates.append((match.start(), match.end(), fixed_template_length(source_template), output_row))

    exact_candidates.sort(key=lambda item: (-item[2], item[0], raw_template_text(item[3].get("ID"))))
    selected_spans: list[tuple[int, int]] = []
    selected_exact_ids: set[tuple[str, str]] = set()
    matches: list[dict[str, object]] = []
    for start, end, _fixed_length, row in exact_candidates:
        identity = (raw_template_text(row.get("ID")), raw_template_text(row.get("OfficialCNTemplate")))
        if identity in selected_exact_ids:
            continue
        if any(start < selected_end and end > selected_start for selected_start, selected_end in selected_spans):
            continue
        selected_spans.append((start, end))
        selected_exact_ids.add(identity)
        matches.append(row)

    normalized_low_value_terms = {normalize_template_text(term) for term in LOW_VALUE_ANNOUNCEMENT_TERMS}
    normalized_matched_terms = sorted(
        {
            normalize_template_text(term)
            for term in (matched_terms or [])
            if normalize_template_text(term) and normalize_template_text(term) not in normalized_low_value_terms
        },
        key=lambda item: (-len(item), item),
    )
    exact_coverage = " ".join(normalize_template_text(row.get("AnnouncementCN")) for row in matches)
    uncovered_terms = [term for term in normalized_matched_terms if term not in exact_coverage]

    similar_candidates: list[tuple[int, int, int, str, dict[str, object]]] = []
    for row in candidate_rows:
        source_template = raw_template_text(row.get("CN"))
        translations = row.get("translations", {})
        identity = (raw_template_text(row.get("ID")), source_template)
        if identity in selected_exact_ids:
            continue
        if not isinstance(translations, dict) or fixed_cjk_length(source_template) < 8:
            continue
        if matched_terms is None:
            overlap = longest_cjk_overlap(source_template, announcement_text)
        else:
            normalized_source = normalize_template_text(source_template)
            overlap = next((term for term in uncovered_terms if term in normalized_source), "")
        if not overlap:
            continue
        output_row = {
            "Priority": 2,
            "MatchType": "official_similar",
            "ID": raw_template_text(row.get("ID")),
            "AnnouncementCN": overlap,
            "OfficialCNTemplate": source_template,
            "translations": {str(language): raw_template_text(target) for language, target in translations.items()},
            "_render_status": {str(language): "evidence" for language in translations},
        }
        context_overlap = longest_cjk_overlap(source_template, announcement_text)
        similar_candidates.append(
            (
                normalized_notice.find(overlap),
                -max(len(overlap), len(context_overlap)),
                len(normalize_template_text(source_template)),
                overlap,
                output_row,
            )
        )

    similar_candidates.sort(key=lambda item: (item[0], item[1], item[2], raw_template_text(item[4].get("ID"))))
    candidates_by_overlap: dict[str, list[dict[str, object]]] = {}
    overlap_order: list[str] = []
    for _notice_position, _negative_context_overlap, _source_length, overlap, row in similar_candidates:
        if overlap not in candidates_by_overlap:
            candidates_by_overlap[overlap] = []
            overlap_order.append(overlap)
        candidates_by_overlap[overlap].append(row)

    seen_similar: set[tuple[str, str]] = set()
    for evidence_round in range(MAX_SIMILAR_ROWS_PER_OVERLAP):
        for overlap in overlap_order:
            rows = candidates_by_overlap[overlap]
            if evidence_round >= len(rows):
                continue
            row = rows[evidence_round]
            identity = (raw_template_text(row.get("ID")), raw_template_text(row.get("OfficialCNTemplate")))
            if identity in seen_similar:
                continue
            seen_similar.add(identity)
            matches.append(row)
            if len(matches) - len(selected_exact_ids) >= MAX_SIMILAR_TEMPLATE_ROWS:
                return matches

    return matches


def validate_official_template_usage(
    template_matches: list[dict[str, object]],
    translated_texts: dict[str, str],
) -> dict[str, object]:
    if not translated_texts:
        return {
            "status": "not_run",
            "checked": 0,
            "matches": 0,
            "mismatches": 0,
            "unverifiable_placeholders": 0,
            "issues": [],
        }

    issues: list[dict[str, str]] = []
    checked = 0
    passed = 0
    unverifiable = 0
    normalized_translations = {
        language.upper(): normalize_template_text(text)
        for language, text in translated_texts.items()
    }
    for row in template_matches:
        if int(row.get("Priority", 0)) != 1:
            continue
        raw_translations = row.get("translations", {})
        raw_statuses = row.get("_render_status", {})
        if not isinstance(raw_translations, dict) or not isinstance(raw_statuses, dict):
            continue
        translations = {str(language).upper(): value for language, value in raw_translations.items()}
        statuses = {str(language).upper(): value for language, value in raw_statuses.items()}
        for language, actual_text in normalized_translations.items():
            checked += 1
            if language not in translations:
                issues.append(
                    {
                        "ID": raw_template_text(row.get("ID")),
                        "language": language,
                        "reason": "missing_official_target",
                        "expected": "",
                    }
                )
                continue
            expected = raw_template_text(translations.get(language))
            render_status = raw_template_text(statuses.get(language))
            if render_status != "rendered":
                unverifiable += 1
                issues.append(
                    {
                        "ID": raw_template_text(row.get("ID")),
                        "language": language,
                        "reason": "unverifiable_placeholder",
                        "expected": expected,
                    }
                )
                continue
            if normalize_template_text(expected) in actual_text:
                passed += 1
                continue
            issues.append(
                {
                    "ID": raw_template_text(row.get("ID")),
                    "language": language,
                    "reason": "official_template_mismatch",
                    "expected": expected,
                }
            )

    mismatches = sum(1 for issue in issues if issue["reason"] == "official_template_mismatch")
    return {
        "status": "warning" if issues else "ok",
        "checked": checked,
        "matches": passed,
        "mismatches": mismatches,
        "unverifiable_placeholders": unverifiable,
        "issues": issues,
    }


def load_translated_material_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        records = records_from_docx_material(path)
    elif suffix in {".xlsx", ".xlsm"}:
        records = load_table_material_records(path)
    elif suffix in {".csv", ".tsv"}:
        records = records_from_delimited_material(path)
    else:
        try:
            return raw_template_text(path.read_text(encoding="utf-8-sig"))
        except UnicodeDecodeError:
            return raw_template_text(path.read_text(encoding="gb18030", errors="ignore"))
    return raw_template_text(" ".join(part for record in records for part in (record.source, record.target) if part))


def sentence_template_candidates_from_sheet_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    language: str | None = None,
    source_only: bool = False,
) -> tuple[list[str], list[dict[str, object]]]:
    if not rows:
        return [], []
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        return [], []

    output_headers = [raw_template_text(header) for header in layout.headers]
    candidates: list[dict[str, object]] = []
    for row_number, row in enumerate(rows[layout.header_row_index + 1 :], start=layout.header_row_index + 2):
        values = list(row)
        row_id = raw_template_text(value_at(values, layout.id_index)) or f"{sheet_title}:{row_number}"
        source_items = structured_text_items(value_at(values, layout.source_index))
        target_items_by_language: dict[str, list[tuple[tuple[object, ...], str]]] = {}
        if not source_only and language:
            target_items_by_language[language.upper()] = structured_text_items(value_at(values, layout.target_index))
        elif not source_only:
            for index in range(2, len(layout.output_indexes)):
                header = output_headers[index]
                if header:
                    target_items_by_language[header] = structured_text_items(
                        value_at(values, layout.output_indexes[index])
                    )

        for source_index, (source_path, source_template) in enumerate(source_items):
            if not is_sentence_template_candidate(source_template):
                continue
            translations: dict[str, str] = {}
            for target_language, target_items in target_items_by_language.items():
                target_lookup = dict(target_items)
                target = target_lookup.get(source_path, "")
                if not target and len(source_items) == len(target_items) and source_index < len(target_items):
                    target = target_items[source_index][1]
                if target:
                    translations[target_language] = target
            if not translations:
                continue
            candidates.append(
                {
                    "ID": row_id,
                    "CN": source_template,
                    "translations": translations,
                }
            )
    languages = [language.upper()] if language else output_headers[2:]
    return languages, candidates


def build_sentence_template_candidates_from_workbook(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    language: str | None = None,
    source_only: bool = False,
) -> tuple[list[str], list[dict[str, object]]]:
    languages: list[str] = []
    candidates: list[dict[str, object]] = []
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
        for worksheet in worksheets:
            sheet_languages, sheet_candidates = sentence_template_candidates_from_sheet_rows(
                rows=list(worksheet.iter_rows(values_only=True)),
                sheet_title=worksheet.title,
                id_column=id_column,
                source_column=source_column,
                target_column=target_column,
                language=language,
                source_only=source_only,
            )
            if sheet_languages and not languages:
                languages = sheet_languages
            candidates.extend(sheet_candidates)
        workbook.close()
    except Exception:
        for raw_sheet_name, rows in iter_raw_xlsx_sheets(input_path):
            if sheet_name and raw_sheet_name != sheet_name:
                continue
            sheet_languages, sheet_candidates = sentence_template_candidates_from_sheet_rows(
                rows=rows,
                sheet_title=raw_sheet_name,
                id_column=id_column,
                source_column=source_column,
                target_column=target_column,
                language=language,
                source_only=source_only,
            )
            if sheet_languages and not languages:
                languages = sheet_languages
            candidates.extend(sheet_candidates)
    return languages, merge_sentence_template_candidates(candidates)


def merge_sentence_template_candidates(
    candidate_rows: list[dict[str, object]],
    required_language: str | None = None,
) -> list[dict[str, object]]:
    merged: dict[str, dict[str, object]] = {}
    order: list[str] = []
    for candidate in candidate_rows:
        source_template = raw_template_text(candidate.get("CN"))
        key = normalize_template_text(source_template)
        if not key:
            continue
        current = merged.get(key)
        if current is None:
            current = {
                "ID": raw_template_text(candidate.get("ID")),
                "CN": source_template,
                "translations": {},
            }
            merged[key] = current
            order.append(key)
        translations = candidate.get("translations", {})
        current_translations = current["translations"]
        if isinstance(translations, dict) and isinstance(current_translations, dict):
            for language, target in translations.items():
                if raw_template_text(target) and language not in current_translations:
                    current_translations[str(language)] = raw_template_text(target)
    rows = [merged[key] for key in order]
    required = raw_template_text(required_language).upper()
    if not required:
        return rows
    return [
        row
        for row in rows
        if isinstance(row.get("translations"), dict)
        and raw_template_text(row["translations"].get(required))
    ]
