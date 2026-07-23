"""AI supplement providers, packets, response merging, and report outputs."""

from __future__ import annotations

import json
import os
import sys
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from glossary_extraction.announcement import announcement_output_values
from glossary_extraction.constants import (
    AI_CONFIDENCE_RANK,
    AI_SUPPLEMENT_EVIDENCE_LIMIT,
    AI_SUPPLEMENT_EVIDENCE_PER_TERM,
    AI_SUPPLEMENT_SCHEMA_VERSION,
    CJK_RE,
    CJK_RUN_RE,
    DEFAULT_OPENAI_RESPONSES_API_URL,
    QUOTED_TERM_RE,
)
from glossary_extraction.excel_io import (
    iter_raw_xlsx_sheets,
    language_table_layout_from_rows,
    value_at,
    write_json_output,
    write_text_output,
)
from glossary_extraction.experience import get_curated_term_state
from glossary_extraction.heuristics import clean_text
from glossary_extraction.models import LanguageTableSpec


class AiSupplementProvider:
    name = "base"

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        raise NotImplementedError


class FileAiSupplementProvider(AiSupplementProvider):
    name = "file"

    def __init__(self, response_path: Path):
        self.response_path = response_path

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        return json.loads(self.response_path.read_text(encoding="utf-8"))


class MockAiSupplementProvider(AiSupplementProvider):
    name = "mock"

    def __init__(self, response: dict[str, object]):
        self.response = response

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        return self.response


class PacketOnlyAiSupplementProvider(AiSupplementProvider):
    name = "packet"

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        return {"supplement_terms": []}


def configure_utf8_stdio() -> None:
    for stream in (getattr(sys, "stdout", None), getattr(sys, "stderr", None)):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


class OpenAiSupplementProvider(AiSupplementProvider):
    name = "openai"

    def __init__(
        self,
        api_key: str,
        model: str,
        api_url: str = DEFAULT_OPENAI_RESPONSES_API_URL,
        timeout_seconds: int = 60,
    ):
        self.api_key = api_key
        self.model = model
        self.api_url = api_url
        self.timeout_seconds = timeout_seconds

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        payload = {
            "model": self.model,
            "input": [
                {
                    "role": "system",
                    "content": (
                        "You are a game localization glossary assistant. Return JSON only. "
                        "Use only announcement_text, matched_terms, and evidence_rows from the packet. "
                        "Do not invent translations without evidence."
                    ),
                },
                {
                    "role": "user",
                    "content": json.dumps(packet, ensure_ascii=False),
                },
            ],
            "text": {
                "format": {
                    "type": "json_schema",
                    "name": "announcement_ai_supplement_response",
                    "schema": ai_supplement_response_json_schema(),
                    "strict": False,
                }
            },
        }
        request = urllib.request.Request(
            self.api_url,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout_seconds) as response:
                raw_response = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"OpenAI AI supplement request failed: HTTP {exc.code} {detail}") from exc
        except urllib.error.URLError as exc:
            raise RuntimeError(f"OpenAI AI supplement request failed: {exc.reason}") from exc
        api_response = json.loads(raw_response)
        output_text = extract_openai_output_text(api_response)
        if not output_text:
            raise RuntimeError("OpenAI AI supplement response did not contain output text")
        return json.loads(output_text)


def ai_supplement_response_json_schema() -> dict[str, object]:
    return {
        "type": "object",
        "properties": {
            "supplement_terms": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "cn": {"type": "string"},
                        "translations": {
                            "type": "object",
                            "additionalProperties": {"type": "string"},
                        },
                        "source_ids": {"type": "array", "items": {"type": "string"}},
                        "confidence": {"type": "string", "enum": ["low", "medium", "high"]},
                        "reason": {"type": "string"},
                        "evidence_ids": {"type": "array", "items": {"type": "string"}},
                        "action": {"type": "string", "enum": ["add_to_main", "report_only", "reject"]},
                    },
                    "required": [
                        "cn",
                        "translations",
                        "source_ids",
                        "confidence",
                        "reason",
                        "evidence_ids",
                        "action",
                    ],
                    "additionalProperties": False,
                },
            }
        },
        "required": ["supplement_terms"],
        "additionalProperties": False,
    }


def extract_openai_output_text(response: dict[str, object]) -> str:
    output_text = response.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text
    output = response.get("output")
    if not isinstance(output, list):
        return ""
    chunks: list[str] = []
    for item in output:
        if not isinstance(item, dict):
            continue
        content = item.get("content")
        if not isinstance(content, list):
            continue
        for part in content:
            if not isinstance(part, dict):
                continue
            text = part.get("text")
            if isinstance(text, str):
                chunks.append(text)
    return "".join(chunks).strip()


def resolve_ai_supplement_provider(
    provider_name: str,
    response_path: Path | None,
    model: str,
    api_url: str,
    timeout_seconds: int,
) -> AiSupplementProvider:
    if response_path is not None and provider_name in {"auto", "file"}:
        return FileAiSupplementProvider(response_path)
    if provider_name == "file":
        raise ValueError("--ai-supplement-provider=file requires --ai-supplement-response")
    if provider_name == "packet":
        return PacketOnlyAiSupplementProvider()
    if provider_name not in {"auto", "openai"}:
        raise ValueError(f"Unsupported AI supplement provider: {provider_name}")
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if api_key:
        return OpenAiSupplementProvider(
            api_key=api_key,
            model=model,
            api_url=api_url,
            timeout_seconds=timeout_seconds,
        )
    if provider_name == "openai":
        raise ValueError("--ai-supplement-provider=openai requires OPENAI_API_KEY")
    return PacketOnlyAiSupplementProvider()


def ai_announcement_query_terms(announcement_text: str, max_terms: int = 800) -> list[str]:
    normalized_notice = clean_text(announcement_text)
    terms: set[str] = set()
    for quoted in QUOTED_TERM_RE.findall(normalized_notice):
        quoted_term = clean_text(quoted)
        if 2 <= len(quoted_term) <= 20 and CJK_RE.search(quoted_term):
            terms.add(quoted_term)
    for run in CJK_RUN_RE.findall(normalized_notice):
        if 2 <= len(run) <= 12:
            terms.add(run)
        upper = min(8, len(run))
        for size in range(upper, 1, -1):
            for start in range(0, len(run) - size + 1):
                terms.add(run[start : start + size])
                if len(terms) >= max_terms:
                    break
            if len(terms) >= max_terms:
                break
        if len(terms) >= max_terms:
            break
    return sorted(terms, key=lambda item: (-len(item), item))[:max_terms]


def compact_announcement_row(row: dict[str, object], headers: list[str]) -> dict[str, object]:
    compact: dict[str, object] = {}
    values = announcement_output_values(row, headers)
    for index, header in enumerate(headers):
        compact[header] = values[index] if index < len(values) else ""
    return compact


def evidence_target_for_row(row: dict[str, object], headers: list[str]) -> tuple[str, str]:
    for header in headers[2:]:
        value = clean_text(row.get(header))
        if value:
            return clean_text(header), value
    en = clean_text(row.get("EN")) or clean_text(row.get("EN2"))
    if en:
        return "EN", en
    source_values = row.get("_AnnouncementValues")
    if isinstance(source_values, list) and len(source_values) >= 3:
        return clean_text(headers[2] if len(headers) >= 3 else "EN"), clean_text(source_values[2])
    return clean_text(headers[2] if len(headers) >= 3 else "EN"), ""


def ai_evidence_candidate_rows_from_sheet_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    language: str,
    source_only: bool = False,
) -> list[dict[str, object]]:
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        return []
    evidence_rows: list[dict[str, object]] = []
    language_header = clean_text(language) or "EN"
    for row_number, row in enumerate(rows[layout.header_row_index + 1 :], start=layout.header_row_index + 2):
        row_values = list(row)
        source_text = clean_text(value_at(row_values, layout.source_index))
        if not source_text or not CJK_RE.search(source_text):
            continue
        target_text = "" if layout.target_index is None else clean_text(value_at(row_values, layout.target_index))
        if not target_text:
            continue
        row_id = clean_text(value_at(row_values, layout.id_index)) or f"{sheet_title}:{row_number}"
        evidence_rows.append(
            {
                "ID": row_id,
                "CN": source_text,
                language_header: target_text,
                "EN": target_text if language_header == "EN" else "",
            }
        )
    return evidence_rows


def build_ai_evidence_candidate_rows_from_workbook(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    language: str,
    source_only: bool = False,
) -> list[dict[str, object]]:
    evidence_rows: list[dict[str, object]] = []
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
        for worksheet in worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            evidence_rows.extend(
                ai_evidence_candidate_rows_from_sheet_rows(
                    rows=rows,
                    sheet_title=worksheet.title,
                    id_column=id_column,
                    source_column=source_column,
                    target_column=target_column,
                    language=language,
                    source_only=source_only,
                )
            )
        workbook.close()
        return evidence_rows
    except Exception:
        for raw_sheet_name, rows in iter_raw_xlsx_sheets(input_path):
            if sheet_name and raw_sheet_name != sheet_name:
                continue
            evidence_rows.extend(
                ai_evidence_candidate_rows_from_sheet_rows(
                    rows=rows,
                    sheet_title=raw_sheet_name,
                    id_column=id_column,
                    source_column=source_column,
                    target_column=target_column,
                    language=language,
                    source_only=source_only,
                )
            )
        return evidence_rows


def build_ai_supplement_packet(
    announcement_text: str,
    matched_rows: list[dict[str, object]],
    candidate_rows: list[dict[str, object]],
    headers: list[str],
    project_name: str = "",
    evidence_limit: int = AI_SUPPLEMENT_EVIDENCE_LIMIT,
    evidence_per_term: int = AI_SUPPLEMENT_EVIDENCE_PER_TERM,
    sentence_template_matches: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    normalized_notice = clean_text(announcement_text)
    matched_terms = {clean_text(row.get("CN")) for row in matched_rows if clean_text(row.get("CN"))}
    query_terms = [term for term in ai_announcement_query_terms(normalized_notice) if term not in matched_terms]
    evidence_rows: list[dict[str, object]] = []
    per_term_counts: Counter[str] = Counter()
    seen_evidence_ids: Counter[str] = Counter()

    for row in candidate_rows:
        source_text = clean_text(row.get("CN"))
        if not source_text or source_text in matched_terms:
            continue
        matched_query = next((term for term in query_terms if term in source_text), "")
        if not matched_query:
            continue
        if per_term_counts[matched_query] >= evidence_per_term:
            continue
        language, target_text = evidence_target_for_row(row, headers)
        if not target_text:
            continue
        raw_id = clean_text(row.get("ID")) or f"evidence-{len(evidence_rows) + 1}"
        seen_evidence_ids[raw_id] += 1
        evidence_id = raw_id if seen_evidence_ids[raw_id] == 1 else f"{raw_id}#{seen_evidence_ids[raw_id]}"
        evidence_rows.append(
            {
                "evidence_id": evidence_id,
                "ID": raw_id,
                "source_text": source_text,
                "target_text": target_text,
                "language": language,
                "reason": f"announcement_overlap:{matched_query}",
            }
        )
        per_term_counts[matched_query] += 1
        if len(evidence_rows) >= evidence_limit:
            break

    uncovered_text = normalized_notice
    for term in sorted(matched_terms, key=len, reverse=True):
        uncovered_text = uncovered_text.replace(term, "")
    sentence_template_matches = sentence_template_matches or []
    official_sentence_matches = [
        row for row in sentence_template_matches if row.get("MatchType") == "official_exact"
    ]
    official_context_evidence = [
        row for row in sentence_template_matches if row.get("MatchType") == "official_similar"
    ]
    packet = {
        "schema_version": AI_SUPPLEMENT_SCHEMA_VERSION,
        "task": "announcement_ai_supplement",
        "instructions": [
            "Reuse official sentence matches first, then official similar-context evidence, then matched atomic terms, and only then use model reasoning for uncovered text.",
            "Only propose terms that appear in announcement_text.",
            "Prefer game-specific system, event, item, mode, character, and proper-name terms.",
            "Use evidence_rows only; do not invent translations without language-table evidence.",
            "Return JSON with supplement_terms: cn, translations, source_ids, confidence, reason, evidence_ids, action.",
        ],
        "project_name": clean_text(project_name),
        "announcement_text": normalized_notice,
        "uncovered_announcement_text": clean_text(uncovered_text),
        "headers": headers,
        "official_sentence_matches": official_sentence_matches,
        "official_context_evidence": official_context_evidence,
        "matched_terms": [compact_announcement_row(row, headers) for row in matched_rows],
        "evidence_rows": evidence_rows,
        "response_schema": {
            "supplement_terms": [
                {
                    "cn": "术语中文",
                    "translations": {"EN": "Term translation"},
                    "source_ids": ["language-table ID"],
                    "confidence": "low|medium|high",
                    "reason": "why this is a term",
                    "evidence_ids": ["evidence_id"],
                    "action": "add_to_main|report_only|reject",
                }
            ]
        },
    }
    return packet


def ai_response_terms(response: dict[str, object]) -> list[dict[str, object]]:
    terms = response.get("supplement_terms", [])
    if not isinstance(terms, list):
        return []
    return [term for term in terms if isinstance(term, dict)]


def evidence_lookup(packet: dict[str, object]) -> dict[str, dict[str, object]]:
    lookup: dict[str, dict[str, object]] = {}
    evidence_rows = packet.get("evidence_rows", [])
    if not isinstance(evidence_rows, list):
        return lookup
    for item in evidence_rows:
        if not isinstance(item, dict):
            continue
        evidence_id = clean_text(item.get("evidence_id"))
        row_id = clean_text(item.get("ID"))
        if evidence_id:
            lookup[evidence_id] = item
        if row_id:
            lookup[row_id] = item
    return lookup


def project_name_translation_missing(project_name: str, rows: list[dict[str, object]], headers: list[str]) -> bool:
    normalized_project = clean_text(project_name)
    if not normalized_project:
        return False
    for row in rows:
        if clean_text(row.get("CN")) != normalized_project:
            continue
        if any(clean_text(row.get(header)) for header in headers[2:]):
            return False
        if clean_text(row.get("EN")) or clean_text(row.get("EN2")):
            return False
    return True


def apply_ai_supplement_response(
    announcement_rows: list[dict[str, object]],
    headers: list[str],
    announcement_text: str,
    packet: dict[str, object],
    response: dict[str, object],
    project_name: str = "",
) -> tuple[list[dict[str, object]], dict[str, object]]:
    normalized_notice = clean_text(announcement_text)
    evidence_by_id = evidence_lookup(packet)
    merged_rows = [dict(row) for row in announcement_rows]
    existing_terms = {clean_text(row.get("CN")) for row in merged_rows if clean_text(row.get("CN"))}
    report_terms: list[dict[str, object]] = []

    for term in ai_response_terms(response):
        cn = clean_text(term.get("cn"))
        action = clean_text(term.get("action")) or "report_only"
        confidence = clean_text(term.get("confidence")).lower()
        translations = term.get("translations", {})
        translations = translations if isinstance(translations, dict) else {}
        evidence_ids = term.get("evidence_ids", [])
        source_ids = term.get("source_ids", [])
        evidence_keys = [
            clean_text(item)
            for item in ([*evidence_ids, *source_ids] if isinstance(evidence_ids, list) and isinstance(source_ids, list) else [])
            if clean_text(item)
        ]
        evidence_items = [evidence_by_id[key] for key in evidence_keys if key in evidence_by_id]
        has_evidence = any(cn and cn in clean_text(item.get("source_text")) for item in evidence_items)
        has_translation = any(clean_text(translations.get(header)) for header in headers[2:])
        if not has_translation:
            has_translation = any(clean_text(value) for value in translations.values())
        missing_languages = [header for header in headers[2:] if not clean_text(translations.get(header))]
        can_add = (
            action == "add_to_main"
            and cn
            and cn in normalized_notice
            and cn not in existing_terms
            and AI_CONFIDENCE_RANK.get(confidence, -1) >= AI_CONFIDENCE_RANK["medium"]
            and has_evidence
            and has_translation
        )
        status = "added_to_main" if can_add else ("rejected" if action == "reject" else "report_only")
        report_terms.append(
            {
                "cn": cn,
                "confidence": confidence,
                "action": action,
                "status": status,
                "reason": clean_text(term.get("reason")),
                "evidence_ids": evidence_keys,
                "missing_languages": missing_languages,
                "translations": {str(key): clean_text(value) for key, value in translations.items()},
            }
        )
        if not can_add:
            continue

        first_evidence = evidence_items[0]
        output_row: dict[str, object] = {
            "ID": clean_text(first_evidence.get("ID")),
            "CN": cn,
        }
        for header in headers[2:]:
            output_row[header] = clean_text(translations.get(header))
        if "EN" in headers and not clean_text(output_row.get("EN")):
            output_row["EN"] = clean_text(translations.get("EN"))
        merged_rows.append(output_row)
        existing_terms.add(cn)

    missing_project_name = project_name_translation_missing(project_name, merged_rows, headers)
    report = {
        "schema_version": AI_SUPPLEMENT_SCHEMA_VERSION,
        "terms": report_terms,
        "project_name": clean_text(project_name),
        "project_name_translation_missing": missing_project_name,
    }
    return merged_rows, report


def build_multilingual_ai_candidate_rows(
    language_table_specs: list[LanguageTableSpec],
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    curated_rules: dict[str, Any],
    announcement_min_hit: int,
    source_only: bool,
) -> list[dict[str, object]]:
    rows_by_cn: dict[str, dict[str, object]] = {}
    for spec in language_table_specs:
        candidate_rows = build_ai_evidence_candidate_rows_from_workbook(
            input_path=spec.path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=spec.language,
            language=spec.language,
            source_only=source_only,
        )
        cn_counts = Counter(clean_text(row.get("CN")) for row in candidate_rows if clean_text(row.get("CN")))
        for candidate in candidate_rows:
            cn = clean_text(candidate.get("CN"))
            if not cn:
                continue
            if cn_counts[cn] < announcement_min_hit:
                continue
            curated_state = get_curated_term_state(curated_rules, cn, create=False)
            if curated_state.get("ignore"):
                continue
            row = rows_by_cn.setdefault(cn, {"ID": candidate.get("ID", ""), "CN": cn})
            target = clean_text(candidate.get(spec.language)) or clean_text(candidate.get("EN"))
            if target:
                row[spec.language] = target
                if spec.language == "EN":
                    row["EN"] = target
    return list(rows_by_cn.values())


def build_ai_supplement_report_markdown(
    report: dict[str, object],
    packet_path: Path | None,
    response_path: Path | None,
    output_path: Path,
) -> str:
    terms = report.get("terms", [])
    terms = terms if isinstance(terms, list) else []
    lines = [
        "# AI Supplement Report",
        "",
        f"status: {report.get('status', 'ok')}",
        f"provider: {report.get('provider', 'unknown')}",
        f"packet: {packet_path or 'disabled'}",
        f"response: {response_path or 'not provided'}",
        f"output: {output_path}",
        f"term_count: {len(terms)}",
        f"added_to_main: {sum(1 for term in terms if isinstance(term, dict) and term.get('status') == 'added_to_main')}",
        f"report_only: {sum(1 for term in terms if isinstance(term, dict) and term.get('status') == 'report_only')}",
        "",
    ]
    provider_error = clean_text(report.get("provider_error"))
    if provider_error:
        lines.extend(["## Provider Error", "", provider_error, ""])
    if report.get("project_name_translation_missing"):
        lines.extend(
            [
                "## Project Name Warning",
                "",
                f"请补充项目名标准译文：{report.get('project_name', '')}",
                "",
            ]
        )
    lines.append("## Terms")
    if not terms:
        lines.append("")
        lines.append("- No AI supplement response terms.")
    for term in terms:
        if not isinstance(term, dict):
            continue
        lines.append(
            f"- {term.get('cn', '')} | status={term.get('status', '')} | confidence={term.get('confidence', '')} | evidence={', '.join(term.get('evidence_ids', [])) if isinstance(term.get('evidence_ids'), list) else ''}"
        )
        missing_languages = term.get("missing_languages", [])
        if isinstance(missing_languages, list) and missing_languages:
            lines.append(f"  missing_languages: {', '.join(str(item) for item in missing_languages)}")
        reason = clean_text(term.get("reason"))
        if reason:
            lines.append(f"  reason: {reason}")
    return "\n".join(lines) + "\n"


def run_ai_supplement_flow(
    announcement_rows: list[dict[str, object]],
    announcement_candidate_rows: list[dict[str, object]],
    announcement_text: str,
    headers: list[str],
    project_name: str,
    packet_output_path: Path,
    report_output_path: Path,
    response_path: Path | None,
    provider: AiSupplementProvider,
    sentence_template_matches: list[dict[str, object]] | None = None,
) -> tuple[list[dict[str, object]], dict[str, object], Path, Path]:
    packet = build_ai_supplement_packet(
        announcement_text=announcement_text,
        matched_rows=announcement_rows,
        candidate_rows=announcement_candidate_rows,
        headers=headers,
        project_name=project_name,
        sentence_template_matches=sentence_template_matches,
    )
    write_json_output(packet_output_path, packet)
    response: dict[str, object] = {"supplement_terms": []}
    provider_error = ""
    try:
        response = provider.generate(packet)
    except Exception as exc:
        provider_error = str(exc)
    merged_rows, report = apply_ai_supplement_response(
        announcement_rows=announcement_rows,
        headers=headers,
        announcement_text=announcement_text,
        packet=packet,
        response=response,
        project_name=project_name,
    )
    report["provider"] = provider.name
    if provider_error:
        report["status"] = "provider_error"
        report["provider_error"] = provider_error
    report_markdown = build_ai_supplement_report_markdown(
        report=report,
        packet_path=packet_output_path,
        response_path=response_path,
        output_path=report_output_path,
    )
    write_text_output(report_output_path, report_markdown)
    return merged_rows, report, packet_output_path, report_output_path
