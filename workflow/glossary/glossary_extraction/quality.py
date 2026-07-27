"""Readback quality gates for clean glossary delivery workbooks."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from glossary_extraction.name_policy import normalized_name, plain_text


VALID_DELIVERY_CATEGORIES = frozenset(
    {
        "活动",
        "UI",
        "动作",
        "装备",
        "道具",
        "资源",
        "品质",
        "属性",
        "技能",
        "技能名",
        "地名",
        "纹章",
        "副本",
        "联盟",
        "英雄",
        "怪物",
        "宠物",
        "世界观",
        "邮件",
    }
)


@dataclass(frozen=True)
class DeliveryQualityReport:
    row_count: int
    blank_cn: int
    blank_target: int
    duplicate_cn: int
    blank_category: int
    invalid_category: int
    name_collisions: int
    structure_errors: int
    hard_blockers: int


def readback_delivery_workbook(
    path: Path,
    target_header: str,
    require_target: bool = True,
    include_en2: bool = False,
) -> DeliveryQualityReport:
    workbook = load_workbook(path, read_only=True, data_only=True)
    expected_headers = ["ID", "CN", target_header]
    if include_en2:
        expected_headers.append("EN2")
    expected_headers.append("分类")

    structure_errors = 0
    if workbook.sheetnames != ["Glossary"]:
        structure_errors += 1
    worksheet = workbook["Glossary"] if "Glossary" in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [plain_text(value) for value in rows[0]] if rows else []
    if headers != expected_headers:
        structure_errors += 1

    data_rows = rows[1:] if rows else []
    cn_index = 1
    target_index = 2
    category_index = len(expected_headers) - 1
    cn_values = [plain_text(row[cn_index] if len(row) > cn_index else "") for row in data_rows]
    target_values = [
        plain_text(row[target_index] if len(row) > target_index else "") for row in data_rows
    ]
    category_values = [
        plain_text(row[category_index] if len(row) > category_index else "") for row in data_rows
    ]

    cn_counter = Counter(value for value in cn_values if value)
    duplicate_cn = sum(count - 1 for count in cn_counter.values() if count > 1)
    names: dict[str, set[str]] = defaultdict(set)
    for cn, target, category in zip(cn_values, target_values, category_values):
        if category not in {"技能名", "地名"}:
            continue
        key = normalized_name(target)
        if key and cn:
            names[key].add(cn)
    name_collisions = sum(1 for cn_set in names.values() if len(cn_set) > 1)

    blank_cn = sum(1 for value in cn_values if not value)
    blank_target = sum(1 for value in target_values if require_target and not value)
    blank_category = sum(1 for value in category_values if not value)
    invalid_category = sum(
        1 for value in category_values if value and value not in VALID_DELIVERY_CATEGORIES
    )
    hard_blockers = (
        blank_cn
        + blank_target
        + duplicate_cn
        + blank_category
        + invalid_category
        + name_collisions
        + structure_errors
    )
    workbook.close()
    return DeliveryQualityReport(
        row_count=len(data_rows),
        blank_cn=blank_cn,
        blank_target=blank_target,
        duplicate_cn=duplicate_cn,
        blank_category=blank_category,
        invalid_category=invalid_category,
        name_collisions=name_collisions,
        structure_errors=structure_errors,
        hard_blockers=hard_blockers,
    )
