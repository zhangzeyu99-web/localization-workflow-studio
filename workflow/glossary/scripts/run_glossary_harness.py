from __future__ import annotations

import argparse
import json
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

import extract_glossary as extractor


def write_fixture_workbook(path: Path, fixture: dict[str, Any]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = fixture.get("sheet", "Sheet0")
    columns = fixture.get("columns", {"id": "ID", "source": "cn", "target": "en"})
    worksheet.append([columns["id"], columns["source"], columns["target"]])
    for row in fixture["rows"]:
        worksheet.append([row[columns["id"]], row[columns["source"]], row[columns["target"]]])
    workbook.save(path)
    workbook.close()


def index_final_rows(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {row["CN"]: row for row in rows}


def write_announcement_language_table(path: Path, table: dict[str, Any]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = table.get("sheet", "Sheet0")
    headers = table.get("headers", ["ID", "CN", table["language"]])
    for metadata_row in table.get("metadata_rows", []):
        worksheet.append(metadata_row)
    worksheet.append(headers)
    for row in table["rows"]:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def evaluate_announcement_fixture(fixture_path: Path, fixture: dict[str, Any]) -> dict[str, Any]:
    expected_rows = fixture.get("expected_rows", [])
    expected_absent = set(fixture.get("expected_absent", []))
    strict_terms = bool(fixture.get("strict_terms", True))
    announcement = fixture["announcement"]

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        notice_path = temp_root / "notice.txt"
        output_path = temp_root / "announcement_terms.xlsx"
        notice_path.write_text(announcement, encoding="utf-8")

        language_specs: list[extractor.LanguageTableSpec] = []
        for table in fixture["language_tables"]:
            table_path = temp_root / f"{table['language'].lower()}_language.xlsx"
            write_announcement_language_table(table_path, table)
            language_specs.append(extractor.LanguageTableSpec(language=table["language"], path=table_path))

        rows, stats = extractor.build_multilingual_announcement_rows(
            language_table_specs=language_specs,
            sheet_name=None,
            id_column="ID",
            source_column="CN",
            curated_rules=extractor.new_curated_rules(),
            announcement_min_hit=int(fixture.get("announcement_min_hit", 1)),
            source_only=False,
            announcement_text=announcement,
            include_empty=False,
        )
        headers = ["ID", "CN", *[spec.language for spec in language_specs]]
        ai_report: dict[str, Any] | None = None
        if "ai_response" in fixture:
            ai_candidate_rows = extractor.build_multilingual_ai_candidate_rows(
                language_table_specs=language_specs,
                sheet_name=None,
                id_column="ID",
                source_column="CN",
                curated_rules=extractor.new_curated_rules(),
                announcement_min_hit=int(fixture.get("announcement_min_hit", 1)),
                source_only=False,
            )
            packet = extractor.build_ai_supplement_packet(
                announcement_text=announcement,
                matched_rows=rows,
                candidate_rows=ai_candidate_rows,
                headers=headers,
                project_name=fixture.get("project_name", ""),
            )
            rows, ai_report = extractor.apply_ai_supplement_response(
                announcement_rows=rows,
                headers=headers,
                announcement_text=announcement,
                packet=packet,
                response=fixture["ai_response"],
                project_name=fixture.get("project_name", ""),
            )
        extractor.write_announcement_glossary_workbook(
            output_path=output_path,
            matched_rows=rows,
            id_header="ID",
            source_header="CN",
            target_header=language_specs[0].language if language_specs else "EN",
            headers=headers,
        )

        validation_candidates = list(temp_root.glob("*_announcement_validation_*.md"))
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        sheet_rows = list(workbook["Glossary"].iter_rows(values_only=True))
        workbook.close()

    produced_headers = list(sheet_rows[0]) if sheet_rows else []
    produced_rows = [dict(zip(produced_headers, row)) for row in sheet_rows[1:]]
    produced = {row.get("CN", ""): row for row in produced_rows}
    expected = {item["CN"]: item for item in expected_rows}
    missing_terms: list[str] = []
    mismatched_terms: list[dict[str, Any]] = []

    for term, expected_item in expected.items():
        predicted = produced.get(term)
        if not predicted:
            missing_terms.append(term)
            continue
        for header, expected_value in expected_item.items():
            if predicted.get(header, "") != expected_value:
                mismatched_terms.append(
                    {
                        "CN": term,
                        "header": header,
                        "expected": expected_value,
                        "actual": predicted.get(header, ""),
                    }
                )

    absent_hits = sorted(term for term in expected_absent if term in produced)
    unexpected_terms = sorted(term for term in produced if strict_terms and term not in expected and term not in expected_absent)
    exact_matches = len(expected) - len(missing_terms) - len(mismatched_terms)
    expected_count = len(expected)
    produced_count = len(produced)

    return {
        "fixture": fixture_path.name,
        "mode": "announcement_lookup",
        "pass": not missing_terms and not mismatched_terms and not absent_hits and not unexpected_terms and not validation_candidates,
        "expected_count": expected_count,
        "produced_count": produced_count,
        "exact_match_count": exact_matches,
        "coverage": round(exact_matches / expected_count, 4) if expected_count else 1.0,
        "headers": produced_headers,
        "missing_terms": missing_terms,
        "mismatched_terms": mismatched_terms,
        "absent_hits": absent_hits,
        "unexpected_terms": unexpected_terms,
        "validation_created": bool(validation_candidates),
        "candidate_terms": stats.get("candidate_terms", 0),
        "duplicate_source_terms": stats.get("duplicate_source_terms", 0),
        "ai_added_to_main": sum(
            1
            for term in (ai_report or {}).get("terms", [])
            if isinstance(term, dict) and term.get("status") == "added_to_main"
        ),
        "project_name_translation_missing": bool((ai_report or {}).get("project_name_translation_missing")),
    }


def evaluate_fixture(fixture_path: Path) -> dict[str, Any]:
    fixture = json.loads(fixture_path.read_text(encoding="utf-8"))
    if fixture.get("mode") == "announcement_lookup":
        return evaluate_announcement_fixture(fixture_path, fixture)

    expected_final = fixture.get("expected_final", [])
    expected_absent = set(fixture.get("expected_absent", []))
    strict_terms = bool(fixture.get("strict_terms", True))
    columns = fixture.get("columns", {"id": "ID", "source": "cn", "target": "en"})
    extract_config = fixture.get("extract", {})

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        workbook_path = temp_root / "fixture.xlsx"
        curated_path = temp_root / "curated.json"
        observations_path = temp_root / "observations.json"
        write_fixture_workbook(workbook_path, fixture)
        curated_rules = fixture.get("curated_rules")
        observations_store = fixture.get("observations_store")
        legacy_memory = fixture.get("memory")
        if legacy_memory and (curated_rules is None and observations_store is None):
            curated_rules, observations_store = extractor.split_legacy_term_memory(legacy_memory)
        extractor.save_curated_rules(curated_path, curated_rules or extractor.new_curated_rules())
        extractor.save_observation_store(observations_path, observations_store or extractor.new_observation_store())

        records, _sheet_name = extractor.load_records(
            input_path=workbook_path,
            sheet_name=fixture.get("sheet"),
            id_column=columns["id"],
            source_column=columns["source"],
            target_column=columns["target"],
        )
        curated = extractor.load_curated_rules(curated_path)
        observations = extractor.load_observation_store(observations_path)
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = extractor.build_term_rows(
            records=records,
            min_hit=int(extract_config.get("min_hit", 1)),
            glossary_hit_threshold=int(extract_config.get("glossary_hit_threshold", 1)),
            curated_rules=curated,
            observations_store=observations,
            input_digest=extractor.file_digest(workbook_path),
        )

    produced = index_final_rows(final_rows)
    expected = {item["CN"]: item for item in expected_final}
    missing_terms: list[str] = []
    mismatched_terms: list[dict[str, Any]] = []

    for term, expected_item in expected.items():
        predicted = produced.get(term)
        if not predicted:
            missing_terms.append(term)
            continue
        expected_en = expected_item.get("EN", "")
        expected_en2 = expected_item.get("EN2", "")
        if predicted.get("EN", "") != expected_en or predicted.get("EN2", "") != expected_en2:
            mismatched_terms.append(
                {
                    "CN": term,
                    "expected": {"EN": expected_en, "EN2": expected_en2},
                    "actual": {"EN": predicted.get("EN", ""), "EN2": predicted.get("EN2", "")},
                }
            )

    absent_hits = sorted(term for term in expected_absent if term in produced)
    unexpected_terms = sorted(term for term in produced if strict_terms and term not in expected and term not in expected_absent)
    exact_matches = len(expected) - len(missing_terms) - len(mismatched_terms)
    expected_count = len(expected)
    produced_count = len(produced)

    return {
        "fixture": fixture_path.name,
        "pass": not missing_terms and not mismatched_terms and not absent_hits and not unexpected_terms,
        "expected_count": expected_count,
        "produced_count": produced_count,
        "exact_match_count": exact_matches,
        "coverage": round(exact_matches / expected_count, 4) if expected_count else 1.0,
        "missing_terms": missing_terms,
        "mismatched_terms": mismatched_terms,
        "absent_hits": absent_hits,
        "unexpected_terms": unexpected_terms,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run glossary regression harness against JSON fixtures.")
    parser.add_argument("fixtures", nargs="+", help="Fixture JSON file(s) to evaluate.")
    parser.add_argument("--report-output", help="Optional path to write the JSON report.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    reports = [evaluate_fixture(Path(path)) for path in args.fixtures]
    summary = {
        "all_passed": all(report["pass"] for report in reports),
        "reports": reports,
    }
    text = json.dumps(summary, ensure_ascii=False, indent=2)
    if args.report_output:
        Path(args.report_output).write_text(text, encoding="utf-8")
    print(text)
    return 0 if summary["all_passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
