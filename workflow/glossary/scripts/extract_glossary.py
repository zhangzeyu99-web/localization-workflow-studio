from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import posixpath
import re
import sys
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CURATED_RULES = REPO_ROOT / "data" / "experience" / "curated_terms.json"
DEFAULT_OBSERVATIONS_STORE = REPO_ROOT / "data" / "experience" / "observed_terms.json"
DEFAULT_LEGACY_EXPERIENCE_STORE = REPO_ROOT / "data" / "experience" / "term_memory.json"
CURATED_VERSION = 1
OBSERVATION_VERSION = 1

SENTENCE_PUNCT_RE = re.compile(r"[，。！？；：,.!?;:\n]")
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
PLACEHOLDER_RE = re.compile(r"\{\d+\}|%[sd]|\\n")
BRACKET_TAG_RE = re.compile(r"\[[^\]]*\]")
HTML_TAG_RE = re.compile(r"<[^>]+>")
SPACE_RE = re.compile(r"\s+")
NON_TERM_RE = re.compile(r"^[\W_]+$|^[IVXⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+$")
CAMEL_SPLIT_RE = re.compile(r"(?<=[a-z])(?=[A-Z])")
EN_COMPARE_RE = re.compile(r"[^a-z0-9+ ]+")
EN_WORD_RE = re.compile(r"[a-z0-9+]+")
NUMBERED_TITLE_RE = re.compile(r"^\s*(?:[0-9]+|[\u4e00\u4e8c\u4e09\u56db\u4e94\u516d\u4e03\u516b\u4e5d\u5341]+)\s*[\u3001.．\s-]+")
HEADER_SCAN_LIMIT = 50

AUTO_ID_HEADERS = ["ID", "id", "\u7d22\u5f15ID", "\u552f\u4e00\u6807\u8bc6ID"]
AUTO_SOURCE_HEADERS = ["CN", "cn", "zh", "source", "Chinese", "\u4e2d\u6587", "\u7b80\u4f53\u4e2d\u6587", "ori_string"]
AUTO_TARGET_HEADERS = ["EN", "en", "target", "translation", "English", "\u82f1\u6587", "\u82f1\u8bed", "\u5185\u5bb9", "text"]
LOW_VALUE_ANNOUNCEMENT_TERMS = {
    "\u73a9\u5bb6",
    "\u6d3b\u52a8",
    "\u4e16\u754c",
    "\u8d2d\u4e70",
    "\u53d1\u9001",
    "\u67e5\u770b",
    "\u83b7\u5f97",
    "\u5956\u52b1",
    "\u9884\u544a",
    "\u7cfb\u7edf",
}
AI_CONFIDENCE_RANK = {"low": 0, "medium": 1, "high": 2}
AI_SUPPLEMENT_SCHEMA_VERSION = 1
AI_SUPPLEMENT_EVIDENCE_LIMIT = 80
AI_SUPPLEMENT_EVIDENCE_PER_TERM = 3
CJK_RUN_RE = re.compile(r"[\u4e00-\u9fff]{2,}")
QUOTED_TERM_RE = re.compile(r"[《「『“\"']([^《》「」『』“”\"']{2,20})[》」』”\"']")

RARITY_TERMS = {
    "普通",
    "精良",
    "精英",
    "卓越",
    "史诗",
    "传说",
    "神话",
    "高级",
}
RESOURCE_TERMS = {
    "钻石",
    "石油",
    "体力",
    "宝石",
    "积分",
    "碎片",
    "材料",
    "芯片",
    "能量",
    "金币",
    "经验",
    "奖励",
    "礼包",
}
STAT_TERMS = {
    "攻击",
    "攻击力",
    "防御",
    "生命",
    "伤害",
    "伤害+",
    "暴击",
    "暴击伤害",
    "闪电",
    "火焰",
    "冰霜",
    "物理",
    "闪电伤害",
    "火焰伤害",
    "冰霜伤害",
    "物理伤害",
    "能量伤害",
}
ACTION_TERMS = {
    "获得",
    "获取",
    "领取",
    "使用",
    "合成",
    "升级",
    "强化",
    "突破",
    "升星",
    "激活",
    "解锁",
    "购买",
    "报名",
    "兑换",
    "刷新",
    "前往",
    "开始",
    "加入",
    "退出",
    "上阵",
    "建造",
    "邀请",
    "重置",
    "探索",
    "挑战",
}
SYSTEM_TERMS = {
    "公会",
    "竞技场",
    "战令",
    "签到",
    "商城",
    "商店",
    "背包",
    "排行",
    "排行榜",
    "活动",
    "防御塔",
    "兵营",
    "基地",
    "据点",
    "任务",
}
OBJECT_TERMS = {
    "英雄",
    "技能",
    "装备",
    "建筑",
    "武器",
    "士兵",
    "坐骑",
    "收藏品",
    "头像",
    "好友",
    "道具",
}
STATUS_TERMS = {
    "当前",
    "暂无",
    "不足",
    "已满",
    "失败",
    "可领取",
    "拥有",
    "最大",
    "剩余",
    "排名",
    "段位",
}

HIGH_CONFUSION_TERMS = (
    RARITY_TERMS
    | RESOURCE_TERMS
    | ACTION_TERMS
    | STATUS_TERMS
    | {
        "战力",
        "品质",
        "等级",
        "积分",
        "攻击",
        "觉醒",
        "奖励",
        "选择",
        "强化",
        "合成",
        "突破",
        "推荐",
    }
)

PROJECT_SIGNAL_GROUPS = {
    "合成/经营": {
        "合成",
        "订单",
        "生产",
        "生产机",
        "生产机器",
        "生成器",
        "仓库",
        "菜品",
        "烹饪",
        "顾客",
        "Merge",
    },
    "花店/装修": {
        "花店",
        "花束",
        "鲜花",
        "玫瑰",
        "百合",
        "装饰",
        "装修",
        "修复",
        "翻新",
        "花园",
        "Florist",
    },
    "休闲/女性向": {
        "女性向",
        "恋爱",
        "恋综",
        "都市",
        "时尚",
        "可爱",
        "漂亮",
        "温馨",
        "甜",
        "咖啡",
        "甜品",
        "裙",
        "珠宝",
        "约会",
        "浪漫",
        "美女",
        "小姐",
        "romance",
        "fashion",
        "cozy",
        "cute",
    },
    "战斗/RPG养成": {
        "战斗",
        "攻击",
        "防御",
        "生命",
        "伤害",
        "暴击",
        "技能",
        "英雄",
        "装备",
        "武器",
        "首领",
        "BOSS",
        "怪物",
        "关卡",
        "挑战",
    },
    "基地/建筑经营": {
        "基地",
        "建筑",
        "兵营",
        "防御塔",
        "建造",
        "升级",
        "营地",
        "总部",
        "据点",
        "采集",
        "生产",
    },
    "活动/商业化": {
        "活动",
        "签到",
        "战令",
        "礼包",
        "充值",
        "商店",
        "商城",
        "购买",
        "限时",
        "奖励",
        "抽奖",
        "召唤",
    },
    "社交/公会竞争": {
        "公会",
        "联盟",
        "好友",
        "排行榜",
        "排名",
        "竞技场",
        "聊天",
        "邀请",
        "成员",
        "队伍",
    },
    "飞行/射击题材": {
        "飞机",
        "战机",
        "飞行员",
        "机库",
        "导弹",
        "空袭",
        "射击",
        "僚机",
        "弹幕",
        "aircraft",
        "fighter",
        "jet",
        "plane",
        "missile",
        "shooter",
    },
    "末日/生存题材": {
        "幸存者",
        "僵尸",
        "末日",
        "避难所",
        "感染",
        "生存",
        "废土",
        "救援",
        "survival",
        "zombie",
        "wasteland",
        "shelter",
    },
    "剧情/叙事": {
        "剧情",
        "章节",
        "对话",
        "故事",
        "任务",
        "探索",
        "冒险",
        "线索",
        "选择",
        "先生",
        "老板",
        "小姐",
        "等等",
        "拜托",
        "story",
        "dialogue",
        "chapter",
    },
}

TEXT_MATERIAL_EXTENSIONS = {".txt", ".md", ".markdown", ".json"}
TABLE_MATERIAL_EXTENSIONS = {".xlsx", ".xlsm"}
DELIMITED_MATERIAL_EXTENSIONS = {".csv", ".tsv"}
IMAGE_MATERIAL_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
DOCX_MATERIAL_EXTENSIONS = {".docx"}

CATEGORY_LABELS = {
    "rarity": "稀有度/品质",
    "resource": "资源/货币/奖励",
    "stat": "战斗属性/数值",
    "action": "UI操作动词",
    "system": "系统/玩法名",
    "object": "角色/装备/对象",
    "status": "状态/进度",
    "other": "其他",
}


@dataclass
class Record:
    row_id: str
    source: str
    target: str


@dataclass
class SheetColumnLayout:
    header_row_index: int
    headers: list[str]
    id_index: int
    source_index: int
    target_index: int | None
    output_indexes: list[int]


@dataclass
class LanguageTableSpec:
    language: str
    path: Path


class AiSupplementProvider:
    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        raise NotImplementedError


class FileAiSupplementProvider(AiSupplementProvider):
    def __init__(self, response_path: Path):
        self.response_path = response_path

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        return json.loads(self.response_path.read_text(encoding="utf-8"))


class MockAiSupplementProvider(AiSupplementProvider):
    def __init__(self, response: dict[str, object]):
        self.response = response

    def generate(self, packet: dict[str, object]) -> dict[str, object]:
        return self.response


def clean_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = html.unescape(text)
    text = HTML_TAG_RE.sub(" ", text)
    text = BRACKET_TAG_RE.sub("", text)
    text = PLACEHOLDER_RE.sub("", text)
    text = SPACE_RE.sub(" ", text).strip()
    return text


def configure_utf8_stdio() -> None:
    for stream in (getattr(sys, "stdout", None), getattr(sys, "stderr", None)):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def normalize_english_for_compare(text: str) -> str:
    text = clean_text(text)
    text = CAMEL_SPLIT_RE.sub(" ", text)
    text = text.lower()
    text = re.sub(r"\s*\+\s*", "+", text)
    text = re.sub(r"[-_/]+", " ", text)
    text = EN_COMPARE_RE.sub(" ", text)
    text = SPACE_RE.sub(" ", text).strip()

    normalized_tokens: list[str] = []
    for token in text.split():
        if token.endswith("ies") and len(token) > 4:
            token = token[:-3] + "y"
        elif token.endswith(("oes", "ses", "xes", "zes", "ches", "shes")) and len(token) > 4:
            token = token[:-2]
        elif token.endswith("s") and len(token) > 3 and not token.endswith(("ss", "us", "is")):
            token = token[:-1]
        normalized_tokens.append(token)
    return " ".join(normalized_tokens)


def is_same_or_extended_usage(example_en: str, actual_en: str) -> bool:
    example_norm = normalize_english_for_compare(example_en)
    actual_norm = normalize_english_for_compare(actual_en)
    if not actual_norm or not example_norm:
        return False
    if actual_norm == example_norm:
        return True
    return f" {example_norm} " in f" {actual_norm} "


def split_usage_buckets(example_en: str, actual_counter: Counter[str]) -> tuple[Counter[str], Counter[str]]:
    example_counter: Counter[str] = Counter()
    manual_counter: Counter[str] = Counter()
    for actual_en, count in actual_counter.items():
        if is_same_or_extended_usage(example_en=example_en, actual_en=actual_en):
            example_counter[actual_en] += count
        else:
            manual_counter[actual_en] += count
    return example_counter, manual_counter


def collect_translation_diff(example_en: str, actual_counter: Counter[str]) -> dict[str, object]:
    same_counter, diff_counter = split_usage_buckets(
        example_en=example_en,
        actual_counter=actual_counter,
    )
    return {
        "has_diff": "Yes" if diff_counter else "No",
        "same_or_format_only_count": sum(same_counter.values()),
        "diff_count": sum(diff_counter.values()),
        "diff_variants": join_counter(diff_counter, limit=8),
        "diff_type": "manual_adaptation" if diff_counter else "",
    }


def token_roots(text: str) -> list[str]:
    roots: list[str] = []
    for token in EN_WORD_RE.findall(normalize_english_for_compare(text)):
        root = token
        if root.endswith("ing") and len(root) > 5:
            root = root[:-3]
        elif root.endswith("ed") and len(root) > 4:
            root = root[:-2]
        elif root.endswith("er") and len(root) > 4:
            root = root[:-2]
        elif root.endswith("ation") and len(root) > 7:
            root = root[:-5] + "e"
        roots.append(root)
    return roots


def titleize_word(word: str) -> str:
    if word.isupper():
        return word
    if word in {"hp", "atk", "def", "dmg", "cp"}:
        return word.upper()
    return word.capitalize()


def choose_en2_value(
    example_en: str,
    exact_diff_counter: Counter[str],
    manual_counter: Counter[str],
) -> str:
    if exact_diff_counter:
        return " | ".join(text for text, _ in exact_diff_counter.most_common(3))
    if not manual_counter:
        return ""

    manual_variants = manual_counter.most_common()
    top_text, top_count = manual_variants[0]
    second_count = manual_variants[1][1] if len(manual_variants) > 1 else 0
    total = sum(manual_counter.values())
    if top_count >= 2 and top_count > second_count and top_count / total >= 0.45:
        top_norm = normalize_english_for_compare(top_text)
        if top_norm and all(
            normalize_english_for_compare(text) == top_norm
            or is_same_or_extended_usage(example_en=top_text, actual_en=text)
            or is_same_or_extended_usage(example_en=text, actual_en=top_text)
            for text, _count in manual_variants[1:]
        ):
            return top_text

    example_roots = set(token_roots(example_en))
    root_counter: Counter[str] = Counter()

    for text, count in manual_counter.items():
        for root in token_roots(text):
            if root in example_roots or root in {"the", "a", "an", "of", "to", "for", "in", "on", "with", "and"}:
                continue
            root_counter[root] += count

    if not root_counter:
        return ""

    top_root, top_count = root_counter.most_common(1)[0]
    second_count = root_counter.most_common(2)[1][1] if len(root_counter) > 1 else 0
    if top_count < 2 or top_count <= second_count:
        return ""
    if top_count / total < 0.45:
        return ""
    return titleize_word(top_root)


def is_short_usage_candidate(record: Record, term: str, example_en: str) -> bool:
    if not record.target:
        return False
    if record.source == term:
        return True
    source_limit = max(8, len(term) + 4)
    target_limit = max(28, len(example_en) + 12) if example_en else 28
    return len(record.source) <= source_limit and len(record.target) <= target_limit


def is_valid_term(term: str) -> bool:
    if len(term) < 2 or len(term) > 12:
        return False
    if SENTENCE_PUNCT_RE.search(term):
        return False
    if NON_TERM_RE.match(term):
        return False
    if not CJK_RE.search(term):
        return False
    if term.startswith(("+", "-", "/", "%")) or term.endswith(("+", "-", "/", "%")):
        return False
    return True


def strip_numbered_title_prefix(value: object) -> str:
    return clean_text(NUMBERED_TITLE_RE.sub("", "" if value is None else str(value)).strip())


def parse_json_like_value(value: object) -> Any | None:
    text = "" if value is None else str(value).strip()
    if not text.startswith("["):
        return None
    try:
        return json.loads(text)
    except (TypeError, ValueError):
        return None


def first_string_from_json_like(value: object) -> str:
    parsed = parse_json_like_value(value)
    if isinstance(parsed, list) and parsed and isinstance(parsed[0], str):
        return strip_numbered_title_prefix(parsed[0])
    return clean_text(value)


def extract_structured_term_pairs(raw_source: object, raw_target: object) -> list[tuple[str, str]]:
    parsed_source = parse_json_like_value(raw_source)
    parsed_target = parse_json_like_value(raw_target)
    if not isinstance(parsed_source, list):
        return [(clean_text(raw_source), clean_text(raw_target))]

    pairs: list[tuple[str, str]] = []
    if parsed_source and isinstance(parsed_source[0], str):
        term = strip_numbered_title_prefix(parsed_source[0])
        target = first_string_from_json_like(raw_target)
        if is_valid_term(term):
            pairs.append((term, target))

    for index, source_item in enumerate(parsed_source):
        if not (isinstance(source_item, list) and source_item and isinstance(source_item[0], str)):
            continue
        term = strip_numbered_title_prefix(source_item[0])
        target = ""
        if isinstance(parsed_target, list) and index < len(parsed_target):
            target_item = parsed_target[index]
            if isinstance(target_item, list) and target_item and isinstance(target_item[0], str):
                target = strip_numbered_title_prefix(target_item[0])
        if is_valid_term(term):
            pairs.append((term, target))

    return pairs or [(clean_text(raw_source), clean_text(raw_target))]


def category_for(term: str) -> str:
    if term in RARITY_TERMS:
        return "rarity"
    if term in RESOURCE_TERMS:
        return "resource"
    if term in STAT_TERMS or term.endswith("伤害") or term.endswith("伤害+"):
        return "stat"
    if term in ACTION_TERMS:
        return "action"
    if term in SYSTEM_TERMS:
        return "system"
    if term in OBJECT_TERMS:
        return "object"
    if term in STATUS_TERMS:
        return "status"
    if any(key in term for key in ("伤害", "攻击", "生命", "防御", "暴击")):
        return "stat"
    if any(key in term for key in ("公会", "竞技场", "战令", "签到", "商城", "商店", "基地", "防御塔", "活动")):
        return "system"
    if any(key in term for key in ("英雄", "技能", "装备", "建筑", "武器", "坐骑")):
        return "object"
    return "other"


def join_counter(counter: Counter[str], limit: int = 5) -> str:
    if not counter:
        return ""
    return " | ".join(f"{text} ({count})" for text, count in counter.most_common(limit))


def risk_for(term: str, variants: int, hits: int, suggested_en: str) -> str:
    if variants > 1 or term in HIGH_CONFUSION_TERMS or not suggested_en:
        return "high"
    if hits >= 30:
        return "medium"
    return "low"


def priority_for(risk: str, hits: int) -> str:
    if risk == "high" or hits >= 80:
        return "P1"
    if hits >= 30:
        return "P2"
    return "P3"


def note_for(
    term: str,
    variants: int,
    exact_hits: int,
    hits: int,
    suggested_en: str,
    has_actual_diff: bool,
) -> str:
    notes: list[str] = []
    if variants > 1:
        notes.append("multiple English variants detected")
    if term in ACTION_TERMS:
        notes.append("action term needs consistency review")
    if term in RARITY_TERMS:
        notes.append("rarity ladder should stay globally aligned")
    if exact_hits == 1 and hits >= 20:
        notes.append("mostly embedded usage, review with context")
    if not suggested_en:
        notes.append("no stable English match found")
    if has_actual_diff:
        notes.append("actual short usages contain manual adaptation")
    return "; ".join(notes)


def counter_to_dict(counter: Counter[str]) -> dict[str, int]:
    return {key: int(value) for key, value in sorted(counter.items()) if key}


def dict_to_counter(value: dict[str, Any] | None) -> Counter[str]:
    counter: Counter[str] = Counter()
    if not value:
        return counter
    for key, raw in value.items():
        try:
            count = int(raw)
        except (TypeError, ValueError):
            continue
        if key and count > 0:
            counter[key] = count
    return counter


def merge_counters(*counters: Counter[str]) -> Counter[str]:
    merged: Counter[str] = Counter()
    for counter in counters:
        merged.update(counter)
    return merged


def new_curated_rules() -> dict[str, Any]:
    return {"version": CURATED_VERSION, "terms": {}}


def new_observation_store() -> dict[str, Any]:
    return {"version": OBSERVATION_VERSION, "terms": {}}


def split_legacy_term_memory(memory: dict[str, Any] | None) -> tuple[dict[str, Any], dict[str, Any]]:
    curated = new_curated_rules()
    observations = new_observation_store()
    if not isinstance(memory, dict):
        return curated, observations

    for term, raw_state in memory.get("terms", {}).items():
        if not isinstance(raw_state, dict):
            continue
        curated_state = {
            "approved_en": clean_text(raw_state.get("approved_en")),
            "approved_en2": clean_text(raw_state.get("approved_en2")),
            "block_en2": bool(raw_state.get("block_en2")),
            "ignore": bool(raw_state.get("ignore")),
            "note": clean_text(raw_state.get("note")),
            "category_override": clean_text(raw_state.get("category_override")),
        }
        observation_state = {
            "observed_exact_candidates": counter_to_dict(dict_to_counter(raw_state.get("observed_exact_candidates"))),
            "observed_example_usages": counter_to_dict(dict_to_counter(raw_state.get("observed_example_usages"))),
            "observed_manual_adaptations": counter_to_dict(dict_to_counter(raw_state.get("observed_manual_adaptations"))),
            "seen_runs": max(0, int(raw_state.get("seen_runs", 0) or 0)),
            "last_seen_at": clean_text(raw_state.get("last_seen_at")),
            "last_input_digest": clean_text(raw_state.get("last_input_digest")),
        }
        if any(
            [
                curated_state["approved_en"],
                curated_state["approved_en2"],
                curated_state["block_en2"],
                curated_state["ignore"],
                curated_state["note"],
                curated_state["category_override"],
            ]
        ):
            curated["terms"][term] = curated_state
        if any(
            [
                observation_state["observed_exact_candidates"],
                observation_state["observed_example_usages"],
                observation_state["observed_manual_adaptations"],
                observation_state["seen_runs"],
                observation_state["last_seen_at"],
                observation_state["last_input_digest"],
            ]
        ):
            observations["terms"][term] = observation_state
    return curated, observations


def load_json_object(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    return payload


def legacy_experience_candidate(path: Path | None) -> Path | None:
    if path is None:
        return None
    candidate = path.with_name("term_memory.json")
    if candidate.exists():
        return candidate
    if DEFAULT_LEGACY_EXPERIENCE_STORE.exists():
        return DEFAULT_LEGACY_EXPERIENCE_STORE
    return None


def default_curated_term_state() -> dict[str, Any]:
    return {
        "approved_en": "",
        "approved_en2": "",
        "block_en2": False,
        "ignore": False,
        "note": "",
        "category_override": "",
    }


def get_curated_term_state(curated_rules: dict[str, Any], term: str, *, create: bool = True) -> dict[str, Any]:
    terms = curated_rules.setdefault("terms", {})
    if create:
        state = terms.setdefault(term, {})
    else:
        state = terms.get(term, {})
        if not isinstance(state, dict):
            state = {}
    defaults = default_curated_term_state()
    if create:
        for key, value in defaults.items():
            state.setdefault(key, value)
        return state
    defaults.update(
        {
            "approved_en": clean_text(state.get("approved_en")),
            "approved_en2": clean_text(state.get("approved_en2")),
            "block_en2": bool(state.get("block_en2")),
            "ignore": bool(state.get("ignore")),
            "note": clean_text(state.get("note")),
            "category_override": clean_text(state.get("category_override")),
        }
    )
    state = defaults
    return state


def get_observation_term_state(observations_store: dict[str, Any], term: str) -> dict[str, Any]:
    terms = observations_store.setdefault("terms", {})
    state = terms.setdefault(term, {})
    state.setdefault("observed_exact_candidates", {})
    state.setdefault("observed_manual_adaptations", {})
    state.setdefault("observed_example_usages", {})
    state.setdefault("seen_runs", 0)
    state.setdefault("last_seen_at", "")
    state.setdefault("last_input_digest", "")
    return state


def sanitize_curated_rules(payload: dict[str, Any] | None) -> dict[str, Any]:
    if not payload:
        return new_curated_rules()
    if "terms" not in payload or not isinstance(payload["terms"], dict):
        payload = {"version": payload.get("version", CURATED_VERSION), "terms": {}}
    curated = new_curated_rules()
    curated["version"] = int(payload.get("version", CURATED_VERSION) or CURATED_VERSION)
    for term in payload["terms"]:
        if not isinstance(term, str):
            continue
        state = get_curated_term_state(curated, term)
        raw = payload["terms"].get(term)
        if isinstance(raw, dict):
            state["approved_en"] = clean_text(raw.get("approved_en"))
            state["approved_en2"] = clean_text(raw.get("approved_en2"))
            state["block_en2"] = bool(raw.get("block_en2"))
            state["ignore"] = bool(raw.get("ignore"))
            state["note"] = clean_text(raw.get("note"))
            state["category_override"] = clean_text(raw.get("category_override"))
    return curated


def sanitize_observation_store(payload: dict[str, Any] | None) -> dict[str, Any]:
    if not payload:
        return new_observation_store()
    if "terms" not in payload or not isinstance(payload["terms"], dict):
        payload = {"version": payload.get("version", OBSERVATION_VERSION), "terms": {}}
    observations = new_observation_store()
    observations["version"] = int(payload.get("version", OBSERVATION_VERSION) or OBSERVATION_VERSION)
    for term in payload["terms"]:
        if not isinstance(term, str):
            continue
        state = get_observation_term_state(observations, term)
        raw = payload["terms"].get(term)
        if isinstance(raw, dict):
            state["observed_exact_candidates"] = counter_to_dict(dict_to_counter(raw.get("observed_exact_candidates")))
            state["observed_example_usages"] = counter_to_dict(dict_to_counter(raw.get("observed_example_usages")))
            state["observed_manual_adaptations"] = counter_to_dict(dict_to_counter(raw.get("observed_manual_adaptations")))
            state["seen_runs"] = max(0, int(raw.get("seen_runs", 0) or 0))
            state["last_seen_at"] = clean_text(raw.get("last_seen_at"))
            state["last_input_digest"] = clean_text(raw.get("last_input_digest"))
    return observations


def load_curated_rules(path: Path | None) -> dict[str, Any]:
    payload = load_json_object(path)
    if payload:
        if any(
            isinstance(state, dict) and any(key.startswith("observed_") for key in state.keys())
            for state in payload.get("terms", {}).values()
        ):
            legacy_curated, _legacy_observations = split_legacy_term_memory(payload)
            return legacy_curated
        return sanitize_curated_rules(payload)

    legacy_path = legacy_experience_candidate(path)
    if legacy_path:
        legacy_payload = load_json_object(legacy_path)
        if legacy_payload:
            legacy_curated, _legacy_observations = split_legacy_term_memory(legacy_payload)
            return legacy_curated
    return new_curated_rules()


def load_observation_store(path: Path | None) -> dict[str, Any]:
    payload = load_json_object(path)
    if payload:
        if any(
            isinstance(state, dict) and any(key.startswith("approved_") or key in {"block_en2", "ignore", "note", "category_override"} for key in state.keys())
            for state in payload.get("terms", {}).values()
        ):
            _legacy_curated, legacy_observations = split_legacy_term_memory(payload)
            return legacy_observations
        return sanitize_observation_store(payload)

    legacy_path = legacy_experience_candidate(path)
    if legacy_path:
        legacy_payload = load_json_object(legacy_path)
        if legacy_payload:
            _legacy_curated, legacy_observations = split_legacy_term_memory(legacy_payload)
            return legacy_observations
    return new_observation_store()


def save_curated_rules(path: Path | None, curated_rules: dict[str, Any]) -> None:
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(sanitize_curated_rules(curated_rules), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def save_observation_store(path: Path | None, observations_store: dict[str, Any]) -> None:
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(sanitize_observation_store(observations_store), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def apply_observation_history(
    observation_state: dict[str, Any],
    exact_translation_counter: Counter[str],
    example_usage_counter: Counter[str],
    manual_adaptation_counter: Counter[str],
) -> tuple[Counter[str], Counter[str], Counter[str]]:
    historical_exact = dict_to_counter(observation_state.get("observed_exact_candidates"))
    historical_examples = dict_to_counter(observation_state.get("observed_example_usages"))
    historical_manual = dict_to_counter(observation_state.get("observed_manual_adaptations"))
    return (
        merge_counters(exact_translation_counter, historical_exact),
        merge_counters(example_usage_counter, historical_examples),
        merge_counters(manual_adaptation_counter, historical_manual),
    )


def apply_curated_preferences(
    curated_state: dict[str, Any],
    term: str,
    suggested_en: str,
    example_en: str,
    en2_value: str,
    exact_translation_counter: Counter[str],
    example_usage_counter: Counter[str],
    manual_adaptation_counter: Counter[str],
) -> tuple[str, str, str, Counter[str], Counter[str], Counter[str]]:
    approved_en = clean_text(curated_state.get("approved_en"))
    approved_en2 = clean_text(curated_state.get("approved_en2"))
    block_en2 = bool(curated_state.get("block_en2"))

    if approved_en:
        suggested_en = approved_en
        example_en = approved_en
    elif not example_en and exact_translation_counter:
        example_en = exact_translation_counter.most_common(1)[0][0]
        suggested_en = example_en

    if approved_en2:
        en2_value = approved_en2
    elif block_en2:
        en2_value = ""
    elif not en2_value and manual_adaptation_counter:
        en2_value = choose_en2_value(
            example_en=example_en,
            exact_diff_counter=Counter(),
            manual_counter=manual_adaptation_counter,
        )

    if curated_state.get("ignore") and term not in HIGH_CONFUSION_TERMS:
        return "", "", "", Counter(), Counter(), Counter()
    return suggested_en, example_en, en2_value, exact_translation_counter, example_usage_counter, manual_adaptation_counter


def update_observation_store(
    observation_state: dict[str, Any],
    *,
    input_digest: str,
    exact_translation_counter: Counter[str],
    example_usage_counter: Counter[str],
    manual_adaptation_counter: Counter[str],
) -> None:
    if observation_state.get("last_input_digest") == input_digest:
        observation_state["last_seen_at"] = datetime.now(timezone.utc).isoformat()
        return

    observed_exact = dict_to_counter(observation_state.get("observed_exact_candidates"))
    observed_example = dict_to_counter(observation_state.get("observed_example_usages"))
    observed_manual = dict_to_counter(observation_state.get("observed_manual_adaptations"))
    observed_exact.update(exact_translation_counter)
    observed_example.update(example_usage_counter)
    observed_manual.update(manual_adaptation_counter)

    observation_state["observed_exact_candidates"] = counter_to_dict(observed_exact)
    observation_state["observed_example_usages"] = counter_to_dict(observed_example)
    observation_state["observed_manual_adaptations"] = counter_to_dict(observed_manual)
    observation_state["seen_runs"] = int(observation_state.get("seen_runs", 0)) + 1
    observation_state["last_seen_at"] = datetime.now(timezone.utc).isoformat()
    observation_state["last_input_digest"] = input_digest


def file_digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def set_widths(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        worksheet.column_dimensions[letter].width = max(10, min(max_len + 2, 42))


def style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    set_widths(worksheet)


def resolve_column_index(headers: list[object], expected_name: str) -> int:
    key = clean_text(expected_name).lower()
    for index, name in enumerate(headers):
        if clean_text(name).lower() == key:
            return index
    available = ", ".join(str(name) for name in headers)
    raise ValueError(f"Missing column '{expected_name}'. Available headers: {available}")


XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XLSX_NS = {"a": XLSX_MAIN_NS, "rel": PACKAGE_REL_NS}


def cell_column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - 64
    return index - 1


def xml_text(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return "".join(node.itertext())


def workbook_sheet_targets(archive: ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("rel:Relationship", XLSX_NS)
    }
    sheets: list[tuple[str, str]] = []
    for sheet in workbook_root.find("a:sheets", XLSX_NS).findall("a:sheet", XLSX_NS):
        rel_id = sheet.attrib[f"{{{XLSX_REL_NS}}}id"]
        target = rel_map[rel_id]
        target_path = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
        sheets.append((sheet.attrib["name"], target_path))
    return sheets


def load_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return [xml_text(item) for item in root.findall("a:si", XLSX_NS)]


def raw_cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value = cell.find("a:v", XLSX_NS)
        if value is None or value.text is None:
            return ""
        try:
            return shared_strings[int(value.text)]
        except (IndexError, ValueError):
            return ""
    if cell_type == "inlineStr":
        return xml_text(cell.find("a:is", XLSX_NS))
    value = cell.find("a:v", XLSX_NS)
    return "" if value is None or value.text is None else value.text


def iter_raw_xlsx_sheets(input_path: Path) -> list[tuple[str, list[list[str]]]]:
    sheets: list[tuple[str, list[list[str]]]] = []
    with ZipFile(input_path) as archive:
        shared_strings = load_shared_strings(archive)
        for sheet_name, target_path in workbook_sheet_targets(archive):
            root = ET.fromstring(archive.read(target_path))
            rows: list[list[str]] = []
            for row in root.findall(".//a:sheetData/a:row", XLSX_NS):
                cells: list[tuple[int, str]] = []
                max_column = -1
                for cell in row.findall("a:c", XLSX_NS):
                    column_index = cell_column_index(cell.attrib.get("r", ""))
                    max_column = max(max_column, column_index)
                    cells.append((column_index, raw_cell_value(cell, shared_strings)))
                values = [""] * (max_column + 1)
                for column_index, value in cells:
                    values[column_index] = value
                rows.append(values)
            sheets.append((sheet_name, rows))
    return sheets


def records_from_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> list[Record]:
    if not rows:
        return []
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        headers = list(rows[0])
        id_index = resolve_column_index(headers, id_column)
        source_index = resolve_column_index(headers, source_column)
        target_index = None if source_only else resolve_column_index(headers, target_column)
        data_rows = rows[1:]
        first_data_row_number = 2
    else:
        id_index = layout.id_index
        source_index = layout.source_index
        target_index = layout.target_index
        data_rows = rows[layout.header_row_index + 1 :]
        first_data_row_number = layout.header_row_index + 2

    records: list[Record] = []
    for row_number, row in enumerate(data_rows, start=first_data_row_number):
        row_values = list(row)
        row_id = "" if id_index >= len(row_values) or row_values[id_index] is None else str(row_values[id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        source = "" if source_index >= len(row_values) else clean_text(row_values[source_index])
        target = "" if target_index is None or target_index >= len(row_values) else clean_text(row_values[target_index])
        if not source:
            continue
        records.append(Record(row_id=row_id, source=source, target=target))
    return records


def load_records_from_raw_xlsx(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> tuple[list[Record], str]:
    sheets = iter_raw_xlsx_sheets(input_path)
    if not sheets:
        return [], ""
    selected_sheet: tuple[str, list[list[str]]] | None = None
    for candidate in sheets:
        if sheet_name is None or candidate[0] == sheet_name:
            selected_sheet = candidate
            break
    if selected_sheet is None:
        available = ", ".join(name for name, _rows in sheets)
        raise ValueError(f"Missing worksheet '{sheet_name}'. Available worksheets: {available}")
    title, rows = selected_sheet
    return records_from_rows(
        rows=rows,
        sheet_title=title,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    ), title


def normalized_header_lookup(headers: list[object]) -> dict[str, int]:
    return {clean_text(name).lower(): index for index, name in enumerate(headers)}


def first_matching_header(headers: list[object], candidates: list[str]) -> int | None:
    lookup = normalized_header_lookup(headers)
    for candidate in candidates:
        key = clean_text(candidate).lower()
        if key in lookup:
            return lookup[key]
    return None


def first_matching_header_fuzzy(headers: list[object], candidates: list[str]) -> int | None:
    exact_index = first_matching_header(headers, candidates)
    if exact_index is not None:
        return exact_index
    normalized_candidates = [clean_text(candidate).lower() for candidate in candidates if clean_text(candidate)]
    for index, header in enumerate(headers):
        header_key = clean_text(header).lower()
        if not header_key:
            continue
        if any(candidate in header_key for candidate in normalized_candidates):
            return index
    return None


def canonical_output_header(requested_header: str, default_header: str) -> str:
    clean_header = clean_text(requested_header)
    return default_header if clean_header.lower() == default_header.lower() else clean_header


def value_at(values: list[object], index: int | None) -> object:
    if index is None or index >= len(values):
        return ""
    return values[index]


def exact_sheet_column_layout(
    headers: list[str],
    header_row_index: int,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> SheetColumnLayout | None:
    try:
        id_index = resolve_column_index(headers, id_column)
        source_index = resolve_column_index(headers, source_column)
        target_index = None if source_only else resolve_column_index(headers, target_column)
    except ValueError:
        return None
    return SheetColumnLayout(
        header_row_index=header_row_index,
        headers=headers,
        id_index=id_index,
        source_index=source_index,
        target_index=target_index,
        output_indexes=list(range(len(headers))),
    )


def auto_sheet_column_layout(
    headers: list[str],
    header_row_index: int,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> SheetColumnLayout | None:
    id_index = first_matching_header_fuzzy(headers, [id_column, *AUTO_ID_HEADERS])
    source_index = first_matching_header_fuzzy(headers, [source_column, *AUTO_SOURCE_HEADERS])
    target_index = None if source_only else first_matching_header_fuzzy(headers, [target_column, *AUTO_TARGET_HEADERS])
    if id_index is None or source_index is None or (not source_only and target_index is None):
        return None
    output_indexes = [id_index, source_index]
    output_headers = ["ID", "CN"]
    if not source_only:
        output_indexes.append(target_index if target_index is not None else -1)
        output_headers.append(canonical_output_header(target_column, "EN"))
    return SheetColumnLayout(
        header_row_index=header_row_index,
        headers=output_headers,
        id_index=id_index,
        source_index=source_index,
        target_index=target_index,
        output_indexes=output_indexes,
    )


def language_table_layout_from_rows(
    rows: list[list[object]],
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> SheetColumnLayout | None:
    for header_row_index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        headers, _empty_values = trim_trailing_empty_columns(list(row))
        if not headers:
            continue
        exact_layout = exact_sheet_column_layout(
            headers=headers,
            header_row_index=header_row_index,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        if exact_layout is not None:
            return exact_layout
        auto_layout = auto_sheet_column_layout(
            headers=headers,
            header_row_index=header_row_index,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        if auto_layout is not None:
            return auto_layout
    return None


def auto_records_from_sheet_rows(sheet_title: str, rows: list[list[object]]) -> list[Record]:
    if not rows:
        return []
    headers = list(rows[0])
    source_index = first_matching_header(
        headers,
        ["简体中文", "中文", "正常对话", "资料", "简介", "内容", "说明", "描述", "cn", "source", "zh", "Chinese", "note", "description"],
    )
    if source_index is None:
        return []
    target_index = first_matching_header(
        headers,
        ["英文", "英语", "en", "English", "优化翻译"],
    )
    id_index = first_matching_header(
        headers,
        ["唯一标识ID", "ID", "id", "章节", "关卡序号"],
    )

    records: list[Record] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = list(row)
        source = "" if source_index >= len(row_values) else clean_text(row_values[source_index])
        if not source:
            continue
        target = "" if target_index is None or target_index >= len(row_values) else clean_text(row_values[target_index])
        row_id = ""
        if id_index is not None and id_index < len(row_values):
            row_id = clean_text(row_values[id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        records.append(Record(row_id=row_id, source=source, target=target))
    return records


def generic_records_from_sheet_rows(sheet_title: str, rows: list[list[object]]) -> list[Record]:
    records: list[Record] = []
    for row_number, row in enumerate(rows[1:], start=2):
        parts = [clean_text(value) for value in row if clean_text(value)]
        if not parts:
            continue
        records.append(Record(row_id=f"{sheet_title}:{row_number}", source=" ".join(parts), target=""))
    return records


def load_project_records(input_path: Path) -> list[Record]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        records: list[Record] = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            records.extend(auto_records_from_sheet_rows(worksheet.title, rows))
        workbook.close()
        return records
    except Exception:
        records = []
        for sheet_title, rows in iter_raw_xlsx_sheets(input_path):
            records.extend(auto_records_from_sheet_rows(sheet_title, rows))
        return records


def load_table_material_records(input_path: Path) -> list[Record]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        records: list[Record] = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            sheet_records = auto_records_from_sheet_rows(worksheet.title, rows)
            records.extend(sheet_records or generic_records_from_sheet_rows(worksheet.title, rows))
        workbook.close()
        return records
    except Exception:
        records = []
        for sheet_title, rows in iter_raw_xlsx_sheets(input_path):
            sheet_records = auto_records_from_sheet_rows(sheet_title, rows)
            records.extend(sheet_records or generic_records_from_sheet_rows(sheet_title, rows))
        return records


def chunk_text_material(text: str, limit: int = 160) -> list[str]:
    cleaned = clean_text(text)
    if not cleaned:
        return []
    raw_chunks = re.split(r"[\r\n]+|(?<=[。！？.!?])\s*", cleaned)
    chunks: list[str] = []
    buffer = ""
    for raw_chunk in raw_chunks:
        chunk = clean_text(raw_chunk)
        if not chunk:
            continue
        if len(chunk) > limit:
            chunks.append(chunk[:limit])
            continue
        if not buffer:
            buffer = chunk
        elif len(buffer) + len(chunk) + 1 <= limit:
            buffer = f"{buffer} {chunk}"
        else:
            chunks.append(buffer)
            buffer = chunk
    if buffer:
        chunks.append(buffer)
    return chunks[:200]


def records_from_text_material(path: Path) -> list[Record]:
    try:
        content = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        content = path.read_text(encoding="gb18030", errors="ignore")
    records: list[Record] = []
    for index, chunk in enumerate(chunk_text_material(content), start=1):
        records.append(Record(row_id=f"{path.name}:{index}", source=chunk, target=""))
    return records


DOCX_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def records_from_docx_material(path: Path) -> list[Record]:
    with ZipFile(path) as archive:
        if "word/document.xml" not in archive.namelist():
            raise ValueError(f"Missing word/document.xml in DOCX file: {path}")
        root = ET.fromstring(archive.read("word/document.xml"))

    records: list[Record] = []
    for index, paragraph in enumerate(root.findall(".//w:p", DOCX_NS), start=1):
        text = clean_text("".join(node.text or "" for node in paragraph.findall(".//w:t", DOCX_NS)))
        if text:
            records.append(Record(row_id=f"{path.name}:{index}", source=text, target=""))
    return records


def records_from_delimited_material(path: Path) -> list[Record]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        content = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        content = path.read_text(encoding="gb18030", errors="ignore")
    rows = list(csv.reader(content.splitlines(), delimiter=delimiter))
    if not rows:
        return []
    return auto_records_from_sheet_rows(path.name, rows) or generic_records_from_sheet_rows(path.name, rows)


def records_from_image_material(path: Path) -> list[Record]:
    source = " ".join(
        part
        for part in [
            "图片资料",
            path.stem.replace("_", " ").replace("-", " "),
            path.parent.name if path.parent else "",
        ]
        if part
    )
    return [Record(row_id=f"{path.name}:image", source=source, target="")]


def load_project_material_records(
    material_paths: list[Path],
    notes: list[str] | None = None,
) -> tuple[list[Record], list[str]]:
    records: list[Record] = []
    sources: list[str] = []
    for note_index, note in enumerate(notes or [], start=1):
        note_text = clean_text(note)
        if note_text:
            records.append(Record(row_id=f"project-note:{note_index}", source=note_text, target=""))
            sources.append(f"备注: {note_text[:40]}")

    for material_path in material_paths:
        path = Path(material_path)
        suffix = path.suffix.lower()
        if not path.exists():
            sources.append(f"缺失资料: {path}")
            continue
        if suffix in TABLE_MATERIAL_EXTENSIONS:
            material_records = load_table_material_records(path)
        elif suffix in DELIMITED_MATERIAL_EXTENSIONS:
            material_records = records_from_delimited_material(path)
        elif suffix in TEXT_MATERIAL_EXTENSIONS:
            material_records = records_from_text_material(path)
        elif suffix in IMAGE_MATERIAL_EXTENSIONS:
            material_records = records_from_image_material(path)
        else:
            material_records = records_from_text_material(path)

        records.extend(material_records)
        sources.append(f"{path.name} ({len(material_records)} 条)")
    return records, sources


def load_announcement_texts(material_paths: list[Path]) -> str:
    chunks: list[str] = []
    for material_path in material_paths:
        path = Path(material_path)
        if not path.exists():
            raise FileNotFoundError(f"Missing announcement material: {path}")

        suffix = path.suffix.lower()
        if suffix in DOCX_MATERIAL_EXTENSIONS:
            records = records_from_docx_material(path)
        elif suffix in TABLE_MATERIAL_EXTENSIONS:
            records = load_table_material_records(path)
        elif suffix in DELIMITED_MATERIAL_EXTENSIONS:
            records = records_from_delimited_material(path)
        elif suffix in TEXT_MATERIAL_EXTENSIONS:
            records = records_from_text_material(path)
        else:
            records = records_from_text_material(path)

        chunks.extend(record.source for record in records if record.source)
    return clean_text(" ".join(chunks))


def build_announcement_candidate_rows(
    records: list[Record],
    curated_rules: dict[str, Any] | None = None,
    min_hit: int = 1,
) -> list[dict[str, object]]:
    curated_rules = curated_rules if curated_rules is not None else new_curated_rules()
    records_by_term: dict[str, list[Record]] = defaultdict(list)
    translations_by_term: dict[str, Counter[str]] = defaultdict(Counter)
    for record in records:
        term = clean_text(record.source)
        if not is_valid_term(term):
            continue
        records_by_term[term].append(record)
        if record.target:
            translations_by_term[term][record.target] += 1

    rows: list[dict[str, object]] = []
    for term, term_records in records_by_term.items():
        if len(term_records) < min_hit:
            continue
        curated_state = get_curated_term_state(curated_rules, term, create=False)
        if curated_state.get("ignore"):
            continue

        approved_en = clean_text(curated_state.get("approved_en"))
        approved_en2 = "" if curated_state.get("block_en2") else clean_text(curated_state.get("approved_en2"))
        common_en = translations_by_term[term].most_common(1)[0][0] if translations_by_term[term] else ""
        en = approved_en or common_en
        example_record = next((record for record in term_records if record.target == en), term_records[0])
        rows.append(
            {
                "ID": example_record.row_id,
                "CN": term,
                "EN": en,
                "EN2": approved_en2,
            }
        )
    return rows


def trim_trailing_empty_columns(headers: list[object], values: list[object] | None = None) -> tuple[list[str], list[object]]:
    last_index = len(headers) - 1
    while last_index >= 0 and not clean_text(headers[last_index]):
        last_index -= 1
    trimmed_headers = [clean_text(header) for header in headers[: last_index + 1]]
    raw_values = list(values or [])
    trimmed_values = raw_values[: len(trimmed_headers)]
    if len(trimmed_values) < len(trimmed_headers):
        trimmed_values.extend([""] * (len(trimmed_headers) - len(trimmed_values)))
    return trimmed_headers, trimmed_values


def announcement_candidate_rows_from_sheet_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    curated_rules: dict[str, Any] | None = None,
    min_hit: int = 1,
    source_only: bool = False,
) -> tuple[list[str], list[dict[str, object]]]:
    if not rows:
        return [], []
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        return [], []

    curated_rules = curated_rules if curated_rules is not None else new_curated_rules()
    records_by_term: dict[str, list[tuple[str, str, list[object], str]]] = defaultdict(list)
    for row_number, row in enumerate(rows[layout.header_row_index + 1 :], start=layout.header_row_index + 2):
        row_values = list(row)
        raw_source = value_at(row_values, layout.source_index)
        row_id = "" if layout.id_index >= len(row_values) or row_values[layout.id_index] is None else str(row_values[layout.id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        raw_target = "" if layout.target_index is None else value_at(row_values, layout.target_index)
        raw_source_text = "" if raw_source is None else str(raw_source).strip()
        for term, target in extract_structured_term_pairs(raw_source, raw_target):
            term = clean_text(term)
            if not is_valid_term(term):
                continue
            values = [value_at(row_values, index) for index in layout.output_indexes]
            if len(values) >= 2:
                values[1] = term
            if len(values) >= 3:
                values[2] = target
            records_by_term[term].append((row_id, clean_text(target), values, raw_source_text))

    candidate_rows: list[dict[str, object]] = []
    for term, entries in records_by_term.items():
        if len(entries) < min_hit:
            continue
        curated_state = get_curated_term_state(curated_rules, term, create=False)
        if curated_state.get("ignore"):
            continue
        row_id, target, values, _raw_source_text = next(
            (entry for entry in entries if entry[3] == term),
            entries[0],
        )
        candidate_rows.append(
            {
                "ID": row_id,
                "CN": term,
                "EN": target,
                "EN2": "" if curated_state.get("block_en2") else clean_text(curated_state.get("approved_en2")),
                "_AnnouncementValues": values,
            }
        )
    return layout.headers, candidate_rows


def build_announcement_candidate_rows_from_workbook(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    curated_rules: dict[str, Any] | None = None,
    min_hit: int = 1,
    source_only: bool = False,
) -> tuple[list[str], list[dict[str, object]]]:
    headers: list[str] = []
    candidate_rows: list[dict[str, object]] = []
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
        for worksheet in worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            sheet_headers, sheet_rows = announcement_candidate_rows_from_sheet_rows(
                rows=rows,
                sheet_title=worksheet.title,
                id_column=id_column,
                source_column=source_column,
                target_column=target_column,
                curated_rules=curated_rules,
                min_hit=min_hit,
                source_only=source_only,
            )
            if sheet_rows and not headers:
                headers = sheet_headers
            candidate_rows.extend(sheet_rows)
        workbook.close()
        return headers, candidate_rows
    except Exception:
        for raw_sheet_name, rows in iter_raw_xlsx_sheets(input_path):
            if sheet_name and raw_sheet_name != sheet_name:
                continue
            sheet_headers, sheet_rows = announcement_candidate_rows_from_sheet_rows(
                rows=rows,
                sheet_title=raw_sheet_name,
                id_column=id_column,
                source_column=source_column,
                target_column=target_column,
                curated_rules=curated_rules,
                min_hit=min_hit,
                source_only=source_only,
            )
            if sheet_rows and not headers:
                headers = sheet_headers
            candidate_rows.extend(sheet_rows)
        return headers, candidate_rows


def is_low_value_announcement_term(term: str) -> bool:
    return clean_text(term) in LOW_VALUE_ANNOUNCEMENT_TERMS


def select_announcement_term_rows(
    term_rows: list[dict[str, object]],
    announcement_text: str,
    include_empty: bool = False,
) -> list[dict[str, object]]:
    normalized_notice = clean_text(announcement_text)
    candidates: list[tuple[int, int, int, int, str, dict[str, object]]] = []
    spans_by_term: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for row in term_rows:
        cn = clean_text(row.get("CN"))
        if not cn:
            continue

        en = clean_text(row.get("EN")) or clean_text(row.get("EN2"))
        if not include_empty and not en:
            continue

        output_row = dict(row)
        output_row["CN"] = cn
        output_row["EN"] = en
        for match in re.finditer(re.escape(cn), normalized_notice):
            span = (match.start(), match.end())
            low_value_rank = 1 if is_low_value_announcement_term(cn) else 0
            candidates.append((low_value_rank, span[0], span[1], -len(cn), cn, output_row))
            spans_by_term[cn].append(span)

    candidates.sort(key=lambda item: (item[0], item[1], item[3], item[4]))
    selected_spans: list[tuple[int, int]] = []
    selected_terms: set[str] = set()
    selected_rows: list[dict[str, object]] = []
    for _low_value_rank, start, end, _negative_length, cn, row in candidates:
        if cn in selected_terms:
            continue
        if any(start < selected_end and end > selected_start for selected_start, selected_end in selected_spans):
            continue
        selected_spans.extend(spans_by_term.get(cn, [(start, end)]))
        selected_terms.add(cn)
        selected_rows.append(row)
    return selected_rows


def load_records(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> tuple[list[Record], str]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        records = records_from_rows(
            rows=rows,
            sheet_title=worksheet.title,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        workbook.close()
        return records, worksheet.title
    except Exception:
        return load_records_from_raw_xlsx(
            input_path=input_path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )



def build_term_rows(
    records: list[Record],
    min_hit: int,
    glossary_hit_threshold: int,
    curated_rules: dict[str, Any] | None = None,
    observations_store: dict[str, Any] | None = None,
    input_digest: str = "",
    include_empty_final_terms: bool = False,
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    curated_rules = curated_rules if curated_rules is not None else new_curated_rules()
    observations_store = observations_store if observations_store is not None else new_observation_store()
    label_counter: Counter[str] = Counter()
    label_translations: dict[str, Counter[str]] = defaultdict(Counter)

    for record in records:
        if is_valid_term(record.source):
            label_counter[record.source] += 1
            if record.target:
                label_translations[record.source][record.target] += 1

    rows_by_term: list[dict[str, object]] = []
    for term in sorted(set(label_counter)):
        hits = 0
        example_record: Record | None = None
        near_translations: Counter[str] = Counter()
        for record in records:
            if term not in record.source:
                continue
            hits += 1
            if example_record is None or len(record.source) < len(example_record.source):
                example_record = record
            if record.target and len(record.source) <= max(18, len(term) + 6):
                near_translations[record.target] += 1

        if hits < min_hit:
            continue

        exact_translations = label_translations.get(term, Counter())
        suggested_en = exact_translations.most_common(1)[0][0] if exact_translations else (
            near_translations.most_common(1)[0][0] if near_translations else ""
        )
        example_en = example_record.target if example_record and example_record.target else suggested_en

        actual_short_counter: Counter[str] = Counter()
        diff_sample: Record | None = None
        for record in records:
            if term not in record.source:
                continue
            if not is_short_usage_candidate(record=record, term=term, example_en=example_en):
                continue
            actual_short_counter[record.target] += 1
            if record.target and not is_same_or_extended_usage(example_en=example_en, actual_en=record.target):
                if diff_sample is None or (len(record.source), len(record.target)) < (len(diff_sample.source), len(diff_sample.target)):
                    diff_sample = record

        example_usage_counter, manual_adaptation_counter = split_usage_buckets(
            example_en=example_en,
            actual_counter=actual_short_counter,
        )
        exact_diff_counter = Counter(
            {
                text: count
                for text, count in exact_translations.items()
                if not is_same_or_extended_usage(example_en=example_en, actual_en=text)
            }
        )
        en2_value = choose_en2_value(
            example_en=example_en,
            exact_diff_counter=exact_diff_counter,
            manual_counter=manual_adaptation_counter,
        )

        curated_state = get_curated_term_state(curated_rules, term, create=False)
        observation_state = get_observation_term_state(observations_store, term)
        exact_translations, example_usage_counter, manual_adaptation_counter = apply_observation_history(
            observation_state=observation_state,
            exact_translation_counter=exact_translations,
            example_usage_counter=example_usage_counter,
            manual_adaptation_counter=manual_adaptation_counter,
        )
        suggested_en, example_en, en2_value, exact_translations, example_usage_counter, manual_adaptation_counter = apply_curated_preferences(
            curated_state=curated_state,
            term=term,
            suggested_en=suggested_en,
            example_en=example_en,
            en2_value=en2_value,
            exact_translation_counter=exact_translations,
            example_usage_counter=example_usage_counter,
            manual_adaptation_counter=manual_adaptation_counter,
        )
        if not suggested_en and exact_translations:
            suggested_en = exact_translations.most_common(1)[0][0]
        if not example_en:
            example_en = suggested_en
        if not suggested_en:
            suggested_en = example_en
        if not example_en and exact_translations:
            example_en = exact_translations.most_common(1)[0][0]

        update_observation_store(
            observation_state,
            input_digest=input_digest,
            exact_translation_counter=exact_translations,
            example_usage_counter=example_usage_counter,
            manual_adaptation_counter=manual_adaptation_counter,
        )

        diff_info = collect_translation_diff(example_en=example_en, actual_counter=actual_short_counter)
        risk = risk_for(term, len(exact_translations or near_translations), hits, suggested_en)
        category = clean_text(curated_state.get("category_override")) or category_for(term)
        note = note_for(
            term=term,
            variants=len(exact_translations or near_translations),
            exact_hits=label_counter[term],
            hits=hits,
            suggested_en=suggested_en,
            has_actual_diff=diff_info["has_diff"] == "Yes",
        )
        if clean_text(curated_state.get("note")):
            note = f"{note}; {clean_text(curated_state.get('note'))}" if note else clean_text(curated_state.get("note"))

        row = {
            "ID": example_record.row_id if example_record else "",
            "CN": term,
            "EN": example_en,
            "EN2": en2_value,
            "SuggestedEN": suggested_en,
            "ExactCandidates": join_counter(exact_translations or near_translations),
            "ExampleUsages": join_counter(example_usage_counter, limit=8),
            "ManualAdaptations": join_counter(manual_adaptation_counter, limit=8),
            "ActualShortUsages": join_counter(actual_short_counter, limit=8),
            "HasActualDiff": diff_info["has_diff"],
            "DiffType": diff_info["diff_type"],
            "DiffVariants": diff_info["diff_variants"],
            "SameOrFormatOnlyCount": diff_info["same_or_format_only_count"],
            "DiffCount": diff_info["diff_count"],
            "Category": category,
            "Risk": risk,
            "Priority": priority_for(risk, hits),
            "HitRows": hits,
            "ExactRows": label_counter[term],
            "ExampleID": example_record.row_id if example_record else "",
            "ExampleSource": example_record.source if example_record else "",
            "ExampleEN": example_record.target if example_record else "",
            "DiffExampleID": diff_sample.row_id if diff_sample else "",
            "DiffExampleSource": diff_sample.source if diff_sample else "",
            "DiffExampleEN": diff_sample.target if diff_sample else "",
            "Note": note,
        }
        if not curated_state.get("ignore"):
            rows_by_term.append(row)

    rows_by_term.sort(
        key=lambda row: (
            {"P1": 0, "P2": 1, "P3": 2}[row["Priority"]],
            {"high": 0, "medium": 1, "low": 2}[row["Risk"]],
            -int(row["HitRows"]),
            row["CN"],
        )
    )

    glossary_rows = [
        row for row in rows_by_term if int(row["HitRows"]) >= glossary_hit_threshold or row["Risk"] == "high"
    ]
    high_risk_rows = [row for row in rows_by_term if row["Risk"] == "high"]
    manual_rows = [row for row in rows_by_term if row["HasActualDiff"] == "Yes"]
    final_rows = list(glossary_rows) if include_empty_final_terms else [
        row for row in glossary_rows if row["EN"] or row["EN2"]
    ]
    return rows_by_term, glossary_rows, high_risk_rows, manual_rows, final_rows


def keyword_evidence(records: list[Record], keywords: set[str]) -> tuple[int, Counter[str]]:
    row_hits = 0
    keyword_counter: Counter[str] = Counter()
    for record in records:
        matched = False
        source_text = record.source.lower()
        for keyword in keywords:
            if keyword.lower() in source_text:
                keyword_counter[keyword] += 1
                matched = True
        if matched:
            row_hits += 1
    return row_hits, keyword_counter


def infer_project_signals(records: list[Record], limit: int = 5) -> list[dict[str, object]]:
    signals: list[dict[str, object]] = []
    for label, keywords in PROJECT_SIGNAL_GROUPS.items():
        row_hits, evidence_counter = keyword_evidence(records, keywords)
        if not row_hits:
            continue
        signals.append(
            {
                "label": label,
                "hit_rows": row_hits,
                "evidence": join_counter(evidence_counter, limit=6),
            }
        )
    signals.sort(key=lambda item: (-int(item["hit_rows"]), str(item["label"])))
    return signals[:limit]


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _header in headers) + " |",
    ]
    for row in rows:
        escaped = [str(value).replace("|", "\\|") for value in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


def source_samples(records: list[Record], signals: list[dict[str, object]], limit: int = 5) -> list[str]:
    if not signals:
        return []
    signal_labels = {str(signal["label"]) for signal in signals}
    keywords: set[str] = set()
    for label, group_keywords in PROJECT_SIGNAL_GROUPS.items():
        if label in signal_labels:
            keywords.update(group_keywords)

    samples: list[str] = []
    seen: set[str] = set()
    for record in records:
        if len(record.source) > 48:
            continue
        if not any(keyword in record.source for keyword in keywords):
            continue
        sample = f"{record.row_id}: {record.source}" if record.row_id else record.source
        if sample in seen:
            continue
        samples.append(sample)
        seen.add(sample)
        if len(samples) >= limit:
            break
    return samples


def category_distribution(rows: list[dict[str, object]]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for row in rows:
        category = str(row.get("Category") or "other")
        counter[CATEGORY_LABELS.get(category, category)] += 1
    return counter


def top_terms(rows: list[dict[str, object]], limit: int = 8) -> list[str]:
    terms: list[str] = []
    for row in sorted(rows, key=lambda item: (-int(item.get("HitRows") or 0), str(item.get("CN") or "")))[:limit]:
        cn = str(row.get("CN") or "")
        en = str(row.get("EN") or "")
        en2 = str(row.get("EN2") or "")
        hit_rows = int(row.get("HitRows") or 0)
        english = en if not en2 else f"{en} / {en2}"
        terms.append(f"{cn} -> {english} ({hit_rows})" if english else f"{cn} ({hit_rows})")
    return terms


def style_guidance(signals: list[dict[str, object]], categories: Counter[str], target_coverage: int) -> list[str]:
    labels = {str(signal["label"]) for signal in signals}
    guidance = [
        "游戏内容/UI 部分尽量精简，适配移动端按钮、弹窗、任务和道具说明。",
        "剧情对话必须自然、地道、通顺，参考美剧日常对白节奏，避免逐字直译。",
        "变量、数字、换行、富文本标签和占位符必须原样保留。",
    ]
    if "合成/经营" in labels:
        guidance.append("合成、订单、生产、仓库等玩法术语保持统一，不在不同系统中来回换词。")
    if "花店/装修" in labels or "休闲/女性向" in labels:
        guidance.append("整体语气偏温暖、轻松、生活化，避免硬核、军工或过度严肃的表达。")
    if "战斗/RPG养成" in labels or categories.get("战斗属性/数值", 0):
        guidance.append("战斗、属性、技能说明优先准确表达机制，不使用夸张营销词替代数值含义。")
    if "基地/建筑经营" in labels:
        guidance.append("建筑、基地、升级线采用稳定系统名；同一建筑不要在 HQ/Base/Headquarters 之间漂移。")
    if "活动/商业化" in labels:
        guidance.append("活动、礼包、商店文案可以有吸引力，但必须短、明确、不过度夸张。")
    if "社交/公会竞争" in labels:
        guidance.append("公会、排行、竞技相关术语保持玩家社区常用表达，如 Guild、Ranking、Arena。")
    if "剧情/叙事" in labels:
        guidance.append("角色对话要保留人物关系、情绪冲突和轻喜剧节奏，可使用自然口语与缩写。")
    if target_coverage:
        guidance.append("已有英文译文视为项目历史用法；当它和术语表冲突时，优先检查是否属于手动适配 EN2。")
    return guidance


def project_type_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    if "飞行/射击题材" in labels and "战斗/RPG养成" in labels:
        return "科幻战机 / 飞行射击 / RPG养成"
    if "飞行/射击题材" in labels:
        return "飞行射击"
    if "战斗/RPG养成" in labels and "社交/公会竞争" in labels and "基地/建筑经营" in labels:
        return "战斗/RPG养成 / 轻SLG"
    if {"合成/经营", "花店/装修", "剧情/叙事"} <= labels:
        return "合成经营 / 花店修复 / 轻剧情休闲"
    if {"合成/经营", "剧情/叙事"} <= labels:
        return "合成经营 / 轻剧情休闲"
    if "花店/装修" in labels:
        return "花店装修 / 休闲经营"
    if "休闲/女性向" in labels:
        return "女性向休闲"
    if "战斗/RPG养成" in labels:
        return "战斗/RPG养成"
    if "剧情/叙事" in labels:
        return "剧情向休闲"
    return "移动游戏"


def target_user_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    if "飞行/射击题材" in labels:
        return "偏中重度、喜欢战机养成、战斗数值、装备强化和活动推进的移动端玩家。"
    if "战斗/RPG养成" in labels:
        return "偏中度、关注战力成长、英雄/装备养成、活动奖励和竞技排名的移动端玩家。"
    if "花店/装修" in labels or "休闲/女性向" in labels:
        return "偏休闲、喜欢合成/装修/经营和轻剧情的女性向或轻度玩家。"
    if "合成/经营" in labels:
        return "喜欢轻策略、收集、订单推进和长期经营成长的休闲玩家。"
    if "剧情/叙事" in labels:
        return "重视角色关系、剧情推进和自然对白体验的玩家。"
    return "移动端游戏玩家。"


def signal_hit_map(signals: list[dict[str, object]]) -> dict[str, int]:
    return {str(signal["label"]): int(signal["hit_rows"]) for signal in signals}


def content_focus_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    hits = signal_hit_map(signals)
    focus: list[str] = []
    if "飞行/射击题材" in labels:
        focus.append("战机、导弹、射击、弹幕等战斗内容")
    if "战斗/RPG养成" in labels:
        focus.append("英雄、装备、技能、属性和战力成长")
    if "基地/建筑经营" in labels:
        focus.append("建造、升级、采集、生产等基地系统")
    if hits.get("合成/经营", 0) >= 20:
        focus.append("合成、订单、生产、仓库等玩法 UI")
    if hits.get("花店/装修", 0) >= 10:
        focus.append("花店修复、装饰和生活化物件")
    if "剧情/叙事" in labels:
        focus.append("角色剧情对话")
    if "活动/商业化" in labels:
        focus.append("活动、礼包和奖励")
    return "；".join(focus) if focus else "系统 UI、玩法说明和剧情文本"


def tone_rule_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    if "飞行/射击题材" in labels:
        return "整体语气冷静、利落、偏科幻军事；战机、装备、导弹、技能和战斗数值要专业清晰，避免可爱化、生活化或过度口语化。"
    if "战斗/RPG养成" in labels:
        return "整体语气清晰、有力量感；战斗、英雄、装备和数值成长要准确直接，避免弱化机制或夸张营销。"
    if "花店/装修" in labels or "休闲/女性向" in labels:
        return "整体语气偏轻松、温暖、生活化；涉及花店、装修、订单、合成、经营时避免硬核或过度严肃的表达。"
    if "合成/经营" in labels:
        return "整体语气轻松、清晰、偏休闲；合成、订单、生产和仓库说明要短句化，避免复杂长句。"
    return "整体语气清晰、自然、符合移动游戏语境；不要为了润色改变玩法含义。"


def build_translation_prompt(
    project_name: str,
    signals: list[dict[str, object]],
    categories: Counter[str],
    key_terms: list[str],
    target_coverage: int,
) -> str:
    project_type = project_type_from_signals(signals)
    tone_rule = tone_rule_from_signals(signals)
    term_rule = "关键术语以随附术语表为准，EN 为标准译法，EN2 为项目中稳定出现的手动适配译法。"
    if not key_terms:
        term_rule = "如未提供术语表，需先从上下文判断固定系统名，保持同一中文术语的英文一致。"
    existing_en_rule = (
        "已有英文译文代表项目历史用法；如现有译法不自然，可以优化，但不要破坏已固定的系统术语。"
        if target_coverage
        else "当前输入可能没有英文列；先按项目类型和术语表建立统一英语风格。"
    )
    return "\n".join(
        [
            f"你是一位资深游戏本地化译者，正在翻译《{project_name}》这款{project_type}游戏。",
            "译文需符合以下要求：",
            "1. 游戏内容/UI/玩法说明尽量精简，适配移动游戏按钮、弹窗、任务、道具和奖励说明；",
            "2. 剧情对话必须自然、地道、通顺，参考美剧日常对白节奏，保留角色语气、冲突、幽默和情绪，不要逐字直译；",
            f"3. {tone_rule}",
            f"4. {term_rule}",
            f"5. {existing_en_rule}",
            "6. 保留所有游戏代码、变量、数字、换行、颜色标签、HTML/富文本标签和占位符，如 {0}、%s、<color> 等；",
            "7. 无法确认的专有名词或信息缺口用 [TBD] 标记，不要自行编造设定。",
        ]
    )


def build_project_brief(
    project_name: str,
    sheet_name: str,
    records: list[Record],
    all_rows: list[dict[str, object]],
    glossary_rows: list[dict[str, object]],
    manual_rows: list[dict[str, object]],
    material_sources: list[str] | None = None,
) -> tuple[str, str]:
    source_rows = len(records)
    target_coverage = sum(1 for record in records if record.target)
    signals = infer_project_signals(records, limit=10)
    categories = category_distribution(glossary_rows or all_rows)
    key_terms = top_terms(glossary_rows or all_rows)
    project_type = project_type_from_signals(signals)
    target_user = target_user_from_signals(signals)
    content_focus = content_focus_from_signals(signals)
    tone_rule = tone_rule_from_signals(signals)
    prompt = build_translation_prompt(
        project_name=project_name,
        signals=signals,
        categories=categories,
        key_terms=key_terms,
        target_coverage=target_coverage,
    )

    markdown = "\n".join(
        [
            f"# {project_name} 翻译提示词与项目元信息",
            "",
            "## 🤖 AI 生成的专属翻译提示词",
            "",
            "```",
            prompt,
            "```",
            "",
            "## 📌 项目元信息",
            "",
            markdown_table(
                ["项目", "信息"],
                [
                    ["游戏类型", project_type],
                    ["目标用户", target_user],
                    ["内容构成", content_focus],
                    ["翻译风格", f"UI/玩法精简适配移动端；剧情自然、地道、通顺，参考美剧日常对白；{tone_rule}"],
                    ["信息来源", "语言表" if not material_sources else "语言表；" + "；".join(material_sources[:6])],
                    ["语言资产", f"{source_rows} 条文本，已有英文 {target_coverage} 条。"],
                    ["生成日期", datetime.now().strftime("%Y-%m-%d")],
                ],
            ),
            "",
        ]
    )
    return markdown, prompt


def write_text_output(output_path: Path, content: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding="utf-8")


def append_rows(worksheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_sheet(worksheet)


def write_detail_workbook(
    output_path: Path,
    sheet_name: str,
    records: list[Record],
    all_rows: list[dict[str, object]],
    glossary_rows: list[dict[str, object]],
    high_risk_rows: list[dict[str, object]],
    manual_rows: list[dict[str, object]],
    curated_rules_path: Path | None,
    observations_store_path: Path | None,
) -> None:
    workbook = Workbook()
    headers = [
        "ID",
        "CN",
        "EN",
        "EN2",
        "SuggestedEN",
        "ExactCandidates",
        "ExampleUsages",
        "ManualAdaptations",
        "ActualShortUsages",
        "HasActualDiff",
        "DiffType",
        "DiffVariants",
        "SameOrFormatOnlyCount",
        "DiffCount",
        "Category",
        "Risk",
        "Priority",
        "HitRows",
        "ExactRows",
        "ExampleID",
        "ExampleSource",
        "ExampleEN",
        "DiffExampleID",
        "DiffExampleSource",
        "DiffExampleEN",
        "Note",
    ]

    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    append_rows(glossary_sheet, headers, glossary_rows)

    high_risk_sheet = workbook.create_sheet("HighRisk")
    append_rows(high_risk_sheet, headers, high_risk_rows)

    manual_sheet = workbook.create_sheet("ManualAdaptation")
    append_rows(manual_sheet, headers, manual_rows)

    all_sheet = workbook.create_sheet("Candidates")
    append_rows(all_sheet, headers, all_rows)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["Item", "Value"])
    for item, value in [
        ("SourceRows", len(records)),
        ("Sheet", sheet_name),
        ("CandidateTerms", len(all_rows)),
        ("GlossaryRows", len(glossary_rows)),
        ("HighRiskRows", len(high_risk_rows)),
        ("ManualAdaptationRows", len(manual_rows)),
        ("CuratedRules", str(curated_rules_path) if curated_rules_path else ""),
        ("ObservationsStore", str(observations_store_path) if observations_store_path else ""),
        ("Rule", "Extract short source terms from the source column and use target column only for English alignment and drift checks."),
        ("ManualAdaptation", "A term is marked as manual adaptation when short target usages introduce a stable wording different from the example EN."),
        ("LearningModel", "Curated rules keep approved EN/EN2 decisions; observation store accumulates seen variants and usage drift."),
    ]:
        notes_sheet.append([item, value])
    style_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def write_final_workbook(output_path: Path, final_rows: list[dict[str, object]]) -> None:
    workbook = Workbook()

    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    final_headers = ["ID", "CN", "EN", "EN2"]
    glossary_sheet.append(final_headers)
    for row in final_rows:
        glossary_sheet.append([row.get(header, "") for header in final_headers])
    style_sheet(glossary_sheet)

    detail_sheet = workbook.create_sheet("Buckets")
    detail_headers = ["ID", "CN", "EN", "EN2", "ExampleUsages", "ManualAdaptations", "Note"]
    detail_sheet.append(detail_headers)
    for row in final_rows:
        detail_sheet.append([row.get(header, "") for header in detail_headers])
    style_sheet(detail_sheet)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["Item", "Value"])
    for item, value in [
        ("Columns", "ID = text id, CN = source term, EN = example English, EN2 = manual adaptation English"),
        ("Rule", "EN2 remains blank when the alternative wording is not stable enough or is explicitly blocked by curated rules."),
        ("RowCount", len(final_rows)),
    ]:
        notes_sheet.append([item, value])
    style_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def display_header_name(header: str, default_header: str) -> str:
    clean_header = clean_text(header)
    return default_header if clean_header.lower() == default_header.lower() else clean_header


def write_announcement_glossary_workbook(
    output_path: Path,
    matched_rows: list[dict[str, object]],
    id_header: str,
    source_header: str,
    target_header: str,
    headers: list[str] | None = None,
) -> None:
    workbook = Workbook()
    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    output_headers = headers or [
        display_header_name(id_header, "ID"),
        display_header_name(source_header, "CN"),
        display_header_name(target_header, "EN"),
    ]
    glossary_sheet.append(output_headers)
    for row in matched_rows:
        source_values = row.get("_AnnouncementValues")
        if isinstance(source_values, list):
            values = source_values[: len(output_headers)]
            if len(values) < len(output_headers):
                values.extend([""] * (len(output_headers) - len(values)))
            glossary_sheet.append(values)
        else:
            values = []
            for header in output_headers:
                if clean_text(header).lower() == "en":
                    values.append(clean_text(row.get("EN")) or clean_text(row.get("EN2")))
                else:
                    values.append(row.get(header, ""))
            glossary_sheet.append(values)
    style_sheet(glossary_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def parse_language_table_spec(raw_spec: str) -> LanguageTableSpec:
    if "=" not in raw_spec:
        raise ValueError(f"Invalid --language-table value '{raw_spec}'. Expected LANG=path.")
    raw_language, raw_path = raw_spec.split("=", 1)
    language = clean_text(raw_language).upper()
    if not language or not re.match(r"^[A-Z0-9_-]+$", language):
        raise ValueError(f"Invalid language code in --language-table value '{raw_spec}'.")
    if not raw_path.strip():
        raise ValueError(f"Missing path in --language-table value '{raw_spec}'.")
    path = Path(raw_path.strip())
    return LanguageTableSpec(language=language, path=path)


def parse_language_table_specs(raw_specs: list[str]) -> list[LanguageTableSpec]:
    specs = [parse_language_table_spec(raw_spec) for raw_spec in raw_specs]
    seen_languages: set[str] = set()
    for spec in specs:
        if spec.language in seen_languages:
            raise ValueError(f"Duplicate --language-table language code: {spec.language}")
        seen_languages.add(spec.language)
    return specs


def build_multilingual_announcement_rows(
    language_table_specs: list[LanguageTableSpec],
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    curated_rules: dict[str, Any],
    announcement_min_hit: int,
    source_only: bool,
    announcement_text: str,
    include_empty: bool = False,
) -> tuple[list[dict[str, object]], dict[str, int]]:
    candidate_by_cn: dict[str, dict[str, object]] = {}
    translations_by_language: dict[str, dict[str, str]] = {spec.language: {} for spec in language_table_specs}
    duplicate_source_terms = 0

    for spec in language_table_specs:
        _headers, candidate_rows = build_announcement_candidate_rows_from_workbook(
            input_path=spec.path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=spec.language,
            curated_rules=curated_rules,
            min_hit=announcement_min_hit,
            source_only=source_only,
        )
        for row in candidate_rows:
            cn = clean_text(row.get("CN"))
            if not cn:
                continue
            target = clean_text(row.get("EN")) or clean_text(row.get("EN2"))
            if target:
                translations_by_language[spec.language][cn] = target
            candidate = candidate_by_cn.get(cn)
            if candidate is None:
                candidate_by_cn[cn] = {
                    "ID": row.get("ID", ""),
                    "CN": cn,
                    "EN": target,
                }
            else:
                if clean_text(candidate.get("ID")) and clean_text(row.get("ID")) and clean_text(candidate.get("ID")) != clean_text(row.get("ID")):
                    duplicate_source_terms += 1
                if not clean_text(candidate.get("EN")) and target:
                    candidate["EN"] = target

    matched_terms = select_announcement_term_rows(
        term_rows=list(candidate_by_cn.values()),
        announcement_text=announcement_text,
        include_empty=include_empty,
    )

    rows: list[dict[str, object]] = []
    for matched in matched_terms:
        cn = clean_text(matched.get("CN"))
        row = {
            "ID": matched.get("ID", ""),
            "CN": cn,
        }
        for spec in language_table_specs:
            row[spec.language] = translations_by_language[spec.language].get(cn, "")
        rows.append(row)

    stats = {
        "candidate_terms": len(candidate_by_cn),
        "duplicate_source_terms": duplicate_source_terms,
    }
    return rows, stats


def ai_announcement_query_terms(announcement_text: str, max_terms: int = 800) -> list[str]:
    normalized_notice = clean_text(announcement_text)
    terms: set[str] = set()
    for quoted in QUOTED_TERM_RE.findall(normalized_notice):
        quoted_term = clean_text(quoted)
        if 2 <= len(quoted_term) <= 20 and CJK_RE.search(quoted_term):
            terms.add(quoted_term)
    for run in CJK_RUN_RE.findall(normalized_notice):
        if 2 <= len(run) <= 12:
            terms.add(run)
        upper = min(8, len(run))
        for size in range(upper, 1, -1):
            for start in range(0, len(run) - size + 1):
                terms.add(run[start : start + size])
                if len(terms) >= max_terms:
                    break
            if len(terms) >= max_terms:
                break
        if len(terms) >= max_terms:
            break
    return sorted(terms, key=lambda item: (-len(item), item))[:max_terms]


def compact_announcement_row(row: dict[str, object], headers: list[str]) -> dict[str, object]:
    compact: dict[str, object] = {}
    values = announcement_output_values(row, headers)
    for index, header in enumerate(headers):
        compact[header] = values[index] if index < len(values) else ""
    return compact


def evidence_target_for_row(row: dict[str, object], headers: list[str]) -> tuple[str, str]:
    for header in headers[2:]:
        value = clean_text(row.get(header))
        if value:
            return clean_text(header), value
    en = clean_text(row.get("EN")) or clean_text(row.get("EN2"))
    if en:
        return "EN", en
    source_values = row.get("_AnnouncementValues")
    if isinstance(source_values, list) and len(source_values) >= 3:
        return clean_text(headers[2] if len(headers) >= 3 else "EN"), clean_text(source_values[2])
    return clean_text(headers[2] if len(headers) >= 3 else "EN"), ""


def ai_evidence_candidate_rows_from_sheet_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    language: str,
    source_only: bool = False,
) -> list[dict[str, object]]:
    layout = language_table_layout_from_rows(
        rows=rows,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    )
    if layout is None:
        return []
    evidence_rows: list[dict[str, object]] = []
    language_header = clean_text(language) or "EN"
    for row_number, row in enumerate(rows[layout.header_row_index + 1 :], start=layout.header_row_index + 2):
        row_values = list(row)
        source_text = clean_text(value_at(row_values, layout.source_index))
        if not source_text or not CJK_RE.search(source_text):
            continue
        target_text = "" if layout.target_index is None else clean_text(value_at(row_values, layout.target_index))
        if not target_text:
            continue
        row_id = clean_text(value_at(row_values, layout.id_index)) or f"{sheet_title}:{row_number}"
        evidence_rows.append(
            {
                "ID": row_id,
                "CN": source_text,
                language_header: target_text,
                "EN": target_text if language_header == "EN" else "",
            }
        )
    return evidence_rows


def build_ai_evidence_candidate_rows_from_workbook(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    language: str,
    source_only: bool = False,
) -> list[dict[str, object]]:
    evidence_rows: list[dict[str, object]] = []
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheets = [workbook[sheet_name]] if sheet_name else list(workbook.worksheets)
        for worksheet in worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            evidence_rows.extend(
                ai_evidence_candidate_rows_from_sheet_rows(
                    rows=rows,
                    sheet_title=worksheet.title,
                    id_column=id_column,
                    source_column=source_column,
                    target_column=target_column,
                    language=language,
                    source_only=source_only,
                )
            )
        workbook.close()
        return evidence_rows
    except Exception:
        for raw_sheet_name, rows in iter_raw_xlsx_sheets(input_path):
            if sheet_name and raw_sheet_name != sheet_name:
                continue
            evidence_rows.extend(
                ai_evidence_candidate_rows_from_sheet_rows(
                    rows=rows,
                    sheet_title=raw_sheet_name,
                    id_column=id_column,
                    source_column=source_column,
                    target_column=target_column,
                    language=language,
                    source_only=source_only,
                )
            )
        return evidence_rows


def build_ai_supplement_packet(
    announcement_text: str,
    matched_rows: list[dict[str, object]],
    candidate_rows: list[dict[str, object]],
    headers: list[str],
    project_name: str = "",
    evidence_limit: int = AI_SUPPLEMENT_EVIDENCE_LIMIT,
    evidence_per_term: int = AI_SUPPLEMENT_EVIDENCE_PER_TERM,
) -> dict[str, object]:
    normalized_notice = clean_text(announcement_text)
    matched_terms = {clean_text(row.get("CN")) for row in matched_rows if clean_text(row.get("CN"))}
    query_terms = [term for term in ai_announcement_query_terms(normalized_notice) if term not in matched_terms]
    evidence_rows: list[dict[str, object]] = []
    per_term_counts: Counter[str] = Counter()
    seen_evidence_ids: Counter[str] = Counter()

    for row in candidate_rows:
        source_text = clean_text(row.get("CN"))
        if not source_text or source_text in matched_terms:
            continue
        matched_query = next((term for term in query_terms if term in source_text), "")
        if not matched_query:
            continue
        if per_term_counts[matched_query] >= evidence_per_term:
            continue
        language, target_text = evidence_target_for_row(row, headers)
        if not target_text:
            continue
        raw_id = clean_text(row.get("ID")) or f"evidence-{len(evidence_rows) + 1}"
        seen_evidence_ids[raw_id] += 1
        evidence_id = raw_id if seen_evidence_ids[raw_id] == 1 else f"{raw_id}#{seen_evidence_ids[raw_id]}"
        evidence_rows.append(
            {
                "evidence_id": evidence_id,
                "ID": raw_id,
                "source_text": source_text,
                "target_text": target_text,
                "language": language,
                "reason": f"announcement_overlap:{matched_query}",
            }
        )
        per_term_counts[matched_query] += 1
        if len(evidence_rows) >= evidence_limit:
            break

    uncovered_text = normalized_notice
    for term in sorted(matched_terms, key=len, reverse=True):
        uncovered_text = uncovered_text.replace(term, "")
    packet = {
        "schema_version": AI_SUPPLEMENT_SCHEMA_VERSION,
        "task": "announcement_ai_supplement",
        "instructions": [
            "Only propose terms that appear in announcement_text.",
            "Prefer game-specific system, event, item, mode, character, and proper-name terms.",
            "Use evidence_rows only; do not invent translations without language-table evidence.",
            "Return JSON with supplement_terms: cn, translations, source_ids, confidence, reason, evidence_ids, action.",
        ],
        "project_name": clean_text(project_name),
        "announcement_text": normalized_notice,
        "uncovered_announcement_text": clean_text(uncovered_text),
        "headers": headers,
        "matched_terms": [compact_announcement_row(row, headers) for row in matched_rows],
        "evidence_rows": evidence_rows,
        "response_schema": {
            "supplement_terms": [
                {
                    "cn": "术语中文",
                    "translations": {"EN": "Term translation"},
                    "source_ids": ["language-table ID"],
                    "confidence": "low|medium|high",
                    "reason": "why this is a term",
                    "evidence_ids": ["evidence_id"],
                    "action": "add_to_main|report_only|reject",
                }
            ]
        },
    }
    return packet


def ai_response_terms(response: dict[str, object]) -> list[dict[str, object]]:
    terms = response.get("supplement_terms", [])
    if not isinstance(terms, list):
        return []
    return [term for term in terms if isinstance(term, dict)]


def evidence_lookup(packet: dict[str, object]) -> dict[str, dict[str, object]]:
    lookup: dict[str, dict[str, object]] = {}
    evidence_rows = packet.get("evidence_rows", [])
    if not isinstance(evidence_rows, list):
        return lookup
    for item in evidence_rows:
        if not isinstance(item, dict):
            continue
        evidence_id = clean_text(item.get("evidence_id"))
        row_id = clean_text(item.get("ID"))
        if evidence_id:
            lookup[evidence_id] = item
        if row_id:
            lookup[row_id] = item
    return lookup


def project_name_translation_missing(project_name: str, rows: list[dict[str, object]], headers: list[str]) -> bool:
    normalized_project = clean_text(project_name)
    if not normalized_project:
        return False
    for row in rows:
        if clean_text(row.get("CN")) != normalized_project:
            continue
        if any(clean_text(row.get(header)) for header in headers[2:]):
            return False
        if clean_text(row.get("EN")) or clean_text(row.get("EN2")):
            return False
    return True


def apply_ai_supplement_response(
    announcement_rows: list[dict[str, object]],
    headers: list[str],
    announcement_text: str,
    packet: dict[str, object],
    response: dict[str, object],
    project_name: str = "",
) -> tuple[list[dict[str, object]], dict[str, object]]:
    normalized_notice = clean_text(announcement_text)
    evidence_by_id = evidence_lookup(packet)
    merged_rows = [dict(row) for row in announcement_rows]
    existing_terms = {clean_text(row.get("CN")) for row in merged_rows if clean_text(row.get("CN"))}
    report_terms: list[dict[str, object]] = []

    for term in ai_response_terms(response):
        cn = clean_text(term.get("cn"))
        action = clean_text(term.get("action")) or "report_only"
        confidence = clean_text(term.get("confidence")).lower()
        translations = term.get("translations", {})
        translations = translations if isinstance(translations, dict) else {}
        evidence_ids = term.get("evidence_ids", [])
        source_ids = term.get("source_ids", [])
        evidence_keys = [
            clean_text(item)
            for item in ([*evidence_ids, *source_ids] if isinstance(evidence_ids, list) and isinstance(source_ids, list) else [])
            if clean_text(item)
        ]
        evidence_items = [evidence_by_id[key] for key in evidence_keys if key in evidence_by_id]
        has_evidence = any(cn and cn in clean_text(item.get("source_text")) for item in evidence_items)
        has_translation = any(clean_text(translations.get(header)) for header in headers[2:])
        if not has_translation:
            has_translation = any(clean_text(value) for value in translations.values())
        missing_languages = [header for header in headers[2:] if not clean_text(translations.get(header))]
        can_add = (
            action == "add_to_main"
            and cn
            and cn in normalized_notice
            and cn not in existing_terms
            and AI_CONFIDENCE_RANK.get(confidence, -1) >= AI_CONFIDENCE_RANK["medium"]
            and has_evidence
            and has_translation
        )
        status = "added_to_main" if can_add else ("rejected" if action == "reject" else "report_only")
        report_terms.append(
            {
                "cn": cn,
                "confidence": confidence,
                "action": action,
                "status": status,
                "reason": clean_text(term.get("reason")),
                "evidence_ids": evidence_keys,
                "missing_languages": missing_languages,
                "translations": {str(key): clean_text(value) for key, value in translations.items()},
            }
        )
        if not can_add:
            continue

        first_evidence = evidence_items[0]
        output_row: dict[str, object] = {
            "ID": clean_text(first_evidence.get("ID")),
            "CN": cn,
        }
        for header in headers[2:]:
            output_row[header] = clean_text(translations.get(header))
        if "EN" in headers and not clean_text(output_row.get("EN")):
            output_row["EN"] = clean_text(translations.get("EN"))
        merged_rows.append(output_row)
        existing_terms.add(cn)

    missing_project_name = project_name_translation_missing(project_name, merged_rows, headers)
    report = {
        "schema_version": AI_SUPPLEMENT_SCHEMA_VERSION,
        "terms": report_terms,
        "project_name": clean_text(project_name),
        "project_name_translation_missing": missing_project_name,
    }
    return merged_rows, report


def build_multilingual_ai_candidate_rows(
    language_table_specs: list[LanguageTableSpec],
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    curated_rules: dict[str, Any],
    announcement_min_hit: int,
    source_only: bool,
) -> list[dict[str, object]]:
    rows_by_cn: dict[str, dict[str, object]] = {}
    for spec in language_table_specs:
        candidate_rows = build_ai_evidence_candidate_rows_from_workbook(
            input_path=spec.path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=spec.language,
            language=spec.language,
            source_only=source_only,
        )
        cn_counts = Counter(clean_text(row.get("CN")) for row in candidate_rows if clean_text(row.get("CN")))
        for candidate in candidate_rows:
            cn = clean_text(candidate.get("CN"))
            if not cn:
                continue
            if cn_counts[cn] < announcement_min_hit:
                continue
            curated_state = get_curated_term_state(curated_rules, cn, create=False)
            if curated_state.get("ignore"):
                continue
            row = rows_by_cn.setdefault(cn, {"ID": candidate.get("ID", ""), "CN": cn})
            target = clean_text(candidate.get(spec.language)) or clean_text(candidate.get("EN"))
            if target:
                row[spec.language] = target
                if spec.language == "EN":
                    row["EN"] = target
    return list(rows_by_cn.values())


def write_json_output(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_ai_supplement_report_markdown(
    report: dict[str, object],
    packet_path: Path | None,
    response_path: Path | None,
    output_path: Path,
) -> str:
    terms = report.get("terms", [])
    terms = terms if isinstance(terms, list) else []
    lines = [
        "# AI Supplement Report",
        "",
        "status: ok",
        f"packet: {packet_path or 'disabled'}",
        f"response: {response_path or 'not provided'}",
        f"output: {output_path}",
        f"term_count: {len(terms)}",
        f"added_to_main: {sum(1 for term in terms if isinstance(term, dict) and term.get('status') == 'added_to_main')}",
        f"report_only: {sum(1 for term in terms if isinstance(term, dict) and term.get('status') == 'report_only')}",
        "",
    ]
    if report.get("project_name_translation_missing"):
        lines.extend(
            [
                "## Project Name Warning",
                "",
                f"请补充项目名标准译文：{report.get('project_name', '')}",
                "",
            ]
        )
    lines.append("## Terms")
    if not terms:
        lines.append("")
        lines.append("- No AI supplement response terms.")
    for term in terms:
        if not isinstance(term, dict):
            continue
        lines.append(
            f"- {term.get('cn', '')} | status={term.get('status', '')} | confidence={term.get('confidence', '')} | evidence={', '.join(term.get('evidence_ids', [])) if isinstance(term.get('evidence_ids'), list) else ''}"
        )
        missing_languages = term.get("missing_languages", [])
        if isinstance(missing_languages, list) and missing_languages:
            lines.append(f"  missing_languages: {', '.join(str(item) for item in missing_languages)}")
        reason = clean_text(term.get("reason"))
        if reason:
            lines.append(f"  reason: {reason}")
    return "\n".join(lines) + "\n"


def run_ai_supplement_flow(
    announcement_rows: list[dict[str, object]],
    announcement_candidate_rows: list[dict[str, object]],
    announcement_text: str,
    headers: list[str],
    project_name: str,
    packet_output_path: Path,
    report_output_path: Path,
    response_path: Path | None,
) -> tuple[list[dict[str, object]], dict[str, object], Path, Path]:
    packet = build_ai_supplement_packet(
        announcement_text=announcement_text,
        matched_rows=announcement_rows,
        candidate_rows=announcement_candidate_rows,
        headers=headers,
        project_name=project_name,
    )
    write_json_output(packet_output_path, packet)
    response: dict[str, object] = {"supplement_terms": []}
    if response_path is not None:
        response = FileAiSupplementProvider(response_path).generate(packet)
    merged_rows, report = apply_ai_supplement_response(
        announcement_rows=announcement_rows,
        headers=headers,
        announcement_text=announcement_text,
        packet=packet,
        response=response,
        project_name=project_name,
    )
    report_markdown = build_ai_supplement_report_markdown(
        report=report,
        packet_path=packet_output_path,
        response_path=response_path,
        output_path=report_output_path,
    )
    write_text_output(report_output_path, report_markdown)
    return merged_rows, report, packet_output_path, report_output_path


def announcement_output_values(row: dict[str, object], headers: list[str]) -> list[object]:
    source_values = row.get("_AnnouncementValues")
    if isinstance(source_values, list):
        values = source_values[: len(headers)]
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        return values
    values: list[object] = []
    for header in headers:
        if clean_text(header).lower() == "en":
            values.append(clean_text(row.get("EN")) or clean_text(row.get("EN2")))
        else:
            values.append(row.get(header, ""))
    return values


def build_announcement_validation_markdown(
    announcement_materials: list[Path],
    language_tables: list[str],
    glossary_output_path: Path,
    rows: list[dict[str, object]],
    headers: list[str],
    stats: dict[str, int] | None = None,
) -> str:
    stats = stats or {}
    cn_values = [clean_text(row.get("CN")) for row in rows if clean_text(row.get("CN"))]
    duplicate_cn = len(cn_values) - len(set(cn_values))
    language_headers = headers[2:]
    empty_translation_cells = 0
    for row in rows:
        values = announcement_output_values(row, headers)
        for index in range(2, len(headers)):
            if index >= len(values) or not clean_text(values[index]):
                empty_translation_cells += 1
    low_value_terms = sum(1 for row in rows if is_low_value_announcement_term(clean_text(row.get("CN"))))

    lines = [
        "# Announcement Glossary Validation",
        "",
        "status: ok",
        f"term_count: {len(rows)}",
        f"languages: {', '.join(language_headers) if language_headers else 'none'}",
        f"duplicate_cn: {duplicate_cn}",
        f"duplicate_source_terms: {int(stats.get('duplicate_source_terms', 0))}",
        f"empty_translation_cells: {empty_translation_cells}",
        f"missing_language_values: {empty_translation_cells}",
        f"low_value_terms: {low_value_terms}",
        f"candidate_terms: {int(stats.get('candidate_terms', len(rows)))}",
        f"output: {glossary_output_path}",
        "",
        "## Announcement Materials",
    ]
    lines.extend(f"- {path}" for path in announcement_materials)
    lines.append("")
    lines.append("## Language Tables")
    lines.extend(f"- {source}" for source in language_tables)
    lines.append("")
    return "\n".join(lines)


def write_announcement_validation_report(
    output_path: Path,
    announcement_materials: list[Path],
    language_tables: list[str],
    glossary_output_path: Path,
    rows: list[dict[str, object]],
    headers: list[str],
    stats: dict[str, int] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        build_announcement_validation_markdown(
            announcement_materials=announcement_materials,
            language_tables=language_tables,
            glossary_output_path=glossary_output_path,
            rows=rows,
            headers=headers,
            stats=stats,
        ),
        encoding="utf-8",
    )


def default_output_paths(input_path: Path, detail_output: str | None, final_output: str | None) -> tuple[Path, Path]:
    date_suffix = datetime.now().strftime("%Y%m%d")
    detail_path = Path(detail_output) if detail_output else input_path.with_name(
        f"{input_path.stem}_glossary_details_{date_suffix}.xlsx"
    )
    final_path = Path(final_output) if final_output else input_path.with_name(
        f"{input_path.stem}_ID_CN_EN_EN2_{date_suffix}.xlsx"
    )
    return detail_path, final_path


def default_project_brief_output_path(input_path: Path, project_brief_output: str | None) -> Path:
    date_suffix = datetime.now().strftime("%Y%m%d")
    return Path(project_brief_output) if project_brief_output else input_path.with_name(
        f"{input_path.stem}_project_brief_{date_suffix}.md"
    )


def default_announcement_output_path(material_paths: list[Path], announcement_output: str | None) -> Path | None:
    if announcement_output:
        return Path(announcement_output)
    if not material_paths:
        return None
    date_suffix = datetime.now().strftime("%Y%m%d")
    first_material = material_paths[0]
    return first_material.with_name(f"{first_material.stem}_announcement_terms_{date_suffix}.xlsx")


def default_announcement_validation_output_path(
    material_paths: list[Path],
    announcement_validation_output: str | None,
) -> Path | None:
    if announcement_validation_output:
        return Path(announcement_validation_output)
    return None


def default_ai_supplement_packet_output_path(
    material_paths: list[Path],
    ai_supplement_packet_output: str | None,
) -> Path | None:
    if ai_supplement_packet_output:
        return Path(ai_supplement_packet_output)
    if not material_paths:
        return None
    date_suffix = datetime.now().strftime("%Y%m%d")
    first_material = material_paths[0]
    return first_material.with_name(f"{first_material.stem}_ai_packet_{date_suffix}.json")


def default_ai_supplement_report_output_path(
    material_paths: list[Path],
    ai_supplement_report_output: str | None,
) -> Path | None:
    if ai_supplement_report_output:
        return Path(ai_supplement_report_output)
    if not material_paths:
        return None
    date_suffix = datetime.now().strftime("%Y%m%d")
    first_material = material_paths[0]
    return first_material.with_name(f"{first_material.stem}_ai_supplement_{date_suffix}.md")


def should_run_announcement_only(args: argparse.Namespace) -> bool:
    return bool(args.announcement_material) and not any(
        [
            args.output,
            args.final_output,
            args.project_brief_output,
            args.translation_prompt_output,
            args.project_material,
            args.project_note,
        ]
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract glossary terms from a game localization language table.")
    parser.add_argument("input_path", nargs="?", help="Path to the source XLSX language table.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the first sheet.")
    parser.add_argument("--id-column", default="ID", help="ID column header. Default: ID")
    parser.add_argument("--source-column", default="cn", help="Source text column header. Default: cn")
    parser.add_argument("--target-column", default="en", help="Target text column header. Default: en")
    parser.add_argument(
        "--language-table",
        action="append",
        default=[],
        help="Announcement lookup language table in LANG=path form. Can be repeated, for example EN=table.xlsx.",
    )
    parser.add_argument(
        "--source-only",
        action="store_true",
        help="Treat the input as source text only and do not require a target text column.",
    )
    parser.add_argument(
        "--include-empty-final-terms",
        action="store_true",
        help="Keep final glossary rows even when EN and EN2 are blank. Useful for source-only extraction.",
    )
    parser.add_argument("--min-hit", type=int, default=5, help="Minimum hit count to keep a candidate. Default: 5")
    parser.add_argument(
        "--glossary-hit-threshold",
        type=int,
        default=10,
        help="Minimum hit count to include a candidate in the delivery glossary unless it is high risk. Default: 10",
    )
    parser.add_argument("--output", help="Path for the detailed workbook output.")
    parser.add_argument("--final-output", help="Path for the clean delivery workbook output.")
    parser.add_argument(
        "--curated-rules",
        default=str(DEFAULT_CURATED_RULES),
        help="Path to the curated glossary rules JSON file. Default: data/experience/curated_terms.json",
    )
    parser.add_argument(
        "--observations-store",
        default=str(DEFAULT_OBSERVATIONS_STORE),
        help="Path to the observed term usage JSON file. Default: data/experience/observed_terms.json",
    )
    parser.add_argument(
        "--project-name",
        help="Project name used in the project brief. Defaults to the input file stem.",
    )
    parser.add_argument(
        "--project-brief-output",
        help="Path for the project audit Markdown output. Defaults to *_project_brief_YYYYMMDD.md.",
    )
    parser.add_argument(
        "--translation-prompt-output",
        help="Optional path for a prompt-only text output extracted from the project brief.",
    )
    parser.add_argument(
        "--project-material",
        action="append",
        default=[],
        help="Additional project material path for brief generation. Can be repeated. Supports txt/md/json/csv/tsv/xlsx and image filenames.",
    )
    parser.add_argument(
        "--project-note",
        action="append",
        default=[],
        help="Additional project note or image observation used for brief generation. Can be repeated.",
    )
    parser.add_argument(
        "--no-project-brief",
        action="store_true",
        help="Disable project audit Markdown generation.",
    )
    parser.add_argument(
        "--announcement-material",
        action="append",
        default=[],
        help="Version announcement material path. Can be repeated. Supports docx/txt/md/json/csv/tsv/xlsx.",
    )
    parser.add_argument(
        "--announcement-output",
        help="Path for the announcement-specific glossary workbook output.",
    )
    parser.add_argument(
        "--announcement-validation-output",
        help="Path for the announcement validation Markdown report.",
    )
    parser.add_argument(
        "--announcement-min-hit",
        type=int,
        default=1,
        help="Minimum hit count used when matching language-table terms against announcement text. Default: 1",
    )
    parser.add_argument(
        "--ai-supplement",
        action="store_true",
        help="Enable optional AI supplement packet/response flow for announcement glossary lookup.",
    )
    parser.add_argument(
        "--ai-supplement-packet-output",
        help="Path for the compact AI supplement JSON packet. Defaults to *_ai_packet_YYYYMMDD.json.",
    )
    parser.add_argument(
        "--ai-supplement-response",
        help="Path to a structured AI supplement response JSON file to merge into the announcement workbook.",
    )
    parser.add_argument(
        "--ai-supplement-report-output",
        help="Path for the AI supplement sidecar report. Defaults to *_ai_supplement_YYYYMMDD.md.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    configure_utf8_stdio()
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        language_table_specs = parse_language_table_specs(args.language_table)
    except ValueError as exc:
        parser.error(str(exc))
    if args.input_path and language_table_specs:
        parser.error("Use either positional input_path or --language-table, not both.")
    if not args.input_path and not language_table_specs:
        parser.error("input_path or at least one --language-table LANG=path is required.")

    announcement_material_paths = [Path(path) for path in args.announcement_material]
    announcement_output_path = default_announcement_output_path(
        material_paths=announcement_material_paths,
        announcement_output=args.announcement_output,
    )
    announcement_validation_output_path = default_announcement_validation_output_path(
        material_paths=announcement_material_paths,
        announcement_validation_output=args.announcement_validation_output,
    )
    ai_supplement_packet_output_path = default_ai_supplement_packet_output_path(
        material_paths=announcement_material_paths,
        ai_supplement_packet_output=args.ai_supplement_packet_output,
    )
    ai_supplement_report_output_path = default_ai_supplement_report_output_path(
        material_paths=announcement_material_paths,
        ai_supplement_report_output=args.ai_supplement_report_output,
    )
    ai_supplement_response_path = Path(args.ai_supplement_response) if args.ai_supplement_response else None
    if any([args.ai_supplement_packet_output, args.ai_supplement_response, args.ai_supplement_report_output]) and not args.ai_supplement:
        parser.error("--ai-supplement is required when using AI supplement packet, response, or report options.")
    if args.ai_supplement and not announcement_material_paths:
        parser.error("--ai-supplement is only supported with --announcement-material.")
    curated_rules_path = Path(args.curated_rules) if args.curated_rules else None
    observations_store_path = Path(args.observations_store) if args.observations_store else None
    curated_rules = load_curated_rules(curated_rules_path)
    observations_store = load_observation_store(observations_store_path)

    if language_table_specs:
        if not announcement_material_paths:
            parser.error("--language-table is only supported with --announcement-material.")
        if any([args.output, args.final_output, args.project_brief_output, args.translation_prompt_output, args.project_material, args.project_note]):
            parser.error("--language-table cannot be combined with full glossary or project brief outputs.")
        if announcement_output_path is None:
            parser.error("--announcement-output could not be resolved.")
        announcement_text = load_announcement_texts(announcement_material_paths)
        announcement_rows, announcement_stats = build_multilingual_announcement_rows(
            language_table_specs=language_table_specs,
            sheet_name=args.sheet,
            id_column=args.id_column,
            source_column=args.source_column,
            curated_rules=curated_rules,
            announcement_min_hit=args.announcement_min_hit,
            source_only=args.source_only,
            announcement_text=announcement_text,
            include_empty=args.include_empty_final_terms,
        )
        announcement_headers = ["ID", "CN", *[spec.language for spec in language_table_specs]]
        ai_supplement_report: dict[str, object] | None = None
        if args.ai_supplement:
            if ai_supplement_packet_output_path is None or ai_supplement_report_output_path is None:
                parser.error("--ai-supplement output paths could not be resolved.")
            ai_candidate_rows = build_multilingual_ai_candidate_rows(
                language_table_specs=language_table_specs,
                sheet_name=args.sheet,
                id_column=args.id_column,
                source_column=args.source_column,
                curated_rules=curated_rules,
                announcement_min_hit=args.announcement_min_hit,
                source_only=args.source_only,
            )
            announcement_rows, ai_supplement_report, _packet_path, _report_path = run_ai_supplement_flow(
                announcement_rows=announcement_rows,
                announcement_candidate_rows=ai_candidate_rows,
                announcement_text=announcement_text,
                headers=announcement_headers,
                project_name=args.project_name or "",
                packet_output_path=ai_supplement_packet_output_path,
                report_output_path=ai_supplement_report_output_path,
                response_path=ai_supplement_response_path,
            )
        write_announcement_glossary_workbook(
            output_path=announcement_output_path,
            matched_rows=announcement_rows,
            id_header=args.id_column,
            source_header=args.source_column,
            target_header=args.target_column,
            headers=announcement_headers,
        )
        if announcement_validation_output_path is not None:
            write_announcement_validation_report(
                output_path=announcement_validation_output_path,
                announcement_materials=announcement_material_paths,
                language_tables=[f"{spec.language}: {spec.path}" for spec in language_table_specs],
                glossary_output_path=announcement_output_path,
                rows=announcement_rows,
                headers=announcement_headers,
                stats=announcement_stats,
            )

        print("INPUT=multi-language")
        print("DETAIL_OUTPUT=disabled")
        print("FINAL_OUTPUT=disabled")
        print("PROJECT_BRIEF_OUTPUT=disabled")
        print("TRANSLATION_PROMPT_OUTPUT=disabled")
        print(f"ANNOUNCEMENT_OUTPUT={announcement_output_path}")
        print(f"ANNOUNCEMENT_VALIDATION_OUTPUT={announcement_validation_output_path or 'disabled'}")
        print(f"ANNOUNCEMENT_MATERIALS={len(announcement_material_paths)}")
        print(f"ANNOUNCEMENT_TERMS={len(announcement_rows)}")
        print(f"LANGUAGE_TABLES={len(language_table_specs)}")
        print(f"CURATED_RULES={curated_rules_path or 'disabled'}")
        print(f"OBSERVATIONS_STORE={observations_store_path or 'disabled'}")
        print(f"AI_SUPPLEMENT_PACKET_OUTPUT={ai_supplement_packet_output_path if args.ai_supplement else 'disabled'}")
        print(f"AI_SUPPLEMENT_REPORT_OUTPUT={ai_supplement_report_output_path if args.ai_supplement else 'disabled'}")
        if ai_supplement_report and ai_supplement_report.get("project_name_translation_missing"):
            print(f"PROJECT_NAME_TRANSLATION_MISSING={clean_text(args.project_name)}")
        return 0

    input_path = Path(args.input_path)
    detail_output_path, final_output_path = default_output_paths(
        input_path=input_path,
        detail_output=args.output,
        final_output=args.final_output,
    )
    project_name = args.project_name or input_path.stem
    project_brief_output_path = default_project_brief_output_path(
        input_path=input_path,
        project_brief_output=args.project_brief_output,
    )
    translation_prompt_output_path = Path(args.translation_prompt_output) if args.translation_prompt_output else None
    announcement_only = should_run_announcement_only(args)
    digest = file_digest(input_path)

    records, sheet_name = load_records(
        input_path=input_path,
        sheet_name=args.sheet,
        id_column=args.id_column,
        source_column=args.source_column,
        target_column=args.target_column,
        source_only=args.source_only,
    )
    if announcement_only and announcement_output_path is not None:
        announcement_headers, announcement_candidate_rows = build_announcement_candidate_rows_from_workbook(
            input_path=input_path,
            sheet_name=args.sheet,
            id_column=args.id_column,
            source_column=args.source_column,
            target_column=args.target_column,
            min_hit=args.announcement_min_hit,
            curated_rules=curated_rules,
            source_only=args.source_only,
        )
        if not announcement_candidate_rows:
            announcement_records = load_project_records(input_path) or records
            announcement_candidate_rows = build_announcement_candidate_rows(
                records=announcement_records,
                min_hit=args.announcement_min_hit,
                curated_rules=curated_rules,
            )
        announcement_text = load_announcement_texts(announcement_material_paths)
        announcement_rows = select_announcement_term_rows(
            term_rows=announcement_candidate_rows,
            announcement_text=announcement_text,
            include_empty=args.include_empty_final_terms,
        )
        output_headers = announcement_headers or [
            display_header_name(args.id_column, "ID"),
            display_header_name(args.source_column, "CN"),
            display_header_name(args.target_column, "EN"),
        ]
        ai_supplement_report: dict[str, object] | None = None
        if args.ai_supplement:
            if ai_supplement_packet_output_path is None or ai_supplement_report_output_path is None:
                parser.error("--ai-supplement output paths could not be resolved.")
            ai_candidate_rows = build_ai_evidence_candidate_rows_from_workbook(
                input_path=input_path,
                sheet_name=args.sheet,
                id_column=args.id_column,
                source_column=args.source_column,
                target_column=args.target_column,
                language=display_header_name(args.target_column, "EN"),
                source_only=args.source_only,
            ) or announcement_candidate_rows
            announcement_rows, ai_supplement_report, _packet_path, _report_path = run_ai_supplement_flow(
                announcement_rows=announcement_rows,
                announcement_candidate_rows=ai_candidate_rows,
                announcement_text=announcement_text,
                headers=output_headers,
                project_name=args.project_name or "",
                packet_output_path=ai_supplement_packet_output_path,
                report_output_path=ai_supplement_report_output_path,
                response_path=ai_supplement_response_path,
            )
        write_announcement_glossary_workbook(
            output_path=announcement_output_path,
            matched_rows=announcement_rows,
            id_header=args.id_column,
            source_header=args.source_column,
            target_header=args.target_column,
            headers=output_headers,
        )
        if announcement_validation_output_path is not None:
            write_announcement_validation_report(
                output_path=announcement_validation_output_path,
                announcement_materials=announcement_material_paths,
                language_tables=[f"{display_header_name(args.target_column, 'EN')}: {input_path}"],
                glossary_output_path=announcement_output_path,
                rows=announcement_rows,
                headers=output_headers,
                stats={"candidate_terms": len(announcement_candidate_rows)},
            )

        print(f"INPUT={input_path}")
        print("DETAIL_OUTPUT=disabled")
        print("FINAL_OUTPUT=disabled")
        print("PROJECT_BRIEF_OUTPUT=disabled")
        print("TRANSLATION_PROMPT_OUTPUT=disabled")
        print(f"ANNOUNCEMENT_OUTPUT={announcement_output_path}")
        print(f"ANNOUNCEMENT_VALIDATION_OUTPUT={announcement_validation_output_path or 'disabled'}")
        print(f"ANNOUNCEMENT_MATERIALS={len(announcement_material_paths)}")
        print(f"ANNOUNCEMENT_TERMS={len(announcement_rows)}")
        print(f"CURATED_RULES={curated_rules_path or 'disabled'}")
        print(f"OBSERVATIONS_STORE={observations_store_path or 'disabled'}")
        print(f"SHEET={sheet_name}")
        print(f"RECORDS={len(records)}")
        print(f"AI_SUPPLEMENT_PACKET_OUTPUT={ai_supplement_packet_output_path if args.ai_supplement else 'disabled'}")
        print(f"AI_SUPPLEMENT_REPORT_OUTPUT={ai_supplement_report_output_path if args.ai_supplement else 'disabled'}")
        if ai_supplement_report and ai_supplement_report.get("project_name_translation_missing"):
            print(f"PROJECT_NAME_TRANSLATION_MISSING={clean_text(args.project_name)}")
        return 0

    all_rows, glossary_rows, high_risk_rows, manual_rows, final_rows = build_term_rows(
        records=records,
        min_hit=args.min_hit,
        glossary_hit_threshold=args.glossary_hit_threshold,
        curated_rules=curated_rules,
        observations_store=observations_store,
        input_digest=digest,
        include_empty_final_terms=args.include_empty_final_terms,
    )

    write_detail_workbook(
        output_path=detail_output_path,
        sheet_name=sheet_name,
        records=records,
        all_rows=all_rows,
        glossary_rows=glossary_rows,
        high_risk_rows=high_risk_rows,
        manual_rows=manual_rows,
        curated_rules_path=curated_rules_path,
        observations_store_path=observations_store_path,
    )
    write_final_workbook(output_path=final_output_path, final_rows=final_rows)
    material_records, material_sources = load_project_material_records(
        material_paths=[Path(path) for path in args.project_material],
        notes=args.project_note,
    )
    project_records = records if args.no_project_brief and translation_prompt_output_path is None else (
        (load_project_records(input_path) or records) + material_records
    )
    project_brief_markdown, translation_prompt = build_project_brief(
        project_name=project_name,
        sheet_name=sheet_name,
        records=project_records,
        all_rows=all_rows,
        glossary_rows=glossary_rows,
        manual_rows=manual_rows,
        material_sources=material_sources,
    )
    if not args.no_project_brief:
        write_text_output(project_brief_output_path, project_brief_markdown)
    if translation_prompt_output_path is not None:
        write_text_output(translation_prompt_output_path, translation_prompt)

    announcement_rows: list[dict[str, object]] = []
    if announcement_material_paths and announcement_output_path is not None:
        announcement_headers, announcement_candidate_rows = build_announcement_candidate_rows_from_workbook(
            input_path=input_path,
            sheet_name=args.sheet,
            id_column=args.id_column,
            source_column=args.source_column,
            target_column=args.target_column,
            min_hit=args.announcement_min_hit,
            curated_rules=curated_rules,
            source_only=args.source_only,
        )
        if not announcement_candidate_rows:
            announcement_records = load_project_records(input_path) or records
            announcement_candidate_rows = build_announcement_candidate_rows(
                records=announcement_records,
                min_hit=args.announcement_min_hit,
                curated_rules=curated_rules,
            )
        announcement_text = load_announcement_texts(announcement_material_paths)
        announcement_rows = select_announcement_term_rows(
            term_rows=announcement_candidate_rows,
            announcement_text=announcement_text,
            include_empty=args.include_empty_final_terms,
        )
        output_headers = announcement_headers or [
            display_header_name(args.id_column, "ID"),
            display_header_name(args.source_column, "CN"),
            display_header_name(args.target_column, "EN"),
        ]
        ai_supplement_report: dict[str, object] | None = None
        if args.ai_supplement:
            if ai_supplement_packet_output_path is None or ai_supplement_report_output_path is None:
                parser.error("--ai-supplement output paths could not be resolved.")
            ai_candidate_rows = build_ai_evidence_candidate_rows_from_workbook(
                input_path=input_path,
                sheet_name=args.sheet,
                id_column=args.id_column,
                source_column=args.source_column,
                target_column=args.target_column,
                language=display_header_name(args.target_column, "EN"),
                source_only=args.source_only,
            ) or announcement_candidate_rows
            announcement_rows, ai_supplement_report, _packet_path, _report_path = run_ai_supplement_flow(
                announcement_rows=announcement_rows,
                announcement_candidate_rows=ai_candidate_rows,
                announcement_text=announcement_text,
                headers=output_headers,
                project_name=project_name,
                packet_output_path=ai_supplement_packet_output_path,
                report_output_path=ai_supplement_report_output_path,
                response_path=ai_supplement_response_path,
            )
        write_announcement_glossary_workbook(
            output_path=announcement_output_path,
            matched_rows=announcement_rows,
            id_header=args.id_column,
            source_header=args.source_column,
            target_header=args.target_column,
            headers=output_headers,
        )
        if announcement_validation_output_path is not None:
            write_announcement_validation_report(
                output_path=announcement_validation_output_path,
                announcement_materials=announcement_material_paths,
                language_tables=[f"{display_header_name(args.target_column, 'EN')}: {input_path}"],
                glossary_output_path=announcement_output_path,
                rows=announcement_rows,
                headers=output_headers,
                stats={"candidate_terms": len(announcement_candidate_rows)},
            )

    save_curated_rules(curated_rules_path, curated_rules)
    save_observation_store(observations_store_path, observations_store)

    print(f"INPUT={input_path}")
    print(f"DETAIL_OUTPUT={detail_output_path}")
    print(f"FINAL_OUTPUT={final_output_path}")
    print(f"PROJECT_BRIEF_OUTPUT={project_brief_output_path if not args.no_project_brief else 'disabled'}")
    print(f"TRANSLATION_PROMPT_OUTPUT={translation_prompt_output_path or 'disabled'}")
    print(f"ANNOUNCEMENT_OUTPUT={announcement_output_path if announcement_output_path else 'disabled'}")
    print(f"ANNOUNCEMENT_VALIDATION_OUTPUT={announcement_validation_output_path if announcement_validation_output_path else 'disabled'}")
    print(f"ANNOUNCEMENT_MATERIALS={len(announcement_material_paths)}")
    print(f"ANNOUNCEMENT_TERMS={len(announcement_rows)}")
    print(f"PROJECT_MATERIALS={len(material_sources)}")
    print(f"CURATED_RULES={curated_rules_path or 'disabled'}")
    print(f"OBSERVATIONS_STORE={observations_store_path or 'disabled'}")
    print(f"SHEET={sheet_name}")
    print(f"RECORDS={len(records)}")
    print(f"CANDIDATES={len(all_rows)}")
    print(f"GLOSSARY_ROWS={len(glossary_rows)}")
    print(f"HIGH_RISK_ROWS={len(high_risk_rows)}")
    print(f"MANUAL_ADAPTATION_ROWS={len(manual_rows)}")
    print(f"AI_SUPPLEMENT_PACKET_OUTPUT={ai_supplement_packet_output_path if args.ai_supplement else 'disabled'}")
    print(f"AI_SUPPLEMENT_REPORT_OUTPUT={ai_supplement_report_output_path if args.ai_supplement else 'disabled'}")
    if 'ai_supplement_report' in locals() and ai_supplement_report and ai_supplement_report.get("project_name_translation_missing"):
        print(f"PROJECT_NAME_TRANSLATION_MISSING={clean_text(project_name)}")
    print(f"FINAL_ROWS={len(final_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
