from __future__ import annotations

import os
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
from app.config import DEFAULT_SETTINGS, save_settings
from app.languages import language_spec
from app.main import app
from app.workflow.glossary import (
    _LEGACY_EN_ALT_HEADERS,
    _normalize_new_english_glossary_artifact,
    _normalize_new_english_glossary_text_artifact,
)
from app.workflow.prompt_snapshots import create_project_glossary_snapshot
from conftest import reset_data_root


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    reset_data_root(Path(os.environ["LWS_DATA_ROOT"]))
    db.init_db()
    save_settings(DEFAULT_SETTINGS)


def test_new_english_glossary_snapshot_contains_only_en_target_column(tmp_path: Path) -> None:
    project = db.insert_project("English snapshot")
    run = db.insert_run(project["id"], "translation", "en")
    db.insert_glossary_term(
        project["id"],
        {
            "term_key": "T-1",
            "source": "战机",
            "target": "Warplane",
            "target_alt": "Fighter",
            "language": "en",
            "confirmed": True,
        },
    )

    artifact = create_project_glossary_snapshot(project["id"], run["id"], tmp_path, language="en")

    workbook = load_workbook(Path(artifact["path"]), read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows[0][:3] == ("ID", "CN", "EN")
    assert "EN2" not in rows[0]
    assert "Fighter" not in rows[1]


def test_english_language_spec_separates_new_output_from_legacy_import_aliases() -> None:
    spec = language_spec("en")

    assert spec.alt_header == ""
    assert "en2" in spec.alt_aliases
    assert "target_alt" in spec.alt_aliases


def test_new_english_glossary_downloads_contain_no_en2(tmp_path: Path) -> None:
    source_path = tmp_path / "language-table.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Language"
    worksheet.append(["ID", "CN", "EN"])
    for index in range(1, 11):
        worksheet.append([f"T-{index}", "\u6218\u673a", "Warplane"])
    workbook.save(source_path)
    workbook.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "English glossary"}).json()
        with source_path.open("rb") as source_file:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": (source_path.name, source_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        response = client.post(
            f"/api/projects/{project['id']}/glossary/extract",
            json={
                "input_artifact_id": source_artifact["id"],
                "id_column": "ID",
                "source_column": "CN",
                "target_column": "EN",
                "language": "en",
                "ai_candidate_supplement": False,
            },
        )

    assert response.status_code == 200, response.text
    artifacts = {artifact["kind"]: artifact for artifact in response.json()["artifacts"]}
    for kind in ("glossary_detail", "glossary_final"):
        artifact_path = Path(artifacts[kind]["path"])
        assert "EN2" not in artifact_path.name.upper()
        generated = load_workbook(artifact_path, read_only=True, data_only=True)
        try:
            for sheet in generated.worksheets:
                headers = [str(cell.value or "").strip().casefold() for cell in sheet[1]]
                assert not set(headers) & _LEGACY_EN_ALT_HEADERS, f"{kind}:{sheet.title}"
            values = [cell for sheet in generated.worksheets for row in sheet.iter_rows(values_only=True) for cell in row]
        finally:
            generated.close()
        assert not any("EN2" in str(value) for value in values if value is not None), kind
    for kind in ("project_brief", "translation_prompt"):
        text = Path(artifacts[kind]["path"]).read_text(encoding="utf-8")
        en2_lines = [line for line in text.splitlines() if "en2" in line.casefold()]
        assert en2_lines == [], f"{kind}: {en2_lines}"


def test_english_glossary_normalization_does_not_rewrite_term_values(tmp_path: Path) -> None:
    path = tmp_path / "generated.xlsx"
    workbook = Workbook()
    glossary = workbook.active
    glossary.title = "Glossary"
    glossary.append(["ID", "CN", "EN", "EN2"])
    glossary.append(["T-1", "保留正文 EN2 标识", "Keep EN2 marker", "Legacy alternate"])
    notes = workbook.create_sheet("Notes")
    notes.append(["Item", "Value"])
    notes.append(["Columns", "ID = text id, CN = source term, EN = example English, EN2 = manual adaptation English"])
    workbook.save(path)
    workbook.close()

    _normalize_new_english_glossary_artifact(path)

    normalized = load_workbook(path, read_only=True, data_only=True)
    try:
        glossary_rows = list(normalized["Glossary"].iter_rows(values_only=True))
        notes_rows = list(normalized["Notes"].iter_rows(values_only=True))
    finally:
        normalized.close()
    assert glossary_rows[0] == ("ID", "CN", "EN")
    assert glossary_rows[1] == ("T-1", "保留正文 EN2 标识", "Keep EN2 marker")
    assert not any("EN2" in str(value) for row in notes_rows for value in row if value is not None)


def test_english_glossary_text_normalization_preserves_non_rule_content(tmp_path: Path) -> None:
    path = tmp_path / "translation_prompt.txt"
    path.write_text(
        "Term example: Keep EN2 marker\n"
        "关键术语以随附术语表为准，EN 为标准译法，EN2 为项目中稳定出现的手动适配译法。\n",
        encoding="utf-8",
    )

    _normalize_new_english_glossary_text_artifact(path)

    normalized = path.read_text(encoding="utf-8")
    assert "Term example: Keep EN2 marker" in normalized
    assert "关键术语以随附术语表为准，EN 为唯一主译。" in normalized


@pytest.mark.parametrize("alternate_header", ["EN2", "target_alt"])
def test_legacy_english_alternate_headers_remain_importable(tmp_path: Path, alternate_header: str) -> None:
    path = tmp_path / f"legacy-{alternate_header}.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Glossary"
    worksheet.append(["ID", "CN", "EN", alternate_header])
    worksheet.append(["T-1", "战机", "Warplane", "Fighter"])
    workbook.save(path)
    workbook.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Legacy glossary"}).json()
        with path.open("rb") as source_file:
            artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": (path.name, source_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        response = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": artifact["id"], "language": "en"},
        )

    assert response.status_code == 200, response.text
    assert response.json()["rows"][0]["target"] == "Warplane"
    assert response.json()["rows"][0]["target_alt"] == ""
