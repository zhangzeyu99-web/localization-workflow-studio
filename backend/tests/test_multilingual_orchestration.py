from __future__ import annotations

import json
import os
import tempfile
import threading
import time
from pathlib import Path

os.environ.setdefault("LWS_DATA_ROOT", str(Path(tempfile.gettempdir()) / "lws-test-data"))
os.environ["LWS_ENABLE_TEST_PROVIDER"] = "1"

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
import app.background_jobs as background_jobs
import app.job_queue as job_queue
import app.jobs as jobs
import app.workflow.multilingual as multilingual
from app.config import DEFAULT_SETTINGS, save_settings
from app.main import app
from conftest import reset_data_root, wait_for_background_jobs


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    reset_data_root(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    wait_for_background_jobs()
    save_settings(DEFAULT_SETTINGS)


def _language_table(path: Path, headers: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", *headers])
    ws.append([1, "开始游戏", *("" for _ in headers)])
    ws.append([2, "领取奖励", *("" for _ in headers)])
    wb.save(path)
    wb.close()


def _add_language_table(project_id: str, path: Path, headers: list[str]) -> dict:
    _language_table(path, headers)
    return db.add_artifact(project_id, "source", path, "language_table", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def test_multilingual_status_maps_existing_child_runs(tmp_path: Path) -> None:
    project = db.insert_project("multi status", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN", "KR"])
    en_run = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={"input_artifact_id": artifact["id"], "task_origin": "translation_run"},
    )

    with TestClient(app) as client:
        response = client.get(
            f"/api/projects/{project['id']}/multilingual/status",
            params={"input_artifact_id": artifact["id"], "languages": "en,ko"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert [item["language"] for item in payload["languages"]] == ["en", "ko"]
    assert payload["languages"][0]["translation_run_id"] == en_run["id"]
    assert payload["languages"][1]["translation_run_id"] is None


def test_multilingual_status_uses_vn_canonical_and_ui_code(tmp_path: Path) -> None:
    project = db.insert_project("Vietnamese status", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source-vn.xlsx", ["EN", "VI"])
    run = db.insert_run(
        project["id"],
        "translation",
        "vi",
        metadata={"input_artifact_id": artifact["id"], "task_origin": "translation_run"},
    )

    with TestClient(app) as client:
        response = client.get(
            f"/api/projects/{project['id']}/multilingual/status",
            params={"input_artifact_id": artifact["id"], "languages": "vi"},
        )

    assert response.status_code == 200
    item = response.json()["languages"][0]
    assert item["language"] == "vn"
    assert item["visible_language"] == "VN"
    assert item["translation_run_id"] == run["id"]


def test_start_multilingual_translation_creates_missing_child_runs(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    project = db.insert_project("multi translate", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN", "KR"])

    def fake_translate(run_id: str, request: object, cancel_event: object | None = None) -> dict:
        _ = request, cancel_event
        run = db.get_run(run_id)
        db.update_run(run_id, status="passed", metadata={**run.get("metadata", {}), "fake_completed": True})
        return {"run": db.get_run(run_id), "artifacts": []}

    monkeypatch.setattr(multilingual, "run_translate_sync", fake_translate)
    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/multilingual/translate/start",
            json={"input_artifact_id": artifact["id"], "languages": ["en", "ko"], "batch_size": 10, "task_code": "T"},
        )
        wait_for_background_jobs()
        status = client.get(
            f"/api/projects/{project['id']}/multilingual/status",
            params={"input_artifact_id": artifact["id"], "languages": "en,ko"},
        ).json()

    assert response.status_code == 200
    assert {item["language"] for item in response.json()["languages"]} == {"en", "ko"}
    assert all(item["translation_run_id"] for item in status["languages"])
    assert all(item["status"] == "passed" for item in status["languages"])


@pytest.mark.parametrize("queue_kind", ["translation", "qa"])
def test_multilingual_worker_error_cannot_overwrite_terminal_task_state(
    tmp_path: Path,
    queue_kind: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project(f"multi terminal {queue_kind}", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / f"terminal-{queue_kind}.xlsx", ["EN"])
    task_id = f"task-multi-terminal-{queue_kind}"
    captured: dict[str, object] = {}

    def capture_start(
        job_kind: str,
        project_id: str,
        input_artifact_id: str,
        request: multilingual.MultilingualQueueRequest,
        child_run_ids: list[str],
    ) -> dict:
        suffix = "translate" if job_kind == "multilingual_translate" else "qa"
        captured["record"] = {
            "job_id": multilingual._queue_job_id(suffix, project_id, input_artifact_id, request.translation_task_id),
            "project_id": project_id,
            "payload": {
                "request": request.model_dump(exclude_none=True),
                "child_run_ids": child_run_ids,
            },
        }
        return {"status": "staged"}

    monkeypatch.setattr(multilingual.background_jobs, "start_multilingual", capture_start)
    if queue_kind == "qa":
        monkeypatch.setattr(multilingual, "_qa_input_artifact", lambda *_args, **_kwargs: artifact)
        started = multilingual.start_multilingual_qa_queue(
            project["id"],
            multilingual.MultilingualQueueRequest(
                input_artifact_id=artifact["id"],
                languages=["en"],
                translation_task_id=task_id,
            ),
        )
        monkeypatch.setattr(multilingual, "run_qa_sync", lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("late QA")))
        execute = multilingual.execute_multilingual_qa_job
    else:
        started = multilingual.start_multilingual_translation_queue(
            project["id"],
            multilingual.MultilingualQueueRequest(
                input_artifact_id=artifact["id"],
                languages=["en"],
                translation_task_id=task_id,
            ),
        )
        monkeypatch.setattr(multilingual, "run_translate_sync", lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("late translation")))
        execute = multilingual.execute_multilingual_translation_job

    run_id = started["created_run_ids"][0]
    from app.workflow.translation_tasks import mark_translation_task_state

    mark_translation_task_state(project["id"], task_id, "delivered")
    db.update_run(run_id, status="running")
    execute(captured["record"], threading.Event())  # type: ignore[arg-type]

    refreshed = db.get_run(run_id)
    assert refreshed["status"] == "canceled"
    assert refreshed["metadata"]["translation_task_state"] == "delivered"


def test_new_translation_task_does_not_reuse_same_source_child_run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project("task isolated queue", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN"])
    old_run = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={
            "input_artifact_id": artifact["id"],
            "task_origin": "translation_run",
            "translation_task_id": "task-old",
        },
    )
    db.update_run(old_run["id"], status="passed")

    def fake_translate(run_id: str, request: object, cancel_event: object | None = None) -> dict:
        _ = request, cancel_event
        db.update_run(run_id, status="passed")
        return {"run": db.get_run(run_id), "artifacts": []}

    monkeypatch.setattr(multilingual, "run_translate_sync", fake_translate)
    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/multilingual/translate/start",
            json={
                "input_artifact_id": artifact["id"],
                "languages": ["en"],
                "translation_task_id": "task-new",
            },
        )
        wait_for_background_jobs()
        status = client.get(
            f"/api/projects/{project['id']}/multilingual/status",
            params={
                "input_artifact_id": artifact["id"],
                "languages": "en",
                "translation_task_id": "task-new",
            },
        ).json()

    assert response.status_code == 200
    created_ids = response.json()["created_run_ids"]
    assert len(created_ids) == 1
    assert created_ids[0] != old_run["id"]
    assert db.get_run(created_ids[0])["metadata"]["translation_task_id"] == "task-new"
    assert status["translation_task_id"] == "task-new"
    assert status["languages"][0]["translation_run_id"] == created_ids[0]


def test_abandon_translation_task_marks_only_matching_runs(tmp_path: Path) -> None:
    project = db.insert_project("task abandon", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN", "KR"])
    task_runs = [
        db.insert_run(
            project["id"],
            kind,
            language,
            metadata={"input_artifact_id": artifact["id"], "translation_task_id": "task-a"},
        )
        for kind, language in [("translation", "en"), ("qa", "ko")]
    ]
    other_run = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={"input_artifact_id": artifact["id"], "translation_task_id": "task-b"},
    )
    for run in [*task_runs, other_run]:
        db.update_run(run["id"], status="needs_input")

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/translation-tasks/task-a/abandon",
        )

    assert response.status_code == 200
    assert set(response.json()["updated_run_ids"]) == {run["id"] for run in task_runs}
    for run in task_runs:
        assert db.get_run(run["id"])["metadata"]["translation_task_state"] == "abandoned"
    assert "translation_task_state" not in db.get_run(other_run["id"])["metadata"]


def test_translation_task_continuation_metadata_preserves_original_source_scope() -> None:
    from app.workflow.translation_tasks import translation_task_continuation_metadata

    metadata = translation_task_continuation_metadata({
        "metadata": {
            "input_artifact_id": "fixed-input",
            "parent_input_artifact_id": "source-root",
            "multilingual_source_artifact_id": "source-root",
            "translation_task_id": "task-a",
        },
    })

    assert metadata == {
        "parent_input_artifact_id": "source-root",
        "multilingual_source_artifact_id": "source-root",
        "translation_task_id": "task-a",
    }


def test_multilingual_translation_uses_persistent_controller_and_duplicate_start_is_idempotent(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project("multi persistent controller", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source-controller.xlsx", ["EN", "KR"])
    release = threading.Event()
    started = threading.Event()

    def fake_translate(run_id: str, request: object, cancel_event: object | None = None) -> dict:
        _ = request, cancel_event
        db.update_run(run_id, status="running")
        started.set()
        release.wait(2.0)
        db.update_run(run_id, status="passed")
        return {"run": db.get_run(run_id), "artifacts": []}

    monkeypatch.setattr(multilingual, "run_translate_sync", fake_translate)
    with TestClient(app) as client:
        first = client.post(
            f"/api/projects/{project['id']}/multilingual/translate/start",
            json={"input_artifact_id": artifact["id"], "languages": ["en", "ko"], "batch_size": 10},
            headers={"X-Operator": "Alice"},
        )
        assert first.status_code == 200, first.text
        assert started.wait(2.0)
        duplicate = client.post(
            f"/api/projects/{project['id']}/multilingual/translate/start",
            json={"input_artifact_id": artifact["id"], "languages": ["en", "ko"], "batch_size": 10},
            headers={"X-Operator": "Bob"},
        )
        assert duplicate.status_code == 200, duplicate.text

        job_id = f"multilingual:translate:{project['id']}:{artifact['id']}"
        controller = job_queue.get_job(job_id)
        assert controller is not None
        assert controller["job_kind"] == "multilingual_translate"
        assert controller["lane"] == "language_table"
        assert controller["operator_name"] == "Alice"
        assert len(job_queue.list_jobs()) == 1
        assert duplicate.json()["active_job_id"] == job_id

        release.set()
        wait_for_background_jobs()


def test_multilingual_translation_executes_only_the_persisted_child_run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project("multi persisted child", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source-persisted-child.xlsx", ["EN"])
    formal = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={"input_artifact_id": artifact["id"], "task_origin": "translation_run"},
    )
    quick = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={"input_artifact_id": artifact["id"], "task_origin": "quick_task"},
    )
    executed: list[str] = []

    def fake_translate(run_id: str, request: object, cancel_event: object | None = None) -> dict:
        _ = request, cancel_event
        executed.append(run_id)
        db.update_run(run_id, status="passed")
        return {"run": db.get_run(run_id), "artifacts": []}

    monkeypatch.setattr(multilingual, "run_translate_sync", fake_translate)
    multilingual.execute_multilingual_translation_job(
        {
            "job_id": "multilingual:translate:persisted-child",
            "project_id": project["id"],
            "payload": {
                "request": {"input_artifact_id": artifact["id"], "languages": ["en"]},
                "child_run_ids": [formal["id"]],
            },
        },
        threading.Event(),
    )

    assert executed == [formal["id"]]
    assert db.get_run(quick["id"])["status"] == "created"


def test_concurrent_multilingual_starts_create_only_one_child_set(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project("multi concurrent start", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source-concurrent.xlsx", ["EN"])
    first_insert_entered = threading.Event()
    second_insert_entered = threading.Event()
    release_first_insert = threading.Event()
    insert_lock = threading.Lock()
    insert_count = 0
    original_insert_run = db.insert_run

    def controlled_insert_run(*args: object, **kwargs: object) -> dict:
        nonlocal insert_count
        with insert_lock:
            insert_count += 1
            call_number = insert_count
        if call_number == 1:
            first_insert_entered.set()
            release_first_insert.wait(2.0)
        elif call_number == 2:
            second_insert_entered.set()
        return original_insert_run(*args, **kwargs)

    monkeypatch.setattr(db, "insert_run", controlled_insert_run)
    monkeypatch.setattr(job_queue, "dispatch_lane", lambda lane: False)
    results: list[dict] = []
    errors: list[BaseException] = []

    def start() -> None:
        try:
            results.append(
                multilingual.start_multilingual_translation_queue(
                    project["id"],
                    multilingual.MultilingualQueueRequest(
                        input_artifact_id=artifact["id"],
                        languages=["en"],
                    ),
                )
            )
        except BaseException as exc:
            errors.append(exc)

    first = threading.Thread(target=start)
    second = threading.Thread(target=start)
    first.start()
    assert first_insert_entered.wait(2.0)
    second.start()
    inserted_concurrently = second_insert_entered.wait(0.5)
    release_first_insert.set()
    first.join(2.0)
    second.join(2.0)

    assert inserted_concurrently is False
    assert errors == []
    assert len(results) == 2
    runs = db.list_runs(project["id"])
    assert len(runs) == 1
    controller = job_queue.get_job(f"multilingual:translate:{project['id']}:{artifact['id']}")
    assert controller is not None
    assert controller["payload"]["child_run_ids"] == [runs[0]["id"]]


def test_multilingual_enqueue_failure_keeps_new_children_created(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project("multi enqueue failure", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source-enqueue-failure.xlsx", ["EN", "KR"])
    monkeypatch.setattr(
        background_jobs.job_queue,
        "enqueue_job",
        lambda **kwargs: (_ for _ in ()).throw(RuntimeError("queue write failed")),
    )

    with pytest.raises(RuntimeError, match="queue write failed"):
        multilingual.start_multilingual_translation_queue(
            project["id"],
            multilingual.MultilingualQueueRequest(
                input_artifact_id=artifact["id"],
                languages=["en", "ko"],
                batch_size=10,
            ),
        )

    assert {run["status"] for run in db.list_runs(project["id"])} == {"created"}


def test_multilingual_start_ignores_legacy_project_lease(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    project = db.insert_project("multi busy", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN", "KR"])
    lease_name = jobs.lease_name_for_project(project["id"])

    def fake_translate(run_id: str, request: object, cancel_event: object | None = None) -> dict:
        _ = request, cancel_event
        db.update_run(run_id, status="passed")
        return {"run": db.get_run(run_id), "artifacts": []}

    monkeypatch.setattr(multilingual, "run_translate_sync", fake_translate)

    with TestClient(app) as client:
        assert db.acquire_job_lease(lease_name, "run:existing", operator_name="Alice")
        try:
            response = client.post(
                f"/api/projects/{project['id']}/multilingual/translate/start",
                json={"input_artifact_id": artifact["id"], "languages": ["en", "ko"], "batch_size": 10, "task_code": "T"},
            )

            assert response.status_code == 200, response.text
            assert len(db.list_runs(project["id"])) == 2
            assert job_queue.get_job(f"multilingual:translate:{project['id']}:{artifact['id']}") is not None
            wait_for_background_jobs()
        finally:
            db.release_job_lease(lease_name, "run:existing")


def test_multilingual_controllers_queue_fifo_without_legacy_capacity_rejection(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    first_project = db.insert_project("multi capacity first", "QA", "")
    second_project = db.insert_project("multi capacity second", "QA", "")
    first_artifact = _add_language_table(first_project["id"], tmp_path / "source-first.xlsx", ["EN"])
    second_artifact = _add_language_table(second_project["id"], tmp_path / "source-second.xlsx", ["EN"])
    release = threading.Event()
    started = threading.Event()

    def fake_translate(run_id: str, request: object, cancel_event: object | None = None) -> dict:
        _ = request, cancel_event
        db.update_run(run_id, status="running")
        started.set()
        release.wait(2.0)
        db.update_run(run_id, status="passed")
        return {"run": db.get_run(run_id), "artifacts": []}

    monkeypatch.setattr(multilingual, "run_translate_sync", fake_translate)
    with TestClient(app) as client:
        first = client.post(
            f"/api/projects/{first_project['id']}/multilingual/translate/start",
            json={"input_artifact_id": first_artifact["id"], "languages": ["en"]},
        )
        assert first.status_code == 200, first.text
        assert started.wait(2.0)
        second = client.post(
            f"/api/projects/{second_project['id']}/multilingual/translate/start",
            json={"input_artifact_id": second_artifact["id"], "languages": ["en"]},
        )
        assert second.status_code == 200, second.text
        second_job = job_queue.get_job(f"multilingual:translate:{second_project['id']}:{second_artifact['id']}")
        assert second_job is not None
        assert second_job["status"] == "queued"
        release.set()
        wait_for_background_jobs()


def test_italian_translation_writes_it_column_in_multilingual_workbook(tmp_path: Path) -> None:
    project = db.insert_project("Italian target column", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN", "IT"])
    save_settings({**DEFAULT_SETTINGS, "provider": "test-fake", "model": "test-fake-localization", "batch_size": 10})

    with TestClient(app) as client:
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "it", "input_artifact_id": artifact["id"]},
        ).json()
        response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake", "batch_size": 10})

    assert response.status_code == 200, response.text
    assert response.json()["run"]["status"] == "passed"
    final_artifact = next(item for item in db.list_artifacts(run_id=run["id"]) if item["kind"] == "qa_final_workbook")
    workbook = load_workbook(final_artifact["path"], read_only=True, data_only=False)
    try:
        sheet = workbook.active
        assert sheet.cell(2, 3).value in {None, ""}
        assert str(sheet.cell(2, 4).value).startswith("TestFake")
    finally:
        workbook.close()


def test_vietnamese_translation_keeps_vn_in_studio_and_vi_in_workflow(tmp_path: Path) -> None:
    project = db.insert_project("Vietnamese target column", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source-vn.xlsx", ["EN", "VI"])
    save_settings({**DEFAULT_SETTINGS, "provider": "test-fake", "model": "test-fake-localization", "batch_size": 10})
    translation_task_id = "task-vietnamese"

    with TestClient(app) as client:
        started = client.post(
            f"/api/projects/{project['id']}/multilingual/translate/start",
            json={
                "input_artifact_id": artifact["id"],
                "languages": ["en", "vi"],
                "batch_size": 10,
                "task_code": "VN",
                "translation_task_id": translation_task_id,
            },
        )
        assert started.status_code == 200, started.text
        assert {item["language"] for item in started.json()["languages"]} == {"en", "vn"}
        deadline = time.monotonic() + 60
        while True:
            status_response = client.get(
                f"/api/projects/{project['id']}/multilingual/status",
                params={
                    "input_artifact_id": artifact["id"],
                    "languages": "en,vi",
                    "translation_task_id": translation_task_id,
                },
            )
            assert status_response.status_code == 200, status_response.text
            status = status_response.json()
            if status["overall_status"] not in {"pending", "running"} and not status["active_job_id"]:
                break
            if time.monotonic() >= deadline:
                pytest.fail(f"multilingual queue did not finish: {json.dumps(status, ensure_ascii=False)}")
            time.sleep(0.05)
        delivered = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={
                "input_artifact_id": artifact["id"],
                "languages": ["en", "vi"],
                "translation_task_id": translation_task_id,
            },
        )

    assert status["overall_status"] == "passed"
    vn_status = next(item for item in status["languages"] if item["language"] == "vn")
    assert vn_status["visible_language"] == "VN"
    run = db.get_run(vn_status["translation_run_id"])
    assert run["language"] == "vn"
    assert run["metadata"]["translation_task_id"] == translation_task_id
    artifacts = db.list_artifacts(run_id=run["id"])
    translation_manifest = next(item for item in artifacts if item["kind"] == "translation_manifest")
    assert json.loads(Path(translation_manifest["path"]).read_text(encoding="utf-8"))["language"] == "vi"
    final_artifact = next(item for item in artifacts if item["kind"] == "qa_final_workbook")
    assert Path(final_artifact["path"]).name == "result_vi.xlsx"
    assert final_artifact["metadata"]["language"] == "vn"
    workbook = load_workbook(final_artifact["path"], read_only=True, data_only=False)
    try:
        sheet = workbook.active
        assert sheet.cell(2, 3).value in {None, ""}
        assert str(sheet.cell(2, 4).value).startswith("TestFake")
    finally:
        workbook.close()
    assert delivered.status_code == 200, delivered.text
    merged_file = next(item for item in delivered.json()["files"] if item["kind"] == "merged_final")
    merged_workbook = load_workbook(merged_file["path"], read_only=True, data_only=False)
    try:
        sheet = merged_workbook.active
        assert [cell.value for cell in sheet[1]][:4] == ["ID", "CN", "EN", "VI"]
        assert str(sheet.cell(2, 3).value).startswith("TestFake")
        assert str(sheet.cell(2, 4).value).startswith("TestFake")
    finally:
        merged_workbook.close()
    assert delivered.json()["merged_languages"] == ["EN", "VI"]
    archived = db.list_translation_entries(project["id"], language="vi")
    assert archived
    assert {item["language"] for item in archived} == {"vn"}


def test_multilingual_status_rejects_cross_project_artifact(tmp_path: Path) -> None:
    first = db.insert_project("A", "QA", "")
    second = db.insert_project("B", "QA", "")
    artifact = _add_language_table(first["id"], tmp_path / "source.xlsx", ["EN"])

    with TestClient(app) as client:
        response = client.get(
            f"/api/projects/{second['id']}/multilingual/status",
            params={"input_artifact_id": artifact["id"], "languages": "en"},
        )

    assert response.status_code == 400



def test_multilingual_status_exposes_large_text_metadata(tmp_path: Path) -> None:
    project = db.insert_project("multi large text", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN"])
    run = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={
            "input_artifact_id": artifact["id"],
            "task_origin": "translation_run",
            "large_text": {
                "mode": "auto",
                "preflight": {"large_pack": True, "unique_items": 6001, "estimated_target_cells": 6001},
                "cache_lint": {"status": "passed", "hard_blockers": 0},
            },
        },
    )

    with TestClient(app) as client:
        response = client.get(
            f"/api/projects/{project['id']}/multilingual/status",
            params={"input_artifact_id": artifact["id"], "languages": "en"},
        )

    assert response.status_code == 200
    item = response.json()["languages"][0]
    assert item["translation_run_id"] == run["id"]
    assert item["large_text"]["preflight"]["large_pack"] is True
    assert item["large_text"]["cache_lint"]["status"] == "passed"


def test_multilingual_qa_skips_languages_without_translated_input(tmp_path: Path) -> None:
    project = db.insert_project("multi qa skip", "QA", "")
    artifact = _add_language_table(project["id"], tmp_path / "source.xlsx", ["EN", "KR"])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/multilingual/qa/start",
            json={"input_artifact_id": artifact["id"], "languages": ["en", "ko"]},
        )

    assert response.status_code == 200
    result = response.json()
    assert result["created_run_ids"] == []
    assert {item["status"] for item in result["languages"]} == {"pending"}
