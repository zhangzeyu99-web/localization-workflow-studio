"""P2: product-side AI line proofreading with deterministic audit gate."""
from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

os.environ["LWS_DATA_ROOT"] = str(Path(tempfile.gettempdir()) / "lws-test-data")

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
import app.workflow.line_proofread as line_proofread
from app.config import DEFAULT_SETTINGS, save_settings
from app.main import app
from app.workflow.line_proofread import audit_suggestion
from conftest import reset_data_root, wait_for_background_jobs


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    reset_data_root(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    wait_for_background_jobs()
    save_settings(DEFAULT_SETTINGS)


def _row(**overrides) -> dict:
    base = {
        "sheet": "Language",
        "row": 2,
        "record_id": "1",
        "source": "欢迎回来，{playerName}",
        "current_target": "Welcome back, {playerName}",
        "term_hits": [],
    }
    base.update(overrides)
    return base


def test_audit_rejects_placeholder_breaking_suggestion() -> None:
    suggestion = {"suggested_target": "Welcome back, player"}
    assert audit_suggestion(suggestion, _row()) == "protected_token_lost"


def test_audit_rejects_term_drift() -> None:
    row = _row(
        source="加入联盟",
        current_target="Join the Alliance",
        term_hits=[{"source": "联盟", "target": "Alliance"}],
    )
    suggestion = {"suggested_target": "Join the Guild"}
    assert audit_suggestion(suggestion, row) == "term_drift"


def test_audit_rejects_number_loss() -> None:
    row = _row(source="获得 300 金币", current_target="Gain 300 gold")
    suggestion = {"suggested_target": "Gain gold"}
    assert audit_suggestion(suggestion, row) == "number_lost"


def test_audit_rejects_noop_and_empty() -> None:
    assert audit_suggestion({"suggested_target": ""}, _row()) == "empty_suggestion"
    assert audit_suggestion({"suggested_target": "Welcome back, {playerName}"}, _row()) == "no_change"


def test_audit_accepts_safe_improvement() -> None:
    row = _row(source="获得 300 金币", current_target="Gain 300 golds")
    suggestion = {"suggested_target": "Gain 300 gold"}
    assert audit_suggestion(suggestion, row) is None


def _sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "领取奖励", ""])
    ws.append([2, "开始游戏", ""])
    ws.append([3, "欢迎回来，{playerName}", ""])
    wb.save(path)
    wb.close()


def _start_translation_run(client: TestClient, tmp_path: Path, project_name: str, *, enable_line_proofread: bool):
    workbook = tmp_path / "source.xlsx"
    _sample_workbook(workbook)
    project = client.post("/api/projects", json={"name": project_name}).json()
    with workbook.open("rb") as fh:
        upload = client.post(
            f"/api/projects/{project['id']}/files?kind=language_table",
            files={"file": ("source.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload.status_code == 200
    run = client.post(
        "/api/runs",
        json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": upload.json()["id"]},
    ).json()
    response = client.post(
        f"/api/runs/{run['id']}/translate",
        json={"provider": "test-fake", "enable_line_proofread": enable_line_proofread},
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_translation_without_switch_skips_line_proofread(tmp_path: Path) -> None:
    with TestClient(app) as client:
        result = _start_translation_run(client, tmp_path, "无逐行校对项目", enable_line_proofread=False)
        assert result["run"]["status"] == "passed"
        assert "line_proofread" not in result["run"]["metadata"]


def test_translation_with_switch_runs_line_proofread_noop(tmp_path: Path) -> None:
    """test-fake reviewer suggests nothing: state recorded, no rerun, run passes."""
    with TestClient(app) as client:
        result = _start_translation_run(client, tmp_path, "逐行校对空建议项目", enable_line_proofread=True)
        assert result["run"]["status"] == "passed"
        state = result["run"]["metadata"]["line_proofread"]
        assert state["status"] == "model_reviewed"
        assert state["reviewed_rows"] == 3
        assert state["suggested"] == 0
        assert state["applied"] == 0
        assert state["suggestions_artifact_id"]
        assert "fixes" not in state


def test_translation_with_switch_applies_audited_fixes(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """Model suggests one safe fix and one placeholder-breaking fix; only the safe one lands."""

    def fake_provider(settings, prompt):
        payload = json.loads(prompt[prompt.index("待审校行：") + len("待审校行："):])
        suggestions = []
        for row in payload:
            if row["source"] == "领取奖励":
                suggestions.append(
                    {
                        "record_id": row["record_id"],
                        "sheet": row["sheet"],
                        "row": row["row"],
                        "severity": "soft",
                        "suggested_target": "Claim Your Rewards",
                        "reason": "更自然的按钮文案",
                    }
                )
            if "playerName" in row["source"]:
                suggestions.append(
                    {
                        "record_id": row["record_id"],
                        "sheet": row["sheet"],
                        "row": row["row"],
                        "severity": "hard",
                        "suggested_target": "Welcome back, player",
                        "reason": "drops placeholder (should be rejected)",
                    }
                )
        return json.dumps({"suggestions": suggestions})

    monkeypatch.setattr(line_proofread, "_call_line_proofread_provider", fake_provider)
    with TestClient(app) as client:
        result = _start_translation_run(client, tmp_path, "逐行校对采纳项目", enable_line_proofread=True)
        assert result["run"]["status"] == "passed"
        state = result["run"]["metadata"]["line_proofread"]
        assert state["reviewed_rows"] == 3
        assert state["suggested"] == 2
        assert state["rejected_by_audit"] == 1
        assert state["applied"] == 1

        final_artifact = next(a for a in result["artifacts"] if a["kind"] == "qa_final_workbook")
        wb = load_workbook(final_artifact["path"], read_only=True, data_only=True)
        try:
            ws = wb["Language"]
            targets = {str(ws.cell(row, 1).value): str(ws.cell(row, 3).value or "") for row in range(2, 5)}
        finally:
            wb.close()
        assert targets["1"] == "Claim Your Rewards"
        assert "{playerName}" in targets["3"]

        suggestions_artifact = db.get_artifact(state["suggestions_artifact_id"])
        records = [json.loads(line) for line in Path(suggestions_artifact["path"]).read_text(encoding="utf-8").splitlines() if line.strip()]
        audits = {record["audit"] for record in records}
        assert "accepted" in audits
        assert "rejected:protected_token_lost" in audits

        events = client.get(f"/api/runs/{result['run']['id']}/events").json()
        assert any("line proofread applied fixes; re-running machine QA" in event["message"] for event in events)
