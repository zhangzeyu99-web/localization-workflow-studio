from __future__ import annotations

import json
import os
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
from app.main import app
from app.workflow.asset_import_export import _parse_multilingual_translation_table, archive_translation_artifact
from app.workflow.glossary_backfill import backfill_project_glossary_from_final
from app.workflow.reference_lookup import lookup_terms, lookup_translation_entries
from conftest import reset_data_root


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    reset_data_root(Path(os.environ["LWS_DATA_ROOT"]))
    db.init_db()


def _create_project(client: TestClient, name: str) -> dict:
    response = client.post("/api/projects", json={"name": name})
    assert response.status_code == 200, response.text
    return response.json()


def _upload_bytes(
    client: TestClient,
    project_id: str,
    filename: str,
    content: bytes,
    kind: str,
    mime: str = "application/octet-stream",
) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/files?kind={kind}",
        files={"file": (filename, content, mime)},
    )
    assert response.status_code == 200, response.text
    return response.json()


def _workbook_bytes(sheets: list[tuple[str, list[list[object]]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        worksheet = workbook.create_sheet(title)
        for row in rows:
            worksheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _translation_rows(response_payload: dict) -> dict[tuple[str, str], dict]:
    return {
        (str(row.get("source") or ""), str(row.get("language") or "")): row
        for row in response_payload["entries"]
    }


def _glossary_rows(response_payload: dict) -> dict[tuple[str, str], dict]:
    return {
        (str(row.get("source") or ""), str(row.get("language") or "")): row
        for row in response_payload["terms"]
    }


def test_parser_contract_distinguishes_detected_empty_target_from_missing_column() -> None:
    detected_empty = _parse_multilingual_translation_table(
        ["ID", "CN", "EN"],
        [("A-1", "开始游戏", None)],
        sheet_name="Language",
        id_column=None,
        source_column=None,
        note_column=None,
        source_artifact_id="artifact",
        source_type="imported",
        include_empty=True,
    )
    missing = _parse_multilingual_translation_table(
        ["ID", "CN"],
        [("A-1", "开始游戏")],
        sheet_name="Language",
        id_column=None,
        source_column=None,
        note_column=None,
        source_artifact_id="artifact",
        source_type="imported",
        include_empty=True,
    )

    assert detected_empty.detected_columns["languages"]["en"]["target"] == {"name": "EN", "index": 2}
    assert detected_empty.rows[0]["target"] == ""
    assert detected_empty.rows[0]["target_column_present"] is True
    assert missing.detected_columns["languages"] == {}
    assert missing.rows == []


def test_official_language_table_template_import_skips_guide_sheet() -> None:
    with TestClient(app) as client:
        project = _create_project(client, "official language template")
        template = client.get("/api/import-templates/language-table")
        assert template.status_code == 200, template.text
        artifact = _upload_bytes(
            client,
            project["id"],
            "language-template.xlsx",
            template.content,
            "language_table",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        imported = client.post(
            f"/api/projects/{project['id']}/translations/import",
            json={"artifact_id": artifact["id"]},
        )

    assert imported.status_code == 200, imported.text
    payload = imported.json()
    assert payload["languages"] == ["en", "ko", "ja"]
    rows = _translation_rows(payload)
    assert rows[("领取奖励", "en")]["target"] == "Claim Reward"
    assert rows[("领取奖励", "en")]["sheet"] == "语言表"
    assert rows[("领取奖励", "ko")]["target"] == "보상 받기"
    assert rows[("领取奖励", "ja")]["target"] == "報酬を受け取る"


def test_official_glossary_template_preview_and_import_skip_guide_sheet() -> None:
    with TestClient(app) as client:
        project = _create_project(client, "official glossary template")
        template = client.get("/api/import-templates/glossary")
        assert template.status_code == 200, template.text
        artifact = _upload_bytes(
            client,
            project["id"],
            "glossary-template.xlsx",
            template.content,
            "term_base",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        preview = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": artifact["id"]},
        )
        imported = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": artifact["id"], "override_protected": True},
        )

    assert preview.status_code == 200, preview.text
    assert preview.json()["languages"] == ["en", "ko", "ja"]
    assert imported.status_code == 200, imported.text
    rows = _glossary_rows(imported.json())
    assert rows[("战力", "en")]["target"] == "CP"
    assert rows[("联盟", "ko")]["target"] == "연맹"
    assert rows[("联盟", "ja")]["target"] == "同盟"


def test_complete_language_table_guard_still_applies_after_skipping_guide_sheet() -> None:
    data_rows: list[list[object]] = [["ID", "CN", "EN"]]
    data_rows.extend([[f"A-{index}", f"源文{index}", f"Target {index}"] for index in range(1001)])
    content = _workbook_bytes(
        [
            ("填写说明", [["请在语言表页填写内容"]]),
            ("语言表", data_rows),
        ]
    )
    with TestClient(app) as client:
        project = _create_project(client, "guarded language template")
        artifact = _upload_bytes(client, project["id"], "full-language-template.xlsx", content, "language_table")
        preview = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": artifact["id"]},
        )

    assert preview.status_code == 400, preview.text
    assert "完整语言表" in str(preview.json()["detail"])
    assert db.list_glossary_terms(project["id"]) == []


def test_complete_language_table_guard_applies_to_csv() -> None:
    lines = ["ID,CN,EN"]
    lines.extend(f"A-{index},源文{index},Target {index}" for index in range(1001))
    content = ("\n".join(lines) + "\n").encode("utf-8-sig")
    with TestClient(app) as client:
        project = _create_project(client, "guarded language csv")
        artifact = _upload_bytes(client, project["id"], "full-language.csv", content, "language_table", "text/csv")
        preview = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": artifact["id"]},
        )

    assert preview.status_code == 400, preview.text
    assert "完整语言表" in str(preview.json()["detail"])
    assert db.list_glossary_terms(project["id"]) == []


def test_complete_language_table_guard_blocks_wide_json_but_allows_long_glossary_export_shape() -> None:
    wide_rows = [
        {"ID": f"A-{index}", "CN": f"源文{index}", "EN": f"Target {index}"}
        for index in range(1001)
    ]
    long_rows = [
        {
            "term_key": f"T-{index}",
            "source": f"术语{index}",
            "target": f"Term {index}",
            "target_alt": "",
            "language": "en",
        }
        for index in range(1001)
    ]
    with TestClient(app) as client:
        wide_project = _create_project(client, "guarded wide json")
        wide_upload = client.post(
            f"/api/projects/{wide_project['id']}/files?kind=term_base",
            files={
                "file": (
                    "language-table.json",
                    json.dumps(wide_rows, ensure_ascii=False).encode("utf-8"),
                    "application/json",
                )
            },
        )

        long_project = _create_project(client, "allowed long glossary json")
        long_artifact = _upload_bytes(
            client,
            long_project["id"],
            "glossary-export.json",
            json.dumps({"terms": long_rows}, ensure_ascii=False).encode("utf-8"),
            "term_base",
            "application/json",
        )
        long_preview = client.post(
            f"/api/projects/{long_project['id']}/glossary/import-preview",
            json={"artifact_id": long_artifact["id"]},
        )

    assert wide_upload.status_code == 400, wide_upload.text
    assert "完整语言表" in str(wide_upload.json()["detail"])
    assert long_preview.status_code == 200, long_preview.text
    assert long_preview.json()["languages"] == ["en"]


def test_implicit_xlsx_sheet_selection_returns_candidates_and_explicit_sheet_works() -> None:
    content = _workbook_bytes(
        [
            ("Data A", [["ID", "CN", "EN"], ["A-1", "开始游戏", "Start A"]]),
            ("Data B", [["ID", "CN", "EN"], ["B-1", "开始游戏", "Start B"]]),
        ]
    )
    with TestClient(app) as client:
        project = _create_project(client, "sheet choice")
        artifact = _upload_bytes(client, project["id"], "multiple.xlsx", content, "language_table")

        ambiguous = client.post(
            f"/api/projects/{project['id']}/translations/import",
            json={"artifact_id": artifact["id"]},
        )
        selected = client.post(
            f"/api/projects/{project['id']}/translations/import",
            json={"artifact_id": artifact["id"], "sheet": "Data B"},
        )

    assert ambiguous.status_code == 400, ambiguous.text
    assert ambiguous.json()["detail"] == {
        "code": "sheet_selection_required",
        "message": "检测到多个可导入的数据工作表，请选择后重试。",
        "candidates": ["Data A", "Data B"],
    }
    assert selected.status_code == 200, selected.text
    assert selected.json()["entries"][0]["entry_key"] == "B-1"
    assert selected.json()["entries"][0]["target"] == "Start B"


@pytest.mark.parametrize("endpoint", ["import-preview", "import"])
def test_glossary_import_rejects_artifact_from_another_project(endpoint: str) -> None:
    content = _workbook_bytes([("Glossary", [["ID", "CN", "EN"], ["T-1", "战力", "CP"]])])
    with TestClient(app) as client:
        owner = _create_project(client, "artifact owner")
        foreign = _create_project(client, "foreign project")
        artifact = _upload_bytes(client, owner["id"], "terms.xlsx", content, "term_base")

        response = client.post(
            f"/api/projects/{foreign['id']}/glossary/{endpoint}",
            json={"artifact_id": artifact["id"]},
        )

    assert response.status_code == 404, response.text
    assert db.list_glossary_terms(foreign["id"]) == []


def test_archive_translation_artifact_imports_only_requested_language(tmp_path: Path) -> None:
    project = db.insert_project("single language archive")
    path = tmp_path / "qa-final.xlsx"
    path.write_bytes(
        _workbook_bytes(
            [
                (
                    "Language",
                    [["ID", "CN", "EN", "KR", "JP"], ["A-1", "开始游戏", "Start Game", "게임 시작", "ゲーム開始"]],
                )
            ]
        )
    )
    artifact = db.add_artifact(project["id"], "QA final", path, "qa_final_workbook")

    result = archive_translation_artifact(project["id"], artifact["id"], language="en")

    assert result["languages"] == ["en"]
    persisted = db.list_translation_entries(project["id"])
    assert [(row["language"], row["target"]) for row in persisted] == [("en", "Start Game")]


def test_generic_target_header_requires_explicit_single_language_mapping() -> None:
    content = _workbook_bytes([("Language", [["ID", "CN", "translation"], ["A-1", "开始游戏", "Start Game"]])])
    with TestClient(app) as client:
        automatic_project = _create_project(client, "generic automatic")
        automatic_artifact = _upload_bytes(client, automatic_project["id"], "generic.xlsx", content, "language_table")
        automatic = client.post(
            f"/api/projects/{automatic_project['id']}/translations/import",
            json={"artifact_id": automatic_artifact["id"], "language": "en"},
        )

        explicit_project = _create_project(client, "generic explicit")
        explicit_artifact = _upload_bytes(client, explicit_project["id"], "generic.xlsx", content, "language_table")
        explicit = client.post(
            f"/api/projects/{explicit_project['id']}/translations/import",
            json={
                "artifact_id": explicit_artifact["id"],
                "language": "ko",
                "target_column": "translation",
            },
        )

    assert automatic.status_code == 400, automatic.text
    assert db.list_translation_entries(automatic_project["id"]) == []
    assert explicit.status_code == 200, explicit.text
    assert explicit.json()["languages"] == ["ko"]
    assert explicit.json()["entries"][0]["target"] == "Start Game"


def test_generic_glossary_target_header_requires_explicit_single_language_mapping() -> None:
    content = _workbook_bytes([("Glossary", [["ID", "CN", "target"], ["T-1", "战力", "Combat Power"]])])
    with TestClient(app) as client:
        automatic_project = _create_project(client, "generic glossary automatic")
        automatic_artifact = _upload_bytes(client, automatic_project["id"], "generic.xlsx", content, "term_base")
        automatic = client.post(
            f"/api/projects/{automatic_project['id']}/glossary/import-preview",
            json={"artifact_id": automatic_artifact["id"], "language": "en"},
        )

        explicit_project = _create_project(client, "generic glossary explicit")
        explicit_artifact = _upload_bytes(client, explicit_project["id"], "generic.xlsx", content, "term_base")
        explicit = client.post(
            f"/api/projects/{explicit_project['id']}/glossary/import",
            json={
                "artifact_id": explicit_artifact["id"],
                "language": "ko",
                "target_column": "target",
            },
        )

    assert automatic.status_code == 400, automatic.text
    assert db.list_glossary_terms(automatic_project["id"]) == []
    assert explicit.status_code == 200, explicit.text
    assert explicit.json()["languages"] == ["ko"]
    assert explicit.json()["terms"][0]["target"] == "Combat Power"


@pytest.mark.parametrize(("target_header", "legacy_alt_header"), [("target", "target_alt"), ("translation", "EN2"), ("译文", "EN2")])
def test_generic_primary_column_is_never_discarded_by_automatic_legacy_alt_promotion(
    target_header: str,
    legacy_alt_header: str,
) -> None:
    content = _workbook_bytes(
        [
            (
                "Data",
                [["ID", "CN", target_header, legacy_alt_header], ["A-1", "开始游戏", "Start Game", "Begin Game"]],
            )
        ]
    )
    with TestClient(app) as client:
        automatic_project = _create_project(client, f"generic plus alt automatic {target_header}")
        automatic_artifact = _upload_bytes(client, automatic_project["id"], "generic-alt.xlsx", content, "language_table")
        automatic = client.post(
            f"/api/projects/{automatic_project['id']}/translations/import",
            json={"artifact_id": automatic_artifact["id"], "language": "en"},
        )

        explicit_project = _create_project(client, f"generic plus alt explicit {target_header}")
        explicit_artifact = _upload_bytes(client, explicit_project["id"], "generic-alt.xlsx", content, "language_table")
        explicit = client.post(
            f"/api/projects/{explicit_project['id']}/translations/import",
            json={
                "artifact_id": explicit_artifact["id"],
                "language": "en",
                "target_column": target_header,
            },
        )

    assert automatic.status_code == 400, automatic.text
    assert db.list_translation_entries(automatic_project["id"]) == []
    assert explicit.status_code == 200, explicit.text
    assert explicit.json()["entries"][0]["target"] == "Start Game"
    assert explicit.json()["entries"][0]["target_alt"] == ""


def test_missing_explicit_target_column_never_falls_back_to_detected_language_column() -> None:
    content = _workbook_bytes([("Data", [["ID", "CN", "EN"], ["A-1", "开始游戏", "Start Game"]])])
    with TestClient(app) as client:
        translation_project = _create_project(client, "missing translation column")
        translation_artifact = _upload_bytes(
            client,
            translation_project["id"],
            "translations.xlsx",
            content,
            "language_table",
        )
        translation = client.post(
            f"/api/projects/{translation_project['id']}/translations/import",
            json={
                "artifact_id": translation_artifact["id"],
                "language": "en",
                "target_column": "MISSING",
            },
        )

        glossary_project = _create_project(client, "missing glossary column")
        glossary_artifact = _upload_bytes(client, glossary_project["id"], "glossary.xlsx", content, "term_base")
        glossary = client.post(
            f"/api/projects/{glossary_project['id']}/glossary/import-preview",
            json={
                "artifact_id": glossary_artifact["id"],
                "language": "en",
                "target_column": "MISSING",
            },
        )

    assert translation.status_code == 400, translation.text
    assert db.list_translation_entries(translation_project["id"]) == []
    assert glossary.status_code == 400, glossary.text
    assert db.list_glossary_terms(glossary_project["id"]) == []


@pytest.mark.parametrize(
    ("endpoint", "kind", "field", "missing_value"),
    [
        ("translations/import", "language_table", "id_column", "MISSING_ID"),
        ("translations/import", "language_table", "target_alt_column", "MISSING_ALT"),
        ("translations/import", "language_table", "note_column", "MISSING_NOTE"),
        ("glossary/import-preview", "term_base", "term_key_column", "MISSING_ID"),
        ("glossary/import-preview", "term_base", "target_alt_column", "MISSING_ALT"),
        ("glossary/import-preview", "term_base", "category_column", "MISSING_CATEGORY"),
        ("glossary/import-preview", "term_base", "note_column", "MISSING_NOTE"),
    ],
)
def test_missing_explicit_optional_column_never_falls_back_to_detected_alias(
    endpoint: str,
    kind: str,
    field: str,
    missing_value: str,
) -> None:
    content = _workbook_bytes(
        [
            (
                "Data",
                [["ID", "CN", "EN", "EN2", "分类", "备注"], ["A-1", "开始游戏", "Start Game", "Begin Game", "UI", "Button"]],
            )
        ]
    )
    with TestClient(app) as client:
        project = _create_project(client, f"missing optional {field}")
        artifact = _upload_bytes(client, project["id"], "mapped.xlsx", content, kind)
        response = client.post(
            f"/api/projects/{project['id']}/{endpoint}",
            json={"artifact_id": artifact["id"], "language": "en", field: missing_value},
        )

    assert response.status_code == 400, response.text
    assert db.list_translation_entries(project["id"]) == []
    assert db.list_glossary_terms(project["id"]) == []


def test_default_single_language_imports_skip_blank_targets() -> None:
    content = "ID,CN,target\nA-1,未翻译,\nA-2,开始游戏,Start Game\n".encode("utf-8-sig")
    with TestClient(app) as client:
        translation_project = _create_project(client, "skip blank translation")
        translation_artifact = _upload_bytes(
            client,
            translation_project["id"],
            "translations.csv",
            content,
            "language_table",
            "text/csv",
        )
        translation = client.post(
            f"/api/projects/{translation_project['id']}/translations/import",
            json={
                "artifact_id": translation_artifact["id"],
                "language": "en",
                "target_column": "target",
            },
        )

        glossary_project = _create_project(client, "skip blank glossary")
        glossary_artifact = _upload_bytes(
            client,
            glossary_project["id"],
            "glossary.csv",
            content,
            "term_base",
            "text/csv",
        )
        glossary = client.post(
            f"/api/projects/{glossary_project['id']}/glossary/import",
            json={
                "artifact_id": glossary_artifact["id"],
                "language": "en",
                "target_column": "target",
            },
        )

    assert translation.status_code == 200, translation.text
    assert translation.json()["imported_count"] == 1
    assert [row["source"] for row in db.list_translation_entries(translation_project["id"])] == ["开始游戏"]
    assert glossary.status_code == 200, glossary.text
    assert glossary.json()["imported_count"] == 1
    assert [row["source"] for row in db.list_glossary_terms(glossary_project["id"])] == ["开始游戏"]


def test_backfill_only_preserves_empty_candidates_and_selects_all_empty_glossary_sheet(tmp_path: Path) -> None:
    content = _workbook_bytes(
        [
            ("填写说明", [["说明"]]),
            ("Glossary", [["ID", "CN", "EN", "EN2"], ["G-1", "能量", "", ""], ["G-2", "耐力", "", ""]]),
        ]
    )
    generated = tmp_path / "all-empty-glossary.xlsx"
    generated.write_bytes(content)
    project = db.insert_project("backfill empty candidates")

    result = backfill_project_glossary_from_final(project["id"], generated)

    with TestClient(app) as client:
        artifact = _upload_bytes(client, project["id"], generated.name, content, "term_base")
        ordinary_import = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": artifact["id"]},
        )

    assert result["candidates"] == 2
    assert result["inserted"] == 2
    assert ordinary_import.status_code == 400, ordinary_import.text
    assert db.list_glossary_terms(project["id"]) == []


@pytest.mark.parametrize("header", ["VN", "VI"])
def test_vietnamese_import_aliases_persist_vn_and_export_vn_header(header: str) -> None:
    content = f"ID,CN,{header}\nA-1,开始游戏,Bắt đầu trò chơi\n".encode("utf-8-sig")
    with TestClient(app) as client:
        project = _create_project(client, f"Vietnamese {header}")
        artifact = _upload_bytes(client, project["id"], f"vietnamese-{header}.csv", content, "language_table", "text/csv")
        imported = client.post(
            f"/api/projects/{project['id']}/translations/import",
            json={"artifact_id": artifact["id"]},
        )
        exported = client.get(f"/api/projects/{project['id']}/translations/export?format=xlsx&language=vn")

    assert imported.status_code == 200, imported.text
    assert imported.json()["languages"] == ["vn"]
    assert imported.json()["entries"][0]["language"] == "vn"
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=True)
    try:
        assert [cell.value for cell in workbook.active[1]][:3] == ["ID", "CN", "VN"]
    finally:
        workbook.close()


@pytest.mark.parametrize("header", ["VN", "VI"])
def test_vietnamese_glossary_candidate_scan_auto_detects_canonical_header(header: str) -> None:
    content = _workbook_bytes([
        (
            "Language",
            [
                ["ID", "CN", header],
                ["A-1", "战力", "Sức mạnh"],
                ["A-2", "提升战力", "Tăng sức mạnh"],
                ["A-3", "战力奖励", "Phần thưởng sức mạnh"],
                ["A-4", "战力排行", "Xếp hạng sức mạnh"],
                ["A-5", "战力系统", "Hệ thống sức mạnh"],
                ["A-6", "战力属性", "Thuộc tính sức mạnh"],
                ["A-7", "战力成长", "Tăng trưởng sức mạnh"],
            ],
        )
    ])
    with TestClient(app) as client:
        project = _create_project(client, f"Vietnamese glossary candidate {header}")
        artifact = _upload_bytes(
            client,
            project["id"],
            f"vietnamese-candidate-{header}.xlsx",
            content,
            "language_table",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        extracted = client.post(
            f"/api/projects/{project['id']}/glossary/extract",
            json={
                "input_artifact_id": artifact["id"],
                "project_name": project["name"],
                "id_column": "ID",
                "source_column": "CN",
                "language": "vn",
                "include_empty_final_terms": True,
                "ai_candidate_supplement": False,
                "update_project_prompt": False,
            },
        )
        batches = client.get(
            f"/api/projects/{project['id']}/glossary/batches",
            params={"language": "vn"},
        )

    assert extracted.status_code == 200, extracted.text
    payload = extracted.json()
    assert payload["run"]["language"] == "vn"
    assert payload["glossary_backfill"]["batch_id"]
    assert batches.status_code == 200, batches.text
    assert batches.json()["active_batch"]["language"] == "vn"
    candidates = batches.json()["candidates"]
    assert candidates
    assert all(candidate["language"] == "vn" for candidate in candidates)
    assert any(candidate["target"] == "Sức mạnh" for candidate in candidates)


def test_legacy_en2_is_promoted_once_without_persisting_target_alt() -> None:
    content = _workbook_bytes(
        [
            (
                "Language",
                [
                    ["ID", "CN", "EN", "EN2"],
                    ["A-1", "开始游戏", "Start Game", "Begin Game"],
                    ["A-2", "领取奖励", "", "Claim Reward"],
                ],
            )
        ]
    )
    with TestClient(app) as client:
        project = _create_project(client, "legacy EN2")
        seeded = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "A-1",
                "source": "开始游戏",
                "target": "Old Start",
                "target_alt": "Historical Begin",
                "language": "en",
            },
        )
        assert seeded.status_code == 200, seeded.text
        artifact = _upload_bytes(client, project["id"], "legacy-en2.xlsx", content, "language_table")
        imported = client.post(
            f"/api/projects/{project['id']}/translations/import",
            json={"artifact_id": artifact["id"]},
        )

    assert imported.status_code == 200, imported.text
    by_key = {row["entry_key"]: row for row in imported.json()["entries"]}
    assert by_key["A-1"]["target"] == "Start Game"
    assert by_key["A-1"]["target_alt"] == ""
    assert by_key["A-2"]["target"] == "Claim Reward"
    assert by_key["A-2"]["target_alt"] == ""
    assert all(row["target_alt"] == "" for row in db.list_translation_entries(project["id"]))


def test_glossary_import_clears_legacy_target_alt_on_existing_record() -> None:
    content = _workbook_bytes(
        [
            (
                "Glossary",
                [
                    ["ID", "CN", "EN", "EN2"],
                    ["T-1", "战力", "Combat Power", "Historical CP"],
                ],
            )
        ]
    )
    with TestClient(app) as client:
        project = _create_project(client, "legacy glossary EN2")
        seeded = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "term_key": "T-1",
                "source": "战力",
                "target": "Old CP",
                "target_alt": "Historical CP",
                "language": "en",
            },
        )
        assert seeded.status_code == 200, seeded.text
        artifact = _upload_bytes(client, project["id"], "legacy-glossary-en2.xlsx", content, "term_base")
        imported = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": artifact["id"], "override_protected": True},
        )

    assert imported.status_code == 200, imported.text
    assert imported.json()["terms"][0]["target"] == "Combat Power"
    assert imported.json()["terms"][0]["target_alt"] == ""
    persisted = db.get_glossary_term(imported.json()["terms"][0]["id"])
    assert persisted["target_alt"] == ""
    assert persisted["confirmed"] is False
    assert persisted["review_status"] == "pending"


@pytest.mark.parametrize("lookup", [lookup_terms, lookup_translation_entries])
def test_lookup_uses_only_primary_target(lookup) -> None:
    rows = [
        {
            "id": "legacy-only",
            "source": "领取奖励",
            "target": "",
            "target_alt": "Legacy Reward",
            "language": "en",
        },
        {
            "id": "primary",
            "source": "开始游戏",
            "target": "Start Game",
            "target_alt": "Begin Game",
            "language": "en",
        },
    ]

    hits = lookup("开始游戏并领取奖励", rows, min_length=2, limit=10)

    assert [hit["id"] for hit in hits] == ["primary"]
    assert hits[0]["target"] == "Start Game"
    assert hits[0]["target_alt"] == ""


def test_multilingual_translation_csv_api_roundtrip() -> None:
    content = "ID,CN,EN,KR\nA-1,开始游戏,Start Game,게임 시작\n".encode("utf-8-sig")
    with TestClient(app) as client:
        project = _create_project(client, "translation csv")
        artifact = _upload_bytes(client, project["id"], "translations.csv", content, "language_table", "text/csv")
        imported = client.post(
            f"/api/projects/{project['id']}/translations/import",
            json={"artifact_id": artifact["id"]},
        )

    assert imported.status_code == 200, imported.text
    rows = _translation_rows(imported.json())
    assert imported.json()["languages"] == ["en", "ko"]
    assert rows[("开始游戏", "en")]["target"] == "Start Game"
    assert rows[("开始游戏", "ko")]["target"] == "게임 시작"


def test_multilingual_glossary_csv_api_roundtrip() -> None:
    content = "ID,CN,EN,KR,分类,备注\nT-1,战力,CP,전투력,属性,统一\n".encode("utf-8-sig")
    with TestClient(app) as client:
        project = _create_project(client, "glossary csv")
        artifact = _upload_bytes(client, project["id"], "glossary.csv", content, "term_base", "text/csv")
        imported = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": artifact["id"]},
        )

    assert imported.status_code == 200, imported.text
    rows = _glossary_rows(imported.json())
    assert imported.json()["languages"] == ["en", "ko"]
    assert rows[("战力", "en")]["target"] == "CP"
    assert rows[("战力", "ko")]["target"] == "전투력"


def test_glossary_json_accepts_export_object_and_raw_array_api_roundtrips() -> None:
    with TestClient(app) as client:
        source = _create_project(client, "json export source")
        for language, target in (("en", "Alliance"), ("ko", "연맹")):
            created = client.post(
                f"/api/projects/{source['id']}/glossary",
                json={"term_key": "T-1", "source": "联盟", "target": target, "language": language},
            )
            assert created.status_code == 200, created.text
        exported = client.get(f"/api/projects/{source['id']}/glossary/export?format=json")
        assert exported.status_code == 200, exported.text

        for name, payload in (("object", exported.json()), ("array", exported.json()["terms"])):
            target_project = _create_project(client, f"json import {name}")
            artifact = _upload_bytes(
                client,
                target_project["id"],
                f"glossary-{name}.json",
                json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                "term_base",
                "application/json",
            )
            imported = client.post(
                f"/api/projects/{target_project['id']}/glossary/import",
                json={"artifact_id": artifact["id"]},
            )
            assert imported.status_code == 200, imported.text
            rows = _glossary_rows(imported.json())
            assert rows[("联盟", "en")]["target"] == "Alliance"
            assert rows[("联盟", "ko")]["target"] == "연맹"


@pytest.mark.parametrize(
    ("kind", "endpoint"),
    [
        ("language_table", "translations/import"),
        ("term_base", "glossary/import"),
    ],
)
def test_legacy_xls_is_explicitly_rejected(kind: str, endpoint: str) -> None:
    with TestClient(app) as client:
        project = _create_project(client, f"legacy xls {kind}")
        artifact = _upload_bytes(client, project["id"], "legacy.xls", b"legacy-xls-content", kind)
        response = client.post(
            f"/api/projects/{project['id']}/{endpoint}",
            json={"artifact_id": artifact["id"]},
        )

    assert response.status_code == 400, response.text
    detail = str(response.json()["detail"])
    assert "XLS" in detail
    assert "XLSX" in detail
