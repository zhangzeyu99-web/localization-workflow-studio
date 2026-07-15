from __future__ import annotations

import ast
import inspect
import shutil
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from app.workflow import qa as qa_workflow
from app.workflow import translation as translation_workflow


def _write_language_workbook(path: Path, translation: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Language"
    worksheet.append(["ID", "CN", "VI"])
    worksheet.append([1, "领取奖励", translation])
    workbook.save(path)
    workbook.close()


def test_vietnamese_change_collection_reads_vi_column(tmp_path: Path) -> None:
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"
    _write_language_workbook(before, "Nhan thuong cu")
    _write_language_workbook(after, "Nhận thưởng")

    changes = qa_workflow._collect_workbook_translation_changes(before, after, language="vn")

    assert changes[0]["previous_translation"] == "Nhan thuong cu"
    assert changes[0]["translation"] == "Nhận thưởng"


def test_vietnamese_manual_and_model_fixes_use_vi_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "language.xlsx"
    _write_language_workbook(workbook_path, "Nhan thuong cu")
    issue = {"id": 1, "sheet": "Language", "row": 2, "severity": "hard"}

    context = qa_workflow._model_fix_row_context(workbook_path, issue, language="vn")
    applied = qa_workflow._apply_workbook_fixes(
        workbook_path,
        [{"sheet": "Language", "row": 2, "record_id": "1", "translation": "Nhận thưởng"}],
        "run-vn",
        language="vn",
    )

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        assert context["current_translation"] == "Nhan thuong cu"
        assert workbook["Language"].cell(2, 3).value == "Nhận thưởng"
    finally:
        workbook.close()
    assert applied[0]["previous_translation"] == "Nhan thuong cu"
    assert applied[0]["translation"] == "Nhận thưởng"


def test_translation_harness_uses_mapped_language_for_prepare_and_apply() -> None:
    source = inspect.getsource(translation_workflow.translate_run)
    tree = ast.parse(source)
    mapped_assignments = [
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.Assign)
        and any(isinstance(target, ast.Name) and target.id == "workflow_language" for target in node.targets)
        and isinstance(node.value, ast.Call)
        and isinstance(node.value.func, ast.Name)
        and node.value.func.id == "workflow_language_code"
    ]
    mapped_lang_arguments = 0
    for node in ast.walk(tree):
        if not isinstance(node, ast.List):
            continue
        for index, element in enumerate(node.elts[:-1]):
            if (
                isinstance(element, ast.Constant)
                and element.value == "--lang"
                and isinstance(node.elts[index + 1], ast.Name)
                and node.elts[index + 1].id == "workflow_language"
            ):
                mapped_lang_arguments += 1

    assert len(mapped_assignments) == 1
    assert mapped_lang_arguments == 2


def test_run_localization_qa_maps_vn_only_at_subprocess_boundary(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook_path = tmp_path / "source.xlsx"
    _write_language_workbook(workbook_path, "Nhan thuong cu")
    review_calls: list[list[str]] = []
    quality_calls: list[list[str]] = []
    mapped_calls: list[str] = []
    studio_helper_languages: list[str] = []
    artifacts: list[dict[str, Any]] = []

    def fake_workflow_language_code(value: Any) -> str:
        mapped_calls.append(str(value))
        return "vi"

    def fake_review(args: list[str], _cwd: Path, _run_id: str) -> None:
        review_calls.append(args)
        machine_dir = Path(args[args.index("--output-dir") + 1])
        shutil.copy2(workbook_path, machine_dir / "result_vi.xlsx")
        _write_language_workbook(machine_dir / "report_vi.xlsx", "Nhan thuong cu")

    def fake_quality(args: list[str], _run_id: str) -> dict[str, Any]:
        quality_calls.append(args)
        return {"passed": True, "issues": [], "failures": []}

    def fake_project_harness(
        _path: Path,
        _harness: dict[str, Any],
        *,
        language: str,
    ) -> dict[str, Any]:
        studio_helper_languages.append(language)
        return {"passed": True, "hard_errors": 0, "soft_warnings": 0, "issues": []}

    def fake_semantic(
        _run_id: str,
        _project_id: str,
        _path: Path,
        _quality: dict[str, Any],
        _project_quality: dict[str, Any],
        *,
        language: str,
        settings: dict[str, Any] | None,
    ) -> dict[str, Any]:
        studio_helper_languages.append(language)
        return {"status": "passed", "passed": True, "hard_errors": 0, "soft_warnings": 0, "issues": []}

    def fake_collect(_before: Path, _after: Path, language: str = "en") -> list[dict[str, Any]]:
        studio_helper_languages.append(language)
        return []

    def fake_add_artifact(
        _project_id: str,
        _name: str,
        path: Path,
        kind: str,
        **kwargs: Any,
    ) -> dict[str, Any]:
        artifact = {
            "id": f"artifact-{len(artifacts) + 1}",
            "path": str(path),
            "kind": kind,
            "metadata": kwargs.get("metadata") or {},
        }
        artifacts.append(artifact)
        return artifact

    monkeypatch.setattr(qa_workflow, "workflow_language_code", fake_workflow_language_code, raising=False)
    monkeypatch.setattr(qa_workflow, "run_subprocess", fake_review)
    monkeypatch.setattr(qa_workflow, "_run_quality_json", fake_quality)
    monkeypatch.setattr(qa_workflow, "run_project_harness_qa", fake_project_harness)
    monkeypatch.setattr(qa_workflow, "run_semantic_qa_report", fake_semantic)
    monkeypatch.setattr(qa_workflow, "_collect_workbook_translation_changes", fake_collect)
    monkeypatch.setattr(qa_workflow.db, "add_artifact", fake_add_artifact)

    result = qa_workflow.run_localization_qa(
        project={"id": "project-vn"},
        run_id="run-vn",
        workbook_path=workbook_path,
        output_dir=tmp_path / "qa",
        glossary_snapshot={"id": "glossary-vn", "path": str(tmp_path / "terms.json")},
        harness_snapshot={"project_harness": {}},
        language="vn",
    )

    assert mapped_calls == ["vn"]
    assert review_calls[0][review_calls[0].index("--lang") + 1] == "vi"
    assert quality_calls[0][quality_calls[0].index("--lang") + 1] == "vi"
    assert Path(result["qa_workbook"]).name == "result_vi.xlsx"
    assert studio_helper_languages == ["vn", "vn", "vn"]
    assert artifacts[-1]["metadata"]["language"] == "vn"
