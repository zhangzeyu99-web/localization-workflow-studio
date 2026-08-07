from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.workflow.announcement_outputs import (
    _read_announcement_translation_workbook,
    _write_announcement_translation_workbook,
)


def test_translation_workbook_replaces_oversized_sentence_adaptations_with_valid_json(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "translation.xlsx"
    segment = {
        "id": "segment-1",
        "index": 0,
        "kind": "line",
        "source": "公告",
        "source_file": "announcement.txt",
    }
    lookup = {
        "en": {
            "sentence_adaptations": [
                {
                    "priority": index,
                    "match_type": "official_similar",
                    "id": f"adaptation-{index}",
                    "announcement_cn": "公告",
                    "official_cn_template": "公告",
                    "target": "x" * 500,
                }
                for index in range(100)
            ]
        }
    }

    _write_announcement_translation_workbook(
        workbook_path,
        {"id": "task-1"},
        [segment],
        ["en"],
        lookup,
    )

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        cell_value = workbook["Translations"]["H2"].value
    finally:
        workbook.close()

    assert json.loads(str(cell_value)) == {}


def test_translation_workbook_reader_ignores_legacy_truncated_sentence_adaptations(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "legacy-truncated.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Translations"
    sheet.append(
        [
            "segment_id",
            "CN",
            "protected_tokens",
            "term_hits_json",
            "sentence_adaptations_json",
            "EN",
        ]
    )
    sheet.append(
        [
            "segment-1",
            "公告",
            "[]",
            "{}",
            '{"en":[{"target":"' + ("x" * 32740),
            "Announcement",
        ]
    )
    workbook.save(workbook_path)
    workbook.close()

    rows = _read_announcement_translation_workbook(workbook_path, ["en"])

    assert rows["segment-1"]["sentence_adaptations"] == {}
    assert rows["segment-1"]["translations"] == {"en": "Announcement"}


def test_translation_workbook_preserves_term_category_for_qa(tmp_path: Path) -> None:
    workbook_path = tmp_path / "translation.xlsx"
    segment = {
        "id": "segment-1",
        "index": 0,
        "kind": "line",
        "source": "联盟成员可领取奖励",
        "source_file": "announcement.txt",
    }
    lookup = {
        "en": {
            "terms": [
                {
                    "source": "联盟成员",
                    "target": "Alliance Member",
                    "category": "普通术语",
                }
            ]
        }
    }

    _write_announcement_translation_workbook(
        workbook_path,
        {"id": "task-1"},
        [segment],
        ["en"],
        lookup,
    )

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        term_hits = json.loads(str(workbook["Translations"]["G2"].value))
    finally:
        workbook.close()

    assert term_hits == {
        "en": [
            {
                "source": "联盟成员",
                "target": "Alliance Member",
                "category": "普通术语",
            }
        ]
    }
