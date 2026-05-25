from __future__ import annotations

import os
import re
import shutil
import tempfile
from pathlib import Path

os.environ["LWS_DATA_ROOT"] = str(Path(tempfile.gettempdir()) / "lws-test-data")

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
import app.workflow as workflow
from app.config import DEFAULT_SETTINGS, save_settings
from app.main import app
from app.providers import mock_translate_batch
from app.workflow import backfill_project_glossary_from_final


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    if data_root.exists():
        shutil.rmtree(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    save_settings(DEFAULT_SETTINGS)


def _sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "领取奖励", ""])
    ws.append([2, "开始游戏", ""])
    ws.append([3, "系统错误", ""])
    ws.append([4, "主线任务", ""])
    ws.append([5, "欢迎回来，{playerName}", ""])
    wb.save(path)
    wb.close()


def _sample_term_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2", "分类", "note"])
    ws.append([1, "最强指挥官", "Strongest Commander", "Top Commander", "title", "confirmed project term"])
    ws.append([2, "联盟", "Alliance", "Guild", "system", "common game term"])
    wb.save(path)
    wb.close()


def _translated_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "领取奖励", "Claim Rewards"])
    ws.append([2, "开始游戏", "Start Game"])
    ws.append([3, "系统错误", "System Error"])
    ws.append([4, "主线任务", "Main Quest"])
    ws.append([5, "欢迎回来，{playerName}", "Welcome back, {playerName}"])
    wb.create_sheet("EmptySheet")
    wb.save(path)
    wb.close()


def _project_harness_failed_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "棰嗗彇濂栧姳", "Forbidden Brand Reward"])
    ws.append([2, "Start game source", "Start Game"])
    wb.save(path)
    wb.close()


def test_mock_provider_preserves_numeric_square_placeholders() -> None:
    result = mock_translate_batch(
        [{"id": 236, "source": "审批文号：新广出审[2016]2751号\\n出版物号:ISBN"}],
        {},
    )
    assert "[2016]" in result[0].translation
    assert "\\n" in result[0].translation


def test_mock_provider_runs_english_workflow_end_to_end(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project_response = client.post(
            "/api/projects",
            json={
                "name": "Synthetic Frontier",
                "type": "科幻 SLG",
                "icon": "🚀",
                "description": "Synthetic public demo project.",
            },
        )
        assert project_response.status_code == 200
        project = project_response.json()

        analysis_response = client.post(
            f"/api/projects/{project['id']}/analyze",
            json={"intro": "A synthetic strategy game for testing localization workflow.", "asset_artifact_ids": []},
        )
        assert analysis_response.status_code == 200
        assert "只返回 JSONL" in analysis_response.json()["prompt"]

        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert upload_response.status_code == 200
        source_artifact = upload_response.json()

        run_response = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "translation",
                "language": "en",
                "input_artifact_id": source_artifact["id"],
                "batch_size": 3,
            },
        )
        assert run_response.status_code == 200
        run = run_response.json()

        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "mock", "allow_mock": True})
        assert translate_response.status_code == 200, translate_response.text
        result = translate_response.json()
        assert result["run"]["status"] == "passed"
        kinds = {artifact["kind"] for artifact in result["artifacts"]}
        assert {
            "raw_translated_workbook",
            "qa_final_workbook",
            "qa_report",
            "qa_result",
            "qa_changes",
            "translation_manifest",
            "glossary_snapshot",
            "prompt_snapshot",
            "project_harness_snapshot",
        }.issubset(kinds)
        final_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "qa_final_workbook")
        assert final_artifact["role"] == "translation_workbook"
        assert final_artifact["origin"] == "generated"
        metadata = result["run"]["metadata"]
        assert metadata["input_artifacts"]["source_workbook"] == source_artifact["id"]
        assert metadata["input_artifacts"]["qa_final_workbook"] == final_artifact["id"]
        assert metadata["semantic_qa"]["status"] == "skipped_no_key"
        assert metadata["translation_archive"]["imported_count"] == 5
        assert result["run"]["metadata"]["harness"]["source"] == "project_harness"
        progress = metadata["translation_progress"]
        assert progress["batch_size"] == 3
        assert progress["total_batches"] == 2
        assert progress["completed_batches"] == 2
        assert progress["percent"] == 100
        batch_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "translation" / "batches_3"
        assert (batch_dir / "batch_00001.jsonl").exists()
        assert (batch_dir / "batch_00002.jsonl").exists()
        events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("completed and persisted" in event["message"] for event in events)
        project_detail_response = client.get(f"/api/projects/{project['id']}")
        assert project_detail_response.status_code == 200
        assert project_detail_response.json()["stats"]["words"] == "33"
        assert project_detail_response.json()["stats"]["archived_rows"] == 5
        assert project_detail_response.json()["stats"]["translation_runs"] == 1
        assert project_detail_response.json()["stats"]["qa_runs"] == 0
        assert project_detail_response.json()["stats"]["langs"] == 1
        resume_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "mock", "allow_mock": True, "batch_size": 3})
        assert resume_response.status_code == 200, resume_response.text
        resume_events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("resume: batch 1/2 already completed" in event["message"] for event in resume_events)


def test_translation_batch_retry_persists_after_transient_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = tmp_path / "retry-language.xlsx"
    _sample_workbook(workbook)
    calls = {"count": 0}

    async def flaky_translate_batch(batch, settings, project_prompt, provider_override=None, protocol_override=None):
        _ = settings, project_prompt, provider_override, protocol_override
        calls["count"] += 1
        if calls["count"] == 1:
            raise RuntimeError("transient provider failure")
        return mock_translate_batch(batch, settings)

    monkeypatch.setattr(workflow, "translate_batch", flaky_translate_batch)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "E2E Retry", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("retry-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": source_artifact["id"], "batch_size": 3},
        ).json()
        response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "mock", "allow_mock": True, "batch_size": 3})
        assert response.status_code == 200, response.text
        events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("failed attempt 1/3" in event["message"] for event in events)
        assert any("batch 1/2 completed and persisted" in event["message"] for event in events)
        batch_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "translation" / "batches_3"
        assert (batch_dir / "batch_00001.jsonl").exists()
        assert not (batch_dir / "batch_00001.error.json").exists()


def test_mock_provider_is_blocked_for_real_project_without_explicit_allow(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "小小战机", "type": "飞行射击"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        source_artifact = upload_response.json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": source_artifact["id"]},
        ).json()

        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "mock"})
        assert translate_response.status_code == 200
        result = translate_response.json()
        assert result["run"]["status"] == "needs_input"
        assert result["run"]["metadata"]["reason"] == "mock provider is blocked for real project translation"
        assert result["artifacts"] == []


def test_assets_register_role_and_origin_with_legacy_kind_mapping(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Asset Registry", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert upload_response.status_code == 200
        uploaded = upload_response.json()
        assert uploaded["role"] == "language_source"
        assert uploaded["origin"] == "uploaded"

        assets_response = client.get(f"/api/projects/{project['id']}/assets?role=language_source")
        assert assets_response.status_code == 200
        assert [asset["id"] for asset in assets_response.json()] == [uploaded["id"]]

        patch_response = client.patch(
            f"/api/artifacts/{uploaded['id']}",
            json={"role": "translation_workbook", "origin": "imported", "label": "Existing EN workbook"},
        )
        assert patch_response.status_code == 200
        assert patch_response.json()["role"] == "translation_workbook"
        assert patch_response.json()["origin"] == "imported"


def test_markdown_reference_material_uploads_and_feeds_analysis(tmp_path: Path) -> None:
    material = tmp_path / "project_brief.md"
    material.write_text(
        "# Tomorrow 2\n\nGame type: Sci-fi SLG.\nTarget players: mobile strategy players.\nGameplay: base building and alliance war.",
        encoding="utf-8",
    )

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Markdown Material", "type": "QA"}).json()
        with material.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief.md", fh, "text/markdown")},
            )
        assert upload_response.status_code == 200
        artifact = upload_response.json()
        assert artifact["kind"] == "asset"
        assert artifact["origin"] == "uploaded"
        assert artifact["mime"] == "text/markdown"
        assert Path(artifact["path"]).exists()

        notes = workflow.analyze_assets([artifact["id"]], DEFAULT_SETTINGS)
        assert notes and "text_material:" in notes[0]
        assert "Sci-fi SLG" in notes[0]

        analysis_response = client.post(
            f"/api/projects/{project['id']}/analyze",
            json={"intro": "", "asset_artifact_ids": [artifact["id"]]},
        )
        assert analysis_response.status_code == 200
        profile = analysis_response.json()["project"]["profile"]
        assert profile["asset_notes"]
        assert "Sci-fi SLG" in profile["asset_notes"][0]


def test_duplicate_reference_material_upload_reuses_existing_artifact(tmp_path: Path) -> None:
    material = tmp_path / "project_brief.md"
    material.write_text("# Tomorrow 2\n\nSame project brief.", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Duplicate Material", "type": "QA"}).json()
        with material.open("rb") as fh:
            first_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief.md", fh, "text/markdown")},
            )
        assert first_response.status_code == 200
        first = first_response.json()
        assert first["duplicate"] is False

        with material.open("rb") as fh:
            second_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief-copy.md", fh, "text/markdown")},
            )
        assert second_response.status_code == 200
        second = second_response.json()
        assert second["duplicate"] is True
        assert second["id"] == first["id"]

        detail = client.get(f"/api/projects/{project['id']}").json()
        assets = [artifact for artifact in detail["artifacts"] if artifact["kind"] == "asset"]
        assert [artifact["id"] for artifact in assets] == [first["id"]]


def test_translation_readiness_skips_filled_translation_workbook(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Readiness Skip", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        artifact = upload_response.json()

        readiness_response = client.get(f"/api/artifacts/{artifact['id']}/translation-readiness?batch_size=90")
        assert readiness_response.status_code == 200
        readiness = readiness_response.json()
        assert readiness["ready_for_qa"] is True
        assert readiness["needs_translation"] is False
        assert readiness["source_rows"] == 5
        assert readiness["translated_rows"] == 5

        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": artifact["id"]},
        ).json()
        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "mock", "allow_mock": True})
        assert translate_response.status_code == 200
        result = translate_response.json()
        assert result["run"]["status"] == "needs_input"
        assert result["run"]["metadata"]["reason"] == "input already contains target translations; run QA instead"
        assert result["artifacts"] == []


def test_glossary_preview_import_and_export(tmp_path: Path) -> None:
    terms = tmp_path / "terms.xlsx"
    _sample_term_workbook(terms)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Import", "type": "QA"}).json()
        with terms.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        term_artifact = upload_response.json()

        preview_response = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": term_artifact["id"]},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["rows"][0]["source"] == "最强指挥官"
        assert preview["rows"][0]["target"] == "Strongest Commander"
        assert preview["rows"][0]["target_alt"] == "Top Commander"
        assert preview["rows"][0]["term_key"] == "1"
        assert preview["rows"][0]["category"] == "title"

        import_response = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": term_artifact["id"]},
        )
        assert import_response.status_code == 200
        assert import_response.json()["imported_count"] == 2

        project_terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert {term["source"] for term in project_terms} == {"最强指挥官", "联盟"}
        assert {term["source_type"] for term in project_terms} == {"imported"}
        assert {term["target_alt"] for term in project_terms} == {"Top Commander", "Guild"}
        assert {term["category"] for term in project_terms} == {"title", "system"}

        manual_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "term_key": "M-1",
                "source": "战机",
                "target": "Warplane",
                "target_alt": "Fighter",
                "category": "unit",
                "note": "manual term",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert manual_response.status_code == 200
        manual = manual_response.json()
        update_response = client.patch(
            f"/api/projects/{project['id']}/glossary/{manual['id']}",
            json={"target": "Fighter Jet", "note": "edited term"},
        )
        assert update_response.status_code == 200

        export_response = client.get(f"/api/projects/{project['id']}/glossary/export?format=json")
        assert export_response.status_code == 200
        exported = export_response.json()
        assert len(exported["terms"]) == 3
        assert all("source_type" not in term and "confirmed" not in term for term in exported["terms"])
        assert any(term["source"] == "战机" and term["target"] == "Fighter Jet" and term["note"] == "edited term" for term in exported["terms"])
        xlsx_response = client.get(f"/api/projects/{project['id']}/glossary/export?format=xlsx")
        assert xlsx_response.status_code == 200
        assert "spreadsheetml" in xlsx_response.headers["content-type"]
        exported_xlsx = tmp_path / "exported_terms.xlsx"
        exported_xlsx.write_bytes(xlsx_response.content)
        wb = load_workbook(exported_xlsx, read_only=True)
        try:
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            assert headers == ["ID", "CN", "EN", "EN2", "分类", "备注"]
        finally:
            wb.close()


def test_glossary_manual_save_replaces_duplicate_cn() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Save Dedup", "type": "QA"}).json()
        created = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "term_key": "28",
                "source": "消灭怪物",
                "target": "",
                "target_alt": "",
                "category": "",
                "note": "manual candidate",
                "source_type": "manual",
                "confirmed": True,
            },
        ).json()
        db.insert_glossary_term(
            project["id"],
            {
                "term_key": "28",
                "source": " 消 灭 怪 物 ",
                "target": "Destroy Monsters",
                "target_alt": "",
                "category": "generated",
                "note": "generated duplicate",
                "source_type": "generated",
                "confirmed": False,
            },
        )

        update_response = client.patch(
            f"/api/projects/{project['id']}/glossary/{created['id']}",
            json={"source": "消灭怪物", "target": "Defeat Monsters", "category": "任务目标", "note": "saved edit"},
        )
        assert update_response.status_code == 200
        assert update_response.json()["id"] == created["id"]

        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        matching = [term for term in terms if re.sub(r"\s+", "", term["source"]) == "消灭怪物"]
        assert len(matching) == 1
        assert matching[0]["id"] == created["id"]
        assert matching[0]["target"] == "Defeat Monsters"
        assert matching[0]["category"] == "任务目标"
        assert matching[0]["note"] == "saved edit"


def test_glossary_manual_add_upserts_existing_cn() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Add Upsert", "type": "QA"}).json()
        first = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "1", "source": "钻石", "target": "Diamonds", "category": "资源", "source_type": "manual", "confirmed": True},
        ).json()

        second_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "1", "source": " 钻 石 ", "target": "Gems", "target_alt": "Diamonds", "category": "货币", "source_type": "manual", "confirmed": True},
        )
        assert second_response.status_code == 200
        assert second_response.json()["id"] == first["id"]

        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert len([term for term in terms if re.sub(r"\s+", "", term["source"]) == "钻石"]) == 1
        assert terms[0]["target"] == "Gems"
        assert terms[0]["target_alt"] == "Diamonds"
        assert terms[0]["category"] == "货币"


def test_duplicate_project_name_returns_existing_project() -> None:
    with TestClient(app) as client:
        first = client.post("/api/projects", json={"name": "明日2", "type": "科幻 SLG"}).json()
        second_response = client.post("/api/projects", json={"name": " 明日2 ", "type": "其他"})

        assert second_response.status_code == 200
        second = second_response.json()
        assert second["id"] == first["id"]
        assert second["type"] == "科幻 SLG"
        assert second["duplicate"] is True
        matching = [project for project in client.get("/api/projects").json() if project["name"] == "明日2"]
        assert len(matching) == 1


def test_glossary_backfill_preserves_existing_terms_and_logs_strategy(tmp_path: Path) -> None:
    generated = tmp_path / "generated_glossary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "战机", "Warplane", "Fighter"])
    ws.append(["G-2", "钻石", "Diamonds", "Gems"])
    ws.append(["G-3", "能量", "", ""])
    wb.save(generated)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Backfill", "type": "QA"}).json()
        client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "战机", "target": "Manual Warplane", "target_alt": "", "source_type": "manual", "confirmed": True},
        ).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()

        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])

        assert result["candidates"] == 3
        assert result["unique_candidates"] == 3
        assert result["inserted"] == 2
        assert result["updated"] == 0
        assert result["pending_confirmation"] == 2
        assert result["skipped_existing"] == 1
        assert result["skipped_duplicate"] == 0
        assert result["conflicts"] == 0
        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert len(terms) == 1
        assert terms[0]["source"] == "战机"
        assert terms[0]["target"] == "Manual Warplane"
        assert terms[0]["target_alt"] == ""
        batches = client.get(f"/api/projects/{project['id']}/glossary/batches").json()
        assert batches["active_batch"]["id"] == result["batch_id"]
        assert batches["active_batch"]["counts"]["pending"] == 2
        candidates = {candidate["source"]: candidate for candidate in batches["candidates"]}
        assert "战机" not in candidates
        assert candidates["钻石"]["action"] == "new"
        assert candidates["能量"]["note"].endswith("人工确认")

        reject_response = client.post(
            f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/reject",
            json={"candidate_ids": [candidates["能量"]["id"]]},
        )
        assert reject_response.status_code == 200
        accept_response = client.post(
            f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/accept",
            json={"candidate_ids": [candidates["钻石"]["id"]]},
        )
        assert accept_response.status_code == 200
        accepted_terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        accepted_by_source = {term["source"]: term for term in accepted_terms}
        assert accepted_by_source["战机"]["target_alt"] == ""
        assert accepted_by_source["钻石"]["target"] == "Diamonds"
        assert "能量" not in accepted_by_source
        events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("Glossary backfill strategy" in event["message"] for event in events)
        assert any("inserted=2" in event["message"] and "updated=0" in event["message"] for event in events)


def test_glossary_backfill_dedupes_generated_terms_by_cn(tmp_path: Path) -> None:
    generated = tmp_path / "generated_glossary_duplicates.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "能量", "", ""])
    ws.append(["G-2", " 能 量 ", "Energy", "Power"])
    wb.save(generated)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Dedup", "type": "QA"}).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()

        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])

        assert result["candidates"] == 2
        assert result["unique_candidates"] == 1
        assert result["skipped_duplicate"] == 1
        assert result["inserted"] == 1
        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert terms == []
        batches = client.get(f"/api/projects/{project['id']}/glossary/batches").json()
        assert len(batches["candidates"]) == 1
        assert batches["candidates"][0]["source"] == "能量"
        assert batches["candidates"][0]["target"] == "Energy"
        assert batches["candidates"][0]["target_alt"] == "Power"


def test_glossary_extract_uses_project_materials_for_brief_and_prompt(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)
    material = tmp_path / "aircraft_context.md"
    material.write_text("Project setting: aircraft, fighter jet, missile, shooter, hero gear progression.", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Material Brief", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            source_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        with material.open("rb") as fh:
            material_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("aircraft_context.md", fh, "text/markdown")},
            )
        source_artifact = source_response.json()
        material_artifact = material_response.json()

        extract_response = client.post(
            f"/api/projects/{project['id']}/glossary/extract",
            json={
                "input_artifact_id": source_artifact["id"],
                "project_name": "Material Brief",
                "id_column": "ID",
                "source_column": "cn",
                "target_column": "en",
                "project_material_artifact_ids": [material_artifact["id"]],
                "project_notes": ["Screenshot shows a dark sci-fi hangar and missile battle UI."],
            },
        )
        assert extract_response.status_code == 200, extract_response.text
        artifacts = {artifact["kind"]: artifact for artifact in extract_response.json()["artifacts"]}
        brief = Path(artifacts["project_brief"]["path"]).read_text(encoding="utf-8")
        prompt = Path(artifacts["translation_prompt"]["path"]).read_text(encoding="utf-8")
        assert "aircraft_context.md" in brief
        assert "\u79d1\u5e7b\u519b\u4e8b" in prompt


def test_existing_translation_workbook_can_run_qa_without_translation_workpack(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Direct QA", "type": "QA"}).json()
        term_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "开始游戏", "target": "Start Game", "target_alt": "Begin", "category": "ui", "source_type": "manual"},
        )
        assert term_response.status_code == 200
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        translated_artifact = upload_response.json()
        assert translated_artifact["role"] == "translation_workbook"

        run_response = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": translated_artifact["id"],
            },
        )
        assert run_response.status_code == 200
        qa_response = client.post(f"/api/runs/{run_response.json()['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text
        result = qa_response.json()
        assert result["run"]["status"] == "passed"
        assert result["quality_summary"]["sources"]["translation_workbook"] == translated_artifact["id"]
        snapshot_id = result["run"]["metadata"]["input_artifacts"]["glossary_snapshot"]
        kinds = {artifact["kind"] for artifact in result["artifacts"]}
        assert "quality_summary" in kinds
        assert "qa_changes" in kinds
        assert "qa_final_workbook" in kinds
        assert "translation_workpack" not in kinds
        snapshot_artifact = next(artifact for artifact in client.get(f"/api/projects/{project['id']}").json()["artifacts"] if artifact["id"] == snapshot_id)
        snapshot_wb = load_workbook(Path(snapshot_artifact["path"]), read_only=True, data_only=True)
        try:
            rows = list(snapshot_wb.active.iter_rows(values_only=True))
            assert any(row[1] == "开始游戏" and row[2] == "Start Game" for row in rows[1:])
        finally:
            snapshot_wb.close()
        events = client.get(f"/api/runs/{run_response.json()['id']}/events").json()
        assert any("--term-base" in event["message"] and "project_glossary_snapshot.xlsx" in event["message"] for event in events)
        project_detail = client.get(f"/api/projects/{project['id']}").json()
        assert int(project_detail["stats"]["words"]) > 0
        assert project_detail["stats"]["langs"] == 1
        archived = client.get(f"/api/projects/{project['id']}/translations").json()
        assert len(archived) == 5
        assert archived[0]["entry_key"] == "1"
        assert archived[0]["target"] == "Claim Rewards"


def test_language_source_with_existing_translations_can_run_direct_qa(tmp_path: Path) -> None:
    workbook = tmp_path / "translated_language_table.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Language Source QA", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("translated_language_table.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        artifact = upload_response.json()
        assert artifact["role"] == "language_source"

        readiness = client.get(f"/api/artifacts/{artifact['id']}/translation-readiness?batch_size=90").json()
        assert readiness["ready_for_qa"] is True

        run_response = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": artifact["id"],
            },
        )
        assert run_response.status_code == 200
        qa_response = client.post(f"/api/runs/{run_response.json()['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text
        result = qa_response.json()
        assert result["run"]["status"] == "passed"
        assert result["quality_summary"]["sources"]["translation_workbook"] == artifact["id"]
        assert result["run"]["metadata"]["input_artifacts"]["translation_workbook"] == artifact["id"]
        archived = client.get(f"/api/projects/{project['id']}/translations").json()
        assert len(archived) == 5
        assert archived[1]["entry_key"] == "2"


def test_translation_archive_import_edit_and_export(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Archive", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        imported = client.post(f"/api/projects/{project['id']}/translations/import", json={"artifact_id": artifact["id"]}).json()
        assert imported["imported_count"] == 5
        entries = client.get(f"/api/projects/{project['id']}/translations").json()
        assert [entry["entry_key"] for entry in entries[:3]] == ["1", "2", "3"]

        updated = client.patch(
            f"/api/projects/{project['id']}/translations/{entries[0]['id']}",
            json={"target": "Claim"},
        ).json()
        assert updated["target"] == "Claim"
        export_json = client.get(f"/api/projects/{project['id']}/translations/export?format=json").json()
        assert export_json["entries"][0]["target"] == "Claim"
        export_xlsx = client.get(f"/api/projects/{project['id']}/translations/export?format=xlsx")
        assert export_xlsx.status_code == 200
        assert export_xlsx.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")


def test_delivery_package_contains_only_task_outputs(tmp_path: Path) -> None:
    terms = tmp_path / "terms.xlsx"
    workbook = tmp_path / "translated.xlsx"
    _sample_term_workbook(terms)
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "小小战机", "type": "飞行射击", "description": "来源：测试文件"}).json()
        client.patch(f"/api/projects/{project['id']}", json={"prompt_text": "项目提示词：准确翻译，术语以项目术语表为准。"})
        with terms.open("rb") as fh:
            term_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        import_response = client.post(f"/api/projects/{project['id']}/glossary/import", json={"artifact_id": term_artifact["id"]})
        assert import_response.status_code == 200
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text

        deliverables_response = client.get(f"/api/projects/{project['id']}/deliverables")
        assert deliverables_response.status_code == 200
        deliverables = deliverables_response.json()["deliverables"]
        assert len(deliverables) == 1
        assert deliverables[0]["task_label"] == f"QA-{run['id'].replace('run_', '')[:6]}"
        assert deliverables[0]["processed_rows"] == 5
        assert deliverables[0]["provider"] == "rules-only"
        assert deliverables[0]["model"] == "-"
        assert deliverables[0]["status"] == "passed"

        package_response = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={run['id']}")
        assert package_response.status_code == 200, package_response.text
        package = package_response.json()
        filenames = [item["filename"] for item in package["files"]]
        assert len(filenames) == 2
        assert re.fullmatch(r"小小战机_EN_\d{12}_QA-[0-9a-f]{6}_final\.xlsx", filenames[0])
        assert re.fullmatch(r"小小战机_EN_\d{12}_QA-[0-9a-f]{6}_changes\.xlsx", filenames[1])
        assert not any(
            "input_copy" in filename
            or "manifest" in filename
            or "jsonl" in filename
            or "project_meta" in filename
            or "translation_prompt" in filename
            or "glossary" in filename
            for filename in filenames
        )
        for item in package["files"]:
            assert Path(item["path"]).exists()
            assert item["download_url"].endswith(item["filename"])
        refreshed = client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"][0]
        assert refreshed["files"]["final"]["download_url"].endswith("_final.xlsx")
        assert refreshed["files"]["changes"]["download_url"].endswith("_changes.xlsx")


def test_delivery_filename_sanitizes_invalid_project_name_without_double_spaces(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "E2E ?? Demo", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        response = client.post(f"/api/runs/{run['id']}/qa")
        assert response.status_code == 200
        package = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={run['id']}").json()
        filename = package["files"][0]["filename"]
        assert "??" not in filename
        assert "  " not in filename
        assert filename.startswith("E2E Demo_EN_")


def test_qa_continuation_inherits_translation_delivery_identity(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "继承任务", "type": "QA"}).json()
        source_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "task_code": "A"},
        ).json()
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        qa_run = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": translated_artifact["id"],
                "task_origin": "translation_continuation",
                "source_run_id": source_run["id"],
                "task_code": "QA",
            },
        ).json()
        qa_response = client.post(f"/api/runs/{qa_run['id']}/qa")
        assert qa_response.status_code == 200

        deliverables = client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"]
        assert deliverables[0]["task_label"] == f"A-{source_run['id'].replace('run_', '')[:6]}"
        package = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={qa_run['id']}").json()
        assert re.fullmatch(r"继承任务_EN_\d{12}_A-[0-9a-f]{6}_final\.xlsx", package["files"][0]["filename"])


def test_failed_runs_are_not_deliverable(tmp_path: Path) -> None:
    workbook = tmp_path / "failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "不可交付", "type": "QA"}).json()
        client.patch(f"/api/projects/{project['id']}/harness", json={"forbidden_translations": ["Forbidden Brand"]})
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200
        assert qa_response.json()["run"]["status"] == "failed"
        assert client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"] == []


def test_failed_qa_exposes_normalized_project_harness_rows(tmp_path: Path) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Failed Row QA", "type": "QA"}).json()
        harness_response = client.patch(
            f"/api/projects/{project['id']}/harness",
            json={"forbidden_translations": ["Forbidden Brand"]},
        )
        assert harness_response.status_code == 200
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        translated_artifact = upload_response.json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"]},
        ).json()

        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200
        assert qa_response.json()["run"]["status"] == "failed"

        issues_response = client.get(f"/api/runs/{run['id']}/quality-issues")
        assert issues_response.status_code == 200
        issues = issues_response.json()["issues"]
        assert issues[0]["source"] == "project_harness"
        assert issues[0]["rule_source"] == "project_harness"
        assert issues[0]["severity"] == "hard"
        assert issues[0]["sheet"] == "Language"
        assert issues[0]["row"] == 2
        assert issues[0]["current_translation"] == "Forbidden Brand Reward"


def test_manual_fix_creates_fixed_workbook_reruns_qa_and_updates_project_harness(tmp_path: Path) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Manual Fix QA", "type": "QA"}).json()
        client.patch(
            f"/api/projects/{project['id']}/harness",
            json={"forbidden_translations": ["Forbidden Brand"]},
        )
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        failed_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"]},
        ).json()
        failed_qa = client.post(f"/api/runs/{failed_run['id']}/qa").json()
        assert failed_qa["run"]["status"] == "failed"

        fix_response = client.post(
            f"/api/runs/{failed_run['id']}/manual-fixes",
            json={
                "fixes": [
                    {
                        "sheet": "Language",
                        "row": 2,
                        "translation": "Reward",
                        "note": "Remove forbidden project-specific wording.",
                    }
                ],
                "rerun_qa": True,
            },
        )
        assert fix_response.status_code == 200, fix_response.text
        result = fix_response.json()
        assert result["fixed_artifact"]["role"] == "translation_workbook"
        assert result["fixed_artifact"]["origin"] == "manual"
        assert result["qa_result"]["run"]["status"] == "passed"
        assert result["qa_result"]["quality_summary"]["sources"]["manual_fix_source_run"] == failed_run["id"]
        assert any(artifact["kind"] == "qa_changes" for artifact in result["qa_result"]["artifacts"])

        harness = client.get(f"/api/projects/{project['id']}/harness").json()["project_harness"]
        assert harness["manual_fixes"][0]["previous_translation"] == "Forbidden Brand Reward"
        assert harness["manual_fixes"][0]["translation"] == "Reward"

        suggestions = client.get(f"/api/projects/{project['id']}/improvements").json()
        assert any(item["category"] == "project_harness" and item["run_id"] == failed_run["id"] for item in suggestions)


def test_model_fix_requires_configured_provider(tmp_path: Path) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        client.patch("/api/settings", json={"provider": "mock", "api_key": ""})
        project = client.post("/api/projects", json={"name": "Model Fix Needs Key", "type": "QA"}).json()
        client.patch(f"/api/projects/{project['id']}/harness", json={"forbidden_translations": ["Forbidden Brand"]})
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        failed_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"]},
        ).json()
        assert client.post(f"/api/runs/{failed_run['id']}/qa").json()["run"]["status"] == "failed"

        response = client.post(f"/api/runs/{failed_run['id']}/model-fixes", json={"max_issues": 20, "rerun_qa": True})
        assert response.status_code == 409
        assert "API key" in response.json()["detail"]


def test_model_fix_applies_provider_suggestions_and_reruns_qa(tmp_path: Path, monkeypatch) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    def fake_model(settings: dict, prompt: str) -> str:
        if "待修复行" not in prompt:
            return '{"passed": true, "issues": []}'
        return '{"fixes":[{"issue_id":"project_harness:0:Language:2:forbidden_translation","sheet":"Language","row":2,"translation":"Reward","note":"remove forbidden phrase"}]}'

    monkeypatch.setattr(workflow, "_call_semantic_provider", fake_model)
    with TestClient(app) as client:
        client.patch("/api/settings", json={"provider": "openai", "api_key": "test-key", "model": "gpt-test"})
        project = client.post("/api/projects", json={"name": "Model Fix QA", "type": "QA"}).json()
        client.patch(f"/api/projects/{project['id']}/harness", json={"forbidden_translations": ["Forbidden Brand"]})
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        failed_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        assert client.post(f"/api/runs/{failed_run['id']}/qa").json()["run"]["status"] == "failed"

        response = client.post(f"/api/runs/{failed_run['id']}/model-fixes", json={"max_issues": 20, "rerun_qa": True})
        assert response.status_code == 200, response.text
        result = response.json()
        assert result["fixed_artifact"]["origin"] == "provider"
        assert result["model_fixes"][0]["translation"] == "Reward"
        assert result["model_fixes"][0]["rule_source"] == "model_fix"
        assert result["qa_result"]["run"]["status"] == "passed"
        assert result["qa_result"]["quality_summary"]["sources"]["model_fix_source_run"] == failed_run["id"]


def test_project_harness_is_project_scoped_and_affects_only_its_run(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        first = client.post(
            "/api/projects",
            json={"name": "Harness A", "type": "科幻 SLG", "icon": "A", "description": "Project A."},
        ).json()
        second = client.post(
            "/api/projects",
            json={"name": "Harness B", "type": "科幻 SLG", "icon": "B", "description": "Project B."},
        ).json()

        harness_response = client.patch(
            f"/api/projects/{first['id']}/harness",
            json={
                "style_guidance": "Use tactical aviation wording for this project only.",
                "target_audience": "Core strategy players",
                "forbidden_translations": ["Mock 1"],
                "hard_rules": [{"label": "No raw mock marker", "description": "Mock marker must not ship."}],
            },
        )
        assert harness_response.status_code == 200
        assert harness_response.json()["project_harness"]["style_guidance"].startswith("Use tactical")

        first_result = _run_mock_translation(client, first["id"], workbook)
        second_result = _run_mock_translation(client, second["id"], workbook)

        assert first_result["run"]["status"] == "failed"
        assert first_result["project_harness_quality"]["hard_errors"] == 1
        assert first_result["project_harness_quality"]["issues"][0]["rule_source"] == "project_harness"
        assert second_result["run"]["status"] == "passed"
        assert second_result["project_harness_quality"]["hard_errors"] == 0
        assert second_result["run"]["metadata"]["harness"]["forbidden_translations"] == 0

        semantic_response = client.post(f"/api/runs/{second_result['run']['id']}/semantic-qa")
        assert semantic_response.status_code == 200
        assert semantic_response.json()["semantic_qa"]["status"] == "needs_model_review"
        assert semantic_response.json()["artifact"]["kind"] == "semantic_qa_context"


def test_improvement_review_writes_suggestions_without_auto_merge(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project = client.post(
            "/api/projects",
            json={"name": "Harness Improvement", "type": "科幻 SLG", "icon": "I", "description": "Project."},
        ).json()
        client.patch(
            f"/api/projects/{project['id']}/harness",
            json={"forbidden_translations": ["Mock 1"]},
        )
        result = _run_mock_translation(client, project["id"], workbook)

        review_response = client.post(f"/api/runs/{result['run']['id']}/improvement-review")
        assert review_response.status_code == 200
        categories = {item["category"] for item in review_response.json()["suggestions"]}
        assert {"project_harness", "upstream_backfeed"}.issubset(categories)

        queue_response = client.get(f"/api/projects/{project['id']}/improvements")
        assert queue_response.status_code == 200
        assert all(item["status"] == "pending_review" for item in queue_response.json())


def _run_mock_translation(client: TestClient, project_id: str, workbook: Path) -> dict:
    with workbook.open("rb") as fh:
        upload_response = client.post(
            f"/api/projects/{project_id}/files?kind=language_table",
            files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload_response.status_code == 200
    source_artifact = upload_response.json()
    run_response = client.post(
        "/api/runs",
        json={
            "project_id": project_id,
            "kind": "translation",
            "language": "en",
            "input_artifact_id": source_artifact["id"],
            "batch_size": 3,
        },
    )
    assert run_response.status_code == 200
    run = run_response.json()
    translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "mock", "allow_mock": True})
    assert translate_response.status_code == 200, translate_response.text
    return translate_response.json()
