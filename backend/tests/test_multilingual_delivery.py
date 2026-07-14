from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

os.environ.setdefault("LWS_DATA_ROOT", str(Path(tempfile.gettempdir()) / "lws-test-data"))

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
from app.config import DEFAULT_SETTINGS, save_settings
from app.main import app
from conftest import reset_data_root, wait_for_background_jobs


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    reset_data_root(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    wait_for_background_jobs()
    save_settings(DEFAULT_SETTINGS)


def _write_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN", "KR"])
    ws.append([1, "开始游戏", "", ""])
    ws.append([2, "领取奖励", "", ""])
    wb.save(path)
    wb.close()


def _write_final(path: Path, language_header: str, values: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", language_header])
    ws.append([1, "开始游戏", values[0]])
    ws.append([2, "领取奖励", values[1]])
    wb.save(path)
    wb.close()


def _write_workbook_with_legacy_en2(path: Path, *, include_kr: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    headers = ["ID", "CN", "EN", "EN2", *(["KR"] if include_kr else [])]
    ws.append(headers)
    ws.append([1, "开始游戏", "Start Game", "Begin Game", *([""] if include_kr else [])])
    ws.append([2, "领取奖励", "Claim Rewards", "Collect Rewards", *([""] if include_kr else [])])
    wb.save(path)
    wb.close()


def _add_passed_final(
    project_id: str,
    source_id: str,
    language: str,
    final_path: Path,
    header: str,
    values: list[str],
    translation_task_id: str | None = None,
) -> dict:
    _write_final(final_path, header, values)
    run = db.insert_run(
        project_id,
        "translation",
        language,
        metadata={
            "input_artifact_id": source_id,
            "parent_input_artifact_id": source_id,
            "task_origin": "translation_run",
            "translation_task_id": translation_task_id,
            "quality_summary": {"passed": True, "hard_errors": 0},
        },
    )
    db.update_run(run["id"], status="passed", metadata={**run["metadata"], "quality_summary": {"passed": True, "hard_errors": 0}})
    db.add_artifact(project_id, f"final {language}", final_path, "qa_final_workbook", run_id=run["id"])
    return db.get_run(run["id"])


def test_merged_delivery_combines_passed_language_outputs(tmp_path: Path) -> None:
    project = db.insert_project("merged delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "en", tmp_path / "en.xlsx", "EN", ["Start Game", "Claim Rewards"])
    _add_passed_final(project["id"], source["id"], "ko", tmp_path / "kr.xlsx", "KR", ["게임 시작", "보상 받기"])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en", "ko"]},
        )

    assert response.status_code == 200
    payload = response.json()
    final_file = next(item for item in payload["files"] if item["kind"] == "merged_final")
    wb = load_workbook(final_file["path"], data_only=True)
    try:
        ws = wb.active
        assert [cell.value for cell in ws[1]][:4] == ["ID", "CN", "EN", "KR"]
        assert ws.cell(row=2, column=3).value == "Start Game"
        assert ws.cell(row=2, column=4).value == "게임 시작"
        assert ws.cell(row=3, column=3).value == "Claim Rewards"
        assert ws.cell(row=3, column=4).value == "보상 받기"
    finally:
        wb.close()


def test_merged_delivery_removes_legacy_en2_column(tmp_path: Path) -> None:
    project = db.insert_project("merged delivery without en2", "QA", "")
    source_path = tmp_path / "source-with-en2.xlsx"
    _write_workbook_with_legacy_en2(source_path, include_kr=True)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "en", tmp_path / "en.xlsx", "EN", ["Start Game", "Claim Rewards"])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en"]},
        )

    assert response.status_code == 200
    final_file = next(item for item in response.json()["files"] if item["kind"] == "merged_final")
    workbook = load_workbook(final_file["path"], read_only=True, data_only=True)
    try:
        headers = [cell.value for cell in workbook.active[1]]
        assert headers == ["ID", "CN", "EN", "KR"]
        assert workbook.active.cell(2, 3).value == "Start Game"
    finally:
        workbook.close()


def test_korean_merged_delivery_normalizes_generic_target_header_without_en2(tmp_path: Path) -> None:
    project = db.insert_project("korean merged delivery without en2", "QA", "")
    source_path = tmp_path / "korean-source-with-generic-target.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Language"
    worksheet.append(["ID", "CN", "target", "EN2"])
    worksheet.append([1, "开始游戏", "", "Begin Game"])
    worksheet.append([2, "领取奖励", "", "Collect Rewards"])
    workbook.save(source_path)
    workbook.close()
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "ko", tmp_path / "kr.xlsx", "KR", ["게임 시작", "보상 받기"])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["ko"]},
        )

    assert response.status_code == 200, response.text
    final_file = next(item for item in response.json()["files"] if item["kind"] == "merged_final")
    delivered = load_workbook(final_file["path"], read_only=True, data_only=True)
    try:
        assert [cell.value for cell in delivered.active[1]] == ["ID", "CN", "KR"]
        assert delivered.active.cell(2, 3).value == "게임 시작"
    finally:
        delivered.close()


def test_merged_delivery_uses_only_requested_translation_task(tmp_path: Path) -> None:
    project = db.insert_project("task isolated delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    task_a_run = _add_passed_final(
        project["id"],
        source["id"],
        "en",
        tmp_path / "task-a.xlsx",
        "EN",
        ["Task A Start", "Task A Reward"],
        translation_task_id="task-a",
    )
    task_b_run = _add_passed_final(
        project["id"],
        source["id"],
        "en",
        tmp_path / "task-b.xlsx",
        "EN",
        ["Task B Start", "Task B Reward"],
        translation_task_id="task-b",
    )

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={
                "input_artifact_id": source["id"],
                "languages": ["en"],
                "translation_task_id": "task-a",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    final_file = next(item for item in payload["files"] if item["kind"] == "merged_final")
    workbook = load_workbook(final_file["path"], read_only=True, data_only=True)
    try:
        assert workbook.active.cell(2, 3).value == "Task A Start"
        assert workbook.active.cell(3, 3).value == "Task A Reward"
    finally:
        workbook.close()
    assert payload["deliverable"]["translation_task_id"] == "task-a"
    assert db.get_run(task_a_run["id"])["metadata"]["translation_task_state"] == "delivered"
    assert "translation_task_state" not in db.get_run(task_b_run["id"])["metadata"]


def test_same_minute_merged_deliveries_keep_distinct_task_files(tmp_path: Path) -> None:
    project = db.insert_project("same minute task delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(
        project["id"], source["id"], "en", tmp_path / "task-a.xlsx", "EN",
        ["Task A Start", "Task A Reward"], translation_task_id="task-a",
    )
    _add_passed_final(
        project["id"], source["id"], "en", tmp_path / "task-b.xlsx", "EN",
        ["Task B Start", "Task B Reward"], translation_task_id="task-b",
    )

    with TestClient(app) as client:
        first = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en"], "translation_task_id": "task-a"},
        ).json()
        second = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en"], "translation_task_id": "task-b"},
        ).json()

    first_path = Path(next(item for item in first["files"] if item["kind"] == "merged_final")["path"])
    second_path = Path(next(item for item in second["files"] if item["kind"] == "merged_final")["path"])
    assert first_path != second_path
    workbook = load_workbook(first_path, read_only=True, data_only=True)
    try:
        assert workbook.active.cell(2, 3).value == "Task A Start"
        assert workbook.active.cell(3, 3).value == "Task A Reward"
    finally:
        workbook.close()


def test_repeated_same_task_merged_deliveries_keep_distinct_files(tmp_path: Path) -> None:
    project = db.insert_project("repeated task delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(
        project["id"], source["id"], "en", tmp_path / "task-a.xlsx", "EN",
        ["Task A Start", "Task A Reward"], translation_task_id="task-a",
    )

    with TestClient(app) as client:
        request = {
            "input_artifact_id": source["id"],
            "languages": ["en"],
            "translation_task_id": "task-a",
        }
        first = client.post(f"/api/projects/{project['id']}/delivery-package/merged", json=request).json()
        second = client.post(f"/api/projects/{project['id']}/delivery-package/merged", json=request).json()

    first_path = Path(next(item for item in first["files"] if item["kind"] == "merged_final")["path"])
    second_path = Path(next(item for item in second["files"] if item["kind"] == "merged_final")["path"])
    assert first_path != second_path
    assert first_path.exists()
    assert second_path.exists()


def test_single_delivery_writes_readback_gate_artifact(tmp_path: Path) -> None:
    project = db.insert_project("single delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    run = _add_passed_final(
        project["id"], source["id"], "en", tmp_path / "en.xlsx", "EN",
        ["Start Game", "Claim Rewards"], translation_task_id="single-task",
    )

    with TestClient(app) as client:
        response = client.post(f"/api/projects/{project['id']}/delivery-package", params={"run_id": run["id"]})

    assert response.status_code == 200
    payload = response.json()
    # Delivery files stay "final files + QA summary" only; the readback gate
    # result is stored as a workbench artifact, not packed into delivery.
    assert all(item["kind"] != "readback_gate" for item in payload["files"])
    gate_artifacts = [item for item in db.list_artifacts(run_id=run["id"]) if item["kind"] == "delivery_readback_gate"]
    assert gate_artifacts
    with open(gate_artifacts[0]["path"], encoding="utf-8") as handle:
        gate_result = json.load(handle)
    assert gate_result["readback_verified"] is True
    assert gate_result["hard_blockers"] == 0
    assert "delivery" not in Path(gate_artifacts[0]["path"]).parent.parts
    assert db.get_run(run["id"])["metadata"]["translation_task_state"] == "delivered"


def test_single_delivery_removes_legacy_en2_column(tmp_path: Path) -> None:
    project = db.insert_project("single delivery without en2", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    final_path = tmp_path / "final-with-en2.xlsx"
    _write_workbook_with_legacy_en2(final_path)
    run = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={
            "input_artifact_id": source["id"],
            "parent_input_artifact_id": source["id"],
            "task_origin": "translation_run",
            "quality_summary": {"passed": True, "hard_errors": 0},
        },
    )
    db.update_run(run["id"], status="passed", metadata={**run["metadata"], "quality_summary": {"passed": True, "hard_errors": 0}})
    db.add_artifact(project["id"], "final en", final_path, "qa_final_workbook", run_id=run["id"])

    with TestClient(app) as client:
        response = client.post(f"/api/projects/{project['id']}/delivery-package", params={"run_id": run["id"]})

    assert response.status_code == 200
    final_file = next(item for item in response.json()["files"] if item["kind"] == "final")
    workbook = load_workbook(final_file["path"], read_only=True, data_only=True)
    try:
        assert [cell.value for cell in workbook.active[1]] == ["ID", "CN", "EN"]
        assert workbook.active.cell(2, 3).value == "Start Game"
    finally:
        workbook.close()


def test_single_delivery_keeps_non_english_target_alt_column(tmp_path: Path) -> None:
    project = db.insert_project("korean delivery keeps generic alt", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    final_path = tmp_path / "ko-with-target-alt.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Language"
    worksheet.append(["ID", "CN", "KR", "target_alt"])
    worksheet.append([1, "开始游戏", "게임 시작", "게임 개시"])
    worksheet.append([2, "领取奖励", "보상 받기", "보상 수령"])
    workbook.save(final_path)
    workbook.close()
    run = db.insert_run(
        project["id"],
        "translation",
        "ko",
        metadata={
            "input_artifact_id": source["id"],
            "parent_input_artifact_id": source["id"],
            "task_origin": "translation_run",
            "quality_summary": {"passed": True, "hard_errors": 0},
        },
    )
    db.update_run(run["id"], status="passed", metadata={**run["metadata"], "quality_summary": {"passed": True, "hard_errors": 0}})
    db.add_artifact(project["id"], "final ko", final_path, "qa_final_workbook", run_id=run["id"])

    with TestClient(app) as client:
        response = client.post(f"/api/projects/{project['id']}/delivery-package", params={"run_id": run["id"]})

    assert response.status_code == 200
    final_file = next(item for item in response.json()["files"] if item["kind"] == "final")
    delivered = load_workbook(final_file["path"], read_only=True, data_only=True)
    try:
        assert [cell.value for cell in delivered.active[1]] == ["ID", "CN", "KR", "target_alt"]
        assert delivered.active.cell(2, 4).value == "게임 개시"
    finally:
        delivered.close()


def test_single_delivery_blocks_when_final_workbook_has_blank_target_cell(tmp_path: Path) -> None:
    project = db.insert_project("single delivery blocked", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    final_path = tmp_path / "en.xlsx"
    _write_final(final_path, "EN", ["Start Game", ""])
    run = db.insert_run(
        project["id"],
        "translation",
        "en",
        metadata={
            "input_artifact_id": source["id"],
            "parent_input_artifact_id": source["id"],
            "task_origin": "translation_run",
            "quality_summary": {"passed": True, "hard_errors": 0},
        },
    )
    db.update_run(run["id"], status="passed", metadata={**run["metadata"], "quality_summary": {"passed": True, "hard_errors": 0}})
    db.add_artifact(project["id"], "final en", final_path, "qa_final_workbook", run_id=run["id"])

    with TestClient(app) as client:
        response = client.post(f"/api/projects/{project['id']}/delivery-package", params={"run_id": run["id"]})

    assert response.status_code == 409


def test_merged_delivery_writes_readback_gate_artifact(tmp_path: Path) -> None:
    project = db.insert_project("merged delivery readback", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "en", tmp_path / "en.xlsx", "EN", ["Start Game", "Claim Rewards"])
    _add_passed_final(project["id"], source["id"], "ko", tmp_path / "kr.xlsx", "KR", ["게임 시작", "보상 받기"])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en", "ko"]},
        )

    assert response.status_code == 200
    payload = response.json()
    assert all(item["kind"] != "readback_gate" for item in payload["files"])
    merged_final = next(item for item in payload["files"] if item["kind"] == "merged_final")
    merged_artifact = db.get_artifact(merged_final["artifact_id"])
    gate_artifact = db.get_artifact(merged_artifact["metadata"]["readback_gate_artifact_id"])
    with open(gate_artifact["path"], encoding="utf-8") as handle:
        gate_result = json.load(handle)
    assert gate_result["readback_verified"] is True


def test_merged_delivery_generates_completed_languages_and_reports_skipped_ones(tmp_path: Path) -> None:
    project = db.insert_project("partial merged delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "en", tmp_path / "en.xlsx", "EN", ["Start Game", "Claim Rewards"])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en", "ko"]},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["merged_languages"] == ["EN"]
    assert payload["skipped_languages"] == ["KR"]
    assert {item["kind"] for item in payload["files"]} == {"merged_final", "qa_summary"}


def test_merged_delivery_skips_structurally_invalid_language_without_blocking_valid_ones(tmp_path: Path) -> None:
    project = db.insert_project("partial invalid merged delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "en", tmp_path / "en.xlsx", "EN", ["Start Game", "Claim Rewards"])
    _add_passed_final(project["id"], source["id"], "ko", tmp_path / "kr.xlsx", "KR", ["게임 시작", ""])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en", "ko"]},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["merged_languages"] == ["EN"]
    assert payload["skipped_languages"] == ["KR"]
    assert payload["language_results"][0]["status"] == "merged"
    assert payload["language_results"][1]["status"] == "skipped"
    assert "结构门禁未通过" in payload["language_results"][1]["reason"]
    with TestClient(app) as client:
        deliverables = client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"]
    merged_task = next(item for item in deliverables if item["task_code"] == "ALL")
    assert merged_task["input_artifact_id"] == source["id"]
    assert merged_task["merged_languages"] == ["EN"]
    assert merged_task["skipped_languages"] == ["KR"]
    assert merged_task["language_results"][1]["status"] == "skipped"
    assert merged_task["delivered_with_issues"] is True


def test_merged_delivery_reports_language_when_every_result_is_structurally_invalid(tmp_path: Path) -> None:
    project = db.insert_project("invalid merged delivery", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(project["id"], "source", source_path, "language_table")
    _add_passed_final(project["id"], source["id"], "ko", tmp_path / "kr.xlsx", "KR", ["게임 시작", ""])

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{project['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["ko"]},
        )

    assert response.status_code == 409
    assert "KR" in response.json()["detail"]
    assert "结构门禁未通过" in response.json()["detail"]
    assert not [item for item in db.list_artifacts(project_id=project["id"]) if item["kind"] == "merged_delivery_workbook"]


def test_merged_delivery_rejects_cross_project_source(tmp_path: Path) -> None:
    first = db.insert_project("A", "QA", "")
    second = db.insert_project("B", "QA", "")
    source_path = tmp_path / "source.xlsx"
    _write_source(source_path)
    source = db.add_artifact(first["id"], "source", source_path, "language_table")

    with TestClient(app) as client:
        response = client.post(
            f"/api/projects/{second['id']}/delivery-package/merged",
            json={"input_artifact_id": source["id"], "languages": ["en"]},
        )

    assert response.status_code == 409
