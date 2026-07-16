from __future__ import annotations

import importlib
import json
import os
from pathlib import Path

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.auth as auth
import app.db as db
import app.main as main_module
from app import route_capabilities
from conftest import reset_data_root

ADMIN_PASSWORD = "Initial-Admin-Password!"
USER_PASSWORD = "Sup3rSecret1!"

PROJECTS_URL = "/api/projects"
USERS_URL = "/api/users"
LOGIN_URL = "/api/auth/login"


def _build_app():
    return importlib.reload(main_module).app


_AUTH_ENV_VARS = ("LWS_AUTH_MODE", "LWS_DEPLOYMENT_MODE", "LWS_ADMIN_USER", "LWS_ADMIN_PASSWORD")


@pytest.fixture(autouse=True)
def reset_auth_environment(monkeypatch: pytest.MonkeyPatch) -> None:
    for name in _AUTH_ENV_VARS:
        monkeypatch.delenv(name, raising=False)
    reset_data_root(Path(os.environ["LWS_DATA_ROOT"]))
    auth.login_rate_limiter._state.clear()  # type: ignore[attr-defined]
    yield
    auth.login_rate_limiter._state.clear()  # type: ignore[attr-defined]
    # See test_users_admin.py's identical teardown comment: app.main.AUTH_REQUIRED
    # is a bare module global re-read by lifespan() on every call, not captured
    # per-app-instance, so a reload() here must be undone before the next test
    # file runs or it leaks "auth required" into unrelated test modules.
    for name in _AUTH_ENV_VARS:
        os.environ.pop(name, None)
    _build_app()


def _required_app(monkeypatch: pytest.MonkeyPatch, *, username: str = "root-admin"):
    monkeypatch.setenv("LWS_AUTH_MODE", "required")
    monkeypatch.setenv("LWS_ADMIN_USER", username)
    monkeypatch.setenv("LWS_ADMIN_PASSWORD", ADMIN_PASSWORD)
    return _build_app()


def _login(client: TestClient, username: str, password: str):
    response = client.post(LOGIN_URL, json={"username": username, "password": password})
    assert response.status_code == 200, response.text
    return response


def _bootstrap_admin_client(client: TestClient, *, username: str = "root-admin") -> None:
    _login(client, username, ADMIN_PASSWORD)
    admin = db.get_user_by_username(username)
    assert admin is not None
    db.update_user(admin["id"], {"must_change_password": False})


def _create_user_via_api(client: TestClient, username: str, role: str, *, password: str = USER_PASSWORD) -> dict:
    response = client.post(
        USERS_URL,
        json={"username": username, "display_name": username, "role": role, "initial_password": password},
    )
    assert response.status_code == 200, response.text
    return response.json()


def _clear_must_change_password(username: str) -> None:
    user = db.get_user_by_username(username)
    assert user is not None
    db.update_user(user["id"], {"must_change_password": False})


def _prepare_legacy_announcement(admin_client: TestClient, project_id: str, tmp_path: Path) -> dict:
    source_path = tmp_path / "legacy-announcement.txt"
    source_path.write_text("英雄公告\n", encoding="utf-8")
    with source_path.open("rb") as handle:
        source_artifact = admin_client.post(
            f"/api/projects/{project_id}/files?kind=asset",
            files={"file": (source_path.name, handle, "text/plain")},
        ).json()

    terms_path = tmp_path / "legacy-announcement-terms.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CN", "EN"])
    sheet.append(["英雄", "Hero"])
    workbook.save(terms_path)
    workbook.close()
    with terms_path.open("rb") as handle:
        terms_artifact = admin_client.post(
            f"/api/projects/{project_id}/files?kind=term_base",
            files={"file": (terms_path.name, handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        ).json()

    prepare_response = admin_client.post(
        f"/api/projects/{project_id}/announcement-docx/prepare",
        json={
            "source_artifact_ids": [source_artifact["id"]],
            "terms_artifact_id": terms_artifact["id"],
            "languages": ["en"],
        },
    )
    assert prepare_response.status_code == 200, prepare_response.text
    prepared = prepare_response.json()
    workpack = next(artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_docx_workpack")
    translation_workbook = next(
        artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_docx_translation_workbook"
    )
    rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]

    response_path = tmp_path / "legacy-announcement-response.jsonl"
    response_path.write_text(
        "\n".join(
            json.dumps({"para_id": row["para_id"], "translation": "Hero announcement"}, ensure_ascii=False)
            for row in rows
        )
        + "\n",
        encoding="utf-8",
    )
    with response_path.open("rb") as handle:
        response_artifact = admin_client.post(
            f"/api/projects/{project_id}/files?kind=asset",
            files={"file": (response_path.name, handle, "application/jsonl")},
        ).json()

    import_response = admin_client.post(
        f"/api/projects/{project_id}/announcement-docx/import-ai",
        json={
            "prepare_run_id": prepared["run"]["id"],
            "response_artifact_ids": [response_artifact["id"]],
            "languages": ["en"],
        },
    )
    assert import_response.status_code == 200, import_response.text
    apply_response = admin_client.post(
        f"/api/projects/{project_id}/announcement-docx/apply",
        json={
            "prepare_run_id": prepared["run"]["id"],
            "translation_workbook_artifact_id": translation_workbook["id"],
        },
    )
    assert apply_response.status_code == 200, apply_response.text
    return {
        "prepare_run_id": prepared["run"]["id"],
        "task_id": prepared["task"]["id"],
        "translation_workbook_artifact_id": translation_workbook["id"],
        "response_artifact_id": response_artifact["id"],
    }


# ---------------------------------------------------------------------------
# role -> capability matrix
# ---------------------------------------------------------------------------


def test_capability_matrix_matches_plan() -> None:
    from app import authz

    assert authz.capability_allowed("admin", authz.PROJECT_READ)
    assert authz.capability_allowed("admin", authz.ASSETS_CURATE)
    assert authz.capability_allowed("admin", authz.PROJECT_MANAGE)
    assert authz.capability_allowed("admin", authz.ADMIN)

    assert authz.capability_allowed("ops", authz.PROJECT_READ)
    assert authz.capability_allowed("ops", authz.TASK_RUN)
    assert authz.capability_allowed("ops", authz.ASSETS_CURATE)
    assert authz.capability_allowed("ops", authz.PROJECT_MANAGE)
    assert not authz.capability_allowed("ops", authz.ADMIN)

    assert authz.capability_allowed("member", authz.PROJECT_READ)
    assert authz.capability_allowed("member", authz.TASK_RUN)
    assert not authz.capability_allowed("member", authz.ASSETS_CURATE)
    assert not authz.capability_allowed("member", authz.PROJECT_MANAGE)
    assert not authz.capability_allowed("member", authz.ADMIN)


# ---------------------------------------------------------------------------
# member vs ops vs admin against real endpoints
# ---------------------------------------------------------------------------


def test_member_forbidden_on_assets_curate_endpoint_even_as_project_member(monkeypatch: pytest.MonkeyPatch) -> None:
    """ASSETS_CURATE (manual glossary term creation): member is a project
    member but still 403s -- capability, not membership, is the blocker.
    """
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project = admin_client.post(PROJECTS_URL, json={"name": "Capability Project", "type": "QA"}).json()
        _create_user_via_api(admin_client, "member-1", "member")
        _clear_must_change_password("member-1")
        member = db.get_user_by_username("member-1")
        db.add_project_member(project["id"], member["id"], added_by="root-admin")

    with TestClient(test_app) as client:
        _login(client, "member-1", USER_PASSWORD)
        response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "按钮", "target": "Button", "language": "en"},
        )
    assert response.status_code == 403, response.text
    assert response.json() == {"detail": "权限不足"}


def test_ops_allowed_on_assets_curate_endpoint_for_own_member_project(monkeypatch: pytest.MonkeyPatch) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project = admin_client.post(PROJECTS_URL, json={"name": "Ops Curate Project", "type": "QA"}).json()
        _create_user_via_api(admin_client, "ops-1", "ops")
        _clear_must_change_password("ops-1")
        ops_user = db.get_user_by_username("ops-1")
        db.add_project_member(project["id"], ops_user["id"], added_by="root-admin")

    with TestClient(test_app) as client:
        _login(client, "ops-1", USER_PASSWORD)
        response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "按钮", "target": "Button", "language": "en"},
        )
    assert response.status_code == 200, response.text
    assert response.json()["source"] == "按钮"


def test_member_allowed_on_task_run_endpoint(monkeypatch: pytest.MonkeyPatch) -> None:
    """TASK_RUN (creating a run): member has this capability and is a
    project member, so this must not be 401/403 -- we don't need the run to
    actually execute anything for this to prove the gate lets it through.
    """
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project = admin_client.post(PROJECTS_URL, json={"name": "Task Run Project", "type": "QA"}).json()
        _create_user_via_api(admin_client, "member-2", "member")
        _clear_must_change_password("member-2")
        member = db.get_user_by_username("member-2")
        db.add_project_member(project["id"], member["id"], added_by="root-admin")

    with TestClient(test_app) as client:
        _login(client, "member-2", USER_PASSWORD)
        response = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en"},
        )
    assert response.status_code not in (401, 403), response.text
    assert response.status_code == 200, response.text


def test_member_cannot_use_other_project_material_for_announcement_terms(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project_a = admin_client.post(PROJECTS_URL, json={"name": "Announcement Scope A", "type": "QA"}).json()
        project_b = admin_client.post(PROJECTS_URL, json={"name": "Announcement Scope B", "type": "QA"}).json()
        _create_user_via_api(admin_client, "announcement-member-1", "member")
        _clear_must_change_password("announcement-member-1")
        member = db.get_user_by_username("announcement-member-1")
        assert member is not None
        db.add_project_member(project_a["id"], member["id"], added_by="root-admin")

        foreign_material_path = tmp_path / "project-b-announcement.txt"
        foreign_material_path.write_text("仅属于项目 B 的公告内容", encoding="utf-8")
        foreign_material = db.add_artifact(
            project_b["id"],
            "Project B announcement",
            foreign_material_path,
            "asset",
            mime="text/plain",
        )

    with TestClient(test_app) as client:
        _login(client, "announcement-member-1", USER_PASSWORD)
        response = client.post(
            f"/api/projects/{project_a['id']}/announcement-terms",
            json={"material_artifact_ids": [foreign_material["id"]], "languages": ["en"]},
        )

    assert response.status_code == 404, response.text
    assert response.json() == {"detail": "project or artifact not found"}


@pytest.mark.parametrize("endpoint", ["apply", "fix-hard-blockers"])
def test_member_cannot_use_other_project_workbook_for_announcement_task(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    endpoint: str,
) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project_a = admin_client.post(PROJECTS_URL, json={"name": "Announcement Apply A", "type": "QA"}).json()
        project_b = admin_client.post(PROJECTS_URL, json={"name": "Announcement Apply B", "type": "QA"}).json()
        _create_user_via_api(admin_client, "announcement-member-2", "member")
        _clear_must_change_password("announcement-member-2")
        member = db.get_user_by_username("announcement-member-2")
        assert member is not None
        db.add_project_member(project_a["id"], member["id"], added_by="root-admin")

        source_path = tmp_path / "project-a-source.txt"
        source_path.write_text("公告内容\n", encoding="utf-8")
        source_artifact = db.add_artifact(project_a["id"], "Project A source", source_path, "asset", mime="text/plain")
        task = db.insert_announcement_task(
            project_a["id"],
            {
                "title": "Project A announcement",
                "source_artifact_id": source_artifact["id"],
                "source_format": "txt",
                "selected_languages": ["en"],
                "status": "translated",
                "current_step": 8,
                "metadata": {
                    "languages": ["en"],
                    "segments": [
                        {
                            "id": "segment-1",
                            "index": 0,
                            "kind": "line",
                            "source_file": source_path.name,
                            "source": "公告内容",
                        }
                    ],
                },
            },
        )

        foreign_workbook_path = tmp_path / "project-b-translation.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Translations"
        sheet.append(["segment_id", "CN", "protected_tokens", "term_hits_json", "EN"])
        sheet.append(["segment-1", "公告内容", "[]", "{}", "Announcement"])
        workbook.save(foreign_workbook_path)
        workbook.close()
        foreign_workbook = db.add_artifact(
            project_b["id"],
            "Project B translation workbook",
            foreign_workbook_path,
            "announcement_translation_workbook",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with TestClient(test_app) as client:
        _login(client, "announcement-member-2", USER_PASSWORD)
        response = client.post(
            f"/api/announcement-tasks/{task['id']}/{endpoint}",
            json={"languages": ["en"], "translation_workbook_artifact_id": foreign_workbook["id"]},
        )

    assert response.status_code == 404, response.text
    assert response.json() == {"detail": "announcement task or artifact not found"}


@pytest.mark.parametrize("endpoint", ["import-ai", "apply", "deliver"])
def test_member_cannot_use_other_project_prepare_run_for_legacy_announcement(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    endpoint: str,
) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project_a = admin_client.post(PROJECTS_URL, json={"name": "Legacy Announcement A", "type": "QA"}).json()
        project_b = admin_client.post(PROJECTS_URL, json={"name": "Legacy Announcement B", "type": "QA"}).json()
        _create_user_via_api(admin_client, "announcement-member-3", "member")
        _clear_must_change_password("announcement-member-3")
        member = db.get_user_by_username("announcement-member-3")
        assert member is not None
        db.add_project_member(project_a["id"], member["id"], added_by="root-admin")
        foreign = _prepare_legacy_announcement(admin_client, project_b["id"], tmp_path)

    with TestClient(test_app) as client:
        _login(client, "announcement-member-3", USER_PASSWORD)
        if endpoint == "import-ai":
            payload = {
                "prepare_run_id": foreign["prepare_run_id"],
                "response_artifact_ids": [foreign["response_artifact_id"]],
                "languages": ["en"],
            }
        elif endpoint == "apply":
            payload = {
                "prepare_run_id": foreign["prepare_run_id"],
                "translation_workbook_artifact_id": foreign["translation_workbook_artifact_id"],
            }
        else:
            payload = {"prepare_run_id": foreign["prepare_run_id"], "date_stamp": "20260716"}
        response = client.post(
            f"/api/projects/{project_a['id']}/announcement-docx/{endpoint}",
            json=payload,
        )

    assert response.status_code == 404, response.text
    assert response.json() == {"detail": "project, run, or artifact not found"}


def test_member_cannot_use_path_project_run_linked_to_other_project_announcement_task(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project_a = admin_client.post(PROJECTS_URL, json={"name": "Legacy Run A", "type": "QA"}).json()
        project_b = admin_client.post(PROJECTS_URL, json={"name": "Legacy Task B", "type": "QA"}).json()
        _create_user_via_api(admin_client, "announcement-member-4", "member")
        _clear_must_change_password("announcement-member-4")
        member = db.get_user_by_username("announcement-member-4")
        assert member is not None
        db.add_project_member(project_a["id"], member["id"], added_by="root-admin")
        foreign = _prepare_legacy_announcement(admin_client, project_b["id"], tmp_path)
        forged_run = db.insert_run(
            project_a["id"],
            kind="announcement_prepare",
            language="en",
            metadata={"task_id": foreign["task_id"], "legacy_prepare": True},
        )

    with TestClient(test_app) as client:
        _login(client, "announcement-member-4", USER_PASSWORD)
        response = client.post(
            f"/api/projects/{project_a['id']}/announcement-docx/import-ai",
            json={
                "prepare_run_id": forged_run["id"],
                "response_artifact_ids": [foreign["response_artifact_id"]],
                "languages": ["en"],
            },
        )

    assert response.status_code == 404, response.text
    assert response.json() == {"detail": "project, run, or artifact not found"}


@pytest.mark.parametrize(
    "material_field",
    ["project_material_artifact_ids", "announcement_material_artifact_ids"],
)
def test_ops_cannot_use_other_project_materials_for_glossary_extract(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    material_field: str,
) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project_a = admin_client.post(PROJECTS_URL, json={"name": "Glossary Scope A", "type": "QA"}).json()
        project_b = admin_client.post(PROJECTS_URL, json={"name": "Glossary Scope B", "type": "QA"}).json()
        _create_user_via_api(admin_client, "glossary-ops-1", "ops")
        _clear_must_change_password("glossary-ops-1")
        ops = db.get_user_by_username("glossary-ops-1")
        assert ops is not None
        db.add_project_member(project_a["id"], ops["id"], added_by="root-admin")

        input_path = tmp_path / "project-a-language.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "cn", "en"])
        sheet.append(["T-1", "英雄", "Hero"])
        workbook.save(input_path)
        workbook.close()
        input_artifact = db.add_artifact(
            project_a["id"],
            "Project A language table",
            input_path,
            "language_table",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        foreign_material_path = tmp_path / "project-b-material.md"
        foreign_material_path.write_text("Project B confidential material.", encoding="utf-8")
        foreign_material = db.add_artifact(
            project_b["id"],
            "Project B material",
            foreign_material_path,
            "asset",
            mime="text/markdown",
        )

    with TestClient(test_app) as client:
        _login(client, "glossary-ops-1", USER_PASSWORD)
        payload = {
            "input_artifact_id": input_artifact["id"],
            "id_column": "ID",
            "source_column": "cn",
            "target_column": "en",
            material_field: [foreign_material["id"]],
            "announcement_only": material_field == "announcement_material_artifact_ids",
            "ai_candidate_supplement": False,
            "update_project_prompt": False,
        }
        response = client.post(
            f"/api/projects/{project_a['id']}/glossary/extract",
            json=payload,
        )

    assert response.status_code == 404, response.text
    assert response.json() == {"detail": "project or artifact not found"}


def test_member_forbidden_on_delete_project(monkeypatch: pytest.MonkeyPatch) -> None:
    """PROJECT_MANAGE (delete project): member is a project member but the
    matrix says member can never delete a project, even one they created.
    """
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as admin_client:
        _bootstrap_admin_client(admin_client)
        project = admin_client.post(PROJECTS_URL, json={"name": "Undeletable By Member", "type": "QA"}).json()
        _create_user_via_api(admin_client, "member-3", "member")
        _clear_must_change_password("member-3")
        member = db.get_user_by_username("member-3")
        db.add_project_member(project["id"], member["id"], added_by="root-admin")

    with TestClient(test_app) as client:
        _login(client, "member-3", USER_PASSWORD)
        response = client.delete(f"/api/projects/{project['id']}")
    assert response.status_code == 403, response.text


def test_admin_bypasses_capability_and_membership_entirely(monkeypatch: pytest.MonkeyPatch) -> None:
    test_app = _required_app(monkeypatch)
    with TestClient(test_app) as client:
        _bootstrap_admin_client(client)
        project = client.post(PROJECTS_URL, json={"name": "Admin Sees All", "type": "QA"}).json()
        # Admin never got a project_members row (see _auto_add_creator_as_member),
        # yet must still be able to read/curate/delete it.
        assert not db.is_project_member(project["id"], db.get_user_by_username("root-admin")["id"])
        detail = client.get(f"/api/projects/{project['id']}")
        assert detail.status_code == 200, detail.text
        curated = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "术语", "target": "Term", "language": "en"},
        )
        assert curated.status_code == 200, curated.text
        deleted = client.delete(f"/api/projects/{project['id']}")
        assert deleted.status_code == 200, deleted.text


# ---------------------------------------------------------------------------
# fail-closed startup assertion
# ---------------------------------------------------------------------------


def test_fail_closed_startup_assertion_rejects_unregistered_route() -> None:
    fake_app = FastAPI()

    @fake_app.get("/api/totally-unregistered-route")
    def _fake() -> dict:
        return {"ok": True}

    with pytest.raises(RuntimeError, match=r"totally-unregistered-route"):
        route_capabilities.assert_full_route_coverage(fake_app)


def test_fail_closed_startup_assertion_passes_for_real_app() -> None:
    # The real app already asserts this at import time in main.py; re-running
    # it here pins the behavior against accidental future regressions (e.g.
    # someone catching/ignoring the RuntimeError at the call site).
    route_capabilities.assert_full_route_coverage(_build_app())


def test_app_metadata_version_matches_release_version_file() -> None:
    test_app = _build_app()
    expected = (Path(__file__).resolve().parents[2] / "VERSION").read_text(encoding="utf-8").strip()
    assert test_app.version == expected
    assert test_app.openapi()["info"]["version"] == expected


def test_every_capability_route_key_is_a_real_capability_constant() -> None:
    from app import authz

    for capability in route_capabilities.CAPABILITY_BY_ROUTE.values():
        assert capability in authz.ALL_CAPABILITIES


# ---------------------------------------------------------------------------
# auth-off (local) mode: zero regression promise
# ---------------------------------------------------------------------------


def test_auth_off_mode_ignores_capability_and_membership_gates_entirely() -> None:
    test_app = _build_app()
    with TestClient(test_app) as client:
        project = client.post(PROJECTS_URL, json={"name": "Local Mode Project", "type": "QA"}).json()
        curated = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "本地", "target": "Local", "language": "en"},
        )
        deleted = client.delete(f"/api/projects/{project['id']}")

    assert curated.status_code == 200, curated.text
    assert deleted.status_code == 200, deleted.text
