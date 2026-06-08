from __future__ import annotations

import os
import json
import re
import tempfile
import zipfile
from pathlib import Path

os.environ["LWS_DATA_ROOT"] = str(Path(tempfile.gettempdir()) / "lws-test-data")

import pytest
from docx import Document
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
import app.workflow as workflow
from app.config import DEFAULT_SETTINGS, save_settings
from app.main import app
from app.providers import TranslationItem, test_fake_translate_batch
from app.workflow import backfill_project_glossary_from_final
from conftest import reset_data_root, wait_for_background_jobs


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    reset_data_root(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    wait_for_background_jobs()
    save_settings(DEFAULT_SETTINGS)


def _sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "领取奖励", ""])
    ws.append([2, "开始游戏", ""])
    ws.append([3, "系统错误", ""])
    ws.append([4, "主线任务", ""])
    ws.append([5, "欢迎回来，{playerName}", ""])
    wb.save(path)
    wb.close()


def _sample_term_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2", "分类", "note"])
    ws.append([1, "最强指挥官", "Strongest Commander", "Top Commander", "title", "confirmed project term"])
    ws.append([2, "联盟", "Alliance", "Guild", "system", "common game term"])
    wb.save(path)
    wb.close()


def _large_language_table_workbook(path: Path, rows: int = 1001) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "KR"])
    for index in range(1, rows + 1):
        ws.append([index, f"完整语言表源文 {index}", f"완성 번역 {index}"])
    wb.save(path)
    wb.close()


def _translated_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "领取奖励", "Claim Rewards"])
    ws.append([2, "开始游戏", "Start Game"])
    ws.append([3, "系统错误", "System Error"])
    ws.append([4, "主线任务", "Main Quest"])
    ws.append([5, "欢迎回来，{playerName}", "Welcome back, {playerName}"])
    wb.create_sheet("EmptySheet")
    wb.save(path)
    wb.close()


def _string_id_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append(["M-001", "暗能电池", ""])
    ws.append(["M-002", "联盟基地", ""])
    wb.save(path)
    wb.close()


def _target_language_workbook(path: Path, target_header: str, targets: list[str] | None = None) -> None:
    values = targets or ["", ""]
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", target_header])
    ws.append(["btn.claim", "领取奖励", values[0]])
    ws.append(["msg.welcome", "欢迎回来，{playerName}", values[1]])
    wb.save(path)
    wb.close()


def _announcement_language_table(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN", "KR", "JP"])
    ws.append(["1001", "秘境", "Trial Realm", "시련의 영역", "試練の境界"])
    ws.append(["1002", "纹章", "Emblem", "문장", "紋章"])
    ws.append(["1003", "商城", "Shop", "상점", "ショップ"])
    wb.save(path)
    wb.close()


def _announcement_docx(path: Path, text: str = "英雄觉醒 2026/5/20") -> None:
    doc = Document()
    doc.add_paragraph(text)
    doc.save(path)



def test_canceled_announcement_tasks_do_not_count_as_active_project_tasks() -> None:
    client = TestClient(app)
    project = client.post(
        "/api/projects",
        json={"name": "Announcement count smoke", "type": "QA", "description": "count active announcement tasks only"},
    ).json()

    task = client.post(
        f"/api/projects/{project['id']}/announcement-tasks",
        json={"title": "notice.txt", "text": "New event announcement."},
    ).json()
    detail = client.get(f"/api/projects/{project['id']}").json()
    assert detail["stats"]["announcement_tasks"] == 1

    client.post(f"/api/announcement-tasks/{task['id']}/cancel")
    detail = client.get(f"/api/projects/{project['id']}").json()

    assert detail["stats"]["announcement_tasks"] == 0
    assert detail["stats"]["tasks"] == detail["stats"]["language_tasks"]

def _announcement_ko_terms(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "KR"])
    ws.append(["term_hero", "英雄", "영웅"])
    ws.append(["term_awaken", "觉醒", "각성"])
    wb.save(path)
    wb.close()


def _announcement_existing_terms_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "FR", "DE", "RU", "IT", "ES", "PT", "TK", "ID", "TH"])
    ws.append(["alliance_txt_21", "公告", "Notice", "Annonce", "Ankündigung", "Объявление", "Annuncio", "Anuncio", "Anúncio", "Duyuru", "Pengumuman", "ประกาศ"])
    ws.append(["time_zone_2", "服务器时间", "Server Time", "Heure du serveur", "Serverzeit", "Время сервера", "Orario server", "Hora del servidor", "Horário de servidor", "Sunucu Saati", "Waktu Server", "เวลาเซิร์ฟเวอร์"])
    wb.save(path)
    wb.close()


def _generated_announcement_terms_like_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "命中次数", "来源", "备注"])
    ws.append(["old_notice", "公告旧术语", "Old Notice Term", 9, "announcement", "already extracted"])
    wb.save(path)
    wb.close()


def _project_harness_failed_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "cn", "en"])
    ws.append([1, "棰嗗彇濂栧姳", "Forbidden Brand Reward"])
    ws.append([2, "Start game source", "Start Game"])
    wb.save(path)
    wb.close()


def test_fake_provider_preserves_numeric_square_placeholders() -> None:
    result = test_fake_translate_batch(
        [{"id": 236, "source": "审批文号：新广出审[2016]2751号\\n出版物号:ISBN"}],
        {},
    )
    assert "[2016]" in result[0].translation
    assert "\\n" in result[0].translation


def test_delete_project_removes_project_records_and_files(tmp_path: Path) -> None:
    upload_path = tmp_path / "language.xlsx"
    _target_language_workbook(upload_path, "EN", ["Claim", "Welcome"])

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Delete Me", "type": "QA"}).json()
        with upload_path.open("rb") as fh:
            artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        client.post(f"/api/projects/{project['id']}/glossary", json={"source": "按钮", "target": "Button", "language": "en"})
        client.post(f"/api/projects/{project['id']}/translations", json={"source": "领取奖励", "target": "Claim Reward", "language": "en"})
        project_path = workflow.project_dir(project["id"])
        assert project_path.exists()
        assert Path(artifact["path"]).exists()

        deleted = client.delete(f"/api/projects/{project['id']}")
        assert deleted.status_code == 200, deleted.text
        assert deleted.json() == {"deleted": True}

        assert client.get(f"/api/projects/{project['id']}").status_code == 404
        assert all(item["id"] != project["id"] for item in client.get("/api/projects").json())
        assert db.list_artifacts(project_id=project["id"]) == []
        assert db.list_runs(project["id"]) == []
        assert db.list_glossary_terms(project["id"]) == []
        assert db.list_translation_entries(project["id"]) == []
        assert not project_path.exists()


def test_announcement_task_can_be_canceled_without_deleting_audit() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Cancel Announcement", "type": "QA"}).json()
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"title": "Stuck Notice", "text": "新公告内容", "languages": ["en", "ko"]},
        ).json()

        response = client.post(f"/api/announcement-tasks/{task['id']}/cancel")

        assert response.status_code == 200, response.text
        payload = response.json()
        canceled = payload["task"]
        assert canceled["status"] == "canceled"
        assert canceled["metadata"]["canceled_at"]
        assert {item["status"] for item in canceled["languages"]} == {"canceled"}

        fetched = client.get(f"/api/announcement-tasks/{task['id']}").json()
        assert fetched["status"] == "canceled"
        assert fetched["source_artifact_id"] == task["source_artifact_id"]


def test_project_stats_count_business_tasks_not_execution_runs() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Stats Scope", "type": "QA"}).json()
        source_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en"},
        ).json()
        client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "source_run_id": source_run["id"]},
        )
        db.insert_run(
            project["id"],
            kind="qa",
            language="en",
            metadata={"manual_fix_source_run_id": source_run["id"]},
        )
        db.insert_run(
            project["id"],
            kind="qa",
            language="en",
            metadata={"model_fix_source_run_id": source_run["id"]},
        )

        stats = client.get(f"/api/projects/{project['id']}").json()["stats"]
        assert stats["language_tasks"] == 1
        assert stats["tasks"] == 1
        assert stats["execution_runs"] == 4


def test_announcement_terms_endpoint_generates_multilingual_terms_with_alias_headers(tmp_path: Path) -> None:
    table_path = tmp_path / "language.xlsx"
    _announcement_language_table(table_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "公告术语", "type": "RPG"}).json()
        with table_path.open("rb") as fh:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        response = client.post(
            f"/api/projects/{project['id']}/announcement-terms",
            json={
                "text": "新增秘境和纹章系统。",
                "language_table_artifact_ids": [table_artifact["id"]],
                "languages": ["en", "ko", "ja"],
                "announcement_min_hit": 1,
                "generate_validation": True,
            },
        )

        assert response.status_code == 200, response.text
        payload = response.json()
        kinds = {artifact["kind"] for artifact in payload["artifacts"]}
        assert {"announcement_terms_workbook", "announcement_terms_validation", "announcement_terms_manifest"}.issubset(kinds)
        workbook_artifact = next(artifact for artifact in payload["artifacts"] if artifact["kind"] == "announcement_terms_workbook")
        wb = load_workbook(workbook_artifact["path"], read_only=True, data_only=True)
        rows = list(wb["Glossary"].iter_rows(values_only=True))
        wb.close()
        assert rows[0] == ("ID", "CN", "EN", "KR", "JP", "命中次数", "来源", "备注")
        assert [row[:5] for row in rows[1:]] == [
            ("1001", "秘境", "Trial Realm", "시련의 영역", "試練の境界"),
            ("1002", "纹章", "Emblem", "문장", "紋章"),
        ]
        assert payload["manifest"]["languages"] == ["en", "ko", "ja"]
        assert payload["summary"]["terms"] == 2


def test_announcement_task_extract_terms_ignores_generated_terms_workbook_as_constraint(tmp_path: Path) -> None:
    table_path = tmp_path / "full_language_table.xlsx"
    generated_terms_path = tmp_path / "notice_announcement_terms_20260601.xlsx"
    notice_path = tmp_path / "notice.txt"

    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN"])
    ws.append(["T1", "\u79d8\u5883", "Trial Realm"])
    wb.save(table_path)
    wb.close()
    _generated_announcement_terms_like_workbook(generated_terms_path)
    notice_path.write_text("\u516c\u544a\u65e7\u672f\u8bed\u548c\u79d8\u5883\u4e0a\u7ebf\u3002", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement Constraint Source", "type": "RPG"}).json()
        with table_path.open("rb") as fh:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("full_language_table.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with generated_terms_path.open("rb") as fh:
            generated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("notice_announcement_terms_20260601.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with notice_path.open("rb") as fh:
            notice_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()

        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={
                "source_artifact_id": notice_artifact["id"],
                "language_table_artifact_ids": [table_artifact["id"], generated_artifact["id"]],
                "languages": ["en"],
                "include_project_archive": False,
            },
        ).json()
        extracted = client.post(
            f"/api/announcement-tasks/{task['id']}/extract-terms",
            json={
                "language_table_artifact_ids": [table_artifact["id"], generated_artifact["id"]],
                "languages": ["en"],
                "include_project_archive": False,
                "ai_supplement": False,
            },
        )

        assert extracted.status_code == 200, extracted.text
        terms = extracted.json()["task"]["metadata"]["terms"]
        assert [term["source"] for term in terms] == ["\u79d8\u5883"]
        assert terms[0]["translations"]["en"] == "Trial Realm"


def test_announcement_force_delivery_with_hard_blockers_generates_package(tmp_path: Path) -> None:
    source_path = tmp_path / "notice_force.txt"
    terms_path = tmp_path / "notice_terms.xlsx"
    source_path.write_text("\u82f1\u96c4\u5956\u52b1 {0}\n", encoding="utf-8")
    _announcement_ko_terms(terms_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "\u516c\u544a\u5f3a\u5236\u4ea4\u4ed8", "type": "RPG"}).json()
        with source_path.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice_force.txt", fh, "text/plain")},
            ).json()
        with terms_path.open("rb") as fh:
            terms_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("notice_terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": source_artifact["id"], "language_table_artifact_ids": [terms_artifact["id"]], "languages": ["ko"]},
        ).json()
        task_id = task["id"]
        for endpoint in ("inspect-constraints", "extract-terms", "lookup-translations"):
            response = client.post(
                f"/api/announcement-tasks/{task_id}/{endpoint}",
                json={"language_table_artifact_ids": [terms_artifact["id"]], "languages": ["ko"], "include_project_archive": True},
            )
            assert response.status_code == 200, response.text
        prepared = client.post(f"/api/announcement-tasks/{task_id}/prepare", json={"languages": ["ko"]}).json()
        workpack = next(artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_workpack")
        rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]
        response_path = tmp_path / "ai_response_ko.jsonl"
        response_path.write_text(
            "\n".join(json.dumps({"para_id": row["para_id"], "translation": "\uc601\uc6c5 \ubcf4\uc0c1"}, ensure_ascii=False) for row in rows) + "\n",
            encoding="utf-8",
        )
        with response_path.open("rb") as fh:
            response_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("ai_response_ko.jsonl", fh, "application/jsonl")},
            ).json()
        imported = client.post(
            f"/api/announcement-tasks/{task_id}/import-ai",
            json={"languages": ["ko"], "response_artifact_ids": [response_artifact["id"]]},
        )
        assert imported.status_code == 200, imported.text

        blocked = client.post(f"/api/announcement-tasks/{task_id}/apply", json={"languages": ["ko"]})
        assert blocked.status_code == 400, blocked.text
        assert "hard blockers" in blocked.json()["detail"]

        normal_delivery = client.post(f"/api/announcement-tasks/{task_id}/deliver", json={"languages": ["ko"], "date_stamp": "20260608"})
        assert normal_delivery.status_code == 400
        assert "hard blockers" in normal_delivery.json()["detail"]

        forced = client.post(f"/api/announcement-tasks/{task_id}/deliver", json={"languages": ["ko"], "date_stamp": "20260608", "force": True})
        assert forced.status_code == 200, forced.text
        summary = forced.json()["summary"]
        assert summary["forced"] is True
        assert summary["hard_blockers"] > 0
        package = next(artifact for artifact in forced.json()["artifacts"] if artifact["kind"] == "announcement_delivery_package")
        assert package["metadata"]["forced"] is True
        with zipfile.ZipFile(package["path"]) as archive:
            names = sorted(archive.namelist())
            assert names == ["KR/notice_force_KR.txt", "QA\u6458\u8981.xlsx"]
            qa_wb = load_workbook(archive.open("QA\u6458\u8981.xlsx"), read_only=True, data_only=True)
            qa_summary = {row[0]: row[1] for row in qa_wb["Summary"].iter_rows(min_row=2, values_only=True)}
            assert qa_summary["hard_blockers"] > 0
            assert qa_summary["outputs"] == 1
            qa_wb.close()


def test_announcement_fix_hard_blockers_repairs_missing_token_and_applies(tmp_path: Path) -> None:
    source_path = tmp_path / "notice_fix.txt"
    terms_path = tmp_path / "notice_terms.xlsx"
    source_path.write_text("\u82f1\u96c4\u5956\u52b1 {0}\n", encoding="utf-8")
    _announcement_ko_terms(terms_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "\u516c\u544aHardBlock\u4fee\u590d", "type": "RPG"}).json()
        with source_path.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice_fix.txt", fh, "text/plain")},
            ).json()
        with terms_path.open("rb") as fh:
            terms_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("notice_terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": source_artifact["id"], "language_table_artifact_ids": [terms_artifact["id"]], "languages": ["ko"]},
        ).json()
        task_id = task["id"]
        for endpoint in ("inspect-constraints", "extract-terms", "lookup-translations"):
            response = client.post(
                f"/api/announcement-tasks/{task_id}/{endpoint}",
                json={"language_table_artifact_ids": [terms_artifact["id"]], "languages": ["ko"], "include_project_archive": True},
            )
            assert response.status_code == 200, response.text
        prepared = client.post(f"/api/announcement-tasks/{task_id}/prepare", json={"languages": ["ko"]}).json()
        workpack = next(artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_workpack")
        rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]
        response_path = tmp_path / "ai_response_ko.jsonl"
        response_path.write_text(
            "\n".join(json.dumps({"para_id": row["para_id"], "translation": "\uc601\uc6c5 \ubcf4\uc0c1"}, ensure_ascii=False) for row in rows) + "\n",
            encoding="utf-8",
        )
        with response_path.open("rb") as fh:
            response_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("ai_response_ko.jsonl", fh, "application/jsonl")},
            ).json()
        imported = client.post(
            f"/api/announcement-tasks/{task_id}/import-ai",
            json={"languages": ["ko"], "response_artifact_ids": [response_artifact["id"]]},
        )
        assert imported.status_code == 200, imported.text
        blocked = client.post(f"/api/announcement-tasks/{task_id}/apply", json={"languages": ["ko"]})
        assert blocked.status_code == 400, blocked.text

        fixed = client.post(f"/api/announcement-tasks/{task_id}/fix-hard-blockers", json={"languages": ["ko"]})
        assert fixed.status_code == 200, fixed.text
        payload = fixed.json()
        assert payload["summary"]["fixed"] == 1
        assert payload["summary"]["remaining_hard_blockers"] == 0
        assert payload["task"]["status"] == "applied"
        output_artifact = next(artifact for artifact in payload["artifacts"] if artifact["kind"] == "announcement_output_file")
        assert "{0}" in Path(output_artifact["path"]).read_text(encoding="utf-8")




def test_announcement_extract_rejects_legacy_txt_constraint_with_human_message(tmp_path: Path) -> None:
    bad_constraint = tmp_path / "notice.txt"
    bad_constraint.write_text("this is announcement source, not a language table", encoding="utf-8")
    notice_path = tmp_path / "announcement.txt"
    notice_path.write_text("\u79d8\u5883\u4e0a\u7ebf\u3002", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Bad Constraint TXT", "type": "RPG"}).json()
        legacy_bad_artifact = workflow.copy_upload(project["id"], bad_constraint, "notice.txt", "language_table")
        with notice_path.open("rb") as fh:
            notice_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("announcement.txt", fh, "text/plain")},
            ).json()
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={
                "source_artifact_id": notice_artifact["id"],
                "language_table_artifact_ids": [legacy_bad_artifact["id"]],
                "languages": ["ko"],
                "include_project_archive": False,
            },
        ).json()
        response = client.post(
            f"/api/announcement-tasks/{task['id']}/extract-terms",
            json={
                "language_table_artifact_ids": [legacy_bad_artifact["id"]],
                "languages": ["ko"],
                "include_project_archive": False,
                "ai_supplement": False,
            },
        )
        assert response.status_code == 400
        detail = response.json()["detail"]
        assert "\u7ea6\u675f\u6587\u4ef6\u683c\u5f0f\u4e0d\u6b63\u786e" in detail
        assert "TXT" in detail or ".txt" in detail
        refreshed = client.get(f"/api/announcement-tasks/{task['id']}").json()
        assert refreshed["status"] != "terms_ready"


def test_announcement_extract_rejects_unrecognized_language_table_headers(tmp_path: Path) -> None:
    table_path = tmp_path / "wrong_headers.xlsx"
    notice_path = tmp_path / "announcement.txt"
    wb = Workbook()
    ws = wb.active
    ws.append(["foo", "bar", "KR"])
    ws.append(["1", "\u79d8\u5883", "\ube44\uacbd"])
    wb.save(table_path)
    wb.close()
    notice_path.write_text("\u79d8\u5883\u4e0a\u7ebf\u3002", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Bad Constraint Headers", "type": "RPG"}).json()
        with table_path.open("rb") as fh:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("wrong_headers.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with notice_path.open("rb") as fh:
            notice_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("announcement.txt", fh, "text/plain")},
            ).json()
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": notice_artifact["id"], "language_table_artifact_ids": [table_artifact["id"]], "languages": ["ko"], "include_project_archive": False},
        ).json()
        response = client.post(
            f"/api/announcement-tasks/{task['id']}/extract-terms",
            json={"language_table_artifact_ids": [table_artifact["id"]], "languages": ["ko"], "include_project_archive": False, "ai_supplement": False},
        )
        assert response.status_code == 400
        detail = response.json()["detail"]
        assert "\u672a\u8bc6\u522b\u5230\u53ef\u53cd\u67e5\u8bcd\u6761" in detail
        assert "CN" in detail


def test_announcement_task_extract_terms_accepts_ai_supplement_response(tmp_path: Path) -> None:
    table_path = tmp_path / "language.xlsx"
    notice_path = tmp_path / "notice.txt"
    response_path = tmp_path / "ai_supplement_response.json"

    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN"])
    ws.append(["T1", "\u79d8\u5883", "Trial Realm"])
    ws.append(["S1", "\u5f00\u542f\u661f\u754c\u88c2\u9699\u6311\u6218", "Unlock Astral Rift Challenge"])
    ws.append(["N1", "\u5b8c\u5168\u65e0\u5173\u7cfb\u7edf", "Unrelated System"])
    wb.save(table_path)
    wb.close()
    notice_path.write_text("\u65b0\u589e\u79d8\u5883\u548c\u661f\u754c\u88c2\u9699\u73a9\u6cd5\u3002", encoding="utf-8")
    response_path.write_text(
        json.dumps(
            {
                "supplement_terms": [
                    {
                        "cn": "\u661f\u754c\u88c2\u9699",
                        "translations": {"EN": "Astral Rift"},
                        "source_ids": ["S1"],
                        "confidence": "high",
                        "reason": "split from language-table sentence",
                        "evidence_ids": ["S1"],
                        "action": "add_to_main",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement AI", "type": "RPG"}).json()
        with table_path.open("rb") as fh:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with notice_path.open("rb") as fh:
            notice_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()
        with response_path.open("rb") as fh:
            response_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("ai_supplement_response.json", fh, "application/json")},
            ).json()

        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": notice_artifact["id"], "language_table_artifact_ids": [table_artifact["id"]], "languages": ["en"]},
        ).json()
        extracted = client.post(
            f"/api/announcement-tasks/{task['id']}/extract-terms",
            json={
                "language_table_artifact_ids": [table_artifact["id"]],
                "languages": ["en"],
                "ai_supplement": True,
                "ai_supplement_response_artifact_id": response_artifact["id"],
            },
        )

        assert extracted.status_code == 200, extracted.text
        payload = extracted.json()
        terms = payload["task"]["metadata"]["terms"]
        assert [term["source"] for term in terms] == ["\u79d8\u5883", "\u661f\u754c\u88c2\u9699"]
        assert terms[1]["translations"]["en"] == "Astral Rift"
        assert payload["summary"]["ai_supplement"]["added_to_main"] == 1
        artifact_kinds = {artifact["kind"] for artifact in payload["artifacts"]}
        assert {"announcement_ai_supplement_packet", "announcement_ai_supplement_report"}.issubset(artifact_kinds)
        task_artifact_kinds = {artifact["kind"] for artifact in payload["task"]["artifacts"]}
        assert {"announcement_ai_supplement_packet", "announcement_ai_supplement_report"}.issubset(task_artifact_kinds)


def test_announcement_task_extract_terms_calls_provider_for_ai_supplement(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    table_path = tmp_path / "language.xlsx"
    notice_path = tmp_path / "notice.txt"

    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN"])
    ws.append(["T1", "秘境", "Trial Realm"])
    ws.append(["S1", "开启星界裂隙挑战", "Unlock Astral Rift Challenge"])
    wb.save(table_path)
    wb.close()
    notice_path.write_text("新增秘境和星界裂隙玩法。", encoding="utf-8")

    calls: list[str] = []

    def fake_provider(settings: dict, prompt: str) -> str:
        calls.append(prompt)
        assert settings["provider"] == "openai"
        assert "announcement_ai_supplement" in prompt
        return json.dumps(
            {
                "supplement_terms": [
                    {
                        "cn": "星界裂隙",
                        "translations": {"EN": "Astral Rift"},
                        "source_ids": ["S1"],
                        "confidence": "high",
                        "reason": "split from language-table sentence",
                        "evidence_ids": ["S1"],
                        "action": "add_to_main",
                    }
                ]
            },
            ensure_ascii=False,
        )

    monkeypatch.setattr(workflow, "_call_semantic_provider", fake_provider)

    with TestClient(app) as client:
        client.patch("/api/settings", json={"provider": "openai", "api_key": "test-key", "model": "gpt-test"})
        project = client.post("/api/projects", json={"name": "Announcement AI Provider", "type": "RPG"}).json()
        with table_path.open("rb") as fh:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with notice_path.open("rb") as fh:
            notice_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()

        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": notice_artifact["id"], "language_table_artifact_ids": [table_artifact["id"]], "languages": ["en"]},
        ).json()
        extracted = client.post(
            f"/api/announcement-tasks/{task['id']}/extract-terms",
            json={
                "language_table_artifact_ids": [table_artifact["id"]],
                "languages": ["en"],
                "ai_supplement": True,
            },
        )

        assert extracted.status_code == 200, extracted.text
        assert len(calls) == 1
        payload = extracted.json()
        terms = payload["task"]["metadata"]["terms"]
        assert [term["source"] for term in terms] == ["秘境", "星界裂隙"]
        assert payload["summary"]["ai_supplement"]["provider"] == "openai"
        assert payload["summary"]["ai_supplement"]["added_to_main"] == 1
        artifact_kinds = {artifact["kind"] for artifact in payload["artifacts"]}
        assert {
            "announcement_ai_supplement_packet",
            "announcement_ai_supplement_response",
            "announcement_ai_supplement_report",
        }.issubset(artifact_kinds)


def test_announcement_task_can_import_edit_and_export_existing_terms(tmp_path: Path) -> None:
    source_path = tmp_path / "notice.txt"
    terms_path = tmp_path / "announcement_terms.xlsx"
    source_path.write_text("公告：服务器时间维护。", encoding="utf-8")
    _announcement_existing_terms_workbook(terms_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "公告术语编辑", "type": "RPG"}).json()
        with source_path.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()
        with terms_path.open("rb") as fh:
            terms_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("announcement_terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": source_artifact["id"]},
        ).json()
        imported = client.post(
            f"/api/announcement-tasks/{task['id']}/import-terms",
            json={"terms_artifact_id": terms_artifact["id"]},
        )
        assert imported.status_code == 200, imported.text
        imported_task = imported.json()["task"]
        assert imported_task["metadata"]["terms_summary"]["terms"] == 2
        assert imported_task["selected_languages"] == ["en", "fr", "de", "ru", "it", "es", "pt", "tr", "idn", "th"]

        terms = imported_task["metadata"]["terms"]
        terms[0]["translations"]["en"] = "Edited Notice"
        terms[0]["source"] = "公告标题"
        saved = client.post(
            f"/api/announcement-tasks/{task['id']}/import-terms",
            json={"languages": imported_task["selected_languages"], "terms": terms},
        )
        assert saved.status_code == 200, saved.text
        workbook_id = saved.json()["task"]["metadata"]["terms_artifact_id"]
        artifact = next(item for item in saved.json()["artifacts"] if item["id"] == workbook_id)
        wb = load_workbook(Path(artifact["path"]), read_only=True, data_only=True)
        try:
            rows = list(wb["Glossary"].iter_rows(values_only=True))
        finally:
            wb.close()
        assert rows[0] == ("ID", "CN", "EN", "FR", "DE", "RU", "IT", "ES", "PT", "TR", "IDN", "TH", "命中次数", "来源", "备注")
        assert rows[1][1] == "公告标题"
        assert rows[1][2] == "Edited Notice"


def test_announcement_terms_import_does_not_treat_identifier_id_as_indonesian(tmp_path: Path) -> None:
    source_path = tmp_path / "notice.txt"
    terms_path = tmp_path / "jp_terms.xlsx"
    source_path.write_text("公告：钓鱼玩法开放。", encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "JP"])
    ws.append(["626900", "钓鱼玩法", "釣りコンテンツ"])
    wb.save(terms_path)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "JP 公告术语", "type": "RPG"}).json()
        with source_path.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()
        with terms_path.open("rb") as fh:
            terms_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("jp_terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": source_artifact["id"]},
        ).json()
        imported = client.post(
            f"/api/announcement-tasks/{task['id']}/import-terms",
            json={"terms_artifact_id": terms_artifact["id"]},
        )
        assert imported.status_code == 200, imported.text
        payload = imported.json()["task"]
        assert payload["selected_languages"] == ["ja"]
        assert payload["metadata"]["terms"][0]["translations"] == {"ja": "釣りコンテンツ"}


def test_announcement_docx_harness_api_prepares_imports_applies_and_delivers(tmp_path: Path) -> None:
    docx_path = tmp_path / "notice.docx"
    terms_path = tmp_path / "notice_terms.xlsx"
    _announcement_docx(docx_path)
    _announcement_ko_terms(terms_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "公告 DOCX", "type": "RPG"}).json()
        with docx_path.open("rb") as fh:
            docx_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.docx", fh, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")},
            ).json()
        with terms_path.open("rb") as fh:
            terms_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("notice_terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        prepare_response = client.post(
            f"/api/projects/{project['id']}/announcement-docx/prepare",
            json={"source_artifact_ids": [docx_artifact["id"]], "terms_artifact_id": terms_artifact["id"], "languages": ["ko"]},
        )
        assert prepare_response.status_code == 200, prepare_response.text
        prepared = prepare_response.json()
        prepare_run_id = prepared["run"]["id"]
        prepare_kinds = {artifact["kind"] for artifact in prepared["artifacts"]}
        assert {"announcement_docx_translation_workbook", "announcement_docx_manifest", "announcement_docx_workpack"}.issubset(prepare_kinds)
        workpack = next(artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_docx_workpack")
        rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]
        assert rows and rows[0]["term_hits"] == [{"source": "英雄", "target": "영웅"}, {"source": "觉醒", "target": "각성"}]

        response_path = tmp_path / "ai_response_ko.jsonl"
        response_path.write_text(
            "\n".join(json.dumps({"para_id": row["para_id"], "translation": "영웅 각성 2026/5/20"}, ensure_ascii=False) for row in rows) + "\n",
            encoding="utf-8",
        )
        with response_path.open("rb") as fh:
            response_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("ai_response_ko.jsonl", fh, "application/jsonl")},
            ).json()

        import_response = client.post(
            f"/api/projects/{project['id']}/announcement-docx/import-ai",
            json={"prepare_run_id": prepare_run_id, "response_artifact_ids": [response_artifact["id"]], "languages": ["ko"]},
        )
        assert import_response.status_code == 200, import_response.text
        assert import_response.json()["summary"]["languages"] == ["KR"]

        translation_workbook = next(artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_docx_translation_workbook")
        apply_response = client.post(
            f"/api/projects/{project['id']}/announcement-docx/apply",
            json={"prepare_run_id": prepare_run_id, "translation_workbook_artifact_id": translation_workbook["id"]},
        )
        assert apply_response.status_code == 200, apply_response.text
        apply_kinds = {artifact["kind"] for artifact in apply_response.json()["artifacts"]}
        assert {"announcement_docx_output_docx", "announcement_docx_qa_summary"}.issubset(apply_kinds)

        deliver_response = client.post(
            f"/api/projects/{project['id']}/announcement-docx/deliver",
            json={"prepare_run_id": prepare_run_id, "date_stamp": "20260526"},
        )
        assert deliver_response.status_code == 200, deliver_response.text
        delivery_artifact = next(artifact for artifact in deliver_response.json()["artifacts"] if artifact["kind"] == "announcement_docx_delivery_package")
        assert Path(delivery_artifact["path"]).exists()


def test_announcement_task_txt_multilingual_flow_uses_archive_priority_and_delivers(tmp_path: Path) -> None:
    source_path = tmp_path / "notice.txt"
    terms_path = tmp_path / "notice_terms.xlsx"
    source_path.write_text("英雄觉醒 2026/5/20\n", encoding="utf-8")
    _announcement_ko_terms(terms_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "公告任务 TXT", "type": "RPG"}).json()
        client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "hero",
                "source": "英雄",
                "target": "히어로",
                "language": "ko",
                "source_type": "qa_passed",
            },
        )
        with source_path.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()
        with terms_path.open("rb") as fh:
            terms_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("notice_terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={"source_artifact_id": source_artifact["id"], "language_table_artifact_ids": [terms_artifact["id"]], "languages": ["ko"]},
        ).json()
        task_id = task["id"]
        assert task["source_format"] == "txt"

        for endpoint in ("inspect-constraints", "extract-terms", "lookup-translations"):
            response = client.post(
                f"/api/announcement-tasks/{task_id}/{endpoint}",
                json={"language_table_artifact_ids": [terms_artifact["id"]], "languages": ["ko"], "include_project_archive": True},
            )
            assert response.status_code == 200, response.text

        prepare_response = client.post(f"/api/announcement-tasks/{task_id}/prepare", json={"languages": ["ko"]})
        assert prepare_response.status_code == 200, prepare_response.text
        prepared = prepare_response.json()
        workpack = next(artifact for artifact in prepared["artifacts"] if artifact["kind"] == "announcement_workpack")
        rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]
        assert rows[0]["term_hits"] == [{"source": "英雄", "target": "히어로"}, {"source": "觉醒", "target": "각성"}]

        response_path = tmp_path / "ai_response_ko.jsonl"
        response_path.write_text(
            "\n".join(json.dumps({"para_id": row["para_id"], "translation": "히어로 각성 2026/5/20"}, ensure_ascii=False) for row in rows) + "\n",
            encoding="utf-8",
        )
        with response_path.open("rb") as fh:
            response_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("ai_response_ko.jsonl", fh, "application/jsonl")},
            ).json()
        imported = client.post(
            f"/api/announcement-tasks/{task_id}/import-ai",
            json={"languages": ["ko"], "response_artifact_ids": [response_artifact["id"]]},
        )
        assert imported.status_code == 200, imported.text
        applied = client.post(f"/api/announcement-tasks/{task_id}/apply", json={"languages": ["ko"]})
        assert applied.status_code == 200, applied.text
        delivered = client.post(f"/api/announcement-tasks/{task_id}/deliver", json={"languages": ["ko"], "date_stamp": "20260526"})
        assert delivered.status_code == 200, delivered.text
        package = next(artifact for artifact in delivered.json()["artifacts"] if artifact["kind"] == "announcement_delivery_package")
        package_path = Path(package["path"])
        assert package_path.exists()
        assert re.fullmatch(r".+_notice_announcement_delivery_20260526\.zip", package_path.name)
        assert ".txt_announcement" not in package_path.name
        with zipfile.ZipFile(package_path) as archive:
            names = sorted(archive.namelist())
        assert names == ["KR/notice_KR.txt", "QA摘要.xlsx"]
        assert not any(name.endswith(".json") or name.endswith(".jsonl") or "manifest" in name.lower() or "workpack" in name.lower() for name in names)
        repeated = client.post(f"/api/announcement-tasks/{task_id}/deliver", json={"languages": ["ko"], "date_stamp": "20260526"})
        assert repeated.status_code == 200, repeated.text
        assert repeated.json()["summary"]["reused"] is True
        assert repeated.json()["summary"]["delivery_artifact_id"] == package["id"]

        deliverables = client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"]
        announcement_deliverable = next(item for item in deliverables if item["task_code"] == "ANN")
        assert announcement_deliverable["status"] == "delivered"
        assert announcement_deliverable["task_type"] == "公告任务"
        assert announcement_deliverable["language"] == "KR"
        assert announcement_deliverable["files"]["package"]["download_url"].startswith("/api/artifacts/")
        assert announcement_deliverable["files"]["qa_summary"]["download_url"].startswith("/api/artifacts/")
        assert announcement_deliverable["files"]["outputs"][0]["download_url"].startswith("/api/artifacts/")

        forced = client.post(f"/api/announcement-tasks/{task_id}/deliver", json={"languages": ["ko"], "date_stamp": "20260526", "force": True})
        assert forced.status_code == 200, forced.text
        forced_package = next(artifact for artifact in forced.json()["artifacts"] if artifact["kind"] == "announcement_delivery_package")
        assert forced_package["id"] != package["id"]
        superseded_package = db.get_artifact(package["id"])
        assert superseded_package["metadata"]["superseded"] is True
        assert superseded_package["metadata"]["superseded_by"] == forced_package["id"]
        visible_packages = [
            artifact for artifact in db.list_artifacts(project_id=project["id"], role="delivery")
            if artifact["kind"] == "announcement_delivery_package" and (artifact.get("metadata") or {}).get("task_id") == task_id
        ]
        assert [artifact["id"] for artifact in visible_packages] == [forced_package["id"]]
        project_detail = client.get(f"/api/projects/{project['id']}").json()
        assert project_detail["stats"]["tasks"] == 1
        assert project_detail["stats"]["announcement_tasks"] == 1
        assert project_detail["stats"]["language_tasks"] == 0
        assert project_detail["stats"]["deliverables"] == 1
        assert project_detail["stats"]["execution_runs"] > project_detail["stats"]["tasks"]

        qa_artifact = next(artifact for artifact in applied.json()["artifacts"] if artifact["kind"] == "announcement_qa_summary")
        qa_wb = load_workbook(qa_artifact["path"], read_only=True, data_only=True)
        try:
            assert qa_wb.sheetnames == ["Summary", "Issues", "Outputs"]
            summary = {row[0]: row[1] for row in qa_wb["Summary"].iter_rows(min_row=2, values_only=True)}
            assert summary["hard_blockers"] == 0
            assert summary["outputs"] == 1
            assert summary["languages"] == "KR"
            outputs = list(qa_wb["Outputs"].iter_rows(min_row=2, values_only=True))
            assert outputs[0][0] == "KR"
            assert outputs[0][1] == "notice_KR.txt"
        finally:
            qa_wb.close()


def test_fake_provider_runs_english_workflow_end_to_end(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project_response = client.post(
            "/api/projects",
            json={
                "name": "Synthetic Frontier",
                "type": "科幻 SLG",
                "icon": "🚀",
                "description": "Synthetic public demo project.",
            },
        )
        assert project_response.status_code == 200
        project = project_response.json()

        analysis_response = client.post(
            f"/api/projects/{project['id']}/analyze",
            json={"intro": "A synthetic strategy game for testing localization workflow.", "asset_artifact_ids": []},
        )
        assert analysis_response.status_code == 200
        assert "只返回 JSONL" in analysis_response.json()["prompt"]
        analyzed_project = analysis_response.json()["project"]
        display_prompt = analyzed_project["profile"]["display_prompts_by_language"]["en"]
        assert "\u7ffb\u8bd1\u76ee\u6807" in display_prompt
        assert "????" not in display_prompt
        assert "JSONL" not in display_prompt

        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert upload_response.status_code == 200
        source_artifact = upload_response.json()

        run_response = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "translation",
                "language": "en",
                "input_artifact_id": source_artifact["id"],
                "batch_size": 3,
            },
        )
        assert run_response.status_code == 200
        run = run_response.json()

        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake"})
        assert translate_response.status_code == 200, translate_response.text
        result = translate_response.json()
        assert result["run"]["status"] == "passed"
        kinds = {artifact["kind"] for artifact in result["artifacts"]}
        assert {
            "raw_translated_workbook",
            "qa_final_workbook",
            "qa_report",
            "qa_result",
            "qa_changes",
            "translation_manifest",
            "glossary_snapshot",
            "prompt_snapshot",
            "project_harness_snapshot",
        }.issubset(kinds)
        final_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "qa_final_workbook")
        assert final_artifact["role"] == "translation_workbook"
        assert final_artifact["origin"] == "generated"
        metadata = result["run"]["metadata"]
        assert metadata["input_artifacts"]["source_workbook"] == source_artifact["id"]
        assert metadata["input_artifacts"]["qa_final_workbook"] == final_artifact["id"]
        assert metadata["semantic_qa"]["status"] == "skipped_no_key"
        assert metadata["translation_archive"]["imported_count"] == 5
        assert result["run"]["metadata"]["harness"]["source"] == "project_harness"
        progress = metadata["translation_progress"]
        assert progress["batch_size"] == 3
        assert progress["total_batches"] == 2
        assert progress["completed_batches"] == 2
        assert progress["percent"] == 100
        batch_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "translation" / "batches_3"
        assert (batch_dir / "batch_00001.jsonl").exists()
        assert (batch_dir / "batch_00002.jsonl").exists()
        events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("completed and persisted" in event["message"] for event in events)
        project_detail_response = client.get(f"/api/projects/{project['id']}")
        assert project_detail_response.status_code == 200
        assert project_detail_response.json()["stats"]["words"] == "33"
        assert project_detail_response.json()["stats"]["archived_rows"] == 5
        assert project_detail_response.json()["stats"]["translation_runs"] == 1
        assert project_detail_response.json()["stats"]["qa_runs"] == 0
        assert project_detail_response.json()["stats"]["langs"] == 1
        resume_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake", "batch_size": 3})
        assert resume_response.status_code == 200, resume_response.text
        resume_events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("resume: batch 1/2 already completed" in event["message"] for event in resume_events)


def test_translation_batch_retry_persists_after_transient_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = tmp_path / "retry-language.xlsx"
    _sample_workbook(workbook)
    calls = {"count": 0}

    async def flaky_translate_batch(batch, settings, project_prompt, provider_override=None, protocol_override=None):
        _ = settings, project_prompt, provider_override, protocol_override
        calls["count"] += 1
        if calls["count"] == 1:
            raise RuntimeError("transient provider failure")
        return test_fake_translate_batch(batch, settings)

    monkeypatch.setattr(workflow, "translate_batch", flaky_translate_batch)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "E2E Retry", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("retry-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": source_artifact["id"], "batch_size": 3},
        ).json()
        response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake", "batch_size": 3})
        assert response.status_code == 200, response.text
        events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("failed attempt 1/3" in event["message"] for event in events)
        assert any("batch 1/2 completed and persisted" in event["message"] for event in events)
        batch_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "translation" / "batches_3"
        assert (batch_dir / "batch_00001.jsonl").exists()
        assert not (batch_dir / "batch_00001.error.json").exists()




def test_language_table_upload_rejects_txt_with_human_message(tmp_path: Path) -> None:
    text_file = tmp_path / "notice.txt"
    text_file.write_text("announcement text", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "TXT Guard", "type": "QA"}).json()
        with text_file.open("rb") as fh:
            response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("notice.txt", fh, "text/plain")},
            )
        assert response.status_code == 400
        detail = response.json()["detail"]
        assert ".txt" in detail
        assert "XLSX" in detail
        assert "TXT" in detail
        assert "Traceback" not in detail
        assert "python.exe" not in detail
        assert "run_translation_harness" not in detail
        assert str(tmp_path) not in detail


def test_translation_rejects_legacy_txt_input_without_raw_harness_error(tmp_path: Path) -> None:
    text_file = tmp_path / "legacy.txt"
    text_file.write_text("long text to translate", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Legacy TXT Guard", "type": "QA"}).json()
        stored = workflow.copy_upload(project["id"], text_file, "legacy.txt", "asset")
        run = db.insert_run(
            project["id"],
            "translation",
            "ko",
            {"input_artifact_id": stored["id"], "task_origin": "translation_run", "batch_size": 3},
        )
        response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake", "batch_size": 3})
        assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["run"]["status"] == "needs_input"
        reason = payload["run"]["metadata"]["reason"]
        assert ".txt" in reason
        assert "XLSX" in reason
        events = client.get(f"/api/runs/{run['id']}/events").json()
        joined = "\n".join(str(event["message"]) for event in events)
        assert "command failed" not in joined
        assert "Traceback" not in joined
        assert "python.exe" not in joined
        assert "run_translation_harness" not in joined
        assert str(tmp_path) not in joined

def test_formal_translation_is_blocked_without_configured_api_key(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "小小战机", "type": "飞行射击"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        source_artifact = upload_response.json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": source_artifact["id"]},
        ).json()

        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "openai"})
        assert translate_response.status_code == 200
        result = translate_response.json()
        assert result["run"]["status"] == "needs_input"
        assert "api_key is required" in result["run"]["metadata"]["reason"]
        assert result["artifacts"] == []


def test_assets_register_role_and_origin_with_legacy_kind_mapping(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Asset Registry", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert upload_response.status_code == 200
        uploaded = upload_response.json()
        assert uploaded["role"] == "language_source"
        assert uploaded["origin"] == "uploaded"

        assets_response = client.get(f"/api/projects/{project['id']}/assets?role=language_source")
        assert assets_response.status_code == 200
        assert [asset["id"] for asset in assets_response.json()] == [uploaded["id"]]

        patch_response = client.patch(
            f"/api/artifacts/{uploaded['id']}",
            json={"role": "translation_workbook", "origin": "imported", "label": "Existing EN workbook"},
        )
        assert patch_response.status_code == 200
        assert patch_response.json()["role"] == "translation_workbook"
        assert patch_response.json()["origin"] == "imported"


def test_markdown_reference_material_uploads_and_feeds_analysis(tmp_path: Path) -> None:
    material = tmp_path / "project_brief.md"
    material.write_text(
        "# Tomorrow 2\n\nGame type: Sci-fi SLG.\nTarget players: mobile strategy players.\nGameplay: base building and alliance war.",
        encoding="utf-8",
    )

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Markdown Material", "type": "QA"}).json()
        with material.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief.md", fh, "text/markdown")},
            )
        assert upload_response.status_code == 200
        artifact = upload_response.json()
        assert artifact["kind"] == "asset"
        assert artifact["origin"] == "uploaded"
        assert artifact["mime"] == "text/markdown"
        assert Path(artifact["path"]).exists()

        notes = workflow.analyze_assets([artifact["id"]], DEFAULT_SETTINGS)
        assert notes and "text_material:" in notes[0]
        assert "Sci-fi SLG" in notes[0]

        analysis_response = client.post(
            f"/api/projects/{project['id']}/analyze",
            json={"intro": "", "asset_artifact_ids": [artifact["id"]]},
        )
        assert analysis_response.status_code == 200
        profile = analysis_response.json()["project"]["profile"]
        assert profile["asset_notes"]
        assert "Sci-fi SLG" in profile["asset_notes"][0]


def test_project_analysis_uses_configured_provider_for_semantic_profile(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    material = tmp_path / "project_brief.md"
    material.write_text(
        "# 明日2\n\n科幻 SLG，基地建设、联盟战争，语气需要硬核军事感。",
        encoding="utf-8",
    )
    calls: list[str] = []

    def fake_provider(settings: dict, prompt: str) -> str:
        calls.append(prompt)
        assert settings["provider"] == "openai"
        assert "科幻 SLG" in prompt
        return json.dumps(
            {
                "game_type": "AI识别：科幻 SLG / 战争策略",
                "target_audience": "AI识别：中重度移动策略玩家",
                "content_scope": "AI识别：基地建设、联盟战争、活动公告",
                "translation_style": "AI识别：硬核军事感，短句清晰，避免可爱化",
                "tone": "AI识别：冷静、硬核、军事化",
            },
            ensure_ascii=False,
        )

    monkeypatch.setattr(workflow, "_call_semantic_provider", fake_provider)

    with TestClient(app) as client:
        client.patch("/api/settings", json={"provider": "openai", "api_key": "test-key", "model": "gpt-test"})
        project = client.post("/api/projects", json={"name": "明日2", "type": "科幻 SLG"}).json()
        with material.open("rb") as fh:
            artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief.md", fh, "text/markdown")},
            ).json()

        response = client.post(
            f"/api/projects/{project['id']}/analyze",
            json={"intro": "面向公告和游戏内语言包。", "asset_artifact_ids": [artifact["id"]], "target_language": "en"},
        )

        assert response.status_code == 200, response.text
        assert len(calls) == 1
        payload = response.json()
        profile = payload["project"]["profile"]
        assert profile["analysis_source"] == "provider"
        assert profile["game_type"] == "AI识别：科幻 SLG / 战争策略"
        assert "硬核军事感" in payload["prompt"]


def test_duplicate_reference_material_upload_reuses_existing_artifact(tmp_path: Path) -> None:
    material = tmp_path / "project_brief.md"
    material.write_text("# Tomorrow 2\n\nSame project brief.", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Duplicate Material", "type": "QA"}).json()
        with material.open("rb") as fh:
            first_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief.md", fh, "text/markdown")},
            )
        assert first_response.status_code == 200
        first = first_response.json()
        assert first["duplicate"] is False

        with material.open("rb") as fh:
            second_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("project_brief-copy.md", fh, "text/markdown")},
            )
        assert second_response.status_code == 200
        second = second_response.json()
        assert second["duplicate"] is True
        assert second["id"] == first["id"]

        detail = client.get(f"/api/projects/{project['id']}").json()
        assets = [artifact for artifact in detail["artifacts"] if artifact["kind"] == "asset"]
        assert [artifact["id"] for artifact in assets] == [first["id"]]


def test_translation_readiness_skips_filled_translation_workbook(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Readiness Skip", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        artifact = upload_response.json()

        readiness_response = client.get(f"/api/artifacts/{artifact['id']}/translation-readiness?batch_size=90")
        assert readiness_response.status_code == 200
        readiness = readiness_response.json()
        assert readiness["ready_for_qa"] is True
        assert readiness["needs_translation"] is False
        assert readiness["source_rows"] == 5
        assert readiness["translated_rows"] == 5

        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": artifact["id"]},
        ).json()
        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake"})
        assert translate_response.status_code == 200
        result = translate_response.json()
        assert result["run"]["status"] == "needs_input"
        assert result["run"]["metadata"]["reason"] == "input already contains target translations; run QA instead"
        assert result["artifacts"] == []


def test_string_ids_run_through_translation_and_qa(tmp_path: Path) -> None:
    workbook = tmp_path / "string-id-language.xlsx"
    _string_id_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "E2E String ID", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("string-id-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        artifact = upload_response.json()

        readiness_response = client.get(f"/api/artifacts/{artifact['id']}/translation-readiness?batch_size=90")
        assert readiness_response.status_code == 200
        readiness = readiness_response.json()
        assert readiness["reason"] == "empty_target_rows"
        assert readiness["source_rows"] == 2
        assert readiness["invalid_id_rows"] == 0
        assert readiness["invalid_id_samples"] == []
        assert readiness["needs_translation"] is True
        assert readiness["ready_for_qa"] is False
        assert readiness["estimated_batches"] == 1

        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "input_artifact_id": artifact["id"]},
        ).json()
        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake"})
        assert translate_response.status_code == 200
        result = translate_response.json()
        assert result["run"]["status"] == "passed"
        assert result["run"]["metadata"]["translation_archive"]["imported_count"] == 2
        assert result["run"]["metadata"]["translation_readiness"]["invalid_id_rows"] == 0
        final_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "qa_final_workbook")

        final_wb = load_workbook(final_artifact["path"], read_only=True, data_only=True)
        try:
            rows = list(final_wb["Language"].iter_rows(min_row=2, max_row=3, values_only=True))
        finally:
            final_wb.close()
        assert rows[0][0] == "M-001"
        assert rows[0][2]

        archived = client.get(f"/api/projects/{project['id']}/translations").json()
        assert [entry["entry_key"] for entry in archived] == ["M-001", "M-002"]


def test_glossary_and_translation_archive_are_language_scoped(tmp_path: Path) -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Multilingual Terms", "type": "QA"}).json()

        en_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "战机", "target": "Warplane", "language": "en"},
        )
        ko_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "战机", "target": "전투기", "language": "ko"},
        )
        assert en_response.status_code == 200, en_response.text
        assert ko_response.status_code == 200, ko_response.text
        assert en_response.json()["id"] != ko_response.json()["id"]

        en_terms = client.get(f"/api/projects/{project['id']}/glossary?language=en").json()
        ko_terms = client.get(f"/api/projects/{project['id']}/glossary?language=ko").json()
        assert [term["target"] for term in en_terms] == ["Warplane"]
        assert [term["target"] for term in ko_terms] == ["전투기"]

        assert client.post(
            f"/api/projects/{project['id']}/translations",
            json={"entry_key": "btn.claim", "source": "领取奖励", "target": "Claim Rewards", "language": "en"},
        ).status_code == 200
        assert client.post(
            f"/api/projects/{project['id']}/translations",
            json={"entry_key": "btn.claim", "source": "领取奖励", "target": "보상 받기", "language": "ko"},
        ).status_code == 200

        en_entries = client.get(f"/api/projects/{project['id']}/translations?language=en").json()
        ko_entries = client.get(f"/api/projects/{project['id']}/translations?language=ko").json()
        assert [entry["target"] for entry in en_entries] == ["Claim Rewards"]
        assert [entry["target"] for entry in ko_entries] == ["보상 받기"]


def test_korean_fake_translation_workflow_end_to_end(tmp_path: Path) -> None:
    workbook = tmp_path / "ko-language.xlsx"
    _target_language_workbook(workbook, "KO")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "E2E Korean", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("ko-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        readiness_response = client.get(f"/api/artifacts/{source_artifact['id']}/translation-readiness?language=ko&batch_size=90")
        assert readiness_response.status_code == 200, readiness_response.text
        readiness = readiness_response.json()
        assert readiness["target_language"] == "ko"
        assert readiness["needs_translation"] is True
        assert readiness["source_rows"] == 2

        run_response = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "ko", "input_artifact_id": source_artifact["id"]},
        )
        assert run_response.status_code == 200, run_response.text
        run = run_response.json()
        translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake"})
        assert translate_response.status_code == 200, translate_response.text
        result = translate_response.json()

        assert result["run"]["status"] == "passed"
        assert result["run"]["language"] == "ko"
        assert result["run"]["metadata"]["translation_archive"]["imported_count"] == 2
        final_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "qa_final_workbook")
        wb = load_workbook(final_artifact["path"], read_only=True, data_only=True)
        try:
            assert wb["Language"].cell(2, 3).value
        finally:
            wb.close()


def test_translation_readiness_accepts_jp_and_kr_target_header_aliases(tmp_path: Path) -> None:
    jp_workbook = tmp_path / "jp-language.xlsx"
    kr_workbook = tmp_path / "kr-language.xlsx"
    _target_language_workbook(jp_workbook, "JP")
    _target_language_workbook(kr_workbook, "KR")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "JP KR Readiness", "type": "QA"}).json()

        for language, workbook in (("ja", jp_workbook), ("ko", kr_workbook)):
            with workbook.open("rb") as fh:
                source_artifact = client.post(
                    f"/api/projects/{project['id']}/files?kind=language_table",
                    files={"file": (workbook.name, fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                ).json()

            readiness_response = client.get(f"/api/artifacts/{source_artifact['id']}/translation-readiness?language={language}&batch_size=90")
            assert readiness_response.status_code == 200, readiness_response.text
            readiness = readiness_response.json()
            assert readiness["target_language"] == language
            assert readiness["source_rows"] == 2
            assert readiness["needs_translation"] is True
            assert readiness["reason"] == "empty_target_rows"


def test_project_language_workflow_accepts_non_en_kr_jp_languages(tmp_path: Path) -> None:
    fr_workbook = tmp_path / "fr-language.xlsx"
    _target_language_workbook(fr_workbook, "FR")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "FR Readiness", "type": "QA"}).json()
        with fr_workbook.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": (fr_workbook.name, fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()

        readiness_response = client.get(f"/api/artifacts/{source_artifact['id']}/translation-readiness?language=fr&batch_size=90")
        assert readiness_response.status_code == 200, readiness_response.text
        readiness = readiness_response.json()
        assert readiness["target_language"] == "fr"
        assert readiness["source_rows"] == 2
        assert readiness["needs_translation"] is True
        assert readiness["reason"] == "empty_target_rows"

        run_response = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "fr", "input_artifact_id": source_artifact["id"]},
        )
        assert run_response.status_code == 200, run_response.text
        assert run_response.json()["language"] == "fr"


def test_glossary_preview_import_and_export(tmp_path: Path) -> None:
    terms = tmp_path / "terms.xlsx"
    _sample_term_workbook(terms)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Import", "type": "QA"}).json()
        with terms.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        term_artifact = upload_response.json()

        preview_response = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": term_artifact["id"]},
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["rows"][0]["source"] == "最强指挥官"
        assert preview["rows"][0]["target"] == "Strongest Commander"
        assert preview["rows"][0]["target_alt"] == "Top Commander"
        assert preview["rows"][0]["term_key"] == "1"
        assert preview["rows"][0]["category"] == "title"

        import_response = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": term_artifact["id"]},
        )
        assert import_response.status_code == 200
        assert import_response.json()["imported_count"] == 2

        project_terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert {term["source"] for term in project_terms} == {"最强指挥官", "联盟"}
        assert {term["source_type"] for term in project_terms} == {"imported"}
        assert {term["target_alt"] for term in project_terms} == {"Top Commander", "Guild"}
        assert {term["category"] for term in project_terms} == {"title", "system"}

        manual_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "term_key": "M-1",
                "source": "战机",
                "target": "Warplane",
                "target_alt": "Fighter",
                "category": "unit",
                "note": "manual term",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert manual_response.status_code == 200
        manual = manual_response.json()
        update_response = client.patch(
            f"/api/projects/{project['id']}/glossary/{manual['id']}",
            json={"target": "Fighter Jet", "note": "edited term"},
        )
        assert update_response.status_code == 200

        export_response = client.get(f"/api/projects/{project['id']}/glossary/export?format=json")
        assert export_response.status_code == 200
        exported = export_response.json()
        assert len(exported["terms"]) == 3
        assert all("source_type" not in term and "confirmed" not in term for term in exported["terms"])
        assert any(term["source"] == "战机" and term["target"] == "Fighter Jet" and term["note"] == "edited term" for term in exported["terms"])
        xlsx_response = client.get(f"/api/projects/{project['id']}/glossary/export?format=xlsx")
        assert xlsx_response.status_code == 200
        assert "spreadsheetml" in xlsx_response.headers["content-type"]
        exported_xlsx = tmp_path / "exported_terms.xlsx"
        exported_xlsx.write_bytes(xlsx_response.content)
        wb = load_workbook(exported_xlsx, read_only=True)
        try:
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            assert headers == ["ID", "CN", "EN", "EN2", "分类", "备注"]
        finally:
            wb.close()


def test_large_language_table_is_blocked_from_project_glossary_import(tmp_path: Path) -> None:
    language_table = tmp_path / "full-language-table.xlsx"
    _large_language_table_workbook(language_table)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Guard", "type": "QA"}).json()
        with language_table.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("full-language-table.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        term_artifact = upload_response.json()

        preview_response = client.post(
            f"/api/projects/{project['id']}/glossary/import-preview",
            json={"artifact_id": term_artifact["id"], "language": "ko"},
        )
        assert preview_response.status_code == 400
        assert "完整语言表" in preview_response.json()["detail"]
        assert "高频词扫描" in preview_response.json()["detail"]

        import_response = client.post(
            f"/api/projects/{project['id']}/glossary/import",
            json={"artifact_id": term_artifact["id"], "language": "ko"},
        )
        assert import_response.status_code == 400
        assert client.get(f"/api/projects/{project['id']}/glossary?language=ko").json() == []


def test_large_language_table_upload_is_rejected_as_term_base(tmp_path: Path) -> None:
    language_table = tmp_path / "full-language-table.xlsx"
    _large_language_table_workbook(language_table)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Upload Guard", "type": "QA"}).json()
        with language_table.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("full-language-table.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert upload_response.status_code == 400
        assert "完整语言表" in upload_response.json()["detail"]
        assert client.get(f"/api/projects/{project['id']}/assets?role=glossary_source").json() == []


def test_glossary_manual_save_replaces_duplicate_cn() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Save Dedup", "type": "QA"}).json()
        created = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "term_key": "28",
                "source": "消灭怪物",
                "target": "",
                "target_alt": "",
                "category": "",
                "note": "manual candidate",
                "source_type": "manual",
                "confirmed": True,
            },
        ).json()
        db.insert_glossary_term(
            project["id"],
            {
                "term_key": "28",
                "source": " 消 灭 怪 物 ",
                "target": "Destroy Monsters",
                "target_alt": "",
                "category": "generated",
                "note": "generated duplicate",
                "source_type": "generated",
                "confirmed": False,
            },
        )

        update_response = client.patch(
            f"/api/projects/{project['id']}/glossary/{created['id']}",
            json={"source": "消灭怪物", "target": "Defeat Monsters", "category": "任务目标", "note": "saved edit"},
        )
        assert update_response.status_code == 200
        assert update_response.json()["id"] == created["id"]

        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        matching = [term for term in terms if re.sub(r"\s+", "", term["source"]) == "消灭怪物"]
        assert len(matching) == 1
        assert matching[0]["id"] == created["id"]
        assert matching[0]["target"] == "Defeat Monsters"
        assert matching[0]["category"] == "任务目标"
        assert matching[0]["note"] == "saved edit"


def test_glossary_manual_add_upserts_existing_cn() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Add Upsert", "type": "QA"}).json()
        first = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "1", "source": "钻石", "target": "Diamonds", "category": "资源", "source_type": "manual", "confirmed": True},
        ).json()

        second_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "1", "source": " 钻 石 ", "target": "Gems", "target_alt": "Diamonds", "category": "货币", "source_type": "manual", "confirmed": True},
        )
        assert second_response.status_code == 200
        assert second_response.json()["id"] == first["id"]

        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert len([term for term in terms if re.sub(r"\s+", "", term["source"]) == "钻石"]) == 1
        assert terms[0]["target"] == "Gems"
        assert terms[0]["target_alt"] == "Diamonds"
        assert terms[0]["category"] == "货币"


def test_duplicate_project_name_returns_existing_project() -> None:
    with TestClient(app) as client:
        first = client.post("/api/projects", json={"name": "明日2", "type": "科幻 SLG"}).json()
        second_response = client.post("/api/projects", json={"name": " 明日2 ", "type": "其他"})

        assert second_response.status_code == 200
        second = second_response.json()
        assert second["id"] == first["id"]
        assert second["type"] == "科幻 SLG"
        assert second["duplicate"] is True
        matching = [project for project in client.get("/api/projects").json() if project["name"] == "明日2"]
        assert len(matching) == 1


def test_glossary_backfill_preserves_existing_terms_and_logs_strategy(tmp_path: Path) -> None:
    generated = tmp_path / "generated_glossary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "战机", "Warplane", "Fighter"])
    ws.append(["G-2", "钻石", "Diamonds", "Gems"])
    ws.append(["G-3", "能量", "", ""])
    wb.save(generated)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Backfill", "type": "QA"}).json()
        client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "战机", "target": "Manual Warplane", "target_alt": "", "source_type": "manual", "confirmed": True},
        ).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()

        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])

        assert result["candidates"] == 3
        assert result["unique_candidates"] == 3
        assert result["inserted"] == 2
        assert result["updated"] == 0
        assert result["pending_confirmation"] == 2
        assert result["skipped_existing"] == 1
        assert result["skipped_duplicate"] == 0
        assert result["conflicts"] == 0
        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert len(terms) == 1
        assert terms[0]["source"] == "战机"
        assert terms[0]["target"] == "Manual Warplane"
        assert terms[0]["target_alt"] == ""
        batches = client.get(f"/api/projects/{project['id']}/glossary/batches").json()
        assert batches["active_batch"]["id"] == result["batch_id"]
        assert batches["active_batch"]["counts"]["pending"] == 2
        candidates = {candidate["source"]: candidate for candidate in batches["candidates"]}
        filled_candidate = next(candidate for candidate in batches["candidates"] if candidate["target"] == "Diamonds")
        empty_candidate = next(candidate for candidate in batches["candidates"] if not candidate["target"])
        assert filled_candidate["translation_status"] == "suggested"
        assert filled_candidate["translation_source"] == "language_table"
        assert empty_candidate["translation_status"] == "needs_translation"
        assert empty_candidate["translation_source"] == "none"
        assert "战机" not in candidates
        assert candidates["钻石"]["action"] == "new"
        assert candidates["能量"]["note"].endswith("人工确认")

        reject_response = client.post(
            f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/reject",
            json={"candidate_ids": [candidates["能量"]["id"]]},
        )
        assert reject_response.status_code == 200
        accept_response = client.post(
            f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/accept",
            json={"candidate_ids": [candidates["钻石"]["id"]]},
        )
        assert accept_response.status_code == 200
        accepted_terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        accepted_by_source = {term["source"]: term for term in accepted_terms}
        assert accepted_by_source["战机"]["target_alt"] == ""
        assert accepted_by_source["钻石"]["target"] == "Diamonds"
        assert "能量" not in accepted_by_source
        events = client.get(f"/api/runs/{run['id']}/events").json()
        assert any("Glossary backfill strategy" in event["message"] for event in events)
        assert any("inserted=2" in event["message"] and "updated=0" in event["message"] for event in events)


def test_glossary_backfill_dedupes_generated_terms_by_cn(tmp_path: Path) -> None:
    generated = tmp_path / "generated_glossary_duplicates.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "能量", "", ""])
    ws.append(["G-2", " 能 量 ", "Energy", "Power"])
    wb.save(generated)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Dedup", "type": "QA"}).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()

        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])

        assert result["candidates"] == 2
        assert result["unique_candidates"] == 1
        assert result["skipped_duplicate"] == 1
        assert result["inserted"] == 1
        terms = client.get(f"/api/projects/{project['id']}/glossary").json()
        assert terms == []
        batches = client.get(f"/api/projects/{project['id']}/glossary/batches").json()
        assert len(batches["candidates"]) == 1
        assert batches["candidates"][0]["source"] == "能量"
        assert batches["candidates"][0]["target"] == "Energy"
        assert batches["candidates"][0]["target_alt"] == "Power"


def test_glossary_accept_blocks_candidates_without_en(tmp_path: Path) -> None:
    generated = tmp_path / "generated_glossary_accept_gate.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "FILLED_TERM", "Filled Term", ""])
    ws.append(["G-2", "EMPTY_TERM", "", ""])
    wb.save(generated)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Accept Gate", "type": "QA"}).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()
        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])
        candidates = {candidate["source"]: candidate for candidate in client.get(f"/api/projects/{project['id']}/glossary/batches").json()["candidates"]}

        response = client.post(
            f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/accept",
            json={"candidate_ids": [candidates["FILLED_TERM"]["id"], candidates["EMPTY_TERM"]["id"]]},
        )

        assert response.status_code == 200
        payload = response.json()
        assert payload["resolved_count"] == 1
        assert payload["blocked_count"] == 1
        assert payload["blocked_candidates"][0]["id"] == candidates["EMPTY_TERM"]["id"]
        terms = {term["source"]: term for term in client.get(f"/api/projects/{project['id']}/glossary").json()}
        assert "FILLED_TERM" in terms
        assert "EMPTY_TERM" not in terms


def test_glossary_candidate_missing_translation_endpoint_blocks_without_real_provider(tmp_path: Path) -> None:
    generated = tmp_path / "generated_glossary_missing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "EMPTY_TERM", "", ""])
    wb.save(generated)
    wb.close()

    settings = DEFAULT_SETTINGS.copy()
    settings["provider"] = "openai"
    settings["api_key"] = ""
    save_settings(settings)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Candidate Translate", "type": "QA"}).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()
        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])

        response = client.post(f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/translate-missing")

        assert response.status_code == 400
        assert "api_key" in response.json()["detail"]
        unchanged = client.get(f"/api/projects/{project['id']}/glossary/batches").json()["candidates"][0]
        assert unchanged["target"] == ""
        assert unchanged["translation_status"] == "needs_translation"


def test_glossary_candidate_translate_missing_fills_only_blank_en(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    generated = tmp_path / "generated_glossary_partial.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2"])
    ws.append(["G-1", "FILLED_TERM", "Existing Translation", "Existing Alt"])
    ws.append(["G-2", "EMPTY_TERM", "", ""])
    wb.save(generated)
    wb.close()

    settings = DEFAULT_SETTINGS.copy()
    settings["provider"] = "openai"
    settings["api_key"] = "test-key"
    save_settings(settings)
    seen_rows: list[dict[str, object]] = []

    async def fake_translate_batch(rows: list[dict[str, object]], provider_settings: dict[str, object], prompt: str, **kwargs: object) -> list[TranslationItem]:
        _ = provider_settings, prompt, kwargs
        seen_rows.extend(rows)
        return [TranslationItem(id=int(row["id"]), translation=f"Translated {row['source']}") for row in rows]

    monkeypatch.setattr(workflow, "translate_batch", fake_translate_batch)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Glossary Candidate Fill", "type": "QA"}).json()
        run = client.post("/api/runs", json={"project_id": project["id"], "kind": "glossary", "language": "en"}).json()
        result = backfill_project_glossary_from_final(project["id"], generated, run["id"])

        response = client.post(f"/api/projects/{project['id']}/glossary/batches/{result['batch_id']}/translate-missing")

        assert response.status_code == 200
        payload = response.json()
        assert payload["translated_count"] == 1
        assert [row["source"] for row in seen_rows] == ["EMPTY_TERM"]
        candidates = {candidate["source"]: candidate for candidate in payload["candidates"]}
        assert candidates["FILLED_TERM"]["target"] == "Existing Translation"
        assert candidates["FILLED_TERM"]["target_alt"] == "Existing Alt"
        assert candidates["FILLED_TERM"]["translation_source"] == "language_table"
        assert candidates["EMPTY_TERM"]["target"] == "Translated EMPTY_TERM"
        assert candidates["EMPTY_TERM"]["target_alt"] == ""
        assert candidates["EMPTY_TERM"]["translation_source"] == "model"


def test_glossary_extract_uses_project_materials_for_brief_and_prompt(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)
    material = tmp_path / "aircraft_context.md"
    material.write_text("Project setting: aircraft, fighter jet, missile, shooter, hero gear progression.", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Material Brief", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            source_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        with material.open("rb") as fh:
            material_response = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("aircraft_context.md", fh, "text/markdown")},
            )
        source_artifact = source_response.json()
        material_artifact = material_response.json()

        extract_response = client.post(
            f"/api/projects/{project['id']}/glossary/extract",
            json={
                "input_artifact_id": source_artifact["id"],
                "project_name": "Material Brief",
                "id_column": "ID",
                "source_column": "cn",
                "target_column": "en",
                "project_material_artifact_ids": [material_artifact["id"]],
                "project_notes": ["Screenshot shows a dark sci-fi hangar and missile battle UI."],
            },
        )
        assert extract_response.status_code == 200, extract_response.text
        artifacts = {artifact["kind"]: artifact for artifact in extract_response.json()["artifacts"]}
        brief = Path(artifacts["project_brief"]["path"]).read_text(encoding="utf-8")
        prompt = Path(artifacts["translation_prompt"]["path"]).read_text(encoding="utf-8")
        assert "aircraft_context.md" in brief
        assert "\u79d1\u5e7b\u519b\u4e8b" in prompt


def test_glossary_extract_can_lookup_announcement_terms_only(tmp_path: Path) -> None:
    workbook = tmp_path / "language_table.xlsx"
    notice = tmp_path / "notice.txt"
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN", "FR"])
    ws.append(["T1", "秘境", "Trial Realm", "Royaume d'épreuve"])
    ws.append(["T2", "纹章", "Emblem", "Emblème"])
    ws.append(["T3", "商城", "Shop", "Boutique"])
    wb.save(workbook)
    wb.close()
    notice.write_text("本次更新新增秘境和纹章系统。", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement Lookup", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            source_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("language_table.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with notice.open("rb") as fh:
            notice_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": ("notice.txt", fh, "text/plain")},
            ).json()

        extract_response = client.post(
            f"/api/projects/{project['id']}/glossary/extract",
            json={
                "input_artifact_id": source_artifact["id"],
                "id_column": "ID",
                "source_column": "CN",
                "target_column": "EN",
                "announcement_material_artifact_ids": [notice_artifact["id"]],
                "announcement_only": True,
            },
        )

        assert extract_response.status_code == 200, extract_response.text
        payload = extract_response.json()
        assert payload["output"]["ANNOUNCEMENT_TERMS"] == "2"
        artifacts = {artifact["kind"]: artifact for artifact in payload["artifacts"]}
        assert set(artifacts) == {"announcement_glossary"}
        assert artifacts["announcement_glossary"]["role"] == "glossary_source"
        output_wb = load_workbook(Path(artifacts["announcement_glossary"]["path"]), read_only=True, data_only=True)
        try:
            rows = list(output_wb["Glossary"].iter_rows(values_only=True))
        finally:
            output_wb.close()
        assert rows[0] == ("ID", "CN", "EN", "FR")
        assert [row[1] for row in rows[1:]] == ["秘境", "纹章"]


def test_generic_project_prompt_does_not_inherit_warplane_style() -> None:
    with TestClient(app) as client:
        project = client.post(
            "/api/projects",
            json={
                "name": "Generic Strategy",
                "type": "\u79d1\u5e7b SLG",
                "description": "\u8054\u76df\u3001\u57fa\u5730\u3001\u6307\u6325\u5b98\u517b\u6210\u548c\u8d44\u6e90\u5faa\u73af\u3002",
            },
        ).json()

        response = client.post(
            f"/api/projects/{project['id']}/analyze",
            json={"intro": "\u901a\u7528\u79d1\u5e7b\u7b56\u7565\u9879\u76ee\uff0c\u4e3b\u7ebf\u662f\u8054\u76df\u8fd0\u8425\u548c\u57fa\u5730\u5efa\u8bbe\u3002", "asset_artifact_ids": []},
        )

        assert response.status_code == 200, response.text
        prompt = response.json()["prompt"]
        assert "\u6218\u673a" not in prompt
        assert "\u5bfc\u5f39" not in prompt
        assert "\u51c6\u786e\u7ffb\u8bd1\u4e3a\u81ea\u7136\u82f1\u6587" in prompt


def test_existing_translation_workbook_can_run_qa_without_translation_workpack(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Direct QA", "type": "QA"}).json()
        term_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "开始游戏", "target": "Start Game", "target_alt": "Begin", "category": "ui", "source_type": "manual"},
        )
        assert term_response.status_code == 200
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        translated_artifact = upload_response.json()
        assert translated_artifact["role"] == "translation_workbook"

        run_response = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": translated_artifact["id"],
            },
        )
        assert run_response.status_code == 200
        qa_response = client.post(f"/api/runs/{run_response.json()['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text
        result = qa_response.json()
        assert result["run"]["status"] == "passed"
        assert result["quality_summary"]["sources"]["translation_workbook"] == translated_artifact["id"]
        snapshot_id = result["run"]["metadata"]["input_artifacts"]["glossary_snapshot"]
        kinds = {artifact["kind"] for artifact in result["artifacts"]}
        assert "quality_summary" in kinds
        assert "qa_changes" in kinds
        assert "qa_final_workbook" in kinds
        assert "translation_workpack" not in kinds
        snapshot_artifact = next(artifact for artifact in client.get(f"/api/projects/{project['id']}").json()["artifacts"] if artifact["id"] == snapshot_id)
        snapshot_wb = load_workbook(Path(snapshot_artifact["path"]), read_only=True, data_only=True)
        try:
            rows = list(snapshot_wb.active.iter_rows(values_only=True))
            assert any(row[1] == "开始游戏" and row[2] == "Start Game" for row in rows[1:])
        finally:
            snapshot_wb.close()
        events = client.get(f"/api/runs/{run_response.json()['id']}/events").json()
        joined_events = "\n".join(str(event["message"]) for event in events)
        assert "--term-base" not in joined_events
        assert "project_glossary_snapshot.xlsx" not in joined_events
        assert "Traceback" not in joined_events
        assert "running local workflow step" in joined_events
        log_path = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run_response.json()["id"] / "logs" / "subprocess.log"
        assert log_path.exists()
        log_text = log_path.read_text(encoding="utf-8")
        assert "--term-base" in log_text
        assert "project_glossary_snapshot" in log_text
        project_detail = client.get(f"/api/projects/{project['id']}").json()
        assert int(project_detail["stats"]["words"]) > 0
        assert project_detail["stats"]["langs"] == 1
        archived = client.get(f"/api/projects/{project['id']}/translations").json()
        assert len(archived) == 5
        assert archived[0]["entry_key"] == "1"
        assert archived[0]["target"] == "Claim Rewards"


def test_language_source_with_existing_translations_can_run_direct_qa(tmp_path: Path) -> None:
    workbook = tmp_path / "translated_language_table.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Language Source QA", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={"file": ("translated_language_table.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        artifact = upload_response.json()
        assert artifact["role"] == "language_source"

        readiness = client.get(f"/api/artifacts/{artifact['id']}/translation-readiness?batch_size=90").json()
        assert readiness["ready_for_qa"] is True

        run_response = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": artifact["id"],
            },
        )
        assert run_response.status_code == 200
        qa_response = client.post(f"/api/runs/{run_response.json()['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text
        result = qa_response.json()
        assert result["run"]["status"] == "passed"
        assert result["quality_summary"]["sources"]["translation_workbook"] == artifact["id"]
        assert result["run"]["metadata"]["input_artifacts"]["translation_workbook"] == artifact["id"]
        archived = client.get(f"/api/projects/{project['id']}/translations").json()
        assert len(archived) == 5
        assert archived[1]["entry_key"] == "2"


def test_quick_task_artifacts_detect_languages_and_stay_temporary(tmp_path: Path) -> None:
    workbook = tmp_path / "quick.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "KR", "JP"])
    ws.append(["btn.claim", "领取奖励", "보상 받기", "報酬受取"])
    wb.save(workbook)
    wb.close()
    reference = tmp_path / "quick-reference.txt"
    reference.write_text("Style: concise sci-fi UI copy.", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Quick Target Detect", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            quick_input = client.post(
                f"/api/projects/{project['id']}/files?kind=quick_input",
                files={"file": ("quick.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with reference.open("rb") as fh:
            quick_reference = client.post(
                f"/api/projects/{project['id']}/files?kind=quick_reference",
                files={"file": ("quick-reference.txt", fh, "text/plain")},
            ).json()

        assert quick_input["role"] == "quick_input"
        assert quick_reference["role"] == "quick_reference"
        assert client.get(f"/api/projects/{project['id']}/assets?role=language_source").json() == []
        targets = client.get(f"/api/artifacts/{quick_input['id']}/translation-targets").json()
        assert targets["source_detected"] is True
        assert targets["detected_languages"] == ["ko", "ja"]
        assert "idn" not in targets["detected_languages"]

        run = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "ko",
                "input_artifact_id": quick_input["id"],
                "reference_artifact_ids": [quick_reference["id"]],
                "task_origin": "quick_task",
                "task_code": "QA",
            },
        ).json()
        assert run["metadata"]["task_origin"] == "quick_task"
        assert run["metadata"]["reference_artifact_ids"] == [quick_reference["id"]]


def test_quick_task_can_translate_txt_and_deliver_same_format(tmp_path: Path) -> None:
    source = tmp_path / "quick.txt"
    source.write_text("\u5f00\u59cb\u6e38\u620f\n\n\u4fdd\u5b58 {0}\n", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Quick TXT", "type": "quick-task"}).json()
        with source.open("rb") as fh:
            quick_input = client.post(
                f"/api/projects/{project['id']}/files?kind=quick_input",
                files={"file": ("quick.txt", fh, "text/plain")},
            ).json()

        targets = client.get(f"/api/artifacts/{quick_input['id']}/translation-targets").json()
        assert targets["supported_file"] is True
        assert targets["source_detected"] is True
        readiness = client.get(f"/api/artifacts/{quick_input['id']}/translation-readiness?language=ko&batch_size=1").json()
        assert readiness["source_rows"] == 2
        assert readiness["needs_translation"] is True

        run = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "translation",
                "language": "ko",
                "input_artifact_id": quick_input["id"],
                "task_origin": "quick_task",
                "batch_size": 1,
            },
        ).json()
        result = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake", "batch_size": 1}).json()
        assert result["run"]["status"] == "passed"
        final_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "final_text")
        final_text = Path(final_artifact["path"]).read_text(encoding="utf-8")
        assert "TestFake" in final_text
        assert "{0}" in final_text
        assert final_text.count("\n") == 3

        package = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={run['id']}").json()
        assert len(package["files"]) == 1
        assert package["files"][0]["filename"].endswith("_final.txt")


def test_disabled_test_fake_provider_does_not_fall_back_to_real_api(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("LWS_ENABLE_TEST_PROVIDER", raising=False)
    source = tmp_path / "quick.txt"
    source.write_text("\u5f00\u59cb\u6e38\u620f\n", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "No Fake Fallback", "type": "quick-task"}).json()
        with source.open("rb") as fh:
            quick_input = client.post(
                f"/api/projects/{project['id']}/files?kind=quick_input",
                files={"file": ("quick.txt", fh, "text/plain")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "ko", "input_artifact_id": quick_input["id"], "task_origin": "quick_task"},
        ).json()
        result = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake"}).json()
        assert result["run"]["status"] == "needs_input"
        assert "\u6d4b\u8bd5 provider \u672a\u542f\u7528" in result["run"]["metadata"]["reason"]


def test_quick_task_qa_creates_reference_snapshot(tmp_path: Path) -> None:
    workbook = tmp_path / "quick-translated.xlsx"
    _target_language_workbook(workbook, "EN", ["Claim Reward", "Welcome back, {playerName}"])
    reference = tmp_path / "style.txt"
    reference.write_text("Use short UI text.", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Quick QA", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            quick_input = client.post(
                f"/api/projects/{project['id']}/files?kind=quick_input",
                files={"file": ("quick-translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        with reference.open("rb") as fh:
            quick_reference = client.post(
                f"/api/projects/{project['id']}/files?kind=quick_reference",
                files={"file": ("style.txt", fh, "text/plain")},
            ).json()
        run = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": quick_input["id"],
                "reference_artifact_ids": [quick_reference["id"]],
                "task_origin": "quick_task",
                "task_code": "QA",
            },
        ).json()
        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text
        result = qa_response.json()
        assert result["run"]["metadata"]["task_origin"] == "quick_task"
        snapshot_id = result["run"]["metadata"]["input_artifacts"]["quick_reference_snapshot"]
        snapshot = client.get(f"/api/artifacts/{snapshot_id}/download")
        assert snapshot.status_code == 200
        assert "quick_task_reference" in snapshot.text


def test_translation_archive_import_edit_and_export(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Archive", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        imported = client.post(f"/api/projects/{project['id']}/translations/import", json={"artifact_id": artifact["id"]}).json()
        assert imported["imported_count"] == 5
        entries = client.get(f"/api/projects/{project['id']}/translations").json()
        assert [entry["entry_key"] for entry in entries[:3]] == ["1", "2", "3"]

        updated = client.patch(
            f"/api/projects/{project['id']}/translations/{entries[0]['id']}",
            json={"target": "Claim"},
        ).json()
        assert updated["target"] == "Claim"
        export_json = client.get(f"/api/projects/{project['id']}/translations/export?format=json").json()
        assert export_json["entries"][0]["target"] == "Claim"
        export_xlsx = client.get(f"/api/projects/{project['id']}/translations/export?format=xlsx")
        assert export_xlsx.status_code == 200
        assert export_xlsx.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")


def test_delivery_package_contains_only_task_outputs(tmp_path: Path) -> None:
    terms = tmp_path / "terms.xlsx"
    workbook = tmp_path / "translated.xlsx"
    _sample_term_workbook(terms)
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "小小战机", "type": "飞行射击", "description": "来源：测试文件"}).json()
        client.patch(f"/api/projects/{project['id']}", json={"prompt_text": "项目提示词：准确翻译，术语以项目术语表为准。"})
        with terms.open("rb") as fh:
            term_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("terms.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        import_response = client.post(f"/api/projects/{project['id']}/glossary/import", json={"artifact_id": term_artifact["id"]})
        assert import_response.status_code == 200
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200, qa_response.text

        deliverables_response = client.get(f"/api/projects/{project['id']}/deliverables")
        assert deliverables_response.status_code == 200
        deliverables = deliverables_response.json()["deliverables"]
        assert len(deliverables) == 1
        assert deliverables[0]["task_label"] == f"QA-{run['id'].replace('run_', '')[:6]}"
        assert deliverables[0]["processed_rows"] == 5
        assert deliverables[0]["provider"] == "rules-only"
        assert deliverables[0]["model"] == "-"
        assert deliverables[0]["status"] == "passed"

        package_response = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={run['id']}")
        assert package_response.status_code == 200, package_response.text
        package = package_response.json()
        filenames = [item["filename"] for item in package["files"]]
        assert len(filenames) == 2
        assert re.fullmatch(r"小小战机_EN_\d{12}_QA-[0-9a-f]{6}_final\.xlsx", filenames[0])
        assert re.fullmatch(r"小小战机_EN_\d{12}_QA-[0-9a-f]{6}_changes\.xlsx", filenames[1])
        assert not any(
            "input_copy" in filename
            or "manifest" in filename
            or "jsonl" in filename
            or "project_meta" in filename
            or "translation_prompt" in filename
            or "glossary" in filename
            for filename in filenames
        )
        for item in package["files"]:
            assert Path(item["path"]).exists()
            assert item["download_url"].endswith(item["filename"])
        refreshed = client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"][0]
        assert refreshed["files"]["final"]["download_url"].endswith("_final.xlsx")
        assert refreshed["files"]["changes"]["download_url"].endswith("_changes.xlsx")


def test_delivery_filename_sanitizes_invalid_project_name_without_double_spaces(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "E2E ?? Demo", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        response = client.post(f"/api/runs/{run['id']}/qa")
        assert response.status_code == 200
        package = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={run['id']}").json()
        filename = package["files"][0]["filename"]
        assert "??" not in filename
        assert "  " not in filename
        assert filename.startswith("E2E Demo_EN_")


def test_delivery_filename_uses_visible_language_code_for_korean(tmp_path: Path) -> None:
    workbook = tmp_path / "translated-kr.xlsx"
    _target_language_workbook(workbook, "KR", ["Claim Reward", "Welcome back, {playerName}"])

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "KR Delivery", "type": "QA"}).json()
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated-kr.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "ko", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        response = client.post(f"/api/runs/{run['id']}/qa")
        assert response.status_code == 200, response.text
        package = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={run['id']}").json()
        filenames = [item["filename"] for item in package["files"]]
        assert re.fullmatch(r"KR Delivery_KR_\d{12}_QA-[0-9a-f]{6}_final\.xlsx", filenames[0])
        assert re.fullmatch(r"KR Delivery_KR_\d{12}_QA-[0-9a-f]{6}_changes\.xlsx", filenames[1])
        assert not any("_KO_" in filename for filename in filenames)


def test_qa_continuation_inherits_translation_delivery_identity(tmp_path: Path) -> None:
    workbook = tmp_path / "translated.xlsx"
    _translated_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "继承任务", "type": "QA"}).json()
        source_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "translation", "language": "en", "task_code": "A"},
        ).json()
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("translated.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        qa_run = client.post(
            "/api/runs",
            json={
                "project_id": project["id"],
                "kind": "qa",
                "language": "en",
                "input_artifact_id": translated_artifact["id"],
                "task_origin": "translation_continuation",
                "source_run_id": source_run["id"],
                "task_code": "QA",
            },
        ).json()
        qa_response = client.post(f"/api/runs/{qa_run['id']}/qa")
        assert qa_response.status_code == 200

        deliverables = client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"]
        assert deliverables[0]["task_label"] == f"A-{source_run['id'].replace('run_', '')[:6]}"
        package = client.post(f"/api/projects/{project['id']}/delivery-package?run_id={qa_run['id']}").json()
        assert re.fullmatch(r"继承任务_EN_\d{12}_A-[0-9a-f]{6}_final\.xlsx", package["files"][0]["filename"])


def test_failed_runs_are_not_deliverable(tmp_path: Path) -> None:
    workbook = tmp_path / "failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "不可交付", "type": "QA"}).json()
        client.patch(f"/api/projects/{project['id']}/harness", json={"forbidden_translations": ["Forbidden Brand"]})
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200
        assert qa_response.json()["run"]["status"] == "failed"
        assert client.get(f"/api/projects/{project['id']}/deliverables").json()["deliverables"] == []


def test_failed_qa_exposes_normalized_project_harness_rows(tmp_path: Path) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Failed Row QA", "type": "QA"}).json()
        harness_response = client.patch(
            f"/api/projects/{project['id']}/harness",
            json={"forbidden_translations": ["Forbidden Brand"]},
        )
        assert harness_response.status_code == 200
        with workbook.open("rb") as fh:
            upload_response = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        translated_artifact = upload_response.json()
        run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"]},
        ).json()

        qa_response = client.post(f"/api/runs/{run['id']}/qa")
        assert qa_response.status_code == 200
        assert qa_response.json()["run"]["status"] == "failed"

        issues_response = client.get(f"/api/runs/{run['id']}/quality-issues")
        assert issues_response.status_code == 200
        issues = issues_response.json()["issues"]
        assert issues[0]["source"] == "project_harness"
        assert issues[0]["rule_source"] == "project_harness"
        assert issues[0]["severity"] == "hard"
        assert issues[0]["sheet"] == "Language"
        assert issues[0]["row"] == 2
        assert issues[0]["current_translation"] == "Forbidden Brand Reward"


def test_manual_fix_creates_fixed_workbook_reruns_qa_and_updates_project_harness(tmp_path: Path) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Manual Fix QA", "type": "QA"}).json()
        client.patch(
            f"/api/projects/{project['id']}/harness",
            json={"forbidden_translations": ["Forbidden Brand"]},
        )
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        failed_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"]},
        ).json()
        failed_qa = client.post(f"/api/runs/{failed_run['id']}/qa").json()
        assert failed_qa["run"]["status"] == "failed"

        fix_response = client.post(
            f"/api/runs/{failed_run['id']}/manual-fixes",
            json={
                "fixes": [
                    {
                        "sheet": "Language",
                        "row": 2,
                        "translation": "Reward",
                        "note": "Remove forbidden project-specific wording.",
                    }
                ],
                "rerun_qa": True,
            },
        )
        assert fix_response.status_code == 200, fix_response.text
        result = fix_response.json()
        assert result["fixed_artifact"]["role"] == "translation_workbook"
        assert result["fixed_artifact"]["origin"] == "manual"
        assert result["qa_result"]["run"]["status"] == "passed"
        assert result["qa_result"]["quality_summary"]["sources"]["manual_fix_source_run"] == failed_run["id"]
        assert any(artifact["kind"] == "qa_changes" for artifact in result["qa_result"]["artifacts"])

        harness = client.get(f"/api/projects/{project['id']}/harness").json()["project_harness"]
        assert harness["manual_fixes"][0]["previous_translation"] == "Forbidden Brand Reward"
        assert harness["manual_fixes"][0]["translation"] == "Reward"

        suggestions = client.get(f"/api/projects/{project['id']}/improvements").json()
        assert any(item["category"] == "project_harness" and item["run_id"] == failed_run["id"] for item in suggestions)


def test_model_fix_requires_configured_provider(tmp_path: Path) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    with TestClient(app) as client:
        client.patch("/api/settings", json={"provider": "openai", "api_key": ""})
        project = client.post("/api/projects", json={"name": "Model Fix Needs Key", "type": "QA"}).json()
        client.patch(f"/api/projects/{project['id']}/harness", json={"forbidden_translations": ["Forbidden Brand"]})
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        failed_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"]},
        ).json()
        assert client.post(f"/api/runs/{failed_run['id']}/qa").json()["run"]["status"] == "failed"

        response = client.post(f"/api/runs/{failed_run['id']}/model-fixes", json={"max_issues": 20, "rerun_qa": True})
        assert response.status_code == 409
        assert "API key" in response.json()["detail"]


def test_model_fix_applies_provider_suggestions_and_reruns_qa(tmp_path: Path, monkeypatch) -> None:
    workbook = tmp_path / "project-failed.xlsx"
    _project_harness_failed_workbook(workbook)

    def fake_model(settings: dict, prompt: str) -> str:
        if "待修复行" not in prompt:
            return '{"passed": true, "issues": []}'
        return '{"fixes":[{"issue_id":"project_harness:0:Language:2:forbidden_translation","sheet":"Language","row":2,"translation":"Reward","note":"remove forbidden phrase"}]}'

    monkeypatch.setattr(workflow, "_call_semantic_provider", fake_model)
    with TestClient(app) as client:
        client.patch("/api/settings", json={"provider": "openai", "api_key": "test-key", "model": "gpt-test"})
        project = client.post("/api/projects", json={"name": "Model Fix QA", "type": "QA"}).json()
        client.patch(f"/api/projects/{project['id']}/harness", json={"forbidden_translations": ["Forbidden Brand"]})
        with workbook.open("rb") as fh:
            translated_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("project-failed.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        failed_run = client.post(
            "/api/runs",
            json={"project_id": project["id"], "kind": "qa", "language": "en", "input_artifact_id": translated_artifact["id"], "task_code": "QA"},
        ).json()
        assert client.post(f"/api/runs/{failed_run['id']}/qa").json()["run"]["status"] == "failed"

        response = client.post(f"/api/runs/{failed_run['id']}/model-fixes", json={"max_issues": 20, "rerun_qa": True})
        assert response.status_code == 200, response.text
        result = response.json()
        assert result["fixed_artifact"]["origin"] == "provider"
        assert result["model_fixes"][0]["translation"] == "Reward"
        assert result["model_fixes"][0]["rule_source"] == "model_fix"
        assert result["qa_result"]["run"]["status"] == "passed"
        assert result["qa_result"]["quality_summary"]["sources"]["model_fix_source_run"] == failed_run["id"]


def test_project_harness_is_project_scoped_and_affects_only_its_run(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        first = client.post(
            "/api/projects",
            json={"name": "Harness A", "type": "科幻 SLG", "icon": "A", "description": "Project A."},
        ).json()
        second = client.post(
            "/api/projects",
            json={"name": "Harness B", "type": "科幻 SLG", "icon": "B", "description": "Project B."},
        ).json()

        harness_response = client.patch(
            f"/api/projects/{first['id']}/harness",
            json={
                "style_guidance": "Use tactical aviation wording for this project only.",
                "target_audience": "Core strategy players",
                "forbidden_translations": ["TestFake 1"],
                "hard_rules": [{"label": "No raw fake marker", "description": "Test fake marker must not ship."}],
            },
        )
        assert harness_response.status_code == 200
        assert harness_response.json()["project_harness"]["style_guidance"].startswith("Use tactical")

        first_result = _run_fake_translation(client, first["id"], workbook)
        second_result = _run_fake_translation(client, second["id"], workbook)

        assert first_result["run"]["status"] == "failed"
        assert first_result["project_harness_quality"]["hard_errors"] == 1
        assert first_result["project_harness_quality"]["issues"][0]["rule_source"] == "project_harness"
        assert second_result["run"]["status"] == "passed"
        assert second_result["project_harness_quality"]["hard_errors"] == 0
        assert second_result["run"]["metadata"]["harness"]["forbidden_translations"] == 0

        semantic_response = client.post(f"/api/runs/{second_result['run']['id']}/semantic-qa")
        assert semantic_response.status_code == 200
        assert semantic_response.json()["semantic_qa"]["status"] == "needs_model_review"
        assert semantic_response.json()["artifact"]["kind"] == "semantic_qa_context"


def test_improvement_review_writes_suggestions_without_auto_merge(tmp_path: Path) -> None:
    workbook = tmp_path / "sample-language.xlsx"
    _sample_workbook(workbook)

    with TestClient(app) as client:
        project = client.post(
            "/api/projects",
            json={"name": "Harness Improvement", "type": "科幻 SLG", "icon": "I", "description": "Project."},
        ).json()
        client.patch(
            f"/api/projects/{project['id']}/harness",
            json={"forbidden_translations": ["TestFake 1"]},
        )
        result = _run_fake_translation(client, project["id"], workbook)

        review_response = client.post(f"/api/runs/{result['run']['id']}/improvement-review")
        assert review_response.status_code == 200
        categories = {item["category"] for item in review_response.json()["suggestions"]}
        assert {"project_harness", "upstream_backfeed"}.issubset(categories)

        queue_response = client.get(f"/api/projects/{project['id']}/improvements")
        assert queue_response.status_code == 200
        assert all(item["status"] == "pending_review" for item in queue_response.json())


def test_glossary_and_translation_wide_views_group_by_cn() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Wide Multilingual Assets", "type": "RPG"}).json()
        client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "T-1", "source": "战机", "target": "Warplane", "target_alt": "Fighter", "language": "en", "category": "unit", "note": "core"},
        )
        client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "K-1", "source": "战机", "target": "전투기", "language": "ko", "category": "기체", "note": "korean"},
        )
        client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"term_key": "J-1", "source": "战机", "target": "戦闘機", "language": "ja", "category": "unit", "note": "core"},
        )
        client.post(
            f"/api/projects/{project['id']}/translations",
            json={"entry_key": "A-1", "source": "领取奖励", "target": "Claim Rewards", "target_alt": "Claim", "language": "en", "note": "button"},
        )
        client.post(
            f"/api/projects/{project['id']}/translations",
            json={"entry_key": "A-1", "source": "领取奖励", "target": "보상 받기", "language": "ko", "note": "button"},
        )

        glossary_response = client.get(f"/api/projects/{project['id']}/glossary/wide")
        assert glossary_response.status_code == 200, glossary_response.text
        glossary = glossary_response.json()
        assert glossary["languages"] == ["en", "ko", "ja"]
        assert glossary["coverage"] == {"en": 1, "ko": 1, "ja": 1}
        assert glossary["row_count"] == 1
        row = glossary["rows"][0]
        assert row["source"] == "战机"
        assert row["translations"]["en"]["target"] == "Warplane"
        assert row["translations"]["en"]["target_alt"] == "Fighter"
        assert row["translations"]["ko"]["target"] == "전투기"
        assert row["translations"]["ko"].get("target_alt", "") == ""
        assert {item["field"] for item in row["conflicts"]} >= {"term_key", "category", "note"}

        translations_response = client.get(f"/api/projects/{project['id']}/translations/wide")
        assert translations_response.status_code == 200, translations_response.text
        translations = translations_response.json()
        assert translations["languages"] == ["en", "ko"]
        assert translations["coverage"] == {"en": 1, "ko": 1}
        translated = translations["rows"][0]
        assert translated["source"] == "领取奖励"
        assert translated["translations"]["en"]["target_alt"] == "Claim"
        assert translated["translations"]["ko"].get("target_alt", "") == ""

        legacy_response = client.get(f"/api/projects/{project['id']}/glossary?language=ko")
        assert legacy_response.status_code == 200
        assert legacy_response.json()[0]["target"] == "전투기"


def test_multilingual_glossary_and_archive_import_once_into_wide_views(tmp_path: Path) -> None:
    glossary_path = tmp_path / "terms-multi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", "EN", "EN2", "KO", "JA", "FR", "DE", "ID", "TH", "AR", "分类", "备注"])
    ws.append(["T-1", "战机", "Warplane", "Fighter", "전투기", "戦闘機", "Avion de chasse", "Kampfflugzeug", "Pesawat Tempur", "เครื่องบินรบ", "طائرة مقاتلة", "unit", "core term"])
    ws.append(["T-2", "钻石", "Diamonds", "", "다이아몬드", "", "Diamants", "", "", "", "", "currency", ""])
    wb.save(glossary_path)
    wb.close()

    archive_path = tmp_path / "archive-multi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"
    ws.append(["ID", "CN", "EN", "EN2", "KO", "JA", "FR", "DE", "ID", "TH", "AR", "备注"])
    ws.append(["A-1", "领取奖励", "Claim Rewards", "Claim", "보상 받기", "報酬を受け取る", "Recevoir les récompenses", "Belohnung abholen", "Klaim Hadiah", "รับรางวัล", "استلام المكافآت", "button"])
    wb.save(archive_path)
    wb.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Auto Import Multilingual", "type": "RPG"}).json()
        with glossary_path.open("rb") as fh:
            glossary_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=term_base",
                files={"file": ("terms-multi.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        glossary_import = client.post(f"/api/projects/{project['id']}/glossary/import", json={"artifact_id": glossary_artifact["id"]})
        assert glossary_import.status_code == 200, glossary_import.text
        assert glossary_import.json()["imported_count"] == 11

        ko_terms = client.get(f"/api/projects/{project['id']}/glossary?language=ko").json()
        assert {term["source"]: term["target"] for term in ko_terms} == {"战机": "전투기", "钻石": "다이아몬드"}
        assert all(term["target_alt"] == "" for term in ko_terms)
        glossary_wide = client.get(f"/api/projects/{project['id']}/glossary/wide").json()
        assert glossary_wide["languages"] == ["en", "ko", "ja", "fr", "de", "idn", "th", "ar"]
        row = next(item for item in glossary_wide["rows"] if item["source"] == "战机")
        assert row["translations"]["en"]["target_alt"] == "Fighter"
        assert row["translations"]["ko"]["target"] == "전투기"
        assert row["translations"]["ja"]["target"] == "戦闘機"
        assert row["translations"]["fr"]["target"] == "Avion de chasse"
        assert row["translations"]["de"]["target"] == "Kampfflugzeug"
        assert row["translations"]["idn"]["target"] == "Pesawat Tempur"
        assert row["translations"]["th"]["target"] == "เครื่องบินรบ"
        assert row["translations"]["ar"]["target"] == "طائرة مقاتلة"

        with archive_path.open("rb") as fh:
            archive_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=final_workbook",
                files={"file": ("archive-multi.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            ).json()
        archive_import = client.post(f"/api/projects/{project['id']}/translations/import", json={"artifact_id": archive_artifact["id"]})
        assert archive_import.status_code == 200, archive_import.text
        assert archive_import.json()["imported_count"] == 8
        archive_wide = client.get(f"/api/projects/{project['id']}/translations/wide").json()
        assert archive_wide["languages"] == ["en", "ko", "ja", "fr", "de", "idn", "th", "ar"]
        archive_row = archive_wide["rows"][0]
        assert archive_row["translations"]["en"]["target_alt"] == "Claim"
        assert archive_row["translations"]["ko"].get("target_alt", "") == ""
        assert archive_row["translations"]["ja"]["target"] == "報酬を受け取る"
        assert archive_row["translations"]["fr"]["target"] == "Recevoir les récompenses"
        assert archive_row["translations"]["de"]["target"] == "Belohnung abholen"
        assert archive_row["translations"]["idn"]["target"] == "Klaim Hadiah"
        assert archive_row["translations"]["th"]["target"] == "รับรางวัล"
        assert archive_row["translations"]["ar"]["target"] == "استلام المكافآت"
        glossary_export = workflow.export_glossary(project["id"], "xlsx")
        assert isinstance(glossary_export, Path)
        assert re.fullmatch(r"Auto Import Multilingual_glossary_ALL_\d{8}\.xlsx", glossary_export.name)
        wb = load_workbook(glossary_export, read_only=True, data_only=True)
        try:
            headers = [cell.value for cell in next(wb["Glossary"].iter_rows(min_row=1, max_row=1))]
            assert headers == ["ID", "CN", "EN", "EN2", "KR", "JP", "FR", "DE", "IDN", "TH", "AR", "分类", "备注"]
        finally:
            wb.close()

        ko_glossary_export = workflow.export_glossary(project["id"], "xlsx", language="ko")
        assert isinstance(ko_glossary_export, Path)
        assert re.fullmatch(r"Auto Import Multilingual_glossary_KR_\d{8}\.xlsx", ko_glossary_export.name)
        wb = load_workbook(ko_glossary_export, read_only=True, data_only=True)
        try:
            headers = [cell.value for cell in next(wb["Glossary"].iter_rows(min_row=1, max_row=1))]
            assert headers == ["ID", "CN", "KR", "分类", "备注"]
        finally:
            wb.close()

        archive_export = workflow.export_translation_archive(project["id"], "xlsx")
        assert isinstance(archive_export, Path)
        assert re.fullmatch(r"Auto Import Multilingual_translations_ALL_\d{8}\.xlsx", archive_export.name)
        wb = load_workbook(archive_export, read_only=True, data_only=True)
        try:
            headers = [cell.value for cell in next(wb["Translations"].iter_rows(min_row=1, max_row=1))]
            assert headers == ["ID", "CN", "EN", "EN2", "KR", "JP", "FR", "DE", "IDN", "TH", "AR", "备注"]
        finally:
            wb.close()


def test_announcement_lookup_uses_glossary_and_qa_passed_archive(tmp_path: Path) -> None:
    notice = tmp_path / "notice.txt"
    notice.write_text("本次更新新增秘境玩法，并开放纹章系统。", encoding="utf-8")

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement Lookup", "type": "RPG"}).json()
        term_response = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={"source": "秘境", "target": "Trial Realm", "language": "en", "category": "system"},
        )
        assert term_response.status_code == 200
        archive_response = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "archive-1",
                "source": "新增秘境玩法",
                "target": "New Trial Realm gameplay",
                "language": "en",
                "sheet": "Archive",
                "row_number": 12,
                "source_type": "qa_passed",
            },
        )
        assert archive_response.status_code == 200
        client.post(
            f"/api/projects/{project['id']}/translations",
            json={"entry_key": "archive-2", "source": "商城", "target": "Shop", "language": "en"},
        )
        with notice.open("rb") as fh:
            material = client.post(f"/api/projects/{project['id']}/files?kind=asset", files={"file": ("notice.txt", fh, "text/plain")}).json()

        response = client.post(
            f"/api/projects/{project['id']}/announcement-lookup",
            json={"material_artifact_ids": [material["id"]], "language": "en"},
        )
        assert response.status_code == 200, response.text
        result = response.json()

        assert result["run"]["kind"] == "announcement_lookup"
        assert result["run"]["status"] == "passed"
        assert result["summary"]["matched_terms"] == 1
        assert result["summary"]["matched_translations"] == 1
        kinds = {artifact["kind"] for artifact in result["artifacts"]}
        assert {"announcement_lookup_workbook", "announcement_lookup_manifest", "announcement_lookup_prompt_context"}.issubset(kinds)
        assert result["manifest"]["language"] == "en"
        assert result["manifest"]["matched_terms"][0]["target"] == "Trial Realm"
        assert result["manifest"]["matched_translations"][0]["target"] == "New Trial Realm gameplay"

        workbook_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "announcement_lookup_workbook")
        wb = load_workbook(Path(workbook_artifact["path"]), read_only=True, data_only=True)
        assert {"Overview", "MatchedTerms", "MatchedTranslations", "PromptContext"}.issubset(set(wb.sheetnames))
        assert wb["MatchedTerms"]["A2"].value == "秘境"
        assert wb["MatchedTerms"]["B2"].value == "Trial Realm"
        assert wb["MatchedTranslations"]["B2"].value == "新增秘境玩法"
        wb.close()

        prompt_artifact = next(artifact for artifact in result["artifacts"] if artifact["kind"] == "announcement_lookup_prompt_context")
        prompt_text = Path(prompt_artifact["path"]).read_text(encoding="utf-8")
        assert "Trial Realm" in prompt_text
        assert "New Trial Realm gameplay" in prompt_text
        assert "不含正文译文" in prompt_text


def test_announcement_lookup_is_language_scoped_and_does_not_mutate_terms_or_archive() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement Language Scope", "type": "STG"}).json()
        client.post(f"/api/projects/{project['id']}/glossary", json={"source": "战机", "target": "Warplane", "language": "en"})
        client.post(f"/api/projects/{project['id']}/glossary", json={"source": "战机", "target": "전투기", "language": "ko"})
        client.post(f"/api/projects/{project['id']}/translations", json={"source": "战机升级", "target": "Warplane Upgrade", "language": "en"})
        client.post(f"/api/projects/{project['id']}/translations", json={"source": "战机升级", "target": "전투기 업그레이드", "language": "ko"})
        before_terms = len(client.get(f"/api/projects/{project['id']}/glossary").json())
        before_entries = len(client.get(f"/api/projects/{project['id']}/translations").json())

        response = client.post(
            f"/api/projects/{project['id']}/announcement-lookup",
            json={"text": "战机升级活动开启。", "language": "ko"},
        )
        assert response.status_code == 200, response.text
        manifest = response.json()["manifest"]
        assert manifest["matched_terms"][0]["target"] == "전투기"
        assert manifest["matched_translations"][0]["target"] == "전투기 업그레이드"
        assert all(row["language"] == "ko" for row in manifest["matched_terms"])
        assert all(row["language"] == "ko" for row in manifest["matched_translations"])
        assert "Warplane" not in json.dumps(manifest, ensure_ascii=False)
        assert len(client.get(f"/api/projects/{project['id']}/glossary").json()) == before_terms
        assert len(client.get(f"/api/projects/{project['id']}/translations").json()) == before_entries


def test_announcement_lookup_requires_text_or_material() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement Empty Input", "type": "QA"}).json()
        response = client.post(f"/api/projects/{project['id']}/announcement-lookup", json={"language": "ja"})
        assert response.status_code == 400
        assert "requires" in response.json()["detail"]


def test_announcement_lookup_generates_empty_pack_when_no_hits() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Announcement No Hits", "type": "QA"}).json()
        response = client.post(
            f"/api/projects/{project['id']}/announcement-lookup",
            json={"text": "这是完全没有项目术语命中的公告正文。", "language": "ja"},
        )
        assert response.status_code == 200, response.text
        result = response.json()
        assert result["summary"]["matched_terms"] == 0
        assert result["summary"]["matched_translations"] == 0
        assert result["summary"]["constraint_status"] == "missing"
        assert any(artifact["kind"] == "announcement_lookup_prompt_context" for artifact in result["artifacts"])


def _run_fake_translation(client: TestClient, project_id: str, workbook: Path) -> dict:
    with workbook.open("rb") as fh:
        upload_response = client.post(
            f"/api/projects/{project_id}/files?kind=language_table",
            files={"file": ("sample-language.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload_response.status_code == 200
    source_artifact = upload_response.json()
    run_response = client.post(
        "/api/runs",
        json={
            "project_id": project_id,
            "kind": "translation",
            "language": "en",
            "input_artifact_id": source_artifact["id"],
            "batch_size": 3,
        },
    )
    assert run_response.status_code == 200
    run = run_response.json()
    translate_response = client.post(f"/api/runs/{run['id']}/translate", json={"provider": "test-fake"})
    assert translate_response.status_code == 200, translate_response.text
    return translate_response.json()


def test_project_prompt_patch_mirrors_display_and_execution_prompts() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Prompt Mirror", "type": "QA"}).json()
        manual_prompt = "\u4eba\u5de5\u4fee\u8ba2\u9879\u76ee\u63d0\u793a\u8bcd\uff1a\u672f\u8bed\u4f18\u5148\uff0cUI \u8868\u8fbe\u7b80\u6d01\u3002"
        response = client.patch(
            f"/api/projects/{project['id']}",
            json={"prompt_text": manual_prompt},
        )
        assert response.status_code == 200, response.text
        updated = response.json()
        assert updated["prompt_text"] == manual_prompt
        assert updated["profile"]["prompts_by_language"]["en"] == manual_prompt
        assert updated["profile"]["display_prompts_by_language"]["en"] == manual_prompt

        jp_prompt = "\u65e5\u8bed\u5c55\u793a\u63d0\u793a\u8bcd\uff1a\u4fdd\u6301\u9879\u76ee\u672f\u8bed\u4e00\u81f4\u3002"
        profile = dict(updated["profile"])
        profile["prompts_by_language"] = {**profile["prompts_by_language"], "ja": jp_prompt}
        profile["display_prompts_by_language"] = {**profile["display_prompts_by_language"], "ja": jp_prompt}
        response = client.patch(f"/api/projects/{project['id']}", json={"profile": profile})
        assert response.status_code == 200, response.text
        updated = response.json()
        assert updated["profile"]["prompts_by_language"]["ja"] == jp_prompt
        assert updated["profile"]["display_prompts_by_language"]["ja"] == jp_prompt
