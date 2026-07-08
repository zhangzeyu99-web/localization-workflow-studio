"""Guard against top-level name shadowing across app.workflow submodules.

backend/app/workflow/__init__.py merges every submodule's globals into one
shared namespace and injects it back into each submodule. When two submodules
define the same top-level name, the later-loaded module silently wins in every
module (see the _MODULES load order in __init__.py). This already caused a
real bug: qa._cell_text (no strip) shadowed delivery._cell_text (strip),
breaking whitespace-tolerant ID matching in merged deliveries.

This test scans all injected submodules with ast and fails when a new
duplicate top-level definition appears outside the reviewed whitelist below.
"""
from __future__ import annotations

import ast
from collections import defaultdict
from pathlib import Path

WORKFLOW_DIR = Path(__file__).resolve().parents[1] / "app" / "workflow"

# Reviewed duplicates that are safe to keep. Every entry documents why the
# shadowing is harmless. "identical" entries are additionally verified below
# to still have identical source, so semantic drift breaks this test.
#
# name -> (frozenset of modules defining it, kind, reason)
#   kind "identical": exact same definition in both modules; winner is
#       interchangeable with the loser.
#   kind "delegate": one side is a thin wrapper delegating to the other's
#       implementation; behaviour is the same whichever wins.
ALLOWED_DUPLICATES: dict[str, tuple[frozenset[str], str, str]] = {
    "_harness_summary": (
        frozenset({"translation_readiness", "qa"}),
        "identical",
        "verbatim copies of the same summary helper",
    ),
    "_safe_delivery_name": (
        frozenset({"naming", "delivery"}),
        "identical",
        "both are one-line wrappers around delivery_naming.safe_delivery_name",
    ),
    "_safe_source_stem": (
        frozenset({"announcement_outputs", "naming"}),
        "identical",
        "both are one-line wrappers around delivery_naming.source_stem",
    ),
    "_today_stamp": (
        frozenset({"announcement_outputs", "naming"}),
        "identical",
        "verbatim copies of the same date-stamp helper",
    ),
    "_visible_language_code": (
        frozenset({"announcement_outputs", "naming"}),
        "identical",
        "both are one-line wrappers around languages.visible_language_code",
    ),
    "read_jsonl": (
        frozenset({"jsonl_helpers", "glossary"}),
        "delegate",
        "glossary.read_jsonl delegates to jsonl_helpers.read_jsonl",
    ),
    "write_jsonl": (
        frozenset({"jsonl_helpers", "glossary"}),
        "delegate",
        "glossary.write_jsonl delegates to jsonl_helpers.write_jsonl",
    ),
    "LANGUAGE_ORDER": (
        frozenset({"common", "table_helpers"}),
        "identical",
        "same constant duplicated in both modules",
    ),
    "AUTO_LANGUAGE_TARGET_ALIASES": (
        frozenset({"common", "table_helpers"}),
        "identical",
        "same constant duplicated in both modules",
    ),
    "AUTO_LANGUAGE_ALT_ALIASES": (
        frozenset({"common", "table_helpers"}),
        "identical",
        "same constant duplicated in both modules",
    ),
}


def _injected_module_names() -> list[str]:
    """Read the _MODULES load order straight from __init__.py."""
    tree = ast.parse((WORKFLOW_DIR / "__init__.py").read_text(encoding="utf-8"))
    for node in tree.body:
        if isinstance(node, ast.Assign):
            targets = [t.id for t in node.targets if isinstance(t, ast.Name)]
            if "_MODULES" in targets and isinstance(node.value, ast.List):
                return [
                    elt.id for elt in node.value.elts if isinstance(elt, ast.Name)
                ]
    raise AssertionError("_MODULES list not found in app/workflow/__init__.py")


def _top_level_definitions(module: str) -> dict[str, str]:
    """Map top-level name -> unparsed source of its definition."""
    tree = ast.parse((WORKFLOW_DIR / f"{module}.py").read_text(encoding="utf-8"))
    result: dict[str, str] = {}
    for node in tree.body:
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
            names = [node.name]
        elif isinstance(node, ast.Assign):
            names = [t.id for t in node.targets if isinstance(t, ast.Name)]
        elif isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
            names = [node.target.id]
        else:
            continue
        for name in names:
            # The injection loop in __init__.py skips dunder-prefixed names.
            if name.startswith("__"):
                continue
            result[name] = ast.unparse(node)
    return result


def test_no_unreviewed_top_level_shadowing() -> None:
    modules = _injected_module_names()
    definitions: dict[str, dict[str, str]] = {
        module: _top_level_definitions(module) for module in modules
    }
    owners: dict[str, list[str]] = defaultdict(list)
    for module in modules:
        for name in definitions[module]:
            owners[name].append(module)

    duplicated = {name: mods for name, mods in owners.items() if len(mods) > 1}

    unreviewed = {
        name: mods
        for name, mods in duplicated.items()
        if name not in ALLOWED_DUPLICATES
        or frozenset(mods) != ALLOWED_DUPLICATES[name][0]
    }
    assert not unreviewed, (
        "Unreviewed top-level name shadowing across app.workflow submodules. "
        "The shared-namespace injection in __init__.py makes the later-loaded "
        "definition silently win everywhere. Rename one side (e.g. the "
        "delivery._cell_text -> _delivery_cell_text fix) or, if genuinely "
        f"harmless, add it to ALLOWED_DUPLICATES with a reason: {unreviewed}"
    )

    stale = {
        name: sorted(expected_mods)
        for name, (expected_mods, _, _) in ALLOWED_DUPLICATES.items()
        if name not in duplicated
    }
    assert not stale, f"ALLOWED_DUPLICATES entries no longer duplicated, remove them: {stale}"


def test_whitelisted_identical_duplicates_stay_identical() -> None:
    for name, (mods, kind, _reason) in ALLOWED_DUPLICATES.items():
        if kind != "identical":
            continue
        sources = {_top_level_definitions(module)[name] for module in mods}
        assert len(sources) == 1, (
            f"{name} is whitelisted as identical across {sorted(mods)} but the "
            "definitions diverged; the namespace injection would silently pick "
            "one of them. Re-align the copies or rename one side."
        )


def test_merge_language_column_matches_ids_with_whitespace(tmp_path: Path) -> None:
    """Regression: the shadowing bug made merged deliveries drop rows whose IDs
    carried stray whitespace, because qa._cell_text (no strip) replaced
    delivery._cell_text (strip). Importing the full package here reproduces the
    injection before exercising the merge."""
    from openpyxl import Workbook, load_workbook

    import app.workflow  # noqa: F401  # trigger the shared-namespace injection
    from app.languages import visible_language_code
    from app.workflow import delivery

    visible = visible_language_code("en")

    target_path = tmp_path / "target.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", visible])
    ws.append([" 1001 ", "开始游戏", ""])  # whitespace on the delivery side
    ws.append(["1002", "退出游戏", ""])
    wb.save(target_path)
    wb.close()

    source_path = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", visible])
    ws.append(["1001", "Start Game"])
    ws.append([" 1002 ", "Quit Game"])  # whitespace on the source side
    wb.save(source_path)
    wb.close()

    copied = delivery._merge_language_column(target_path, source_path, "en")
    assert copied == 2

    wb = load_workbook(target_path, data_only=True)
    try:
        ws = wb["Language"]
        assert ws.cell(row=2, column=3).value == "Start Game"
        assert ws.cell(row=3, column=3).value == "Quit Game"
    finally:
        wb.close()
