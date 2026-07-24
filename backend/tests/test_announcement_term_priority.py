from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.db as db
from app.config import DEFAULT_SETTINGS, save_settings
from app.main import app
from app.workflow.announcement_outputs import _announcement_segment_term_hits
from app.workflow.announcement_segments import _select_announcement_constraint_rows
from conftest import reset_data_root, wait_for_background_jobs


@pytest.fixture(autouse=True)
def reset_test_state() -> None:
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    reset_data_root(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    wait_for_background_jobs()
    save_settings(DEFAULT_SETTINGS)


def _create_game_name_task(client: TestClient, project_id: str, *, include_project_archive: bool) -> dict:
    response = client.post(
        f"/api/projects/{project_id}/announcement-tasks",
        json={
            "text": "《菇勇者传说》联动公告",
            "languages": ["en"],
            "include_project_archive": include_project_archive,
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def _extract_terms(client: TestClient, task_id: str, *, include_project_archive: bool) -> dict:
    response = client.post(
        f"/api/announcement-tasks/{task_id}/extract-terms",
        json={
            "languages": ["en"],
            "include_project_archive": include_project_archive,
            "ai_supplement": False,
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def _write_multilingual_constraint_table(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CN", "EN", "FR"])
    sheet.append(["菇勇者传说", "", "Légende du Champignon"])
    sheet.append(["联动公告", "", "Annonce de collaboration"])
    workbook.save(path)
    workbook.close()


def test_long_term_suppresses_only_overlapping_short_term_occurrences() -> None:
    notice = "《菇勇者传说》联动公告\n菇勇者们，请查收联动奖励。"
    candidates = [
        {
            "id": "game-name",
            "source": "菇勇者传说",
            "translations": {"en": "Legend of Mushroom"},
        },
        {
            "id": "nickname",
            "source": "菇勇者",
            "translations": {"en": "Shroomie"},
        },
    ]

    selected = _select_announcement_constraint_rows(notice, candidates, ["en"], min_hit=1)

    assert [
        (row["source"], row["hit_count"], row["first_position"])
        for row in selected
    ] == [
        ("菇勇者传说", 1, notice.index("菇勇者传说")),
        ("菇勇者", 1, notice.rindex("菇勇者")),
    ]

    lookup = {
        "en": {
            "terms": [
                {
                    "source": row["source"],
                    "target": row["translations"]["en"],
                }
                for row in selected
            ]
        }
    }
    assert _announcement_segment_term_hits(
        {"source": "《菇勇者传说》联动公告"},
        "en",
        lookup,
    ) == [{"source": "菇勇者传说", "target": "Legend of Mushroom"}]
    assert _announcement_segment_term_hits(
        {"source": "菇勇者们，请查收联动奖励。"},
        "en",
        lookup,
    ) == [{"source": "菇勇者", "target": "Shroomie"}]


def test_announcement_uses_confirmed_project_glossary_without_translation_archive() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Legend of Mushroom", "type": "RPG"}).json()
        glossary = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "source": "菇勇者传说",
                "target": "Legend of Mushroom",
                "language": "en",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert glossary.status_code == 200, glossary.text
        task = _create_game_name_task(client, project["id"], include_project_archive=True)

        extracted = _extract_terms(client, task["id"], include_project_archive=True)

        terms = extracted["task"]["metadata"]["terms"]
        assert len(terms) == 1
        assert terms[0]["id"] == glossary.json()["id"]
        assert terms[0]["source"] == "菇勇者传说"
        assert terms[0]["translations"] == {"en": "Legend of Mushroom"}
        assert terms[0]["translation_sources"]["en"]["type"] == "project_glossary"
        assert terms[0]["hit_count"] == 1
        assert terms[0]["first_position"] == 1


def test_confirmed_project_glossary_wins_over_conflicting_qa_archive() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Legend of Mushroom priority", "type": "RPG"}).json()
        glossary = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "source": "菇勇者传说",
                "target": "Legend of Mushroom",
                "language": "en",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert glossary.status_code == 200, glossary.text
        archived = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "old-wrong-game-name",
                "source": "菇勇者传说",
                "target": "Shroomie Legendary",
                "language": "en",
                "source_type": "qa_passed",
            },
        )
        assert archived.status_code == 200, archived.text
        task = _create_game_name_task(client, project["id"], include_project_archive=True)
        extracted = _extract_terms(client, task["id"], include_project_archive=True)
        assert extracted["task"]["metadata"]["terms"][0]["translations"]["en"] == "Legend of Mushroom"

        lookup = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": True},
        )

        assert lookup.status_code == 200, lookup.text
        term = lookup.json()["manifest"]["lookup"]["en"]["terms"][0]
        assert term["target"] == "Legend of Mushroom"
        assert term["source_type"] == "project_glossary"


def test_edited_announcement_term_survives_disabling_conflicting_qa_archive() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Edited announcement term", "type": "RPG"}).json()
        archived = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "old-wrong-game-name",
                "source": "菇勇者传说",
                "target": "Shroomie Legendary",
                "language": "en",
                "source_type": "qa_passed",
            },
        )
        assert archived.status_code == 200, archived.text
        task = _create_game_name_task(client, project["id"], include_project_archive=True)
        imported = client.post(
            f"/api/announcement-tasks/{task['id']}/import-terms",
            json={
                "languages": ["en"],
                "terms": [
                    {
                        "id": "game-name",
                        "source": "菇勇者传说",
                        "translations": {"en": "Legend of Mushroom"},
                        "hit_count": 1,
                        "first_position": 1,
                    }
                ],
            },
        )
        assert imported.status_code == 200, imported.text

        lookup_with_archive = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": True},
        )
        assert lookup_with_archive.status_code == 200, lookup_with_archive.text
        imported_term = lookup_with_archive.json()["manifest"]["lookup"]["en"]["terms"][0]
        assert imported_term["target"] == "Legend of Mushroom"
        assert imported_term["source_type"] == "announcement_term"

        lookup = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": False},
        )

        assert lookup.status_code == 200, lookup.text
        term = lookup.json()["manifest"]["lookup"]["en"]["terms"][0]
        assert term["target"] == "Legend of Mushroom"
        assert term["source_type"] == "announcement_term"


def test_lookup_without_project_archive_filters_cached_project_sources_per_language(tmp_path: Path) -> None:
    table_path = tmp_path / "constraints.xlsx"
    _write_multilingual_constraint_table(table_path)

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Cached source isolation", "type": "RPG"}).json()
        glossary = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "source": "菇勇者传说",
                "target": "Legend of Mushroom",
                "language": "en",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert glossary.status_code == 200, glossary.text
        archived = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "cached-announcement",
                "source": "联动公告",
                "target": "Collab Notice",
                "language": "en",
                "source_type": "qa_passed",
            },
        )
        assert archived.status_code == 200, archived.text
        with table_path.open("rb") as table_file:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={
                    "file": (
                        table_path.name,
                        table_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert table_artifact.status_code == 200, table_artifact.text
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={
                "text": "《菇勇者传说》联动公告",
                "languages": ["en", "fr"],
                "language_table_artifact_ids": [table_artifact.json()["id"]],
                "include_project_archive": True,
            },
        )
        assert task.status_code == 200, task.text
        extracted = client.post(
            f"/api/announcement-tasks/{task.json()['id']}/extract-terms",
            json={
                "languages": ["en", "fr"],
                "language_table_artifact_ids": [table_artifact.json()["id"]],
                "include_project_archive": True,
                "ai_supplement": False,
            },
        )
        assert extracted.status_code == 200, extracted.text
        extracted_by_source = {
            term["source"]: term
            for term in extracted.json()["task"]["metadata"]["terms"]
        }
        assert extracted_by_source["菇勇者传说"]["translation_sources"] == {
            "en": {"type": "project_glossary", "priority": 0},
            "fr": {"type": "language_table", "priority": 1},
        }
        assert extracted_by_source["联动公告"]["translation_sources"]["en"]["type"] == "qa_archive"
        assert extracted_by_source["联动公告"]["translation_sources"]["fr"]["type"] == "language_table"

        lookup = client.post(
            f"/api/announcement-tasks/{task.json()['id']}/lookup-translations",
            json={"languages": ["en", "fr"], "include_project_archive": False},
        )

        assert lookup.status_code == 200, lookup.text
        english = {
            term["source"]: term
            for term in lookup.json()["manifest"]["lookup"]["en"]["terms"]
        }
        french = {
            term["source"]: term
            for term in lookup.json()["manifest"]["lookup"]["fr"]["terms"]
        }
        assert english["菇勇者传说"]["target"] == ""
        assert english["菇勇者传说"]["source_type"] == "missing"
        assert "en" not in english["菇勇者传说"]["translations"]
        assert "en" not in english["菇勇者传说"]["translation_sources"]
        assert english["联动公告"]["target"] == ""
        assert english["联动公告"]["source_type"] == "missing"
        assert "en" not in english["联动公告"]["translations"]
        assert "en" not in english["联动公告"]["translation_sources"]
        assert french["菇勇者传说"]["target"] == "Légende du Champignon"
        assert french["菇勇者传说"]["source_type"] == "language_table"
        assert french["菇勇者传说"]["translation_sources"]["fr"]["type"] == "language_table"
        assert french["联动公告"]["target"] == "Annonce de collaboration"
        assert french["联动公告"]["source_type"] == "language_table"
        assert french["联动公告"]["translation_sources"]["fr"]["type"] == "language_table"


def test_lookup_without_project_archive_discards_unattributed_legacy_cached_target() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Legacy cached source isolation", "type": "RPG"}).json()
        archived = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "legacy-wrong-game-name",
                "source": "菇勇者传说",
                "target": "Shroomie Legendary",
                "language": "en",
                "source_type": "qa_passed",
            },
        )
        assert archived.status_code == 200, archived.text
        task = _create_game_name_task(client, project["id"], include_project_archive=True)
        extracted = _extract_terms(client, task["id"], include_project_archive=True)
        legacy_metadata = dict(extracted["task"]["metadata"])
        for term in legacy_metadata["terms"]:
            term.pop("translation_sources", None)
            term.pop("sources", None)
        db.update_announcement_task(task["id"], metadata=legacy_metadata)

        lookup = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": False},
        )

        assert lookup.status_code == 200, lookup.text
        term = lookup.json()["manifest"]["lookup"]["en"]["terms"][0]
        assert term["target"] == ""
        assert term["source_type"] == "missing"
        assert "en" not in term["translations"]
        assert "en" not in term["translation_sources"]


def test_lookup_with_project_archive_rebuilds_unattributed_legacy_cached_target(tmp_path: Path) -> None:
    table_path = tmp_path / "official-game-name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CN", "EN"])
    sheet.append(["菇勇者传说", "Legend of Mushroom"])
    workbook.save(table_path)
    workbook.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Legacy cached priority", "type": "RPG"}).json()
        archived = client.post(
            f"/api/projects/{project['id']}/translations",
            json={
                "entry_key": "legacy-wrong-game-name",
                "source": "菇勇者传说",
                "target": "Shroomie Legendary",
                "language": "en",
                "source_type": "qa_passed",
            },
        )
        assert archived.status_code == 200, archived.text
        glossary = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "source": "菇勇者传说",
                "target": "Legend of Mushroom",
                "language": "en",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert glossary.status_code == 200, glossary.text
        with table_path.open("rb") as table_file:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={
                    "file": (
                        table_path.name,
                        table_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert table_artifact.status_code == 200, table_artifact.text
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={
                "text": "《菇勇者传说》联动公告",
                "languages": ["en"],
                "language_table_artifact_ids": [table_artifact.json()["id"]],
                "include_project_archive": True,
            },
        ).json()
        extracted = _extract_terms(client, task["id"], include_project_archive=True)
        legacy_metadata = dict(extracted["task"]["metadata"])
        for term in legacy_metadata["terms"]:
            if term["source"] != "菇勇者传说":
                continue
            term["translations"]["en"] = "Shroomie Legendary"
            term.pop("translation_sources", None)
            term.pop("sources", None)
        db.update_announcement_task(task["id"], metadata=legacy_metadata)

        lookup = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": True},
        )

        assert lookup.status_code == 200, lookup.text
        term = lookup.json()["manifest"]["lookup"]["en"]["terms"][0]
        assert term["target"] == "Legend of Mushroom"
        assert term["source_type"] == "project_glossary"
        assert term["translation_sources"]["en"]["type"] == "project_glossary"

        prepared = client.post(f"/api/announcement-tasks/{task['id']}/prepare", json={"languages": ["en"]})
        assert prepared.status_code == 200, prepared.text
        workpack = next(artifact for artifact in prepared.json()["artifacts"] if artifact["kind"] == "announcement_workpack")
        workpack_rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]
        assert workpack_rows[0]["term_hits"] == [{"source": "菇勇者传说", "target": "Legend of Mushroom"}]

        translated = client.post(
            f"/api/announcement-tasks/{task['id']}/translate",
            json={"languages": ["en"], "provider": "test-fake"},
        )
        assert translated.status_code == 200, translated.text
        response = next(artifact for artifact in translated.json()["artifacts"] if artifact["kind"] == "announcement_ai_response")
        response_rows = [json.loads(line) for line in Path(response["path"]).read_text(encoding="utf-8").splitlines()]
        assert response_rows[0]["translation"] == "TestFake row Legend of Mushroom"


def test_lookup_preserves_ai_supplement_without_translation_source_metadata(tmp_path: Path) -> None:
    table_path = tmp_path / "official-game-name.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["CN", "EN"])
    sheet.append(["菇勇者传说", "Legend of Mushroom"])
    workbook.save(table_path)
    workbook.close()

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "AI supplement preservation", "type": "RPG"}).json()
        with table_path.open("rb") as table_file:
            table_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=language_table",
                files={
                    "file": (
                        table_path.name,
                        table_file,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
        assert table_artifact.status_code == 200, table_artifact.text
        task = client.post(
            f"/api/projects/{project['id']}/announcement-tasks",
            json={
                "text": "《菇勇者传说》联动公告",
                "languages": ["en"],
                "language_table_artifact_ids": [table_artifact.json()["id"]],
                "include_project_archive": True,
            },
        ).json()
        extracted = _extract_terms(client, task["id"], include_project_archive=True)
        ai_metadata = dict(extracted["task"]["metadata"])
        term = ai_metadata["terms"][0]
        term["translations"]["en"] = "AI Supplement Name"
        term["source_type"] = "ai_supplement"
        term.pop("translation_sources", None)
        term.pop("sources", None)
        db.update_announcement_task(task["id"], metadata=ai_metadata)

        lookup = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": True},
        )

        assert lookup.status_code == 200, lookup.text
        lookup_term = lookup.json()["manifest"]["lookup"]["en"]["terms"][0]
        assert lookup_term["target"] == "AI Supplement Name"
        assert lookup_term["source_type"] == "ai_supplement"


def test_wrong_game_name_ai_output_remains_a_hard_blocker_after_apply(tmp_path: Path) -> None:
    response_path = tmp_path / "wrong_game_name.jsonl"
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Game name QA", "type": "RPG"}).json()
        glossary = client.post(
            f"/api/projects/{project['id']}/glossary",
            json={
                "source": "菇勇者传说",
                "target": "Legend of Mushroom",
                "language": "en",
                "source_type": "manual",
                "confirmed": True,
            },
        )
        assert glossary.status_code == 200, glossary.text
        task = _create_game_name_task(client, project["id"], include_project_archive=True)
        _extract_terms(client, task["id"], include_project_archive=True)
        lookup = client.post(
            f"/api/announcement-tasks/{task['id']}/lookup-translations",
            json={"languages": ["en"], "include_project_archive": True},
        )
        assert lookup.status_code == 200, lookup.text
        prepared = client.post(f"/api/announcement-tasks/{task['id']}/prepare", json={"languages": ["en"]})
        assert prepared.status_code == 200, prepared.text
        workpack = next(artifact for artifact in prepared.json()["artifacts"] if artifact["kind"] == "announcement_workpack")
        workpack_rows = [json.loads(line) for line in Path(workpack["path"]).read_text(encoding="utf-8").splitlines()]
        response_path.write_text(
            "\n".join(
                json.dumps(
                    {"para_id": row["para_id"], "translation": "Shroomie Legendary"},
                    ensure_ascii=False,
                )
                for row in workpack_rows
            )
            + "\n",
            encoding="utf-8",
        )
        with response_path.open("rb") as response_file:
            response_artifact = client.post(
                f"/api/projects/{project['id']}/files?kind=asset",
                files={"file": (response_path.name, response_file, "application/jsonl")},
            ).json()
        imported = client.post(
            f"/api/announcement-tasks/{task['id']}/import-ai",
            json={"languages": ["en"], "response_artifact_ids": [response_artifact["id"]]},
        )
        assert imported.status_code == 200, imported.text

        applied = client.post(f"/api/announcement-tasks/{task['id']}/apply", json={"languages": ["en"]})

        assert applied.status_code == 200, applied.text
        payload = applied.json()
        assert payload["summary"]["hard_blockers"] == 1
        assert payload["summary"]["auto_fixed"] == 0
        assert payload["task"]["metadata"]["qa_issues"][0]["check_type"] == "term_missing"
        output = next(artifact for artifact in payload["artifacts"] if artifact["kind"] == "announcement_output_file")
        assert Path(output["path"]).read_text(encoding="utf-8").strip() == "Shroomie Legendary"
