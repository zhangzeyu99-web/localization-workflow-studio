from __future__ import annotations

import asyncio
import re
from io import BytesIO
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Any

os.environ["LWS_DATA_ROOT"] = str(Path(tempfile.gettempdir()) / "lws-test-data")

import httpx
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.db as db
import app.workflow as workflow
from app.config import DEFAULT_SETTINGS, normalize_settings, save_settings
from app.main import app
from app.providers import TranslationItem, call_text, openai_responses_translate_batch, translate_batch
from conftest import reset_data_root, wait_for_background_jobs


@pytest.fixture(autouse=True)
def reset_test_state(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("LWS_MAX_UPLOAD_MB", "1")
    data_root = Path(os.environ["LWS_DATA_ROOT"])
    reset_data_root(data_root)
    db.init_db()
    save_settings(DEFAULT_SETTINGS)
    yield
    wait_for_background_jobs()
    save_settings(DEFAULT_SETTINGS)


def test_subprocess_failure_writes_structured_backend_error_without_raw_user_text(tmp_path: Path) -> None:
    project = db.insert_project("Subprocess Structured Error", "QA", "", "🎮")
    run = db.insert_run(project["id"], "translation", "en", metadata={})

    with pytest.raises(workflow.UserFacingWorkflowError) as raised:
        workflow.run_subprocess(
            [sys.executable, "-c", "import sys; print('public stdout'); print('Traceback raw secret', file=sys.stderr); sys.exit(7)"],
            tmp_path,
            run["id"],
        )

    assert "Traceback raw secret" not in str(raised.value)
    log_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "logs"
    error_file = log_dir / "subprocess_error.json"
    assert error_file.exists()
    payload = json.loads(error_file.read_text(encoding="utf-8"))
    assert payload["returncode"] == 7
    assert "Traceback raw secret" in payload["stderr"]
    assert (log_dir / "subprocess_events.jsonl").exists()




def test_subprocess_event_output_summarizes_qa_dict() -> None:
    from app.workflow.subprocess_runner import _safe_subprocess_event_output

    raw = "{'passed': False, 'total_cases': 61, 'issue_counts': {'title_case_overuse': 13, 'clipped_word': 25}}"

    summary = _safe_subprocess_event_output(raw)

    assert summary == "QA \u672a\u901a\u8fc7\uff1a\u53d1\u73b0 38 \u4e2a\u95ee\u9898\uff0c\u8be6\u60c5\u5df2\u5199\u5165 QA \u6458\u8981\u3002"
    assert "issue_counts" not in summary
    assert "title_case_overuse" not in summary

def test_subprocess_reads_structured_result_key_output(tmp_path: Path) -> None:
    project = db.insert_project("Subprocess Structured Result", "QA", "", "🎮")
    run = db.insert_run(project["id"], "glossary", "en", metadata={})
    script = (
        "import json, os, pathlib; "
        "pathlib.Path(os.environ['LWS_SUBPROCESS_RESULT_PATH']).write_text("
        "json.dumps({'key_output': {'FINAL_OUTPUT': 'final.xlsx', 'DETAIL_OUTPUT': 'detail.xlsx'}}, ensure_ascii=False), encoding='utf-8')"
    )

    proc = workflow.run_subprocess([sys.executable, "-c", script], tmp_path, run["id"])
    parsed = workflow.parse_key_output(proc.stdout)

    assert parsed == {"FINAL_OUTPUT": "final.xlsx", "DETAIL_OUTPUT": "detail.xlsx"}
    log_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "logs"
    payload = json.loads((log_dir / "subprocess_result.json").read_text(encoding="utf-8"))
    assert payload["result"]["key_output"]["FINAL_OUTPUT"] == "final.xlsx"


def test_subprocess_reads_structured_error_user_message(tmp_path: Path) -> None:
    project = db.insert_project("Subprocess Structured Error Message", "QA", "", "🎮")
    run = db.insert_run(project["id"], "glossary", "en", metadata={})
    script = (
        "import json, os, pathlib, sys; "
        "pathlib.Path(os.environ['LWS_SUBPROCESS_ERROR_PATH']).write_text("
        "json.dumps({'user_message': '请上传完整语言表，不要上传公告原文。'}, ensure_ascii=False), encoding='utf-8'); "
        "print('Traceback raw secret', file=sys.stderr); sys.exit(9)"
    )

    with pytest.raises(workflow.UserFacingWorkflowError) as raised:
        workflow.run_subprocess([sys.executable, "-c", script], tmp_path, run["id"])

    assert str(raised.value) == "请上传完整语言表，不要上传公告原文。"
    log_dir = Path(os.environ["LWS_DATA_ROOT"]) / "runs" / run["id"] / "logs"
    payload = json.loads((log_dir / "subprocess_error.json").read_text(encoding="utf-8"))
    assert payload["error"]["user_message"] == "请上传完整语言表，不要上传公告原文。"


def test_manifest_invalidates_when_source_language_prompt_or_settings_change() -> None:
    rows = [{"id": 1, "source": "开始游戏"}]
    base_settings = {**DEFAULT_SETTINGS, "provider": "test-fake", "preset": "balanced", "batch_size": 1, "max_batch_input_tokens": 12000}
    manifest = workflow._build_batch_manifest(rows, "Prompt A", base_settings, batch_size=1, language="en")

    assert workflow._manifest_matches_rows(manifest, rows, "Prompt A", base_settings, 1, "en")
    assert not workflow._manifest_matches_rows(manifest, [{"id": 1, "source": "领取奖励"}], "Prompt A", base_settings, 1, "en")
    assert not workflow._manifest_matches_rows(manifest, rows, "Prompt B", base_settings, 1, "en")
    assert not workflow._manifest_matches_rows(manifest, rows, "Prompt A", {**base_settings, "preset": "deep"}, 1, "en")
    assert not workflow._manifest_matches_rows(manifest, rows, "Prompt A", base_settings, 1, "ko")


def test_import_templates_download_readable_workbooks() -> None:
    expected_headers = {
        "language-table": ["ID", "CN", "EN", "EN2", "KR", "JP", "备注"],
        "glossary": ["ID", "CN", "EN", "EN2", "KR", "JP", "分类", "备注"],
        "announcement-language-table": ["ID", "CN", "EN", "EN2", "KR", "JP", "备注"],
        "announcement-terms": ["ID", "CN", "EN", "KR", "JP", "命中次数", "来源", "备注"],
    }
    with TestClient(app) as client:
        for kind, headers in expected_headers.items():
            response = client.get(f"/api/import-templates/{kind}")
            assert response.status_code == 200, response.text
            wb = load_workbook(BytesIO(response.content), data_only=True)
            try:
                assert "填写说明" in wb.sheetnames
                ws = wb[wb.sheetnames[1]]
                actual = [ws.cell(1, col).value for col in range(1, len(headers) + 1)]
                assert actual == headers
                assert ws.max_row >= 3
            finally:
                wb.close()


def _write_qa_source_workbook(path: Path, translated: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "EN"])
    ws.append([1, "开始游戏", "Start Game" if translated else ""])
    ws.append([2, "领取奖励", "Claim Rewards" if translated else ""])
    wb.save(path)
    wb.close()


def test_translation_qa_uses_translated_artifact_instead_of_original_input(tmp_path: Path) -> None:
    project = db.insert_project("QA source selection", "QA", "", "🎮")
    run = db.insert_run(project["id"], "translation", "en", metadata={})
    original_path = tmp_path / "original.xlsx"
    translated_path = tmp_path / "translated.xlsx"
    _write_qa_source_workbook(original_path, translated=False)
    _write_qa_source_workbook(translated_path, translated=True)
    original = db.add_artifact(project["id"], "original", original_path, "language_table", run_id=run["id"])
    translated = db.add_artifact(project["id"], "translated", translated_path, "raw_translated_workbook", run_id=run["id"])
    run = db.update_run(
        run["id"],
        metadata={
            "input_artifact_id": original["id"],
            "input_artifacts": {
                "source_workbook": original["id"],
                "raw_translated_workbook": translated["id"],
            },
        },
    )

    selected = workflow._workbook_artifact_for_quality_run(run)

    assert selected["id"] == translated["id"]


def test_translation_qa_without_translated_output_is_rejected(tmp_path: Path) -> None:
    project = db.insert_project("QA source missing output", "QA", "", "🎮")
    run = db.insert_run(project["id"], "translation", "en", metadata={})
    original_path = tmp_path / "original.xlsx"
    _write_qa_source_workbook(original_path, translated=False)
    original = db.add_artifact(project["id"], "original", original_path, "language_table", run_id=run["id"])
    run = db.update_run(
        run["id"],
        metadata={
            "input_artifact_id": original["id"],
            "input_artifacts": {"source_workbook": original["id"]},
        },
    )

    with pytest.raises(ValueError, match="先完成 AI 翻译"):
        workflow._workbook_artifact_for_quality_run(run)


def test_delivery_skips_empty_workbook_but_keeps_failed_review_artifact(tmp_path: Path) -> None:
    project = db.insert_project("Delivery guard", "QA", "", "🎮")
    failed_run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": False}})
    failed_path = tmp_path / "failed.xlsx"
    _write_qa_source_workbook(failed_path, translated=True)
    db.add_artifact(project["id"], "failed final", failed_path, "qa_final_workbook", run_id=failed_run["id"])
    db.update_run(failed_run["id"], status="failed", metadata={"quality_summary": {"passed": False}})

    empty_run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": True}})
    empty_path = tmp_path / "empty.xlsx"
    _write_qa_source_workbook(empty_path, translated=False)
    db.add_artifact(project["id"], "empty final", empty_path, "qa_final_workbook", run_id=empty_run["id"])
    db.update_run(empty_run["id"], status="passed", metadata={"quality_summary": {"passed": True}})

    passed_run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": True}})
    passed_path = tmp_path / "passed.xlsx"
    _write_qa_source_workbook(passed_path, translated=True)
    db.add_artifact(project["id"], "passed final", passed_path, "qa_final_workbook", run_id=passed_run["id"])
    db.update_run(passed_run["id"], status="passed", metadata={"quality_summary": {"passed": True}})

    deliverables = workflow.list_project_deliverables(project["id"])

    assert [item["run_id"] for item in deliverables] == [passed_run["id"], failed_run["id"]]


def test_deliverable_summaries_flag_delivered_with_issues(tmp_path: Path) -> None:
    project = db.insert_project("Delivery issue flag", "QA", "", "🎮")
    failed_run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": False, "hard_errors": 2}})
    failed_path = tmp_path / "failed-flag.xlsx"
    _write_qa_source_workbook(failed_path, translated=True)
    db.add_artifact(project["id"], "failed final", failed_path, "qa_final_workbook", run_id=failed_run["id"])
    db.update_run(failed_run["id"], status="failed", metadata={"quality_summary": {"passed": False, "hard_errors": 2}})

    passed_run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": True}})
    passed_path = tmp_path / "passed-flag.xlsx"
    _write_qa_source_workbook(passed_path, translated=True)
    db.add_artifact(project["id"], "passed final", passed_path, "qa_final_workbook", run_id=passed_run["id"])
    db.update_run(passed_run["id"], status="passed", metadata={"quality_summary": {"passed": True}})

    deliverables = {item["run_id"]: item for item in workflow.list_project_deliverables(project["id"])}

    assert deliverables[failed_run["id"]]["delivered_with_issues"] is True
    assert deliverables[passed_run["id"]]["delivered_with_issues"] is False


def test_failed_qa_delivery_archives_translation_with_issue_source(tmp_path: Path) -> None:
    project = db.insert_project("Delivery archive failed QA", "QA", "", "🎮")
    failed_run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": False, "hard_errors": 1}})
    failed_path = tmp_path / "failed-delivery.xlsx"
    _write_qa_source_workbook(failed_path, translated=True)
    final_artifact = db.add_artifact(project["id"], "failed final", failed_path, "qa_final_workbook", run_id=failed_run["id"])
    db.update_run(failed_run["id"], status="failed", metadata={"quality_summary": {"passed": False, "hard_errors": 1}})

    result = workflow.build_delivery_package(project["id"], run_id=failed_run["id"])
    entries = db.list_translation_entries(project["id"], language="en")

    assert result["archive"]["artifact_id"] == final_artifact["id"]
    assert result["archive"]["source_type"] == "delivered_with_issues"
    assert result["archive"]["imported_count"] == 2
    assert len(entries) == 2
    assert {entry["source_type"] for entry in entries} == {"delivered_with_issues"}
    assert {entry["target"] for entry in entries} == {"Start Game", "Claim Rewards"}


def test_artifact_payload_exposes_file_existence(tmp_path: Path) -> None:
    project = db.insert_project("Artifact exists", "QA", "", "🎮")
    existing_path = tmp_path / "existing.xlsx"
    _write_qa_source_workbook(existing_path, translated=True)
    existing = db.add_artifact(project["id"], "existing", existing_path, "final_workbook")
    missing = db.add_artifact(project["id"], "missing", tmp_path / "missing.xlsx", "final_workbook")

    assert db.get_artifact(existing["id"])["exists"] is True
    assert db.get_artifact(missing["id"])["exists"] is False


def test_version_endpoint_returns_runtime_version() -> None:
    with TestClient(app) as client:
        response = client.get("/api/version")

    assert response.status_code == 200
    payload = response.json()
    assert payload["version"]
    assert "data_root" in payload
    assert isinstance(payload.get("frontend_assets"), list)


def test_api_responses_are_marked_no_store() -> None:
    with TestClient(app) as client:
        version_response = client.get("/api/version")
        health_response = client.get("/api/health")

    assert version_response.headers.get("cache-control") == "no-store"
    assert health_response.headers.get("cache-control") == "no-store"


def test_context_cap_keeps_project_prompt_under_batch_budget() -> None:
    rows = [{"id": 1, "source": "开始游戏"}]
    long_prompt = "项目规则开始\n" + ("超长项目背景 " * 2000) + "\n输出协议：只返回 JSONL"
    settings = {
        **DEFAULT_SETTINGS,
        "provider": "test-fake",
        "batch_size": 1,
        "max_batch_input_tokens": 1000,
        "max_project_context_tokens": 240,
    }

    manifest = workflow._build_batch_manifest(rows, long_prompt, settings, batch_size=1, language="en")

    context = manifest["project_context"]
    assert context["trimmed"] is True
    assert context["managed_estimated_tokens"] <= 260
    assert context["original_estimated_tokens"] > context["managed_estimated_tokens"]
    assert all(batch["estimated_input_tokens"] <= settings["max_batch_input_tokens"] for batch in manifest["batches"])


def test_orchestrator_passes_managed_project_context_to_provider(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    project = db.insert_project("Managed Context", "QA", "", "🎯")
    run = db.insert_run(project["id"], "translation", "en", metadata={})
    rows = [{"id": 1, "source": "开始 {count}"}]
    long_prompt = "项目规则开始\n" + ("超长项目背景 " * 2000) + "\n输出协议：只返回 JSONL"
    settings = {
        **DEFAULT_SETTINGS,
        "provider": "test-fake",
        "batch_size": 1,
        "max_batch_input_tokens": 1000,
        "max_project_context_tokens": 240,
        "api_budget_warning_tokens": 20_000_000,
    }
    captured: dict[str, Any] = {}

    async def fake_translate_batch(batch: list[dict[str, Any]], provider_settings: dict[str, Any], project_prompt: str) -> list[TranslationItem]:
        _ = provider_settings
        captured["prompt"] = project_prompt
        return [TranslationItem(id=row["id"], translation="Translated {count}") for row in batch]

    monkeypatch.setattr(workflow, "translate_batch", fake_translate_batch)

    result = asyncio.run(
        workflow._translate_rows_with_orchestration(
            run_id=run["id"],
            rows=rows,
            settings=settings,
            project_prompt=long_prompt,
            work_dir=tmp_path,
            batch_size=1,
            language="en",
            confirm_api_budget=True,
        )
    )

    assert result[0]["translation"] == "Translated {count}"
    assert workflow._estimate_text_tokens(captured["prompt"]) <= 260
    assert "[context trimmed:" in captured["prompt"]


def test_manifest_fingerprint_mismatch_does_not_reuse_completed_batch(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    project = db.insert_project("Fingerprint", "QA", "", "🎮")
    run = db.insert_run(project["id"], "translation", "en", metadata={})
    rows = [{"id": 1, "source": "开始 {count}"}]
    settings = {**DEFAULT_SETTINGS, "provider": "test-fake", "batch_size": 1, "api_budget_warning_tokens": 20_000_000}
    state = {"calls": 0}

    async def fake_translate_batch(batch: list[dict[str, Any]], provider_settings: dict[str, Any], project_prompt: str) -> list[TranslationItem]:
        _ = provider_settings, project_prompt
        state["calls"] += 1
        return [TranslationItem(id=row["id"], translation=f"Translated {state['calls']} {{count}}") for row in batch]

    monkeypatch.setattr(workflow, "translate_batch", fake_translate_batch)
    first = asyncio.run(
        workflow._translate_rows_with_orchestration(
            run_id=run["id"],
            rows=rows,
            settings=settings,
            project_prompt="Prompt A",
            work_dir=tmp_path,
            batch_size=1,
            language="en",
            confirm_api_budget=True,
        )
    )
    second = asyncio.run(
        workflow._translate_rows_with_orchestration(
            run_id=run["id"],
            rows=[{"id": 1, "source": "领取 {count}"}],
            settings=settings,
            project_prompt="Prompt A",
            work_dir=tmp_path,
            batch_size=1,
            language="en",
            confirm_api_budget=True,
        )
    )

    assert first[0]["translation"] == "Translated 1 {count}"
    assert second[0]["translation"] == "Translated 2 {count}"
    assert state["calls"] == 2


def test_openai_responses_body_uses_max_output_tokens(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, Any] = {}

    class FakeAsyncClient:
        def __init__(self, timeout: Any) -> None:
            captured["timeout"] = timeout

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> httpx.Response:
            captured["url"] = url
            captured["headers"] = headers
            captured["body"] = json
            return httpx.Response(200, json={"output_text": '{"id":1,"translation":"Start"}'})

    monkeypatch.setattr(httpx, "AsyncClient", FakeAsyncClient)

    result = asyncio.run(
        openai_responses_translate_batch(
            [{"id": 1, "source": "开始"}],
            {**DEFAULT_SETTINGS, "api_key": "sk-test", "max_output_tokens": 1234, "provider_timeout_seconds": 77},
            "Prompt",
        )
    )

    assert result == [TranslationItem(id=1, translation="Start")]
    assert captured["url"].endswith("/v1/responses")
    assert captured["body"]["max_output_tokens"] == 1234
    assert captured["timeout"] == 77


def test_openai_provider_does_not_fallback_to_chat_completions(monkeypatch: pytest.MonkeyPatch) -> None:
    async def fake_responses(rows: list[dict[str, Any]], settings: dict[str, Any], project_prompt: str) -> list[TranslationItem]:
        _ = rows, settings, project_prompt
        return [TranslationItem(id=1, translation="OK")]

    monkeypatch.setattr("app.providers.openai_responses_translate_batch", fake_responses)
    result = asyncio.run(
        translate_batch(
            [{"id": 1, "source": "开始"}],
            {**DEFAULT_SETTINGS, "provider": "openai", "protocol": "chat-completions", "api_key": "sk-test"},
            "Prompt",
        )
    )
    assert result == [TranslationItem(id=1, translation="OK")]


def test_openai_chat_relay_presets_map_to_expected_reasoning_levels() -> None:
    expected = {"balanced": "medium", "deep": "high", "critical": "xhigh"}
    for preset, reasoning in expected.items():
        settings = normalize_settings(
            {
                **DEFAULT_SETTINGS,
                "provider": "openai-chat",
                "preset": preset,
                "base_url": "https://relay.example.com/api",
                "model": "gpt-5.5",
                "reasoning_effort": "",
            }
        )

        assert settings["provider"] == "openai-chat"
        assert settings["protocol"] == "chat-completions"
        assert settings["base_url"] == "https://relay.example.com/api"
        assert settings["model"] == "gpt-5.5"
        assert settings["reasoning_effort"] == reasoning


def test_settings_extracts_api_key_and_base_url_from_pasted_relay_block() -> None:
    settings = normalize_settings(
        {
            **DEFAULT_SETTINGS,
            "provider": "openai-chat",
            "preset": "balanced",
            "base_url": "",
            "api_key": (
                "Codex账号：\n"
                "base_url: https://relay.example.com/api\n"
                "key： cr_abc1234567890abcdef\n"
            ),
            "model": "gpt-5.5",
        }
    )

    assert settings["api_key"] == "cr_abc1234567890abcdef"
    assert settings["base_url"] == "https://relay.example.com/api"


def test_openai_chat_relay_uses_chat_completions_body(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, Any] = {}

    class FakeAsyncClient:
        def __init__(self, timeout: Any) -> None:
            captured["timeout"] = timeout

        async def __aenter__(self) -> "FakeAsyncClient":
            return self

        async def __aexit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
            return None

        async def post(self, url: str, headers: dict[str, str], json: dict[str, Any]) -> httpx.Response:
            captured["url"] = url
            captured["headers"] = headers
            captured["body"] = json
            return httpx.Response(200, json={"choices": [{"message": {"content": '{"id":1,"translation":"Start"}'}}]})

    monkeypatch.setattr(httpx, "AsyncClient", FakeAsyncClient)

    result = asyncio.run(
        translate_batch(
            [{"id": 1, "source": "开始"}],
            {
                **DEFAULT_SETTINGS,
                "provider": "openai-chat",
                "api_key": "sk-test",
                "base_url": "https://relay.example.com/api",
                "model": "gpt-5.5",
                "reasoning_effort": "xhigh",
                "max_output_tokens": 1234,
                "provider_timeout_seconds": 77,
            },
            "Prompt",
        )
    )

    assert result == [TranslationItem(id=1, translation="Start")]
    assert captured["url"].endswith("/v1/" + "chat/completions")
    assert captured["body"]["model"] == "gpt-5.5"
    assert captured["body"]["reasoning"] == {"effort": "xhigh"}
    assert captured["body"]["max_tokens"] == 1234
    user_content = captured["body"]["messages"][1]["content"]
    assert "\\u5f00\\u59cb" in user_content
    assert "开始" not in user_content
    assert captured["timeout"] == 77


def test_semantic_text_provider_uses_openai_responses(monkeypatch: pytest.MonkeyPatch) -> None:
    captured: dict[str, Any] = {}

    def fake_post(url: str, headers: dict[str, str], json: dict[str, Any], timeout: int) -> httpx.Response:
        captured["url"] = url
        captured["headers"] = headers
        captured["body"] = json
        captured["timeout"] = timeout
        return httpx.Response(200, json={"output_text": '{"passed":true,"issues":[]}'})

    monkeypatch.setattr(httpx, "post", fake_post)

    text = call_text(
        {**DEFAULT_SETTINGS, "provider": "openai", "protocol": "chat-completions", "api_key": "sk-test", "max_output_tokens": 2222},
        "Check this workbook.",
    )

    assert text == '{"passed":true,"issues":[]}'
    assert captured["url"].endswith("/v1/responses")
    assert captured["body"]["max_output_tokens"] == 2222
    assert ("chat" + "/completions") not in captured["url"]


def test_semantic_qa_payload_parser_accepts_wrapped_json() -> None:
    parsed = workflow._parse_semantic_qa_payload('说明：{"passed":true,"issues":[]}\\n{"ignored":true}')

    assert parsed == {"passed": True, "issues": []}


def test_model_fix_resolves_generated_confirmation_sheet_by_record_id(tmp_path: Path) -> None:
    workbook = tmp_path / "language.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "CN", "EN"])
    for row_id in range(8381, 8389):
        ws.append([row_id, f"source {row_id}", f"old {row_id}"])
    confirmation_title = "\u9700\u786e\u8ba4"
    review_ws = wb.create_sheet(confirmation_title)
    review_ws.append(["ID", "source", "current", "suggestion", "reason"])
    review_ws.append([8381, "source 8381", "old 8381", "short fixed", "too long"])
    review_ws.append([8387, "source 8387", "old 8387", "short fixed", "too long"])
    wb.save(workbook)
    wb.close()

    issue = {"id": 8387, "sheet": confirmation_title, "row": 3, "severity": "hard", "check_type": "ui_length_overflow"}
    context = workflow._model_fix_row_context(workbook, issue)

    assert context["sheet"] == "Sheet1"
    assert context["row"] == 8
    assert context["record_id"] == "8387"

    applied = workflow._apply_workbook_fixes(
        workbook,
        [{"issue_id": "issue-1", "sheet": confirmation_title, "row": 3, "record_id": 8387, "translation": "short fixed"}],
        "run_test",
    )
    fixed = load_workbook(workbook, read_only=True, data_only=True)
    try:
        fixed_ws = fixed["Sheet1"]
        assert fixed_ws.cell(8, 3).value == "short fixed"
        assert fixed_ws.cell(3, 3).value == "old 8382"
        assert applied[0]["row"] == 8
        assert applied[0]["sheet"] == "Sheet1"
    finally:
        fixed.close()


def test_quality_dedupes_generated_confirmation_sheet_duplicates() -> None:
    quality = {
        "passed": False,
        "issues": [
            {
                "id": 124,
                "sheet": "Sheet1",
                "row": 125,
                "severity": "error",
                "check_type": "term_missing",
                "message": "Source term 部队 must use Troops",
                "source": "没有支援部队",
                "translation": "No Reinforcements",
            },
            {
                "id": 124,
                "sheet": "\u9700\u786e\u8ba4",
                "row": 5,
                "severity": "error",
                "check_type": "term_missing",
                "message": "Source term 部队 must use Troops",
                "source": "没有支援部队",
                "translation": "No Reinforcements",
            },
        ],
        "issue_counts": {"term_missing": 2},
        "hard_errors": 2,
        "soft_warnings": 0,
    }

    deduped = workflow._dedupe_quality_payload(quality)
    normalized = workflow._dedupe_quality_issues(workflow._normalize_quality_issues("global_harness", deduped))

    assert len(deduped["issues"]) == 1
    assert deduped["issues"][0]["sheet"] == "Sheet1"
    assert deduped["issue_counts"] == {"term_missing": 1}
    assert deduped["hard_errors"] == 1
    assert len(normalized) == 1
    assert normalized[0]["sheet"] == "Sheet1"


def test_language_api_uses_visible_kr_jp_and_aliases() -> None:
    with TestClient(app) as client:
        response = client.get("/api/languages")
        assert response.status_code == 200
        languages = {item["code"]: item for item in response.json()["languages"]}
        assert languages["ko"]["visible_code"] == "KR"
        assert languages["ja"]["visible_code"] == "JP"
        assert "kr" in languages["ko"]["aliases"]
        assert "jp" in languages["ja"]["aliases"]
        assert languages["en"]["alt_header"] == "EN2"
        assert languages["ko"]["alt_header"] == ""


def test_sqlite_pragmas_and_unique_indexes_block_duplicate_entry_keys() -> None:
    project = db.insert_project("DB Guard", "QA", "", "🎮")
    db.upsert_translation_entry(project["id"], {"entry_key": "1001", "source": "开始", "target": "Start", "language": "en"})
    db.upsert_translation_entry(project["id"], {"entry_key": "1001", "source": "开始", "target": "Begin", "language": "en"})
    entries = db.list_translation_entries(project["id"], language="en")

    with db.connect() as conn:
        foreign_keys = conn.execute("PRAGMA foreign_keys").fetchone()[0]
        indexes = {row["name"] for row in conn.execute("PRAGMA index_list(translation_entries)").fetchall()}

    assert foreign_keys == 1
    assert len([entry for entry in entries if entry["entry_key"] == "1001"]) == 1
    assert "idx_translation_entries_project_language_entry_key_unique" in indexes


def test_upload_streaming_rejects_oversized_file_and_sanitizes_name() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Upload Guard", "type": "QA"}).json()
        oversized = b"x" * (1024 * 1024 + 1)
        rejected = client.post(
            f"/api/projects/{project['id']}/files?kind=asset",
            files={"file": ("bad:name.xlsx", oversized, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert rejected.status_code == 413

        accepted = client.post(
            f"/api/projects/{project['id']}/files?kind=asset",
            files={"file": ("bad:name.xlsx", b"ok", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert accepted.status_code == 200
        payload = accepted.json()
        assert ":" not in payload["label"]
        assert Path(payload["path"]).name == payload["label"]


def test_chunked_upload_assembles_file_and_creates_artifact() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Chunk Upload", "type": "QA"}).json()
        content = b"first chunk\nsecond chunk\nthird chunk\n"
        chunks = [content[:12], content[12:25], content[25:]]
        final_payload: dict[str, Any] | None = None
        for index, chunk in enumerate(chunks):
            response = client.post(
                f"/api/projects/{project['id']}/files/chunk",
                data={
                    "upload_id": "chunk-test-upload",
                    "filename": "large_terms.xlsx",
                    "kind": "asset",
                    "purpose": "",
                    "index": str(index),
                    "total": str(len(chunks)),
                },
                files={"file": ("chunk.part", chunk, "application/octet-stream")},
            )
            assert response.status_code == 200
            final_payload = response.json()
        assert final_payload is not None
        assert final_payload["complete"] is True
        artifact = final_payload["artifact"]
        assert artifact["label"] == "large_terms.xlsx"
        assert Path(artifact["path"]).read_bytes() == content


def test_persistent_long_text_lease_allows_single_running_job_and_cancel() -> None:
    assert db.acquire_job_lease("long_text", "job_a")
    assert not db.acquire_job_lease("long_text", "job_b")
    lease = db.get_job_lease("long_text")
    assert lease["job_id"] == "job_a"
    assert lease["status"] == "running"

    db.cancel_job_lease("long_text", "job_a")
    canceled = db.get_job_lease("long_text")
    assert canceled["cancel_requested"] is True

    db.release_job_lease("long_text", "job_a", status="completed")
    assert db.acquire_job_lease("long_text", "job_b")


def test_cancel_translation_run_matches_run_prefixed_lease_job_id() -> None:
    """Regression test: routers/runs.py stores the lease under job_id=f"run:{run_id}",
    so cancel_translation_run must cancel that same job_id, not the bare run_id
    (which never matches and previously left the lease's cancel_requested unset).

    Since M2, the lease name is project-scoped (``long_text:{project_id}``)
    rather than the single global ``long_text`` name.
    """
    import app.workflow.translation as translation
    from app.jobs import lease_name_for_project

    project = db.insert_project("cancel lease matching", "QA", "")
    run = db.insert_run(project["id"], "translation", "en", metadata={})
    run_id = run["id"]
    lease_name = lease_name_for_project(project["id"])

    assert db.acquire_job_lease(lease_name, f"run:{run_id}")

    translation.cancel_translation_run(run_id)

    lease = db.get_job_lease(lease_name)
    assert lease["job_id"] == f"run:{run_id}"
    assert lease["cancel_requested"] is True

    updated = db.get_run(run_id)
    assert updated["status"] == "canceled"
    assert updated["metadata"]["cancel_requested_at"]


def test_core_python_files_do_not_have_utf8_bom() -> None:
    for relative in ("backend/app/config.py", "backend/app/workflow/common.py"):
        assert not Path(relative).read_bytes().startswith(b"\xef\xbb\xbf")


def test_workflow_modules_do_not_use_legacy_common_star_imports() -> None:
    for path in Path("backend/app/workflow").glob("*.py"):
        text = path.read_text(encoding="utf-8")
        assert "from .common import *" not in text
        assert "ruff: noqa: F403,F405" not in text


def test_no_read_modify_write_metadata_spread_pattern_regression() -> None:
    """Guard against the N-5 race that motivated ``db.merge_run_metadata``:
    a whole-dict ``metadata={**db.get_run(...).get("metadata", {}), ...}``
    replace is a read-modify-write race under concurrent writers (M1 fixed
    ~15 call sites; M2 raises the odds of concurrent writers by letting
    different projects run in parallel, so this must stay at zero).
    """
    offenders: list[str] = []
    for path in Path("backend/app").rglob("*.py"):
        text = path.read_text(encoding="utf-8")
        if "metadata={**" in text:
            offenders.append(str(path))
    assert not offenders, f"found read-modify-write metadata spread pattern in: {offenders}"


def _assert_error_detail_is_safe(detail: str) -> None:
    assert detail.strip()
    assert "Traceback" not in detail
    assert "python.exe" not in detail
    assert not re.search(r"[A-Za-z]:[\\/]", detail)


def _large_language_table_bytes(rows: int = 1001) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Language"
    ws.append(["ID", "CN", "KR"])
    for index in range(1, rows + 1):
        ws.append([index, f"完整语言表源文 {index}", ""])
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


def test_error_responses_do_not_leak_traceback_or_server_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    def raise_internal(*args: Any, **kwargs: Any) -> Any:
        raise Exception(
            "Traceback (most recent call last):\n"
            '  File "D:\\secret\\workflow.py", line 10, in run\n'
            "command failed: C:\\Python314\\python.exe run_translation_harness.py"
        )

    monkeypatch.setattr("app.routers.glossary.extract_glossary", raise_internal)
    monkeypatch.setenv("LWS_MAX_UPLOAD_MB", "8")
    with TestClient(app) as client:
        missing_download = client.get("/api/artifacts/no-such-artifact/download")
        assert missing_download.status_code == 404
        _assert_error_detail_is_safe(missing_download.json()["detail"])

        project = client.post("/api/projects", json={"name": "Leak Guard", "type": "QA"}).json()
        rejected_upload = client.post(
            f"/api/projects/{project['id']}/files?kind=term_base",
            files={
                "file": (
                    "full-language-table.xlsx",
                    _large_language_table_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert rejected_upload.status_code == 400
        _assert_error_detail_is_safe(rejected_upload.json()["detail"])

        crashed = client.post(
            f"/api/projects/{project['id']}/glossary/extract",
            json={"input_artifact_id": "artifact-x", "language": "en"},
        )
        assert crashed.status_code == 500
        _assert_error_detail_is_safe(crashed.json()["detail"])


@pytest.mark.parametrize(
    ("headers", "rows", "expected"),
    [
        # Glossary-shaped file: term-like columns, few rows.
        (["ID", "CN", "EN", "分类", "note"], 3, False),
        # Exactly at the 1000-source-row threshold: classifier requires MORE than threshold.
        (["ID", "CN", "KR"], 1000, False),
        # Just above the threshold: complete language table.
        (["ID", "CN", "KR"], 1001, True),
        # Many rows but no target language column: not a language table.
        (["ID", "CN"], 1001, False),
    ],
)
def test_complete_language_table_classifier_boundary(tmp_path: Path, headers: list[str], rows: int, expected: bool) -> None:
    from app.workflow.asset_import_export import is_complete_language_table_for_glossary_import

    path = tmp_path / "classifier-input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for index in range(1, rows + 1):
        ws.append([index, f"源文 {index}"] + [""] * (len(headers) - 2))
    wb.save(path)
    wb.close()

    assert is_complete_language_table_for_glossary_import(path) is expected


def test_deployment_check_frontend_asset_comparison(tmp_path: Path) -> None:
    import importlib.util

    script_path = Path("scripts/deployment_check.py").resolve()
    spec = importlib.util.spec_from_file_location("deployment_check", script_path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)

    matching = module.compare_frontend_assets(["index-abc.css", "index-abc.js"], ["index-abc.js", "index-abc.css"])
    assert matching["ok"] is True
    assert matching["missing_on_server"] == []
    assert matching["missing_locally"] == []

    mismatch = module.compare_frontend_assets(["index-old.js"], ["index-new.js"])
    assert mismatch["ok"] is False
    assert mismatch["missing_on_server"] == ["index-new.js"]
    assert mismatch["missing_locally"] == ["index-old.js"]
    assert "different frontend build" in mismatch["error"]

    empty_local = module.compare_frontend_assets(["index-abc.js"], [])
    assert empty_local["ok"] is False
    assert "frontend build" in empty_local["error"]

    (tmp_path / "index-abc.js").write_text("js", encoding="utf-8")
    (tmp_path / "index-abc.css").write_text("css", encoding="utf-8")
    (tmp_path / "nested").mkdir()
    assert module.local_frontend_assets(tmp_path) == ["index-abc.css", "index-abc.js"]


def test_user_facing_error_from_route_maps_status_and_sanitized_detail(monkeypatch: pytest.MonkeyPatch) -> None:
    from app.errors import ArtifactError, ProviderError

    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "UserFacing Handler", "type": "QA"}).json()
        payload = {
            "project_id": project["id"],
            "kind": "translation",
            "language": "en",
            "input_artifact_id": "artifact-x",
            "batch_size": 10,
            "task_code": "T",
        }

        def raise_artifact_error(*args: Any, **kwargs: Any) -> None:
            raise ArtifactError("找不到所选文件，请重新上传或重新选择文件。")

        monkeypatch.setattr("app.routers.runs._validate_run_input_artifact", raise_artifact_error)
        artifact_response = client.post("/api/runs", json=payload)
        assert artifact_response.status_code == 404
        assert artifact_response.json()["detail"] == "找不到所选文件，请重新上传或重新选择文件。"

        def raise_provider_error(*args: Any, **kwargs: Any) -> None:
            raise ProviderError("provider exploded at D:\\srv\\app\\worker.py")

        monkeypatch.setattr("app.routers.runs._validate_run_input_artifact", raise_provider_error)
        provider_response = client.post("/api/runs", json=payload)
        assert provider_response.status_code == 502
        _assert_error_detail_is_safe(provider_response.json()["detail"])


def test_upload_unsupported_format_returns_readable_error() -> None:
    with TestClient(app) as client:
        project = client.post("/api/projects", json={"name": "Upload Format Guard", "type": "QA"}).json()
        rejected = client.post(
            f"/api/projects/{project['id']}/files?kind=language_table",
            files={"file": ("notes.docx", b"not a workbook", "application/octet-stream")},
        )
        assert rejected.status_code == 400
        detail = rejected.json()["detail"]
        assert "不支持 .docx" in detail
        assert "XLSX" in detail
        _assert_error_detail_is_safe(detail)


def test_deliverable_disappears_when_final_file_deleted_on_disk(tmp_path: Path) -> None:
    project = db.insert_project("Delivery missing file", "QA", "", "🎮")
    run = db.insert_run(project["id"], "qa", "en", metadata={"quality_summary": {"passed": True}})
    final_path = tmp_path / "delivered.xlsx"
    _write_qa_source_workbook(final_path, translated=True)
    db.add_artifact(project["id"], "delivered final", final_path, "qa_final_workbook", run_id=run["id"])
    db.update_run(run["id"], status="passed", metadata={"quality_summary": {"passed": True}})

    with TestClient(app) as client:
        listed = client.get(f"/api/projects/{project['id']}/deliverables")
        assert listed.status_code == 200
        assert [item["run_id"] for item in listed.json()["deliverables"]] == [run["id"]]

        final_path.unlink()

        refreshed = client.get(f"/api/projects/{project['id']}/deliverables")
        assert refreshed.status_code == 200
        assert refreshed.json()["deliverables"] == []

