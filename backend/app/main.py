from __future__ import annotations

import mimetypes
import json
import os
import shutil
import sys
import hashlib
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any

import uvicorn
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import load_workbook

if __package__ is None or __package__ == "":
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from app import db
    from app.config import DATA_ROOT, load_settings, public_settings, save_settings
    from app.jobs import active_job_id, cancel_singleton_job, start_singleton_job
    from app.languages import require_supported_language
    from app.schemas import (
        AnnouncementLookupRequest,
        AnnouncementDocxApplyRequest,
        AnnouncementDocxDeliverRequest,
        AnnouncementDocxImportAiRequest,
        AnnouncementDocxPrepareRequest,
        AnnouncementTaskActionRequest,
        AnnouncementTaskApplyRequest,
        AnnouncementTaskCreateRequest,
        AnnouncementTaskDeliverRequest,
        AnnouncementTaskImportAiRequest,
        AnnouncementTaskTermsRequest,
        AnnouncementTaskTranslateRequest,
        AnnouncementTermsRequest,
        ArtifactUpdate,
        GlossaryExtractRequest,
        GlossaryBatchResolveRequest,
        GlossaryCandidateUpdate,
        GlossaryImportRequest,
        GlossaryTermPayload,
        GlossaryTermUpdate,
        ManualFixRequest,
        ModelFixRequest,
        ProjectHarnessUpdate,
        ProjectAnalysisRequest,
        ProjectCreate,
        ProjectUpdate,
        RunCreate,
        SettingsUpdate,
        TranslateRequest,
        TranslationArchiveImportRequest,
        TranslationEntryPayload,
        TranslationEntryUpdate,
    )
    from app.workflow import (
        analyze_assets,
        apply_manual_fixes,
        apply_model_fixes,
        build_delivery_package,
        create_improvement_review,
        create_semantic_qa_context,
        export_glossary,
        export_translation_archive,
        extract_glossary,
        harness_overview,
        import_glossary,
        import_translation_archive,
        inspect_translation_readiness,
        list_glossary_wide,
        list_project_deliverables,
        list_improvements,
        list_quality_issues,
        list_translation_archive_wide,
        preview_glossary_import,
        project_dir,
        read_project_harness,
        apply_announcement_task,
        cancel_announcement_task,
        cancel_announcement_translation_task,
        create_announcement_task,
        deliver_announcement_task,
        extract_announcement_terms,
        generate_announcement_terms_package,
        get_announcement_task,
        import_announcement_ai_response,
        import_announcement_terms,
        inspect_announcement_constraints,
        legacy_apply_announcement_docx,
        legacy_deliver_announcement_docx,
        legacy_import_announcement_docx_ai,
        legacy_prepare_announcement_docx,
        list_announcement_tasks,
        lookup_announcement_translations,
        prepare_announcement_translation,
        run_announcement_lookup,
        cancel_translation_run,
        reconcile_interrupted_background_jobs,
        translation_batch_file,
        translation_run_progress,
        translate_announcement_task,
        run_qa_sync,
        run_translate_sync,
        translate_missing_glossary_candidates_sync,
        write_project_harness,
        write_project_prompt,
    )
else:
    from . import db
    from .config import DATA_ROOT, load_settings, public_settings, save_settings
    from .jobs import active_job_id, cancel_singleton_job, start_singleton_job
    from .languages import require_supported_language
    from .schemas import (
        AnnouncementLookupRequest,
        AnnouncementDocxApplyRequest,
        AnnouncementDocxDeliverRequest,
        AnnouncementDocxImportAiRequest,
        AnnouncementDocxPrepareRequest,
        AnnouncementTaskActionRequest,
        AnnouncementTaskApplyRequest,
        AnnouncementTaskCreateRequest,
        AnnouncementTaskDeliverRequest,
        AnnouncementTaskImportAiRequest,
        AnnouncementTaskTermsRequest,
        AnnouncementTaskTranslateRequest,
        AnnouncementTermsRequest,
        ArtifactUpdate,
        GlossaryExtractRequest,
        GlossaryBatchResolveRequest,
        GlossaryCandidateUpdate,
        GlossaryImportRequest,
        GlossaryTermPayload,
        GlossaryTermUpdate,
        ManualFixRequest,
        ModelFixRequest,
        ProjectHarnessUpdate,
        ProjectAnalysisRequest,
        ProjectCreate,
        ProjectUpdate,
        RunCreate,
        SettingsUpdate,
        TranslateRequest,
        TranslationArchiveImportRequest,
        TranslationEntryPayload,
        TranslationEntryUpdate,
    )
    from .workflow import (
        analyze_assets,
        apply_manual_fixes,
        apply_model_fixes,
        build_delivery_package,
        create_improvement_review,
        create_semantic_qa_context,
        export_glossary,
        export_translation_archive,
        extract_glossary,
        harness_overview,
        import_glossary,
        import_translation_archive,
        inspect_translation_readiness,
        list_glossary_wide,
        list_project_deliverables,
        list_improvements,
        list_quality_issues,
        list_translation_archive_wide,
        preview_glossary_import,
        project_dir,
        read_project_harness,
        apply_announcement_task,
        cancel_announcement_task,
        cancel_announcement_translation_task,
        create_announcement_task,
        deliver_announcement_task,
        extract_announcement_terms,
        generate_announcement_terms_package,
        get_announcement_task,
        import_announcement_ai_response,
        import_announcement_terms,
        inspect_announcement_constraints,
        legacy_apply_announcement_docx,
        legacy_deliver_announcement_docx,
        legacy_import_announcement_docx_ai,
        legacy_prepare_announcement_docx,
        list_announcement_tasks,
        lookup_announcement_translations,
        prepare_announcement_translation,
        run_announcement_lookup,
        cancel_translation_run,
        reconcile_interrupted_background_jobs,
        translation_batch_file,
        translation_run_progress,
        translate_announcement_task,
        run_qa_sync,
        run_translate_sync,
        translate_missing_glossary_candidates_sync,
        write_project_harness,
        write_project_prompt,
    )


@asynccontextmanager
async def lifespan(app: FastAPI):
    _ = app
    db.init_db()
    reconcile_interrupted_background_jobs()
    yield



def _cors_origins() -> list[str]:
    defaults = ["http://localhost:5173", "http://127.0.0.1:5173"]
    configured = os.environ.get("LWS_CORS_ORIGINS", "")
    extra = [origin.strip() for origin in configured.split(",") if origin.strip()]
    return [*defaults, *[origin for origin in extra if origin not in defaults]]


app = FastAPI(title="Localization Workflow Studio", version="0.4.9", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins(),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _query_language(language: str | None) -> str | None:
    if not language:
        return None
    try:
        return require_supported_language(language)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/health")
def health() -> dict[str, Any]:
    return {"ok": True, "data_root": str(DATA_ROOT)}


@app.get("/api/settings")
def get_settings() -> dict[str, Any]:
    return public_settings()


@app.patch("/api/settings")
def patch_settings(payload: SettingsUpdate) -> dict[str, Any]:
    current = load_settings()
    updates = payload.model_dump(exclude_none=True)
    if "api_key" in updates and updates["api_key"] in {"", "configured"}:
        updates.pop("api_key")
    current.update(updates)
    saved = save_settings(current)
    return public_settings(saved)


@app.get("/api/projects")
def get_projects() -> list[dict[str, Any]]:
    return [_with_project_stats(project) for project in db.list_projects()]


@app.post("/api/projects")
def create_project(payload: ProjectCreate) -> dict[str, Any]:
    existing = db.find_project_by_name(payload.name)
    if existing:
        return {**_with_project_stats(existing), "duplicate": True}
    project = db.insert_project(payload.name, payload.type, payload.description, payload.icon)
    project_dir(project["id"])
    return {**_with_project_stats(project), "duplicate": False}


@app.get("/api/projects/{project_id}")
def get_project(project_id: str) -> dict[str, Any]:
    try:
        return _with_project_stats(db.get_project(project_id), include_details=True)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.patch("/api/projects/{project_id}")
def update_project(project_id: str, payload: ProjectUpdate) -> dict[str, Any]:
    try:
        updates = payload.model_dump(exclude_none=True)
        return _with_project_stats(db.update_project(project_id, updates), include_details=True)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str) -> dict[str, bool]:
    try:
        run_ids = [run["id"] for run in db.list_runs(project_id)]
        db.delete_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    shutil.rmtree(DATA_ROOT / "projects" / project_id, ignore_errors=True)
    for run_id in run_ids:
        shutil.rmtree(DATA_ROOT / "runs" / run_id, ignore_errors=True)
    return {"deleted": True}


@app.get("/api/projects/{project_id}/harness")
def get_project_harness(project_id: str) -> dict[str, Any]:
    try:
        return harness_overview(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.patch("/api/projects/{project_id}/harness")
def patch_project_harness(project_id: str, payload: ProjectHarnessUpdate) -> dict[str, Any]:
    try:
        harness = write_project_harness(project_id, payload.model_dump(exclude_none=True))
        return {"global_harness": harness_overview(project_id)["global_harness"], "project_harness": harness}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.post("/api/projects/{project_id}/analyze")
def analyze_project(project_id: str, payload: ProjectAnalysisRequest) -> dict[str, Any]:
    try:
        project = db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    try:
        target_language = require_supported_language(payload.target_language)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    notes = analyze_assets(payload.asset_artifact_ids, load_settings())
    profile_path, prompt_path, brief_path, prompt = write_project_prompt(project, payload.intro, notes, target_language=target_language)
    artifacts = [
        db.add_artifact(project_id, "Project profile", profile_path, "project_profile", mime="application/json"),
        db.add_artifact(project_id, "Translation prompt", prompt_path, "translation_prompt", mime="text/plain"),
        db.add_artifact(project_id, "Project brief", brief_path, "project_brief", mime="text/markdown"),
    ]
    return {"project": _with_project_stats(db.get_project(project_id), include_details=True), "artifacts": artifacts, "prompt": prompt}


@app.post("/api/projects/{project_id}/files")
def upload_project_file(project_id: str, file: UploadFile = File(...), kind: str = "upload") -> dict[str, Any]:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    safe_name = _safe_filename(file.filename or "upload.bin")
    upload_bytes = file.file.read()
    digest = hashlib.sha256(upload_bytes).hexdigest()
    if kind == "asset":
        duplicate = _find_duplicate_project_upload(project_id, kind, digest)
        if duplicate:
            duplicate["duplicate"] = True
            return duplicate
    destination = _unique_path(project_dir(project_id) / "uploads" / safe_name)
    with destination.open("wb") as fh:
        fh.write(upload_bytes)
    mime = file.content_type or mimetypes.guess_type(str(destination))[0] or "application/octet-stream"
    artifact = db.add_artifact(
        project_id,
        safe_name,
        destination,
        kind,
        mime=mime,
        origin="uploaded",
        metadata={"sha256": digest, "original_filename": safe_name},
    )
    artifact["duplicate"] = False
    return artifact


@app.get("/api/projects/{project_id}/assets")
def list_project_assets(project_id: str, role: str | None = None, origin: str | None = None, run_id: str | None = None) -> list[dict[str, Any]]:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    return db.list_artifacts(project_id=project_id, run_id=run_id, role=role, origin=origin)


@app.get("/api/artifacts/{artifact_id}/translation-readiness")
def artifact_translation_readiness(artifact_id: str, batch_size: int | None = None, language: str = "en") -> dict[str, Any]:
    try:
        return inspect_translation_readiness(artifact_id, batch_size=batch_size, language=require_supported_language(language))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/glossary")
def list_project_glossary(project_id: str, language: str | None = None) -> list[dict[str, Any]]:
    return db.list_glossary_terms(project_id, language=_query_language(language))


@app.get("/api/projects/{project_id}/glossary/wide")
def list_project_glossary_wide(project_id: str) -> dict[str, Any]:
    try:
        return list_glossary_wide(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.get("/api/projects/{project_id}/glossary/batches")
def list_project_glossary_batches(project_id: str, language: str | None = None) -> dict[str, Any]:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    language_code = _query_language(language)
    batches = db.list_glossary_batches(project_id, language=language_code)
    latest = batches[0] if batches else None
    candidates = db.list_glossary_candidates(project_id, batch_id=latest["id"], language=language_code) if latest else []
    return {"batches": batches, "active_batch": latest, "candidates": candidates}


@app.patch("/api/projects/{project_id}/glossary/candidates/{candidate_id}")
def update_project_glossary_candidate(project_id: str, candidate_id: str, payload: GlossaryCandidateUpdate) -> dict[str, Any]:
    candidate = _require_project_candidate(project_id, candidate_id)
    _ = candidate
    data = payload.model_dump(exclude_unset=True)
    if "language" in data:
        data["language"] = _query_language(data.get("language")) or "en"
    return db.update_glossary_candidate(candidate_id, data)


@app.post("/api/projects/{project_id}/glossary/batches/{batch_id}/accept")
def accept_project_glossary_candidates(project_id: str, batch_id: str, payload: GlossaryBatchResolveRequest) -> dict[str, Any]:
    _require_project_batch(project_id, batch_id)
    return db.accept_glossary_candidates(project_id, batch_id, payload.candidate_ids or None)


@app.post("/api/projects/{project_id}/glossary/batches/{batch_id}/translate-missing")
def translate_missing_project_glossary_candidates(project_id: str, batch_id: str) -> dict[str, Any]:
    _require_project_batch(project_id, batch_id)
    try:
        return translate_missing_glossary_candidates_sync(project_id, batch_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/glossary/batches/{batch_id}/reject")
def reject_project_glossary_candidates(project_id: str, batch_id: str, payload: GlossaryBatchResolveRequest) -> dict[str, Any]:
    _require_project_batch(project_id, batch_id)
    return db.reject_glossary_candidates(project_id, batch_id, payload.candidate_ids or None)


@app.post("/api/projects/{project_id}/glossary")
def create_glossary_term(project_id: str, payload: GlossaryTermPayload) -> dict[str, Any]:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    data = payload.model_dump()
    data["language"] = _query_language(data.get("language")) or "en"
    return db.upsert_glossary_term(project_id, data)


@app.patch("/api/projects/{project_id}/glossary/{term_id}")
def update_glossary_term(project_id: str, term_id: str, payload: GlossaryTermUpdate) -> dict[str, Any]:
    _require_project_term(project_id, term_id)
    data = payload.model_dump(exclude_unset=True)
    if "language" in data:
        data["language"] = _query_language(data.get("language")) or "en"
    updated = db.update_glossary_term(term_id, data)
    db.dedupe_project_glossary_terms(project_id, preferred_term_id=term_id, merge_duplicates=False, language=updated.get("language"))
    return db.get_glossary_term(updated["id"])


@app.delete("/api/projects/{project_id}/glossary/{term_id}")
def delete_glossary_term(project_id: str, term_id: str) -> dict[str, bool]:
    _require_project_term(project_id, term_id)
    db.delete_glossary_term(term_id)
    return {"deleted": True}


@app.post("/api/projects/{project_id}/glossary/import-preview")
def preview_project_glossary_import(project_id: str, payload: GlossaryImportRequest) -> dict[str, Any]:
    try:
        return preview_glossary_import(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project, artifact, or column not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/glossary/import")
def import_project_glossary(project_id: str, payload: GlossaryImportRequest) -> dict[str, Any]:
    try:
        payload.language = _query_language(payload.language) or "en"
        return import_glossary(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project, artifact, or column not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/glossary/export")
def export_project_glossary(project_id: str, format: str = "xlsx", language: str | None = None) -> Any:
    try:
        exported = export_glossary(project_id, format, language=_query_language(language))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if isinstance(exported, dict):
        return exported
    media_type = "text/csv" if exported.suffix.lower() == ".csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(exported, media_type=media_type, filename=exported.name)


@app.get("/api/projects/{project_id}/translations")
def list_project_translations(project_id: str, language: str | None = None) -> list[dict[str, Any]]:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    return db.list_translation_entries(project_id, language=_query_language(language))


@app.get("/api/projects/{project_id}/translations/wide")
def list_project_translations_wide(project_id: str) -> dict[str, Any]:
    try:
        return list_translation_archive_wide(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.post("/api/projects/{project_id}/translations")
def create_translation_entry(project_id: str, payload: TranslationEntryPayload) -> dict[str, Any]:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    data = payload.model_dump()
    data["language"] = _query_language(data.get("language")) or "en"
    return db.upsert_translation_entry(project_id, data)


@app.patch("/api/projects/{project_id}/translations/{entry_id}")
def update_translation_entry(project_id: str, entry_id: str, payload: TranslationEntryUpdate) -> dict[str, Any]:
    _require_project_translation(project_id, entry_id)
    data = payload.model_dump(exclude_unset=True)
    if "language" in data:
        data["language"] = _query_language(data.get("language")) or "en"
    return db.update_translation_entry(entry_id, data)


@app.delete("/api/projects/{project_id}/translations/{entry_id}")
def delete_translation_entry(project_id: str, entry_id: str) -> dict[str, bool]:
    _require_project_translation(project_id, entry_id)
    db.delete_translation_entry(entry_id)
    return {"deleted": True}


@app.post("/api/projects/{project_id}/translations/import")
def import_project_translations(project_id: str, payload: TranslationArchiveImportRequest) -> dict[str, Any]:
    try:
        return import_translation_archive(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project or artifact not found") from exc
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/translations/export")
def export_project_translations(project_id: str, format: str = "xlsx", language: str | None = None) -> Any:
    try:
        exported = export_translation_archive(project_id, format, language=_query_language(language))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if isinstance(exported, dict):
        return exported
    media_type = "text/csv" if exported.suffix.lower() == ".csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(exported, media_type=media_type, filename=exported.name)


@app.get("/api/projects/{project_id}/deliverables")
def get_project_deliverables(project_id: str) -> dict[str, Any]:
    try:
        deliverables = list_project_deliverables(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    for deliverable in deliverables:
        _attach_delivery_downloads(project_id, deliverable)
    return {"project_id": project_id, "deliverables": deliverables}


@app.post("/api/projects/{project_id}/delivery-package")
def create_project_delivery(project_id: str, run_id: str | None = None) -> dict[str, Any]:
    try:
        package = build_delivery_package(project_id, run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    for item in package["files"]:
        item["download_url"] = f"/api/projects/{project_id}/delivery/{item['filename']}"
    _attach_delivery_downloads(project_id, package["deliverable"])
    return package


@app.get("/api/projects/{project_id}/delivery/{filename}")
def download_project_delivery(project_id: str, filename: str) -> FileResponse:
    try:
        db.get_project(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    safe_name = _safe_filename(filename)
    path = project_dir(project_id) / "delivery" / safe_name
    if not path.exists():
        raise HTTPException(status_code=404, detail="delivery file missing")
    media_type = "text/plain"
    if path.suffix.lower() == ".md":
        media_type = "text/markdown"
    elif path.suffix.lower() == ".xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.post("/api/projects/{project_id}/announcement-lookup")
def create_announcement_lookup(project_id: str, payload: AnnouncementLookupRequest) -> dict[str, Any]:
    try:
        payload.language = _query_language(payload.language) or "en"
        return run_announcement_lookup(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/announcement-terms")
def create_announcement_terms(project_id: str, payload: AnnouncementTermsRequest) -> dict[str, Any]:
    try:
        return generate_announcement_terms_package(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/announcement-docx/prepare")
def prepare_announcement_docx(project_id: str, payload: AnnouncementDocxPrepareRequest) -> dict[str, Any]:
    try:
        return legacy_prepare_announcement_docx(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project, run, or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/announcement-docx/import-ai")
def import_announcement_docx_ai(project_id: str, payload: AnnouncementDocxImportAiRequest) -> dict[str, Any]:
    try:
        return legacy_import_announcement_docx_ai(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project, run, or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/announcement-docx/apply")
def apply_announcement_docx(project_id: str, payload: AnnouncementDocxApplyRequest) -> dict[str, Any]:
    try:
        return legacy_apply_announcement_docx(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project, run, or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/announcement-docx/deliver")
def deliver_announcement_docx(project_id: str, payload: AnnouncementDocxDeliverRequest) -> dict[str, Any]:
    try:
        return legacy_deliver_announcement_docx(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project, run, or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/announcement-tasks")
def list_project_announcement_tasks(project_id: str) -> list[dict[str, Any]]:
    try:
        return list_announcement_tasks(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.post("/api/projects/{project_id}/announcement-tasks")
def create_project_announcement_task(project_id: str, payload: AnnouncementTaskCreateRequest) -> dict[str, Any]:
    try:
        return create_announcement_task(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/announcement-tasks/{task_id}")
def get_project_announcement_task(task_id: str) -> dict[str, Any]:
    try:
        return get_announcement_task(task_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task not found") from exc


@app.post("/api/announcement-tasks/{task_id}/cancel")
def cancel_project_announcement_task(task_id: str) -> dict[str, Any]:
    try:
        return cancel_announcement_task(task_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task not found") from exc


@app.post("/api/announcement-tasks/{task_id}/inspect-constraints")
def inspect_project_announcement_constraints(task_id: str, payload: AnnouncementTaskActionRequest) -> dict[str, Any]:
    try:
        return inspect_announcement_constraints(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/extract-terms")
def extract_project_announcement_terms(task_id: str, payload: AnnouncementTaskActionRequest) -> dict[str, Any]:
    try:
        return extract_announcement_terms(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/import-terms")
def import_project_announcement_terms(task_id: str, payload: AnnouncementTaskTermsRequest) -> dict[str, Any]:
    try:
        return import_announcement_terms(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/lookup-translations")
def lookup_project_announcement_translations(task_id: str, payload: AnnouncementTaskActionRequest) -> dict[str, Any]:
    try:
        return lookup_announcement_translations(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/prepare")
def prepare_project_announcement_translation(task_id: str, payload: AnnouncementTaskActionRequest) -> dict[str, Any]:
    try:
        return prepare_announcement_translation(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/translate")
def translate_project_announcement(task_id: str, payload: AnnouncementTaskTranslateRequest) -> dict[str, Any]:
    try:
        return translate_announcement_task(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


def _start_announcement_translation_background(task_id: str, payload: AnnouncementTaskTranslateRequest) -> dict[str, Any]:
    try:
        task = get_announcement_task(task_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task not found") from exc
    active = active_job_id()
    job_id = f"announcement:{task_id}"
    if active and active != job_id:
        raise HTTPException(status_code=409, detail=f"another long-text AI job is active: {active}")
    db.update_announcement_task(task_id, status="queued", current_step=7, metadata={**(task.get("metadata") or {}), "queued_at": db.now_iso()})

    def worker(cancel_event: Any) -> None:
        try:
            translate_announcement_task(task_id, payload, cancel_event=cancel_event)
        except Exception as exc:
            try:
                current = db.get_announcement_task(task_id)
                if current.get("status") not in {"translated", "canceled", "needs_input", "awaiting_ai_response", "prepared"}:
                    db.update_announcement_task(task_id, status="failed", current_step=7, metadata={**(current.get("metadata") or {}), "error": str(exc)})
            except Exception:
                pass

    started, active_conflict = start_singleton_job(job_id, worker)
    if not started and active_conflict:
        raise HTTPException(status_code=409, detail=f"another long-text AI job is active: {active_conflict}")
    return {"task": get_announcement_task(task_id), "summary": {"status": "queued"}}


@app.post("/api/announcement-tasks/{task_id}/translate/start")
def translate_project_announcement_start(task_id: str, payload: AnnouncementTaskTranslateRequest) -> dict[str, Any]:
    return _start_announcement_translation_background(task_id, payload)


@app.post("/api/announcement-tasks/{task_id}/translate/resume")
def translate_project_announcement_resume(task_id: str, payload: AnnouncementTaskTranslateRequest) -> dict[str, Any]:
    return _start_announcement_translation_background(task_id, payload)


@app.post("/api/announcement-tasks/{task_id}/translate/cancel")
def translate_project_announcement_cancel(task_id: str) -> dict[str, Any]:
    try:
        cancel_singleton_job(f"announcement:{task_id}")
        return {"task": cancel_announcement_translation_task(task_id)["task"], "summary": {"status": "prepared", "reason": "announcement_translation_canceled"}}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task not found") from exc


@app.post("/api/announcement-tasks/{task_id}/import-ai")
def import_project_announcement_ai(task_id: str, payload: AnnouncementTaskImportAiRequest) -> dict[str, Any]:
    try:
        return import_announcement_ai_response(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/apply")
def apply_project_announcement(task_id: str, payload: AnnouncementTaskApplyRequest) -> dict[str, Any]:
    try:
        return apply_announcement_task(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/announcement-tasks/{task_id}/deliver")
def deliver_project_announcement(task_id: str, payload: AnnouncementTaskDeliverRequest) -> dict[str, Any]:
    try:
        return deliver_announcement_task(task_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="announcement task or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/projects/{project_id}/glossary/extract")
def extract_project_glossary(project_id: str, payload: GlossaryExtractRequest) -> dict[str, Any]:
    try:
        payload.language = _query_language(payload.language) or "en"
        return extract_glossary(project_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project or artifact not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/runs")
def create_run(payload: RunCreate) -> dict[str, Any]:
    try:
        db.get_project(payload.project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc
    try:
        language = require_supported_language(payload.language)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    active = [
        run
        for run in db.list_runs(payload.project_id)
        if run["kind"] == payload.kind and run["status"] in {"queued", "running"}
    ]
    if active:
        raise HTTPException(status_code=409, detail=f"{payload.kind} run already active for this project")
    metadata = {
        "input_artifact_id": payload.input_artifact_id,
        "term_artifact_id": payload.term_artifact_id,
        "batch_size": payload.batch_size,
        "task_origin": payload.task_origin or ("direct_import" if payload.kind == "qa" else "translation_run"),
        "source_run_id": payload.source_run_id,
        "task_code": _resolve_task_code(payload),
    }
    return db.insert_run(payload.project_id, payload.kind, language, metadata)


@app.get("/api/runs")
def list_runs(project_id: str | None = None) -> list[dict[str, Any]]:
    return db.list_runs(project_id)


@app.get("/api/runs/{run_id}")
def get_run(run_id: str) -> dict[str, Any]:
    try:
        run = db.get_run(run_id)
        run["events"] = db.list_events(run_id)
        run["artifacts"] = db.list_artifacts(run_id=run_id)
        return run
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc


@app.post("/api/runs/{run_id}/translate")
def translate(run_id: str, payload: TranslateRequest) -> dict[str, Any]:
    try:
        return run_translate_sync(run_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run or artifact not found") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


def _start_translation_background(run_id: str, payload: TranslateRequest) -> dict[str, Any]:
    try:
        run = db.get_run(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    if run["kind"] != "translation":
        raise HTTPException(status_code=400, detail="run is not a translation run")
    active = active_job_id()
    job_id = f"run:{run_id}"
    if active and active != job_id:
        raise HTTPException(status_code=409, detail=f"another long-text AI job is active: {active}")
    if run["status"] == "running" and active == job_id:
        return get_run(run_id)
    metadata = run.get("metadata", {})
    db.update_run(run_id, status="queued", metadata={**metadata, "queued_at": db.now_iso()})

    def worker(cancel_event: Any) -> None:
        try:
            run_translate_sync(run_id, payload, cancel_event=cancel_event)
        except Exception as exc:
            try:
                current = db.get_run(run_id)
                if current.get("status") not in {"failed", "canceled", "needs_input", "passed"}:
                    db.update_run(run_id, status="failed", metadata={**current.get("metadata", {}), "error": str(exc)})
            except Exception:
                pass

    started, active_conflict = start_singleton_job(job_id, worker)
    if not started and active_conflict:
        run = db.get_run(run_id)
        db.update_run(run_id, status=run.get("status") or "created", metadata={**run.get("metadata", {}), "queue_error": f"active job: {active_conflict}"})
        raise HTTPException(status_code=409, detail=f"another long-text AI job is active: {active_conflict}")
    db.add_event(run_id, "translation background job started")
    return get_run(run_id)


@app.post("/api/runs/{run_id}/translate/start")
def translate_start(run_id: str, payload: TranslateRequest) -> dict[str, Any]:
    return _start_translation_background(run_id, payload)


@app.post("/api/runs/{run_id}/translate/resume")
def translate_resume(run_id: str, payload: TranslateRequest) -> dict[str, Any]:
    return _start_translation_background(run_id, payload)


@app.post("/api/runs/{run_id}/translate/cancel")
def translate_cancel(run_id: str) -> dict[str, Any]:
    try:
        cancel_singleton_job(f"run:{run_id}")
        cancel_translation_run(run_id)
        return get_run(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc


@app.get("/api/runs/{run_id}/translate/progress")
def translate_progress(run_id: str) -> dict[str, Any]:
    try:
        return translation_run_progress(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc


@app.get("/api/runs/{run_id}/translate/batches/{batch_index}/{kind}")
def translate_batch_download(run_id: str, batch_index: int, kind: str) -> FileResponse:
    try:
        path = translation_batch_file(run_id, batch_index, kind)
        return FileResponse(path, filename=path.name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="batch file not found") from exc


@app.post("/api/runs/{run_id}/qa")
def qa(run_id: str) -> dict[str, Any]:
    try:
        return run_qa_sync(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run or artifact not found") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/runs/{run_id}/quality-issues")
def quality_issues(run_id: str) -> dict[str, Any]:
    try:
        return list_quality_issues(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc


@app.post("/api/runs/{run_id}/manual-fixes")
def manual_fixes(run_id: str, payload: ManualFixRequest) -> dict[str, Any]:
    try:
        return apply_manual_fixes(run_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run, artifact, sheet, or column not found") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/runs/{run_id}/model-fixes")
def model_fixes(run_id: str, payload: ModelFixRequest) -> dict[str, Any]:
    try:
        return apply_model_fixes(run_id, payload)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run, artifact, sheet, or column not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/runs/{run_id}/semantic-qa")
def semantic_qa(run_id: str) -> dict[str, Any]:
    try:
        return create_semantic_qa_context(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc


@app.get("/api/runs/{run_id}/events")
def get_events(run_id: str) -> list[dict[str, Any]]:
    return db.list_events(run_id)


@app.get("/api/projects/{project_id}/improvements")
def get_project_improvements(project_id: str) -> list[dict[str, Any]]:
    try:
        db.get_project(project_id)
        return list_improvements(project_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="project not found") from exc


@app.post("/api/runs/{run_id}/improvement-review")
def improvement_review(run_id: str) -> dict[str, Any]:
    try:
        return create_improvement_review(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc


@app.get("/api/artifacts/{artifact_id}/download")
def download_artifact(artifact_id: str) -> FileResponse:
    try:
        artifact = db.get_artifact(artifact_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="artifact not found") from exc
    path = Path(artifact["path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="artifact file missing")
    try:
        path.resolve().relative_to(DATA_ROOT.resolve())
    except ValueError as exc:
        raise HTTPException(status_code=403, detail="artifact path is outside data root") from exc
    return FileResponse(path, media_type=artifact["mime"], filename=path.name)


@app.patch("/api/artifacts/{artifact_id}")
def patch_artifact(artifact_id: str, payload: ArtifactUpdate) -> dict[str, Any]:
    try:
        return db.update_artifact(artifact_id, payload.model_dump(exclude_none=True))
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="artifact not found") from exc


def _attach_delivery_downloads(project_id: str, deliverable: dict[str, Any]) -> None:
    files = deliverable.get("files") if isinstance(deliverable.get("files"), dict) else {}
    for item in files.values():
        if item.get("path"):
            item["download_url"] = f"/api/projects/{project_id}/delivery/{item['filename']}"
        else:
            item["download_url"] = ""


def _resolve_task_code(payload: RunCreate) -> str:
    if payload.source_run_id:
        try:
            source = db.get_run(payload.source_run_id)
            if source["project_id"] == payload.project_id:
                source_code = str((source.get("metadata") or {}).get("task_code") or "").upper()
                if source_code in {"A", "T", "QA"}:
                    return source_code
                if source["kind"] == "translation":
                    return "T"
                if source["kind"] == "qa":
                    return "QA"
        except KeyError:
            pass
    task_code = str(payload.task_code or "").upper()
    if task_code in {"A", "T", "QA"}:
        return task_code
    if payload.kind == "translation":
        return "T"
    if payload.kind == "qa":
        return "QA"
    return str(payload.kind or "TASK").upper()


def _safe_filename(name: str) -> str:
    cleaned = "".join(ch for ch in name if ch not in '<>:"/\\|?*').strip()
    return cleaned or "upload.bin"


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    index = 2
    while True:
        candidate = parent / f"{stem}-{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def _find_duplicate_project_upload(project_id: str, kind: str, digest: str) -> dict[str, Any] | None:
    for artifact in db.list_artifacts(project_id=project_id):
        if artifact.get("kind") != kind:
            continue
        metadata = dict(artifact.get("metadata") or {})
        existing_digest = metadata.get("sha256")
        if not existing_digest:
            existing_digest = _file_sha256(Path(artifact.get("path") or ""))
            if existing_digest:
                metadata["sha256"] = existing_digest
                artifact = db.update_artifact(artifact["id"], {"metadata": metadata})
        if existing_digest == digest:
            return artifact
    return None


def _file_sha256(path: Path) -> str | None:
    if not path.exists() or not path.is_file():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _require_project_term(project_id: str, term_id: str) -> dict[str, Any]:
    try:
        term = db.get_glossary_term(term_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="glossary term not found") from exc
    if term["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="glossary term not found")
    return term


def _require_project_batch(project_id: str, batch_id: str) -> dict[str, Any]:
    try:
        batch = db.get_glossary_batch(batch_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="glossary batch not found") from exc
    if batch["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="glossary batch not found")
    return batch


def _require_project_candidate(project_id: str, candidate_id: str) -> dict[str, Any]:
    try:
        candidate = db.get_glossary_candidate(candidate_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="glossary candidate not found") from exc
    if candidate["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="glossary candidate not found")
    return candidate


def _require_project_translation(project_id: str, entry_id: str) -> dict[str, Any]:
    try:
        entry = db.get_translation_entry(entry_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="translation entry not found") from exc
    if entry["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="translation entry not found")
    return entry


def _with_project_stats(project: dict[str, Any], include_details: bool = False) -> dict[str, Any]:
    artifacts = db.list_artifacts(project_id=project["id"])
    runs = db.list_runs(project["id"])
    terms = db.list_glossary_terms(project["id"])
    translation_entries = db.list_translation_entries(project["id"])
    announcement_tasks = list_announcement_tasks(project["id"])
    archive_metrics = _translation_archive_metrics(translation_entries)
    translation_runs = len([run for run in runs if run["kind"] == "translation"])
    qa_runs = len([run for run in runs if run["kind"] == "qa"])
    project["stats"] = {
        "tasks": len(runs),
        "announcement_tasks": len(announcement_tasks),
        "translation_runs": translation_runs,
        "qa_runs": qa_runs,
        "words": str(archive_metrics["source_chars"]),
        "archived_rows": archive_metrics["archived_rows"],
        "langs": len(archive_metrics["languages"]),
        "glossary": len(terms),
    }
    if include_details:
        project["artifacts"] = artifacts
        project["runs"] = runs
        project["glossary"] = terms
        project["translations"] = translation_entries
        project["announcement_tasks"] = announcement_tasks
        project["harness"] = read_project_harness(project["id"])
    return project


def _translation_archive_metrics(entries: list[dict[str, Any]]) -> dict[str, Any]:
    source_chars = 0
    archived_rows = 0
    languages: set[str] = set()
    for entry in entries:
        source = str(entry.get("source") or "").strip()
        target = str(entry.get("target") or "").strip()
        if not source or not target:
            continue
        archived_rows += 1
        source_chars += len("".join(source.split()))
        languages.add(str(entry.get("language") or "en").lower())
    return {"source_chars": source_chars, "archived_rows": archived_rows, "languages": languages}


def _translation_workbook_metrics(artifact: dict[str, Any], runs_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if artifact.get("role") != "translation_workbook":
        return {"source_chars": 0, "valid_rows": 0, "language": ""}
    path = Path(artifact["path"])
    if not path.exists():
        return {"source_chars": 0, "valid_rows": 0, "language": ""}
    run = runs_by_id.get(artifact.get("run_id") or "")
    language = (artifact.get("metadata") or {}).get("language") or (run or {}).get("language") or "en"
    source_chars = 0
    valid_rows = 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                try:
                    header_row = next(ws.iter_rows(min_row=1, max_row=1))
                except StopIteration:
                    continue
                headers = {
                    str(cell.value or "").strip().lower(): index
                    for index, cell in enumerate(header_row, start=1)
                    if cell.value is not None
                }
                source_col = _first_header(headers, ["cn", "source", "original", "中文", "原文"])
                target_col = _first_header(headers, ["en", "translation", "target", "英文", "译文"])
                if source_col is None or target_col is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    source = _row_value(row, source_col)
                    target = _row_value(row, target_col)
                    if source and target:
                        valid_rows += 1
                        source_chars += len("".join(str(source).split()))
        finally:
            wb.close()
    except Exception:
        return {"source_chars": 0, "valid_rows": 0, "language": ""}
    return {"source_chars": source_chars, "valid_rows": valid_rows, "language": language}


def _first_header(headers: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        hit = headers.get(name.lower())
        if hit is not None:
            return hit
    return None


def _row_value(row: tuple[Any, ...], column: int) -> str:
    if column < 1 or column > len(row):
        return ""
    value = row[column - 1]
    return "" if value is None else str(value).strip()


if __name__ == "__main__":
    uvicorn.run("app.main:app", host="127.0.0.1", port=8000, reload=False, app_dir=str(Path(__file__).resolve().parents[1]))
