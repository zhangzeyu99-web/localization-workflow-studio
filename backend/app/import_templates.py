from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import DATA_ROOT


TEMPLATE_DIR = DATA_ROOT / "templates"


TEMPLATE_KINDS: dict[str, dict[str, Any]] = {
    "language-table": {
        "filename": "语言表导入模板.xlsx",
        "sheet": "语言表",
        "headers": ["ID", "CN", "EN", "EN2", "KR", "JP", "备注"],
        "rows": [
            ["1001", "开始游戏", "", "", "", "", "翻译任务：只填 ID 和 CN 也可以，系统会生成译文"],
            ["1002", "领取奖励", "Claim Reward", "", "보상 받기", "報酬を受け取る", "校对/归档任务：已有译文可填入对应语言列"],
        ],
        "notes": [
            "适用入口：新翻译任务 STEP 4、快速任务投入内容、译文归档导入。",
            "必填列：ID、CN。ID 用于回填，必须稳定且不重复。",
            "语言列：EN / KR / JP；英文可选 EN2 作为第二译法，韩语和日语默认不使用 KR2/JP2。",
            "如果只是要翻译，目标语言列可以留空；如果是校对或归档，请填入已有译文。",
        ],
    },
    "glossary": {
        "filename": "项目术语表导入模板.xlsx",
        "sheet": "项目术语表",
        "headers": ["ID", "CN", "EN", "EN2", "KR", "JP", "分类", "备注"],
        "rows": [
            ["term_001", "战力", "CP", "Power", "전투력", "戦力", "战斗属性/数值", "短 UI 词优先统一"],
            ["term_002", "联盟", "Alliance", "", "연맹", "同盟", "系统/组织", "专有名词保持一致"],
        ],
        "notes": [
            "适用入口：新翻译任务 STEP 3、项目术语表导入。",
            "必填列：CN，以及至少一个目标语言列。ID 可填业务 ID；没有 ID 时系统会按中文去重。",
            "语言列：EN / EN2 / KR / JP。EN2 只给英文第二译法使用。",
            "分类和备注可选，但建议填写，方便后续 QA 判断术语是否必须命中。",
        ],
    },
    "announcement-language-table": {
        "filename": "公告约束语言表模板.xlsx",
        "sheet": "完整语言表",
        "headers": ["ID", "CN", "EN", "EN2", "KR", "JP", "备注"],
        "rows": [
            ["2001", "限时活动", "Limited-time Event", "", "기간 한정 이벤트", "期間限定イベント", "公告反查会用 CN 命中原文，再取目标语言译文"],
            ["2002", "累计充值", "Cumulative Top-up", "", "누적 충전", "累計チャージ", "完整语言表或术语交付表都可以按此结构上传"],
        ],
        "notes": [
            "适用入口：公告翻译 STEP 2 约束来源。",
            "用途：从公告原文反查已有中文词条，再按目标语言取译文，生成公告临时术语表。",
            "必填列：CN，以及本次目标语言列。ID 建议保留，便于追溯来源。",
            "不要在这里上传公告原文 TXT/DOCX；公告原文请放在 STEP 1。",
        ],
    },
    "announcement-terms": {
        "filename": "公告术语表导入模板.xlsx",
        "sheet": "公告术语表",
        "headers": ["ID", "CN", "EN", "KR", "JP", "命中次数", "来源", "备注"],
        "rows": [
            ["ann_001", "限时活动", "Limited-time Event", "기간 한정 이벤트", "期間限定イベント", 3, "语言表反查", "本次公告必须统一"],
            ["ann_002", "新英雄", "New Hero", "신규 영웅", "新英雄", 1, "人工补充", "可在 STEP 4 编辑后保存"],
        ],
        "notes": [
            "适用入口：公告翻译 STEP 4 上传已提取术语表。",
            "用途：模拟或复用已经提取好的公告临时术语，不会自动写回项目术语库。",
            "必填列：CN，以及至少一个目标语言列。命中次数、来源、备注是辅助信息。",
            "如果还没提取术语，优先点击“提取术语并 AI 复查”；这个模板用于人工补充或迁移已有结果。",
        ],
    },
}


def build_import_template(kind: str) -> Path:
    template = TEMPLATE_KINDS.get(str(kind or "").strip().lower())
    if not template:
        raise ValueError("unknown import template")
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    path = TEMPLATE_DIR / template["filename"]

    wb = Workbook()
    ws = wb.active
    ws.title = template["sheet"]
    _write_table_sheet(ws, template["headers"], template["rows"])

    guide = wb.create_sheet("填写说明", 0)
    _write_guide_sheet(guide, template)

    wb.save(path)
    wb.close()
    return path


def _write_table_sheet(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="3B2F6B")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    widths = {
        "ID": 16,
        "CN": 24,
        "EN": 28,
        "EN2": 22,
        "KR": 24,
        "JP": 24,
        "分类": 18,
        "备注": 44,
        "命中次数": 12,
        "来源": 16,
    }
    for index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(index)].width = widths.get(header, 20)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_guide_sheet(ws: Any, template: dict[str, Any]) -> None:
    ws["A1"] = template["filename"].replace(".xlsx", "")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "怎么填"
    ws["A3"].font = Font(bold=True)
    for offset, note in enumerate(template["notes"], start=4):
        ws.cell(offset, 1, f"{offset - 3}. {note}")
    ws["A10"] = "字段说明"
    ws["A10"].font = Font(bold=True)
    field_notes = [
        ("ID", "业务 ID / key。建议填写，后续回填和排查更稳定。"),
        ("CN", "中文原文或中文术语。导入、反查和 QA 都以这列为核心。"),
        ("EN / KR / JP", "目标语言译文列。只填本次需要的语言也可以。"),
        ("EN2", "英文第二译法，可选；韩语/日语默认不需要第二译法列。"),
        ("分类 / 来源 / 备注", "辅助信息，可选；建议写清用途、来源或人工确认点。"),
    ]
    for row_index, (field, desc) in enumerate(field_notes, start=11):
        ws.cell(row_index, 1, field).font = Font(bold=True)
        ws.cell(row_index, 2, desc)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
