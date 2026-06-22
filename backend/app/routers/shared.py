from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook

from .. import db
from ..delivery_naming import safe_filename
from ..download_urls import attach_delivery_item_downloads
from ..languages import SOURCE_HEADER_ALIASES, require_supported_language, target_aliases
from ..schemas import (
    RunCreate,
)
from ..workflow import (
    read_project_harness,
    list_announcement_tasks,
    user_facing_error,
)

def _query_language(language: str | None) -> str | None:
    if not language:
        return None
    try:
        return require_supported_language(language)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=user_facing_error(exc)) from exc

def _attach_delivery_downloads(project_id: str, deliverable: dict[str, Any]) -> None:
    files = deliverable.get("files") if isinstance(deliverable.get("files"), dict) else {}
    for value in files.values():
        items = value if isinstance(value, list) else [value]
        attach_delivery_item_downloads(project_id, [item for item in items if isinstance(item, dict)])


def _resolve_task_code(payload: RunCreate) -> str:
    if payload.source_run_id:
        try:
            source = db.get_run(payload.source_run_id)
            if source["project_id"] == payload.project_id:
                source_code = str((source.get("metadata") or {}).get("task_code") or "").upper()
                if source_code in {"A", "T", "QA"}:
                    return source_code
                if source["kind"] == "translation":
                    return "T"
                if source["kind"] == "qa":
                    return "QA"
        except KeyError:
            pass
    task_code = str(payload.task_code or "").upper()
    if task_code in {"A", "T", "QA"}:
        return task_code
    if payload.kind == "translation":
        return "T"
    if payload.kind == "qa":
        return "QA"
    return str(payload.kind or "TASK").upper()


def _safe_filename(name: str) -> str:
    return safe_filename(name)


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    index = 2
    while True:
        candidate = parent / f"{stem}-{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1




_TABLE_UPLOAD_SUFFIXES = {".xlsx", ".xls", ".csv"}
_QUICK_INPUT_SUFFIXES = _TABLE_UPLOAD_SUFFIXES | {".txt", ".md", ".markdown"}
_TERM_UPLOAD_SUFFIXES = {".xlsx", ".xls", ".csv", ".json"}
_ANNOUNCEMENT_SOURCE_SUFFIXES = {".docx", ".txt", ".xlsx"}
_AI_RESPONSE_SUFFIXES = {".json", ".jsonl"}


def _allowed_upload_suffixes(kind: str) -> set[str] | None:
    if kind == "quick_input":
        return _QUICK_INPUT_SUFFIXES
    if kind in {"language_table", "final_workbook"}:
        return _TABLE_UPLOAD_SUFFIXES
    if kind in {"term_base", "glossary_final"}:
        return _TERM_UPLOAD_SUFFIXES
    if kind == "announcement_terms_workbook":
        return {".xlsx", ".xls"}
    if kind in {"announcement_ai_response", "announcement_ai_supplement_response"}:
        return _AI_RESPONSE_SUFFIXES
    return None


def _upload_kind_error(kind: str, suffix: str) -> str:
    ext = suffix or "unknown"
    if kind == "quick_input":
        return (
            f"当前入口不支持 {ext} 文件。"
            "快速任务请上传 XLSX/XLS/CSV 语言表，或 TXT 文本。"
        )
    if kind in {"language_table", "final_workbook"}:
        return (
            f"\u5f53\u524d\u5165\u53e3\u4e0d\u652f\u6301 {ext} \u6587\u4ef6\u3002"
            "\u8bed\u8a00\u5305\u7ffb\u8bd1\u8bf7\u4e0a\u4f20 XLSX/XLS/CSV \u8bed\u8a00\u8868\uff1b"
            "TXT/DOCX \u957f\u6587\u672c\u8bf7\u4f7f\u7528\u516c\u544a\u7ffb\u8bd1/\u5916\u6587\u672c\u6d41\u7a0b\u3002"
        )
    if kind in {"term_base", "glossary_final"}:
        return "\u672f\u8bed\u8868\u8bf7\u4e0a\u4f20 XLSX/XLS/CSV/JSON \u6587\u4ef6\u3002"
    if kind == "announcement_terms_workbook":
        return "\u5df2\u63d0\u53d6\u516c\u544a\u672f\u8bed\u8868\u8bf7\u4e0a\u4f20 XLSX \u6587\u4ef6\u3002"
    return "\u6587\u4ef6\u7c7b\u578b\u4e0e\u5f53\u524d\u5165\u53e3\u4e0d\u5339\u914d\u3002"


def _validate_upload_kind_filename(kind: str, filename: str) -> None:
    allowed = _allowed_upload_suffixes(kind)
    if not allowed:
        return
    suffix = Path(filename).suffix.lower()
    if suffix not in allowed:
        raise HTTPException(status_code=400, detail=_upload_kind_error(kind, suffix))


def _validate_run_input_artifact(payload: RunCreate) -> None:
    if payload.kind not in {"translation", "qa"} or not payload.input_artifact_id:
        return
    try:
        artifact = db.get_artifact(payload.input_artifact_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="input artifact not found") from exc
    if artifact["project_id"] != payload.project_id:
        raise HTTPException(status_code=400, detail="input artifact does not belong to project")
    suffix = Path(str(artifact.get("path") or artifact.get("label") or "")).suffix.lower()
    if payload.kind == "translation" and payload.task_origin == "quick_task" and suffix in {".txt", ".md", ".markdown"}:
        return
    if suffix not in _TABLE_UPLOAD_SUFFIXES:
        raise HTTPException(status_code=400, detail=_upload_kind_error("language_table", suffix))

def _find_duplicate_project_upload(project_id: str, kind: str, digest: str) -> dict[str, Any] | None:
    for artifact in db.list_artifacts(project_id=project_id):
        if artifact.get("kind") != kind:
            continue
        metadata = dict(artifact.get("metadata") or {})
        existing_digest = metadata.get("sha256")
        if not existing_digest:
            existing_digest = _file_sha256(Path(artifact.get("path") or ""))
            if existing_digest:
                metadata["sha256"] = existing_digest
                artifact = db.update_artifact(artifact["id"], {"metadata": metadata})
        if existing_digest == digest:
            return artifact
    return None


def _file_sha256(path: Path) -> str | None:
    if not path.exists() or not path.is_file():
        return None
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _require_project_term(project_id: str, term_id: str) -> dict[str, Any]:
    try:
        term = db.get_glossary_term(term_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="glossary term not found") from exc
    if term["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="glossary term not found")
    return term


def _require_project_batch(project_id: str, batch_id: str) -> dict[str, Any]:
    try:
        batch = db.get_glossary_batch(batch_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="glossary batch not found") from exc
    if batch["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="glossary batch not found")
    return batch


def _require_project_candidate(project_id: str, candidate_id: str) -> dict[str, Any]:
    try:
        candidate = db.get_glossary_candidate(candidate_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="glossary candidate not found") from exc
    if candidate["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="glossary candidate not found")
    return candidate


def _require_project_translation(project_id: str, entry_id: str) -> dict[str, Any]:
    try:
        entry = db.get_translation_entry(entry_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="translation entry not found") from exc
    if entry["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="translation entry not found")
    return entry


def _with_project_stats(project: dict[str, Any], include_details: bool = False) -> dict[str, Any]:
    runs = db.list_runs(project["id"])
    announcement_tasks = list_announcement_tasks(project["id"])
    active_announcement_tasks = [task for task in announcement_tasks if task.get("status") != "canceled"]
    translation_runs = len([run for run in runs if run["kind"] == "translation"])
    qa_runs = len([run for run in runs if run["kind"] == "qa"])
    language_tasks = _language_business_task_count(runs)
    artifacts: list[dict[str, Any]] = []
    terms: list[dict[str, Any]] = []
    translation_entries: list[dict[str, Any]] = []
    if include_details:
        artifacts = db.list_artifacts(project_id=project["id"])
        terms = db.list_glossary_terms(project["id"])
        translation_entries = db.list_translation_entries(project["id"])
        archive_metrics = _translation_archive_metrics(translation_entries)
        deliverable_count = len([
            run for run in runs
            if run["kind"] in {"translation", "qa"}
            and any(
                artifact["run_id"] == run["id"]
                and artifact["kind"] == "qa_final_workbook"
                for artifact in artifacts
            )
        ])
        glossary_count = len(terms)
    else:
        archive_metrics = _translation_archive_metrics_fast(project["id"])
        deliverable_count = _project_delivery_count_fast(project["id"])
        glossary_count = _project_glossary_count_fast(project["id"])
    announcement_deliverable_count = len([
        task for task in active_announcement_tasks
        if task.get("status") == "delivered" and (task.get("metadata") or {}).get("delivery_artifact_id")
    ])
    business_tasks = language_tasks + len(active_announcement_tasks)
    project["stats"] = {
        "tasks": business_tasks,
        "execution_runs": len(runs),
        "language_tasks": language_tasks,
        "deliverables": deliverable_count + announcement_deliverable_count,
        "announcement_tasks": len(active_announcement_tasks),
        "translation_runs": translation_runs,
        "qa_runs": qa_runs,
        "words": str(archive_metrics["source_chars"]),
        "archived_rows": archive_metrics["archived_rows"],
        "langs": len(archive_metrics["languages"]),
        "glossary": glossary_count,
    }
    if include_details:
        project["artifacts"] = artifacts
        project["runs"] = runs
        project["glossary"] = terms
        project["translations"] = translation_entries
        project["announcement_tasks"] = announcement_tasks
        project["harness"] = read_project_harness(project["id"])
    return project


def _translation_archive_metrics_fast(project_id: str) -> dict[str, Any]:
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT
              COUNT(*) AS archived_rows,
              COUNT(DISTINCT language) AS language_count
            FROM translation_entries
            WHERE project_id = ?
            """,
            (project_id,),
        ).fetchone()
    languages = {f"lang_{index}" for index in range(int(row["language_count"] or 0))}
    return {
        # The project list is loaded on every page open/focus.  Avoid scanning
        # and measuring every archived source string here; exact character
        # counts are still returned by the detail endpoint.
        "source_chars": 0,
        "archived_rows": int(row["archived_rows"] or 0),
        "languages": languages,
    }


def _project_glossary_count_fast(project_id: str) -> int:
    with db.connect() as conn:
        row = conn.execute(
            "SELECT COUNT(*) AS count FROM glossary_terms WHERE project_id = ? AND confirmed = 1",
            (project_id,),
        ).fetchone()
    return int(row["count"] or 0)


def _project_delivery_count_fast(project_id: str) -> int:
    with db.connect() as conn:
        row = conn.execute(
            """
            SELECT COUNT(DISTINCT artifacts.run_id) AS count
            FROM artifacts
            JOIN runs ON runs.id = artifacts.run_id
            WHERE artifacts.project_id = ?
              AND artifacts.kind = 'qa_final_workbook'
              AND runs.kind IN ('translation', 'qa')
            """,
            (project_id,),
        ).fetchone()
    return int(row["count"] or 0)


def _language_business_task_count(runs: list[dict[str, Any]]) -> int:
    runs_by_id = {str(run["id"]): run for run in runs}
    business_ids: set[str] = set()
    for run in runs:
        if run["kind"] not in {"translation", "qa"}:
            continue
        metadata = run.get("metadata") or {}
        if metadata.get("announcement_task_id") or metadata.get("task_id"):
            continue
        business_ids.add(_language_business_root_id(run, runs_by_id))
    return len(business_ids)


def _language_business_root_id(run: dict[str, Any], runs_by_id: dict[str, dict[str, Any]], seen: set[str] | None = None) -> str:
    seen = seen or set()
    run_id = str(run["id"])
    if run_id in seen:
        return run_id
    seen.add(run_id)
    metadata = run.get("metadata") or {}
    for key in ("source_run_id", "manual_fix_source_run_id", "model_fix_source_run_id"):
        source_id = str(metadata.get(key) or "")
        source_run = runs_by_id.get(source_id)
        if source_run and source_run.get("kind") in {"translation", "qa"}:
            return _language_business_root_id(source_run, runs_by_id, seen)
    return run_id


def _translation_archive_metrics(entries: list[dict[str, Any]]) -> dict[str, Any]:
    source_chars = 0
    archived_rows = 0
    languages: set[str] = set()
    for entry in entries:
        source = str(entry.get("source") or "").strip()
        target = str(entry.get("target") or "").strip()
        if not source or not target:
            continue
        archived_rows += 1
        source_chars += len("".join(source.split()))
        languages.add(str(entry.get("language") or "en").lower())
    return {"source_chars": source_chars, "archived_rows": archived_rows, "languages": languages}


def _translation_workbook_metrics(artifact: dict[str, Any], runs_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if artifact.get("role") != "translation_workbook":
        return {"source_chars": 0, "valid_rows": 0, "language": ""}
    path = Path(artifact["path"])
    if not path.exists():
        return {"source_chars": 0, "valid_rows": 0, "language": ""}
    run = runs_by_id.get(artifact.get("run_id") or "")
    language = (artifact.get("metadata") or {}).get("language") or (run or {}).get("language") or "en"
    source_chars = 0
    valid_rows = 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                try:
                    header_row = next(ws.iter_rows(min_row=1, max_row=1))
                except StopIteration:
                    continue
                headers = {
                    str(cell.value or "").strip().lower(): index
                    for index, cell in enumerate(header_row, start=1)
                    if cell.value is not None
                }
                source_col = _first_header(headers, list(SOURCE_HEADER_ALIASES))
                target_col = _first_header(headers, target_aliases(language))
                if source_col is None or target_col is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    source = _row_value(row, source_col)
                    target = _row_value(row, target_col)
                    if source and target:
                        valid_rows += 1
                        source_chars += len("".join(str(source).split()))
        finally:
            wb.close()
    except Exception:
        return {"source_chars": 0, "valid_rows": 0, "language": ""}
    return {"source_chars": source_chars, "valid_rows": valid_rows, "language": language}


def _first_header(headers: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        hit = headers.get(name.lower())
        if hit is not None:
            return hit
    return None


def _row_value(row: tuple[Any, ...], column: int) -> str:
    if column < 1 or column > len(row):
        return ""
    value = row[column - 1]
    return "" if value is None else str(value).strip()

__all__ = [name for name in globals() if not name.startswith("__")]
