from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .. import db
from ..config import GLOSSARY_ROOT, REAL_PROVIDERS, load_settings, normalize_provider_name
from ..languages import language_spec, require_supported_language
from ..providers import translate_batch
from ..translation_batches import manage_project_prompt_context as _manage_project_prompt_context
from . import jsonl_helpers as _jsonl_helpers
from .common import project_dir, run_dir
from .materials import analyze_assets
from .naming import _safe_source_stem, _today_stamp
from .glossary_ai import supplement_language_table_glossary_candidates_with_ai
from .glossary_backfill import backfill_project_glossary_from_final
from .subprocess_runner import parse_key_output, run_subprocess, user_facing_error


_LEGACY_EN_ALT_HEADERS = frozenset({"en2", "en 2", "english2", "英语2", "英文2", "target_alt"})
_GENERATED_EN2_NOTE_REPLACEMENTS = {
    "LearningModel": "Curated rules keep approved EN decisions; observation store accumulates seen variants and usage drift.",
    "Columns": "ID = text id, CN = source term, EN = approved English.",
    "Rule": "Only the EN target column is included in newly generated artifacts.",
}
_GENERATED_EN2_TEXT_REPLACEMENTS = {
    "关键术语以随附术语表为准，EN 为标准译法，EN2 为项目中稳定出现的手动适配译法。":
        "关键术语以随附术语表为准，EN 为唯一主译。",
}


def _detect_target_column(path: Path, requested: str | None, language: str, sheet: str | None = None) -> str:
    explicit = str(requested or "").strip()
    if explicit:
        return explicit
    spec = language_spec(language)
    aliases = {str(alias).strip().casefold() for alias in (*spec.target_aliases, spec.target_header, spec.visible_code)}
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return spec.visible_code
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheets = [workbook[sheet]] if sheet and sheet in workbook.sheetnames else workbook.worksheets
        for worksheet in worksheets:
            for row in worksheet.iter_rows(min_row=1, max_row=20, values_only=True):
                for value in row:
                    header = str(value or "").strip()
                    if header.casefold() in aliases:
                        return header
    finally:
        workbook.close()
    return spec.visible_code


def _normalize_new_english_glossary_artifact(path: Path) -> None:
    if not path.exists():
        return
    workbook = load_workbook(path)
    try:
        changed = False
        for worksheet in workbook.worksheets:
            alternate_columns = [
                cell.column
                for cell in worksheet[1]
                if str(cell.value or "").strip().casefold() in _LEGACY_EN_ALT_HEADERS
            ]
            for column in reversed(alternate_columns):
                worksheet.delete_cols(column)
                changed = True
            if worksheet.title != "Notes":
                continue
            for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=2):
                label = str(row[0].value or "").strip()
                replacement = _GENERATED_EN2_NOTE_REPLACEMENTS.get(label)
                if replacement and isinstance(row[1].value, str) and "en2" in row[1].value.casefold():
                    row[1].value = replacement
                    changed = True
        if changed:
            workbook.save(path)
    finally:
        workbook.close()


def _normalize_new_english_glossary_text_artifact(path: Path) -> None:
    if not path.exists():
        return
    text = path.read_text(encoding="utf-8")
    normalized = text
    for legacy_text, replacement in _GENERATED_EN2_TEXT_REPLACEMENTS.items():
        normalized = normalized.replace(legacy_text, replacement)
    if normalized != text:
        path.write_text(normalized, encoding="utf-8")


def extract_glossary(project_id: str, request: Any) -> dict[str, Any]:
    project = db.get_project(project_id)
    artifact = db.get_artifact(request.input_artifact_id)
    language = require_supported_language(getattr(request, "language", "en") or "en")
    spec = language_spec(language)
    material_artifact_ids = list(getattr(request, "project_material_artifact_ids", []) or [])
    announcement_material_artifact_ids = list(getattr(request, "announcement_material_artifact_ids", []) or [])
    for material_artifact_id in [*material_artifact_ids, *announcement_material_artifact_ids]:
        if db.get_artifact(material_artifact_id)["project_id"] != project_id:
            raise KeyError(material_artifact_id)
    announcement_only = bool(getattr(request, "announcement_only", False))
    announcement_min_hit = max(1, int(getattr(request, "announcement_min_hit", 1) or 1))
    if announcement_only and not announcement_material_artifact_ids:
        raise ValueError("announcement_only requires announcement_material_artifact_ids")
    project_notes = [str(note).strip() for note in getattr(request, "project_notes", []) or [] if str(note).strip()]
    # Load settings once for this whole request: the material-notes analysis
    # and the AI candidate supplement below (if enabled) must observe the
    # same snapshot, even though the extraction subprocess in between may
    # take a while to run.
    settings = load_settings()
    material_notes = analyze_assets(material_artifact_ids, settings) if material_artifact_ids else []
    run = db.insert_run(
        project_id,
        kind="glossary",
        language=language,
        metadata={
            "input_artifact_id": request.input_artifact_id,
            "project_material_artifact_ids": material_artifact_ids,
            "announcement_material_artifact_ids": announcement_material_artifact_ids,
            "announcement_only": announcement_only,
            "announcement_min_hit": announcement_min_hit,
            "project_notes": project_notes,
            "project_material_notes": material_notes,
        },
    )
    db.update_run(run["id"], status="running")
    output_dir = run_dir(run["id"]) / "glossary"
    output_dir.mkdir(parents=True, exist_ok=True)
    input_path = Path(artifact["path"])
    target_column = _detect_target_column(input_path, request.target_column, language, request.sheet)
    detail_output = output_dir / f"{input_path.stem}_glossary_details.xlsx"
    final_suffix = f"{spec.target_header}_{spec.alt_header}" if spec.alt_header else spec.target_header
    final_output = output_dir / f"{input_path.stem}_ID_CN_{final_suffix}.xlsx"
    brief_output = output_dir / "project_brief.md"
    prompt_output = output_dir / "translation_prompt.txt"
    announcement_base = _safe_source_stem(input_path.name)
    announcement_output = output_dir / f"{announcement_base}_announcement_terms_{_today_stamp()}.xlsx" if announcement_material_artifact_ids else None
    ai_supplement = bool(getattr(request, "ai_supplement", False))
    ai_supplement_packet_output = output_dir / f"{announcement_base}_ai_supplement_packet_{_today_stamp()}.json" if ai_supplement else None
    ai_supplement_report_output = output_dir / f"{announcement_base}_ai_supplement_report_{_today_stamp()}.md" if ai_supplement else None
    args = [
        sys.executable,
        str(GLOSSARY_ROOT / "scripts" / "extract_glossary.py"),
        str(input_path),
        "--id-column",
        request.id_column,
        "--source-column",
        request.source_column,
        "--target-column",
        target_column,
        "--curated-rules",
        str(project_dir(project_id) / "glossary" / "curated_terms.json"),
        "--observations-store",
        str(project_dir(project_id) / "glossary" / "observed_terms.json"),
    ]
    if not announcement_only:
        args.extend(
            [
                "--output",
                str(detail_output),
                "--final-output",
                str(final_output),
                "--project-name",
                request.project_name or project["name"],
                "--project-brief-output",
                str(brief_output),
                "--translation-prompt-output",
                str(prompt_output),
            ]
        )
    if request.sheet:
        args.extend(["--sheet", request.sheet])
    if request.source_only:
        args.append("--source-only")
    if request.include_empty_final_terms:
        args.append("--include-empty-final-terms")
    for announcement_artifact_id in announcement_material_artifact_ids:
        announcement_artifact = db.get_artifact(announcement_artifact_id)
        args.extend(["--announcement-material", announcement_artifact["path"]])
    if announcement_output is not None:
        args.extend(["--announcement-output", str(announcement_output), "--announcement-min-hit", str(announcement_min_hit)])
    if ai_supplement:
        if not announcement_material_artifact_ids:
            raise ValueError("ai_supplement requires announcement_material_artifact_ids")
        args.append("--ai-supplement")
        if ai_supplement_packet_output is not None:
            args.extend(["--ai-supplement-packet-output", str(ai_supplement_packet_output)])
        if ai_supplement_report_output is not None:
            args.extend(["--ai-supplement-report-output", str(ai_supplement_report_output)])
        response_artifact_id = str(getattr(request, "ai_supplement_response_artifact_id", "") or "").strip()
        if response_artifact_id:
            response_artifact = db.get_artifact(response_artifact_id)
            if response_artifact["project_id"] != project_id:
                raise KeyError(response_artifact_id)
            args.extend(["--ai-supplement-provider", "file", "--ai-supplement-response", response_artifact["path"]])
        else:
            # The workbench owns provider calls via backend/app/providers.py.
            # Do not let the embedded CLI auto-read OPENAI_API_KEY and bypass
            # the configured GPT/Claude/Codex-relay settings.
            args.extend(["--ai-supplement-provider", "packet"])
    if not announcement_only:
        for material_artifact_id in material_artifact_ids:
            material_artifact = db.get_artifact(material_artifact_id)
            args.extend(["--project-material", material_artifact["path"]])
        for note in [*project_notes, *material_notes]:
            args.extend(["--project-note", note])
    try:
        proc = run_subprocess(args, GLOSSARY_ROOT, run["id"])
        parsed = parse_key_output(proc.stdout)
        if language == "en":
            for generated_path in (detail_output, final_output, announcement_output):
                if generated_path is not None:
                    _normalize_new_english_glossary_artifact(generated_path)
            for generated_path in (brief_output, prompt_output):
                _normalize_new_english_glossary_text_artifact(generated_path)
        artifacts = []
        backfill: dict[str, Any] = {}
        if not announcement_only:
            artifacts.extend(
                [
                    db.add_artifact(project_id, "Glossary details", detail_output, "glossary_detail", run_id=run["id"]),
                    db.add_artifact(project_id, f"ID CN {final_suffix} glossary", final_output, "glossary_final", run_id=run["id"]),
                    db.add_artifact(project_id, "Project brief", brief_output, "project_brief", run_id=run["id"], mime="text/markdown"),
                    db.add_artifact(project_id, "Translation prompt", prompt_output, "translation_prompt", run_id=run["id"], mime="text/plain"),
                ]
            )
            backfill = backfill_project_glossary_from_final(project_id, final_output, run["id"], language=language)
            if bool(getattr(request, "ai_candidate_supplement", True)):
                backfill["ai_supplement"] = supplement_language_table_glossary_candidates_with_ai(
                    project_id=project_id,
                    batch_id=str(backfill.get("batch_id") or ""),
                    input_path=input_path,
                    language=language,
                    run_id=run["id"],
                    settings=settings,
                )
        if announcement_output is not None and announcement_output.exists():
            artifacts.append(
                db.add_artifact(
                    project_id,
                    "Announcement glossary lookup",
                    announcement_output,
                    "announcement_glossary",
                    run_id=run["id"],
                )
            )
        if ai_supplement_packet_output is not None and ai_supplement_packet_output.exists():
            artifacts.append(db.add_artifact(project_id, "公告 AI 补充包", ai_supplement_packet_output, "announcement_ai_supplement_packet", run_id=run["id"], mime="application/json"))
        if ai_supplement_report_output is not None and ai_supplement_report_output.exists():
            artifacts.append(db.add_artifact(project_id, "公告 AI 补充报告", ai_supplement_report_output, "announcement_ai_supplement_report", run_id=run["id"], mime="text/markdown"))
        if (
            not announcement_only
            and prompt_output.exists()
            and bool(getattr(request, "update_project_prompt", True))
        ):
            prompt = prompt_output.read_text(encoding="utf-8")
            db.update_project(project_id, {"prompt_text": prompt})
        db.update_run(
            run["id"],
            status="passed",
            metadata={
                "output": parsed,
                "glossary_backfill": backfill,
                "announcement": {
                    "material_artifact_ids": announcement_material_artifact_ids,
                    "only": announcement_only,
                    "output": str(announcement_output) if announcement_output else "",
                    "terms": int(parsed.get("ANNOUNCEMENT_TERMS") or 0),
                    "ai_supplement_packet": parsed.get("AI_SUPPLEMENT_PACKET_OUTPUT") or "disabled",
                    "ai_supplement_report": parsed.get("AI_SUPPLEMENT_REPORT_OUTPUT") or "disabled",
                },
            },
        )
        return {"run": db.get_run(run["id"]), "artifacts": artifacts, "output": parsed, "glossary_backfill": backfill}
    except Exception as exc:
        friendly = user_facing_error(exc)
        db.add_event(run["id"], friendly, level="error")
        db.update_run(run["id"], status="failed", metadata={"error": friendly})
        raise


async def translate_missing_glossary_candidates(project_id: str, batch_id: str) -> dict[str, Any]:
    project = db.get_project(project_id)
    batch = db.get_glossary_batch(batch_id)
    if batch["project_id"] != project_id:
        raise KeyError(batch_id)
    language = require_supported_language(batch.get("language") or "en")
    settings = load_settings()
    provider = normalize_provider_name(settings.get("provider"))
    if provider in REAL_PROVIDERS and not settings.get("api_key"):
        raise ValueError(f"{provider} api_key is required to translate glossary candidates")

    pending = db.list_glossary_candidates(project_id, batch_id=batch_id, status="pending", language=language)
    missing = [candidate for candidate in pending if not str(candidate.get("target") or "").strip()]
    if not missing:
        return {
            "batch": batch,
            "translated_count": 0,
            "skipped_count": len(pending),
            "candidates": db.list_glossary_candidates(project_id, batch_id=batch_id, language=language),
        }

    rows: list[dict[str, Any]] = []
    id_to_candidate: dict[int, dict[str, Any]] = {}
    for index, candidate in enumerate(missing, start=1):
        rows.append(
            {
                "id": index,
                "source": str(candidate.get("source") or ""),
                "text_type": "glossary_term_candidate",
                "term_key": str(candidate.get("term_key") or ""),
                "note": str(candidate.get("note") or ""),
            }
        )
        id_to_candidate[index] = candidate

    prompt = _glossary_candidate_translation_prompt(project, rows, language=language, settings=settings)
    try:
        items = await translate_batch(rows, settings, prompt)
    except Exception as exc:
        raise ValueError(f"glossary candidate translation failed: {user_facing_error(exc)}") from exc
    translated_count = 0
    for item in items:
        candidate = id_to_candidate.get(int(item.id))
        if candidate is None:
            continue
        translation = str(item.translation or "").strip()
        if not translation:
            continue
        metadata = dict(candidate.get("metadata") or {})
        metadata["model"] = {
            "provider": provider,
            "model": settings.get("model") or "",
            "preset": settings.get("preset") or "",
            "translated_at": db.now_iso(),
        }
        db.update_glossary_candidate(
            candidate["id"],
            {
                "target": translation,
                "translation_status": "suggested",
                "translation_source": "model",
                "metadata": metadata,
            },
        )
        translated_count += 1

    if batch.get("run_id"):
        db.add_event(batch["run_id"], f"Glossary candidate translation filled {translated_count} missing {language.upper()} values.")
    return {
        "batch": db.get_glossary_batch(batch_id),
        "translated_count": translated_count,
        "skipped_count": len(pending) - len(missing),
        "candidates": db.list_glossary_candidates(project_id, batch_id=batch_id, language=language),
    }


def translate_missing_glossary_candidates_sync(project_id: str, batch_id: str) -> dict[str, Any]:
    return asyncio.run(translate_missing_glossary_candidates(project_id, batch_id))


def _glossary_candidate_translation_prompt(
    project: dict[str, Any], rows: list[dict[str, Any]], language: str = "en", settings: dict[str, Any] | None = None
) -> str:
    language = require_supported_language(language)
    spec = language_spec(language)
    profile = project.get("profile") or {}
    prompt_text = str((profile.get("prompts_by_language") or {}).get(language) or project.get("prompt_text") or "").strip()
    prompt_text = _manage_project_prompt_context(prompt_text, settings if settings is not None else load_settings())
    profile_summary = {
        key: profile.get(key)
        for key in ("game_type", "target_audience", "content_scope", "translation_style", "tone", "language_assets")
        if profile.get(key)
    }
    existing_terms = [
        {"source": term.get("source"), "target": term.get("target"), "target_alt": term.get("target_alt")}
        for term in db.list_glossary_terms(project["id"], language=language)[:200]
        if str(term.get("source") or "").strip() and str(term.get("target") or "").strip()
    ]
    term_instruction = (
        f"Translate only the {spec.target_header} term. Do not create {spec.alt_header}, notes, categories, explanations, or markdown. "
        if spec.alt_header
        else f"Translate only the {spec.target_header} term. Do not create notes, categories, explanations, or markdown. "
    )
    return (
        f"Translate short game glossary term candidates from Chinese to {spec.prompt_name}. "
        "Return JSONL only; each line must be {\"id\": number, \"translation\": string}. "
        f"{term_instruction}"
        "Keep UI terms concise and consistent with existing glossary.\n\n"
        f"Project: {project.get('name', '')}\n"
        f"Profile: {json.dumps(profile_summary, ensure_ascii=False)}\n"
        f"Project prompt:\n{prompt_text}\n\n"
        f"Existing glossary examples:\n{json.dumps(existing_terms, ensure_ascii=False)}\n\n"
        f"Candidates:\n" + "\n".join(json.dumps(row, ensure_ascii=False) for row in rows)
    )


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return _jsonl_helpers.read_jsonl(path)


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    _jsonl_helpers.write_jsonl(path, rows)



__all__ = [name for name in globals() if not name.startswith("__")]
