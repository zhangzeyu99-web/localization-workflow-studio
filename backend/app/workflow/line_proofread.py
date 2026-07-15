"""Product-side AI line-by-line proofreading (深度逐行校对).

Boundary: opt-in only, via ``TranslateRequest.enable_line_proofread``. The
model only *suggests* fixes; a deterministic audit gate rejects unsafe
suggestions (protected-token loss, term drift, number loss, no-op edits)
before anything is written back, and the caller must re-run machine QA on the
proofread workbook afterwards. Mirrors the agent-side deep proofreading rule:
subagents propose, the deterministic pipeline decides.
"""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .. import db
from ..config import TEST_FAKE_PROVIDER, normalize_provider_name
from ..languages import SOURCE_HEADER_ALIASES, require_supported_language, target_aliases
from ..providers import call_text
from ..translation_batches import manage_project_prompt_context
from .jsonl_helpers import write_jsonl
from .large_text import TOKEN_RE, is_auto_protected_token
from .qa import (
    WORKBOOK_ID_HEADER_ALIASES,
    _apply_workbook_fixes,
    _cell_text,
    _first_col,
    _header_map,
)
from .subprocess_runner import UserFacingWorkflowError, user_facing_error

LINE_PROOFREAD_BATCH_SIZE = 50
LINE_PROOFREAD_RULE_SOURCE = "line_proofread"
_NUMBER_TOKEN_RE = re.compile(r"\d+(?:[.,]\d+)?")


def _workbook_review_rows(path: Path, language: str) -> list[dict[str, Any]]:
    language = require_supported_language(language)
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            headers = _header_map(ws)
            source_col = _first_col(headers, list(SOURCE_HEADER_ALIASES))
            target_col = _first_col(headers, target_aliases(language))
            id_col = _first_col(headers, WORKBOOK_ID_HEADER_ALIASES)
            if source_col is None or target_col is None:
                continue
            for row_index in range(2, ws.max_row + 1):
                source = _cell_text(ws.cell(row_index, source_col).value)
                target = _cell_text(ws.cell(row_index, target_col).value)
                if not source or not target:
                    continue
                rows.append(
                    {
                        "sheet": ws.title,
                        "row": row_index,
                        "record_id": _cell_text(ws.cell(row_index, id_col).value) if id_col else "",
                        "source": source,
                        "current_target": target,
                    }
                )
    finally:
        wb.close()
    return rows


def _term_hits_by_record_id(workpack_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    mapping: dict[str, list[dict[str, Any]]] = {}
    for row in workpack_rows:
        record_id = str(row.get("id") or "").strip()
        hits = row.get("term_hits") or []
        if record_id and isinstance(hits, list) and hits:
            mapping[record_id] = [hit for hit in hits if isinstance(hit, dict)]
    return mapping


def _line_proofread_prompt(project: dict[str, Any], language: str, batch: list[dict[str, Any]], settings: dict[str, Any]) -> str:
    prompt = manage_project_prompt_context(str(project.get("prompt_text") or ""), settings)
    payload = [
        {
            "record_id": row["record_id"],
            "sheet": row["sheet"],
            "row": row["row"],
            "source": row["source"],
            "current_target": row["current_target"],
            "term_hits": [
                {"source": hit.get("source", ""), "target": hit.get("target", "")}
                for hit in row.get("term_hits") or []
            ],
        }
        for row in batch
    ]
    return (
        "你是游戏本地化逐行审校模型。逐行对照原文和当前译文，只报告确有问题的行："
        "漏译、误译、术语不一致、语境不自然、数字/占位符错误。"
        "没有问题的行不要出现在结果里。修改必须保留原文中的变量、占位符、HTML/BBCode 标签、转义换行和数字。"
        "term_hits 是术语约束：命中术语的译法必须沿用 hit.target，不要改名。\n\n"
        "返回严格 JSON：{\"suggestions\":[{\"record_id\":\"...\",\"sheet\":\"...\",\"row\":2,"
        "\"severity\":\"hard|soft\",\"suggested_target\":\"...\",\"reason\":\"...\"}]}。"
        "必须沿用输入行的 record_id/sheet/row。\n"
        f"项目：{project.get('name', '')}\n"
        f"目标语言：{language}\n"
        f"项目提示词：\n{prompt}\n\n"
        f"待审校行：\n{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def _call_line_proofread_provider(settings: dict[str, Any], prompt: str) -> str:
    provider = normalize_provider_name(settings.get("provider"))
    if provider == TEST_FAKE_PROVIDER:
        # Deterministic no-op reviewer so test/dev flows can exercise the full
        # pipeline; tests monkeypatch this function to inject suggestions.
        return json.dumps({"suggestions": []})
    return call_text(settings, prompt, system="Return strict JSON only.")


def _parse_suggestions(text: str) -> list[dict[str, Any]]:
    from .semantic_qa import _parse_semantic_qa_payload

    payload = _parse_semantic_qa_payload(text)
    raw = payload.get("suggestions")
    if not isinstance(raw, list):
        return []
    return [item for item in raw if isinstance(item, dict)]


def _protected_source_tokens(source: str) -> list[str]:
    return sorted({token for token in TOKEN_RE.findall(source) if is_auto_protected_token(token)}, key=len, reverse=True)


def audit_suggestion(suggestion: dict[str, Any], row: dict[str, Any]) -> str | None:
    """Deterministic audit gate. Returns a rejection reason, or None to accept."""
    suggested = str(suggestion.get("suggested_target") or "").strip()
    if not suggested:
        return "empty_suggestion"
    current = str(row.get("current_target") or "")
    if suggested == current.strip():
        return "no_change"
    source = str(row.get("source") or "")
    for token in _protected_source_tokens(source):
        if source.count(token) > suggested.count(token):
            return "protected_token_lost"
    for hit in row.get("term_hits") or []:
        term_target = str(hit.get("target") or "").strip()
        if term_target and term_target in current and term_target not in suggested:
            return "term_drift"
    for number in _NUMBER_TOKEN_RE.findall(source):
        if number in current and number not in suggested:
            return "number_lost"
    return None


def run_line_proofread(
    *,
    run_id: str,
    project: dict[str, Any],
    language: str,
    qa_workbook_path: Path,
    workpack_rows: list[dict[str, Any]],
    settings: dict[str, Any],
    output_dir: Path,
    cancel_event: Any = None,
) -> dict[str, Any]:
    """Review the QA workbook line by line and apply audited fixes to a copy.

    Returns a state dict; when ``applied > 0`` the caller must re-run machine
    QA on ``state["workbook_path"]`` before trusting the result.
    """
    language = require_supported_language(language)
    output_dir.mkdir(parents=True, exist_ok=True)
    review_rows = _workbook_review_rows(qa_workbook_path, language)
    hits_by_id = _term_hits_by_record_id(workpack_rows)
    for row in review_rows:
        row["term_hits"] = hits_by_id.get(str(row.get("record_id") or "").strip(), [])

    rows_by_position = {(row["sheet"], row["row"]): row for row in review_rows}
    rows_by_record_id = {str(row["record_id"]): row for row in review_rows if str(row.get("record_id") or "").strip()}

    suggestions: list[dict[str, Any]] = []
    batches = [review_rows[i : i + LINE_PROOFREAD_BATCH_SIZE] for i in range(0, len(review_rows), LINE_PROOFREAD_BATCH_SIZE)]
    for index, batch in enumerate(batches, start=1):
        if cancel_event is not None and cancel_event.is_set():
            raise RuntimeError("translation canceled")
        db.add_event(run_id, f"line proofread: reviewing batch {index}/{len(batches)} ({len(batch)} rows)")
        prompt = _line_proofread_prompt(project, language, batch, settings)
        try:
            text = _call_line_proofread_provider(settings, prompt)
        except Exception as exc:
            raise UserFacingWorkflowError(f"逐行校对模型调用失败：{user_facing_error(exc)}") from exc
        suggestions.extend(_parse_suggestions(text))

    accepted_fixes: list[dict[str, Any]] = []
    audit_records: list[dict[str, Any]] = []
    seen_keys: set[tuple[str, int]] = set()
    for suggestion in suggestions:
        sheet = str(suggestion.get("sheet") or "")
        row_index = int(suggestion.get("row") or 0)
        row = rows_by_position.get((sheet, row_index)) or rows_by_record_id.get(str(suggestion.get("record_id") or "").strip())
        record = {
            "record_id": str(suggestion.get("record_id") or ""),
            "sheet": sheet,
            "row": row_index,
            "severity": str(suggestion.get("severity") or "soft"),
            "suggested_target": str(suggestion.get("suggested_target") or ""),
            "reason": str(suggestion.get("reason") or ""),
        }
        if row is None:
            record["audit"] = "rejected:unknown_row"
            audit_records.append(record)
            continue
        key = (str(row["sheet"]), int(row["row"]))
        if key in seen_keys:
            record["audit"] = "rejected:duplicate_row"
            audit_records.append(record)
            continue
        rejection = audit_suggestion(suggestion, row)
        if rejection:
            record["audit"] = f"rejected:{rejection}"
            audit_records.append(record)
            continue
        seen_keys.add(key)
        record["audit"] = "accepted"
        record["previous_target"] = row["current_target"]
        audit_records.append(record)
        accepted_fixes.append(
            {
                "issue_id": "",
                "sheet": row["sheet"],
                "row": row["row"],
                "record_id": row["record_id"],
                "source_text": row["source"],
                "translation": str(suggestion.get("suggested_target") or "").strip(),
                "note": record["reason"] or "line_proofread",
                "rule_source": LINE_PROOFREAD_RULE_SOURCE,
            }
        )

    suggestions_path = output_dir / "line_proofread_suggestions.jsonl"
    write_jsonl(suggestions_path, audit_records)
    suggestions_artifact = db.add_artifact(
        project["id"],
        "Line proofread suggestions",
        suggestions_path,
        "line_proofread_suggestions",
        run_id=run_id,
        mime="application/jsonl",
        origin="generated",
    )

    applied_fixes: list[dict[str, Any]] = []
    proofread_workbook: Path | None = None
    if accepted_fixes:
        proofread_workbook = output_dir / f"{qa_workbook_path.stem}_line_proofread.xlsx"
        shutil.copy2(qa_workbook_path, proofread_workbook)
        applied_fixes = _apply_workbook_fixes(proofread_workbook, accepted_fixes, run_id, language=language)

    rejected = sum(1 for record in audit_records if str(record.get("audit", "")).startswith("rejected:"))
    state = {
        "status": "model_reviewed",
        "provider": normalize_provider_name(settings.get("provider")),
        "model": str(settings.get("model") or ""),
        "reviewed_rows": len(review_rows),
        "batches": len(batches),
        "suggested": len(suggestions),
        "rejected_by_audit": rejected,
        "applied": len(applied_fixes),
        "suggestions_artifact_id": suggestions_artifact["id"],
        "workbook_path": str(proofread_workbook) if proofread_workbook else "",
        "fixes": applied_fixes,
    }
    db.add_event(
        run_id,
        "line proofread finished: "
        f"reviewed={state['reviewed_rows']}, suggested={state['suggested']}, rejected={rejected}, applied={state['applied']}",
    )
    return state
