from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

from .. import db
from ..delivery_naming import safe_delivery_name
from ..download_urls import artifact_download_url
from ..languages import PROJECT_LANGUAGE_ORDER, SOURCE_HEADER_ALIASES, require_supported_language, target_aliases
from .announcement_outputs import _announcement_task_source_stem, _artifact_display_label, _visible_language_code
from .announcement_segments import _normalize_announcement_languages
from .asset_import_export import archive_translation_artifact
from .common import project_dir
from .qa import _first_col, _row_cell, write_qa_changes_report
from .subprocess_runner import user_facing_error

DELIVERED_WITH_ISSUES_SOURCE_TYPE = "delivered_with_issues"


def list_project_deliverables(project_id: str) -> list[dict[str, Any]]:
    project = db.get_project(project_id)
    deliverables: list[dict[str, Any]] = []
    for run in db.list_runs(project_id):
        if run["kind"] not in {"translation", "qa"} or run["status"] in {"created", "running", "queued", "canceled"}:
            continue
        final_artifact = _deliverable_final_artifact(run)
        if not final_artifact or not Path(final_artifact["path"]).exists():
            continue
        summary = _deliverable_summary(project, run, final_artifact)
        if final_artifact["kind"] != "final_text" and int(summary.get("translated_rows") or 0) <= 0:
            continue
        deliverables.append(summary)
    deliverables.extend(_merged_deliverable_summaries(project))
    deliverables.extend(_announcement_deliverable_summaries(project))
    return deliverables


def _merged_deliverable_summaries(project: dict[str, Any]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for artifact in db.list_artifacts(project_id=project["id"]):
        if artifact.get("kind") != "merged_delivery_workbook":
            continue
        path = Path(str(artifact.get("path") or ""))
        if not path.exists():
            continue
        metadata = artifact.get("metadata") or {}
        summary_artifact = None
        summary_id = str(metadata.get("summary_artifact_id") or "")
        if summary_id:
            try:
                candidate = db.get_artifact(summary_id)
                if Path(candidate["path"]).exists():
                    summary_artifact = candidate
            except KeyError:
                summary_artifact = None
        languages = metadata.get("merged_languages") if isinstance(metadata.get("merged_languages"), list) else []
        skipped = metadata.get("skipped_languages") if isinstance(metadata.get("skipped_languages"), list) else []
        summaries.append(
            {
                "run_id": artifact["id"],
                "task_code": "ALL",
                "task_id": _short_run_id(artifact["id"]),
                "task_label": f"ALL-{_short_run_id(artifact['id'])}",
                "task_type": "多语言合并交付",
                "language": " / ".join(languages) or "ALL",
                "created_at": artifact.get("created_at", ""),
                "updated_at": artifact.get("created_at", ""),
                "status": "delivered",
                "processed_rows": int(metadata.get("processed_rows") or 0),
                "source_rows": int(metadata.get("source_rows") or 0),
                "translated_rows": int(metadata.get("translated_rows") or 0),
                "provider": "-",
                "model": "-",
                "input_label": str(metadata.get("input_label") or "多语言合并交付"),
                "qa_status": "mixed" if skipped else "passed",
                "qa_hard_errors": int(metadata.get("qa_hard_errors") or 0),
                "qa_soft_warnings": 0,
                "delivered_with_issues": int(metadata.get("qa_hard_errors") or 0) > 0,
                "files": {
                    "final": _artifact_delivery_file("merged_final", artifact),
                    "qa_summary": _artifact_delivery_file("qa_summary", summary_artifact) if summary_artifact else None,
                    "outputs": [],
                },
                "source_artifacts": {
                    "merged_delivery_workbook": artifact["id"],
                    "merged_delivery_summary": summary_id,
                },
            }
        )
    return summaries


def _announcement_deliverable_summaries(project: dict[str, Any]) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for task in db.list_announcement_tasks(project["id"]):
        if task.get("status") != "delivered":
            continue
        metadata = task.get("metadata") or {}
        delivery_artifact_id = str(metadata.get("delivery_artifact_id") or "")
        if not delivery_artifact_id:
            continue
        try:
            package_artifact = db.get_artifact(delivery_artifact_id)
        except KeyError:
            continue
        package_metadata = package_artifact.get("metadata") or {}
        if package_metadata.get("superseded"):
            continue
        if not Path(package_artifact["path"]).exists():
            continue
        languages = _normalize_announcement_languages(task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
        output_files = []
        output_artifact_ids = metadata.get("output_artifact_ids") if isinstance(metadata.get("output_artifact_ids"), dict) else {}
        for language in languages:
            artifact_id = str(output_artifact_ids.get(language) or "")
            if not artifact_id:
                continue
            try:
                artifact = db.get_artifact(artifact_id)
            except KeyError:
                continue
            if Path(artifact["path"]).exists():
                output_files.append(_artifact_delivery_file(f"output_{_visible_language_code(language)}", artifact))
        qa_file = None
        qa_artifact_id = str(metadata.get("qa_summary_artifact_id") or "")
        if qa_artifact_id:
            try:
                qa_artifact = db.get_artifact(qa_artifact_id)
                if Path(qa_artifact["path"]).exists():
                    qa_file = _artifact_delivery_file("qa_summary", qa_artifact)
            except KeyError:
                qa_file = None
        package_file = _artifact_delivery_file("package", package_artifact)
        language_label = " / ".join(_visible_language_code(language) for language in languages) or "-"
        source_rows = int(metadata.get("segment_count") or metadata.get("terms_count") or metadata.get("term_count") or 0)
        return_label = task.get("title") or _announcement_task_source_stem(task)
        summaries.append(
            {
                "run_id": package_artifact.get("run_id") or task["id"],
                "task_code": "ANN",
                "task_id": _short_run_id(task["id"]),
                "task_label": f"ANN-{_short_run_id(task['id'])}",
                "task_type": "公告任务",
                "language": language_label,
                "created_at": task.get("created_at", ""),
                "updated_at": task.get("updated_at", ""),
                "status": "delivered",
                "processed_rows": source_rows,
                "source_rows": source_rows,
                "translated_rows": source_rows,
                "provider": "-",
                "model": "-",
                "input_label": return_label,
                "qa_status": "passed",
                "qa_hard_errors": int(metadata.get("hard_blockers") or 0),
                "qa_soft_warnings": 0,
                "delivered_with_issues": bool(package_metadata.get("forced") or metadata.get("forced"))
                or int(package_metadata.get("hard_blockers") or metadata.get("hard_blockers") or 0) > 0,
                "files": {
                    "package": package_file,
                    "qa_summary": qa_file,
                    "outputs": output_files,
                },
                "source_artifacts": {
                    "announcement_delivery_package": package_artifact["id"],
                    "announcement_outputs": [item.get("artifact_id") for item in output_files if item.get("artifact_id")],
                    "announcement_qa_summary": qa_artifact_id,
                },
            }
        )
    return summaries


def build_delivery_package(project_id: str, run_id: str | None = None) -> dict[str, Any]:
    project = db.get_project(project_id)
    deliverables = list_project_deliverables(project_id)
    if not deliverables:
        raise ValueError("QA 未通过，暂无最终交付 workbook")
    selected = deliverables[0]
    if run_id:
        selected = next((item for item in deliverables if item["run_id"] == run_id), None)
        if not selected:
            raise ValueError("指定任务未通过 QA，暂无最终交付")
    run = db.get_run(selected["run_id"])
    final_source = _deliverable_final_artifact(run)
    if not final_source or not Path(final_source["path"]).exists():
        raise ValueError("暂无最终交付文件")
    if final_source["kind"] != "final_text" and _workbook_processed_rows(Path(final_source["path"]), run.get("language") or "en")["translated_rows"] <= 0:
        raise ValueError("最终译文为空，不能生成交付。请先完成翻译或 QA。")
    changes_source = _run_artifact(run["id"], "qa_changes")

    output_dir = project_dir(project_id) / "delivery"
    output_dir.mkdir(parents=True, exist_ok=True)
    if final_source["kind"] == "final_text":
        final_path = _delivery_final_output_path(project, run, final_source)
        shutil.copy2(final_source["path"], final_path)
        summary = _deliverable_summary(project, run, final_source)
        summary["files"] = {"final": _delivery_file("final", final_path)}
        return {"project_id": project_id, "project_name": project["name"], "deliverable": summary, "files": list(summary["files"].values()), "archive": None}

    final_path, changes_path = _delivery_output_paths(project, run)
    shutil.copy2(final_source["path"], final_path)
    _normalize_delivery_workbook_headers(final_path, run.get("language") or "en")
    if changes_source and Path(changes_source["path"]).exists():
        shutil.copy2(changes_source["path"], changes_path)
    else:
        empty_changes = write_qa_changes_report(output_dir, [])
        empty_changes.replace(changes_path)

    summary = _deliverable_summary(project, run, final_source)
    summary["files"] = {
        "final": _delivery_file("final", final_path),
        "changes": _delivery_file("changes", changes_path),
    }
    archive_result = _archive_delivery_translation(project_id, run, final_source)
    return {"project_id": project_id, "project_name": project["name"], "deliverable": summary, "files": list(summary["files"].values()), "archive": archive_result}


def build_merged_delivery_package(project_id: str, input_artifact_id: str, languages: list[str]) -> dict[str, Any]:
    project = db.get_project(project_id)
    source_artifact = db.get_artifact(input_artifact_id)
    if source_artifact.get("project_id") != project_id:
        raise ValueError("输入文件不属于当前项目")
    source_path = Path(str(source_artifact.get("path") or ""))
    if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("多语言合并交付当前只支持 XLSX 语言表")
    if not source_path.exists():
        raise ValueError("原始语言表文件不可读，无法生成合并交付")
    selected_languages = _normalize_delivery_languages(languages)
    output_dir = project_dir(project_id) / "delivery"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y%m%d%H%M")
    output_path = output_dir / f"{_safe_delivery_name(project['name'])}_ALL_{timestamp}_final.xlsx"
    shutil.copy2(source_path, output_path)

    merged: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    for language in selected_languages:
        run = _find_merge_source_run(project_id, input_artifact_id, language)
        if not run:
            skipped.append({"language": _visible_language_code(language), "reason": "未找到可交付的翻译/QA 结果"})
            continue
        final_artifact = _deliverable_final_artifact(run)
        if not final_artifact or not Path(final_artifact["path"]).exists():
            skipped.append({"language": _visible_language_code(language), "run_id": run["id"], "reason": "缺少最终译文文件"})
            continue
        try:
            copied = _merge_language_column(output_path, Path(final_artifact["path"]), language)
            if copied <= 0:
                skipped.append({"language": _visible_language_code(language), "run_id": run["id"], "reason": "没有可合并的译文列"})
                continue
            quality = run.get("metadata", {}).get("quality_summary") or {}
            merged.append(
                {
                    "language": _visible_language_code(language),
                    "run_id": run["id"],
                    "rows": copied,
                    "status": "passed" if run.get("status") == "passed" else "deliverable_with_issues",
                    "hard_errors": int(quality.get("hard_errors") or 0),
                }
            )
        except Exception as exc:
            skipped.append({"language": _visible_language_code(language), "run_id": run["id"], "reason": user_facing_error(exc)})
    if not merged:
        raise ValueError("没有可合并的已完成语言，请先完成翻译或 QA")

    summary_path = _write_merged_delivery_summary(output_dir, merged, skipped, output_path)
    summary_artifact = db.add_artifact(
        project_id,
        f"{project['name']} ALL QA摘要",
        summary_path,
        "merged_delivery_summary",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        metadata={"merged_languages": [item["language"] for item in merged], "skipped_languages": [item["language"] for item in skipped]},
    )
    final_artifact = db.add_artifact(
        project_id,
        f"{project['name']} ALL 合并交付",
        output_path,
        "merged_delivery_workbook",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        metadata={
            "input_artifact_id": input_artifact_id,
            "input_label": _artifact_display_label(source_artifact),
            "summary_artifact_id": summary_artifact["id"],
            "merged_languages": [item["language"] for item in merged],
            "skipped_languages": [item["language"] for item in skipped],
            "source_rows": _workbook_processed_rows(output_path)["source_rows"],
            "translated_rows": sum(int(item.get("rows") or 0) for item in merged),
            "processed_rows": sum(int(item.get("rows") or 0) for item in merged),
            "qa_hard_errors": sum(int(item.get("hard_errors") or 0) for item in merged),
        },
    )
    files = [_artifact_delivery_file("merged_final", final_artifact), _artifact_delivery_file("qa_summary", summary_artifact)]
    return {
        "project_id": project_id,
        "project_name": project["name"],
        "merged_languages": [item["language"] for item in merged],
        "skipped_languages": [item["language"] for item in skipped],
        "files": files,
        "deliverable": _merged_deliverable_summaries(project)[0] if _merged_deliverable_summaries(project) else {},
    }


def _deliverable_summary(project: dict[str, Any], run: dict[str, Any], final_artifact: dict[str, Any]) -> dict[str, Any]:
    changes_artifact = _run_artifact(run["id"], "qa_changes")
    final_path, changes_path = _delivery_output_paths(project, run)
    if final_artifact["kind"] == "final_text":
        final_path = _delivery_final_output_path(project, run, final_artifact)
    task_code, task_run_id = _effective_task_identity(run)
    metadata = run.get("metadata", {})
    quality_summary = metadata.get("quality_summary") or {}
    # Same decision as _archive_delivery_translation's source_type.
    qa_passed = bool(quality_summary.get("passed", run.get("status") == "passed"))
    provider, model = _deliverable_provider_model(metadata, quality_summary)
    input_label = _input_artifact_label(run, run["project_id"])
    processed = (
        {"processed_rows": int(metadata.get("translated_rows") or metadata.get("source_rows") or 0), "source_rows": int(metadata.get("source_rows") or 0), "translated_rows": int(metadata.get("translated_rows") or 0)}
        if final_artifact["kind"] == "final_text"
        else _workbook_processed_rows(Path(final_artifact["path"]), run.get("language") or "en")
    )
    files = {"final": _delivery_file("final", final_path) if final_path.exists() else _expected_delivery_file("final", final_path)}
    if final_artifact["kind"] != "final_text":
        files["changes"] = _delivery_file("changes", changes_path) if changes_path.exists() else _expected_delivery_file("changes", changes_path)
    return {
        "run_id": run["id"],
        "task_code": task_code,
        "task_id": _short_run_id(task_run_id),
        "task_label": f"{task_code}-{_short_run_id(task_run_id)}",
        "task_type": _task_type_label(task_code),
        "language": _visible_language_code(run.get("language") or "en"),
        "created_at": run.get("created_at", ""),
        "updated_at": run.get("updated_at", ""),
        "status": run.get("status", ""),
        "processed_rows": processed["processed_rows"] or int(metadata.get("translated_rows") or 0),
        "source_rows": processed["source_rows"],
        "translated_rows": processed["translated_rows"],
        "provider": provider,
        "model": model,
        "input_label": input_label,
        "qa_status": "passed" if qa_passed else "failed",
        "qa_hard_errors": int(quality_summary.get("hard_errors") or 0),
        "qa_soft_warnings": _soft_warning_count(quality_summary),
        "delivered_with_issues": not qa_passed,
        "files": files,
        "source_artifacts": {
            "qa_final_workbook": final_artifact["id"] if final_artifact["kind"] == "qa_final_workbook" else "",
            "final_text": final_artifact["id"] if final_artifact["kind"] == "final_text" else "",
            "qa_changes": changes_artifact["id"] if changes_artifact else "",
        },
    }


def _deliverable_final_artifact(run: dict[str, Any]) -> dict[str, Any] | None:
    return _run_artifact(run["id"], "qa_final_workbook") or _run_artifact(run["id"], "final_text")


def _archive_delivery_translation(project_id: str, run: dict[str, Any], final_artifact: dict[str, Any]) -> dict[str, Any]:
    quality_summary = run.get("metadata", {}).get("quality_summary") or {}
    source_type = "qa_passed" if bool(quality_summary.get("passed", run.get("status") == "passed")) else DELIVERED_WITH_ISSUES_SOURCE_TYPE
    archive_result = archive_translation_artifact(
        project_id,
        final_artifact["id"],
        language=run.get("language") or "en",
        source_type=source_type,
    )
    db.add_event(run["id"], f"delivery translation archive updated: source_type={source_type} rows={archive_result['imported_count']}")
    return {**archive_result, "source_type": source_type}


def _effective_task_identity(run: dict[str, Any], seen: set[str] | None = None) -> tuple[str, str]:
    seen = seen or set()
    if run["id"] in seen:
        return _fallback_task_code(run), run["id"]
    seen.add(run["id"])
    metadata = run.get("metadata", {})
    source_run_id = metadata.get("manual_fix_source_run_id") or metadata.get("model_fix_source_run_id") or metadata.get("source_run_id")
    if source_run_id:
        try:
            source_run = db.get_run(str(source_run_id))
            if source_run["project_id"] == run["project_id"]:
                source_code, source_id = _effective_task_identity(source_run, seen)
                if run["kind"] == "qa" and source_run["kind"] in {"translation", "qa"}:
                    return source_code, source_id
        except KeyError:
            pass
    task_code = str(metadata.get("task_code") or "").upper()
    if task_code not in {"A", "T", "QA"}:
        task_code = _fallback_task_code(run)
    return task_code, run["id"]


def _fallback_task_code(run: dict[str, Any]) -> str:
    if run["kind"] == "translation":
        return "T"
    if run["kind"] == "qa":
        return "QA"
    return str(run["kind"] or "TASK").upper()


def _task_type_label(task_code: str) -> str:
    return {"A": "完整工作流", "T": "翻译任务", "QA": "校对任务"}.get(task_code, task_code)


def _short_run_id(run_id: str) -> str:
    return str(run_id).removeprefix("run_")[:6]


def _delivery_output_paths(project: dict[str, Any], run: dict[str, Any]) -> tuple[Path, Path]:
    output_dir = project_dir(project["id"]) / "delivery"
    output_dir.mkdir(parents=True, exist_ok=True)
    task_code, task_run_id = _effective_task_identity(run)
    timestamp = _delivery_timestamp(run.get("created_at", ""))
    language = _visible_language_code(run.get("language") or "en")
    prefix = f"{_safe_delivery_name(project['name'])}_{language}_{timestamp}_{task_code}-{_short_run_id(task_run_id)}"
    return output_dir / f"{prefix}_final.xlsx", output_dir / f"{prefix}_changes.xlsx"


def _delivery_final_output_path(project: dict[str, Any], run: dict[str, Any], source_artifact: dict[str, Any]) -> Path:
    output_dir = project_dir(project["id"]) / "delivery"
    output_dir.mkdir(parents=True, exist_ok=True)
    task_code, task_run_id = _effective_task_identity(run)
    timestamp = _delivery_timestamp(run.get("created_at", ""))
    language = _visible_language_code(run.get("language") or "en")
    suffix = Path(str(source_artifact.get("path") or "")).suffix.lower() or ".txt"
    prefix = f"{_safe_delivery_name(project['name'])}_{language}_{timestamp}_{task_code}-{_short_run_id(task_run_id)}"
    return output_dir / f"{prefix}_final{suffix}"


def _normalize_delivery_workbook_headers(path: Path, language: Any) -> None:
    code = require_supported_language(language or "en")
    target = _visible_language_code(code)
    aliases = {alias.strip().lower() for alias in target_aliases(code)}
    if not aliases:
        return
    wb = load_workbook(path)
    changed = False
    try:
        for ws in wb.worksheets:
            for cell in ws[1]:
                value = str(cell.value or "").strip()
                if value and value.lower() in aliases and value != target:
                    cell.value = target
                    changed = True
        if changed:
            wb.save(path)
    finally:
        wb.close()


def _delivery_timestamp(value: str) -> str:
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            return parsed.strftime("%Y%m%d%H%M")
        return parsed.astimezone(ZoneInfo("Asia/Shanghai")).strftime("%Y%m%d%H%M")
    except Exception:
        return datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y%m%d%H%M")


def _run_artifact(run_id: str, kind: str) -> dict[str, Any] | None:
    artifacts = [artifact for artifact in db.list_artifacts(run_id=run_id) if artifact["kind"] == kind]
    return artifacts[0] if artifacts else None


def _input_artifact_label(run: dict[str, Any], project_id: str, seen: set[str] | None = None) -> str:
    seen = seen or set()
    run_id = str(run.get("id") or "")
    if run_id:
        if run_id in seen:
            return "-"
        seen.add(run_id)
    metadata = run.get("metadata") or {}
    source_run_id = metadata.get("manual_fix_source_run_id") or metadata.get("model_fix_source_run_id") or metadata.get("source_run_id")
    if source_run_id:
        try:
            source_run = db.get_run(str(source_run_id))
            if source_run.get("project_id") == project_id:
                source_label = _input_artifact_label(source_run, project_id, seen)
                if source_label and source_label != "-":
                    return source_label
        except KeyError:
            pass
    input_artifacts = metadata.get("input_artifacts") if isinstance(metadata.get("input_artifacts"), dict) else {}
    candidates = [
        input_artifacts.get("source_workbook"),
        metadata.get("input_artifact_id"),
        input_artifacts.get("translation_workbook"),
    ]
    for artifact_id in candidates:
        if not artifact_id:
            continue
        try:
            artifact = db.get_artifact(str(artifact_id))
            if artifact["project_id"] == project_id:
                return _artifact_display_label(artifact)
        except KeyError:
            continue
    return "-"


def _workbook_processed_rows(path: Path, language: Any | None = None) -> dict[str, int]:
    stats = {"source_rows": 0, "translated_rows": 0, "processed_rows": 0}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = {
                    str(value).strip().lower(): index
                    for index, value in enumerate(header_row, start=1)
                    if value is not None and str(value).strip()
                }
                source_col = _first_col(headers, list(SOURCE_HEADER_ALIASES))
                target_col = _first_col(headers, _target_headers_for_processed_rows(language))
                if source_col is None or target_col is None:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    has_source = bool(_row_cell(row, source_col))
                    has_target = bool(_row_cell(row, target_col))
                    if has_source:
                        stats["source_rows"] += 1
                    if has_target:
                        stats["translated_rows"] += 1
                    if has_source and has_target:
                        stats["processed_rows"] += 1
        finally:
            wb.close()
    except Exception:
        return stats
    return stats


def _target_headers_for_processed_rows(language: Any | None = None) -> list[str]:
    if language:
        code = require_supported_language(language)
        return [*target_aliases(code), _visible_language_code(code), "target", "translation", "译文"]
    headers: list[str] = ["target", "translation", "译文"]
    for code in PROJECT_LANGUAGE_ORDER:
        headers.extend(target_aliases(code))
        headers.append(_visible_language_code(code))
    return headers


def _soft_warning_count(summary: dict[str, Any]) -> int:
    total = 0
    for key in ("global_harness_quality", "project_harness_quality", "semantic_qa"):
        payload = summary.get(key) if isinstance(summary.get(key), dict) else {}
        total += int(payload.get("soft_warnings") or payload.get("warnings") or 0)
    return total


def _deliverable_provider_model(metadata: dict[str, Any], quality_summary: dict[str, Any]) -> tuple[str, str]:
    model_info = metadata.get("model") if isinstance(metadata.get("model"), dict) else {}
    if model_info.get("provider"):
        return str(model_info.get("provider") or "-"), str(model_info.get("model") or "-")
    semantic_qa = metadata.get("semantic_qa") if isinstance(metadata.get("semantic_qa"), dict) else quality_summary.get("semantic_qa", {})
    if not isinstance(semantic_qa, dict):
        return "-", "-"
    status = str(semantic_qa.get("status") or "")
    if status == "skipped_no_key":
        return "rules-only", "-"
    return str(semantic_qa.get("provider") or "-"), str(semantic_qa.get("model") or "-")


def _safe_delivery_name(name: str) -> str:
    return safe_delivery_name(name)


def _delivery_file(kind: str, path: Path) -> dict[str, str]:
    return {"kind": kind, "filename": path.name, "path": str(path)}


def _artifact_delivery_file(kind: str, artifact: dict[str, Any]) -> dict[str, str]:
    path = Path(str(artifact.get("path") or ""))
    return {
        "kind": kind,
        "filename": path.name,
        "path": str(path),
        "artifact_id": str(artifact.get("id") or ""),
        "download_url": artifact_download_url(str(artifact["project_id"]), str(artifact["id"])),
    }


def _expected_delivery_file(kind: str, path: Path) -> dict[str, str]:
    return {"kind": kind, "filename": path.name, "path": ""}


def _write_empty_workbook(path: Path, headers: list[str], note: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery"
    ws.append(headers)
    ws.append(["", note, ""])
    wb.save(path)
    wb.close()


def _latest_artifact(project_id: str, role: str | None = None, kind: str | None = None) -> dict[str, Any] | None:
    artifacts = db.list_artifacts(project_id=project_id, role=role)
    if kind:
        artifacts = [artifact for artifact in artifacts if artifact["kind"] == kind]
    return artifacts[0] if artifacts else None


def _normalize_delivery_languages(languages: list[str]) -> list[str]:
    normalized: list[str] = []
    for item in languages:
        value = str(item or "").strip()
        if not value:
            continue
        code = require_supported_language(value)
        if code not in normalized:
            normalized.append(code)
    if not normalized:
        raise ValueError("请选择至少一种目标语言")
    return normalized


def _find_merge_source_run(project_id: str, input_artifact_id: str, language: str) -> dict[str, Any] | None:
    accepted_status = {"passed", "failed", "needs_input"}
    for run in db.list_runs(project_id):
        if run.get("kind") not in {"qa", "translation"}:
            continue
        if require_supported_language(run.get("language") or "en") != language:
            continue
        if run.get("status") not in accepted_status:
            continue
        metadata = run.get("metadata") or {}
        candidates = {
            str(metadata.get("input_artifact_id") or ""),
            str(metadata.get("parent_input_artifact_id") or ""),
            str(metadata.get("multilingual_source_artifact_id") or ""),
        }
        input_artifacts = metadata.get("input_artifacts") if isinstance(metadata.get("input_artifacts"), dict) else {}
        candidates.update(str(value) for value in input_artifacts.values() if value)
        if input_artifact_id not in candidates:
            continue
        if _deliverable_final_artifact(run):
            return run
    return None


def _merge_language_column(target_path: Path, source_path: Path, language: str) -> int:
    code = require_supported_language(language)
    visible = _visible_language_code(code)
    wb = load_workbook(target_path)
    source_wb = load_workbook(source_path, data_only=False)
    copied = 0
    try:
        for target_ws in wb.worksheets:
            source_ws = source_wb[target_ws.title] if target_ws.title in source_wb.sheetnames else (source_wb.worksheets[0] if len(source_wb.worksheets) == 1 else None)
            if source_ws is None:
                continue
            target_headers = _delivery_header_map(target_ws)
            source_headers = _delivery_header_map(source_ws)
            target_id_col = _first_header_index(target_headers, ["id"])
            source_id_col = _first_header_index(source_headers, ["id"])
            source_lang_col = _first_header_index(source_headers, [visible, *target_aliases(code)])
            if source_lang_col is None:
                continue
            target_lang_col = _first_header_index(target_headers, [visible, *target_aliases(code)])
            if target_lang_col is None:
                target_lang_col = target_ws.max_column + 1
                target_ws.cell(row=1, column=target_lang_col).value = visible
            if target_id_col and source_id_col:
                source_by_id = {
                    _delivery_cell_text(source_ws.cell(row=row_index, column=source_id_col).value): source_ws.cell(row=row_index, column=source_lang_col).value
                    for row_index in range(2, source_ws.max_row + 1)
                }
                for row_index in range(2, target_ws.max_row + 1):
                    row_id = _delivery_cell_text(target_ws.cell(row=row_index, column=target_id_col).value)
                    if not row_id or row_id not in source_by_id:
                        continue
                    value = source_by_id[row_id]
                    if value not in (None, ""):
                        target_ws.cell(row=row_index, column=target_lang_col).value = value
                        copied += 1
            else:
                max_row = min(target_ws.max_row, source_ws.max_row)
                for row_index in range(2, max_row + 1):
                    value = source_ws.cell(row=row_index, column=source_lang_col).value
                    if value not in (None, ""):
                        target_ws.cell(row=row_index, column=target_lang_col).value = value
                        copied += 1
        wb.save(target_path)
    finally:
        wb.close()
        source_wb.close()
    return copied


# NOTE: 命名带 _delivery_ 前缀是刻意的。app/workflow/__init__.py 会把所有子模块的
# 顶层符号合并注入共享命名空间，qa.py 也定义了 _header_map / _cell_text（且 _cell_text
# 不做 strip），后加载的 qa 版本会静默覆盖本模块版本，导致合并交付时带空格的 ID
# 匹配失败。改名以避开注入覆盖，保证交付合并始终使用 strip 语义。
def _delivery_header_map(ws: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        value = _delivery_cell_text(cell.value).lower()
        if value:
            headers[value] = int(cell.column)
    return headers


def _first_header_index(headers: dict[str, int], aliases: list[str]) -> int | None:
    for alias in aliases:
        key = _delivery_cell_text(alias).lower()
        if key in headers:
            return headers[key]
    return None


def _delivery_cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _write_merged_delivery_summary(output_dir: Path, merged: list[dict[str, Any]], skipped: list[dict[str, Any]], final_path: Path) -> Path:
    path = output_dir / f"{final_path.stem}_QA摘要.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["项目", "值"])
    ws.append(["合并语言数", len(merged)])
    ws.append(["跳过/失败语言数", len(skipped)])
    ws.append(["输出文件", final_path.name])
    ws.append([])
    ws.append(["language", "status", "run_id", "rows", "hard_errors", "note"])
    for item in merged:
        ws.append([item.get("language"), item.get("status"), item.get("run_id"), item.get("rows"), item.get("hard_errors"), "已写入合并交付"])
    for item in skipped:
        ws.append([item.get("language"), "skipped", item.get("run_id", ""), 0, "", item.get("reason", "")])
    issues = wb.create_sheet("Issues")
    issues.append(["severity", "language", "run_id", "message"])
    for item in skipped:
        issues.append(["warning", item.get("language"), item.get("run_id", ""), item.get("reason", "")])
    outputs = wb.create_sheet("Outputs")
    outputs.append(["kind", "filename", "path"])
    outputs.append(["merged_final", final_path.name, str(final_path)])
    wb.save(path)
    wb.close()
    return path



__all__ = [name for name in globals() if not name.startswith("__")]
