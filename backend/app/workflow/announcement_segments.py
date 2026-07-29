from __future__ import annotations

import hashlib
import importlib.util
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .. import db
from ..config import GLOSSARY_ROOT
from ..languages import ANNOUNCEMENT_LANGUAGE_ORDER, require_supported_language, target_aliases
from .announcement_outputs import _file_sha256, _safe_source_stem, _visible_language_code
from .announcement_shared import (
    _announcement_task_metadata,
    _announcement_term_occurs,
    _is_low_value_announcement_term,
    _rank_translation_lookup_source,
    _suppress_overlapping_lookup_hits,
)
from .common import _CJK_RE
from .table_helpers import _wide_source_key


_XLSX_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
_LANGUAGE_TABLE_HEADER_SUMMARY_CACHE: dict[tuple[str, int, int, tuple[str, ...]], tuple[list[str], dict[str, int | None]]] = {}


class _RawXlsxCell:
    def __init__(self, value: Any, row_number: int, column_index: int) -> None:
        self.value = value
        self.coordinate = f"{_xlsx_column_name(column_index + 1)}{row_number}"


class _RawXlsxSheet:
    def __init__(self, title: str, rows: list[list[Any]]) -> None:
        self.title = title
        self._rows = rows
        self.max_row = len(rows)

    def iter_rows(self, min_row: int = 1, max_row: int | None = None, values_only: bool = False) -> Any:
        start = max(0, int(min_row or 1) - 1)
        stop = int(max_row) if max_row is not None else len(self._rows)
        for row_offset, row in enumerate(self._rows[start:stop], start=start + 1):
            if values_only:
                yield tuple(row)
            else:
                yield tuple(_RawXlsxCell(value, row_offset, index) for index, value in enumerate(row))


def _xlsx_column_name(number: int) -> str:
    name = ""
    while number > 0:
        number, remainder = divmod(number - 1, 26)
        name = chr(65 + remainder) + name
    return name or "A"


def _load_xlsx_sheets(path: Path) -> tuple[list[Any], Any]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        return list(wb.worksheets), wb.close
    except Exception:
        raw_sheets = _glossary_extractor_module().iter_raw_xlsx_sheets(path)
        return [_RawXlsxSheet(title, rows) for title, rows in raw_sheets], lambda: None


def _announcement_source_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return "docx"
    if suffix in {".txt", ".md", ".markdown"}:
        return "txt"
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return "xlsx"
    return suffix.lstrip(".")


def _announcement_source_manifest(artifact: dict[str, Any]) -> dict[str, Any]:
    path = Path(artifact["path"])
    return {"artifact_id": artifact["id"], "label": artifact.get("label", ""), "path": str(path), "format": _announcement_source_format(path), "sha256": _file_sha256(path) if path.exists() else ""}


def _announcement_task_source_text(task: dict[str, Any]) -> str:
    from .materials import _compact_lookup_text, _read_lookup_material_text

    artifact = db.get_artifact(task["source_artifact_id"])
    return _compact_lookup_text(_read_lookup_material_text(Path(artifact["path"])))


def _announcement_task_segments(task: dict[str, Any]) -> list[dict[str, Any]]:
    metadata = _announcement_task_metadata(task)
    if metadata.get("segments"):
        return list(metadata["segments"])
    artifact = db.get_artifact(task["source_artifact_id"])
    path = Path(artifact["path"])
    fmt = _announcement_source_format(path)
    if fmt == "docx":
        return _docx_announcement_segments(path)
    if fmt == "xlsx":
        return _xlsx_announcement_segments(path)
    return _txt_announcement_segments(path)


def _docx_announcement_segments(path: Path) -> list[dict[str, Any]]:
    from docx import Document

    doc = Document(str(path))
    rows = []
    for index, paragraph in enumerate(doc.paragraphs):
        source = paragraph.text
        if not source.strip():
            continue
        rows.append({"id": _segment_id(path.name, index, source), "source_file": path.name, "index": index, "kind": "paragraph", "source": source, "style": paragraph.style.name if paragraph.style else ""})
    return rows


def _txt_announcement_segments(path: Path) -> list[dict[str, Any]]:
    text = _read_announcement_text(path)
    rows = []
    for index, line in enumerate(text.splitlines()):
        if not line.strip():
            continue
        rows.append({"id": _segment_id(path.name, index, line), "source_file": path.name, "index": index, "kind": "line", "source": line})
    return rows


def _read_announcement_text(path: Path) -> str:
    from .materials import _read_lookup_text_file

    return _read_lookup_text_file(path)


def _is_quick_text_path(path: Path) -> bool:
    return path.suffix.lower() in {".txt", ".md", ".markdown"}


def _quick_text_translation_rows(path: Path) -> list[dict[str, Any]]:
    return [
        {"id": segment["id"], "source": segment["source"], "index": segment["index"], "source_file": segment["source_file"]}
        for segment in _txt_announcement_segments(path)
    ]


def _write_quick_text_output(source_path: Path, translated_rows: list[dict[str, Any]], language: str, output_dir: Path) -> Path:
    translations = {str(row.get("id")): str(row.get("translation") or "") for row in translated_rows}
    segments = _txt_announcement_segments(source_path)
    by_index = {int(segment["index"]): translations.get(str(segment["id"]), segment["source"]) for segment in segments}
    source_text = _read_announcement_text(source_path)
    raw_lines = source_text.splitlines(keepends=True)
    if not raw_lines and source_text.strip():
        raw_lines = [source_text]
    output_parts: list[str] = []
    for index, raw in enumerate(raw_lines):
        if raw.endswith("\r\n"):
            content, newline = raw[:-2], "\r\n"
        elif raw.endswith("\n"):
            content, newline = raw[:-1], "\n"
        elif raw.endswith("\r"):
            content, newline = raw[:-1], "\r"
        else:
            content, newline = raw, ""
        output_parts.append((by_index.get(index, content) if content.strip() else content) + newline)
    suffix = source_path.suffix.lower() if source_path.suffix.lower() in {".txt", ".md", ".markdown"} else ".txt"
    output_path = output_dir / f"{_safe_source_stem(source_path.name)}_{_visible_language_code(language)}{suffix}"
    output_path.write_text("".join(output_parts), encoding="utf-8")
    return output_path


def _xlsx_announcement_segments(path: Path) -> list[dict[str, Any]]:
    sheets, close = _load_xlsx_sheets(path)
    rows: list[dict[str, Any]] = []
    try:
        for ws in sheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = str(cell.value or "").strip()
                    if not value or not _CJK_RE.search(value):
                        continue
                    key = f"{ws.title}!{cell.coordinate}"
                    rows.append({"id": _segment_id(path.name, len(rows), f"{key}:{value}"), "source_file": path.name, "index": len(rows), "kind": "cell", "sheet": ws.title, "coordinate": cell.coordinate, "source": value})
    finally:
        close()
    return rows


def _segment_id(source_file: str, index: int, source: str) -> str:
    digest_source = "\0".join([source_file, str(index), source])
    digest = hashlib.sha1(digest_source.encode("utf-8")).hexdigest()[:10]
    return f"{Path(source_file).stem}:{index:04d}:{digest}"


def _detect_announcement_constraint_languages(project_id: str, metadata: dict[str, Any]) -> list[str]:
    found: set[str] = set()
    for artifact_id in _announcement_constraint_artifact_ids(metadata):
        try:
            artifact = db.get_artifact(artifact_id)
        except KeyError:
            continue
        if artifact["project_id"] != project_id:
            continue
        if _is_generated_announcement_terms_artifact(artifact):
            continue
        found.update(_detect_language_columns(Path(artifact["path"])))
    if metadata.get("include_project_archive", True):
        found.update(_project_glossary_languages(project_id))
        found.update(_translation_archive_languages(project_id))
    return [language for language in ANNOUNCEMENT_LANGUAGE_ORDER if language in found]


def _announcement_language_constraint_summary(project_id: str, metadata: dict[str, Any], languages: list[str]) -> dict[str, Any]:
    table_counts = _language_table_count_summary_from_artifacts(
        project_id,
        _announcement_constraint_artifact_ids(metadata),
        languages,
    )
    include_project_archive = metadata.get("include_project_archive", True)
    glossary_counts = _project_glossary_count_summary(project_id, languages) if include_project_archive else {}
    archive_counts = _translation_archive_count_summary(project_id, languages) if include_project_archive else {}
    return {
        language: {
            "project_glossary": glossary_counts.get(language, 0),
            "language_table": table_counts.get(language),
            "qa_archive": archive_counts.get(language, 0),
        }
        for language in languages
    }


def _announcement_constraint_artifact_ids(metadata: dict[str, Any]) -> list[str]:
    """Return selected constraint artifacts once, preserving user selection order."""
    result: list[str] = []
    for artifact_id in [*metadata.get("language_table_artifact_ids", []), *metadata.get("constraint_artifact_ids", [])]:
        value = str(artifact_id or "").strip()
        if value and value not in result:
            result.append(value)
    return result


def _translation_archive_count_summary(project_id: str, languages: list[str]) -> dict[str, int]:
    if not languages:
        return {}
    placeholders = ",".join("?" for _ in languages)
    with db.connect() as conn:
        rows = conn.execute(
            f"""
            SELECT language, COUNT(*) AS count
            FROM translation_entries
            WHERE project_id = ? AND active = 1 AND language IN ({placeholders})
            GROUP BY language
            """,
            [project_id, *languages],
        ).fetchall()
    return {str(row["language"]): int(row["count"] or 0) for row in rows}


def _translation_archive_languages(project_id: str) -> set[str]:
    with db.connect() as conn:
        rows = conn.execute(
            """
            SELECT DISTINCT language
            FROM translation_entries
            WHERE project_id = ? AND active = 1
            """,
            (project_id,),
        ).fetchall()
    return {str(row["language"]) for row in rows}


def _project_glossary_languages(project_id: str) -> set[str]:
    return {
        str(row.get("language") or "")
        for row in db.list_glossary_terms(project_id)
        if str(row.get("target") or row.get("target_alt") or "").strip()
    }


def _project_glossary_count_summary(project_id: str, languages: list[str]) -> dict[str, int]:
    return {
        language: sum(
            1
            for row in db.list_glossary_terms(project_id, language=language)
            if str(row.get("target") or row.get("target_alt") or "").strip()
        )
        for language in languages
    }


def _normalize_announcement_languages(raw: Any, fallback: list[str] | tuple[str, ...] = ()) -> list[str]:
    values = list(raw or []) or list(fallback or [])
    normalized: list[str] = []
    for value in values:
        try:
            code = require_supported_language(value)
        except ValueError:
            continue
        if code not in normalized:
            normalized.append(code)
    return [code for code in ANNOUNCEMENT_LANGUAGE_ORDER if code in normalized]


_SOURCE_HEADER_ALIASES = [
    "cn",
    "zh",
    "zh-cn",
    "zh_cn",
    "zhcn",
    "cnzh",
    "cn-zh",
    "source",
    "chinese",
    "中文",
    "中文key",
    "原文",
    "简体中文",
    "term",
    "术语",
]

def _detect_language_columns(path: Path) -> list[str]:
    if path.suffix.lower() not in _XLSX_SUFFIXES:
        return []
    found, _counts = _language_table_header_summary(path, ANNOUNCEMENT_LANGUAGE_ORDER)
    return found


def _language_table_header_summary(path: Path, languages: list[str]) -> tuple[list[str], dict[str, int | None]]:
    stat = path.stat()
    ordered_languages = tuple(languages)
    cache_key = (str(path.resolve()), int(stat.st_mtime_ns), int(stat.st_size), ordered_languages)
    cached = _LANGUAGE_TABLE_HEADER_SUMMARY_CACHE.get(cache_key)
    if cached is not None:
        return cached
    found: set[str] = set()
    counts: dict[str, int | None] = {language: 0 for language in languages}
    sheets, close = _load_xlsx_sheets(path)
    try:
        for ws in sheets:
            layout = _language_table_header_layout(ws, languages)
            if not layout:
                continue
            estimated_rows: int | None = None
            exact_counts: dict[str, int] | None = None
            if ws.max_row is not None and int(ws.max_row or 0) <= 5000:
                exact_counts = {language: 0 for language in languages}
                source_indices = [int(index) for index in layout.get("source_indices") or [layout["source_index"]]]
                language_indices = layout.get("language_indices") or {}
                for values in ws.iter_rows(min_row=int(layout["header_row"]) + 1, values_only=True):
                    sources = [
                        str(values[index] or "").strip()
                        for index in source_indices
                        if index < len(values)
                        and str(values[index] or "").strip()
                        and str(values[index] or "").strip().lower() not in {"string", "text", "c", "s", "int", "float", "number", "bool", "#ignore"}
                    ]
                    if not sources:
                        continue
                    for language, index in language_indices.items():
                        if index is not None and index < len(values) and str(values[index] or "").strip():
                            exact_counts[language] = exact_counts.get(language, 0) + len(sources)
            elif ws.max_row is not None:
                estimated_rows = max(0, int(ws.max_row or 0) - int(layout["header_row"] or 1))
            for code, index in (layout.get("language_indices") or {}).items():
                if index is None:
                    continue
                found.add(code)
                if exact_counts is not None:
                    counts[code] = int(counts.get(code) or 0) + int(exact_counts.get(code, 0))
                elif estimated_rows is None:
                    counts[code] = None
                elif counts.get(code) is not None:
                    counts[code] = int(counts.get(code) or 0) + estimated_rows
    finally:
        close()
    result = ([code for code in languages if code in found], counts)
    if len(_LANGUAGE_TABLE_HEADER_SUMMARY_CACHE) > 64:
        _LANGUAGE_TABLE_HEADER_SUMMARY_CACHE.clear()
    _LANGUAGE_TABLE_HEADER_SUMMARY_CACHE[cache_key] = result
    return result


def _language_column_index(normalized_headers: dict[str, int], language: str) -> int | None:
    aliases = [alias.lower() for alias in target_aliases(language)]
    for alias in aliases:
        if alias in normalized_headers:
            return normalized_headers[alias]
    return None


def _header_index(headers: list[str], *, prefer_last: bool = False) -> dict[str, int]:
    normalized: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = str(header or "").strip().lower()
        if not key:
            continue
        if prefer_last or key not in normalized:
            normalized[key] = index
    return normalized


def _language_table_source_indices(headers: list[str]) -> list[int]:
    normalized_first = _header_index(headers)
    primary = _column_by_alias(normalized_first, _SOURCE_HEADER_ALIASES)
    chinese_prc_indices = [
        index
        for index, header in enumerate(headers)
        if (key := str(header or "").strip().lower()) == "chinese_prc"
        or (key.startswith("chinese_prc.") and key.removeprefix("chinese_prc.").isdigit())
    ]
    indices = [primary, *chinese_prc_indices]
    return [index for position, index in enumerate(indices) if index is not None and index not in indices[:position]]


def _reserved_language_table_indices(headers: list[str]) -> list[int]:
    normalized_first = _header_index(headers)
    indices = [
        _column_by_alias(normalized_first, ["id", "key", "编号", "序号"]),
        *_language_table_source_indices(headers),
    ]
    return [index for index in indices if index is not None]


def _column_by_alias(normalized_headers: dict[str, int], aliases: list[str]) -> int | None:
    for alias in aliases:
        key = alias.lower()
        if key in normalized_headers:
            return normalized_headers[key]
    return None


def _announcement_constraint_rows(project_id: str, metadata: dict[str, Any], languages: list[str]) -> list[dict[str, Any]]:
    rows = _language_table_rows_from_artifacts(project_id, _announcement_constraint_artifact_ids(metadata), languages)
    by_source: dict[str, dict[str, Any]] = {
        _wide_source_key(row.get("source")): row
        for row in rows
        if _wide_source_key(row.get("source"))
    }
    for row in by_source.values():
        translation_sources = row.setdefault("translation_sources", {})
        for language, target in (row.get("translations") or {}).items():
            if str(target or "").strip():
                translation_sources.setdefault(language, {"type": "language_table", "priority": 1})
    if metadata.get("include_project_archive", True):
        for language in languages:
            for entry in db.list_translation_entries(project_id, language=language):
                source = str(entry.get("source") or "").strip()
                if not source:
                    continue
                key = _wide_source_key(source)
                row = by_source.setdefault(
                    key,
                    {
                        "id": entry.get("entry_key") or entry.get("id"),
                        "source": source,
                        "translations": {},
                        "translation_sources": {},
                        "sources": [],
                    },
                )
                target = str(entry.get("target") or "").strip()
                if target:
                    translation_sources = row.setdefault("translation_sources", {})
                    current = str((row.get("translations") or {}).get(language) or "").strip()
                    current_source = translation_sources.get(language) if isinstance(translation_sources.get(language), dict) else {}
                    current_priority = int(current_source.get("priority", 99))
                    archive_rank = _rank_translation_lookup_source(entry.get("source_type", ""))
                    current_archive_rank = int(current_source.get("archive_rank", 99))
                    if not current or 2 < current_priority or (current_priority == 2 and archive_rank < current_archive_rank):
                        row["translations"][language] = target
                        translation_sources[language] = {
                            "type": "qa_archive",
                            "priority": 2,
                            "archive_rank": archive_rank,
                        }
                row.setdefault("sources", []).append({"type": "qa_archive", "language": language, "entry_id": entry.get("id")})
            for term in db.list_glossary_terms(project_id, language=language):
                source = str(term.get("source") or "").strip()
                target = str(term.get("target") or term.get("target_alt") or "").strip()
                if not source or not target:
                    continue
                key = _wide_source_key(source)
                row = by_source.setdefault(
                    key,
                    {
                        "id": term.get("term_key") or term.get("id"),
                        "source": source,
                        "translations": {},
                        "translation_sources": {},
                        "sources": [],
                    },
                )
                translation_sources = row.setdefault("translation_sources", {})
                current_source = translation_sources.get(language) if isinstance(translation_sources.get(language), dict) else {}
                if int(current_source.get("priority", 99)) > 0:
                    row["translations"][language] = target
                    translation_sources[language] = {"type": "project_glossary", "priority": 0}
                row.setdefault("sources", []).append({"type": "project_glossary", "language": language, "term_id": term.get("id")})
    return list(by_source.values())


def _announcement_sentence_template_matches(
    project_id: str,
    metadata: dict[str, Any],
    languages: list[str],
    source_text: str,
    matched_terms: list[str],
) -> list[dict[str, Any]]:
    extractor = _glossary_extractor_module()
    candidates: list[dict[str, Any]] = []
    for artifact_id in _announcement_constraint_artifact_ids(metadata):
        artifact = db.get_artifact(artifact_id)
        if artifact["project_id"] != project_id:
            raise KeyError(artifact_id)
        if _is_generated_announcement_terms_artifact(artifact):
            continue
        path = Path(artifact["path"])
        if path.suffix.lower() not in _XLSX_SUFFIXES:
            continue
        for language in languages:
            visible_code = _visible_language_code(language)
            _found_languages, artifact_candidates = extractor.build_sentence_template_candidates_from_workbook(
                input_path=path,
                sheet_name=None,
                id_column="ID",
                source_column="CN",
                target_column=visible_code,
                language=visible_code,
                source_only=False,
            )
            candidates.extend(artifact_candidates)
    merged = extractor.merge_sentence_template_candidates(candidates)
    raw_matches = extractor.build_sentence_template_matches(
        candidate_rows=merged,
        announcement_text=source_text,
        matched_terms=matched_terms,
    )
    matches: list[dict[str, Any]] = []
    selected = set(languages)
    for row in raw_matches:
        translations: dict[str, str] = {}
        render_status: dict[str, str] = {}
        raw_status = row.get("_render_status") if isinstance(row.get("_render_status"), dict) else {}
        for raw_language, raw_target in (row.get("translations") or {}).items():
            try:
                language = require_supported_language(str(raw_language))
            except ValueError:
                continue
            target = str(raw_target or "").strip()
            if language not in selected or not target:
                continue
            translations[language] = target
            render_status[language] = str(raw_status.get(raw_language) or "").strip()
        if not translations:
            continue
        matches.append(
            {
                "priority": int(row.get("Priority") or 0),
                "match_type": str(row.get("MatchType") or "").strip(),
                "id": str(row.get("ID") or "").strip(),
                "announcement_cn": str(row.get("AnnouncementCN") or "").strip(),
                "official_cn_template": str(row.get("OfficialCNTemplate") or "").strip(),
                "translations": translations,
                "render_status": render_status,
            }
        )
    return matches


def _language_table_rows_from_artifacts(project_id: str, artifact_ids: list[str], languages: list[str]) -> list[dict[str, Any]]:
    by_source: dict[str, dict[str, Any]] = {}
    scanned_artifacts = 0
    for artifact_id in artifact_ids:
        if not artifact_id:
            continue
        artifact = db.get_artifact(artifact_id)
        if artifact["project_id"] != project_id:
            raise KeyError(artifact_id)
        if _is_generated_announcement_terms_artifact(artifact):
            continue
        path = Path(artifact["path"])
        scanned_artifacts += 1
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise ValueError(f"约束文件格式不正确：{path.name} 不是 XLSX 语言表。请上传完整语言表/术语交付表，不要把公告原文或 TXT 放在约束来源。")
        artifact_rows = _read_language_table_rows(path, languages)
        if not artifact_rows:
            visible = " / ".join(_visible_language_code(language) for language in languages) or "目标语言"
            raise ValueError(f"约束文件未识别到可反查词条：{path.name}。请检查表头是否包含 ID、CN/简体中文/原文，以及 {visible} 目标语言列。")
        for row in artifact_rows:
            key = _wide_source_key(row.get("source"))
            if not key:
                continue
            existing = by_source.setdefault(key, {"id": row.get("id", ""), "source": row.get("source", ""), "translations": {}, "sources": []})
            if row.get("id") and not existing.get("id"):
                existing["id"] = row["id"]
            for language, target in (row.get("translations") or {}).items():
                if target and not existing["translations"].get(language):
                    existing["translations"][language] = target
            existing["sources"].append({"type": "language_table", "artifact_id": artifact_id})
    if scanned_artifacts and not by_source:
        visible = " / ".join(_visible_language_code(language) for language in languages) or "目标语言"
        raise ValueError(f"约束文件未识别到可反查词条。请确认表头包含 ID、CN/简体中文/原文，以及 {visible} 目标语言列。")
    return list(by_source.values())


def _language_table_count_summary_from_artifacts(project_id: str, artifact_ids: list[str], languages: list[str]) -> dict[str, int | None]:
    """Estimate language-table coverage for Step 2 without reading every row.

    Step 2 is only a constraint/language recognition checkpoint. Full row parsing
    is intentionally deferred to term extraction, where the announcement source
    text is available and the rows are actually needed.
    """
    counts: dict[str, int | None] = {language: 0 for language in languages}
    scanned_artifacts = 0
    parsed_layout = False
    for artifact_id in artifact_ids:
        if not artifact_id:
            continue
        artifact = db.get_artifact(artifact_id)
        if artifact["project_id"] != project_id:
            raise KeyError(artifact_id)
        if _is_generated_announcement_terms_artifact(artifact):
            continue
        path = Path(artifact["path"])
        scanned_artifacts += 1
        if path.suffix.lower() not in _XLSX_SUFFIXES:
            raise ValueError(f"约束文件格式不正确：{path.name} 不是 XLSX 语言表。请上传完整语言表/术语交付表，不要把公告原文或 TXT 放在约束来源。")
        found_languages, artifact_counts = _language_table_header_summary(path, ANNOUNCEMENT_LANGUAGE_ORDER)
        parsed_layout = parsed_layout or bool(found_languages)
        for language, count in artifact_counts.items():
            if language not in languages:
                continue
            if language not in found_languages:
                continue
            if count is None:
                counts[language] = None
            elif counts.get(language) is not None:
                counts[language] = int(counts.get(language) or 0) + int(count or 0)
    if scanned_artifacts and not parsed_layout:
        visible = " / ".join(_visible_language_code(language) for language in languages) or "目标语言"
        raise ValueError(f"约束文件未识别到可反查词条。请确认表头包含 ID、CN/简体中文/原文，以及 {visible} 目标语言列。")
    return counts


def _is_generated_announcement_terms_artifact(artifact: dict[str, Any]) -> bool:
    if artifact.get("kind") == "announcement_terms_workbook":
        return True
    text = " ".join(
        str(part or "")
        for part in (
            artifact.get("label"),
            artifact.get("path"),
            (artifact.get("metadata") or {}).get("original_filename"),
        )
    ).lower()
    if "announcement_terms" not in text and "公告术语" not in text:
        return False
    return _workbook_looks_like_announcement_terms(Path(str(artifact.get("path") or "")))


def _workbook_looks_like_announcement_terms(path: Path) -> bool:
    if path.suffix.lower() not in _XLSX_SUFFIXES or not path.exists():
        return True
    sheets, close = _load_xlsx_sheets(path)
    try:
        for ws in sheets:
            try:
                headers = [str(value or "").strip().lower() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            except StopIteration:
                continue
            has_source = any(header in {alias.lower() for alias in _SOURCE_HEADER_ALIASES} for header in headers)
            has_hit_count = any(header in {"hit count", "hit_count", "hits", "命中次数"} or "命中" in header for header in headers)
            has_origin = any(header in {"来源", "origin", "source"} for header in headers)
            if has_source and has_hit_count and has_origin:
                return True
    finally:
        close()
    return False


def _read_language_table_rows(path: Path, languages: list[str]) -> list[dict[str, Any]]:
    if path.suffix.lower() not in _XLSX_SUFFIXES:
        return []
    sheets, close = _load_xlsx_sheets(path)
    rows: list[dict[str, Any]] = []
    try:
        for ws in sheets:
            layout = _language_table_header_layout(ws, languages)
            if not layout:
                continue
            id_idx = layout["id_index"]
            source_idx = layout["source_index"]
            source_indices = layout.get("source_indices") or [source_idx]
            lang_indices = layout["language_indices"]
            for values in ws.iter_rows(min_row=int(layout["header_row"]) + 1, values_only=True):
                translations = {}
                for language, index in lang_indices.items():
                    if index is not None and index < len(values):
                        value = str(values[index] or "").strip()
                        if value:
                            translations[language] = value
                for source_idx in source_indices:
                    source = str(values[source_idx] or "").strip() if source_idx < len(values) else ""
                    if not source:
                        continue
                    if source.lower() in {"string", "text", "c", "s", "int", "float", "number", "bool", "#ignore"}:
                        continue
                    rows.append({"id": str(values[id_idx] or "").strip() if id_idx is not None and id_idx < len(values) else "", "source": source, "translations": translations, "sheet": ws.title})
    finally:
        close()
    return rows


def _read_announcement_sentence_adaptations(path: Path, languages: list[str]) -> list[dict[str, Any]]:
    if path.suffix.lower() not in _XLSX_SUFFIXES:
        return []
    sheets, close = _load_xlsx_sheets(path)
    try:
        worksheet = next((sheet for sheet in sheets if str(sheet.title).strip().casefold() == "sentencetemplates"), None)
        if worksheet is None:
            return []
        values = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        normalized = _header_index(headers)
        priority_idx = _column_by_alias(normalized, ["priority", "优先级"])
        match_type_idx = _column_by_alias(normalized, ["matchtype", "match_type", "匹配类型"])
        id_idx = _column_by_alias(normalized, ["id", "key", "编号"])
        announcement_idx = _column_by_alias(normalized, ["announcementcn", "announcement_cn", "公告中文"])
        official_idx = _column_by_alias(normalized, ["officialcntemplate", "official_cn_template", "官方中文模板"])
        if None in {priority_idx, match_type_idx, id_idx, announcement_idx, official_idx}:
            raise ValueError("SentenceTemplates missing required columns")
        language_indices = {
            language: _language_column_index(_header_index(headers, prefer_last=True), language)
            for language in languages
        }
        output: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            match_type = str(row[match_type_idx] or "").strip() if match_type_idx < len(row) else ""
            if not match_type:
                continue
            if match_type not in {"official_exact", "official_similar"}:
                raise ValueError(f"unsupported SentenceTemplates match type: {match_type}")
            translations = {
                language: str(row[index] or "").strip()
                for language, index in language_indices.items()
                if index is not None and index < len(row) and str(row[index] or "").strip()
            }
            if not translations:
                continue
            output.append(
                {
                    "priority": int(row[priority_idx] or 0),
                    "match_type": match_type,
                    "id": str(row[id_idx] or "").strip(),
                    "announcement_cn": str(row[announcement_idx] or "").strip(),
                    "official_cn_template": str(row[official_idx] or "").strip(),
                    "translations": translations,
                    "render_status": {language: "imported" for language in translations},
                }
            )
        return output
    finally:
        close()


def _language_table_header_layout(ws: Any, languages: list[str], *, max_scan_rows: int = 12) -> dict[str, Any] | None:
    best: dict[str, Any] | None = None
    max_row = min(int(ws.max_row or max_scan_rows), max_scan_rows)
    for row_number, values in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        headers = [str(value or "").strip() for value in values]
        normalized_first = _header_index(headers)
        normalized_last = _header_index(headers, prefer_last=True)
        id_idx = _column_by_alias(normalized_first, ["id", "key", "编号", "序号"])
        source_indices = _language_table_source_indices(headers)
        if not source_indices:
            continue
        source_idx = source_indices[0]
        reserved_indices = set(index for index in [id_idx, *source_indices] if index is not None)
        language_indices = {
            language: index if index is not None and index not in reserved_indices else None
            for language in languages
            for index in [_language_column_index(normalized_last, language)]
        }
        language_count = sum(1 for index in language_indices.values() if index is not None)
        if language_count <= 0 and len(languages) == 1:
            adjacent_index = source_idx + 1
            if adjacent_index < len(headers) and not headers[adjacent_index]:
                has_adjacent_values = any(
                    source_idx < len(data_row)
                    and adjacent_index < len(data_row)
                    and str(data_row[source_idx] or "").strip()
                    and str(data_row[adjacent_index] or "").strip()
                    for data_row in ws.iter_rows(
                        min_row=row_number + 1,
                        max_row=min(int(ws.max_row or row_number + 20), row_number + 20),
                        values_only=True,
                    )
                )
                if has_adjacent_values:
                    language_indices[languages[0]] = adjacent_index
                    language_count = 1
        if language_count <= 0:
            continue
        score = language_count * 10 + (1 if id_idx is not None else 0) + row_number / 100
        candidate = {
            "header_row": row_number,
            "id_index": id_idx,
            "source_index": source_idx,
            "source_indices": source_indices,
            "language_indices": language_indices,
            "score": score,
        }
        if not best or score > float(best["score"]):
            best = candidate
    return best


def _select_announcement_constraint_rows(text: str, candidates: list[dict[str, Any]], languages: list[str], *, min_hit: int) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    for row in candidates:
        source = str(row.get("source") or "").strip()
        if len(source) < 2:
            continue
        if _is_low_value_announcement_term(source):
            continue
        hit_count, first_position = _announcement_term_occurs(text, source)
        if hit_count < min_hit:
            continue
        selected.append(
            {
                "id": row.get("id", ""),
                "source": source,
                "translations": {language: (row.get("translations") or {}).get(language, "") for language in languages},
                "translation_sources": {
                    language: (row.get("translation_sources") or {}).get(language)
                    for language in languages
                    if (row.get("translation_sources") or {}).get(language)
                },
                "sources": list(row.get("sources") or []),
                "hit_count": hit_count,
                "first_position": first_position,
            }
        )
    selected.sort(key=lambda row: (int(row.get("first_position") or 0), -len(str(row.get("source") or "")), str(row.get("source") or "")))
    return [
        row
        for row in _suppress_overlapping_lookup_hits(selected, text=text)
        if int(row.get("hit_count") or 0) >= min_hit
    ]


def _glossary_extractor_module() -> Any:
    global _GLOSSARY_EXTRACTOR_MODULE
    if _GLOSSARY_EXTRACTOR_MODULE is not None:
        return _GLOSSARY_EXTRACTOR_MODULE
    script_path = GLOSSARY_ROOT / "scripts" / "extract_glossary.py"
    spec = importlib.util.spec_from_file_location("lws_embedded_extract_glossary", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load glossary extractor: {script_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules.setdefault("lws_embedded_extract_glossary", module)
    spec.loader.exec_module(module)
    _GLOSSARY_EXTRACTOR_MODULE = module
    return module

__all__ = [name for name in globals() if not name.startswith("__")]
