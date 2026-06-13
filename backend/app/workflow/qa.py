from __future__ import annotations

import json
import re
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .. import db
from ..config import LOCALIZATION_ROOT, load_settings
from ..languages import require_supported_language, target_aliases
from ..translation_batches import manage_project_prompt_context as _manage_project_prompt_context
from .common import GLOBAL_HARNESS_CONTRACT, HARNESS_SCHEMA_VERSION, RowId, project_dir, read_project_harness, run_dir, write_project_harness
from .semantic_qa import run_semantic_qa_report
from .subprocess_runner import run_subprocess, run_subprocess_allow_failure


def _harness_summary(harness: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": "project_harness",
        "schema_version": harness.get("schema_version", HARNESS_SCHEMA_VERSION),
        "updated_at": harness.get("updated_at", ""),
        "style_guidance": bool(harness.get("style_guidance")),
        "hard_rules": len(harness.get("hard_rules", [])),
        "soft_rules": len(harness.get("soft_rules", [])),
        "fixed_terms": len(harness.get("fixed_terms", [])),
        "forbidden_translations": len(harness.get("forbidden_translations", [])),
        "reference_examples": len(harness.get("reference_examples", [])),
    }

def run_project_harness_qa(final_workbook: Path, harness: dict[str, Any], language: str = "en") -> dict[str, Any]:
    language = require_supported_language(language)
    issues: list[dict[str, Any]] = []
    if not final_workbook.exists():
        return {
            "source": "project_harness",
            "passed": False,
            "hard_errors": 1,
            "soft_warnings": 0,
            "issues": [{"severity": "hard", "message": "final workbook missing", "rule_source": "project_harness"}],
        }

    forbidden = [str(item).strip() for item in harness.get("forbidden_translations", []) if str(item).strip()]
    fixed_terms = [item for item in harness.get("fixed_terms", []) if isinstance(item, dict)]
    hard_rules = [item for item in harness.get("hard_rules", []) if isinstance(item, dict) and item.get("enabled", True)]

    wb = load_workbook(final_workbook, data_only=False)
    try:
        for ws in wb.worksheets:
            headers = _header_map(ws)
            source_col = _first_col(headers, ["cn", "source", "original", "原文", "中文"])
            target_col = _first_col(headers, target_aliases(language))
            if target_col is None:
                continue
            for row_index in range(2, ws.max_row + 1):
                source = _cell_text(ws.cell(row_index, source_col).value) if source_col else ""
                target = _cell_text(ws.cell(row_index, target_col).value)
                if not target:
                    continue
                for phrase in forbidden:
                    if phrase in target:
                        issues.append(
                            _project_issue(
                                ws.title,
                                row_index,
                                "forbidden_translation",
                                f"Translation contains forbidden phrase: {phrase}",
                                target,
                            )
                        )
                for term in fixed_terms:
                    source_term = str(term.get("source", "")).strip()
                    target_term = str(term.get("target", "")).strip()
                    if source_term and target_term and source_term in source and target_term not in target:
                        issues.append(
                            _project_issue(
                                ws.title,
                                row_index,
                                "fixed_term_missing",
                                f"Source term '{source_term}' must use '{target_term}'",
                                target,
                            )
                        )
                for rule in hard_rules:
                    pattern = str(rule.get("pattern", "")).strip()
                    if not pattern:
                        continue
                    try:
                        matched = re.search(pattern, target) is not None
                    except re.error as exc:
                        issues.append(
                            _project_issue(
                                ws.title,
                                row_index,
                                "invalid_project_rule",
                                f"Invalid project hard-rule pattern '{pattern}': {exc}",
                                target,
                                severity="soft",
                            )
                        )
                        continue
                    if matched:
                        message = str(rule.get("description") or rule.get("label") or f"Project rule matched: {pattern}")
                        issues.append(_project_issue(ws.title, row_index, "project_hard_rule", message, target))
    finally:
        wb.close()

    hard_errors = len([issue for issue in issues if issue["severity"] == "hard"])
    soft_warnings = len([issue for issue in issues if issue["severity"] == "soft"])
    return {
        "source": "project_harness",
        "passed": hard_errors == 0,
        "hard_errors": hard_errors,
        "soft_warnings": soft_warnings,
        "issues": issues[:100],
        "active_overlay": _harness_summary(harness),
    }


def list_improvements(project_id: str) -> list[dict[str, Any]]:
    path = project_dir(project_id) / "profile" / "improvement_suggestions.json"
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def create_project_improvement(project_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    db.get_project(project_id)
    category = str(payload.get("category") or "soft_rule").strip() or "soft_rule"
    title = str(payload.get("title") or "手动改进建议").strip() or "手动改进建议"
    detail = str(payload.get("detail") or "").strip()
    item = _improvement_item(category, "manual", title, detail)
    _append_improvement_items(project_id, [item])
    return item


def create_improvement_review(run_id: str) -> dict[str, Any]:
    run = db.get_run(run_id)
    project_id = run["project_id"]
    metadata = run.get("metadata", {})
    suggestions = list_improvements(project_id)
    quality = metadata.get("quality", {})
    project_quality = metadata.get("project_harness_quality", {})
    if project_quality.get("hard_errors"):
        suggestions.append(
            _improvement_item(
                "project_harness",
                run_id,
                "Review project-specific hard rules and manual fixes",
                "Project harness QA produced hard errors; update this project's harness only after human review.",
            )
        )
    if quality and not quality.get("passed", True):
        suggestions.append(
            _improvement_item(
                "studio_integration",
                run_id,
                "Review reusable QA adapter coverage",
                "Global quality gate failed; inspect whether Studio needs better reporting or retry controls.",
            )
        )
    suggestions.append(
        _improvement_item(
            "upstream_backfeed",
            run_id,
            "Prepare upstream backfeed candidate",
            "If this run exposed a reusable gap, create a human-reviewed issue or PR against the source workflow repo.",
        )
    )
    path = project_dir(project_id) / "profile" / "improvement_suggestions.json"
    path.write_text(json.dumps(suggestions, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"project_id": project_id, "run_id": run_id, "suggestions": suggestions}


def list_quality_issues(run_id: str) -> dict[str, Any]:
    run = db.get_run(run_id)
    metadata = run.get("metadata", {})
    summary = metadata.get("quality_summary", {})
    issues: list[dict[str, Any]] = []
    for source_key, payload in (
        ("global_harness", metadata.get("quality") or summary.get("global_harness_quality") or {}),
        ("project_harness", metadata.get("project_harness_quality") or summary.get("project_harness_quality") or {}),
        ("semantic_qa", metadata.get("semantic_qa") or summary.get("semantic_qa") or {}),
    ):
        issues.extend(_normalize_quality_issues(source_key, payload))
    hard_errors = len([issue for issue in issues if issue["severity"] == "hard"])
    return {
        "run_id": run_id,
        "project_id": run["project_id"],
        "status": run["status"],
        "hard_errors": hard_errors,
        "issues": issues,
    }


def apply_manual_fixes(run_id: str, request: Any) -> dict[str, Any]:
    run = db.get_run(run_id)
    project_id = run["project_id"]
    fixes = [fix.model_dump() if hasattr(fix, "model_dump") else dict(fix) for fix in getattr(request, "fixes", [])]
    if not fixes:
        raise ValueError("manual fixes are required")

    source_artifact = _workbook_artifact_for_quality_run(run)
    source_path = Path(source_artifact["path"])
    if not source_path.exists():
        raise FileNotFoundError(str(source_path))

    output_dir = run_dir(run_id) / "manual_fixes"
    output_dir.mkdir(parents=True, exist_ok=True)
    fixed_path = output_dir / f"{source_path.stem}_manual_fixed.xlsx"
    shutil.copy2(source_path, fixed_path)

    applied = _apply_workbook_fixes(fixed_path, fixes, run_id)
    fixed_artifact = db.add_artifact(
        project_id,
        "Manual fixed workbook",
        fixed_path,
        "manual_fixed_workbook",
        run_id=run_id,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        origin="manual",
        metadata={"source_run_id": run_id, "source_artifact_id": source_artifact["id"], "manual_fix_count": len(applied)},
    )
    harness = read_project_harness(project_id)
    write_project_harness(
        project_id,
        {
            "manual_fixes": [*harness.get("manual_fixes", []), *applied],
            "qa_summary": {
                **harness.get("qa_summary", {}),
                "last_manual_fix_run": run_id,
                "last_manual_fix_artifact": fixed_artifact["id"],
            },
        },
    )
    _append_improvement_items(
        project_id,
        [
            _improvement_item(
                "project_harness",
                run_id,
                "Review manual fixes for project harness reuse",
                "Manual QA fixes were applied; review whether they should become project-specific rules, fixed terms, or reference examples.",
            )
        ],
    )

    result: dict[str, Any] = {
        "source_run": run,
        "fixed_artifact": fixed_artifact,
        "manual_fixes": applied,
        "qa_result": None,
    }
    if getattr(request, "rerun_qa", True):
        qa_run = db.insert_run(
            project_id,
            kind="qa",
            language=run.get("language", "en"),
            metadata={
                "input_artifact_id": fixed_artifact["id"],
                "manual_fix_source_run_id": run_id,
                "manual_fix_source_artifact_id": source_artifact["id"],
                "manual_fix_count": len(applied),
                "manual_fixes": applied,
            },
        )
        result["qa_result"] = run_qa_sync(qa_run["id"])
    return result


def create_semantic_qa_context(run_id: str) -> dict[str, Any]:
    run = db.get_run(run_id)
    project = db.get_project(run["project_id"])
    output_dir = run_dir(run_id) / "semantic_qa"
    output_dir.mkdir(parents=True, exist_ok=True)
    report = {
        "status": "needs_model_review",
        "message": "Semantic QA context is prepared; model review is not auto-marked as passed.",
        "run_id": run_id,
        "project_id": project["id"],
        "sources": {
            "global_harness": GLOBAL_HARNESS_CONTRACT,
            "project_harness": read_project_harness(project["id"]),
        },
        "run_quality": {
            "global": run.get("metadata", {}).get("quality", {}),
            "project_harness": run.get("metadata", {}).get("project_harness_quality", {}),
        },
    }
    path = output_dir / "semantic_qa_context.json"
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    artifact = db.add_artifact(project["id"], "Semantic QA context", path, "semantic_qa_context", run_id=run_id, mime="application/json")
    return {"run": db.get_run(run_id), "artifact": artifact, "semantic_qa": report}


def run_qa_sync(run_id: str) -> dict[str, Any]:
    run = db.get_run(run_id)
    project = db.get_project(run["project_id"])
    metadata = run.get("metadata", {})
    workbook_artifact = _workbook_artifact_for_quality_run(run)
    workbook_path = Path(workbook_artifact["path"])
    if not workbook_path.exists():
        raise ValueError("译文表文件不存在，请重新上传或重新生成翻译结果后再运行 QA。")
    db.update_run(run_id, status="running")

    output_dir = run_dir(run_id) / "qa"
    output_dir.mkdir(parents=True, exist_ok=True)
    language = require_supported_language(run.get("language") or "en")
    from .prompt_snapshots import create_project_glossary_snapshot, create_prompt_and_harness_snapshots, create_quick_reference_snapshot

    glossary_snapshot = create_project_glossary_snapshot(project["id"], run_id, output_dir / "snapshots", language=language)
    snapshots = create_prompt_and_harness_snapshots(project["id"], run_id, output_dir / "snapshots", language=language)
    reference_snapshot = create_quick_reference_snapshot(project["id"], run_id, metadata.get("reference_artifact_ids"), output_dir / "snapshots")
    qa_result = run_localization_qa(
        project=project,
        run_id=run_id,
        workbook_path=workbook_path,
        output_dir=output_dir,
        glossary_snapshot=glossary_snapshot,
        harness_snapshot=snapshots["harness_snapshot"],
        workbook_artifact=workbook_artifact,
        run_metadata=metadata,
        manual_fixes=metadata.get("manual_fixes") or [],
        language=language,
    )
    input_artifacts = {
        "translation_workbook": workbook_artifact["id"],
        "glossary_snapshot": glossary_snapshot["id"],
        "prompt_snapshot": snapshots["prompt_artifact"]["id"],
        "harness_snapshot": snapshots["harness_artifact"]["id"],
    }
    if reference_snapshot:
        input_artifacts["quick_reference_snapshot"] = reference_snapshot["artifact"]["id"]
    if qa_result.get("qa_final_artifact"):
        input_artifacts["qa_final_workbook"] = qa_result["qa_final_artifact"]["id"]
    status = "passed" if qa_result["quality_summary"]["passed"] else "failed"
    archive_result = None
    if status == "passed" and qa_result.get("qa_final_artifact"):
        from .asset_import_export import archive_translation_artifact

        archive_result = archive_translation_artifact(
            project["id"],
            qa_result["qa_final_artifact"]["id"],
            language=run.get("language") or "en",
            source_type="qa_passed",
        )
    db.update_run(
        run_id,
        status=status,
        metadata={
            **metadata,
            "task_origin": metadata.get("task_origin") or "direct_import",
            "input_artifacts": input_artifacts,
            "quality": qa_result["quality"],
            "project_harness_quality": qa_result["project_harness_quality"],
            "semantic_qa": qa_result["semantic_qa"],
            "quality_summary": qa_result["quality_summary"],
            "translation_archive": archive_result,
        },
    )
    return {"run": db.get_run(run_id), "artifacts": qa_result["artifacts"], "quality_summary": qa_result["quality_summary"]}


def run_localization_qa(
    project: dict[str, Any],
    run_id: str,
    workbook_path: Path,
    output_dir: Path,
    glossary_snapshot: dict[str, Any],
    harness_snapshot: dict[str, Any],
    workbook_artifact: dict[str, Any] | None = None,
    run_metadata: dict[str, Any] | None = None,
    manual_fixes: list[dict[str, Any]] | None = None,
    language: str = "en",
) -> dict[str, Any]:
    language = require_supported_language(language)
    output_dir.mkdir(parents=True, exist_ok=True)
    machine_dir = output_dir / "machine_review"
    machine_dir.mkdir(parents=True, exist_ok=True)
    review_args = [
        sys.executable,
        str(LOCALIZATION_ROOT / "process_language.py"),
        "--input",
        str(workbook_path),
        "--lang",
        language,
        "--output-dir",
        str(machine_dir),
        "--auto-fix",
        "--term-base",
        glossary_snapshot["path"],
    ]
    run_subprocess(review_args, LOCALIZATION_ROOT, run_id)
    qa_workbook = machine_dir / f"result_{language}.xlsx"
    qa_report = machine_dir / f"report_{language}.xlsx"
    _normalize_review_workbook_sheet_names(qa_workbook, workbook_path)
    quality_args = [
        sys.executable,
        str(LOCALIZATION_ROOT / "scripts" / "run_quality_harness.py"),
        "--workbook",
        str(qa_workbook),
        "--term-base",
        glossary_snapshot["path"],
        "--lang",
        language,
        "--json",
    ]
    if language == "en":
        quality_args.insert(2, str(LOCALIZATION_ROOT / "fixtures" / "quality_regression.json"))
    quality = _run_quality_json(quality_args, run_id)
    project_harness_quality = run_project_harness_qa(qa_workbook, harness_snapshot["project_harness"], language=language)
    semantic_qa = run_semantic_qa_report(run_id, project["id"], qa_workbook, quality, project_harness_quality, language=language)
    hard_errors = _hard_error_count(quality) + int(project_harness_quality.get("hard_errors", 0)) + int(semantic_qa.get("hard_errors", 0))
    passed = hard_errors == 0
    summary = {
        "version": 1,
        "run_id": run_id,
        "project_id": project["id"],
        "passed": passed,
        "hard_errors": hard_errors,
        "sources": {
            "translation_workbook": (workbook_artifact or {}).get("id", ""),
            "qa_workbook": str(qa_workbook),
            "glossary_snapshot": glossary_snapshot["id"],
            "global_harness": GLOBAL_HARNESS_CONTRACT,
            "project_harness": "project_harness_snapshot.json",
            "semantic_qa": semantic_qa.get("status", ""),
        },
        "global_harness_quality": quality,
        "project_harness_quality": project_harness_quality,
        "semantic_qa": semantic_qa,
    }
    metadata = run_metadata or {}
    if metadata.get("manual_fix_source_run_id"):
        summary["sources"]["manual_fix_source_run"] = metadata["manual_fix_source_run_id"]
    if metadata.get("model_fix_source_run_id"):
        summary["sources"]["model_fix_source_run"] = metadata["model_fix_source_run_id"]
    summary_path = output_dir / "quality_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    auto_fixes = _collect_workbook_translation_changes(workbook_path, qa_workbook)
    changes_path = write_qa_changes_report(output_dir, manual_fixes or [], auto_fixes)
    artifacts = [
        db.add_artifact(
            project["id"],
            "QA reviewed workbook",
            qa_workbook,
            "qa_result",
            run_id=run_id,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        db.add_artifact(
            project["id"],
            "QA report",
            qa_report,
            "qa_report",
            run_id=run_id,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        db.add_artifact(project["id"], "Quality summary", summary_path, "quality_summary", run_id=run_id, mime="application/json"),
        db.add_artifact(
            project["id"],
            "QA changes",
            changes_path,
            "qa_changes",
            run_id=run_id,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    ]
    qa_final_artifact = db.add_artifact(
        project["id"],
        "QA final workbook",
        qa_workbook,
        "qa_final_workbook",
        run_id=run_id,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        origin="generated",
        metadata={
            "language": language,
            "source_workbook": str(workbook_path),
            "glossary_snapshot": glossary_snapshot["id"],
            "qa_passed": passed,
            "hard_errors": hard_errors,
        },
    )
    artifacts.append(qa_final_artifact)
    return {
        "artifacts": artifacts,
        "qa_final_artifact": qa_final_artifact,
        "quality": quality,
        "project_harness_quality": project_harness_quality,
        "semantic_qa": semantic_qa,
        "quality_summary": summary,
        "qa_workbook": qa_workbook,
    }


def write_qa_changes_report(output_dir: Path, manual_fixes: list[dict[str, Any]], auto_fixes: list[dict[str, Any]] | None = None) -> Path:
    path = output_dir / "qa_changes.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Changes"
    ws.append(["工作表", "行号", "问题ID", "修改前", "修改后", "规则来源", "备注"])
    if manual_fixes:
        for fix in manual_fixes:
            ws.append(
                [
                    fix.get("sheet", ""),
                    fix.get("row", ""),
                    fix.get("issue_id", ""),
                    fix.get("previous_translation", ""),
                    fix.get("translation", ""),
                    fix.get("rule_source", "manual_fix"),
                    fix.get("note", ""),
                ]
            )
    for fix in auto_fixes or []:
        ws.append(
            [
                fix.get("sheet", ""),
                fix.get("row", ""),
                fix.get("issue_id", "auto_fix"),
                fix.get("previous_translation", ""),
                fix.get("translation", ""),
                "localization_auto_fix",
                fix.get("note", ""),
            ]
        )
    if not manual_fixes and not auto_fixes:
        ws.append(["", "", "", "", "", "qa", "未应用修改"])
    wb.save(path)
    wb.close()
    return path


def _run_quality_json(args: list[str], run_id: str) -> dict[str, Any]:
    proc = run_subprocess_allow_failure(args, LOCALIZATION_ROOT, run_id)
    if not proc.stdout.strip():
        raise RuntimeError(f"quality harness returned no JSON: {proc.stderr}")
    return json.loads(proc.stdout)


def _hard_error_count(quality: dict[str, Any]) -> int:
    if quality.get("passed"):
        return 0
    hard_issues = [
        issue
        for issue in quality.get("issues", [])
        if str(issue.get("severity", "hard")).lower() not in {"warning", "soft", "info"}
    ]
    issue_count = len(hard_issues) + len(quality.get("failures", []))
    if issue_count:
        return issue_count
    return 1 if not quality.get("issues") and not quality.get("failures") else 0


def _collect_workbook_translation_changes(before_path: Path, after_path: Path) -> list[dict[str, Any]]:
    if not before_path.exists() or not after_path.exists():
        return []
    before_wb = load_workbook(before_path, read_only=True, data_only=True)
    after_wb = load_workbook(after_path, read_only=True, data_only=True)
    changes: list[dict[str, Any]] = []
    try:
        before_ws = before_wb[before_wb.sheetnames[0]]
        after_ws = after_wb[before_ws.title] if before_ws.title in after_wb.sheetnames else after_wb[after_wb.sheetnames[0]]
        before_headers = _header_map(before_ws)
        after_headers = _header_map(after_ws)
        before_id_col = _first_col(before_headers, ["id", "key", "编号", "序号"])
        after_id_col = _first_col(after_headers, ["id", "key", "编号", "序号"])
        before_target_col = _first_col(before_headers, ["en", "translation", "target", "译文", "英文"])
        after_target_col = _first_col(after_headers, ["en", "translation", "target", "译文", "英文"])
        if before_id_col is None or after_id_col is None or before_target_col is None or after_target_col is None:
            return []
        after_by_id: dict[str, tuple[int, str]] = {}
        for row_index, row in enumerate(after_ws.iter_rows(min_row=2, values_only=True), start=2):
            row_id = _row_cell(row, after_id_col)
            if row_id:
                after_by_id[row_id] = (row_index, _row_cell(row, after_target_col))
        for row_index, row in enumerate(before_ws.iter_rows(min_row=2, values_only=True), start=2):
            row_id = _row_cell(row, before_id_col)
            if not row_id or row_id not in after_by_id:
                continue
            after_row, after_text = after_by_id[row_id]
            before_text = _row_cell(row, before_target_col)
            if before_text != after_text:
                changes.append(
                    {
                        "sheet": after_ws.title,
                        "row": after_row,
                        "issue_id": "auto_fix",
                        "previous_translation": before_text,
                        "translation": after_text,
                        "note": f"localization workflow auto-fix for ID {row_id}",
                    }
                )
    finally:
        before_wb.close()
        after_wb.close()
    return changes


def _normalize_review_workbook_sheet_names(review_path: Path, source_path: Path) -> None:
    if not review_path.exists() or not source_path.exists():
        return
    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    review_wb = load_workbook(review_path)
    try:
        source_title = source_wb.sheetnames[0] if source_wb.sheetnames else ""
        if source_title and review_wb.sheetnames:
            first = review_wb[review_wb.sheetnames[0]]
            if first.title != source_title and source_title not in review_wb.sheetnames:
                first.title = source_title
                review_wb.save(review_path)
    finally:
        source_wb.close()
        review_wb.close()


def _row_cell(row: tuple[Any, ...], column: int) -> str:
    if column < 1 or column > len(row):
        return ""
    value = row[column - 1]
    return "" if value is None else str(value).strip()


def _normalize_translation_id(value: Any) -> RowId | None:
    if value is None:
        return None
    if isinstance(value, bool):
        text = str(value).strip()
        return text or None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else str(value).strip()
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"-?(0|[1-9]\d*)", text):
        return int(text)
    return text


def _is_supported_translation_id(value: Any) -> bool:
    return _normalize_translation_id(value) is not None


def _normalize_quality_issues(source_key: str, payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, issue in enumerate(payload.get("issues", []) or []):
        severity = str(issue.get("severity") or "hard").lower()
        if severity in {"warning", "soft", "info"}:
            severity = "soft" if severity != "info" else "info"
        else:
            severity = "hard"
        rows.append(
            {
                "id": str(issue.get("id") or f"{source_key}:{index}:{issue.get('sheet', '')}:{issue.get('row', '')}:{issue.get('check_type', '')}"),
                "source": source_key,
                "rule_source": issue.get("rule_source") or issue.get("source") or source_key,
                "severity": severity,
                "sheet": issue.get("sheet") or "",
                "row": int(issue.get("row") or 0),
                "check_type": issue.get("check_type") or issue.get("type") or issue.get("code") or "quality_issue",
                "message": issue.get("message") or issue.get("detail") or "",
                "current_translation": issue.get("translation") or issue.get("target") or issue.get("actual") or "",
            }
        )
    for index, failure in enumerate(payload.get("failures", []) or []):
        rows.append(
            {
                "id": str(failure.get("id") or f"{source_key}:failure:{index}"),
                "source": source_key,
                "rule_source": source_key,
                "severity": "hard",
                "sheet": failure.get("sheet") or "",
                "row": int(failure.get("row") or 0),
                "check_type": "fixture_failure",
                "message": failure.get("message") or f"Fixture failure: {failure.get('id', index)}",
                "current_translation": failure.get("actual") or "",
            }
        )
    return rows


def _workbook_artifact_for_quality_run(run: dict[str, Any]) -> dict[str, Any]:
    metadata = run.get("metadata", {})
    input_artifacts = metadata.get("input_artifacts") if isinstance(metadata.get("input_artifacts"), dict) else {}
    candidate_ids: list[Any] = []
    if run.get("kind") == "translation":
        candidate_ids.extend(
            [
                input_artifacts.get("qa_final_workbook"),
                input_artifacts.get("raw_translated_workbook"),
                input_artifacts.get("translation_workbook"),
            ]
        )
    else:
        candidate_ids.extend(
            [
                input_artifacts.get("qa_final_workbook"),
                input_artifacts.get("translation_workbook"),
                metadata.get("input_artifact_id"),
            ]
        )

    for input_artifact_id in candidate_ids:
        if not input_artifact_id:
            continue
        artifact = db.get_artifact(str(input_artifact_id))
        if artifact["project_id"] != run["project_id"]:
            continue
        if artifact["role"] in {"translation_workbook", "translation_draft", "language_source", "quick_input"}:
            return artifact
    if run.get("kind") == "translation":
        artifacts = db.list_artifacts(run_id=run["id"], role="translation_workbook") or db.list_artifacts(run_id=run["id"], role="translation_draft")
        if artifacts:
            return artifacts[0]
        raise ValueError("请先完成 AI 翻译，再运行 QA。")
    artifacts = db.list_artifacts(run_id=run["id"], role="translation_workbook") or db.list_artifacts(run_id=run["id"], role="language_source")
    if artifacts:
        return artifacts[0]
    raise KeyError("translation workbook artifact not found")


def _model_fix_row_context(path: Path, issue: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = str(issue.get("sheet") or wb.sheetnames[0])
        requested_ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
        issue_record_id = issue.get("id") or issue.get("record_id") or ""
        row_index = int(issue.get("row") or 0)
        resolved = _resolve_workbook_row_for_issue(wb, requested_ws, row_index, issue_record_id)
        ws, row_index = resolved if resolved else (requested_ws or wb[wb.sheetnames[0]], row_index)
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {
            str(value).strip().lower(): index
            for index, value in enumerate(header_row, start=1)
            if value is not None and str(value).strip()
        }
        source_col = _first_col(headers, ["cn", "source", "original", "原文", "中文"])
        target_col = _first_col(headers, ["en", "target", "translation", "译文", "英文"])
        id_col = _first_col(headers, ["id", "key", "编号", "序号"])
        row_values = next(ws.iter_rows(min_row=row_index, max_row=row_index, values_only=True), ())
        return {
            "issue_id": issue.get("id", ""),
            "sheet": ws.title,
            "row": row_index,
            "record_id": _row_cell(row_values, id_col) if id_col else "",
            "source_text": _row_cell(row_values, source_col) if source_col else "",
            "current_translation": _row_cell(row_values, target_col) if target_col else issue.get("current_translation", ""),
            "severity": issue.get("severity", "hard"),
            "check_type": issue.get("check_type", ""),
            "message": issue.get("message", ""),
            "rule_source": issue.get("rule_source") or issue.get("source") or "qa",
        }
    finally:
        wb.close()


def _resolve_workbook_row_for_issue(wb: Any, requested_ws: Any | None, row_index: int, record_id: Any) -> tuple[Any, int] | None:
    normalized_record_id = _normalize_translation_id(record_id)
    if requested_ws is not None and row_index >= 2:
        headers = _header_map(requested_ws)
        id_col = _first_col(headers, ["id", "key", "编号", "序号"])
        if id_col is None or normalized_record_id is None:
            return requested_ws, row_index
        current_id = _normalize_translation_id(requested_ws.cell(row_index, id_col).value)
        if current_id == normalized_record_id:
            return requested_ws, row_index
    if normalized_record_id is None:
        return (requested_ws, row_index) if requested_ws is not None and row_index >= 2 else None
    for ws in wb.worksheets:
        headers = _header_map(ws)
        id_col = _first_col(headers, ["id", "key", "编号", "序号"])
        if id_col is None:
            continue
        for candidate_row in range(2, ws.max_row + 1):
            if _normalize_translation_id(ws.cell(candidate_row, id_col).value) == normalized_record_id:
                return ws, candidate_row
    return (requested_ws, row_index) if requested_ws is not None and row_index >= 2 else None


def _model_fix_prompt(project: dict[str, Any], run: dict[str, Any], rows: list[dict[str, Any]]) -> str:
    language = require_supported_language(run.get("language") or "en")
    profile = project.get("profile") or {}
    prompt = str((profile.get("prompts_by_language") or {}).get(language) or project.get("prompt_text") or "").strip()
    prompt = _manage_project_prompt_context(prompt, load_settings())
    harness = read_project_harness(project["id"])
    return (
        "你是游戏本地化 QA 修复模型。请根据项目提示词、项目规则、术语要求和 QA 问题，"
        "只修复译文，不改原文，不解释过程。必须保留变量、数字、HTML/BBCode 标签、换行和占位符。"
        "如果无法确定，保留原译文并在 note 写明需要人工确认。\n\n"
        "返回严格 JSON：{\"fixes\":[{\"issue_id\":\"...\",\"record_id\":\"...\",\"sheet\":\"...\",\"row\":2,\"translation\":\"...\",\"note\":\"...\"}]}。"
        "必须优先沿用待修复行里的 issue_id 和 record_id；sheet/row 仅作辅助定位。\n"
        f"项目：{project.get('name','')}\n"
        f"任务：{run.get('id','')}\n"
        f"项目提示词：\n{prompt}\n\n"
        f"项目规则：\n{json.dumps(harness, ensure_ascii=False)}\n\n"
        f"待修复行：\n{json.dumps(rows, ensure_ascii=False, indent=2)}"
    )


def _normalize_model_fixes(payload: dict[str, Any], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    fixes_by_issue = {str(row["issue_id"]): row for row in rows}
    fixes_by_position = {(str(row["sheet"]), int(row["row"])): row for row in rows}
    fixes: list[dict[str, Any]] = []
    seen: set[tuple[str, int]] = set()
    for item in payload.get("fixes", []) if isinstance(payload.get("fixes"), list) else []:
        issue_id = str(item.get("issue_id") or "")
        sheet = str(item.get("sheet") or "")
        row_index = int(item.get("row") or 0)
        source = fixes_by_issue.get(issue_id) or fixes_by_position.get((sheet, row_index))
        if not source:
            continue
        key = (str(source["sheet"]), int(source["row"]))
        if key in seen:
            continue
        translation = str(item.get("translation") or "").strip()
        if not translation:
            continue
        seen.add(key)
        fixes.append(
            {
                "issue_id": source["issue_id"],
                "sheet": source["sheet"],
                "row": source["row"],
                "record_id": source.get("record_id", ""),
                "source_text": source.get("source_text", ""),
                "translation": translation,
                "note": str(item.get("note") or f"model_fix:{source['check_type']}").strip(),
                "rule_source": "model_fix",
            }
        )
    return fixes


def _apply_workbook_fixes(path: Path, fixes: list[dict[str, Any]], source_run_id: str) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    applied: list[dict[str, Any]] = []
    try:
        for fix in fixes:
            row_index = int(fix.get("row") or 0)
            if row_index < 2:
                raise ValueError(f"invalid workbook row: {row_index}")
            sheet_name = str(fix.get("sheet") or wb.sheetnames[0]).strip()
            if sheet_name not in wb.sheetnames:
                resolved = _resolve_workbook_row_for_issue(wb, None, row_index, fix.get("record_id"))
                if not resolved:
                    raise KeyError(f"sheet not found: {sheet_name}")
                ws, row_index = resolved
                sheet_name = ws.title
            else:
                ws = wb[sheet_name]
                resolved = _resolve_workbook_row_for_issue(wb, ws, row_index, fix.get("record_id"))
                if resolved:
                    ws, row_index = resolved
                    sheet_name = ws.title
            target_col = _first_col(_header_map(ws), ["en", "translation", "target", "译文", "英文"])
            if target_col is None:
                raise KeyError(f"target column not found in sheet: {sheet_name}")
            cell = ws.cell(row_index, target_col)
            previous = _cell_text(cell.value)
            translation = str(fix.get("translation") or "").strip()
            cell.value = translation
            applied.append(
                {
                    "id": db.new_id("fix"),
                    "source_run_id": source_run_id,
                    "issue_id": fix.get("issue_id") or "",
                    "sheet": sheet_name,
                    "row": row_index,
                    "column": target_col,
                    "previous_translation": previous,
                    "translation": translation,
                    "note": str(fix.get("note") or "").strip(),
                    "rule_source": str(fix.get("rule_source") or "manual_fix").strip() or "manual_fix",
                    "applied_at": db.now_iso(),
                }
            )
        wb.save(path)
    finally:
        wb.close()
    return applied


def _append_improvement_items(project_id: str, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    suggestions = list_improvements(project_id)
    suggestions.extend(items)
    path = project_dir(project_id) / "profile" / "improvement_suggestions.json"
    path.write_text(json.dumps(suggestions, ensure_ascii=False, indent=2), encoding="utf-8")
    return suggestions


def _header_map(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    if getattr(ws, "max_row", 0) < 1 or getattr(ws, "max_column", 0) < 1:
        return result
    try:
        header_row = ws[1]
    except IndexError:
        return result
    for cell in header_row:
        if cell.value is None:
            continue
        result[str(cell.value).strip().lower()] = int(cell.column)
    return result


def _first_col(headers: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        hit = headers.get(name.lower())
        if hit is not None:
            return hit
    return None


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value)


def _project_issue(
    sheet: str,
    row: int,
    check_type: str,
    message: str,
    translation: str,
    severity: str = "hard",
) -> dict[str, Any]:
    return {
        "source": "project_harness",
        "rule_source": "project_harness",
        "severity": severity,
        "sheet": sheet,
        "row": row,
        "check_type": check_type,
        "message": message,
        "translation": translation,
    }


def _improvement_item(category: str, run_id: str, title: str, detail: str) -> dict[str, Any]:
    return {
        "id": db.new_id("imp"),
        "category": category,
        "run_id": run_id,
        "title": title,
        "detail": detail,
        "status": "pending_review",
        "created_at": db.now_iso(),
    }

__all__ = [name for name in globals() if not name.startswith("__")]
