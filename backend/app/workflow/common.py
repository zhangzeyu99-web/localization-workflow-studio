from __future__ import annotations

import json
import re
import threading
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

from .. import db, translation_batches as _translation_batches
from ..config import DATA_ROOT
from ..languages import PROJECT_LANGUAGE_ORDER, SOURCE_HEADER_ALIASES, alt_aliases, target_aliases, visible_language_code

_AsyncTokenRateLimiter = _translation_batches.AsyncTokenRateLimiter
_build_batch_manifest = _translation_batches.build_batch_manifest
_cap_context_text = _translation_batches.cap_context_text
_estimate_row_tokens = _translation_batches.estimate_row_tokens
_estimate_text_tokens = _translation_batches.estimate_text_tokens
_load_or_create_batch_manifest = _translation_batches.load_or_create_batch_manifest
_manage_project_prompt_context = _translation_batches.manage_project_prompt_context
_manifest_matches_rows = _translation_batches.manifest_matches_rows
_project_context_summary = _translation_batches.project_context_summary
_provider_retry_delay_seconds = _translation_batches.provider_retry_delay_seconds

HARNESS_SCHEMA_VERSION = 1

# Per-project_id threading.Lock registry guarding the read-modify-write of
# this project's shared JSON state files (project_harness.json,
# improvement_suggestions.json). Multiple background jobs for the SAME
# project can run concurrently (e.g. a harness PATCH racing a running QA
# job's manual-fix write); without this lock, two concurrent
# read-modify-write cycles can silently lose one writer's update. A
# process-level dict + lock (not asyncio.Lock) is required because jobs run
# on separate threads, each with its own event loop.
_project_file_locks_guard = threading.Lock()
_project_file_locks: dict[str, threading.Lock] = {}


def _project_file_lock(project_id: str) -> threading.Lock:
    with _project_file_locks_guard:
        lock = _project_file_locks.get(project_id)
        if lock is None:
            lock = threading.Lock()
            _project_file_locks[project_id] = lock
        return lock
_GLOSSARY_EXTRACTOR_MODULE: Any | None = None
GLOBAL_HARNESS_CONTRACT: dict[str, Any] = {
    "source": "global_harness",
    "workpack": "translation_workpack.jsonl",
    "response_protocol": "jsonl:{id:int|str,translation:str}",
    "hard_gates": ["id", "placeholder", "tag", "newline", "input_fingerprint"],
    "qa_sources": ["workflow/localization/utils/quality_harness.py", "workflow/localization/fixtures/quality_regression.json"],
}

_CJK_RE = re.compile(r"[\u3400-\u9fff]")
RowId = int | str
LANGUAGE_ORDER = PROJECT_LANGUAGE_ORDER
TERM_REFERENCE_RULE = "术语译法以随附术语表、行级 term_hits 和译文归档命中为准。"
AUTO_LANGUAGE_TARGET_ALIASES = {code: tuple(target_aliases(code)) for code in LANGUAGE_ORDER}
AUTO_LANGUAGE_ALT_ALIASES = {code: tuple(alt_aliases(code)) for code in LANGUAGE_ORDER}
_GENERIC_TARGET_ALIASES = {"target", "translation", "译文"}
_TARGET_DETECTION_ALIASES: dict[str, set[str]] = {
    code: {
        *{alias.strip().lower() for alias in target_aliases(code) if alias.strip().lower() not in _GENERIC_TARGET_ALIASES},
        visible_language_code(code).lower(),
    }
    for code in LANGUAGE_ORDER
}
_STRUCTURAL_TARGET_HEADERS = {
    "id", "key", "编号", "序号",
    *{alias.lower() for alias in SOURCE_HEADER_ALIASES},
    "category", "type", "分类", "类别", "类型",
    "note", "notes", "comment", "备注",
    "target", "translation", "译文",
}


def _looks_like_untranslated_seed(text: str, language: str) -> bool:
    value = str(text or "")
    if not _CJK_RE.search(value):
        return False
    if language == "ja":
        return False
    return True


def _mkdir_with_parents_retry(path: Path, *, attempts: int = 3) -> None:
    """``Path.mkdir(parents=True, exist_ok=True)`` that tolerates a concurrent
    deletion of an ancestor directory racing with this call (e.g. a project
    delete, or test teardown wiping the data root, happening between the
    parent and child directory checks). Without this, a background job that
    is mid-write while its project directory is removed can raise
    ``FileNotFoundError`` even though ``exist_ok=True`` was requested.
    """
    last_error: OSError | None = None
    for _ in range(attempts):
        try:
            path.mkdir(parents=True, exist_ok=True)
            return
        except FileNotFoundError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error


def project_dir(project_id: str) -> Path:
    path = DATA_ROOT / "projects" / project_id
    _mkdir_with_parents_retry(path)
    for child in ("uploads", "profile", "glossary", "runs", "assets", "translations"):
        _mkdir_with_parents_retry(path / child)
    return path


def run_dir(run_id: str) -> Path:
    path = DATA_ROOT / "runs" / run_id
    _mkdir_with_parents_retry(path)
    return path


def default_project_harness(project: dict[str, Any]) -> dict[str, Any]:
    return {
        "schema_version": HARNESS_SCHEMA_VERSION,
        "source": "project_harness",
        "project_id": project["id"],
        "project_name": project["name"],
        "project_metadata": {},
        "style_guidance": "",
        "target_audience": "",
        "tone": "",
        "forbidden_translations": [],
        "fixed_terms": [],
        "hard_rules": [],
        "soft_rules": [],
        "reference_examples": [],
        "manual_fixes": [],
        "qa_summary": {},
        "updated_at": "",
    }


def project_harness_path(project_id: str) -> Path:
    return project_dir(project_id) / "profile" / "project_harness.json"


def _sanitize_harness(payload: dict[str, Any]) -> dict[str, Any]:
    text_fields = ("style_guidance", "target_audience", "tone")
    list_fields = (
        "forbidden_translations",
        "fixed_terms",
        "hard_rules",
        "soft_rules",
        "reference_examples",
        "manual_fixes",
    )
    for key in text_fields:
        payload[key] = str(payload.get(key) or "").strip()
    for key in list_fields:
        value = payload.get(key)
        payload[key] = value if isinstance(value, list) else []
    cleaned_fixed_terms = []
    for item in payload.get("fixed_terms", []):
        if not isinstance(item, dict):
            continue
        source = str(item.get("source") or "").strip()
        target = str(item.get("target") or "").strip()
        if not source or not target:
            continue
        if all(ch in {"?", "\ufffd"} or ch.isspace() for ch in source):
            continue
        cleaned_fixed_terms.append(item)
    payload["fixed_terms"] = cleaned_fixed_terms
    if not isinstance(payload.get("project_metadata"), dict):
        payload["project_metadata"] = {}
    if not isinstance(payload.get("qa_summary"), dict):
        payload["qa_summary"] = {}
    return payload


def read_project_harness(project_id: str) -> dict[str, Any]:
    project = db.get_project(project_id)
    default = default_project_harness(project)
    path = project_harness_path(project_id)
    if not path.exists():
        return default
    payload = json.loads(path.read_text(encoding="utf-8"))
    merged = {**default, **payload}
    merged["project_id"] = project_id
    merged["project_name"] = project["name"]
    merged["schema_version"] = HARNESS_SCHEMA_VERSION
    return _sanitize_harness(merged)


def update_project_harness(project_id: str, updater: Callable[[dict[str, Any]], dict[str, Any]]) -> dict[str, Any]:
    """Atomically read-modify-write this project's project_harness.json.

    ``updater`` receives the current harness dict and returns the fields to
    merge into it (same merge semantics as ``write_project_harness``'s
    ``updates`` argument: ``None`` values are ignored). The whole
    read+updater+write runs under this project's file lock so two
    concurrent writers (e.g. a harness PATCH racing a manual-fix job) can't
    silently clobber one another's update.
    """
    with _project_file_lock(project_id):
        payload = read_project_harness(project_id)
        updates = updater(payload)
        for key, value in updates.items():
            if value is not None:
                payload[key] = value
        payload["updated_at"] = db.now_iso()
        payload = _sanitize_harness(payload)
        path = project_harness_path(project_id)
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload


def write_project_harness(project_id: str, updates: dict[str, Any]) -> dict[str, Any]:
    return update_project_harness(project_id, lambda _current: updates)


def improvement_suggestions_path(project_id: str) -> Path:
    return project_dir(project_id) / "profile" / "improvement_suggestions.json"


def read_improvement_suggestions(project_id: str) -> list[dict[str, Any]]:
    path = improvement_suggestions_path(project_id)
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def update_improvement_suggestions(
    project_id: str, updater: Callable[[list[dict[str, Any]]], list[dict[str, Any]]]
) -> list[dict[str, Any]]:
    """Atomically read-modify-write this project's improvement_suggestions.json.

    Shares the same per-project lock as ``update_project_harness`` since
    both files are this project's shared mutable state and the goal is to
    stop any concurrent read-modify-write on either file from losing an
    update, not to maximize file-level parallelism.
    """
    with _project_file_lock(project_id):
        current = read_improvement_suggestions(project_id)
        updated = updater(current)
        path = improvement_suggestions_path(project_id)
        path.write_text(json.dumps(updated, ensure_ascii=False, indent=2), encoding="utf-8")
        return updated


def harness_overview(project_id: str) -> dict[str, Any]:
    return {
        "global_harness": GLOBAL_HARNESS_CONTRACT,
        "project_harness": read_project_harness(project_id),
        "boundary": (
            "global_harness stores reusable workflow contracts and gates; "
            "project_harness stores this project's private requirements only."
        ),
    }


def _workbook_text_stats(path: Path) -> dict[str, int]:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        lowered = [header.lower() for header in headers]
        source_idx = next((idx for idx, header in enumerate(lowered) if header in SOURCE_HEADER_ALIASES), None)
        target_idx = next(
            (idx for idx, header in enumerate(lowered) if header in {"en", "target", "translation", "译文", "英文"}),
            None,
        )
        source_rows = 0
        translated_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if source_idx is not None and row[source_idx] not in (None, ""):
                source_rows += 1
            if target_idx is not None and row[target_idx] not in (None, ""):
                translated_rows += 1
        wb.close()
        return {"source_rows": source_rows, "translated_rows": translated_rows}
    except Exception:
        return {"source_rows": 0, "translated_rows": 0}


def _language_assets_summary(project_id: str) -> str:
    candidates = [
        *db.list_artifacts(project_id=project_id, role="translation_workbook"),
        *db.list_artifacts(project_id=project_id, role="language_source"),
    ]
    best = {"source_rows": 0, "translated_rows": 0}
    for artifact in candidates:
        path = Path(str(artifact.get("path") or ""))
        if path.suffix.lower() != ".xlsx" or not path.exists():
            continue
        stats = _workbook_text_stats(path)
        if stats["source_rows"] > best["source_rows"]:
            best = stats
        elif stats["translated_rows"] > best["translated_rows"]:
            best["translated_rows"] = stats["translated_rows"]
    if best["source_rows"]:
        return f"{best['source_rows']} 条文本，已有英文 {best['translated_rows']} 条。"
    return "暂未统计语言表行数。"


def _project_material_labels(project_id: str) -> list[str]:
    labels: list[str] = []
    for role in ("glossary_source", "language_source", "translation_workbook"):
        for artifact in db.list_artifacts(project_id=project_id, role=role):
            label = str(artifact.get("label") or "").strip()
            if label and label not in labels:
                labels.append(label)
    return labels

# Export imported helpers and private compatibility names for sibling modules.
__all__ = [name for name in globals() if not name.startswith("__")]
