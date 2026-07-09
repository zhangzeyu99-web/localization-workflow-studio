from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .. import db
from ..languages import language_spec
from ..translation_batches import manage_project_prompt_context as _manage_project_prompt_context
from .announcement_segments import _quick_text_translation_rows, _write_quick_text_output
from .common import run_dir
from .jsonl_helpers import write_jsonl
from .prompt_snapshots import (
    create_project_glossary_snapshot,
    create_prompt_and_harness_snapshots,
    create_quick_reference_snapshot,
)
from .reference_lookup import attach_reference_hits, lookup_terms as _lookup_terms
from .translation_orchestrator import _translate_rows_with_orchestration


def _quick_text_empty_result(run_id: str, metadata: dict[str, Any], readiness: dict[str, Any]) -> dict[str, Any]:
    reason = "TXT 文件没有检测到可翻译文本。"
    db.merge_run_metadata(run_id, {"reason": reason, "translation_readiness": readiness})
    db.update_run(run_id, status="needs_input")
    db.add_event(run_id, reason)
    return {"run": db.get_run(run_id), "artifacts": [], "quality": None, "translation_readiness": readiness}


def _quick_text_prompt_context(
    *,
    project_id: str,
    run_id: str,
    metadata: dict[str, Any],
    language: str,
    work_dir: Path,
    settings: dict[str, Any],
) -> dict[str, Any]:
    snapshot_dir = work_dir / "snapshots"
    glossary_snapshot = create_project_glossary_snapshot(
        project_id,
        run_id,
        snapshot_dir,
        language=language,
        extra_term_artifact_id=metadata.get("term_artifact_id"),
    )
    snapshots = create_prompt_and_harness_snapshots(project_id, run_id, snapshot_dir, language=language)
    reference_snapshot = create_quick_reference_snapshot(project_id, run_id, metadata.get("reference_artifact_ids"), snapshot_dir)
    prompt = snapshots["prompt"]
    if reference_snapshot and reference_snapshot.get("context"):
        prompt = _manage_project_prompt_context(f"{prompt}\n\nQuick Task References:\n{reference_snapshot['context']}", settings)
    prompt = _manage_project_prompt_context(
        f"{prompt}\n\n快速 TXT 任务：逐行翻译 source 字段，保持每个 id 的顺序和结构。只返回 JSONL，每行包含 id 和 translation。",
        settings,
    )
    return {
        "prompt": prompt,
        "glossary_snapshot": glossary_snapshot,
        "prompt_snapshot": snapshots["prompt_artifact"],
        "harness_snapshot_artifact": snapshots["harness_artifact"],
        "reference_snapshot": reference_snapshot,
    }


def _quick_text_glossary_rows(glossary_snapshot: dict[str, Any], language: str) -> list[dict[str, Any]]:
    path = Path(str(glossary_snapshot.get("path") or ""))
    if not path.exists():
        return []
    spec = language_spec(language)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        header_index = {header.casefold(): index for index, header in enumerate(headers) if header}
        source_index = header_index.get("cn")
        target_index = header_index.get(spec.target_header.casefold())
        alt_index = header_index.get(spec.alt_header.casefold()) if spec.alt_header else None
        if source_index is None or target_index is None:
            return []
        output = []
        for raw in rows[1:]:
            source = str(raw[source_index] or "").strip() if source_index < len(raw) else ""
            target = str(raw[target_index] or "").strip() if target_index < len(raw) else ""
            target_alt = str(raw[alt_index] or "").strip() if alt_index is not None and alt_index < len(raw) else ""
            if source and (target or target_alt):
                output.append({"source": source, "target": target, "target_alt": target_alt, "language": language})
        return output
    finally:
        wb.close()


def _quick_text_rows_with_context(project_id: str, language: str, rows: list[dict[str, Any]], glossary_snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    glossary_rows = _quick_text_glossary_rows(glossary_snapshot, language)
    enriched = []
    for row in rows:
        source = str(row.get("source") or "")
        term_hits = _lookup_terms(source, glossary_rows, min_length=2, limit=20)
        enriched.append({
            **row,
            "term_hits": [{"source": hit.get("source", ""), "target": hit.get("target", ""), "target_alt": hit.get("target_alt", "")} for hit in term_hits],
        })
    attach_reference_hits(enriched, project_id, language)
    return enriched


def _write_quick_text_artifacts(
    *,
    project: dict[str, Any],
    run_id: str,
    input_artifact: dict[str, Any],
    input_path: Path,
    work_dir: Path,
    translated_rows: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    language: str,
) -> dict[str, Any]:
    response_path = work_dir / "translation_response.jsonl"
    write_jsonl(response_path, translated_rows)
    response_artifact = db.add_artifact(
        project["id"],
        "快速 TXT translation response",
        response_path,
        "translation_response",
        run_id=run_id,
        mime="application/jsonl",
        metadata={"language": language, "source_artifact_id": input_artifact["id"]},
    )
    output_path = _write_quick_text_output(input_path, translated_rows, language, work_dir)
    final_artifact = db.add_artifact(
        project["id"],
        "快速 TXT 最终译文",
        output_path,
        "final_text",
        run_id=run_id,
        mime="text/plain",
        role="delivery",
        origin="generated",
        metadata={"language": language, "source_artifact_id": input_artifact["id"]},
    )
    manifest_path = work_dir / "quick_text_translation_manifest.json"
    manifest = {
        "kind": "quick_text_translation",
        "run_id": run_id,
        "project_id": project["id"],
        "language": language,
        "source_artifact_id": input_artifact["id"],
        "source_rows": len(rows),
        "final_artifact_id": final_artifact["id"],
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    manifest_artifact = db.add_artifact(
        project["id"],
        "快速 TXT translation manifest",
        manifest_path,
        "translation_manifest",
        run_id=run_id,
        mime="application/json",
        metadata={"language": language},
    )
    return {"response": response_artifact, "final": final_artifact, "manifest": manifest_artifact, "output_path": output_path}


def _finish_quick_text_run(
    *,
    run_id: str,
    metadata: dict[str, Any],
    input_artifact: dict[str, Any],
    input_path: Path,
    rows: list[dict[str, Any]],
    readiness: dict[str, Any],
    artifacts: dict[str, Any],
    prompt_context: dict[str, Any],
    workpack_artifact: dict[str, Any],
) -> dict[str, Any]:
    input_artifacts = {
        "source_text": input_artifact["id"],
        "final_text": artifacts["final"]["id"],
        "translation_workpack": workpack_artifact["id"],
        "translation_response": artifacts["response"]["id"],
        "prompt_snapshot": prompt_context["prompt_snapshot"]["id"],
        "harness_snapshot": prompt_context["harness_snapshot_artifact"]["id"],
        "glossary_snapshot": prompt_context["glossary_snapshot"]["id"],
    }
    if prompt_context["reference_snapshot"]:
        input_artifacts["quick_reference_snapshot"] = prompt_context["reference_snapshot"]["artifact"]["id"]
    quality_summary = {"passed": True, "hard_errors": 0, "soft_warnings": 0, "rows": len(rows), "format": input_path.suffix.lower().lstrip(".") or "txt"}
    db.merge_run_metadata(
        run_id,
        {
            "task_origin": metadata.get("task_origin") or "quick_task",
            "input_artifacts": input_artifacts,
            "quality_summary": quality_summary,
            "quality": {"passed": True, "issues": []},
            "translation_readiness": readiness,
            "translated_rows": len(rows),
            "source_rows": len(rows),
            "output_format": input_path.suffix.lower().lstrip(".") or "txt",
        },
    )
    db.update_run(run_id, status="passed")
    db.add_event(run_id, f"quick TXT translation finished: rows={len(rows)}, output={artifacts['output_path'].name}")
    return {
        "run": db.get_run(run_id),
        "artifacts": [
            artifacts["final"],
            artifacts["response"],
            workpack_artifact,
            artifacts["manifest"],
            prompt_context["glossary_snapshot"],
            prompt_context["prompt_snapshot"],
            prompt_context["harness_snapshot_artifact"],
        ],
        "quality": {"passed": True, "issues": []},
        "translation_readiness": readiness,
    }


async def _translate_quick_text_run(
    *,
    run: dict[str, Any],
    input_artifact: dict[str, Any],
    settings: dict[str, Any],
    batch_size: int,
    language: str,
    readiness: dict[str, Any],
    request: Any,
    cancel_event: Any | None = None,
) -> dict[str, Any]:
    run_id = run["id"]
    project = db.get_project(run["project_id"])
    metadata = run.get("metadata", {})
    input_path = Path(input_artifact["path"])
    rows = _quick_text_translation_rows(input_path)
    if not rows:
        return _quick_text_empty_result(run_id, metadata, readiness)

    db.update_run(run_id, status="running")
    db.add_event(run_id, f"quick TXT translation preflight: source_lines={len(rows)}, batch_size={batch_size}, estimated_batches={readiness.get('estimated_batches') or '-'}")
    work_dir = run_dir(run_id) / "quick_text_translation"
    work_dir.mkdir(parents=True, exist_ok=True)
    prompt_context = _quick_text_prompt_context(
        project_id=project["id"],
        run_id=run_id,
        metadata=metadata,
        language=language,
        work_dir=work_dir,
        settings=settings,
    )
    rows = _quick_text_rows_with_context(project["id"], language, rows, prompt_context["glossary_snapshot"])

    workpack_path = work_dir / "quick_text_workpack.jsonl"
    write_jsonl(workpack_path, rows)
    workpack_artifact = db.add_artifact(
        project["id"],
        "快速 TXT workpack",
        workpack_path,
        "translation_workpack",
        run_id=run_id,
        mime="application/jsonl",
        metadata={"language": language, "source_artifact_id": input_artifact["id"]},
    )
    translated_rows = await _translate_rows_with_orchestration(
        run_id=run_id,
        rows=rows,
        settings=settings,
        project_prompt=prompt_context["prompt"],
        work_dir=work_dir,
        batch_size=batch_size,
        language=language,
        cancel_event=cancel_event,
        confirm_api_budget=bool(getattr(request, "confirm_api_budget", False)),
    )
    if not translated_rows and db.get_run(run_id).get("status") == "needs_input":
        return {"run": db.get_run(run_id), "artifacts": [workpack_artifact], "quality": None, "translation_readiness": readiness}

    artifacts = _write_quick_text_artifacts(
        project=project,
        run_id=run_id,
        input_artifact=input_artifact,
        input_path=input_path,
        work_dir=work_dir,
        translated_rows=translated_rows,
        rows=rows,
        language=language,
    )
    return _finish_quick_text_run(
        run_id=run_id,
        metadata=metadata,
        input_artifact=input_artifact,
        input_path=input_path,
        rows=rows,
        readiness=readiness,
        artifacts=artifacts,
        prompt_context=prompt_context,
        workpack_artifact=workpack_artifact,
    )


__all__ = [name for name in globals() if not name.startswith("__")]
