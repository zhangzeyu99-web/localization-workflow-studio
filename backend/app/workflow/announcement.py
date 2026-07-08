from __future__ import annotations

import asyncio
import hashlib
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import Workbook

from .. import db
from ..config import REAL_PROVIDERS, load_settings, normalize_provider_name
from ..languages import language_spec, require_supported_language
from ..translation_batches import manage_project_prompt_context as _manage_project_prompt_context
from .announcement_ai import (
    _announcement_terms_languages,
    _announcement_terms_validation,
    _apply_announcement_ai_supplement,
    _filter_announcement_terms_languages,
    _normalize_announcement_terms_payload,
    _save_announcement_terms,
    _write_announcement_terms_workbook,
)
from .announcement_outputs import (
    _announcement_delivery_base_name,
    _announcement_response_artifact_map,
    _announcement_task_source_stem,
    _announcement_translation_prompt,
    _announcement_workpack_rows,
    _artifact_source_stem,
    _import_announcement_response_into_workbook,
    _mime_for_path,
    _project_archive_by_language,
    _read_announcement_translation_workbook,
    _repair_announcement_translation_workbook,
    _safe_file_stem,
    _validate_announcement_translation_rows,
    _write_announcement_outputs,
    _write_announcement_qa_summary,
    _write_announcement_translation_workbook,
)
from .announcement_segments import (
    _announcement_constraint_rows,
    _announcement_language_constraint_summary,
    _announcement_source_format,
    _announcement_source_manifest,
    _announcement_task_segments,
    _announcement_task_source_text,
    _detect_announcement_constraint_languages,
    _detect_language_columns,
    _normalize_announcement_languages,
    _read_language_table_rows,
    _select_announcement_constraint_rows,
)
from .announcement_shared import (
    ANNOUNCEMENT_STEP,
    _announcement_task_metadata,
)
from .common import project_dir, run_dir
from .delivery import DELIVERED_WITH_ISSUES_SOURCE_TYPE
from .jsonl_helpers import read_jsonl, write_jsonl
from .materials import _compact_lookup_text, _read_lookup_material_text
from .naming import _today_stamp, _visible_language_code
from .prompt_snapshots import create_prompt_and_harness_snapshots
from .reference_lookup import lookup_terms as _lookup_terms, lookup_translation_entries as _lookup_translation_entries
from .subprocess_runner import user_facing_error
from .table_helpers import _wide_source_key
from .translation import _translate_rows_with_orchestration


def _announcement_prompt_context(project: dict[str, Any], language: str, terms: list[dict[str, Any]], translations: list[dict[str, Any]]) -> str:
    spec = language_spec(language)
    lines = [
        "公告翻译检索上下文（不含正文译文）",
        f"项目：{project['name']}",
        f"目标语言：{spec.prompt_name} / {spec.target_header}",
        "用途：供下游长文本翻译工作流使用；必须优先遵守项目术语和 QA 通过译文参考。",
        "",
        "【命中的项目术语】",
    ]
    if terms:
        for term in terms:
            alt = f" / {term['target_alt']}" if term.get("target_alt") else ""
            note = f"（{term['note']}）" if term.get("note") else ""
            lines.append(f"- {term['source']} => {term['target']}{alt}{note}")
    else:
        lines.append("- 无命中；当前语言缺少术语约束或公告文本未命中现有术语。")
    lines.extend(["", "【命中的 QA/归档译文参考】"])
    if translations:
        for entry in translations:
            alt = f" / {entry['target_alt']}" if entry.get("target_alt") else ""
            meta = f"{entry.get('sheet') or 'Archive'}:{entry.get('row_number') or ''}".rstrip(":")
            lines.append(f"- {entry['source']} => {entry['target']}{alt}（{meta} / {entry.get('source_type') or 'archive'}）")
    else:
        lines.append("- 无命中；下游翻译需仅依赖项目提示词和通用语言质量要求。")
    lines.extend(
        [
            "",
            "【硬性要求】",
            "- 不修改变量、占位符、HTML/富文本标签、数字、专名和换行结构。",
            "- 术语命中时使用上述译法；同一概念在整篇公告中保持一致。",
            "- 本检索包不代表公告正文已翻译，正文翻译和 QA 由后续工作流完成。",
        ]
    )
    return "\n".join(lines).strip() + "\n"


def _write_announcement_lookup_workbook(path: Path, project: dict[str, Any], language: str, text: str, terms: list[dict[str, Any]], translations: list[dict[str, Any]], prompt_context: str) -> None:
    spec = language_spec(language)
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    overview.append(["Field", "Value"])
    overview.append(["Project", project["name"]])
    overview.append(["Project ID", project["id"]])
    overview.append(["Language", language])
    overview.append(["Target header", spec.target_header])
    overview.append(["Text chars", len(text)])
    overview.append(["Matched terms", len(terms)])
    overview.append(["Matched translations", len(translations)])

    term_sheet = wb.create_sheet("MatchedTerms")
    term_sheet.append(["CN", spec.target_header, *([spec.alt_header] if spec.alt_header else []), "分类", "备注", "first_position", "hit_count", "term_id"])
    for term in terms:
        row = [term["source"], term["target"]]
        if spec.alt_header:
            row.append(term.get("target_alt", ""))
        row.extend([term.get("category", ""), term.get("note", ""), term["first_position"], term["hit_count"], term.get("id", "")])
        term_sheet.append(row)

    translation_sheet = wb.create_sheet("MatchedTranslations")
    translation_sheet.append(["ID", "CN", spec.target_header, *([spec.alt_header] if spec.alt_header else []), "source_type", "sheet", "row_number", "first_position", "hit_count", "entry_id"])
    for entry in translations:
        row = [entry.get("entry_key", ""), entry["source"], entry["target"]]
        if spec.alt_header:
            row.append(entry.get("target_alt", ""))
        row.extend([entry.get("source_type", ""), entry.get("sheet", ""), entry.get("row_number", 0), entry["first_position"], entry["hit_count"], entry.get("id", "")])
        translation_sheet.append(row)

    context_sheet = wb.create_sheet("PromptContext")
    context_sheet.append(["Prompt context"])
    for line in prompt_context.splitlines():
        context_sheet.append([line])
    wb.save(path)
    wb.close()


def run_announcement_lookup(project_id: str, request: Any) -> dict[str, Any]:
    project = db.get_project(project_id)
    language = require_supported_language(getattr(request, "language", "en") or "en")
    material_ids = list(getattr(request, "material_artifact_ids", []) or [])
    direct_text = str(getattr(request, "text", "") or "")
    if not material_ids and not direct_text.strip():
        raise ValueError("announcement lookup requires material_artifact_ids or text")

    text_parts: list[str] = []
    materials: list[dict[str, Any]] = []
    if direct_text.strip():
        text_parts.append(direct_text)
    for artifact_id in material_ids:
        artifact = db.get_artifact(artifact_id)
        if artifact["project_id"] != project_id:
            raise KeyError(artifact_id)
        path = Path(artifact["path"])
        material_text = _read_lookup_material_text(path) if path.exists() else ""
        text_parts.append(material_text)
        materials.append({"id": artifact["id"], "label": artifact.get("label", ""), "kind": artifact.get("kind", ""), "chars": len(material_text)})

    text = _compact_lookup_text("\n".join(text_parts))
    if not text:
        raise ValueError("announcement lookup text is empty")

    max_terms = max(0, min(int(getattr(request, "max_terms", 300) or 300), 1000))
    max_translation_rows = max(0, min(int(getattr(request, "max_translation_rows", 300) or 300), 1000))
    min_term_length = max(1, int(getattr(request, "min_term_length", 2) or 2))
    min_translation_length = max(1, int(getattr(request, "min_translation_length", 4) or 4))

    run = db.insert_run(
        project_id,
        kind="announcement_lookup",
        language=language,
        metadata={
            "request": {
                "material_artifact_ids": material_ids,
                "has_inline_text": bool(direct_text.strip()),
                "language": language,
                "include_glossary": bool(getattr(request, "include_glossary", True)),
                "include_translation_archive": bool(getattr(request, "include_translation_archive", True)),
            }
        },
    )
    db.add_event(run["id"], "announcement lookup started")
    try:
        glossary_rows = db.list_glossary_terms(project_id, language=language) if bool(getattr(request, "include_glossary", True)) else []
        archive_rows = db.list_translation_entries(project_id, language=language) if bool(getattr(request, "include_translation_archive", True)) else []
        matched_terms = _lookup_terms(text, glossary_rows, min_length=min_term_length, limit=max_terms)
        matched_translations = _lookup_translation_entries(text, archive_rows, min_length=min_translation_length, limit=max_translation_rows)
        text_fingerprint = hashlib.sha256(text.encode("utf-8")).hexdigest()
        constraint_status = "available" if matched_terms or matched_translations else "missing"
        summary = {
            "language": language,
            "text_chars": len(text),
            "text_fingerprint": text_fingerprint,
            "materials": len(materials),
            "matched_terms": len(matched_terms),
            "matched_translations": len(matched_translations),
            "constraint_status": constraint_status,
        }
        prompt_context = _announcement_prompt_context(project, language, matched_terms, matched_translations)
        manifest = {
            "kind": "announcement_lookup",
            "project_id": project_id,
            "project_name": project["name"],
            "language": language,
            "materials": materials,
            "text_fingerprint": text_fingerprint,
            "limits": {
                "min_term_length": min_term_length,
                "min_translation_length": min_translation_length,
                "max_terms": max_terms,
                "max_translation_rows": max_translation_rows,
            },
            "summary": summary,
            "matched_terms": matched_terms,
            "matched_translations": matched_translations,
        }

        output = run_dir(run["id"]) / "announcement_lookup"
        output.mkdir(parents=True, exist_ok=True)
        lang_code = _visible_language_code(language)
        workbook_path = output / f"announcement_lookup_{lang_code}.xlsx"
        manifest_path = output / f"announcement_lookup_manifest_{lang_code}.json"
        prompt_path = output / f"announcement_lookup_prompt_context_{lang_code}.txt"
        _write_announcement_lookup_workbook(workbook_path, project, language, text, matched_terms, matched_translations, prompt_context)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
        prompt_path.write_text(prompt_context, encoding="utf-8")

        artifact_metadata = {"language": language, "summary": summary, "text_fingerprint": text_fingerprint}
        artifacts = [
            db.add_artifact(project_id, f"Announcement lookup workbook ({lang_code})", workbook_path, "announcement_lookup_workbook", run_id=run["id"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", metadata=artifact_metadata),
            db.add_artifact(project_id, f"Announcement lookup manifest ({lang_code})", manifest_path, "announcement_lookup_manifest", run_id=run["id"], mime="application/json", metadata=artifact_metadata),
            db.add_artifact(project_id, f"Announcement lookup prompt context ({lang_code})", prompt_path, "announcement_lookup_prompt_context", run_id=run["id"], mime="text/plain", metadata=artifact_metadata),
        ]
        db.add_event(run["id"], f"announcement lookup matched terms={len(matched_terms)} translations={len(matched_translations)}")
        db.merge_run_metadata(run["id"], {"summary": summary, "manifest_path": str(manifest_path)})
        run = db.update_run(run["id"], status="passed")
        return {"run": run, "summary": summary, "artifacts": artifacts, "manifest": manifest}
    except Exception as exc:
        friendly = user_facing_error(exc)
        db.add_event(run["id"], f"announcement lookup failed: {friendly}", level="error")
        db.merge_run_metadata(run["id"], {"error": friendly})
        db.update_run(run["id"], status="failed")
        raise



def list_announcement_tasks(project_id: str) -> list[dict[str, Any]]:
    db.get_project(project_id)
    return [_hydrate_announcement_task(task) for task in db.list_announcement_tasks(project_id)]


def get_announcement_task(task_id: str) -> dict[str, Any]:
    return _hydrate_announcement_task(db.get_announcement_task(task_id))


def cancel_announcement_task(task_id: str) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    metadata["canceled_at"] = db.now_iso()
    task = db.update_announcement_task(
        task_id,
        status="canceled",
        current_step=task.get("current_step") or 1,
        metadata=metadata,
    )
    for item in task.get("languages") or []:
        lang_meta = dict(item.get("metadata") or {})
        lang_meta["canceled_at"] = metadata["canceled_at"]
        db.upsert_announcement_task_language(
            task_id,
            task["project_id"],
            item["language"],
            status="canceled",
            current_step=item.get("current_step") or task.get("current_step") or 1,
            metadata=lang_meta,
        )
    return {"task": _hydrate_announcement_task(db.get_announcement_task(task_id))}


def cancel_announcement_translation_task(task_id: str) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    metadata["translation_cancel_requested_at"] = db.now_iso()
    metadata["reason"] = "announcement_translation_canceled"
    task = db.update_announcement_task(
        task_id,
        status="prepared",
        current_step=ANNOUNCEMENT_STEP["translate"],
        metadata=metadata,
    )
    for item in task.get("languages") or []:
        if item.get("status") in {"queued", "running"}:
            lang_meta = dict(item.get("metadata") or {})
            lang_meta["translation_cancel_requested_at"] = metadata["translation_cancel_requested_at"]
            db.upsert_announcement_task_language(
                task_id,
                task["project_id"],
                str(item["language"]),
                status="prepared",
                current_step=ANNOUNCEMENT_STEP["translate"],
                metadata=lang_meta,
            )
    return {"task": _hydrate_announcement_task(db.get_announcement_task(task_id))}


def create_announcement_task(project_id: str, request: Any) -> dict[str, Any]:
    project = db.get_project(project_id)
    source_artifact_id = str(getattr(request, "source_artifact_id", "") or "").strip()
    text = str(getattr(request, "text", "") or "")
    if not source_artifact_id and text.strip():
        source_artifact_id = _create_inline_announcement_source(project_id, text, getattr(request, "title", "") or "announcement")["id"]
    if not source_artifact_id:
        raise ValueError("announcement task requires source_artifact_id or text")
    source_artifact = db.get_artifact(source_artifact_id)
    if source_artifact["project_id"] != project_id:
        raise KeyError(source_artifact_id)
    source_format = _announcement_source_format(Path(source_artifact["path"]))
    if source_format not in {"docx", "txt", "xlsx"}:
        raise ValueError("announcement source must be DOCX, TXT, or XLSX")

    metadata = {
        "project_name": project["name"],
        "output_policy": str(getattr(request, "output_policy", "same_format") or "same_format"),
        "language_table_artifact_ids": list(getattr(request, "language_table_artifact_ids", []) or []),
        "constraint_artifact_ids": list(getattr(request, "constraint_artifact_ids", []) or []),
        "include_project_archive": bool(getattr(request, "include_project_archive", True)),
        "source": _announcement_source_manifest(source_artifact),
    }
    detected_languages = _detect_announcement_constraint_languages(project_id, metadata)
    requested_languages = _normalize_announcement_languages(getattr(request, "languages", []) or [], fallback=detected_languages)
    task = db.insert_announcement_task(
        project_id,
        {
            "title": str(getattr(request, "title", "") or source_artifact.get("label") or "公告翻译").strip(),
            "source_artifact_id": source_artifact_id,
            "source_format": source_format,
            "selected_languages": requested_languages,
            "status": "source_ready",
            "current_step": ANNOUNCEMENT_STEP["constraints"],
            "metadata": {**metadata, "detected_languages": detected_languages},
        },
    )
    return _hydrate_announcement_task(task)


def inspect_announcement_constraints(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    original_metadata = _announcement_task_metadata(task)
    metadata = _merge_announcement_constraint_request(original_metadata, request)
    detected = _detect_announcement_constraint_languages(task["project_id"], metadata)
    selected = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=detected)
    metadata["detected_languages"] = detected
    metadata["language_constraints"] = _announcement_language_constraint_summary(task["project_id"], metadata, selected)
    confirmed = bool(getattr(request, "confirm_languages", False))
    next_step = ANNOUNCEMENT_STEP["terms"] if confirmed else ANNOUNCEMENT_STEP["languages"]
    next_status = "languages_ready" if confirmed else "constraints_ready" if detected or selected else "missing_constraints"
    current_step = int(task.get("current_step") or 0)
    if current_step > next_step:
        return {"task": _hydrate_announcement_task(task), "detected_languages": detected, "selected_languages": task.get("selected_languages") or selected, "constraints": metadata["language_constraints"]}
    if (
        current_step == next_step
        and str(task.get("status") or "") == next_status
        and list(task.get("selected_languages") or []) == selected
        and _same_announcement_constraint_inputs(original_metadata, metadata)
        and (original_metadata.get("language_constraints") or {}) == metadata["language_constraints"]
    ):
        return {"task": _hydrate_announcement_task(task), "detected_languages": detected, "selected_languages": selected, "constraints": metadata["language_constraints"]}
    task = db.update_announcement_task(
        task_id,
        status=next_status,
        current_step=next_step,
        selected_languages=selected,
        metadata=metadata,
    )
    for language in selected:
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status=next_status, current_step=next_step)
    return {"task": _hydrate_announcement_task(task), "detected_languages": detected, "selected_languages": selected, "constraints": metadata["language_constraints"]}


def _same_announcement_constraint_inputs(left: dict[str, Any], right: dict[str, Any]) -> bool:
    def ids(value: Any) -> list[str]:
        result: list[str] = []
        for item in list(value or []):
            text = str(item or "").strip()
            if text and text not in result:
                result.append(text)
        return result

    return (
        ids(left.get("language_table_artifact_ids")) == ids(right.get("language_table_artifact_ids"))
        and ids(left.get("constraint_artifact_ids")) == ids(right.get("constraint_artifact_ids"))
        and bool(left.get("include_project_archive", True)) == bool(right.get("include_project_archive", True))
    )


def extract_announcement_terms(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    project_id = task["project_id"]
    metadata = _merge_announcement_constraint_request(_announcement_task_metadata(task), request)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("detected_languages") or [])
    source_text = _announcement_task_source_text(task)
    if not source_text:
        raise ValueError("announcement source text is empty")
    min_hit = max(1, int(getattr(request, "announcement_min_hit", 1) or 1))
    candidates = _announcement_constraint_rows(project_id, metadata, languages)
    rows = _select_announcement_constraint_rows(source_text, candidates, languages, min_hit=min_hit)

    run = db.insert_run(project_id, kind="announcement_terms", language=languages[0] if languages else "en", metadata={"task_id": task_id, "languages": languages})
    db.update_run(run["id"], status="running")
    output = run_dir(run["id"]) / "announcement_terms"
    output.mkdir(parents=True, exist_ok=True)
    base = _announcement_task_source_stem(task)
    stamp = _today_stamp()
    workbook_path = output / f"{base}_announcement_terms_{stamp}.xlsx"
    manifest_path = output / f"{base}_announcement_terms_manifest_{stamp}.json"
    validation_path = output / f"{base}_announcement_terms_validation_{stamp}.md"
    rows, ai_summary = _apply_announcement_ai_supplement(
        project_id=project_id,
        output_dir=output,
        base_name=base,
        source_text=source_text,
        rows=rows,
        candidates=candidates,
        languages=languages,
        request=request,
        project_name=db.get_project(project_id).get("name", ""),
    )
    _write_announcement_terms_workbook(workbook_path, rows, languages)
    summary = {"terms": len(rows), "languages": languages, "source_chars": len(source_text)}
    if ai_summary:
        summary["ai_supplement"] = {
            key: ai_summary[key]
            for key in ("enabled", "response_artifact_id", "provider", "provider_status", "provider_error", "term_count", "added_to_main", "report_only", "project_name_translation_missing")
        }
    manifest = {"kind": "announcement_terms", "task_id": task_id, "project_id": project_id, "languages": languages, "summary": summary, "terms": rows}
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    validation_path.write_text(_announcement_terms_validation(summary, rows, languages), encoding="utf-8")
    artifacts = [
        db.add_artifact(project_id, "公告术语表", workbook_path, "announcement_terms_workbook", run_id=run["id"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", metadata={"task_id": task_id, "languages": languages}),
        db.add_artifact(project_id, "公告术语 manifest", manifest_path, "announcement_terms_manifest", run_id=run["id"], mime="application/json", metadata={"task_id": task_id, "languages": languages}),
        db.add_artifact(project_id, "公告术语 validation", validation_path, "announcement_terms_validation", run_id=run["id"], mime="text/markdown", metadata={"task_id": task_id, "languages": languages}),
    ]
    if ai_summary:
        packet_artifact = db.add_artifact(project_id, "公告 AI 补充包", Path(ai_summary["packet_path"]), "announcement_ai_supplement_packet", run_id=run["id"], mime="application/json", metadata={"task_id": task_id, "languages": languages})
        response_artifact = None
        if ai_summary.get("response_path"):
            response_artifact = db.add_artifact(project_id, "公告 AI 补充响应", Path(ai_summary["response_path"]), "announcement_ai_supplement_response", run_id=run["id"], mime="application/json", metadata={"task_id": task_id, "languages": languages, "provider": ai_summary.get("provider", "")})
        report_artifact = db.add_artifact(project_id, "公告 AI 补充报告", Path(ai_summary["report_path"]), "announcement_ai_supplement_report", run_id=run["id"], mime="text/markdown", metadata={"task_id": task_id, "languages": languages})
        artifacts.extend([artifact for artifact in (packet_artifact, response_artifact, report_artifact) if artifact])
        summary["ai_supplement"]["packet_artifact_id"] = packet_artifact["id"]
        if response_artifact:
            summary["ai_supplement"]["response_artifact_id"] = response_artifact["id"]
        summary["ai_supplement"]["report_artifact_id"] = report_artifact["id"]
    metadata.update({"languages": languages, "terms": rows, "terms_artifact_id": artifacts[0]["id"], "terms_manifest_artifact_id": artifacts[1]["id"], "terms_validation_artifact_id": artifacts[2]["id"], "terms_summary": summary})
    if ai_summary:
        metadata["ai_supplement"] = summary["ai_supplement"]
    task = db.update_announcement_task(task_id, status="terms_ready", current_step=ANNOUNCEMENT_STEP["lookup"], selected_languages=languages, metadata=metadata)
    for language in languages:
        missing = sum(1 for row in rows if not str((row.get("translations") or {}).get(language) or "").strip())
        db.upsert_announcement_task_language(task_id, project_id, language, status="terms_ready", current_step=ANNOUNCEMENT_STEP["lookup"], metadata={"terms": len(rows), "missing_terms": missing})
    db.merge_run_metadata(run["id"], {"summary": summary, "task_id": task_id})
    db.update_run(run["id"], status="passed")
    return {"task": _hydrate_announcement_task(task), "run": db.get_run(run["id"]), "summary": summary, "artifacts": artifacts, "manifest": manifest}


def import_announcement_terms(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    project_id = task["project_id"]
    metadata = _announcement_task_metadata(task)
    source_artifact_id = str(getattr(request, "terms_artifact_id", "") or "").strip()
    requested_languages = list(getattr(request, "languages", []) or [])
    raw_terms = list(getattr(request, "terms", []) or [])

    rows: list[dict[str, Any]] = []
    detected_languages: list[str] = []
    if source_artifact_id:
        artifact = db.get_artifact(source_artifact_id)
        if artifact["project_id"] != project_id:
            raise KeyError(source_artifact_id)
        detected_languages = _detect_language_columns(Path(artifact["path"]))
        languages = _normalize_announcement_languages(requested_languages, fallback=detected_languages or task.get("selected_languages") or metadata.get("languages") or [])
        rows = _read_language_table_rows(Path(artifact["path"]), languages)
        metadata["imported_terms_artifact_id"] = source_artifact_id
    else:
        rows = _normalize_announcement_terms_payload(raw_terms)
        detected_languages = _announcement_terms_languages(rows)
        languages = _normalize_announcement_languages(requested_languages, fallback=detected_languages or task.get("selected_languages") or metadata.get("languages") or [])
        rows = _filter_announcement_terms_languages(rows, languages)

    if not rows:
        raise ValueError("announcement terms are empty")
    if not languages:
        languages = _announcement_terms_languages(rows)
    if not languages:
        raise ValueError("announcement terms contain no target languages")

    return _save_announcement_terms(task_id, rows, languages, run_kind="announcement_terms_import")


def lookup_announcement_translations(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    project = db.get_project(task["project_id"])
    metadata = _merge_announcement_constraint_request(_announcement_task_metadata(task), request)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    terms = list(metadata.get("terms") or [])
    if not terms:
        raise ValueError("extract terms before lookup")
    archive_by_language = _project_archive_by_language(task["project_id"], languages)
    lookup: dict[str, Any] = {}
    run = db.insert_run(task["project_id"], kind="announcement_translation_lookup", language=languages[0] if languages else "en", metadata={"task_id": task_id, "languages": languages})
    output = run_dir(run["id"]) / "announcement_lookup"
    output.mkdir(parents=True, exist_ok=True)
    artifacts: list[dict[str, Any]] = []
    for language in languages:
        rows = []
        for term in terms:
            source = str(term.get("source") or "").strip()
            archive_entry = archive_by_language.get(language, {}).get(_wide_source_key(source))
            table_target = str((term.get("translations") or {}).get(language) or "").strip()
            target = str((archive_entry or {}).get("target") or "").strip() or table_target
            rows.append({**term, "language": language, "target": target, "source_type": "qa_archive" if archive_entry else "language_table" if table_target else "missing"})
        missing = [row for row in rows if not str(row.get("target") or "").strip()]
        prompt_context = _announcement_prompt_context(
            project,
            language,
            [
                {"source": row.get("source", ""), "target": row.get("target", ""), "target_alt": "", "category": "", "note": f"{row.get('hit_count', 0)} hit(s)"}
                for row in rows
                if row.get("target")
            ],
            [],
        )
        context_path = output / f"{_announcement_task_source_stem(task)}_prompt_context_{_visible_language_code(language)}.txt"
        context_path.write_text(prompt_context, encoding="utf-8")
        artifacts.append(db.add_artifact(task["project_id"], f"公告 prompt context ({_visible_language_code(language)})", context_path, "announcement_lookup_prompt_context", run_id=run["id"], mime="text/plain", metadata={"task_id": task_id, "language": language}))
        lookup[language] = {"terms": rows, "missing_terms": [{"source": row.get("source", ""), "id": row.get("id", "")} for row in missing], "prompt_context_artifact_id": artifacts[-1]["id"]}
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status="lookup_ready", current_step=ANNOUNCEMENT_STEP["prepare"], metadata={"terms": len(rows), "missing_terms": len(missing), "prompt_context_artifact_id": artifacts[-1]["id"]})
    summary = {"languages": languages, "terms": len(terms), "missing_terms": sum(len(lookup[language]["missing_terms"]) for language in languages)}
    manifest_path = output / "announcement_lookup_manifest.json"
    manifest = {"kind": "announcement_translation_lookup", "task_id": task_id, "project_id": task["project_id"], "summary": summary, "lookup": lookup}
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    artifacts.append(db.add_artifact(task["project_id"], "公告译文反查 manifest", manifest_path, "announcement_lookup_manifest", run_id=run["id"], mime="application/json", metadata={"task_id": task_id, "languages": languages}))
    metadata.update({"lookup": lookup, "lookup_manifest_artifact_id": artifacts[-1]["id"], "lookup_summary": summary})
    task = db.update_announcement_task(task_id, status="lookup_ready", current_step=ANNOUNCEMENT_STEP["prepare"], selected_languages=languages, metadata=metadata)
    db.merge_run_metadata(run["id"], {"summary": summary})
    db.update_run(run["id"], status="passed")
    return {"task": _hydrate_announcement_task(task), "run": db.get_run(run["id"]), "summary": summary, "artifacts": artifacts, "manifest": manifest}


def prepare_announcement_translation(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    project = db.get_project(task["project_id"])
    metadata = _announcement_task_metadata(task)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    if not languages:
        raise ValueError("select at least one target language")
    segments = _announcement_task_segments(task)
    if not segments:
        raise ValueError("announcement source contains no translatable text")
    lookup = metadata.get("lookup") or {}
    run = db.insert_run(task["project_id"], kind="announcement_prepare", language=languages[0], metadata={"task_id": task_id, "languages": languages})
    output = run_dir(run["id"]) / "announcement_prepare"
    output.mkdir(parents=True, exist_ok=True)
    source_stem = _announcement_task_source_stem(task)
    workbook_path = output / f"{source_stem}_announcement_translation_workbook.xlsx"
    manifest_path = output / "announcement_manifest.json"
    _write_announcement_translation_workbook(workbook_path, task, segments, languages, lookup)
    manifest = {"kind": "announcement_prepare", "task_id": task_id, "project_id": task["project_id"], "source_format": task["source_format"], "languages": languages, "segments": segments}
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    artifacts = [
        db.add_artifact(task["project_id"], "公告翻译中转表", workbook_path, "announcement_translation_workbook", run_id=run["id"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", metadata={"task_id": task_id, "languages": languages}),
        db.add_artifact(task["project_id"], "公告翻译 manifest", manifest_path, "announcement_docx_manifest", run_id=run["id"], mime="application/json", metadata={"task_id": task_id, "languages": languages}),
    ]
    workpacks: dict[str, str] = {}
    prompts: dict[str, str] = {}
    for language in languages:
        prompt_snapshot = create_prompt_and_harness_snapshots(task["project_id"], run["id"], output / "snapshots" / language, language=language)
        raw_prompt = _announcement_translation_prompt(project, language, prompt_snapshot["prompt"], lookup.get(language, {}))
        prompt = _manage_project_prompt_context(raw_prompt, load_settings())
        lang_code = _visible_language_code(language)
        prompt_path = output / f"{source_stem}_prompt_{lang_code}.txt"
        prompt_path.write_text(prompt, encoding="utf-8")
        prompts[language] = db.add_artifact(task["project_id"], f"公告翻译提示词 ({lang_code})", prompt_path, "prompt_snapshot", run_id=run["id"], mime="text/plain", metadata={"task_id": task_id, "language": language})["id"]
        workpack_path = output / f"{source_stem}_workpack_{lang_code}.jsonl"
        write_jsonl(workpack_path, _announcement_workpack_rows(segments, language, lookup))
        workpack_artifact = db.add_artifact(task["project_id"], f"公告 workpack ({lang_code})", workpack_path, "announcement_workpack", run_id=run["id"], mime="application/jsonl", metadata={"task_id": task_id, "language": language})
        workpacks[language] = workpack_artifact["id"]
        artifacts.append(workpack_artifact)
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status="prepared", current_step=ANNOUNCEMENT_STEP["translate"], metadata={"workpack_artifact_id": workpack_artifact["id"], "prompt_artifact_id": prompts[language], "translation_workbook_artifact_id": artifacts[0]["id"]})
    metadata.update({"segments": segments, "prepare_run_id": run["id"], "translation_workbook_artifact_id": artifacts[0]["id"], "manifest_artifact_id": artifacts[1]["id"], "workpack_artifact_ids": workpacks, "prompt_artifact_ids": prompts})
    task = db.update_announcement_task(task_id, status="prepared", current_step=ANNOUNCEMENT_STEP["translate"], selected_languages=languages, metadata=metadata)
    db.merge_run_metadata(run["id"], {"summary": {"segments": len(segments), "languages": languages}})
    db.update_run(run["id"], status="passed")
    return {"task": _hydrate_announcement_task(task), "run": db.get_run(run["id"]), "summary": {"segments": len(segments), "languages": languages}, "artifacts": artifacts, "manifest": manifest}


def translate_announcement_task(task_id: str, request: Any, cancel_event: Any | None = None) -> dict[str, Any]:
    return asyncio.run(_translate_announcement_task(task_id, request, cancel_event=cancel_event))


async def _translate_announcement_task(task_id: str, request: Any, cancel_event: Any | None = None) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    if not metadata.get("translation_workbook_artifact_id") or not metadata.get("workpack_artifact_ids"):
        raise ValueError("prepare announcement translation before AI translation")
    settings = load_settings()
    provider = normalize_provider_name(getattr(request, "provider", None) or settings.get("provider"))
    if getattr(request, "provider", None):
        settings["provider"] = provider
    if getattr(request, "protocol", None):
        settings["protocol"] = request.protocol
    if provider in REAL_PROVIDERS and not settings.get("api_key"):
        for language in languages:
            db.upsert_announcement_task_language(task_id, task["project_id"], language, status="awaiting_ai_response", current_step=ANNOUNCEMENT_STEP["translate"])
        return {"task": _hydrate_announcement_task(db.update_announcement_task(task_id, status="awaiting_ai_response", current_step=ANNOUNCEMENT_STEP["translate"], metadata=metadata)), "summary": {"status": "awaiting_ai_response", "reason": f"{provider} api_key is required; upload AI response instead"}, "artifacts": []}
    run = db.insert_run(task["project_id"], kind="announcement_translate", language=languages[0] if languages else "en", metadata={"task_id": task_id, "languages": languages, "provider": provider})
    metadata = {**metadata, "translate_run_id": run["id"]}
    db.update_announcement_task(task_id, status="running", current_step=ANNOUNCEMENT_STEP["translate"], metadata=metadata)
    output = run_dir(run["id"]) / "announcement_translate"
    output.mkdir(parents=True, exist_ok=True)
    artifacts: list[dict[str, Any]] = []
    response_artifacts: dict[str, str] = {}
    source_stem = _announcement_task_source_stem(task)
    for language in languages:
        lang_code = _visible_language_code(language)
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status="running", current_step=ANNOUNCEMENT_STEP["translate"])
        workpack_artifact = db.get_artifact(metadata["workpack_artifact_ids"][language])
        rows = read_jsonl(Path(workpack_artifact["path"]))
        provider_rows = [{"id": row["id"], "source": row["source"], "term_hits": row.get("term_hits") or []} for row in rows]
        prompt = Path(db.get_artifact(metadata.get("prompt_artifact_ids", {}).get(language, ""))["path"]).read_text(encoding="utf-8") if metadata.get("prompt_artifact_ids", {}).get(language) else ""
        translated = await _translate_rows_with_orchestration(
            run_id=run["id"],
            rows=provider_rows,
            settings=settings,
            project_prompt=prompt,
            work_dir=output / language,
            batch_size=int(getattr(request, "batch_size", None) or settings.get("batch_size") or 90),
            language=language,
            cancel_event=cancel_event,
            confirm_api_budget=bool(getattr(request, "confirm_api_budget", False)),
        )
        if not translated and db.get_run(run["id"]).get("status") == "needs_input":
            task = db.update_announcement_task(task_id, status="needs_input", current_step=ANNOUNCEMENT_STEP["translate"], metadata=metadata)
            return {"task": _hydrate_announcement_task(task), "run": db.get_run(run["id"]), "summary": {"status": "needs_input", "reason": "api_budget_confirmation_required"}, "artifacts": artifacts}
        response_path = output / f"{source_stem}_ai_response_{lang_code}.jsonl"
        write_jsonl(response_path, [{"para_id": item["id"], "translation": item["translation"]} for item in translated])
        artifact = db.add_artifact(task["project_id"], f"公告 AI response ({lang_code})", response_path, "announcement_ai_response", run_id=run["id"], mime="application/jsonl", metadata={"task_id": task_id, "language": language, "provider": provider})
        response_artifacts[language] = artifact["id"]
        artifacts.append(artifact)
    import_result = import_announcement_ai_response(task_id, _SimpleRequest(languages=languages, response_artifacts_by_language=response_artifacts))
    db.merge_run_metadata(run["id"], {"response_artifact_ids": response_artifacts})
    db.update_run(run["id"], status="passed")
    return {"task": import_result["task"], "run": db.get_run(run["id"]), "summary": {"status": "translated", "languages": languages}, "artifacts": [*artifacts, *import_result.get("artifacts", [])]}


def import_announcement_ai_response(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    workbook_artifact = db.get_artifact(metadata.get("translation_workbook_artifact_id", ""))
    response_map = _announcement_response_artifact_map(request, languages)
    if not response_map:
        raise ValueError("response_artifact_ids or response_artifacts_by_language is required")
    imported_languages: list[str] = []
    for language in languages:
        artifact_id = response_map.get(language)
        if not artifact_id:
            continue
        response_artifact = db.get_artifact(artifact_id)
        if response_artifact["project_id"] != task["project_id"]:
            raise KeyError(artifact_id)
        _import_announcement_response_into_workbook(Path(workbook_artifact["path"]), Path(response_artifact["path"]), language)
        imported_languages.append(language)
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status="translated", current_step=ANNOUNCEMENT_STEP["apply"], metadata={"response_artifact_id": artifact_id, "translation_workbook_artifact_id": workbook_artifact["id"]})
    metadata.setdefault("response_artifact_ids", {}).update(response_map)
    task = db.update_announcement_task(task_id, status="translated", current_step=ANNOUNCEMENT_STEP["apply"], metadata=metadata)
    return {"task": _hydrate_announcement_task(task), "summary": {"languages": [language_spec(code).target_header for code in imported_languages], "imported": len(imported_languages)}, "artifacts": [workbook_artifact]}


def apply_announcement_task(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    workbook_artifact_id = str(getattr(request, "translation_workbook_artifact_id", "") or metadata.get("translation_workbook_artifact_id") or "")
    if not workbook_artifact_id:
        raise ValueError("translation workbook is required")
    workbook_artifact = db.get_artifact(workbook_artifact_id)
    segments = metadata.get("segments") or _announcement_task_segments(task)
    workbook_path = Path(workbook_artifact["path"])
    rows = _read_announcement_translation_workbook(workbook_path, languages)
    issues = _validate_announcement_translation_rows(segments, rows, languages)
    run = db.insert_run(task["project_id"], kind="announcement_apply", language=languages[0] if languages else "en", metadata={"task_id": task_id, "languages": languages})
    output = run_dir(run["id"]) / "announcement_apply"
    output.mkdir(parents=True, exist_ok=True)
    qa_path = output / "QA摘要.xlsx"
    auto_fixed_count = 0
    if issues:
        fixed_path = output / f"{workbook_path.stem}_auto_fixed.xlsx"
        shutil.copy2(workbook_path, fixed_path)
        auto_fixed_count = _repair_announcement_translation_workbook(fixed_path, issues, languages)
        if auto_fixed_count:
            fixed_artifact = db.add_artifact(
                task["project_id"],
                "公告严重问题自动修复中转表",
                fixed_path,
                "announcement_translation_workbook",
                run_id=run["id"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                origin="generated",
                metadata={
                    "task_id": task_id,
                    "languages": languages,
                    "source_artifact_id": workbook_artifact_id,
                    "fixed_count": auto_fixed_count,
                },
            )
            workbook_artifact = fixed_artifact
            workbook_path = fixed_path
            rows = _read_announcement_translation_workbook(workbook_path, languages)
            issues = _validate_announcement_translation_rows(segments, rows, languages)
    output_files = _write_announcement_outputs(task, segments, rows, languages, output / "outputs")
    _write_announcement_qa_summary(qa_path, issues, output_files)
    hard_blockers = len(issues)
    artifacts = [db.add_artifact(task["project_id"], "公告 QA 摘要", qa_path, "announcement_qa_summary", run_id=run["id"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", metadata={"task_id": task_id, "hard_blockers": hard_blockers})]
    output_artifacts: dict[str, str] = {}
    for language, path in output_files:
        artifact = db.add_artifact(task["project_id"], f"公告成品 ({_visible_language_code(language)})", path, "announcement_output_file", run_id=run["id"], mime=_mime_for_path(path), metadata={"task_id": task_id, "language": language, "hard_blockers": hard_blockers})
        artifacts.append(artifact)
        output_artifacts[language] = artifact["id"]
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status="applied", current_step=ANNOUNCEMENT_STEP["deliver"], metadata={"output_artifact_id": artifact["id"], "qa_summary_artifact_id": artifacts[0]["id"], "hard_blockers": hard_blockers})
    metadata.update({
        "qa_summary_artifact_id": artifacts[0]["id"],
        "output_artifact_ids": output_artifacts,
        "translation_workbook_artifact_id": workbook_artifact["id"],
        "hard_blockers": hard_blockers,
        "qa_issues": issues,
        "auto_fixed_hard_blockers": auto_fixed_count,
    })
    task = db.update_announcement_task(task_id, status="applied", current_step=ANNOUNCEMENT_STEP["deliver"], metadata=metadata)
    db.merge_run_metadata(run["id"], {"outputs": output_artifacts, "hard_blockers": hard_blockers, "auto_fixed_hard_blockers": auto_fixed_count})
    db.update_run(run["id"], status="passed")
    return {"task": _hydrate_announcement_task(task), "run": db.get_run(run["id"]), "summary": {"hard_blockers": hard_blockers, "auto_fixed": auto_fixed_count, "can_deliver": True, "languages": languages}, "artifacts": artifacts}


def fix_announcement_hard_blockers(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    workbook_artifact_id = str(getattr(request, "translation_workbook_artifact_id", "") or metadata.get("translation_workbook_artifact_id") or "")
    if not workbook_artifact_id:
        raise ValueError("translation workbook is missing; prepare/import translations before fixing")
    source_artifact = db.get_artifact(workbook_artifact_id)
    source_path = Path(source_artifact["path"])
    if not source_path.exists():
        raise FileNotFoundError(str(source_path))
    issues = metadata.get("qa_issues") if isinstance(metadata.get("qa_issues"), list) else []
    if not issues:
        segments = metadata.get("segments") or _announcement_task_segments(task)
        rows = _read_announcement_translation_workbook(source_path, languages)
        issues = _validate_announcement_translation_rows(segments, rows, languages)
    if not issues:
        return {"task": _hydrate_announcement_task(task), "summary": {"fixed": 0, "remaining_hard_blockers": 0, "message": "no hard blockers"}, "artifacts": []}

    run = db.insert_run(task["project_id"], kind="announcement_fix", language=languages[0] if languages else "en", metadata={"task_id": task_id, "languages": languages, "source_artifact_id": workbook_artifact_id})
    output = run_dir(run["id"]) / "announcement_fix"
    output.mkdir(parents=True, exist_ok=True)
    fixed_path = output / f"{Path(source_path).stem}_hardblock_fixed.xlsx"
    shutil.copy2(source_path, fixed_path)
    fixed_count = _repair_announcement_translation_workbook(fixed_path, issues, languages)
    fixed_artifact = db.add_artifact(
        task["project_id"],
        "公告严重问题修复中转表",
        fixed_path,
        "announcement_translation_workbook",
        run_id=run["id"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        origin="generated",
        metadata={"task_id": task_id, "languages": languages, "source_artifact_id": workbook_artifact_id, "fixed_count": fixed_count},
    )
    metadata.update({
        "translation_workbook_artifact_id": fixed_artifact["id"],
        "hardblock_fix_artifact_id": fixed_artifact["id"],
        "hardblock_fix_count": fixed_count,
    })
    task = db.update_announcement_task(task_id, status="translated", current_step=ANNOUNCEMENT_STEP["apply"], metadata=metadata)
    db.merge_run_metadata(run["id"], {"fixed_count": fixed_count, "fixed_artifact_id": fixed_artifact["id"]})
    db.update_run(run["id"], status="passed")

    apply_request = _SimpleRequest(languages=languages, translation_workbook_artifact_id=fixed_artifact["id"])
    try:
        applied = apply_announcement_task(task_id, apply_request)
        remaining = int((applied.get("summary") or {}).get("hard_blockers") or 0)
        applied["summary"] = {**(applied.get("summary") or {}), "fixed": fixed_count, "remaining_hard_blockers": remaining}
        applied["artifacts"] = [fixed_artifact, *(applied.get("artifacts") or [])]
        return applied
    except ValueError as exc:
        if "hard blockers" not in str(exc):
            raise
        current_task = _hydrate_announcement_task(db.get_announcement_task(task_id))
        remaining = int((current_task.get("metadata") or {}).get("hard_blockers") or 0)
        return {
            "task": current_task,
            "run": db.get_run(run["id"]),
            "summary": {"fixed": fixed_count, "remaining_hard_blockers": remaining, "message": "some hard blockers still need manual review"},
            "artifacts": [fixed_artifact],
        }


def deliver_announcement_task(task_id: str, request: Any) -> dict[str, Any]:
    task = db.get_announcement_task(task_id)
    metadata = _announcement_task_metadata(task)
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or task.get("selected_languages") or [], fallback=metadata.get("languages") or [])
    force = bool(getattr(request, "force", False))
    hard_blockers = _announcement_hard_blocker_count(task, metadata)
    forced_by_hard_blockers = hard_blockers > 0
    stamp = str(getattr(request, "date_stamp", "") or datetime.now().strftime("%Y%m%d"))
    existing_artifact = _find_existing_announcement_delivery(task, languages, stamp)
    if existing_artifact and not force:
        existing_run = None
        if existing_artifact.get("run_id"):
            try:
                existing_run = db.get_run(str(existing_artifact["run_id"]))
            except KeyError:
                existing_run = None
        existing_languages = _normalize_announcement_languages((existing_artifact.get("metadata") or {}).get("languages") or [], fallback=languages)
        metadata["delivery_artifact_id"] = existing_artifact["id"]
        task = db.update_announcement_task(task_id, status="delivered", current_step=ANNOUNCEMENT_STEP["deliver"], metadata=metadata)
        return {
            "task": _hydrate_announcement_task(task),
            "run": existing_run,
            "summary": {"languages": existing_languages or languages, "delivery_artifact_id": existing_artifact["id"], "reused": True, "date_stamp": stamp},
            "artifacts": [existing_artifact],
        }
    superseded_artifacts = _matching_announcement_delivery_artifacts(task, languages, stamp) if force else []
    run = db.insert_run(task["project_id"], kind="announcement_deliver", language=languages[0] if languages else "en", metadata={"task_id": task_id, "languages": languages})
    output = run_dir(run["id"]) / "announcement_delivery"
    output.mkdir(parents=True, exist_ok=True)
    output_artifact_ids = metadata.get("output_artifact_ids") or {}
    qa_artifact_id = metadata.get("qa_summary_artifact_id")
    if not output_artifact_ids:
        generated = _force_materialize_announcement_outputs_for_delivery(task, metadata, languages, output, run["id"])
        output_artifact_ids = generated["output_artifact_ids"]
        qa_artifact_id = generated["qa_summary_artifact_id"]
        metadata.update(generated)
        hard_blockers = max(hard_blockers, int(generated.get("hard_blockers") or 0))
        forced_by_hard_blockers = hard_blockers > 0
    zip_path = output / f"{_announcement_delivery_base_name(task)}_announcement_delivery_{stamp}.zip"
    with ZipFile(zip_path, "w") as archive:
        for language in languages:
            artifact_id = output_artifact_ids.get(language)
            if not artifact_id:
                continue
            artifact = db.get_artifact(artifact_id)
            archive.write(artifact["path"], f"{_visible_language_code(language)}/{Path(artifact['path']).name}")
        if qa_artifact_id:
            qa_artifact = db.get_artifact(qa_artifact_id)
            archive.write(qa_artifact["path"], "QA摘要.xlsx")
    artifact_metadata = {"task_id": task_id, "languages": languages, "date_stamp": stamp}
    if forced_by_hard_blockers:
        artifact_metadata.update({"forced": True, "hard_blockers": hard_blockers, "source_type": DELIVERED_WITH_ISSUES_SOURCE_TYPE})
    artifact = db.add_artifact(task["project_id"], "公告交付总包", zip_path, "announcement_delivery_package", run_id=run["id"], mime="application/zip", metadata=artifact_metadata)
    for old_artifact in superseded_artifacts:
        if old_artifact["id"] == artifact["id"]:
            continue
        db.update_artifact(old_artifact["id"], {"metadata": {**(old_artifact.get("metadata") or {}), "superseded": True, "superseded_by": artifact["id"], "superseded_at": datetime.now().isoformat(timespec="seconds")}})
    metadata["delivery_artifact_id"] = artifact["id"]
    task = db.update_announcement_task(task_id, status="delivered", current_step=ANNOUNCEMENT_STEP["deliver"], metadata=metadata)
    for language in languages:
        db.upsert_announcement_task_language(task_id, task["project_id"], language, status="delivered", current_step=ANNOUNCEMENT_STEP["deliver"])
    db.merge_run_metadata(run["id"], {"delivery_artifact_id": artifact["id"], "forced": forced_by_hard_blockers, "hard_blockers": hard_blockers})
    db.update_run(run["id"], status="passed")
    return {"task": _hydrate_announcement_task(task), "run": db.get_run(run["id"]), "summary": {"languages": languages, "delivery_artifact_id": artifact["id"], "date_stamp": stamp, "forced": forced_by_hard_blockers, "hard_blockers": hard_blockers}, "artifacts": [artifact]}


def _announcement_hard_blocker_count(task: dict[str, Any], metadata: dict[str, Any]) -> int:
    counts = [int(metadata.get("hard_blockers") or 0)]
    qa_issues = metadata.get("qa_issues")
    if isinstance(qa_issues, list):
        counts.append(sum(1 for issue in qa_issues if str((issue or {}).get("severity") or "hard").lower() == "hard"))
    for artifact in db.list_artifacts(project_id=task["project_id"], include_superseded=True):
        artifact_metadata = artifact.get("metadata") or {}
        if str(artifact_metadata.get("task_id") or "") != str(task["id"]):
            continue
        if artifact["kind"] in {"announcement_qa_summary", "announcement_docx_qa_summary"}:
            counts.append(int(artifact_metadata.get("hard_blockers") or 0))
    for child in task.get("languages") or []:
        child_metadata = child.get("metadata") or {}
        counts.append(int(child_metadata.get("hard_blockers") or 0))
    return max(counts)


def _force_materialize_announcement_outputs_for_delivery(task: dict[str, Any], metadata: dict[str, Any], languages: list[str], output: Path, run_id: str) -> dict[str, Any]:
    workbook_artifact_id = str(metadata.get("translation_workbook_artifact_id") or "")
    if not workbook_artifact_id:
        raise ValueError("translation workbook is missing; prepare/import translations before delivery")
    workbook_artifact = db.get_artifact(workbook_artifact_id)
    segments = metadata.get("segments") or _announcement_task_segments(task)
    rows = _read_announcement_translation_workbook(Path(workbook_artifact["path"]), languages)
    issues = metadata.get("qa_issues") if isinstance(metadata.get("qa_issues"), list) else _validate_announcement_translation_rows(segments, rows, languages)
    output_files = _write_announcement_outputs(task, segments, rows, languages, output / "forced_outputs")
    qa_path = output / "QA摘要.xlsx"
    _write_announcement_qa_summary(qa_path, issues, output_files)
    qa_artifact = db.add_artifact(task["project_id"], "公告 QA 摘要", qa_path, "announcement_qa_summary", run_id=run_id, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", metadata={"task_id": task["id"], "hard_blockers": len(issues), "forced_delivery": True})
    output_artifact_ids: dict[str, str] = {}
    for language, path in output_files:
        artifact = db.add_artifact(task["project_id"], f"公告成品 ({_visible_language_code(language)})", path, "announcement_output_file", run_id=run_id, mime=_mime_for_path(path), metadata={"task_id": task["id"], "language": language, "forced_delivery": True})
        output_artifact_ids[language] = artifact["id"]
        db.upsert_announcement_task_language(task["id"], task["project_id"], language, status="applied_with_blockers", current_step=ANNOUNCEMENT_STEP["deliver"], metadata={"output_artifact_id": artifact["id"], "qa_summary_artifact_id": qa_artifact["id"], "forced_delivery": True})
    return {
        "qa_summary_artifact_id": qa_artifact["id"],
        "output_artifact_ids": output_artifact_ids,
        "hard_blockers": len(issues),
        "forced_delivery": True,
    }


def _find_existing_announcement_delivery(task: dict[str, Any], languages: list[str], date_stamp: str) -> dict[str, Any] | None:
    for artifact in _matching_announcement_delivery_artifacts(task, languages, date_stamp):
        if (artifact.get("metadata") or {}).get("superseded"):
            continue
        return artifact
    return None


def _matching_announcement_delivery_artifacts(task: dict[str, Any], languages: list[str], date_stamp: str) -> list[dict[str, Any]]:
    task_id = str(task.get("id") or "")
    expected_languages = set(_normalize_announcement_languages(languages, fallback=[]))
    matches: list[dict[str, Any]] = []
    for artifact in db.list_artifacts(project_id=task["project_id"], role="delivery", include_superseded=True):
        if artifact["kind"] not in {"announcement_delivery_package", "announcement_docx_delivery_package"}:
            continue
        metadata = artifact.get("metadata") or {}
        if str(metadata.get("task_id") or "") != task_id:
            continue
        artifact_languages = set(_normalize_announcement_languages(metadata.get("languages") or [], fallback=[]))
        if expected_languages and artifact_languages and artifact_languages != expected_languages:
            continue
        if _announcement_delivery_artifact_date(artifact) != date_stamp:
            continue
        if not Path(artifact["path"]).exists():
            continue
        matches.append(artifact)
    return matches


def _announcement_delivery_artifact_date(artifact: dict[str, Any]) -> str:
    metadata = artifact.get("metadata") or {}
    if metadata.get("date_stamp"):
        return str(metadata["date_stamp"])
    match = re.search(r"_announcement_delivery_(\d{8})\.zip$", Path(str(artifact.get("path") or "")).name)
    return match.group(1) if match else ""


def generate_announcement_terms_package(project_id: str, request: Any) -> dict[str, Any]:
    text = str(getattr(request, "text", "") or "")
    material_ids = list(getattr(request, "material_artifact_ids", []) or [])
    if not text.strip() and not material_ids:
        raise ValueError("announcement terms requires text or material_artifact_ids")
    if material_ids:
        text = "\n".join([text, *[_read_lookup_material_text(Path(db.get_artifact(artifact_id)["path"])) for artifact_id in material_ids]]).strip()
    if not text.strip():
        raise ValueError("announcement text is empty")
    languages = _normalize_announcement_languages(getattr(request, "languages", []) or [], fallback=["en"])
    metadata = {
        "language_table_artifact_ids": list(getattr(request, "language_table_artifact_ids", []) or []),
        "constraint_artifact_ids": [],
        "include_project_archive": False,
    }
    candidates = _announcement_constraint_rows(project_id, metadata, languages)
    rows = _select_announcement_constraint_rows(_compact_lookup_text(text), candidates, languages, min_hit=max(1, int(getattr(request, "announcement_min_hit", 1) or 1)))
    run = db.insert_run(project_id, kind="announcement_terms", language=languages[0] if languages else "en", metadata={"languages": languages})
    output = run_dir(run["id"]) / "announcement_terms"
    output.mkdir(parents=True, exist_ok=True)
    base = "announcement"
    if material_ids:
        try:
            base = _artifact_source_stem(db.get_artifact(material_ids[0]))
        except KeyError:
            base = "announcement"
    stamp = _today_stamp()
    workbook_path = output / f"{base}_announcement_terms_{stamp}.xlsx"
    manifest_path = output / f"{base}_announcement_terms_manifest_{stamp}.json"
    validation_path = output / f"{base}_announcement_terms_validation_{stamp}.md"
    rows, ai_summary = _apply_announcement_ai_supplement(
        project_id=project_id,
        output_dir=output,
        base_name=base,
        source_text=_compact_lookup_text(text),
        rows=rows,
        candidates=candidates,
        languages=languages,
        request=request,
        project_name=db.get_project(project_id).get("name", ""),
    )
    _write_announcement_terms_workbook(workbook_path, rows, languages)
    summary = {"terms": len(rows), "languages": languages, "source_chars": len(text)}
    if ai_summary:
        summary["ai_supplement"] = {
            key: ai_summary[key]
            for key in ("enabled", "response_artifact_id", "provider", "provider_status", "provider_error", "term_count", "added_to_main", "report_only", "project_name_translation_missing")
        }
    manifest = {"kind": "announcement_terms", "project_id": project_id, "languages": languages, "summary": summary, "terms": rows}
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    validation_path.write_text(_announcement_terms_validation(summary, rows, languages), encoding="utf-8")
    artifacts = [
        db.add_artifact(project_id, "公告术语表", workbook_path, "announcement_terms_workbook", run_id=run["id"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", metadata={"languages": languages}),
        db.add_artifact(project_id, "公告术语 validation", validation_path, "announcement_terms_validation", run_id=run["id"], mime="text/markdown", metadata={"languages": languages}),
        db.add_artifact(project_id, "公告术语 manifest", manifest_path, "announcement_terms_manifest", run_id=run["id"], mime="application/json", metadata={"languages": languages}),
    ]
    if ai_summary:
        packet_artifact = db.add_artifact(project_id, "公告 AI 补充包", Path(ai_summary["packet_path"]), "announcement_ai_supplement_packet", run_id=run["id"], mime="application/json", metadata={"languages": languages})
        response_artifact = None
        if ai_summary.get("response_path"):
            response_artifact = db.add_artifact(project_id, "公告 AI 补充响应", Path(ai_summary["response_path"]), "announcement_ai_supplement_response", run_id=run["id"], mime="application/json", metadata={"languages": languages, "provider": ai_summary.get("provider", "")})
        report_artifact = db.add_artifact(project_id, "公告 AI 补充报告", Path(ai_summary["report_path"]), "announcement_ai_supplement_report", run_id=run["id"], mime="text/markdown", metadata={"languages": languages})
        artifacts.extend([artifact for artifact in (packet_artifact, response_artifact, report_artifact) if artifact])
        summary["ai_supplement"]["packet_artifact_id"] = packet_artifact["id"]
        if response_artifact:
            summary["ai_supplement"]["response_artifact_id"] = response_artifact["id"]
        summary["ai_supplement"]["report_artifact_id"] = report_artifact["id"]
    db.merge_run_metadata(run["id"], {"summary": summary})
    db.update_run(run["id"], status="passed")
    return {"run": db.get_run(run["id"]), "summary": summary, "artifacts": artifacts, "manifest": manifest}


def legacy_prepare_announcement_docx(project_id: str, request: Any) -> dict[str, Any]:
    if not list(getattr(request, "source_artifact_ids", []) or []):
        raise ValueError("source_artifact_ids is required")
    source_artifact_id = list(getattr(request, "source_artifact_ids", []))[0]
    create_request = _SimpleRequest(
        source_artifact_id=source_artifact_id,
        title="公告 DOCX",
        languages=list(getattr(request, "languages", []) or []),
        constraint_artifact_ids=[getattr(request, "terms_artifact_id")],
        language_table_artifact_ids=[getattr(request, "terms_artifact_id")],
        include_project_archive=False,
    )
    task = create_announcement_task(project_id, create_request)
    extract_announcement_terms(task["id"], _SimpleRequest(languages=create_request.languages, language_table_artifact_ids=create_request.language_table_artifact_ids, include_project_archive=False))
    lookup_announcement_translations(task["id"], _SimpleRequest(languages=create_request.languages, include_project_archive=False))
    prepared = prepare_announcement_translation(task["id"], _SimpleRequest(languages=create_request.languages))
    run = prepared["run"]
    db.merge_run_metadata(run["id"], {"task_id": task["id"], "legacy_prepare": True})
    artifacts = []
    for artifact in prepared["artifacts"]:
        if artifact["kind"] == "announcement_translation_workbook":
            artifact = db.update_artifact(artifact["id"], {"label": "Announcement DOCX translation workbook", "metadata": {**artifact.get("metadata", {}), "legacy_kind": "announcement_docx_translation_workbook"}})
            artifact["kind"] = "announcement_docx_translation_workbook"
        elif artifact["kind"] == "announcement_workpack":
            artifact = db.update_artifact(artifact["id"], {"label": artifact["label"].replace("公告", "Announcement DOCX"), "metadata": {**artifact.get("metadata", {}), "legacy_kind": "announcement_docx_workpack"}})
            artifact["kind"] = "announcement_docx_workpack"
        elif artifact["kind"] == "announcement_docx_manifest":
            pass
        artifacts.append(artifact)
    return {**prepared, "run": db.get_run(run["id"]), "artifacts": artifacts}


def legacy_import_announcement_docx_ai(project_id: str, request: Any) -> dict[str, Any]:
    run = db.get_run(getattr(request, "prepare_run_id"))
    task_id = str((run.get("metadata") or {}).get("task_id") or "")
    if not task_id:
        raise KeyError("task_id")
    result = import_announcement_ai_response(task_id, request)
    return {"summary": result["summary"], "task": result["task"], "artifacts": result["artifacts"]}


def legacy_apply_announcement_docx(project_id: str, request: Any) -> dict[str, Any]:
    run = db.get_run(getattr(request, "prepare_run_id"))
    task_id = str((run.get("metadata") or {}).get("task_id") or "")
    result = apply_announcement_task(task_id, request)
    legacy_artifacts = []
    for artifact in result["artifacts"]:
        if artifact["kind"] == "announcement_output_file" and str(artifact.get("path", "")).lower().endswith(".docx"):
            artifact["kind"] = "announcement_docx_output_docx"
        elif artifact["kind"] == "announcement_qa_summary":
            artifact["kind"] = "announcement_docx_qa_summary"
        legacy_artifacts.append(artifact)
    result["artifacts"] = legacy_artifacts
    return result


def legacy_deliver_announcement_docx(project_id: str, request: Any) -> dict[str, Any]:
    run = db.get_run(getattr(request, "prepare_run_id"))
    task_id = str((run.get("metadata") or {}).get("task_id") or "")
    result = deliver_announcement_task(task_id, request)
    for artifact in result["artifacts"]:
        if artifact["kind"] == "announcement_delivery_package":
            artifact["kind"] = "announcement_docx_delivery_package"
    return result


class _SimpleRequest:
    def __init__(self, **kwargs: Any) -> None:
        self.__dict__.update(kwargs)


def _hydrate_announcement_task(task: dict[str, Any]) -> dict[str, Any]:
    metadata = task.get("metadata") or {}
    translate_run_id = str(metadata.get("translate_run_id") or "")
    if translate_run_id:
        try:
            translate_run = db.get_run(translate_run_id)
            metadata = {
                **metadata,
                "translate_run_status": translate_run.get("status"),
                "translation_progress": (translate_run.get("metadata") or {}).get("translation_progress"),
                "api_budget_estimate": (translate_run.get("metadata") or {}).get("api_budget_estimate"),
            }
            task["metadata"] = metadata
        except KeyError:
            pass
    artifact_ids: set[str] = set()
    for key in (
        "terms_artifact_id",
        "terms_manifest_artifact_id",
        "terms_validation_artifact_id",
        "translation_workbook_artifact_id",
        "manifest_artifact_id",
        "qa_summary_artifact_id",
        "delivery_artifact_id",
    ):
        if metadata.get(key):
            artifact_ids.add(str(metadata[key]))
    for mapping_key in ("workpack_artifact_ids", "prompt_artifact_ids", "response_artifact_ids", "output_artifact_ids"):
        value = metadata.get(mapping_key)
        if isinstance(value, dict):
            artifact_ids.update(str(item) for item in value.values() if item)
    ai_supplement = metadata.get("ai_supplement")
    if isinstance(ai_supplement, dict):
        for key in ("packet_artifact_id", "report_artifact_id"):
            if ai_supplement.get(key):
                artifact_ids.add(str(ai_supplement[key]))
    artifacts: list[dict[str, Any]] = []
    for artifact_id in sorted(artifact_ids):
        try:
            artifacts.append(db.get_artifact(artifact_id))
        except KeyError:
            continue
    task["artifacts"] = artifacts
    return task



def _merge_announcement_constraint_request(metadata: dict[str, Any], request: Any) -> dict[str, Any]:
    merged = dict(metadata)
    for key in ("language_table_artifact_ids", "constraint_artifact_ids"):
        value = list(getattr(request, key, []) or [])
        if value:
            merged[key] = value
    if hasattr(request, "include_project_archive"):
        merged["include_project_archive"] = bool(getattr(request, "include_project_archive", True))
    return merged


def _create_inline_announcement_source(project_id: str, text: str, title: str) -> dict[str, Any]:
    output = project_dir(project_id) / "announcements" / "inline"
    output.mkdir(parents=True, exist_ok=True)
    path = output / f"{_safe_file_stem(title or 'announcement')}_{hashlib.sha1(text.encode('utf-8')).hexdigest()[:8]}.txt"
    path.write_text(text, encoding="utf-8")
    return db.add_artifact(project_id, "Inline announcement source", path, "asset", mime="text/plain", origin="uploaded")

__all__ = [name for name in globals() if not name.startswith("__")]
