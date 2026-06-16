from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .. import db
from ..config import load_settings
from ..languages import SOURCE_HEADER_ALIASES, require_supported_language, target_aliases
from .announcement_segments import _is_quick_text_path, _quick_text_translation_rows, _txt_announcement_segments
from .common import (
    HARNESS_SCHEMA_VERSION,
    LANGUAGE_ORDER,
    _STRUCTURAL_TARGET_HEADERS,
    _TARGET_DETECTION_ALIASES,
    _looks_like_untranslated_seed,
)
from .qa import _first_col, _header_map, _is_supported_translation_id, _row_cell
from .subprocess_runner import user_facing_error

def inspect_translation_targets(artifact_id: str) -> dict[str, Any]:
    artifact = db.get_artifact(artifact_id)
    path = Path(artifact["path"])
    result: dict[str, Any] = {
        "artifact_id": artifact_id,
        "label": artifact.get("label", ""),
        "supported_file": path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"} or _is_quick_text_path(path),
        "source_detected": False,
        "detected_languages": [],
        "suggested_language": None,
        "sheets": [],
    }
    if _is_quick_text_path(path) and path.exists():
        segments = _txt_announcement_segments(path)
        result["source_detected"] = bool(segments)
        result["sheets"] = [{"sheet": path.name, "languages": [], "source_detected": bool(segments)}]
        return result
    if not result["supported_file"] or not path.exists():
        return result
    detected: set[str] = set()
    sheet_rows: list[dict[str, Any]] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1), ())]
                normalized = [header.lower() for header in headers]
                source_detected = any(header in SOURCE_HEADER_ALIASES for header in normalized)
                result["source_detected"] = bool(result["source_detected"] or source_detected)
                sheet_languages: list[str] = []
                for index, header in enumerate(normalized):
                    if not header or header in _STRUCTURAL_TARGET_HEADERS:
                        continue
                    for code in LANGUAGE_ORDER:
                        if header in _TARGET_DETECTION_ALIASES.get(code, set()):
                            detected.add(code)
                            sheet_languages.append(code)
                            break
                if sheet_languages or source_detected:
                    sheet_rows.append({"sheet": ws.title, "languages": sorted(set(sheet_languages), key=LANGUAGE_ORDER.index), "source_detected": source_detected})
        finally:
            wb.close()
    except Exception as exc:
        result["reason"] = f"inspect_failed:{user_facing_error(exc)}"
        return result
    languages = [code for code in LANGUAGE_ORDER if code in detected]
    result["detected_languages"] = languages
    result["suggested_language"] = languages[0] if languages else None
    result["sheets"] = sheet_rows
    return result



def _scan_workbook_translation_readiness(path: Path, language: str, summary: dict[str, Any]) -> tuple[bool, str | None]:
    found_target_column = False
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                try:
                    headers = _header_map(ws)
                except Exception:
                    continue
                source_col = _first_col(headers, list(SOURCE_HEADER_ALIASES))
                target_col = _first_col(headers, target_aliases(language))
                id_col = _first_col(headers, ["id", "编号", "序号"])
                if source_col is None:
                    continue
                if target_col is not None:
                    found_target_column = True
                for row in ws.iter_rows(min_row=2, values_only=True):
                    source = _row_cell(row, source_col)
                    if not source:
                        continue
                    summary["source_rows"] += 1
                    raw_id = _row_cell(row, id_col) if id_col is not None else ""
                    if not _is_supported_translation_id(raw_id):
                        summary["invalid_id_rows"] += 1
                        if len(summary["invalid_id_samples"]) < 5:
                            summary["invalid_id_samples"].append(raw_id or "<missing>")
                    target = _row_cell(row, target_col) if target_col is not None else ""
                    if not target:
                        summary["empty_target_rows"] += 1
                    elif _looks_like_untranslated_seed(target, language):
                        summary["cjk_target_rows"] += 1
                    else:
                        summary["translated_rows"] += 1
        finally:
            wb.close()
    except Exception as exc:
        return found_target_column, f"inspect_failed:{user_facing_error(exc)}"
    return found_target_column, None


def _set_readiness_action(summary: dict[str, Any], *, input_mode: str, next_step: int, user_message: str, format_errors: list[str] | None = None) -> dict[str, Any]:
    summary["input_mode"] = input_mode
    summary["next_step"] = next_step
    summary["user_message"] = user_message
    summary["format_errors"] = format_errors or []
    return summary


def _finalize_translation_readiness(summary: dict[str, Any], *, found_target_column: bool, effective_batch_size: int) -> dict[str, Any]:
    source_rows = int(summary["source_rows"])
    empty_rows = int(summary["empty_target_rows"])
    cjk_rows = int(summary["cjk_target_rows"])
    translated_rows = int(summary["translated_rows"])
    summary["estimated_batches"] = math.ceil(source_rows / effective_batch_size) if source_rows else 0
    if not source_rows:
        summary["reason"] = "no_source_rows"
        return _set_readiness_action(
            summary,
            input_mode="invalid",
            next_step=4,
            user_message="未检测到中文原文列，请上传包含 ID 和 CN/简体中文/原文 的语言表。",
            format_errors=["missing_source_rows"],
        )
    if int(summary["invalid_id_rows"]):
        summary["estimated_batches"] = 0
        summary["reason"] = "invalid_id_rows"
        return _set_readiness_action(
            summary,
            input_mode="invalid",
            next_step=4,
            user_message="有行缺少可回写 ID，请修正后重新上传。",
            format_errors=["invalid_id_rows"],
        )
    if not found_target_column:
        summary["reason"] = "target_column_missing"
        language = str(summary.get("target_language") or "en").upper()
        return _set_readiness_action(
            summary,
            input_mode="invalid",
            next_step=4,
            user_message=f"未检测到 {language} 译文列；请上传包含 ID、CN/简体中文/原文 和目标语言列的语言表，或先切换目标语言。",
            format_errors=["target_column_missing"],
        )
    if empty_rows == 0 and cjk_rows == 0 and translated_rows > 0:
        summary["ready_for_qa"] = True
        summary["reason"] = "existing_target_translation"
        return _set_readiness_action(
            summary,
            input_mode="ready_for_qa",
            next_step=8,
            user_message="检测到已有完整译文，可跳过模型翻译，直接进入校对。",
        )
    summary["needs_translation"] = True
    summary["ready_for_translation"] = True
    if empty_rows and cjk_rows:
        summary["reason"] = "empty_or_cjk_target_rows"
    elif empty_rows:
        summary["reason"] = "empty_target_rows"
    elif cjk_rows:
        summary["reason"] = "cjk_target_rows"
    else:
        summary["reason"] = "needs_translation"
    return _set_readiness_action(
        summary,
        input_mode="needs_translation",
        next_step=5,
        user_message="检测到待翻译内容，请先扫描术语候选。",
    )

def inspect_translation_readiness(artifact_id: str, batch_size: int | None = None, language: str = "en") -> dict[str, Any]:
    language = require_supported_language(language)
    artifact = db.get_artifact(artifact_id)
    path = Path(artifact["path"])
    effective_batch_size = max(1, min(int(batch_size or load_settings().get("batch_size") or 90), 200))
    summary = {
        "artifact_id": artifact_id,
        "label": artifact.get("label", ""),
        "target_language": language,
        "source_rows": 0,
        "translated_rows": 0,
        "empty_target_rows": 0,
        "cjk_target_rows": 0,
        "invalid_id_rows": 0,
        "invalid_id_samples": [],
        "needs_translation": False,
        "ready_for_translation": False,
        "ready_for_qa": False,
        "reason": "unsupported_file",
        "batch_size": effective_batch_size,
        "estimated_batches": 0,
        "input_mode": "invalid",
        "next_step": 4,
        "format_errors": ["unsupported_file"],
        "user_message": "当前文件类型不适合作为语言表，请上传 XLSX/XLS/CSV 语言表。",
    }
    if _is_quick_text_path(path) and path.exists():
        rows = _quick_text_translation_rows(path)
        summary["source_rows"] = len(rows)
        summary["empty_target_rows"] = len(rows)
        summary["needs_translation"] = bool(rows)
        summary["ready_for_translation"] = bool(rows)
        summary["reason"] = "needs_translation" if rows else "no_source_rows"
        summary["estimated_batches"] = math.ceil(len(rows) / effective_batch_size) if rows else 0
        if rows:
            return _set_readiness_action(
                summary,
                input_mode="needs_translation",
                next_step=5,
                user_message="检测到待翻译文本，可进入翻译。",
            )
        return _set_readiness_action(
            summary,
            input_mode="invalid",
            next_step=4,
            user_message="未检测到可翻译文本，请重新上传。",
            format_errors=["no_source_rows"],
        )
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"} or not path.exists():
        return summary

    found_target_column, scan_error = _scan_workbook_translation_readiness(path, language, summary)
    if scan_error:
        summary["reason"] = scan_error
        return _set_readiness_action(
            summary,
            input_mode="invalid",
            next_step=4,
            user_message="语言表读取失败，请确认文件未损坏且表头格式正确。",
            format_errors=["inspect_failed"],
        )
    return _finalize_translation_readiness(summary, found_target_column=found_target_column, effective_batch_size=effective_batch_size)



def _project_harness_prompt_parts(harness: dict[str, Any]) -> list[str]:
    parts: list[str] = []
    if harness.get("target_audience"):
        parts.append(f"- Target audience: {harness['target_audience']}")
    if harness.get("tone"):
        parts.append(f"- Tone: {harness['tone']}")
    if harness.get("style_guidance"):
        parts.append(f"- Style guidance: {harness['style_guidance']}")
    forbidden = [str(item).strip() for item in harness.get("forbidden_translations", []) if str(item).strip()]
    if forbidden:
        parts.append("- Forbidden translations: " + "; ".join(forbidden))
    fixed_terms = []
    for item in harness.get("fixed_terms", []):
        if not isinstance(item, dict):
            continue
        source = str(item.get("source", "")).strip()
        target = str(item.get("target", "")).strip()
        if source and target:
            fixed_terms.append(f"{source} => {target}")
    if fixed_terms:
        parts.append("- Fixed terms: " + "; ".join(fixed_terms))
    for label, key in (("Hard project rules", "hard_rules"), ("Soft project rules", "soft_rules")):
        rules = [
            str(rule.get("description") or rule.get("label") or "").strip()
            for rule in harness.get(key, [])
            if isinstance(rule, dict) and rule.get("enabled", True) and str(rule.get("description") or rule.get("label") or "").strip()
        ]
        if rules:
            parts.append(f"- {label}: " + "; ".join(rules))
    examples = []
    for item in harness.get("reference_examples", []):
        if not isinstance(item, dict):
            continue
        source = str(item.get("source", "")).strip()
        target = str(item.get("target", "")).strip()
        if source and target:
            examples.append(f"{source} => {target}")
    if examples:
        parts.append("- Accepted examples: " + "; ".join(examples[:10]))
    return parts

def _harness_summary(harness: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": "project_harness",
        "schema_version": harness.get("schema_version", HARNESS_SCHEMA_VERSION),
        "updated_at": harness.get("updated_at", ""),
        "style_guidance": bool(harness.get("style_guidance")),
        "hard_rules": len(harness.get("hard_rules", [])),
        "soft_rules": len(harness.get("soft_rules", [])),
        "fixed_terms": len(harness.get("fixed_terms", [])),
        "forbidden_translations": len(harness.get("forbidden_translations", [])),
        "reference_examples": len(harness.get("reference_examples", [])),
    }

__all__ = [name for name in globals() if not name.startswith("__")]
