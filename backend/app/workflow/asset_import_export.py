from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .. import db
from ..languages import SOURCE_HEADER_ALIASES, alt_aliases, normalize_language, require_supported_language, target_aliases, ui_language_code
from .common import project_dir
from .naming import _safe_delivery_name, _today_stamp
from .table_helpers import (
    IGNORED_AUTO_SHEET_TITLES,
    ImportContractError,
    LANGUAGE_ORDER,
    ParsedImportTable,
    UnsupportedImportFormatError,
    XLSX_IMPORT_SUFFIXES,
    _auto_language_indices,
    _column_index,
    _mapping_rows_to_matrix,
    _normalized_header_indices,
    _primary_target,
    _read_csv_matrix,
    _read_glossary_rows,
    _read_json_mapping_rows,
    _select_xlsx_data_sheet,
    _sheet_headers,
    _value_at,
    _wide_source_key,
    _worksheet_has_data,
)

_LARGE_LANGUAGE_TABLE_ROW_THRESHOLD = 1000
_LANGUAGE_TABLE_SOURCE_ALIASES = [alias for alias in SOURCE_HEADER_ALIASES if alias not in {"term", "术语"}]
COMPLETE_LANGUAGE_TABLE_GLOSSARY_IMPORT_MESSAGE = "这个文件看起来是完整语言表，不是项目术语表。请到「生成术语」或翻译流程 STEP5 做高频词扫描并生成术语候选，候选确认后才会进入项目术语库。"
COMPLETE_LANGUAGE_TABLE_PROJECT_MATERIAL_MESSAGE = "这个文件看起来是完整语言表，请上传到 STEP4「语言表」。它不会作为项目资料参与术语提取。"
INVALID_GLOSSARY_TEMPLATE_MESSAGE = "术语表格式有误，请重新上传。请先下载导入模板，按模板列填写：ID、CN、EN 或 KR/JP、分类、备注。"


def _has_complete_language_table_rows(headers: list[str], raw_rows: Any, row_threshold: int) -> bool:
    normalized = _normalized_header_indices(headers)
    term_key_idx = _column_index(normalized, None, ["id", "key", "编号", "序号"], required=False)
    source_idx = _column_index(normalized, None, _LANGUAGE_TABLE_SOURCE_ALIASES, required=False)
    if term_key_idx is None or source_idx is None:
        return False
    if not _auto_language_indices(headers, {term_key_idx, source_idx}):
        return False
    source_rows = 0
    for row in raw_rows:
        if _value_at(row, source_idx):
            source_rows += 1
            if source_rows > row_threshold:
                return True
    return False


def is_complete_language_table_for_glossary_import(path: Path, sheet: str | None = None, row_threshold: int = _LARGE_LANGUAGE_TABLE_ROW_THRESHOLD) -> bool:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        headers, raw_rows = _read_csv_matrix(path)
        return _has_complete_language_table_rows(headers, raw_rows, row_threshold)
    if suffix == ".json":
        mappings = _read_json_mapping_rows(path, ("terms", "rows", "entries"))
        if mappings and all(str(row.get("language") or "").strip() for row in mappings):
            return False
        headers, raw_rows = _mapping_rows_to_matrix(mappings)
        return _has_complete_language_table_rows(headers, raw_rows, row_threshold)
    if suffix not in XLSX_IMPORT_SUFFIXES:
        return False
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheets = [wb[sheet]] if sheet else [
            worksheet
            for worksheet in wb.worksheets
            if worksheet.title.strip().casefold() not in IGNORED_AUTO_SHEET_TITLES
        ]
        for worksheet in worksheets:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1), None)
            if header_row is None:
                continue
            headers = [str(cell.value or "").strip() for cell in header_row]
            if _has_complete_language_table_rows(headers, worksheet.iter_rows(min_row=2, values_only=True), row_threshold):
                return True
        return False
    finally:
        wb.close()


def guard_complete_language_table_for_glossary_import(path: Path, sheet: str | None = None) -> None:
    if is_complete_language_table_for_glossary_import(path, sheet=sheet):
        raise ValueError(COMPLETE_LANGUAGE_TABLE_GLOSSARY_IMPORT_MESSAGE)


def guard_complete_language_table_for_project_material(path: Path, sheet: str | None = None) -> None:
    if is_complete_language_table_for_glossary_import(path, sheet=sheet):
        raise ValueError(COMPLETE_LANGUAGE_TABLE_PROJECT_MATERIAL_MESSAGE)


def preview_glossary_import(project_id: str, request: Any, import_all: bool = False) -> dict[str, Any]:
    project = db.get_project(project_id)
    _ = project
    artifact = db.get_artifact(request.artifact_id)
    if artifact["project_id"] != project_id:
        raise KeyError("artifact")
    path = Path(artifact["path"])
    guard_complete_language_table_for_glossary_import(path, sheet=getattr(request, "sheet", None))
    language = require_supported_language(getattr(request, "language", "en") or "en")
    auto_languages = bool(getattr(request, "auto_languages", True))
    if auto_languages and not getattr(request, "target_column", None) and not getattr(request, "target_alt_column", None):
        try:
            rows, columns, languages = _read_multilingual_glossary_rows(
                path,
                sheet=getattr(request, "sheet", None),
                term_key_column=getattr(request, "term_key_column", None),
                source_column=getattr(request, "source_column", None),
                category_column=getattr(request, "category_column", None),
                note_column=getattr(request, "note_column", None),
                limit=None if import_all else int(getattr(request, "limit", 100) or 100),
            )
        except ImportContractError:
            raise
        except (KeyError, StopIteration, ValueError) as exc:
            raise ValueError(INVALID_GLOSSARY_TEMPLATE_MESSAGE) from exc
        if languages:
            _ensure_glossary_template_rows(rows)
            return {"artifact": artifact, "columns": columns, "rows": rows, "total_rows": len(rows), "language": "auto", "languages": languages}
        raise ValueError(INVALID_GLOSSARY_TEMPLATE_MESSAGE)
    try:
        rows, columns = _read_glossary_rows(
            path,
            sheet=getattr(request, "sheet", None),
            term_key_column=getattr(request, "term_key_column", None),
            source_column=getattr(request, "source_column", None),
            target_column=getattr(request, "target_column", None),
            target_alt_column=getattr(request, "target_alt_column", None),
            category_column=getattr(request, "category_column", None),
            note_column=getattr(request, "note_column", None),
            language=language,
            limit=None if import_all else int(getattr(request, "limit", 100) or 100),
        )
    except ImportContractError:
        raise
    except (KeyError, StopIteration, ValueError) as exc:
        raise ValueError(INVALID_GLOSSARY_TEMPLATE_MESSAGE) from exc
    _ensure_glossary_template_rows(rows)
    return {"artifact": artifact, "columns": columns, "rows": rows, "total_rows": len(rows), "language": language}


def import_glossary(project_id: str, request: Any) -> dict[str, Any]:
    from ..glossary_archive_batches import analyze_glossary_archive, commit_glossary_archive

    legacy_request = request.model_copy(
        update={"mode": "merge", "confirmed_glossary": True}
    ) if hasattr(request, "model_copy") else request
    analysis = analyze_glossary_archive(project_id, legacy_request)
    if int(analysis["summary"].get("skip") or 0) == int(analysis["summary"].get("source_rows") or 0):
        raise ValueError(INVALID_GLOSSARY_TEMPLATE_MESSAGE)
    result = commit_glossary_archive(project_id, analysis["token"])
    return {**result, "preview": analysis}


def _ensure_glossary_template_rows(rows: list[dict[str, Any]]) -> None:
    valid = [
        row for row in rows
        if str(row.get("source") or "").strip()
        and (str(row.get("target") or "").strip() or str(row.get("target_alt") or "").strip())
    ]
    if not valid:
        raise ValueError(INVALID_GLOSSARY_TEMPLATE_MESSAGE)


def export_glossary(project_id: str, fmt: str, language: str | None = None) -> dict[str, Any] | Path:
    project = db.get_project(project_id)
    language = require_supported_language(language or "en") if language else None
    terms = db.list_glossary_terms(project_id, language=language)
    if fmt == "json":
        return {
            "project_id": project_id,
            "language": language,
            "terms": [dict(zip(("term_key", "source", "target", "target_alt", "category", "note"), _glossary_export_row(term))) | {"language": term.get("language", "en")} for term in terms],
        }
    output_dir = project_dir(project_id) / "glossary" / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = _export_language_suffix(language)
    if language:
        columns = ["ID", "CN", ui_language_code(language), "分类", "备注"]
        rows = [_glossary_export_row(term, include_alt=False) for term in terms]
    else:
        wide = list_glossary_wide(project_id)
        languages = list(wide.get("languages") or [])
        columns = ["ID", "CN", *_wide_language_columns(languages), "分类", "备注"]
        rows = _glossary_wide_export_rows(wide, languages)
    if fmt == "csv":
        path = output_dir / _export_filename(project, "glossary", suffix, "csv")
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(columns)
            writer.writerows(rows)
        return path
    path = output_dir / _export_filename(project, "glossary", suffix, "xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(columns)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


def _export_language_suffix(language: str | None) -> str:
    return ui_language_code(language) if language else "ALL"


def _export_filename(project: dict[str, Any], kind: str, suffix: str, ext: str) -> str:
    return f"{_safe_delivery_name(project['name'])}_{kind}_{suffix}_{_today_stamp()}.{ext}"


def _glossary_export_row(term: dict[str, Any], *, include_alt: bool = True) -> list[Any]:
    row = [
        term.get("term_key", ""),
        term.get("source", ""),
        term.get("target", ""),
    ]
    if include_alt:
        row.append(term.get("target_alt", ""))
    row.extend([term.get("category", ""), term.get("note", "")])
    return row


def _wide_language_columns(languages: list[str]) -> list[str]:
    return [ui_language_code(code) for code in languages]


def _glossary_wide_export_rows(wide: dict[str, Any], languages: list[str]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in wide["rows"]:
        translations = row.get("translations") or {}
        values = [row.get("term_key", ""), row.get("source", "")]
        for code in languages:
            entry = translations.get(code) or {}
            values.append(entry.get("target", ""))
        values.extend([row.get("category", ""), row.get("note", "")])
        rows.append(values)
    return rows


def import_translation_archive(project_id: str, request: Any, source_type: str = "imported") -> dict[str, Any]:
    from ..translation_archive_batches import analyze_translation_archive, commit_translation_archive

    analysis = analyze_translation_archive(project_id, request, source_type=source_type)
    result = commit_translation_archive(project_id, analysis["token"])
    return {**result, "artifact_id": request.artifact_id}


def archive_translation_artifact(project_id: str, artifact_id: str, language: str = "en", source_type: str = "qa_passed") -> dict[str, Any]:
    class Request:
        pass

    request = Request()
    request.artifact_id = artifact_id
    request.language = language
    request.sheet = None
    request.id_column = None
    request.source_column = None
    request.target_column = None
    request.target_alt_column = None
    request.note_column = None
    request.auto_languages = False
    from ..translation_archive_batches import analyze_translation_archive, commit_translation_archive

    analysis = analyze_translation_archive(project_id, request, source_type=source_type)
    if not analysis["can_commit"]:
        return {
            "project_id": project_id,
            "artifact_id": artifact_id,
            "batch_id": analysis["batch_id"],
            "status": "blocked",
            "imported_count": 0,
            "entries": [],
            "languages": analysis["languages"],
            "summary": analysis["summary"],
            "conflicts": analysis["conflicts"],
        }
    return {**commit_translation_archive(project_id, analysis["token"]), "artifact_id": artifact_id}


def export_translation_archive(project_id: str, fmt: str, language: str | None = None) -> dict[str, Any] | Path:
    project = db.get_project(project_id)
    language = require_supported_language(language or "en") if language else None
    entries = db.list_translation_entries(project_id, language=language)
    if fmt == "json":
        return {"project_id": project_id, "language": language, "entries": [_translation_export_payload(entry) for entry in entries]}
    output_dir = project_dir(project_id) / "translations" / "exports"
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = _export_language_suffix(language)
    if language:
        columns = ["ID", "CN", ui_language_code(language), "备注"]
        rows = [_translation_export_row(entry, include_alt=False) for entry in entries]
    else:
        wide = list_translation_archive_wide(project_id)
        languages = list(wide.get("languages") or [])
        columns = ["ID", "CN", *_wide_language_columns(languages), "备注"]
        rows = _translation_wide_export_rows(wide, languages)
    if fmt == "csv":
        path = output_dir / _export_filename(project, "translations", suffix, "csv")
        with path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(columns)
            writer.writerows(rows)
        return path
    path = output_dir / _export_filename(project, "translations", suffix, "xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"
    ws.append(columns)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


def list_glossary_wide(project_id: str) -> dict[str, Any]:
    db.get_project(project_id)
    rows = _wide_rows(
        db.list_glossary_terms(project_id),
        key_field="term_key",
        shared_fields=("term_key", "category", "note"),
    )
    return {"project_id": project_id, **rows}


def list_translation_archive_wide(project_id: str) -> dict[str, Any]:
    db.get_project(project_id)
    rows = _wide_rows(
        db.list_translation_entries(project_id),
        key_field="entry_key",
        shared_fields=("entry_key", "note"),
    )
    return {"project_id": project_id, **rows}


def _wide_rows(items: list[dict[str, Any]], *, key_field: str, shared_fields: tuple[str, ...]) -> dict[str, Any]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        source_key = _wide_source_key(item.get("source"))
        if not source_key:
            continue
        grouped.setdefault(source_key, []).append(item)

    wide_rows: list[dict[str, Any]] = []
    coverage: dict[str, int] = {}
    for source_key, group in grouped.items():
        translations: dict[str, dict[str, Any]] = {}
        for code in LANGUAGE_ORDER:
            candidates = [item for item in group if normalize_language(item.get("language") or "en") == code and (str(item.get("target") or "").strip() or str(item.get("target_alt") or "").strip())]
            if not candidates:
                continue
            selected = sorted(candidates, key=lambda item: str(item.get("updated_at") or ""), reverse=True)[0]
            payload = {
                "id": selected.get("id", ""),
                "language": code,
                "target": selected.get("target", ""),
                "target_alt": selected.get("target_alt", ""),
            }
            translations[code] = payload
            coverage[code] = coverage.get(code, 0) + 1
        shared = {field: _first_non_blank(group, field) for field in shared_fields}
        wide_rows.append(
            {
                "source_key": source_key,
                "source": _first_non_blank(group, "source"),
                **shared,
                "translations": translations,
                "languages": [code for code in LANGUAGE_ORDER if code in translations],
                "conflicts": _wide_conflicts(group, ("source", *shared_fields)),
            }
        )

    languages = [code for code in LANGUAGE_ORDER if coverage.get(code, 0) > 0]
    wide_rows.sort(key=lambda row: (str(row.get("source") or ""), str(row.get(key_field) or "")))
    return {"languages": languages, "coverage": {code: coverage[code] for code in languages}, "row_count": len(wide_rows), "rows": wide_rows}



def _first_non_blank(rows: list[dict[str, Any]], field: str) -> str:
    for row in rows:
        value = str(row.get(field) or "").strip()
        if value and value != "-":
            return value
    return ""


def _wide_conflicts(rows: list[dict[str, Any]], fields: tuple[str, ...]) -> list[dict[str, Any]]:
    conflicts: list[dict[str, Any]] = []
    for field in fields:
        values: list[str] = []
        for row in rows:
            value = str(row.get(field) or "").strip()
            if value and value != "-" and value not in values:
                values.append(value)
        if len(values) > 1:
            conflicts.append({"field": field, "values": values})
    return conflicts



def _multilingual_glossary_layout(
    headers: list[str],
    *,
    term_key_column: str | None,
    source_column: str | None,
    category_column: str | None,
    note_column: str | None,
) -> tuple[int | None, int, int | None, int | None, dict[str, tuple[int, int | None]]]:
    normalized = _normalized_header_indices(headers)
    term_key_idx = _column_index(normalized, term_key_column, ["id", "key", "编号", "序号"], required=False)
    source_idx = _column_index(normalized, source_column, list(SOURCE_HEADER_ALIASES))
    category_idx = _column_index(normalized, category_column, ["category", "type", "分类", "类别", "类型"], required=False)
    note_idx = _column_index(normalized, note_column, ["note", "notes", "comment", "备注"], required=False)
    reserved = {index for index in (term_key_idx, source_idx, category_idx, note_idx) if index is not None}
    return term_key_idx, source_idx, category_idx, note_idx, _auto_language_indices(headers, reserved)


def _parse_multilingual_glossary_matrix(
    headers: list[str],
    raw_rows: Any,
    *,
    term_key_column: str | None,
    source_column: str | None,
    category_column: str | None,
    note_column: str | None,
    limit: int | None,
) -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
    term_key_idx, source_idx, category_idx, note_idx, language_indices = _multilingual_glossary_layout(
        headers,
        term_key_column=term_key_column,
        source_column=source_column,
        category_column=category_column,
        note_column=note_column,
    )
    if not language_indices:
        return [], {}, []
    rows: list[dict[str, Any]] = []
    source_rows = 0
    for row in raw_rows:
        source = _value_at(row, source_idx)
        if not source:
            continue
        source_rows += 1
        if limit is not None and source_rows > limit:
            break
        for code, (target_idx, alt_idx) in language_indices.items():
            target = _primary_target(_value_at(row, target_idx), _value_at(row, alt_idx))
            if not target:
                continue
            rows.append(
                {
                    "term_key": _value_at(row, term_key_idx) if term_key_idx is not None else "",
                    "source": source,
                    "target": target,
                    "target_alt": "",
                    "language": code,
                    "category": _value_at(row, category_idx) if category_idx is not None else "",
                    "note": _value_at(row, note_idx) if note_idx is not None else "",
                }
            )
    columns = {
        "term_key": headers[term_key_idx] if term_key_idx is not None else "",
        "source": headers[source_idx],
        "languages": {code: {"target": headers[target_idx], "target_alt": ""} for code, (target_idx, _alt_idx) in language_indices.items()},
        "category": headers[category_idx] if category_idx is not None else "",
        "note": headers[note_idx] if note_idx is not None else "",
    }
    languages = [code for code in LANGUAGE_ORDER if any(row.get("language") == code for row in rows)]
    return rows, columns, languages


def _mapping_pick(row: dict[str, Any], explicit: str | None, aliases: list[str]) -> tuple[str, str]:
    normalized = {str(key or "").strip().lower(): (str(key or "").strip(), value) for key, value in row.items()}
    names = [explicit] if explicit else aliases
    for name in names:
        if not name:
            continue
        hit = normalized.get(name.strip().lower())
        if hit and hit[1] not in (None, ""):
            return str(hit[1]).strip(), hit[0]
    return "", ""


def _read_multilingual_glossary_json_rows(
    path: Path,
    *,
    term_key_column: str | None,
    source_column: str | None,
    category_column: str | None,
    note_column: str | None,
    limit: int | None,
) -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
    mappings = _read_json_mapping_rows(path, ("terms", "rows", "entries"))
    if not any(str(row.get("language") or "").strip() for row in mappings):
        headers, raw_rows = _mapping_rows_to_matrix(mappings)
        return _parse_multilingual_glossary_matrix(
            headers,
            raw_rows,
            term_key_column=term_key_column,
            source_column=source_column,
            category_column=category_column,
            note_column=note_column,
            limit=limit,
        )
    rows: list[dict[str, Any]] = []
    columns: dict[str, Any] = {"term_key": "", "source": "", "languages": {}, "category": "", "note": ""}
    for mapping in mappings:
        raw_language = str(mapping.get("language") or "").strip()
        if not raw_language:
            continue
        try:
            code = require_supported_language(raw_language)
        except ValueError:
            continue
        source, source_header = _mapping_pick(mapping, source_column, list(SOURCE_HEADER_ALIASES))
        target, target_header = _mapping_pick(mapping, None, ["target", *target_aliases(code)])
        legacy_alt, _ = _mapping_pick(mapping, None, ["target_alt", *alt_aliases(code)])
        target = _primary_target(target, legacy_alt)
        if not source or not target:
            continue
        if limit is not None and len(rows) >= limit:
            break
        term_key, term_key_header = _mapping_pick(mapping, term_key_column, ["term_key", "id", "key", "编号", "序号"])
        category, category_header = _mapping_pick(mapping, category_column, ["category", "type", "分类", "类别", "类型"])
        note, note_header = _mapping_pick(mapping, note_column, ["note", "notes", "comment", "备注"])
        rows.append(
            {
                "term_key": term_key,
                "source": source,
                "target": target,
                "target_alt": "",
                "language": code,
                "category": category,
                "note": note,
            }
        )
        columns["term_key"] = columns["term_key"] or term_key_header
        columns["source"] = columns["source"] or source_header
        columns["category"] = columns["category"] or category_header
        columns["note"] = columns["note"] or note_header
        columns["languages"].setdefault(code, {"target": target_header, "target_alt": ""})
    languages = [code for code in LANGUAGE_ORDER if any(row.get("language") == code for row in rows)]
    return rows, columns, languages


def _read_multilingual_glossary_rows(
    path: Path,
    sheet: str | None = None,
    term_key_column: str | None = None,
    source_column: str | None = None,
    category_column: str | None = None,
    note_column: str | None = None,
    limit: int | None = 100,
) -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv", ".json"))
    if suffix == ".csv":
        headers, raw_rows = _read_csv_matrix(path)
        return _parse_multilingual_glossary_matrix(
            headers,
            raw_rows,
            term_key_column=term_key_column,
            source_column=source_column,
            category_column=category_column,
            note_column=note_column,
            limit=limit,
        )
    if suffix == ".json":
        return _read_multilingual_glossary_json_rows(
            path,
            term_key_column=term_key_column,
            source_column=source_column,
            category_column=category_column,
            note_column=note_column,
            limit=limit,
        )
    if suffix not in XLSX_IMPORT_SUFFIXES:
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv", ".json"))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        def is_candidate(worksheet: Any) -> bool:
            headers = _sheet_headers(worksheet)
            try:
                _, source_idx, _, _, language_indices = _multilingual_glossary_layout(
                    headers,
                    term_key_column=term_key_column,
                    source_column=source_column,
                    category_column=category_column,
                    note_column=note_column,
                )
            except (KeyError, ValueError):
                return False
            return bool(language_indices) and _worksheet_has_data(
                worksheet,
                source_idx,
                [index for pair in language_indices.values() for index in pair],
            )

        worksheet = _select_xlsx_data_sheet(workbook, sheet, is_candidate, allow_none=True)
        if worksheet is None:
            return [], {}, []
        return _parse_multilingual_glossary_matrix(
            _sheet_headers(worksheet),
            worksheet.iter_rows(min_row=2, values_only=True),
            term_key_column=term_key_column,
            source_column=source_column,
            category_column=category_column,
            note_column=note_column,
            limit=limit,
        )
    finally:
        workbook.close()


def _translation_layout(
    headers: list[str],
    *,
    id_column: str | None,
    source_column: str | None,
    target_column: str | None,
    target_alt_column: str | None,
    note_column: str | None,
    language: str,
) -> tuple[int | None, int, int, int | None, int | None]:
    normalized = _normalized_header_indices(headers)
    id_idx = _column_index(normalized, id_column, ["id", "key", "编号", "序号"], required=False)
    source_idx = _column_index(normalized, source_column, list(SOURCE_HEADER_ALIASES))
    target_idx = _column_index(normalized, target_column, target_aliases(language), required=bool(target_column))
    target_alt_idx = _column_index(normalized, target_alt_column, alt_aliases(language), required=False)
    if target_idx is None and target_alt_idx is not None:
        target_idx, target_alt_idx = target_alt_idx, None
    if target_idx is None:
        raise KeyError(f"target column not found for language {language}")
    note_idx = _column_index(normalized, note_column, ["note", "notes", "comment", "备注"], required=False)
    return id_idx, source_idx, target_idx, target_alt_idx, note_idx


def _parse_translation_matrix(
    headers: list[str],
    raw_rows: Any,
    *,
    sheet_name: str,
    id_column: str | None,
    source_column: str | None,
    target_column: str | None,
    target_alt_column: str | None,
    note_column: str | None,
    language: str,
    source_artifact_id: str,
    source_type: str,
) -> list[dict[str, Any]]:
    id_idx, source_idx, target_idx, target_alt_idx, note_idx = _translation_layout(
        headers,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        target_alt_column=target_alt_column,
        note_column=note_column,
        language=language,
    )
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(raw_rows, start=2):
        source = _value_at(row, source_idx)
        target = _primary_target(_value_at(row, target_idx), _value_at(row, target_alt_idx))
        if not source or not target:
            continue
        rows.append(
            {
                "entry_key": _value_at(row, id_idx) if id_idx is not None else "",
                "source": source,
                "target": target,
                "target_alt": "",
                "language": language,
                "sheet": sheet_name,
                "row_number": row_index,
                "note": _value_at(row, note_idx) if note_idx is not None else "",
                "source_type": source_type,
                "source_artifact_id": source_artifact_id,
            }
        )
    return rows


def _multilingual_translation_layout(
    headers: list[str],
    *,
    id_column: str | None,
    source_column: str | None,
    note_column: str | None,
) -> tuple[int | None, int, int | None, dict[str, tuple[int, int | None]]]:
    normalized = _normalized_header_indices(headers)
    id_idx = _column_index(normalized, id_column, ["id", "key", "编号", "序号"], required=False)
    source_idx = _column_index(normalized, source_column, list(SOURCE_HEADER_ALIASES))
    note_idx = _column_index(normalized, note_column, ["note", "notes", "comment", "备注"], required=False)
    reserved = {index for index in (id_idx, source_idx, note_idx) if index is not None}
    return id_idx, source_idx, note_idx, _auto_language_indices(headers, reserved)


def _parse_multilingual_translation_matrix(
    headers: list[str],
    raw_rows: Any,
    *,
    sheet_name: str,
    id_column: str | None,
    source_column: str | None,
    note_column: str | None,
    source_artifact_id: str,
    source_type: str,
) -> list[dict[str, Any]]:
    return _parse_multilingual_translation_table(
        headers,
        raw_rows,
        sheet_name=sheet_name,
        id_column=id_column,
        source_column=source_column,
        note_column=note_column,
        source_artifact_id=source_artifact_id,
        source_type=source_type,
    ).rows


def _parse_multilingual_translation_table(
    headers: list[str],
    raw_rows: Any,
    *,
    sheet_name: str,
    id_column: str | None,
    source_column: str | None,
    note_column: str | None,
    source_artifact_id: str,
    source_type: str,
    include_empty: bool = False,
) -> ParsedImportTable:
    id_idx, source_idx, note_idx, language_indices = _multilingual_translation_layout(
        headers,
        id_column=id_column,
        source_column=source_column,
        note_column=note_column,
    )
    detected_columns: dict[str, Any] = {
        "id": {"name": headers[id_idx], "index": id_idx} if id_idx is not None else None,
        "source": {"name": headers[source_idx], "index": source_idx},
        "note": {"name": headers[note_idx], "index": note_idx} if note_idx is not None else None,
        "languages": {
            code: {
                "target": {"name": headers[target_idx], "index": target_idx},
                "legacy_alt": {"name": headers[alt_idx], "index": alt_idx} if alt_idx is not None else None,
            }
            for code, (target_idx, alt_idx) in language_indices.items()
        },
    }
    if not language_indices:
        return ParsedImportTable(rows=[], detected_columns=detected_columns, sheet=sheet_name, include_empty=include_empty)
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(raw_rows, start=2):
        source = _value_at(row, source_idx)
        if not source:
            continue
        for code, (target_idx, alt_idx) in language_indices.items():
            target = _primary_target(_value_at(row, target_idx), _value_at(row, alt_idx))
            if not target and not include_empty:
                continue
            parsed_row = {
                "entry_key": _value_at(row, id_idx) if id_idx is not None else "",
                "source": source,
                "target": target,
                "target_alt": "",
                "language": code,
                "sheet": sheet_name,
                "row_number": row_index,
                "note": _value_at(row, note_idx) if note_idx is not None else "",
                "source_type": source_type,
                "source_artifact_id": source_artifact_id,
            }
            if include_empty:
                parsed_row["target_column_present"] = True
            rows.append(parsed_row)
    return ParsedImportTable(rows=rows, detected_columns=detected_columns, sheet=sheet_name, include_empty=include_empty)


def _read_multilingual_translation_rows(
    path: Path,
    sheet: str | None = None,
    id_column: str | None = None,
    source_column: str | None = None,
    note_column: str | None = None,
    source_artifact_id: str = "",
    source_type: str = "imported",
) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv"))
    if suffix == ".csv":
        headers, raw_rows = _read_csv_matrix(path)
        return _parse_multilingual_translation_matrix(
            headers,
            raw_rows,
            sheet_name="",
            id_column=id_column,
            source_column=source_column,
            note_column=note_column,
            source_artifact_id=source_artifact_id,
            source_type=source_type,
        )
    if suffix not in XLSX_IMPORT_SUFFIXES:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        def is_candidate(worksheet: Any) -> bool:
            headers = _sheet_headers(worksheet)
            try:
                _, source_idx, _, language_indices = _multilingual_translation_layout(
                    headers,
                    id_column=id_column,
                    source_column=source_column,
                    note_column=note_column,
                )
            except (KeyError, ValueError):
                return False
            return bool(language_indices) and _worksheet_has_data(
                worksheet,
                source_idx,
                [index for pair in language_indices.values() for index in pair],
            )

        worksheet = _select_xlsx_data_sheet(workbook, sheet, is_candidate, allow_none=True)
        if worksheet is None:
            return []
        return _parse_multilingual_translation_matrix(
            _sheet_headers(worksheet),
            worksheet.iter_rows(min_row=2, values_only=True),
            sheet_name=worksheet.title,
            id_column=id_column,
            source_column=source_column,
            note_column=note_column,
            source_artifact_id=source_artifact_id,
            source_type=source_type,
        )
    finally:
        workbook.close()


def _read_translation_rows(
    path: Path,
    sheet: str | None = None,
    id_column: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
    target_alt_column: str | None = None,
    note_column: str | None = None,
    language: str = "en",
    source_artifact_id: str = "",
    source_type: str = "imported",
) -> list[dict[str, Any]]:
    language = require_supported_language(language)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv"))
    if suffix == ".json":
        raw_rows = _read_json_mapping_rows(path, ("entries", "rows"))
        rows = []
        for row in raw_rows:
            normalized = {str(key or "").strip().lower(): value for key, value in row.items()}
            parsed = _translation_row_from_mapping(
                normalized,
                int(row.get("row_number") or 0),
                str(row.get("sheet") or "").strip(),
                language,
                source_artifact_id,
                source_type,
            )
            if parsed["source"] and parsed["target"]:
                rows.append(parsed)
        return rows
    if suffix == ".csv":
        headers, raw_rows = _read_csv_matrix(path)
        return _parse_translation_matrix(
            headers,
            raw_rows,
            sheet_name="",
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            target_alt_column=target_alt_column,
            note_column=note_column,
            language=language,
            source_artifact_id=source_artifact_id,
            source_type=source_type,
        )
    if suffix not in XLSX_IMPORT_SUFFIXES:
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv"))

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        def is_candidate(worksheet: Any) -> bool:
            headers = _sheet_headers(worksheet)
            try:
                _, source_idx, target_idx, target_alt_idx, _ = _translation_layout(
                    headers,
                    id_column=id_column,
                    source_column=source_column,
                    target_column=target_column,
                    target_alt_column=target_alt_column,
                    note_column=note_column,
                    language=language,
                )
            except (KeyError, ValueError):
                return False
            return _worksheet_has_data(worksheet, source_idx, (target_idx, target_alt_idx))

        worksheet = _select_xlsx_data_sheet(workbook, sheet, is_candidate)
        return _parse_translation_matrix(
            _sheet_headers(worksheet),
            worksheet.iter_rows(min_row=2, values_only=True),
            sheet_name=worksheet.title,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            target_alt_column=target_alt_column,
            note_column=note_column,
            language=language,
            source_artifact_id=source_artifact_id,
            source_type=source_type,
        )
    finally:
        workbook.close()


def _translation_row_from_mapping(
    row: dict[str, Any],
    row_number: int,
    sheet: str,
    language: str,
    source_artifact_id: str,
    source_type: str,
) -> dict[str, Any]:
    language = require_supported_language(language)

    def pick(*names: str) -> str:
        for name in names:
            value = row.get(name.lower())
            if value not in (None, ""):
                return str(value).strip()
        return ""

    target = _primary_target(pick("target", *target_aliases(language)), pick("target_alt", *alt_aliases(language)))
    return {
        "entry_key": pick("id", "key", "entry_key", "编号", "序号"),
        "source": pick(*SOURCE_HEADER_ALIASES),
        "target": target,
        "target_alt": "",
        "language": language,
        "sheet": sheet,
        "row_number": row_number,
        "note": pick("note", "notes", "comment", "备注"),
        "source_type": source_type,
        "source_artifact_id": source_artifact_id,
    }


def _translation_export_payload(entry: dict[str, Any]) -> dict[str, Any]:
    return {
        "entry_key": entry.get("entry_key", ""),
        "source": entry.get("source", ""),
        "target": entry.get("target", ""),
        "target_alt": entry.get("target_alt", ""),
        "language": entry.get("language", "en"),
        "note": entry.get("note", ""),
    }


def _translation_export_row(entry: dict[str, Any], *, include_alt: bool = True) -> list[Any]:
    row = [
        entry.get("entry_key", ""),
        entry.get("source", ""),
        entry.get("target", ""),
    ]
    if include_alt:
        row.append(entry.get("target_alt", ""))
    row.append(entry.get("note", ""))
    return row


def _translation_wide_export_rows(wide: dict[str, Any], languages: list[str]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in wide["rows"]:
        translations = row.get("translations") or {}
        values = [row.get("entry_key", ""), row.get("source", "")]
        for code in languages:
            entry = translations.get(code) or {}
            values.append(entry.get("target", ""))
        values.append(row.get("note", ""))
        rows.append(values)
    return rows
