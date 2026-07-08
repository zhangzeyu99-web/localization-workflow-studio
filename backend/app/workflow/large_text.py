"""Product-side large-text helpers.

This module intentionally keeps behavioral parity with the local harness
source of truth at ``workflow/localization/utils/large_text_multilingual_gate.py``.
That file remains the canonical implementation for preflight sizing, cache
lint parsing rules (numbers, word multipliers, CJK filtering, machine-like
bracket tokens) and readback checks used by local agent workflows. This
module ports the same pure logic so the product backend does not depend on
cross-tree imports or local-agent file-system assumptions; parity is
verified by ``backend/tests/test_large_text_productization.py`` which
compares representative rows against the workflow gate module directly.

Do not let this module drift from the workflow gate without updating the
parity tests. If the two disagree, the workflow gate wins.
"""
from __future__ import annotations

import json
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

WORKFLOW_VERSION = "large_text_product_v1"
ALLOWED_MODES = {"auto", "strict", "off"}

TOKEN_RE = re.compile(r"\\n|\{[^{}\s]+\}|%[sdif]|##\d+|</?[A-Za-z][^>\s]*[^>]*>|\[[A-Za-z0-9_:/#=.,-]+\]")
CJK_RE = re.compile(r"[\u3400-\u9fff]")
NUMBER_RE = re.compile(
    r"\d+(?:[,.]\d+)?(?:\s*(?:千|万|萬|亿|億|(?i:thousand|million|billion|ribu|rb|juta|miliar|millones|millón|milhao|milhão|milhões|mil)\b)|[KkMBWw](?![A-Za-z]))%?"
    r"|\d{1,3}(?:[,\s.]\d{3})+(?:[,.]\d+)?%?"
    r"|\d+(?:[,.]\d+)?%?"
)
WORD_MULTIPLIERS = {
    "thousand": Decimal("1000"),
    "ribu": Decimal("1000"),
    "rb": Decimal("1000"),
    "mil": Decimal("1000"),
    "million": Decimal("1000000"),
    "juta": Decimal("1000000"),
    "millones": Decimal("1000000"),
    "millón": Decimal("1000000"),
    "milhao": Decimal("1000000"),
    "milhão": Decimal("1000000"),
    "milhões": Decimal("1000000"),
    "billion": Decimal("1000000000"),
    "miliar": Decimal("1000000000"),
}
NUMBER_WORDS = {
    "zero": Decimal("0"),
    "one": Decimal("1"),
    "once": Decimal("1"),
    "single": Decimal("1"),
    "two": Decimal("2"),
    "three": Decimal("3"),
    "four": Decimal("4"),
    "five": Decimal("5"),
    "six": Decimal("6"),
    "seven": Decimal("7"),
    "eight": Decimal("8"),
    "nine": Decimal("9"),
    "ten": Decimal("10"),
    "satu": Decimal("1"),
    "sekali": Decimal("1"),
    "dua": Decimal("2"),
    "tiga": Decimal("3"),
    "empat": Decimal("4"),
    "lima": Decimal("5"),
    "seis": Decimal("6"),
    "tujuh": Decimal("7"),
    "delapan": Decimal("8"),
    "sembilan": Decimal("9"),
    "sepuluh": Decimal("10"),
    "uno": Decimal("1"),
    "una": Decimal("1"),
    "un": Decimal("1"),
    "dos": Decimal("2"),
    "tres": Decimal("3"),
    "cuatro": Decimal("4"),
    "cinco": Decimal("5"),
    "siete": Decimal("7"),
    "ocho": Decimal("8"),
    "nueve": Decimal("9"),
    "diez": Decimal("10"),
    "um": Decimal("1"),
    "uma": Decimal("1"),
    "dois": Decimal("2"),
    "duas": Decimal("2"),
    "quatro": Decimal("4"),
    "sete": Decimal("7"),
    "oito": Decimal("8"),
    "nove": Decimal("9"),
    "dez": Decimal("10"),
}
CJK_ALLOWED_LANGS = {"cn", "zh", "zh-cn", "ja", "jp"}
SOURCE_HEADERS = {"CN", "ZH", "SOURCE", "TEXT", "原文"}
# Sheets emitted by the local QA harness (workflow/localization/process_language.py)
# alongside the primary translation sheet. They carry review/reference columns
# only (no full target-language coverage) and must not be treated as delivery
# content when checking for missing target columns or blank target cells.
REVIEW_ONLY_SHEET_TITLES = {"需确认", "术语行筛选"}


def normalize_large_text_mode(value: str | None) -> str:
    mode = str(value or "auto").strip().lower()
    if mode not in ALLOWED_MODES:
        return "auto"
    return mode


def row_key(row: dict[str, Any], fallback: int) -> str:
    return str(row.get("key") or row.get("id") or row.get("para_id") or fallback)


def source_text(row: dict[str, Any]) -> str:
    return str(row.get("cn") or row.get("CN") or row.get("source") or "")


def row_translation(row: dict[str, Any], lang: str) -> str:
    translations = row.get("translations")
    if isinstance(translations, dict):
        value = translations.get(lang) or translations.get(lang.upper())
        return "" if value is None else str(value)
    value = row.get(lang) or row.get(lang.upper())
    return "" if value is None else str(value)


def explicit_translation_langs(row: dict[str, Any]) -> set[str]:
    translations = row.get("translations")
    if isinstance(translations, dict):
        return {str(key).lower() for key in translations}
    return set()


def build_large_text_preflight(
    rows: list[dict[str, Any]],
    *,
    target_languages: list[str],
    source_rows: int | None = None,
    workbook_count: int = 1,
    full_proofread: bool = False,
) -> dict[str, Any]:
    target_langs = [str(lang).strip().lower() for lang in target_languages if str(lang).strip()]
    unique_keys = {row_key(row, index) for index, row in enumerate(rows, 1)}
    long_text_items = [row for row in rows if int(row.get("char_len") or len(source_text(row))) > 300]
    estimated_cells = len(unique_keys) * len(target_langs)
    reasons: list[str] = []
    if len(unique_keys) > 5000:
        reasons.append("unique_items>5000")
    if len(target_langs) > 4:
        reasons.append("target_languages>4")
    if workbook_count > 1:
        reasons.append("workbook_count>1")
    if long_text_items:
        reasons.append("long_text_items>0")
    if full_proofread:
        reasons.append("full_proofread_requested")
    recommended_shards = 1
    if reasons:
        recommended_shards = max(2, min(8, (estimated_cells // 25000) + 2))
    return {
        "workflow": WORKFLOW_VERSION,
        "unique_items": len(unique_keys),
        "source_rows": source_rows,
        "target_languages": target_langs,
        "target_language_count": len(target_langs),
        "estimated_target_cells": estimated_cells,
        "long_text_items": len(long_text_items),
        "workbook_count": workbook_count,
        "large_pack": bool(reasons),
        "large_pack_reasons": reasons,
        "recommended_translation_shards": recommended_shards,
        "recommended_deep_proofread_shards": max(recommended_shards, 4) if full_proofread else recommended_shards,
    }


def is_auto_protected_token(token: str) -> bool:
    if token.startswith("[") and token.endswith("]"):
        inner = token[1:-1]
        return bool(re.search(r"[\d_:/#=.,-]", inner) or (inner.isupper() and len(inner) <= 12))
    return True


def protected_tokens(row: dict[str, Any]) -> list[str]:
    tokens: list[str] = []
    raw_tokens = row.get("tokens") or row.get("protected_tokens") or []
    if isinstance(raw_tokens, str):
        try:
            parsed = json.loads(raw_tokens)
            raw_tokens = parsed if isinstance(parsed, list) else [raw_tokens]
        except json.JSONDecodeError:
            raw_tokens = [raw_tokens]
    if isinstance(raw_tokens, list):
        tokens.extend(str(token) for token in raw_tokens if str(token))
    tokens.extend(token for token in TOKEN_RE.findall(source_text(row)) if is_auto_protected_token(token))
    return sorted(set(tokens), key=len, reverse=True)


def parse_number_token(token: str) -> Decimal | None:
    raw_with_spaces = token.strip()
    raw = raw_with_spaces.replace(" ", "")
    if not raw:
        return None

    if raw.endswith("%"):
        raw = raw[:-1]
        raw_with_spaces = raw_with_spaces[:-1].strip()

    suffix = ""
    word_multiplier: Decimal | None = None
    lowered = raw_with_spaces.lower()
    for word, multiplier in sorted(WORD_MULTIPLIERS.items(), key=lambda item: len(item[0]), reverse=True):
        match = re.search(rf"\s+{re.escape(word)}\.?$", lowered)
        if match:
            word_multiplier = multiplier
            raw = raw_with_spaces[: match.start()].strip().replace(" ", "")
            break
    if raw and raw[-1] in "KkMBWw":
        suffix = raw[-1].upper()
        raw = raw[:-1]
    elif raw.endswith(("千", "万", "萬", "亿", "億")):
        suffix = raw[-1]
        raw = raw[:-1]
    if not raw:
        return None

    has_comma = "," in raw
    has_dot = "." in raw
    has_unit = bool(suffix or word_multiplier)
    if has_unit and (has_comma or has_dot):
        if has_comma and has_dot:
            last_comma = raw.rfind(",")
            last_dot = raw.rfind(".")
            last_sep = max(last_comma, last_dot)
            after = raw[last_sep + 1 :]
            before = re.sub(r"[,.]", "", raw[:last_sep])
            raw = before + "." + after
        elif has_comma:
            raw = raw.replace(",", ".")
    elif has_comma or has_dot:
        last_comma = raw.rfind(",")
        last_dot = raw.rfind(".")
        last_sep = max(last_comma, last_dot)
        sep = raw[last_sep]
        after = raw[last_sep + 1 :]
        before = raw[:last_sep]
        thousands_like = len(after) == 3 and all(group.isdigit() for group in re.split(r"[,.]", before) if group)
        if has_comma and has_dot:
            if thousands_like:
                raw = re.sub(r"[,.]", "", raw)
            else:
                raw = re.sub(r"[,.]", "", before) + "." + after
        elif sep == "," and not thousands_like:
            raw = before.replace(",", "") + "." + after
        elif thousands_like:
            raw = re.sub(r"[,.]", "", raw)
        elif sep == ".":
            raw = before.replace(".", "") + "." + after

    try:
        multiplier = word_multiplier or {
            "千": Decimal("1000"),
            "万": Decimal("10000"),
            "萬": Decimal("10000"),
            "亿": Decimal("100000000"),
            "億": Decimal("100000000"),
            "K": Decimal("1000"),
            "M": Decimal("1000000"),
            "B": Decimal("1000000000"),
            "W": Decimal("10000"),
        }.get(suffix, Decimal(1))
        return (Decimal(raw) * multiplier).normalize()
    except InvalidOperation:
        return None


def numeric_values(text: str) -> set[Decimal]:
    text = text or ""
    text = re.sub(r"(\d(?:[\d,.]*))(?:<[^>]+>)+\s*([千万萬亿億KkMBWw])", r"\1\2", text)
    text = re.sub(r"(?<=\d)\uff0c(?=\d{3}(?!\d))", ",", text)
    text = text.replace("\uff0c", " ")
    values: set[Decimal] = set()
    for token in NUMBER_RE.findall(text):
        parsed = parse_number_token(token)
        if parsed is not None:
            values.add(parsed)
    lowered = text.lower()
    for word, value in NUMBER_WORDS.items():
        if re.search(rf"\b{re.escape(word)}\b", lowered):
            values.add(value)
    return values


def source_numeric_values(row: dict[str, Any]) -> set[Decimal]:
    src = source_text(row)
    src = re.sub(r"\d+(?:[,.]\d+)?\s*月", "", src)
    src = re.sub(r"(?<=\d)\.(?=\d点)", " ", src)
    values = numeric_values(src)
    if CJK_RE.search(src):
        values = {
            value
            for value in values
            if not (
                value == value.to_integral_value()
                and Decimal("0") <= value <= Decimal("10")
                and not re.search(rf"(?<!\d){int(value)}\s*%", src)
            )
        }
    return values


def numeric_value_present(value: Decimal, targets: set[Decimal]) -> bool:
    if value in targets:
        return True
    if abs(value) >= Decimal("1000"):
        for target in targets:
            if target == 0:
                continue
            if abs(target - value) / abs(value) <= Decimal("0.005"):
                return True
    if Decimal("1") <= value <= Decimal("99"):
        if (Decimal("1900") + value) in targets or (Decimal("2000") + value) in targets:
            return True
    return False


def add_issue(issues: list[dict[str, Any]], issue_type: str, key: str, lang: str, detail: str) -> None:
    issues.append({"severity": "hard", "type": issue_type, "key": key, "lang": lang, "detail": detail})


def _term_hits(row: dict[str, Any]) -> list[dict[str, Any]]:
    raw = row.get("term_hits") or row.get("term_hits_json") or []
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            return []
        raw = parsed
    if not isinstance(raw, list):
        return []
    return [hit for hit in raw if isinstance(hit, dict)]


def _accepted_term_variants(hit: dict[str, Any], lang: str) -> list[str]:
    values: list[str] = []
    translations = hit.get("translations")
    if isinstance(translations, dict) and translations.get(lang):
        values.append(str(translations[lang]))
    if hit.get(lang):
        values.append(str(hit[lang]))

    variants = hit.get("accepted_variants") or hit.get("variants")
    if isinstance(variants, dict):
        raw = variants.get(lang) or []
    else:
        raw = variants or []
    if isinstance(raw, str):
        values.append(raw)
    elif isinstance(raw, list):
        values.extend(str(value) for value in raw if str(value))

    return sorted({value for value in values if value}, key=len, reverse=True)


def _check_required_terms(issues: list[dict[str, Any]], row: dict[str, Any], key: str, lang: str, target: str) -> None:
    for hit in _term_hits(row):
        if not (hit.get("required") or hit.get("strict")):
            continue
        variants = _accepted_term_variants(hit, lang)
        if variants and not any(variant in target for variant in variants):
            source = hit.get("source") or hit.get("CN") or hit.get("term") or ""
            add_issue(issues, "term_missing", key, lang, f"required term not used: {source}")


def build_translation_cache_rows(
    workpack_rows: list[dict[str, Any]],
    translated_rows: list[dict[str, Any]],
    language: str,
) -> list[dict[str, Any]]:
    by_id = {str(row.get("id")): str(row.get("translation") or "") for row in translated_rows}
    lang = language.lower()
    cache_rows: list[dict[str, Any]] = []
    for index, row in enumerate(workpack_rows, 1):
        key = row_key(row, index)
        item = dict(row)
        item["key"] = key
        item["source"] = source_text(row)
        item["translations"] = {lang: by_id.get(str(row.get("id")), "")}
        cache_rows.append(item)
    return cache_rows


def cache_lint_rows(cache_rows: list[dict[str, Any]], *, target_languages: list[str]) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    seen: set[str] = set()
    unauthorized: Counter[str] = Counter()
    langs = [lang.lower() for lang in target_languages]
    requested = set(langs)
    for index, row in enumerate(cache_rows, 1):
        key = row_key(row, index)
        if key in seen:
            add_issue(issues, "duplicate_key", key, "", "cache contains duplicate source key")
        seen.add(key)
        extras = explicit_translation_langs(row) - requested
        for lang in sorted(extras):
            unauthorized[lang] += 1
            add_issue(issues, "unauthorized_language", key, lang, "translation cache contains a language that was not requested")

        src_numbers = source_numeric_values(row)
        tokens = protected_tokens(row)
        for lang in langs:
            target = row_translation(row, lang).strip()
            if not target:
                add_issue(issues, "empty_translation", key, lang, "target translation is empty")
                continue
            if lang not in CJK_ALLOWED_LANGS and CJK_RE.search(target):
                add_issue(issues, "cjk_residue", key, lang, "target translation still contains Chinese/Japanese ideographs")
            for token in tokens:
                if token and token not in target:
                    add_issue(issues, "protected_token_missing", key, lang, f"missing protected token {token}")
            target_numbers = numeric_values(target)
            missing_numbers = {number for number in src_numbers if not numeric_value_present(number, target_numbers)}
            if missing_numbers:
                add_issue(issues, "number_missing", key, lang, f"missing numeric value(s): {sorted(str(value) for value in missing_numbers)}")
            _check_required_terms(issues, row, key, lang, target)
    by_type = Counter(issue["type"] for issue in issues)
    return {
        "workflow": WORKFLOW_VERSION,
        "checked_items": len(cache_rows),
        "target_languages": langs,
        "hard_blockers": len(issues),
        "hard_by_type": dict(sorted(by_type.items())),
        "unauthorized_languages": dict(sorted(unauthorized.items())),
        "issues": issues,
        "ok_to_apply": len(issues) == 0,
    }


def _looks_like_translation_sheet(headers: list[str], target_langs: list[str]) -> bool:
    header_set = {header for header in headers if header}
    if header_set.intersection(set(target_langs)):
        return True
    return bool(header_set.intersection(SOURCE_HEADERS))


def readback_gate_files(paths: list[Path], *, target_languages: list[str]) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    files: list[dict[str, Any]] = []
    targets = [lang.upper() for lang in target_languages]
    for path in paths:
        path = Path(path)
        files.append({"name": path.name, "bytes": path.stat().st_size if path.exists() else 0})
        if not path.exists():
            add_issue(issues, "delivery_file_missing", path.name, "", "delivery file does not exist")
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                if sheet.title in REVIEW_ONLY_SHEET_TITLES:
                    continue
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [str(value).strip().upper() if value is not None else "" for value in first_row]
                if not _looks_like_translation_sheet(headers, targets):
                    continue
                col_by_lang = {header: index for index, header in enumerate(headers) if header}
                target_columns: list[tuple[str, int]] = []
                for lang in targets:
                    col_index = col_by_lang.get(lang)
                    if col_index is None:
                        add_issue(issues, "target_column_missing", f"{path.name}:{sheet.title}", lang, "target language column is missing")
                        continue
                    target_columns.append((lang, col_index))
                if not target_columns:
                    continue
                for row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                    for lang, col_index in target_columns:
                        value = row_values[col_index] if col_index < len(row_values) else None
                        if value is None or str(value).strip() == "":
                            add_issue(issues, "blank_target_cell", f"{path.name}:{sheet.title}!R{row_index}C{col_index + 1}", lang, "target cell is blank")
        finally:
            workbook.close()
    by_type = Counter(issue["type"] for issue in issues)
    return {
        "workflow": WORKFLOW_VERSION,
        "files": files,
        "target_languages": target_languages,
        "hard_blockers": len(issues),
        "hard_by_type": dict(sorted(by_type.items())),
        "issues": issues,
        "readback_verified": len(issues) == 0,
    }


LONG_TASK_REVIEW_SECONDS = 3600


def _gate_line(label: str, gate: dict[str, Any]) -> str:
    status = str(gate.get("status") or ("passed" if int(gate.get("hard_blockers") or 0) == 0 else "failed"))
    reason = f", reason={gate.get('reason')}" if status in {"skipped", "waived"} and gate.get("reason") else ""
    return f"- {label}: status={status}{reason}, hard={gate.get('hard_blockers', 'n/a')}"


def render_large_text_retro(metrics: dict[str, Any]) -> str:
    progress = metrics.get("translation_progress") or {}
    elapsed = int(progress.get("elapsed_seconds") or 0)
    long_status = "triggered" if elapsed >= LONG_TASK_REVIEW_SECONDS else "not_triggered"
    return f"""# 大文本处理复盘

## 执行规模

- unique_items: {(metrics.get("preflight") or {}).get("unique_items", "n/a")}
- estimated_target_cells: {(metrics.get("preflight") or {}).get("estimated_target_cells", "n/a")}
- total_rows: {progress.get("total_rows", "n/a")}
- completed_rows: {progress.get("completed_rows", "n/a")}

## 执行门禁结果

{_gate_line("cache-lint", metrics.get("cache_lint") or {})}
{_gate_line("readback-gate", metrics.get("readback_gate") or {})}

## 长任务复盘触发

- status={long_status}, threshold=3600s, elapsed={elapsed}s
- review_focus=判断耗时是否只是任务规模导致；检查失败/重试/跳过门禁/意外修复；重复出现或可机器检查的问题沉淀为测试、gate 或文档，偶发问题只记录。
"""


__all__ = [name for name in globals() if not name.startswith("_")]
