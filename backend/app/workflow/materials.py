from __future__ import annotations

import csv
import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook

from .. import db
from ..config import REAL_PROVIDERS, normalize_provider_name
from ..languages import visible_language_code
from ..providers import call_image_text
from ..translation_batches import cap_context_text as _cap_context_text
from .common import project_dir
from .project_analysis import _project_brief_from_text
from .semantic_qa import _parse_semantic_qa_payload
from .subprocess_runner import user_facing_error
from .table_helpers import _auto_language_indices, _column_index, _normalized_header_indices, _value_at

PROJECT_TEXT_MATERIAL_EXTENSIONS = {".md", ".markdown", ".txt"}
PROJECT_DOCX_MATERIAL_EXTENSIONS = {".docx"}
PROJECT_PDF_MATERIAL_EXTENSIONS = {".pdf"}
PROJECT_TABLE_MATERIAL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
PROJECT_DELIMITED_MATERIAL_EXTENSIONS = {".csv", ".tsv"}
PROJECT_JSON_MATERIAL_EXTENSIONS = {".json", ".jsonl"}
PROJECT_IMAGE_MATERIAL_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
PROJECT_VIDEO_MATERIAL_EXTENSIONS = {".mp4", ".mov", ".mkv", ".avi"}
PROJECT_AUDIO_MATERIAL_EXTENSIONS = {".mp3", ".wav", ".m4a"}


def analyze_assets(artifact_ids: list[str], settings: dict[str, Any]) -> list[str]:
    packet = build_project_material_packet("", artifact_ids, settings, run_visual_analysis=False)
    return [str(material.get("note") or f"{material.get('label')}={material.get('status')}") for material in packet.get("materials", [])]


def _text_asset_note(path: Path) -> str:
    try:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
    except OSError:
        return "text_material:read_failed"
    compact = re.sub(r"\s+", " ", text).strip()
    if not compact:
        return "text_material:empty"
    return f"text_material:{compact[:800]}"


def build_project_material_packet(
    project_id: str,
    artifact_ids: list[str],
    settings: dict[str, Any],
    *,
    run_visual_analysis: bool = True,
) -> dict[str, Any]:
    materials: list[dict[str, Any]] = []
    context_parts: list[str] = []
    warnings: list[str] = []
    language_table_candidates: list[dict[str, Any]] = []
    for artifact_id in artifact_ids:
        artifact = db.get_artifact(artifact_id)
        material = _parse_project_material_artifact(artifact, settings, run_visual_analysis=run_visual_analysis, project_id=project_id)
        materials.append(material)
        if material.get("excerpt"):
            context_parts.append(f"## {material.get('label')}\n{material.get('excerpt')}")
        if material.get("language_table_candidate"):
            language_table_candidates.append({
                "artifact_id": artifact_id,
                "label": material.get("label"),
                "languages": material.get("detected_languages", []),
                "rows": material.get("rows", 0),
            })
        if material.get("warning"):
            warnings.append(str(material["warning"]))
    parsed = [m for m in materials if str(m.get("status") or "").startswith(("parsed", "vision_analyzed"))]
    image_materials = [m for m in materials if m.get("material_type") == "image"]
    brief_materials = [m for m in materials if m.get("project_brief_candidate")]
    unsupported = [m for m in materials if str(m.get("status") or "").startswith(("archived_only", "unsupported"))]
    return {
        "version": 1,
        "source": "project_material_packet",
        "project_id": project_id,
        "artifact_ids": artifact_ids,
        "materials": materials,
        "context": _cap_context_text("\n\n".join(context_parts), 9000, "project material packet"),
        "summary": {
            "total": len(materials),
            "parsed": len(parsed),
            "unsupported": len(unsupported),
            "images": len(image_materials),
            "project_brief_candidates": len(brief_materials),
            "language_table_candidates": len(language_table_candidates),
            "warnings": len(warnings),
        },
        "language_table_candidates": language_table_candidates,
        "warnings": warnings,
    }


def _parse_project_material_artifact(artifact: dict[str, Any], settings: dict[str, Any], *, run_visual_analysis: bool, project_id: str = "") -> dict[str, Any]:
    path = Path(artifact["path"])
    suffix = path.suffix.lower()
    base = {
        "artifact_id": artifact.get("id"),
        "label": artifact.get("label") or path.name,
        "filename": path.name,
        "created_at": artifact.get("created_at", ""),
        "kind": artifact.get("kind", ""),
        "role": artifact.get("role", ""),
        "origin": artifact.get("origin", ""),
        "size": artifact.get("size", 0),
        "suffix": suffix,
        "status": "parsed",
        "note": "",
        "excerpt": "",
    }
    try:
        if suffix in PROJECT_TEXT_MATERIAL_EXTENSIONS:
            result = _parse_project_text_material(path)
        elif suffix in PROJECT_DOCX_MATERIAL_EXTENSIONS:
            result = _parse_project_docx_material(path)
        elif suffix in PROJECT_PDF_MATERIAL_EXTENSIONS:
            result = _parse_project_pdf_material(path, settings)
        elif suffix in PROJECT_TABLE_MATERIAL_EXTENSIONS:
            result = _parse_project_xlsx_material(path)
        elif suffix in PROJECT_DELIMITED_MATERIAL_EXTENSIONS:
            result = _parse_project_delimited_material(path)
        elif suffix in PROJECT_JSON_MATERIAL_EXTENSIONS:
            result = _parse_project_json_material(path)
        elif suffix in PROJECT_IMAGE_MATERIAL_EXTENSIONS:
            result = _parse_project_image_material(path, settings, run_visual_analysis=run_visual_analysis)
        elif suffix in PROJECT_VIDEO_MATERIAL_EXTENSIONS:
            result = _parse_project_video_material(path, settings, run_visual_analysis=run_visual_analysis, project_id=project_id, artifact_id=str(artifact.get("id") or "video"))
        elif suffix in PROJECT_AUDIO_MATERIAL_EXTENSIONS:
            result = {"material_type": "audio", "status": "archived_only:audio_analysis_not_supported", "warning": f"{path.name} 已归档，v1 未进行音频内容分析。", "excerpt": f"音频资料：{path.name}。v1 未分析内容。"}
        else:
            result = {"material_type": "unknown", "status": "archived_only:unknown_type", "warning": f"{path.name} 文件类型暂未分析。", "excerpt": f"未分析资料：{path.name}"}
    except Exception as exc:
        result = {"material_type": "unknown", "status": "parse_failed", "warning": f"{path.name} 解析失败：{user_facing_error(exc)}", "excerpt": ""}
    item = {**base, **result}
    item["note"] = _project_material_note(item)
    return item


def _project_material_note(material: dict[str, Any]) -> str:
    label = material.get("label") or material.get("filename") or "material"
    status = material.get("status") or "parsed"
    if material.get("language_table_candidate"):
        langs = "/".join(str(x) for x in material.get("detected_languages") or [])
        return f"{label}=language_table_candidate:{material.get('rows', 0)} rows:{langs}"
    if material.get("material_type") == "image":
        return f"{label}={status}:{str(material.get('visual_summary') or material.get('excerpt') or '')[:300]}"
    return f"{label}={status}:{str(material.get('excerpt') or '')[:300]}"


def _parse_project_text_material(path: Path) -> dict[str, Any]:
    text = _read_lookup_text_file(path)
    compact = re.sub(r"\s+", " ", text).strip()
    headings = re.findall(r"(?m)^#{1,4}\s+(.+)$", text)[:20]
    table_rows = re.findall(r"(?m)^\|.+\|$", text)[:40]
    excerpt = "\n".join(part for part in [
        "\n".join(f"# {heading}" for heading in headings[:8]),
        "\n".join(table_rows[:20]),
        compact[:3000],
    ] if part).strip()
    result: dict[str, Any] = {"material_type": "text", "status": "parsed_text", "chars": len(text), "headings": headings, "table_rows": table_rows[:20], "excerpt": excerpt}
    project_brief = _project_brief_from_text(text) if path.suffix.lower() in {".md", ".markdown"} else None
    if project_brief:
        result["project_brief"] = project_brief
        result["project_brief_candidate"] = True
        result["status"] = "parsed_project_brief"
    return result


def _parse_project_docx_material(path: Path) -> dict[str, Any]:
    with ZipFile(path) as archive:
        if "word/document.xml" not in archive.namelist():
            raise ValueError(f"Missing word/document.xml in DOCX file: {path}")
        root = ET.fromstring(archive.read("word/document.xml"))
    paragraphs: list[str] = []
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    for paragraph in root.findall(".//w:p", ns):
        text = re.sub(r"\s+", " ", "".join(node.text or "" for node in paragraph.findall(".//w:t", ns))).strip()
        if text:
            paragraphs.append(text)
    return {"material_type": "docx", "status": "parsed_docx", "paragraphs": len(paragraphs), "excerpt": "\n".join(paragraphs[:120])[:5000]}


def _parse_project_pdf_material(path: Path, settings: dict[str, Any]) -> dict[str, Any]:
    if not (settings.get("multimodal", {}) or {}).get("pdf", True):
        return {"material_type": "pdf", "status": "archived_only:pdf_not_supported", "warning": f"{path.name} 已归档，当前设置未启用 PDF 文字提取。", "excerpt": ""}
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as exc:
        return {"material_type": "pdf", "status": "archived_only:pdf_parser_missing", "warning": f"缺少 PDF 解析依赖：{exc}", "excerpt": ""}
    reader = PdfReader(str(path))
    chunks: list[str] = []
    for page in reader.pages[:20]:
        text = re.sub(r"\s+", " ", page.extract_text() or "").strip()
        if text:
            chunks.append(text)
    return {"material_type": "pdf", "status": "parsed_pdf_text", "pages": len(reader.pages), "excerpt": "\n".join(chunks)[:5000]}


def _parse_project_xlsx_material(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    detected_languages: set[str] = set()
    total_rows = 0
    samples: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets[:12]:
            rows_iter = ws.iter_rows(values_only=True)
            first = next(rows_iter, None)
            headers = [str(value or "").strip() for value in (first or [])]
            normalized = _normalized_header_indices(headers)
            id_idx = _column_index(normalized, None, ["id", "key", "编号", "序号"], required=False)
            source_idx = _column_index(normalized, None, ["source", "original", "cn", "zh", "chinese", "原文", "中文", "简体中文"], required=False)
            reserved = {idx for idx in (id_idx, source_idx) if idx is not None}
            language_indices = _auto_language_indices(headers, reserved)
            detected_languages.update(language_indices.keys())
            row_count = 0
            for row in rows_iter:
                if any(str(cell or "").strip() for cell in row):
                    row_count += 1
                if len(samples) < 30 and source_idx is not None:
                    source = _value_at(row, source_idx)
                    if source:
                        item = {"sheet": ws.title, "source": source}
                        for code, (target_idx, _alt_idx) in language_indices.items():
                            target = _value_at(row, target_idx)
                            if target:
                                item[visible_language_code(code)] = target
                        samples.append(item)
            total_rows += row_count
            sheets.append({"name": ws.title, "headers": headers[:40], "rows": row_count, "source_column": headers[source_idx] if source_idx is not None and source_idx < len(headers) else "", "languages": sorted(visible_language_code(code) for code in language_indices)})
    finally:
        wb.close()
    language_table_candidate = total_rows > 0 and bool(detected_languages) and any(sheet.get("source_column") for sheet in sheets)
    excerpt = json.dumps({"sheets": sheets[:8], "samples": samples[:20]}, ensure_ascii=False, indent=2)
    return {
        "material_type": "table",
        "status": "parsed_xlsx",
        "rows": total_rows,
        "sheets": sheets,
        "samples": samples,
        "detected_languages": sorted(visible_language_code(code) for code in detected_languages),
        "language_table_candidate": language_table_candidate,
        "excerpt": excerpt[:7000],
    }


def _parse_project_delimited_material(path: Path) -> dict[str, Any]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    raw = _read_lookup_text_file(path)
    rows = list(csv.reader(raw.splitlines(), delimiter=delimiter))
    headers = [str(value or "").strip() for value in rows[0]] if rows else []
    samples = rows[1:31]
    return {"material_type": "table", "status": "parsed_delimited", "rows": max(0, len(rows) - 1), "headers": headers, "excerpt": json.dumps({"headers": headers, "samples": samples}, ensure_ascii=False)[:4000]}


def _parse_project_json_material(path: Path) -> dict[str, Any]:
    text = _read_lookup_text_file(path)
    try:
        payload = json.loads(text)
        excerpt = json.dumps(payload, ensure_ascii=False, indent=2)[:5000]
    except Exception:
        excerpt = text[:5000]
    return {"material_type": "json", "status": "parsed_json", "excerpt": excerpt}


def _parse_project_image_material(path: Path, settings: dict[str, Any], *, run_visual_analysis: bool) -> dict[str, Any]:
    if not run_visual_analysis:
        return {"material_type": "image", "status": "archived_only:image_not_requested", "excerpt": f"图片资料：{path.name}"}
    provider = normalize_provider_name(settings.get("provider"))
    if provider not in REAL_PROVIDERS or not settings.get("api_key"):
        return {"material_type": "image", "status": "archived_only:image_api_key_missing", "warning": f"{path.name} 已归档，未配置支持图片分析的 API。", "excerpt": f"图片资料：{path.name}；未配置 API，未做视觉分析。"}
    prompt = (
        "请分析这张游戏本地化项目资料图。返回严格 JSON："
        "{\"ui_type\":\"\", \"theme\":\"\", \"worldview\":\"\", \"characters_or_scene\":\"\", "
        "\"visible_text\":\"\", \"localization_style_hints\":\"\", \"confidence\":\"low|medium|high\"}。"
        "只基于图片可见信息，不要编造。"
    )
    try:
        raw = call_image_text(settings, prompt, path, system="Return strict JSON only.")
        payload = _parse_semantic_qa_payload(raw)
        if not isinstance(payload, dict):
            raise ValueError("image provider returned non-object")
        summary = "；".join(str(payload.get(key) or "").strip() for key in ("ui_type", "theme", "worldview", "characters_or_scene", "visible_text", "localization_style_hints") if str(payload.get(key) or "").strip())
        return {"material_type": "image", "status": "vision_analyzed", "visual_analysis": payload, "visual_summary": summary, "excerpt": summary[:2500]}
    except Exception as exc:
        return {"material_type": "image", "status": "vision_failed", "warning": f"{path.name} 图片分析失败：{user_facing_error(exc)}", "excerpt": f"图片资料：{path.name}；视觉分析失败。"}


def _extract_video_keyframes(path: Path, output_dir: Path, *, max_frames: int = 4) -> tuple[list[Path], dict[str, Any]]:
    try:
        import cv2  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"缺少视频抽帧依赖 OpenCV：{exc}") from exc
    output_dir.mkdir(parents=True, exist_ok=True)
    cap = cv2.VideoCapture(str(path))
    if not cap.isOpened():
        raise RuntimeError("视频无法打开，可能是格式不受支持或文件损坏")
    try:
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
        fps = float(cap.get(cv2.CAP_PROP_FPS) or 0)
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
        if frame_count > 0:
            ratios = [0.08, 0.32, 0.58, 0.84][:max_frames]
            indices = sorted({min(frame_count - 1, max(0, int(frame_count * ratio))) for ratio in ratios})
        else:
            indices = list(range(max_frames))
        frames: list[Path] = []
        for order, index in enumerate(indices, start=1):
            if frame_count > 0:
                cap.set(cv2.CAP_PROP_POS_FRAMES, index)
            ok, frame = cap.read()
            if not ok or frame is None:
                continue
            frame_path = output_dir / f"frame_{order:02d}.png"
            if cv2.imwrite(str(frame_path), frame):
                frames.append(frame_path)
        if not frames:
            raise RuntimeError("没有抽取到可用关键帧")
        return frames, {"frame_count": frame_count, "fps": fps, "width": width, "height": height, "frames": [frame.name for frame in frames]}
    finally:
        cap.release()


def _parse_project_video_material(path: Path, settings: dict[str, Any], *, run_visual_analysis: bool, project_id: str, artifact_id: str) -> dict[str, Any]:
    if not run_visual_analysis:
        return {"material_type": "video", "status": "archived_only:video_not_requested", "excerpt": f"视频资料：{path.name}"}
    provider = normalize_provider_name(settings.get("provider"))
    if provider not in REAL_PROVIDERS or not settings.get("api_key"):
        return {"material_type": "video", "status": "archived_only:video_api_key_missing", "warning": f"{path.name} 已归档，未配置支持视频关键帧分析的 API。", "excerpt": f"视频资料：{path.name}；未配置 API，未做画面分析。"}
    frame_root = (project_dir(project_id) / "profile" / "video_frames" / artifact_id) if project_id else (path.parent / ".lws_video_frames" / artifact_id)
    try:
        frames, frame_meta = _extract_video_keyframes(path, frame_root, max_frames=4)
    except Exception as exc:
        return {"material_type": "video", "status": "video_frame_extract_failed", "warning": f"{path.name} 抽帧失败：{user_facing_error(exc)}", "excerpt": f"视频资料：{path.name}；抽帧失败。"}
    prompt = (
        "请分析这个游戏本地化项目视频的关键帧。返回严格 JSON："
        "{\"ui_type\":\"\", \"theme\":\"\", \"worldview\":\"\", \"characters_or_scene\":\"\", "
        "\"visible_text\":\"\", \"localization_style_hints\":\"\", \"confidence\":\"low|medium|high\"}。"
        "只基于当前帧可见信息，不要编造；如果文字看不清，visible_text 写空。"
    )
    analyses: list[dict[str, Any]] = []
    warnings: list[str] = []
    for frame in frames:
        try:
            raw = call_image_text(settings, prompt, frame, system="Return strict JSON only.")
            payload = _parse_semantic_qa_payload(raw)
            if not isinstance(payload, dict):
                payload = {"summary": str(raw)[:1000]}
            payload["frame"] = frame.name
            analyses.append(payload)
        except Exception as exc:
            warnings.append(f"{frame.name}: {user_facing_error(exc)}")
    if not analyses:
        return {"material_type": "video", "status": "vision_failed_video", "warning": f"{path.name} 关键帧视觉分析失败：{'；'.join(warnings[:3])}", "frame_meta": frame_meta, "excerpt": f"视频资料：{path.name}；关键帧分析失败。"}
    summary_parts: list[str] = []
    for payload in analyses:
        summary = "；".join(str(payload.get(key) or "").strip() for key in ("ui_type", "theme", "worldview", "characters_or_scene", "visible_text", "localization_style_hints") if str(payload.get(key) or "").strip())
        if summary:
            summary_parts.append(f"{payload.get('frame')}: {summary}")
    warning = f"{path.name} 部分关键帧分析失败：{'；'.join(warnings[:3])}" if warnings else ""
    return {
        "material_type": "video",
        "status": "vision_analyzed_video",
        "frames_analyzed": len(analyses),
        "frame_meta": frame_meta,
        "visual_analysis": analyses,
        "warning": warning,
        "excerpt": "\n".join(summary_parts)[:4000],
    }


def _read_lookup_text_file(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def _read_lookup_csv_text(path: Path) -> str:
    raw = _read_lookup_text_file(path)
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = csv.reader(raw.splitlines(), delimiter=delimiter)
    cells: list[str] = []
    for row in rows:
        cells.extend(str(cell).strip() for cell in row if str(cell).strip())
    return "\n".join(cells)


def _read_lookup_xlsx_text(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        cells: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells.extend(str(cell).strip() for cell in row if cell not in (None, ""))
        return "\n".join(cells)
    finally:
        wb.close()


def _read_lookup_docx_text(path: Path) -> str:
    with ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    texts = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
    return "\n".join(texts)


def _read_lookup_material_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".markdown", ".json"}:
        return _read_lookup_text_file(path)
    if suffix in {".csv", ".tsv"}:
        return _read_lookup_csv_text(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_lookup_xlsx_text(path)
    if suffix == ".docx":
        return _read_lookup_docx_text(path)
    return ""


def _compact_lookup_text(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()

__all__ = [name for name in globals() if not name.startswith("__")]
