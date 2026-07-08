from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .. import db
from ..config import load_settings
from ..languages import language_spec, require_supported_language
from ..translation_batches import (
    cap_context_text as _cap_context_text,
    estimate_text_tokens as _estimate_text_tokens,
    manage_project_prompt_context as _manage_project_prompt_context,
    project_context_summary as _project_context_summary,
)
from .common import GLOBAL_HARNESS_CONTRACT, _language_assets_summary, project_dir, read_project_harness, run_dir
from .asset_import_export import _glossary_export_row
from .table_helpers import _read_glossary_rows
from .materials import build_project_material_packet
from .project_analysis import (
    _apply_project_analysis_provider,
    _apply_project_brief_markdown_profile,
    _build_project_profile,
    _is_stale_project_prompt_text,
    _profile_material_summary,
    _project_analysis_report_markdown,
    _project_brief_markdown,
    _project_display_prompt_from_profile,
    _project_execution_prompt_from_profile,
    _save_generated_project_harness,
)
from .translation_readiness import _harness_summary, _project_harness_prompt_parts

def write_project_prompt(
    project: dict[str, Any],
    intro: str,
    asset_notes: list[str],
    target_language: str = "en",
    material_packet: dict[str, Any] | None = None,
    settings: dict[str, Any] | None = None,
) -> tuple[Path, Path, Path, Path, Path, str]:
    target_language = require_supported_language(target_language)
    root = project_dir(project["id"]) / "profile"
    root.mkdir(parents=True, exist_ok=True)
    # Load settings once for this whole call: both the material-packet build
    # (when the caller doesn't already have one) and the provider-backed
    # analysis below must observe the same snapshot.
    settings = settings if settings is not None else load_settings()
    material_packet = material_packet or build_project_material_packet(project["id"], [], settings, run_visual_analysis=False)
    profile = _build_project_profile(project, intro, asset_notes, target_language=target_language, material_packet=material_packet)
    profile = _apply_project_brief_markdown_profile(profile, material_packet) or _apply_project_analysis_provider(project, intro, asset_notes, profile, material_packet, settings=settings)
    profile["material_packet"] = _profile_material_summary(material_packet)
    if profile.get("brief_source") == "md_primary" and profile.get("brief_supplements"):
        existing_sources = profile.get("source_materials") if isinstance(profile.get("source_materials"), list) else []
        supplement_sources = (profile.get("brief_supplements") or {}).get("sources") or []
        profile["source_materials"] = list(dict.fromkeys([*existing_sources, *supplement_sources]))
    else:
        profile["source_materials"] = [material.get("label") for material in material_packet.get("materials", []) if isinstance(material, dict)]
    prompt = _project_execution_prompt_from_profile(profile)
    display_prompt = _project_display_prompt_from_profile(profile)
    _save_generated_project_harness(project, profile)
    profile_path = root / f"project_profile_{target_language}.json"
    prompt_path = root / f"translation_prompt_{target_language}.txt"
    brief_path = root / f"project_brief_{target_language}.md"
    packet_path = root / "project_material_packet.json"
    report_path = root / "project_analysis_report.md"
    profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    prompt_path.write_text(prompt, encoding="utf-8")
    brief_path.write_text(_project_brief_markdown(profile, display_prompt), encoding="utf-8")
    packet_path.write_text(json.dumps(material_packet, ensure_ascii=False, indent=2), encoding="utf-8")
    report_path.write_text(_project_analysis_report_markdown(profile, material_packet), encoding="utf-8")
    project_profile = dict(project.get("profile") or {})
    prompts = dict(project_profile.get("prompts_by_language") or {})
    prompts[target_language] = prompt
    display_prompts = dict(project_profile.get("display_prompts_by_language") or {})
    display_prompts[target_language] = display_prompt
    profiles = dict(project_profile.get("profiles_by_language") or {})
    profiles[target_language] = profile
    shared_profile_keys = (
        "game_type", "target_audience", "content_scope", "translation_style", "tone",
        "display_game_type", "display_target_audience", "display_content_scope", "display_worldview",
        "display_translation_style", "display_focus", "display_key_terms",
        "language_assets", "source_materials", "asset_notes", "material_packet", "analysis_source",
        "analysis_warning", "analysis_provider", "analysis_model", "confidence", "missing_info", "generated_date",
        "brief_source", "brief_artifact_id", "brief_prompt_text", "brief_meta", "brief_supplements",
    )
    project_profile.update({key: profile[key] for key in shared_profile_keys if key in profile})
    project_profile["prompts_by_language"] = prompts
    project_profile["display_prompts_by_language"] = display_prompts
    project_profile["profiles_by_language"] = profiles
    updates: dict[str, Any] = {"profile": project_profile}
    if target_language == "en" or not str(project.get("prompt_text") or "").strip():
        updates["prompt_text"] = prompt
    db.update_project(project["id"], updates)
    return profile_path, prompt_path, brief_path, packet_path, report_path, prompt


def _profile_for_prompt_repair(project: dict[str, Any], language: str) -> dict[str, Any]:
    spec = language_spec(language)
    profile = dict(project.get("profile") or {})
    profiles_by_language = profile.get("profiles_by_language") if isinstance(profile.get("profiles_by_language"), dict) else {}
    language_profile = profiles_by_language.get(language) if isinstance(profiles_by_language.get(language), dict) else {}
    repaired = {**profile, **language_profile}
    repaired.update(
        {
            "project_name": repaired.get("project_name") or project.get("name") or "当前项目",
            "project_type": repaired.get("project_type") or project.get("type") or "",
            "description": repaired.get("description") or project.get("description") or "",
            "game_type": repaired.get("game_type") or repaired.get("display_game_type") or project.get("type") or "游戏本地化项目",
            "target_audience": repaired.get("target_audience") or repaired.get("display_target_audience") or "目标语区游戏玩家",
            "content_scope": repaired.get("content_scope") or repaired.get("display_content_scope") or "UI、系统、任务、道具、活动和剧情文本",
            "translation_style": repaired.get("translation_style") or repaired.get("display_translation_style") or "短句清晰、自然准确，适合游戏 UI",
            "tone": repaired.get("tone") or repaired.get("display_worldview") or "",
            "target_language": language,
            "target_language_label": spec.label,
            "target_language_name": spec.prompt_name,
            "language_assets": repaired.get("language_assets") or _language_assets_summary(project["id"]),
            "generated_date": repaired.get("generated_date") or db.now_iso()[:10],
        }
    )
    return repaired


def _repair_stale_project_prompt(project: dict[str, Any], language: str, prompt_path: Path) -> str:
    profile = _profile_for_prompt_repair(project, language)
    prompt = _project_execution_prompt_from_profile(profile)
    display_prompt = _project_display_prompt_from_profile(profile)
    prompt_path.parent.mkdir(parents=True, exist_ok=True)
    prompt_path.write_text(prompt, encoding="utf-8")

    project_profile = dict(project.get("profile") or {})
    prompts = dict(project_profile.get("prompts_by_language") or {})
    display_prompts = dict(project_profile.get("display_prompts_by_language") or {})
    prompts[language] = prompt
    display_prompts[language] = display_prompt
    project_profile["prompts_by_language"] = prompts
    project_profile["display_prompts_by_language"] = display_prompts
    updates: dict[str, Any] = {"profile": project_profile}
    if language == "en":
        updates["prompt_text"] = prompt
    db.update_project(project["id"], updates)
    return prompt


def compile_project_harness_prompt(project: dict[str, Any], base_prompt: str, output_dir: Path) -> tuple[Path, Path, str, dict[str, Any]]:
    harness = read_project_harness(project["id"])
    parts = [base_prompt.strip()]
    project_parts = _project_harness_prompt_parts(harness)
    if project_parts:
        parts.append(
            "Project Harness (project-specific; apply only to this project, do not generalize):\n"
            + "\n".join(project_parts)
        )
    raw_compiled = "\n\n".join(part for part in parts if part)
    settings = load_settings()
    compiled = _manage_project_prompt_context(raw_compiled, settings)
    prompt_path = output_dir / "compiled_project_harness_prompt.txt"
    snapshot_path = output_dir / "project_harness_snapshot.json"
    snapshot = {
        "global_harness": GLOBAL_HARNESS_CONTRACT,
        "project_harness": harness,
        "summary": _harness_summary(harness),
        "context_budget": _project_context_summary(raw_compiled, settings),
    }
    prompt_path.write_text(compiled, encoding="utf-8")
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    return prompt_path, snapshot_path, compiled, snapshot


def create_project_glossary_snapshot(
    project_id: str,
    run_id: str,
    output_dir: Path | None = None,
    language: str = "en",
    extra_term_artifact_id: str | None = None,
) -> dict[str, Any]:
    language = require_supported_language(language)
    spec = language_spec(language)
    output = output_dir or run_dir(run_id) / "snapshots"
    output.mkdir(parents=True, exist_ok=True)
    path = output / ("project_glossary_snapshot.xlsx" if language == "en" else f"project_glossary_snapshot_{language}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Glossary"
    ws.append(["ID", "CN", spec.target_header, *(["EN2"] if spec.alt_header else []), "分类", "备注"])
    terms = db.list_glossary_terms(project_id, language=language)
    seen_sources: set[str] = set()
    for term in reversed(terms):
        ws.append(_glossary_export_row(term, include_alt=bool(spec.alt_header)))
        source_key = str(term.get("source") or "").strip()
        if source_key:
            seen_sources.add(source_key)
    extra_term_count = 0
    extra_term_rows_read = 0
    extra_artifact = None
    extra_term_error = ""
    if extra_term_artifact_id:
        try:
            candidate = db.get_artifact(extra_term_artifact_id)
            if candidate.get("project_id") == project_id and Path(str(candidate.get("path") or "")).exists():
                extra_artifact = candidate
                extra_rows, _ = _read_glossary_rows(Path(candidate["path"]), limit=None, language=language)
                extra_term_rows_read = len(extra_rows)
                for row in extra_rows:
                    source = str(row.get("source") or "").strip()
                    target = str(row.get("target") or "").strip()
                    if not source or not target or source in seen_sources:
                        continue
                    ws.append([
                        row.get("term_key", ""),
                        source,
                        target,
                        *( [row.get("target_alt", "")] if spec.alt_header else [] ),
                        row.get("category", ""),
                        row.get("note", ""),
                    ])
                    seen_sources.add(source)
                    extra_term_count += 1
        except Exception as exc:
            extra_term_error = str(exc)
            db.add_event(run_id, f"extra term artifact ignored: {exc}", level="warning")
    wb.save(path)
    wb.close()
    return db.add_artifact(
        project_id,
        "Project glossary snapshot",
        path,
        "glossary_snapshot",
        run_id=run_id,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        origin="generated",
        metadata={
            "term_count": len(terms) + extra_term_count,
            "project_term_count": len(terms),
            "extra_term_count": extra_term_count,
            "extra_term_rows_read": extra_term_rows_read,
            "extra_term_artifact_id": extra_artifact["id"] if extra_artifact else "",
            "extra_term_artifact_selected": bool(extra_artifact),
            "extra_term_error": extra_term_error,
            "source": "project_glossary_with_task_terms" if extra_artifact else "project_glossary",
            "language": language,
        },
    )


def create_prompt_and_harness_snapshots(project_id: str, run_id: str, output_dir: Path | None = None, language: str = "en") -> dict[str, Any]:
    language = require_supported_language(language)
    project = db.get_project(project_id)
    output = output_dir or run_dir(run_id) / "snapshots"
    output.mkdir(parents=True, exist_ok=True)
    profile = project.get("profile") or {}
    prompt_path = project_dir(project_id) / "profile" / f"translation_prompt_{language}.txt"
    stored_prompt = str((profile.get("prompts_by_language") or {}).get(language) or "")
    if stored_prompt.strip():
        prompt_path.parent.mkdir(parents=True, exist_ok=True)
        if not prompt_path.exists() or prompt_path.read_text(encoding="utf-8") != stored_prompt:
            prompt_path.write_text(stored_prompt, encoding="utf-8")
        base_prompt = stored_prompt
    else:
        if not prompt_path.exists():
            write_project_prompt(project, project.get("description", ""), [], target_language=language)
        base_prompt = prompt_path.read_text(encoding="utf-8")
    if _is_stale_project_prompt_text(base_prompt) or _is_stale_project_prompt_text(stored_prompt):
        base_prompt = _repair_stale_project_prompt(project, language, prompt_path)
    compiled_path, harness_path, compiled_prompt, harness_snapshot = compile_project_harness_prompt(project, base_prompt, output)
    prompt_artifact = db.add_artifact(
        project_id,
        "Prompt snapshot",
        compiled_path,
        "prompt_snapshot",
        run_id=run_id,
        mime="text/plain",
        origin="generated",
        metadata={"source": "project_prompt_and_harness", "language": language},
    )
    harness_artifact = db.add_artifact(
        project_id,
        "Project harness snapshot",
        harness_path,
        "project_harness_snapshot",
        run_id=run_id,
        mime="application/json",
        origin="generated",
        metadata={"source": "project_harness", "language": language},
    )
    return {
        "prompt": compiled_prompt,
        "prompt_artifact": prompt_artifact,
        "harness_artifact": harness_artifact,
        "harness_snapshot": harness_snapshot,
        "prompt_path": compiled_path,
        "harness_path": harness_path,
    }


def _quick_reference_excerpt(artifact: dict[str, Any], max_chars: int = 5000) -> str:
    path = Path(str(artifact.get("path") or ""))
    if not path.exists():
        return ""
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md", ".csv", ".json", ".jsonl"}:
            return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                lines: list[str] = []
                remaining = max_chars
                for ws in wb.worksheets[:3]:
                    lines.append(f"[Sheet] {ws.title}")
                    for row in ws.iter_rows(max_row=16, values_only=True):
                        text = " | ".join(str(cell).strip() for cell in row if cell not in (None, ""))
                        if not text:
                            continue
                        lines.append(text)
                        remaining -= len(text)
                        if remaining <= 0:
                            break
                    if remaining <= 0:
                        break
                return "\n".join(lines)[:max_chars]
            finally:
                wb.close()
    except Exception:
        return "[reference read failed]"
    return f"[binary reference: {path.name}]"


def create_quick_reference_snapshot(project_id: str, run_id: str, reference_artifact_ids: list[str] | None, output_dir: Path | None = None) -> dict[str, Any] | None:
    ids = [str(item).strip() for item in (reference_artifact_ids or []) if str(item).strip()]
    if not ids:
        return None
    output = output_dir or run_dir(run_id) / "snapshots"
    output.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, Any]] = []
    context_parts = [
        "Temporary reference material for this quick task only.",
        "Use it as style/term/context guidance for this run. Do not treat it as permanent project memory unless the user imports it separately.",
    ]
    for artifact_id in ids:
        artifact = db.get_artifact(artifact_id)
        if artifact["project_id"] != project_id:
            raise KeyError(artifact_id)
        excerpt = _quick_reference_excerpt(artifact)
        item = {
            "id": artifact["id"],
            "label": artifact.get("label", ""),
            "kind": artifact.get("kind", ""),
            "role": artifact.get("role", ""),
            "origin": artifact.get("origin", ""),
            "original_filename": (artifact.get("metadata") or {}).get("original_filename", ""),
            "sha256": (artifact.get("metadata") or {}).get("sha256", ""),
            "size": artifact.get("size", 0),
            "excerpt": excerpt,
        }
        rows.append(item)
        context_parts.append(f"\nReference: {item['original_filename'] or item['label']} ({item['kind']})\n{excerpt}")
    snapshot_path = output / "quick_reference_snapshot.json"
    snapshot = {"source": "quick_task_reference", "reference_artifact_ids": ids, "references": rows}
    snapshot_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    context = "\n".join(part for part in context_parts if part).strip()
    settings = load_settings()
    quick_ref_budget = max(500, int(settings.get("max_quick_reference_context_tokens") or 2000))
    managed_context = _cap_context_text(context, quick_ref_budget, "quick task references")
    context_summary = {
        "original_estimated_tokens": _estimate_text_tokens(context),
        "managed_estimated_tokens": _estimate_text_tokens(managed_context),
        "max_quick_reference_context_tokens": quick_ref_budget,
        "trimmed": managed_context != context,
    }
    artifact = db.add_artifact(
        project_id,
        "Quick task reference snapshot",
        snapshot_path,
        "quick_reference_snapshot",
        run_id=run_id,
        mime="application/json",
        origin="generated",
        metadata={"source": "quick_task_reference", "reference_artifact_ids": ids, "reference_count": len(rows), "context_budget": context_summary},
    )
    return {"artifact": artifact, "snapshot": snapshot, "context": managed_context, "context_budget": context_summary}

__all__ = [name for name in globals() if not name.startswith("__")]
