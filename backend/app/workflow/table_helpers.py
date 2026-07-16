from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

from ..languages import PROJECT_LANGUAGE_ORDER, SOURCE_HEADER_ALIASES, alt_aliases, normalize_language, require_supported_language, target_aliases

LANGUAGE_ORDER = PROJECT_LANGUAGE_ORDER
GENERIC_TARGET_ALIASES = frozenset({"target", "translation", "译文"})
AUTO_LANGUAGE_TARGET_ALIASES = {
    code: tuple(alias for alias in target_aliases(code) if alias.strip().lower() not in {"target", "translation", "译文"})
    for code in LANGUAGE_ORDER
}
AUTO_LANGUAGE_ALT_ALIASES = {code: tuple(alt_aliases(code)) for code in LANGUAGE_ORDER}
XLSX_IMPORT_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})
IGNORED_AUTO_SHEET_TITLES = frozenset({"填写说明", "说明", "notes", "readme", "instructions", "buckets"})


@dataclass(frozen=True)
class ParsedImportTable:
    rows: list[dict[str, Any]]
    detected_columns: dict[str, Any]
    sheet: str = ""
    include_empty: bool = False


class ImportContractError(ValueError):
    code = "invalid_import"

    def __init__(self, message: str, *, detail: str | dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.detail = detail if detail is not None else message


class SheetSelectionError(ImportContractError):
    code = "sheet_selection_required"

    def __init__(self, candidates: list[str]) -> None:
        self.candidates = list(candidates)
        message = "检测到多个可导入的数据工作表，请选择后重试。"
        super().__init__(
            f"{message} 候选：{', '.join(self.candidates)}",
            detail={"code": self.code, "message": message, "candidates": self.candidates},
        )


class UnsupportedImportFormatError(ImportContractError):
    code = "unsupported_import_format"

    def __init__(self, suffix: str, supported: tuple[str, ...]) -> None:
        normalized = suffix.upper().lstrip(".") or "UNKNOWN"
        supported_text = "/".join(item.upper().lstrip(".") for item in supported)
        message = f"不支持旧版 {normalized} 文件，请另存为 {supported_text} 后重试。"
        super().__init__(message)
        self.suffix = suffix
        self.supported = supported


def _wide_source_key(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).casefold()


def _normalized_header_indices(headers: list[str]) -> dict[str, int]:
    normalized: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = str(header or "").strip().lower()
        if key and key not in normalized:
            normalized[key] = index
    return normalized


def _column_index(normalized_headers: dict[str, int], explicit: str | None, candidates: list[str], required: bool = True) -> int | None:
    if explicit:
        hit = normalized_headers.get(explicit.strip().lower())
        if hit is not None:
            return hit
        raise KeyError(f"column not found: {explicit}")
    for candidate in candidates:
        hit = normalized_headers.get(candidate.lower())
        if hit is not None:
            return hit
    if required:
        raise KeyError(f"none of columns found: {', '.join(candidates)}")
    return None


def _value_at(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def _primary_target(target: Any, legacy_alt: Any = "") -> str:
    primary = str(target or "").strip()
    return primary or str(legacy_alt or "").strip()


def _sheet_headers(worksheet: Any) -> list[str]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value or "").strip() for value in header_row]


def _worksheet_has_data(worksheet: Any, source_index: int, target_indices: Iterable[int | None]) -> bool:
    targets = [index for index in target_indices if index is not None]
    if not targets:
        return False
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if _value_at(row, source_index) and any(_value_at(row, index) for index in targets):
            return True
    return False


def _worksheet_has_source_data(worksheet: Any, source_index: int) -> bool:
    return any(_value_at(row, source_index) for row in worksheet.iter_rows(min_row=2, values_only=True))


def _select_xlsx_data_sheet(
    workbook: Any,
    requested_sheet: str | None,
    is_candidate: Callable[[Any], bool],
    *,
    allow_none: bool = False,
) -> Any | None:
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise KeyError(f"sheet not found: {requested_sheet}")
        return workbook[requested_sheet]
    candidates = [
        worksheet
        for worksheet in workbook.worksheets
        if worksheet.title.strip().casefold() not in IGNORED_AUTO_SHEET_TITLES and is_candidate(worksheet)
    ]
    if len(candidates) > 1:
        raise SheetSelectionError([worksheet.title for worksheet in candidates])
    if candidates:
        return candidates[0]
    if allow_none:
        return None
    raise ValueError("no valid data sheet found")


def _read_csv_matrix(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        headers = [str(value or "").strip() for value in next(reader, [])]
        return headers, [tuple(row) for row in reader]


def _read_json_mapping_rows(path: Path, collection_keys: tuple[str, ...]) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, list):
        raw_rows = payload
    elif isinstance(payload, dict):
        raw_rows: Any = None
        for key in collection_keys:
            if isinstance(payload.get(key), list):
                raw_rows = payload[key]
                break
        if raw_rows is None:
            raw_rows = [payload]
    else:
        raise ValueError("JSON import payload must be an array or object")
    return [dict(row) for row in raw_rows if isinstance(row, dict)]


def _mapping_rows_to_matrix(rows: list[dict[str, Any]]) -> tuple[list[str], list[tuple[Any, ...]]]:
    headers: list[str] = []
    for row in rows:
        for key in row:
            label = str(key or "").strip()
            if label and label not in headers:
                headers.append(label)
    return headers, [tuple(row.get(header) for header in headers) for row in rows]


def _auto_language_indices(headers: list[str], reserved_indices: set[int] | None = None) -> dict[str, tuple[int, int | None]]:
    reserved = reserved_indices or set()
    normalized_headers: dict[str, int] = {}
    for index, header in enumerate(headers):
        if index in reserved:
            continue
        key = str(header or "").strip().lower()
        if key:
            normalized_headers[key] = index
    generic_target_present = any(alias in normalized_headers for alias in GENERIC_TARGET_ALIASES)
    detected: dict[str, tuple[int, int | None]] = {}
    for code in LANGUAGE_ORDER:
        target_idx = _column_index(normalized_headers, None, list(AUTO_LANGUAGE_TARGET_ALIASES[code]), required=False)
        alt_idx = None
        if code == "en":
            alt_idx = _column_index(normalized_headers, None, list(AUTO_LANGUAGE_ALT_ALIASES["en"]), required=False)
            if target_idx is None and alt_idx is not None and not generic_target_present:
                target_idx, alt_idx = alt_idx, None
        if target_idx is None:
            continue
        detected[code] = (target_idx, alt_idx)
    return detected


def _glossary_layout(
    headers: list[str],
    *,
    term_key_column: str | None,
    source_column: str | None,
    target_column: str | None,
    target_alt_column: str | None,
    category_column: str | None,
    note_column: str | None,
    language: str,
) -> tuple[dict[str, int], int | None, int, int, int | None, int | None, int | None]:
    normalized = _normalized_header_indices(headers)
    term_key_idx = _column_index(normalized, term_key_column, ["id", "key", "编号", "序号"], required=False)
    source_idx = _column_index(normalized, source_column, list(SOURCE_HEADER_ALIASES))
    target_idx = _column_index(normalized, target_column, target_aliases(language), required=bool(target_column))
    target_alt_idx = _column_index(normalized, target_alt_column, alt_aliases(language), required=False)
    if target_idx is None and target_alt_idx is not None:
        target_idx, target_alt_idx = target_alt_idx, None
    if target_idx is None:
        raise KeyError(f"target column not found for language {language}")
    category_idx = _column_index(normalized, category_column, ["category", "type", "分类", "类别", "类型"], required=False)
    note_idx = _column_index(normalized, note_column, ["note", "notes", "comment", "备注"], required=False)
    return normalized, term_key_idx, source_idx, target_idx, target_alt_idx, category_idx, note_idx


def _parse_glossary_matrix(
    headers: list[str],
    raw_rows: Iterable[tuple[Any, ...]],
    *,
    term_key_column: str | None,
    source_column: str | None,
    target_column: str | None,
    target_alt_column: str | None,
    category_column: str | None,
    note_column: str | None,
    language: str,
    limit: int | None,
    include_empty: bool,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    _, term_key_idx, source_idx, target_idx, target_alt_idx, category_idx, note_idx = _glossary_layout(
        headers,
        term_key_column=term_key_column,
        source_column=source_column,
        target_column=target_column,
        target_alt_column=target_alt_column,
        category_column=category_column,
        note_column=note_column,
        language=language,
    )
    rows: list[dict[str, Any]] = []
    for row in raw_rows:
        if limit is not None and len(rows) >= limit:
            break
        source = _value_at(row, source_idx)
        target = _primary_target(_value_at(row, target_idx), _value_at(row, target_alt_idx))
        if not source or (not target and not include_empty):
            continue
        rows.append(
            {
                "term_key": _value_at(row, term_key_idx) if term_key_idx is not None else "",
                "source": source,
                "target": target,
                "target_alt": "",
                "category": _value_at(row, category_idx) if category_idx is not None else "",
                "note": _value_at(row, note_idx) if note_idx is not None else "",
            }
        )
    return rows, {
        "term_key": headers[term_key_idx] if term_key_idx is not None else "",
        "source": headers[source_idx],
        "target": headers[target_idx],
        "target_alt": "",
        "category": headers[category_idx] if category_idx is not None else "",
        "note": headers[note_idx] if note_idx is not None else "",
    }


def _read_glossary_rows(
    path: Path,
    sheet: str | None = None,
    term_key_column: str | None = None,
    source_column: str | None = None,
    target_column: str | None = None,
    target_alt_column: str | None = None,
    category_column: str | None = None,
    note_column: str | None = None,
    language: str = "en",
    limit: int | None = 100,
    include_empty: bool = False,
    allow_header_only: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, str]]:
    language = require_supported_language(language)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv", ".json"))
    if suffix == ".csv":
        headers, raw_rows = _read_csv_matrix(path)
        return _parse_glossary_matrix(
            headers,
            raw_rows,
            term_key_column=term_key_column,
            source_column=source_column,
            target_column=target_column,
            target_alt_column=target_alt_column,
            category_column=category_column,
            note_column=note_column,
            language=language,
            limit=limit,
            include_empty=include_empty,
        )
    if suffix == ".json":
        mappings = _read_json_mapping_rows(path, ("terms", "rows", "entries"))
        mappings = [
            row
            for row in mappings
            if not str(row.get("language") or "").strip() or normalize_language(row.get("language")) == language
        ]
        headers, raw_rows = _mapping_rows_to_matrix(mappings)
        return _parse_glossary_matrix(
            headers,
            raw_rows,
            term_key_column=term_key_column,
            source_column=source_column,
            target_column=target_column,
            target_alt_column=target_alt_column,
            category_column=category_column,
            note_column=note_column,
            language=language,
            limit=limit,
            include_empty=include_empty,
        )
    if suffix not in XLSX_IMPORT_SUFFIXES:
        raise UnsupportedImportFormatError(suffix, (".xlsx", ".csv", ".json"))
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        def is_candidate(worksheet: Any) -> bool:
            headers = _sheet_headers(worksheet)
            try:
                _, _, source_idx, target_idx, target_alt_idx, _, _ = _glossary_layout(
                    headers,
                    term_key_column=term_key_column,
                    source_column=source_column,
                    target_column=target_column,
                    target_alt_column=target_alt_column,
                    category_column=category_column,
                    note_column=note_column,
                    language=language,
                )
            except (KeyError, ValueError):
                return False
            if allow_header_only:
                return True
            if include_empty:
                return _worksheet_has_source_data(worksheet, source_idx)
            return _worksheet_has_data(worksheet, source_idx, (target_idx, target_alt_idx))

        ws = _select_xlsx_data_sheet(wb, sheet, is_candidate)
        headers = _sheet_headers(ws)
        return _parse_glossary_matrix(
            headers,
            ws.iter_rows(min_row=2, values_only=True),
            term_key_column=term_key_column,
            source_column=source_column,
            target_column=target_column,
            target_alt_column=target_alt_column,
            category_column=category_column,
            note_column=note_column,
            language=language,
            limit=limit,
            include_empty=include_empty,
        )
    finally:
        wb.close()
