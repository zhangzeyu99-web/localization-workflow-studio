from __future__ import annotations

import hashlib
import json
import sqlite3
import unicodedata
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from . import db
from .archive_batch_engine import (
    ArchiveBatchError,
    ArchiveEntityAdapter,
    commit_archive_batch,
    list_archive_import_batches,
    rollback_archive_import_batch,
)
from .languages import normalize_language, require_supported_language, target_aliases
from .workflow.asset_import_export import (
    _multilingual_translation_layout,
    _parse_multilingual_translation_table,
    _translation_layout,
)
from .workflow.table_helpers import (
    LANGUAGE_ORDER,
    ParsedImportTable,
    SheetSelectionError,
    _mapping_rows_to_matrix,
    _primary_target,
    _read_csv_matrix,
    _read_json_mapping_rows,
    _select_xlsx_data_sheet,
    _sheet_headers,
    _value_at,
    _worksheet_has_source_data,
)


ARCHIVE_KIND = "translations"
CSV_SHEET_KEY = "__csv__"
JSON_SHEET_KEY = "__json__"
CHANGE_SAMPLE_LIMIT = 50
PROTECTED_TRANSLATION_SOURCES = frozenset(
    {
        "manual",
        "curated",
        "qa_passed",
        "qa_final",
        "delivered_with_issues",
        "archive",
        "translation_archive",
    }
)
TRUSTED_QA_SOURCES = frozenset({"qa_passed", "qa_final"})
TRANSLATION_FIELDS = (
    "id",
    "project_id",
    "entry_key",
    "source",
    "source_key",
    "target",
    "target_alt",
    "language",
    "sheet",
    "row_number",
    "note",
    "source_type",
    "source_artifact_id",
    "active",
    "dataset_key",
    "last_import_batch_id",
    "review_status",
    "created_at",
    "updated_at",
)


@dataclass(frozen=True)
class ParsedTranslationArtifact:
    rows: list[dict[str, Any]]
    columns: dict[str, Any]
    sheet: str
    languages: list[str]


def _json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _json_load(value: Any, default: Any) -> Any:
    try:
        return json.loads(str(value or ""))
    except (TypeError, ValueError, json.JSONDecodeError):
        return default


def _hash_json(value: Any) -> str:
    return hashlib.sha256(_json_dump(value).encode("utf-8")).hexdigest()


def _file_checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_key(value: Any) -> str:
    return "".join(str(value or "").split()).casefold()


def _is_protected_overwrite(before: dict[str, Any], incoming_source_type: str) -> bool:
    before_source = str(before.get("source_type") or "").strip().lower()
    incoming_source = str(incoming_source_type or "").strip().lower()
    if before_source == "delivered_with_issues" and incoming_source in TRUSTED_QA_SOURCES:
        return False
    return before_source in PROTECTED_TRANSLATION_SOURCES


def _sheet_key(value: str) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def _translation_row(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    payload = dict(row)
    return {field: payload.get(field) for field in TRANSLATION_FIELDS}


def _translation_hash(row: sqlite3.Row | dict[str, Any] | None) -> str:
    return _hash_json(_translation_row(row)) if row is not None else ""


def _state_checksum(rows: Iterable[sqlite3.Row | dict[str, Any]]) -> str:
    normalized = sorted((_translation_row(row) for row in rows), key=lambda row: str(row.get("id") or ""))
    return _hash_json(normalized)


def _request_payload(request: Any) -> dict[str, Any]:
    if hasattr(request, "model_dump"):
        return dict(request.model_dump())
    return {key: value for key, value in vars(request).items() if not key.startswith("_")}


def _selected_languages(request: Any) -> list[str]:
    requested = getattr(request, "languages", None) or []
    selected: list[str] = []
    for value in requested:
        language = require_supported_language(str(value))
        if language not in selected:
            selected.append(language)
    return [language for language in LANGUAGE_ORDER if language in selected]


def _single_language_table(
    headers: list[str],
    raw_rows: Iterable[tuple[Any, ...]],
    *,
    sheet_name: str,
    request: Any,
    source_artifact_id: str,
    source_type: str,
) -> ParsedImportTable:
    language = require_supported_language(getattr(request, "language", "en") or "en")
    id_idx, source_idx, target_idx, target_alt_idx, note_idx = _translation_layout(
        headers,
        id_column=getattr(request, "id_column", None),
        source_column=getattr(request, "source_column", None),
        target_column=getattr(request, "target_column", None),
        target_alt_column=getattr(request, "target_alt_column", None),
        note_column=getattr(request, "note_column", None),
        language=language,
    )
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(raw_rows, start=2):
        source = _value_at(row, source_idx)
        if not source:
            continue
        rows.append(
            {
                "entry_key": _value_at(row, id_idx) if id_idx is not None else "",
                "source": source,
                "target": _primary_target(_value_at(row, target_idx), _value_at(row, target_alt_idx)),
                "target_alt": "",
                "language": language,
                "sheet": sheet_name,
                "row_number": row_number,
                "note": _value_at(row, note_idx) if note_idx is not None else "",
                "source_type": source_type,
                "source_artifact_id": source_artifact_id,
                "target_column_present": True,
            }
        )
    columns = {
        "id": {"name": headers[id_idx], "index": id_idx} if id_idx is not None else None,
        "source": {"name": headers[source_idx], "index": source_idx},
        "note": {"name": headers[note_idx], "index": note_idx} if note_idx is not None else None,
        "languages": {
            language: {
                "target": {"name": headers[target_idx], "index": target_idx},
                "legacy_alt": {"name": headers[target_alt_idx], "index": target_alt_idx}
                if target_alt_idx is not None
                else None,
            }
        },
    }
    return ParsedImportTable(rows=rows, detected_columns=columns, sheet=sheet_name, include_empty=True)


def _auto_language_mode(request: Any) -> bool:
    return bool(getattr(request, "auto_languages", True)) and not getattr(request, "target_column", None) and not getattr(
        request, "target_alt_column", None
    )


def _guard_snapshot_identity_values(
    entry_key: Any,
    source: Any,
    *,
    sheet: str,
    row_number: int,
) -> None:
    if str(entry_key or "").strip() and not str(source or "").strip():
        raise ArchiveBatchError(
            422,
            "snapshot_source_required",
            "快照中存在 ID 非空但中文源文为空的行，无法安全判断归档身份。",
            extra={"sheet": sheet, "row": row_number},
        )


def _guard_snapshot_identity_rows(
    headers: list[str],
    raw_rows: Iterable[tuple[Any, ...]],
    *,
    sheet_name: str,
    request: Any,
) -> Iterable[tuple[Any, ...]]:
    if _auto_language_mode(request):
        id_idx, source_idx, _, _ = _multilingual_translation_layout(
            headers,
            id_column=getattr(request, "id_column", None),
            source_column=getattr(request, "source_column", None),
            note_column=getattr(request, "note_column", None),
        )
    else:
        id_idx, source_idx, _, _, _ = _translation_layout(
            headers,
            id_column=getattr(request, "id_column", None),
            source_column=getattr(request, "source_column", None),
            target_column=getattr(request, "target_column", None),
            target_alt_column=getattr(request, "target_alt_column", None),
            note_column=getattr(request, "note_column", None),
            language=require_supported_language(getattr(request, "language", "en") or "en"),
        )
    for row_number, row in enumerate(raw_rows, start=2):
        if id_idx is not None:
            _guard_snapshot_identity_values(
                _value_at(row, id_idx),
                _value_at(row, source_idx),
                sheet=sheet_name,
                row_number=row_number,
            )
        yield row


def _parse_matrix(
    headers: list[str],
    raw_rows: Iterable[tuple[Any, ...]],
    *,
    sheet_name: str,
    request: Any,
    source_artifact_id: str,
    source_type: str,
    guard_snapshot_identity: bool = True,
) -> ParsedImportTable:
    mode = str(getattr(request, "mode", "merge") or "merge").strip().lower()
    if guard_snapshot_identity and mode == "snapshot":
        raw_rows = _guard_snapshot_identity_rows(
            headers,
            raw_rows,
            sheet_name=sheet_name,
            request=request,
        )
    if _auto_language_mode(request):
        return _parse_multilingual_translation_table(
            headers,
            raw_rows,
            sheet_name=sheet_name,
            id_column=getattr(request, "id_column", None),
            source_column=getattr(request, "source_column", None),
            note_column=getattr(request, "note_column", None),
            source_artifact_id=source_artifact_id,
            source_type=source_type,
            include_empty=True,
        )
    return _single_language_table(
        headers,
        raw_rows,
        sheet_name=sheet_name,
        request=request,
        source_artifact_id=source_artifact_id,
        source_type=source_type,
    )


def _xlsx_candidate(worksheet: Any, request: Any, *, allow_id_only: bool = False) -> bool:
    headers = _sheet_headers(worksheet)
    try:
        if _auto_language_mode(request):
            id_idx, source_idx, _, language_indices = _multilingual_translation_layout(
                headers,
                id_column=getattr(request, "id_column", None),
                source_column=getattr(request, "source_column", None),
                note_column=getattr(request, "note_column", None),
            )
            has_identity_data = _worksheet_has_source_data(worksheet, source_idx) or (
                allow_id_only and id_idx is not None and _worksheet_has_source_data(worksheet, id_idx)
            )
            return bool(language_indices) and has_identity_data
        id_idx, source_idx, _, _, _ = _translation_layout(
            headers,
            id_column=getattr(request, "id_column", None),
            source_column=getattr(request, "source_column", None),
            target_column=getattr(request, "target_column", None),
            target_alt_column=getattr(request, "target_alt_column", None),
            note_column=getattr(request, "note_column", None),
            language=require_supported_language(getattr(request, "language", "en") or "en"),
        )
        return _worksheet_has_source_data(worksheet, source_idx) or (
            allow_id_only and id_idx is not None and _worksheet_has_source_data(worksheet, id_idx)
        )
    except (KeyError, ValueError):
        return False


def _guard_uncached_formulas(path: Path, table: ParsedImportTable, selected_languages: list[str]) -> None:
    cached = load_workbook(path, read_only=True, data_only=True)
    formulas = load_workbook(path, read_only=True, data_only=False)
    try:
        cached_sheet = cached[table.sheet]
        formula_sheet = formulas[table.sheet]
        guarded_indices: set[int] = set()
        identity_indices: dict[str, int] = {}
        for field in ("id", "source", "note"):
            column = table.detected_columns.get(field)
            if isinstance(column, dict) and column.get("index") is not None:
                index = int(column["index"])
                guarded_indices.add(index)
                identity_indices[field] = index
        for language in selected_languages:
            mapping = (table.detected_columns.get("languages") or {}).get(language) or {}
            for field in ("target", "legacy_alt"):
                column = mapping.get(field)
                if isinstance(column, dict) and column.get("index") is not None:
                    guarded_indices.add(int(column["index"]))
        for row_number in range(2, max(cached_sheet.max_row, formula_sheet.max_row) + 1):
            for index in sorted(guarded_indices):
                formula_cell = formula_sheet.cell(row=row_number, column=index + 1)
                cached_cell = cached_sheet.cell(row=row_number, column=index + 1)
                if formula_cell.data_type == "f" and cached_cell.value in (None, ""):
                    raise ArchiveBatchError(
                        422,
                        "formula_value_unavailable",
                        "快照包含没有可用缓存值的公式单元格，无法安全解释为空值。",
                        extra={"sheet": table.sheet, "cell": formula_cell.coordinate},
                    )
            id_index = identity_indices.get("id")
            source_index = identity_indices.get("source")
            if id_index is not None and source_index is not None:
                _guard_snapshot_identity_values(
                    cached_sheet.cell(row=row_number, column=id_index + 1).value,
                    cached_sheet.cell(row=row_number, column=source_index + 1).value,
                    sheet=table.sheet,
                    row_number=row_number,
                )
    finally:
        cached.close()
        formulas.close()


def _parse_json_artifact(path: Path, request: Any, artifact_id: str, source_type: str) -> ParsedImportTable:
    mappings = _read_json_mapping_rows(path, ("entries", "rows"))
    mode = str(getattr(request, "mode", "merge") or "merge").strip().lower()
    if not any(str(row.get("language") or "").strip() for row in mappings):
        headers, raw_rows = _mapping_rows_to_matrix(mappings)
        return _parse_matrix(
            headers,
            raw_rows,
            sheet_name=JSON_SHEET_KEY,
            request=request,
            source_artifact_id=artifact_id,
            source_type=source_type,
        )

    rows: list[dict[str, Any]] = []
    detected: dict[str, Any] = {"id": None, "source": None, "note": None, "languages": {}}
    for row_number, mapping in enumerate(mappings, start=2):
        raw_language = str(mapping.get("language") or "").strip()
        if not raw_language:
            continue
        try:
            language = require_supported_language(raw_language)
        except ValueError:
            continue
        normalized = {str(key or "").strip().lower(): value for key, value in mapping.items()}
        entry_key = str(normalized.get("entry_key") or normalized.get("id") or normalized.get("key") or "").strip()
        source = str(normalized.get("source") or normalized.get("cn") or normalized.get("中文") or "").strip()
        effective_row_number = int(mapping.get("row_number") or row_number)
        if mode == "snapshot":
            _guard_snapshot_identity_values(
                entry_key,
                source,
                sheet=JSON_SHEET_KEY,
                row_number=effective_row_number,
            )
        if not source:
            continue
        target_name = next(
            (name for name in ("target", *target_aliases(language)) if name.lower() in normalized),
            "target",
        )
        target = str(normalized.get(target_name.lower()) or "").strip()
        rows.append(
            {
                "entry_key": entry_key,
                "source": source,
                "target": target,
                "target_alt": "",
                "language": language,
                "sheet": JSON_SHEET_KEY,
                "row_number": effective_row_number,
                "note": str(normalized.get("note") or "").strip(),
                "source_type": source_type,
                "source_artifact_id": artifact_id,
                "target_column_present": target_name.lower() in normalized,
            }
        )
        detected["languages"].setdefault(
            language,
            {"target": {"name": target_name, "index": None}, "legacy_alt": None},
        )
    return ParsedImportTable(rows=rows, detected_columns=detected, sheet=JSON_SHEET_KEY, include_empty=True)


def _parse_translation_artifact(artifact: dict[str, Any], request: Any, source_type: str) -> ParsedTranslationArtifact:
    path = Path(artifact["path"])
    suffix = path.suffix.lower()
    mode = str(getattr(request, "mode", "merge") or "merge").strip().lower()
    try:
        selected = _selected_languages(request)
    except ValueError as exc:
        raise ArchiveBatchError(
            422,
            "unsupported_language",
            "所选目标语言不受支持。",
            extra={"reason": str(exc)},
        ) from exc
    try:
        if suffix == ".csv":
            headers, raw_rows = _read_csv_matrix(path)
            table = _parse_matrix(
                headers,
                raw_rows,
                sheet_name=CSV_SHEET_KEY,
                request=request,
                source_artifact_id=artifact["id"],
                source_type=source_type,
            )
        elif suffix == ".json":
            table = _parse_json_artifact(path, request, artifact["id"], source_type)
        elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            workbook = load_workbook(path, read_only=True, data_only=True)
            selection_workbook = workbook
            try:
                if mode == "snapshot":
                    selection_workbook = load_workbook(path, read_only=True, data_only=False)
                selected_worksheet = _select_xlsx_data_sheet(
                    selection_workbook,
                    getattr(request, "sheet", None),
                    lambda candidate: _xlsx_candidate(candidate, request, allow_id_only=mode == "snapshot"),
                )
                worksheet = workbook[selected_worksheet.title]
                headers = _sheet_headers(worksheet)
                table = _parse_matrix(
                    headers,
                    worksheet.iter_rows(min_row=2, values_only=True),
                    sheet_name=_sheet_key(worksheet.title),
                    request=request,
                    source_artifact_id=artifact["id"],
                    source_type=source_type,
                    guard_snapshot_identity=False,
                )
            finally:
                if selection_workbook is not workbook:
                    selection_workbook.close()
                workbook.close()
        else:
            raise ArchiveBatchError(422, "unsupported_import_format", "译文归档仅支持 XLSX、CSV 或 JSON。")
    except SheetSelectionError as exc:
        raise ArchiveBatchError(
            422,
            "sheet_selection_required",
            "检测到多个可导入的数据工作表，请选择后重试。",
            extra={"sheets": exc.candidates},
        ) from exc
    except ArchiveBatchError:
        raise
    except (KeyError, ValueError) as exc:
        raise ArchiveBatchError(
            422,
            "column_mapping_invalid",
            "导入列映射无效。",
            extra={"reason": str(exc)},
        ) from exc

    detected_languages = [language for language in LANGUAGE_ORDER if language in (table.detected_columns.get("languages") or {})]
    if not selected:
        selected = detected_languages
    missing = [language for language in selected if language not in detected_languages]
    if missing:
        raise ArchiveBatchError(
            422,
            "language_column_missing",
            "所选语言在导入文件中没有真实目标列。",
            extra={"languages": missing},
        )
    if not selected:
        raise ArchiveBatchError(422, "dedicated_language_column_required", "未检测到可导入的目标语言列。")
    if mode == "snapshot" and path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        _guard_uncached_formulas(path, table, selected)
    rows = [row for row in table.rows if normalize_language(row.get("language") or "en") in selected]
    return ParsedTranslationArtifact(rows=rows, columns=table.detected_columns, sheet=table.sheet, languages=selected)


def _conflict(code: str, message: str, row: dict[str, Any]) -> dict[str, Any]:
    return {
        "code": code,
        "message": message,
        "language": normalize_language(row.get("language") or "en"),
        "entry_key": str(row.get("entry_key") or "").strip(),
        "source": str(row.get("source") or "").strip(),
        "row_number": int(row.get("row_number") or 0),
    }


def _input_conflicts(rows: list[dict[str, Any]]) -> dict[int, list[dict[str, Any]]]:
    conflicts: dict[int, list[dict[str, Any]]] = {}
    by_id: dict[tuple[str, str], list[int]] = {}
    by_concept_id: dict[str, list[int]] = {}
    by_unkeyed_source: dict[tuple[str, str], list[int]] = {}
    by_source: dict[tuple[str, str], list[int]] = {}
    by_source_global: dict[str, list[int]] = {}
    for index, row in enumerate(rows):
        language = normalize_language(row.get("language") or "en")
        entry_key = str(row.get("entry_key") or "").strip()
        source_key = _source_key(row.get("source"))
        if entry_key:
            by_id.setdefault((language, entry_key), []).append(index)
            by_concept_id.setdefault(entry_key, []).append(index)
        elif source_key:
            by_unkeyed_source.setdefault((language, source_key), []).append(index)
        if source_key:
            by_source.setdefault((language, source_key), []).append(index)
            by_source_global.setdefault(source_key, []).append(index)

    def add(indices: list[int], code: str, message: str) -> None:
        for index in indices:
            conflicts.setdefault(index, []).append(_conflict(code, message, rows[index]))

    for indices in by_id.values():
        if len(indices) > 1:
            add(indices, "duplicate_entry_key", "同一语言中存在重复 ID。")
    for indices in by_concept_id.values():
        shared_values = {
            (_source_key(rows[index].get("source")), str(rows[index].get("note") or "").strip())
            for index in indices
        }
        if len(shared_values) > 1:
            add(indices, "shared_identity_mismatch", "同一 ID 的跨语言行必须使用一致的中文源文和备注。")
    for indices in by_unkeyed_source.values():
        if len(indices) > 1:
            add(indices, "duplicate_source", "同一语言中存在重复的无 ID 中文源文。")
    for indices in by_source.values():
        keys = {str(rows[index].get("entry_key") or "").strip() for index in indices if str(rows[index].get("entry_key") or "").strip()}
        if len(keys) > 1:
            add(indices, "source_multiple_ids", "同一中文源文对应多个 ID。")
        elif keys and any(not str(rows[index].get("entry_key") or "").strip() for index in indices):
            add(indices, "source_mixed_identity", "同一中文源文不能同时使用有 ID 和无 ID 身份。")
    for indices in by_source_global.values():
        languages = {normalize_language(rows[index].get("language") or "en") for index in indices}
        if len(languages) < 2:
            continue
        keys = {
            str(rows[index].get("entry_key") or "").strip()
            for index in indices
            if str(rows[index].get("entry_key") or "").strip()
        }
        if len(keys) > 1:
            add(indices, "source_multiple_ids", "同一中文源文不能跨语言对应多个 ID。")
        elif keys and any(not str(rows[index].get("entry_key") or "").strip() for index in indices):
            add(indices, "source_mixed_identity", "同一中文源文不能跨语言混用有 ID 和无 ID 身份。")
    return conflicts


def _make_after(
    before: dict[str, Any] | None,
    row: dict[str, Any],
    *,
    entity_id: str,
    project_id: str,
    batch_id: str,
    dataset_key: str,
    sheet_key: str,
    source_type: str,
    review_status: str,
    timestamp: str,
    target: str,
    active: int,
) -> dict[str, Any]:
    after = dict(before or {})
    after.update(
        {
            "id": entity_id,
            "project_id": project_id,
            "entry_key": str(row.get("entry_key") or "").strip(),
            "source": str(row.get("source") or "").strip(),
            "source_key": _source_key(row.get("source")),
            "target": target,
            "target_alt": "",
            "language": normalize_language(row.get("language") or "en"),
            "sheet": sheet_key,
            "row_number": int(row.get("row_number") or 0),
            "note": str(row.get("note") or "").strip(),
            "source_type": source_type,
            "source_artifact_id": str(row.get("source_artifact_id") or "").strip(),
            "active": active,
            "dataset_key": dataset_key,
            "last_import_batch_id": batch_id,
            "review_status": review_status,
            "created_at": str((before or {}).get("created_at") or timestamp),
            "updated_at": timestamp,
        }
    )
    return {field: after.get(field) for field in TRANSLATION_FIELDS}


def _make_scope_after(
    before: dict[str, Any],
    *,
    batch_id: str,
    dataset_key: str,
    sheet_key: str,
    timestamp: str,
) -> dict[str, Any]:
    after = dict(before)
    after.update(
        {
            "dataset_key": dataset_key,
            "sheet": sheet_key,
            "last_import_batch_id": batch_id,
            "updated_at": timestamp,
        }
    )
    return {field: after.get(field) for field in TRANSLATION_FIELDS}


def _shared_content_unchanged(before: dict[str, Any], shared: dict[str, Any]) -> bool:
    return (
        str(before.get("entry_key") or "").strip() == str(shared.get("entry_key") or "").strip()
        and str(before.get("source") or "").strip() == str(shared.get("source") or "").strip()
        and str(before.get("note") or "").strip() == str(shared.get("note") or "").strip()
    )


def _make_shared_after(
    before: dict[str, Any],
    shared: dict[str, Any],
    *,
    batch_id: str,
    timestamp: str,
) -> dict[str, Any]:
    after = dict(before)
    source = str(shared.get("source") or "").strip()
    after.update(
        {
            "entry_key": str(shared.get("entry_key") or "").strip(),
            "source": source,
            "source_key": _source_key(source),
            "note": str(shared.get("note") or "").strip(),
            "last_import_batch_id": batch_id,
            "updated_at": timestamp,
        }
    )
    return {field: after.get(field) for field in TRANSLATION_FIELDS}


def _content_unchanged(before: dict[str, Any], row: dict[str, Any]) -> bool:
    return (
        int(before.get("active") or 0) == 1
        and str(before.get("entry_key") or "").strip() == str(row.get("entry_key") or "").strip()
        and str(before.get("source") or "").strip() == str(row.get("source") or "").strip()
        and str(before.get("target") or "").strip() == str(row.get("target") or "").strip()
        and str(before.get("target_alt") or "").strip() == ""
        and str(before.get("note") or "").strip() == str(row.get("note") or "").strip()
    )


def _item_record(
    *,
    batch_id: str,
    ordinal: int,
    row: dict[str, Any],
    entity_id: str,
    action: str,
    before: dict[str, Any] | None,
    expected_after: dict[str, Any] | None,
    conflicts: list[dict[str, Any]],
    timestamp: str,
) -> dict[str, Any]:
    language = normalize_language(row.get("language") or "en")
    entry_key = str(row.get("entry_key") or "").strip()
    source = str(row.get("source") or "").strip()
    return {
        "id": db.new_id("abi"),
        "batch_id": batch_id,
        "ordinal": ordinal,
        "kind": ARCHIVE_KIND,
        "entity_id": entity_id,
        "identity": {"language": language, "entry_key": entry_key, "source_key": _source_key(source)},
        "language": language,
        "entry_key": entry_key,
        "source_key": _source_key(source),
        "source": source,
        "target": str(row.get("target") or "").strip(),
        "target_column_present": bool(row.get("target_column_present")),
        "explicit_empty": bool(row.get("target_column_present")) and not str(row.get("target") or "").strip(),
        "planned_action": action,
        "before_hash": _translation_hash(before),
        "expected_after": expected_after or {},
        "conflicts": conflicts,
        "created_at": timestamp,
    }


def _plan_batch(
    project_id: str,
    artifact: dict[str, Any],
    parsed: ParsedTranslationArtifact,
    request: Any,
    source_type: str,
    batch_id: str,
    timestamp: str,
    existing_rows: list[dict[str, Any]],
) -> tuple[str, list[dict[str, Any]], dict[str, int], list[dict[str, Any]]]:
    mode = str(getattr(request, "mode", "merge") or "merge").strip().lower()
    if mode not in {"merge", "snapshot"}:
        raise ArchiveBatchError(422, "invalid_mode", "mode 必须是 merge 或 snapshot。", batch_id=batch_id)
    requested_dataset = str(getattr(request, "dataset_key", None) or "").strip()
    if mode == "snapshot" and not requested_dataset:
        raise ArchiveBatchError(422, "dataset_key_required", "snapshot 必须明确选择已有 dataset_key。", batch_id=batch_id)
    if mode == "snapshot" and not getattr(request, "languages", None):
        raise ArchiveBatchError(422, "languages_required", "snapshot 必须明确选择目标语言。", batch_id=batch_id)
    known_lineages = {str(row.get("dataset_key") or "").strip() for row in existing_rows}
    if mode == "snapshot" and requested_dataset not in known_lineages:
        raise ArchiveBatchError(422, "dataset_key_not_found", "snapshot 指定的 dataset_key 不存在。", batch_id=batch_id)

    by_id: dict[tuple[str, str], dict[str, Any]] = {}
    by_concept_id: dict[str, list[dict[str, Any]]] = {}
    by_source: dict[tuple[str, str], list[dict[str, Any]]] = {}
    by_source_global: dict[str, list[dict[str, Any]]] = {}
    for existing in existing_rows:
        language = normalize_language(existing.get("language") or "en")
        entry_key = str(existing.get("entry_key") or "").strip()
        source_key = _source_key(existing.get("source"))
        if entry_key:
            by_id[(language, entry_key)] = existing
            by_concept_id.setdefault(entry_key, []).append(existing)
        if source_key:
            by_source.setdefault((language, source_key), []).append(existing)
            by_source_global.setdefault(source_key, []).append(existing)

    row_conflicts = _input_conflicts(parsed.rows)
    matches: dict[int, dict[str, Any] | None] = {}
    concept_siblings: dict[int, list[dict[str, Any]]] = {}
    inferred_lineages: set[str] = set()
    for index, row in enumerate(parsed.rows):
        if row_conflicts.get(index):
            matches[index] = None
            continue
        language = normalize_language(row.get("language") or "en")
        entry_key = str(row.get("entry_key") or "").strip()
        source_key = _source_key(row.get("source"))
        source_matches = by_source.get((language, source_key), [])
        global_source_matches = list(by_source_global.get(source_key, []))
        siblings = list(by_concept_id.get(entry_key, [])) if entry_key else []
        concept_siblings[index] = siblings
        match: dict[str, Any] | None = None
        if entry_key:
            foreign_keyed_matches = [
                candidate
                for candidate in global_source_matches
                if str(candidate.get("entry_key") or "").strip()
                and str(candidate.get("entry_key") or "").strip() != entry_key
            ]
            unkeyed_global_matches = [
                candidate
                for candidate in global_source_matches
                if not str(candidate.get("entry_key") or "").strip()
            ]
            sibling_source_keys = {_source_key(candidate.get("source")) for candidate in siblings}
            sibling_languages = {
                normalize_language(candidate.get("language") or "en") for candidate in siblings
            }
            unkeyed_language_counts: dict[str, int] = {}
            for candidate in unkeyed_global_matches:
                candidate_language = normalize_language(candidate.get("language") or "en")
                unkeyed_language_counts[candidate_language] = unkeyed_language_counts.get(candidate_language, 0) + 1
            ambiguous_unkeyed = any(count > 1 for count in unkeyed_language_counts.values()) or any(
                candidate_language in sibling_languages for candidate_language in unkeyed_language_counts
            )
            unkeyed_is_other_concept = bool(siblings and unkeyed_global_matches and source_key not in sibling_source_keys)
            if foreign_keyed_matches or unkeyed_is_other_concept:
                row_conflicts.setdefault(index, []).append(
                    _conflict("concept_source_conflict", "修改后的中文源文已属于另一个归档概念。", row)
                )
            elif ambiguous_unkeyed:
                row_conflicts.setdefault(index, []).append(
                    _conflict(
                        "existing_identity_ambiguous",
                        "同一中文源文在某个语言中存在多条身份记录，不能自动认领。",
                        row,
                    )
                )
            elif unkeyed_global_matches:
                siblings.extend(unkeyed_global_matches)
                concept_siblings[index] = siblings

            id_match = by_id.get((language, entry_key))
            other_source_matches = [candidate for candidate in source_matches if not id_match or candidate["id"] != id_match["id"]]
            if id_match and other_source_matches:
                row_conflicts.setdefault(index, []).append(
                    _conflict("identity_cross_match", "ID 与中文源文命中了不同归档记录。", row)
                )
            elif id_match:
                match = id_match
            else:
                keyed = [candidate for candidate in source_matches if str(candidate.get("entry_key") or "").strip()]
                unkeyed = [candidate for candidate in source_matches if not str(candidate.get("entry_key") or "").strip()]
                if keyed:
                    row_conflicts.setdefault(index, []).append(
                        _conflict("new_id_source_exists", "新 ID 的中文源文已属于另一条有 ID 记录。", row)
                    )
                elif len(unkeyed) == 1:
                    match = unkeyed[0]
                elif len(unkeyed) > 1:
                    row_conflicts.setdefault(index, []).append(
                        _conflict("existing_identity_ambiguous", "现有无 ID 中文源文存在多条匹配。", row)
                    )
        else:
            keyed = [candidate for candidate in source_matches if str(candidate.get("entry_key") or "").strip()]
            unkeyed = [candidate for candidate in source_matches if not str(candidate.get("entry_key") or "").strip()]
            if keyed:
                row_conflicts.setdefault(index, []).append(
                    _conflict("source_owned_by_id", "无 ID 行的中文源文已属于有 ID 记录。", row)
                )
            elif len(unkeyed) == 1:
                match = unkeyed[0]
            elif len(unkeyed) > 1:
                row_conflicts.setdefault(index, []).append(
                    _conflict("existing_identity_ambiguous", "现有无 ID 中文源文存在多条匹配。", row)
                )
        matches[index] = match
        for sibling in siblings or ([match] if match else []):
            if sibling and str(sibling.get("dataset_key") or "").strip():
                inferred_lineages.add(str(sibling["dataset_key"]).strip())

    if requested_dataset:
        dataset_key = requested_dataset
    elif len(inferred_lineages) == 1:
        dataset_key = next(iter(inferred_lineages))
    elif inferred_lineages:
        dataset_key = sorted(inferred_lineages)[0]
        for index, row in enumerate(parsed.rows):
            if matches.get(index) and str(matches[index].get("dataset_key") or "").strip() in inferred_lineages:
                row_conflicts.setdefault(index, []).append(
                    _conflict("lineage_selection_required", "输入同时命中多个既有 lineage，必须明确选择。", row)
                )
    else:
        dataset_key = f"dataset_{batch_id.removeprefix('aib_')}"

    for index, row in enumerate(parsed.rows):
        scoped_rows = list(concept_siblings.get(index, []))
        match = matches.get(index)
        if match and all(str(candidate.get("id")) != str(match.get("id")) for candidate in scoped_rows):
            scoped_rows.append(match)
        existing_lineages = {
            str(candidate.get("dataset_key") or "").strip()
            for candidate in scoped_rows
            if str(candidate.get("dataset_key") or "").strip()
        }
        if any(lineage != dataset_key for lineage in existing_lineages):
            row_conflicts.setdefault(index, []).append(
                _conflict("cross_lineage_match", "归档身份已属于另一个 lineage。", row)
            )

        entry_key = str(row.get("entry_key") or "").strip()
        if entry_key or match or row_conflicts.get(index):
            continue
        language = normalize_language(row.get("language") or "en")
        row_number = int(row.get("row_number") or 0)
        identity_drift = row_number > 0 and any(
            str(candidate.get("dataset_key") or "").strip() == dataset_key
            and normalize_language(candidate.get("language") or "en") == language
            and _sheet_key(str(candidate.get("sheet") or "")) == parsed.sheet
            and int(candidate.get("row_number") or 0) == row_number
            and _source_key(candidate.get("source")) != _source_key(row.get("source"))
            for candidate in existing_rows
        )
        if identity_drift:
            row_conflicts.setdefault(index, []).append(
                _conflict(
                    "unstable_identity_change",
                    "无 ID 行不能在既有归档范围内通过修改中文源文创建新身份，请先补充稳定 ID。",
                    row,
                )
            )

    summary = {
        "source_rows": len(
            {
                (str(row.get("sheet") or ""), int(row.get("row_number") or 0), str(row.get("source") or ""))
                for row in parsed.rows
            }
        ),
        "insert": 0,
        "update": 0,
        "unchanged": 0,
        "skip": 0,
        "clear": 0,
        "deactivate": 0,
        "protected": 0,
        "conflict": 0,
    }
    items: list[dict[str, Any]] = []
    all_conflicts: list[dict[str, Any]] = []
    seen_entity_ids: set[str] = set()
    override_protected = bool(getattr(request, "override_protected", False))
    for index, row in enumerate(parsed.rows):
        before = matches.get(index)
        if before:
            seen_entity_ids.add(str(before["id"]))
        conflicts = list(row_conflicts.get(index) or [])
        target = str(row.get("target") or "").strip()
        entity_id = str((before or {}).get("id") or db.new_id("tr"))
        expected_after: dict[str, Any] | None = None
        if conflicts:
            action = "conflict"
        elif mode == "merge" and not target:
            if before is not None and not _shared_content_unchanged(before, row):
                action = "update"
                expected_after = _make_shared_after(
                    before,
                    row,
                    batch_id=batch_id,
                    timestamp=timestamp,
                )
            else:
                action = "skip"
                expected_after = before
        elif not before and not target:
            action = "skip"
        elif not before:
            action = "insert"
            expected_after = _make_after(
                None,
                row,
                entity_id=entity_id,
                project_id=project_id,
                batch_id=batch_id,
                dataset_key=dataset_key,
                sheet_key=parsed.sheet,
                source_type=source_type,
                review_status="approved",
                timestamp=timestamp,
                target=target,
                active=1,
            )
        elif mode == "snapshot" and not target:
            if int(before.get("active") or 0) == 0 and not str(before.get("target") or "").strip():
                action = "unchanged"
                expected_after = before
            else:
                action = "clear"
                expected_after = _make_after(
                    before,
                    row,
                    entity_id=entity_id,
                    project_id=project_id,
                    batch_id=batch_id,
                    dataset_key=dataset_key,
                    sheet_key=parsed.sheet,
                    source_type=source_type,
                    review_status="approved",
                    timestamp=timestamp,
                    target="",
                    active=0,
                )
        elif _content_unchanged(before, row):
            if (
                str(before.get("dataset_key") or "").strip() == dataset_key
                and str(before.get("sheet") or "").strip() == parsed.sheet
            ):
                action = "unchanged"
                expected_after = before
            else:
                action = "update"
                expected_after = _make_scope_after(
                    before,
                    batch_id=batch_id,
                    dataset_key=dataset_key,
                    sheet_key=parsed.sheet,
                    timestamp=timestamp,
                )
        else:
            action = "update"
            expected_after = _make_after(
                before,
                row,
                entity_id=entity_id,
                project_id=project_id,
                batch_id=batch_id,
                dataset_key=dataset_key,
                sheet_key=parsed.sheet,
                source_type=source_type,
                review_status="approved",
                timestamp=timestamp,
                target=target,
                active=1,
            )

        protected = bool(before) and action in {"update", "clear"} and _is_protected_overwrite(before, source_type)
        if protected:
            if override_protected and expected_after:
                summary["protected"] += 1
                expected_after["source_type"] = "imported"
                expected_after["review_status"] = "pending"
            elif not override_protected:
                protected_conflict = _conflict("protected_source", "可信来源记录默认禁止覆盖。", row)
                conflicts.append(protected_conflict)
                action = "protected"
        summary[action if action in summary else "conflict"] += 1
        all_conflicts.extend(conflicts)
        items.append(
            _item_record(
                batch_id=batch_id,
                ordinal=len(items),
                row=row,
                entity_id=entity_id if before or action == "insert" else "",
                action=action,
                before=before,
                expected_after=expected_after,
                conflicts=conflicts,
                timestamp=timestamp,
            )
        )

    planned_entity_ids = {str(item["entity_id"]) for item in items if str(item.get("entity_id") or "")}
    for index, item in enumerate(list(items)):
        entry_key = str(item.get("entry_key") or "").strip()
        if (
            not entry_key
            or item["planned_action"] not in {"insert", "update", "unchanged", "skip", "clear"}
            or item["conflicts"]
        ):
            continue
        shared = item["expected_after"] or parsed.rows[index]
        for before in concept_siblings.get(index, []):
            entity_id = str(before["id"])
            if entity_id in planned_entity_ids or _shared_content_unchanged(before, shared):
                continue
            expected_after = _make_shared_after(
                before,
                shared,
                batch_id=batch_id,
                timestamp=timestamp,
            )
            row = {
                "entry_key": expected_after.get("entry_key", ""),
                "source": expected_after.get("source", ""),
                "target": before.get("target", ""),
                "language": before.get("language", "en"),
                "sheet": before.get("sheet", ""),
                "row_number": before.get("row_number", 0),
                "note": expected_after.get("note", ""),
                "source_artifact_id": before.get("source_artifact_id", ""),
                "target_column_present": False,
            }
            conflicts: list[dict[str, Any]] = []
            action = "update"
            if _is_protected_overwrite(before, source_type):
                if override_protected:
                    summary["protected"] += 1
                    expected_after["source_type"] = "imported"
                    expected_after["review_status"] = "pending"
                else:
                    conflicts.append(_conflict("protected_source", "可信来源记录默认禁止覆盖。", row))
                    action = "protected"
            summary[action if action in summary else "conflict"] += 1
            all_conflicts.extend(conflicts)
            items.append(
                _item_record(
                    batch_id=batch_id,
                    ordinal=len(items),
                    row=row,
                    entity_id=entity_id,
                    action=action,
                    before=before,
                    expected_after=expected_after,
                    conflicts=conflicts,
                    timestamp=timestamp,
                )
            )
            planned_entity_ids.add(entity_id)
            seen_entity_ids.add(entity_id)

    if mode == "snapshot":
        for before in existing_rows:
            if (
                int(before.get("active") or 0) != 1
                or str(before.get("dataset_key") or "").strip() != dataset_key
                or _sheet_key(str(before.get("sheet") or "")) != parsed.sheet
                or normalize_language(before.get("language") or "en") not in parsed.languages
                or str(before.get("id")) in seen_entity_ids
            ):
                continue
            row = {
                "entry_key": before.get("entry_key", ""),
                "source": before.get("source", ""),
                "target": "",
                "language": before.get("language", "en"),
                "sheet": before.get("sheet", ""),
                "row_number": before.get("row_number", 0),
                "note": before.get("note", ""),
                "source_artifact_id": artifact["id"],
                "target_column_present": True,
            }
            conflicts: list[dict[str, Any]] = []
            expected_after = dict(before)
            expected_after.update({"active": 0, "last_import_batch_id": batch_id, "updated_at": timestamp})
            protected = str(before.get("source_type") or "").strip().lower() in PROTECTED_TRANSLATION_SOURCES
            action = "deactivate"
            if protected:
                if override_protected:
                    summary["protected"] += 1
                    expected_after["source_type"] = "imported"
                    expected_after["review_status"] = "pending"
                else:
                    conflicts.append(_conflict("protected_source", "可信来源记录默认禁止停用。", row))
                    action = "protected"
            summary[action if action in summary else "conflict"] += 1
            all_conflicts.extend(conflicts)
            items.append(
                _item_record(
                    batch_id=batch_id,
                    ordinal=len(items),
                    row=row,
                    entity_id=str(before["id"]),
                    action=action,
                    before=before,
                    expected_after=expected_after,
                    conflicts=conflicts,
                    timestamp=timestamp,
                )
            )

    summary["conflict"] = len(all_conflicts)
    return dataset_key, items, summary, all_conflicts


def _persist_analysis(
    *,
    batch_id: str,
    project_id: str,
    artifact: dict[str, Any],
    artifact_checksum: str,
    token: str,
    request: Any,
    summary: dict[str, int],
    dataset_key: str,
    sheet_key: str,
    languages: list[str],
    base_state_version: int,
    base_state_checksum: str,
    items: list[dict[str, Any]],
    timestamp: str,
) -> None:
    with db.connect() as conn:
        conn.execute(
            """
            INSERT INTO archive_import_batches
              (id, project_id, kind, artifact_id, artifact_checksum, token, request_json,
               summary_json, result_json, rollback_result_json, mode, dataset_key, sheet_key,
               languages_json, base_state_version, base_state_checksum, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, '{}', '{}', ?, ?, ?, ?, ?, ?, 'analyzed', ?, ?)
            """,
            (
                batch_id,
                project_id,
                ARCHIVE_KIND,
                artifact["id"],
                artifact_checksum,
                token,
                _json_dump(_request_payload(request)),
                _json_dump(summary),
                str(getattr(request, "mode", "merge") or "merge").strip().lower(),
                dataset_key,
                sheet_key,
                _json_dump(languages),
                base_state_version,
                base_state_checksum,
                timestamp,
                timestamp,
            ),
        )
        conn.executemany(
            """
            INSERT INTO archive_import_batch_items
              (id, batch_id, ordinal, kind, entity_id, identity_json, language, entry_key,
               source_key, source, target, target_column_present, explicit_empty,
               planned_action, before_hash, expected_after_json, conflict_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["id"],
                    batch_id,
                    item["ordinal"],
                    item["kind"],
                    item["entity_id"],
                    _json_dump(item["identity"]),
                    item["language"],
                    item["entry_key"],
                    item["source_key"],
                    item["source"],
                    item["target"],
                    1 if item["target_column_present"] else 0,
                    1 if item["explicit_empty"] else 0,
                    item["planned_action"],
                    item["before_hash"],
                    _json_dump(item["expected_after"]),
                    _json_dump(item["conflicts"]),
                    timestamp,
                )
                for item in items
            ],
        )


def analyze_translation_archive(
    project_id: str,
    request: Any,
    *,
    source_type: str = "imported",
) -> dict[str, Any]:
    try:
        db.get_project(project_id)
        artifact = db.get_artifact(request.artifact_id)
    except KeyError as exc:
        raise ArchiveBatchError(404, "project_or_artifact_not_found", "项目或 artifact 不存在。") from exc
    if artifact["project_id"] != project_id:
        raise ArchiveBatchError(404, "project_or_artifact_not_found", "项目或 artifact 不存在。")

    batch_id = db.new_id("aib")
    token = f"ait_{uuid.uuid4().hex}"
    timestamp = db.now_iso()
    parsed = _parse_translation_artifact(artifact, request, source_type)
    path = Path(artifact["path"])
    artifact_checksum = _file_checksum(path)
    with db.connect() as conn:
        rows = [
            _translation_row(row)
            for row in conn.execute(
                "SELECT * FROM translation_entries WHERE project_id = ?",
                (project_id,),
            ).fetchall()
        ]
        state_row = conn.execute(
            "SELECT version FROM archive_state_versions WHERE project_id = ? AND kind = ?",
            (project_id, ARCHIVE_KIND),
        ).fetchone()
        base_state_version = int(state_row["version"] if state_row else 0)
        base_state_checksum = _state_checksum(rows)

    dataset_key, items, summary, conflicts = _plan_batch(
        project_id,
        artifact,
        parsed,
        request,
        source_type,
        batch_id,
        timestamp,
        rows,
    )
    _persist_analysis(
        batch_id=batch_id,
        project_id=project_id,
        artifact=artifact,
        artifact_checksum=artifact_checksum,
        token=token,
        request=request,
        summary=summary,
        dataset_key=dataset_key,
        sheet_key=parsed.sheet,
        languages=parsed.languages,
        base_state_version=base_state_version,
        base_state_checksum=base_state_checksum,
        items=items,
        timestamp=timestamp,
    )
    changes = [
        {
            "ordinal": item["ordinal"],
            "action": item["planned_action"],
            "language": item["language"],
            "entry_key": item["entry_key"],
            "source": item["source"],
            "target": item["target"],
            "explicit_empty": item["explicit_empty"],
        }
        for item in items[:CHANGE_SAMPLE_LIMIT]
    ]
    return {
        "batch_id": batch_id,
        "token": token,
        "artifact": {
            "id": artifact["id"],
            "label": artifact.get("label", ""),
            "kind": artifact.get("kind", ""),
            "checksum": artifact_checksum,
        },
        "sheet": parsed.sheet,
        "mode": str(getattr(request, "mode", "merge") or "merge").strip().lower(),
        "dataset_key": dataset_key,
        "languages": parsed.languages,
        "columns": parsed.columns,
        "summary": summary,
        "changes": changes,
        "conflicts": conflicts,
        "can_commit": not conflicts,
    }


def _insert_expected(conn: sqlite3.Connection, after: dict[str, Any]) -> dict[str, Any]:
    after = {**after, "source_key": _source_key(after.get("source"))}
    conn.execute(
        f"INSERT INTO translation_entries ({', '.join(TRANSLATION_FIELDS)}) VALUES ({', '.join('?' for _ in TRANSLATION_FIELDS)})",
        tuple(after.get(field) for field in TRANSLATION_FIELDS),
    )
    return _translation_row(conn.execute("SELECT * FROM translation_entries WHERE id = ?", (after["id"],)).fetchone())


def _replace_expected(conn: sqlite3.Connection, after: dict[str, Any]) -> dict[str, Any]:
    after = {**after, "source_key": _source_key(after.get("source"))}
    fields = [field for field in TRANSLATION_FIELDS if field != "id"]
    conn.execute(
        f"UPDATE translation_entries SET {', '.join(f'{field} = ?' for field in fields)} WHERE id = ?",
        (*[after.get(field) for field in fields], after["id"]),
    )
    return _translation_row(conn.execute("SELECT * FROM translation_entries WHERE id = ?", (after["id"],)).fetchone())


def _preflight_rollback_uniqueness(conn: sqlite3.Connection, revisions: list[sqlite3.Row], batch_id: str) -> None:
    for revision in revisions:
        before = _json_load(revision["before_json"], {})
        if not before:
            continue
        entry_key = str(before.get("entry_key") or "").strip()
        if entry_key:
            duplicate = conn.execute(
                """
                SELECT id FROM translation_entries
                WHERE project_id = ? AND language = ? AND entry_key = ? AND id <> ?
                LIMIT 1
                """,
                (before["project_id"], before["language"], entry_key, before["id"]),
            ).fetchone()
            if duplicate:
                raise ArchiveBatchError(409, "rollback_constraint_conflict", "回滚会违反归档身份唯一约束。", batch_id=batch_id)
        source_key = _source_key(before.get("source"))
        if source_key:
            candidates = conn.execute(
                "SELECT id, source FROM translation_entries WHERE project_id = ? AND language = ? AND id <> ?",
                (before["project_id"], before["language"], before["id"]),
            ).fetchall()
            if any(_source_key(candidate["source"]) == source_key for candidate in candidates):
                raise ArchiveBatchError(409, "rollback_constraint_conflict", "回滚会造成中文源文身份冲突。", batch_id=batch_id)


def _translation_adapter() -> ArchiveEntityAdapter:
    return ArchiveEntityAdapter(
        kind=ARCHIVE_KIND,
        table="translation_entries",
        fields=TRANSLATION_FIELDS,
        collection_key="entries",
        normalize_row=_translation_row,
        row_hash=_translation_hash,
        state_checksum=_state_checksum,
        insert_expected=_insert_expected,
        replace_expected=_replace_expected,
        preflight_rollback=_preflight_rollback_uniqueness,
        artifact_checksum=_file_checksum,
    )


def commit_translation_archive(project_id: str, token: str, *, compact: bool = False) -> dict[str, Any]:
    return commit_archive_batch(project_id, token, _translation_adapter(), compact=compact)


def list_translation_import_batches(project_id: str, *, compact: bool = False) -> dict[str, Any]:
    return list_archive_import_batches(project_id, ARCHIVE_KIND, compact=compact)


def rollback_translation_import_batch(project_id: str, batch_id: str) -> dict[str, Any]:
    return rollback_archive_import_batch(project_id, batch_id, _translation_adapter())
