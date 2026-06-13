from __future__ import annotations

import argparse
import json
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / ".tmp" / "stability"


class StabilityCheck:
    def __init__(self, base_url: str, keep_project: bool = False) -> None:
        self.base_url = base_url.rstrip("/")
        self.keep_project = keep_project
        self.session = httpx.Client(follow_redirects=True)
        self.results: list[dict[str, Any]] = []
        self.project_id = ""
        self.project_name = f"STABILITY-{time.strftime('%Y%m%d%H%M%S')}"
        self.fixture_dir = OUT_DIR / self.project_name
        self.fixture_dir.mkdir(parents=True, exist_ok=True)

    def run_step(self, name: str, fn) -> Any:
        start = time.perf_counter()
        try:
            data = fn()
            result = {"name": name, "ok": True, "seconds": round(time.perf_counter() - start, 3), "summary": self._summary(data)}
            self.results.append(result)
            print(json.dumps(result, ensure_ascii=False))
            return data
        except Exception as exc:
            result = {"name": name, "ok": False, "seconds": round(time.perf_counter() - start, 3), "error": str(exc)}
            self.results.append(result)
            print(json.dumps(result, ensure_ascii=False))
            raise

    def get(self, path: str, timeout: int = 60) -> Any:
        response = self.session.get(f"{self.base_url}{path}", timeout=timeout)
        if response.status_code >= 400:
            raise RuntimeError(f"GET {path} -> {response.status_code}: {response.text[:1000]}")
        content_type = response.headers.get("content-type", "")
        if "application/json" in content_type:
            return response.json()
        return {"status_code": response.status_code, "bytes": len(response.content), "content_type": content_type}

    def post(self, path: str, payload: dict[str, Any] | None = None, timeout: int = 180) -> Any:
        response = self.session.post(f"{self.base_url}{path}", json=payload or {}, timeout=timeout)
        if response.status_code >= 400:
            raise RuntimeError(f"POST {path} -> {response.status_code}: {response.text[:1200]}")
        return response.json()

    def delete(self, path: str, timeout: int = 60) -> Any:
        response = self.session.delete(f"{self.base_url}{path}", timeout=timeout)
        if response.status_code >= 400:
            raise RuntimeError(f"DELETE {path} -> {response.status_code}: {response.text[:1000]}")
        return response.json()

    def upload(self, project_id: str, path: Path, kind: str, timeout: int = 120) -> Any:
        with path.open("rb") as fh:
            response = self.session.post(
                f"{self.base_url}/api/projects/{project_id}/files?kind={kind}",
                files={"file": (path.name, fh, "application/octet-stream")},
                timeout=timeout,
            )
        if response.status_code >= 400:
            raise RuntimeError(f"UPLOAD {path.name} ({kind}) -> {response.status_code}: {response.text[:1000]}")
        return response.json()

    def poll_run(self, run_id: str, timeout_seconds: int = 240) -> dict[str, Any]:
        deadline = time.time() + timeout_seconds
        run = self.get(f"/api/runs/{run_id}")
        while run.get("status") in {"queued", "running"} and time.time() < deadline:
            time.sleep(3)
            run = self.get(f"/api/runs/{run_id}")
        if run.get("status") in {"queued", "running"}:
            raise RuntimeError(f"run timeout: {run_id} status={run.get('status')}")
        return run

    def poll_announcement_task(self, task_id: str, timeout_seconds: int = 240) -> dict[str, Any]:
        deadline = time.time() + timeout_seconds
        task = self.get(f"/api/announcement-tasks/{task_id}")
        while task.get("status") in {"queued", "running"} and time.time() < deadline:
            time.sleep(3)
            task = self.get(f"/api/announcement-tasks/{task_id}")
        if task.get("status") in {"queued", "running"}:
            raise RuntimeError(f"announcement task timeout: {task_id} status={task.get('status')}")
        return task

    def _summary(self, data: Any) -> Any:
        if isinstance(data, dict):
            keep = {}
            for key in (
                "id",
                "name",
                "label",
                "kind",
                "status",
                "imported_count",
                "languages",
                "source_rows",
                "translated_rows",
                "needs_translation",
                "reason",
                "estimated_batches",
                "project_id",
                "task_id",
            ):
                if key in data:
                    keep[key] = data[key]
            if "rows" in data:
                keep["rows"] = len(data.get("rows") or [])
            if "artifacts" in data:
                keep["artifacts"] = len(data.get("artifacts") or [])
            if "summary" in data:
                keep["summary"] = data.get("summary")
            if "metadata" in data and isinstance(data["metadata"], dict):
                progress = data["metadata"].get("translation_progress")
                quality = data["metadata"].get("quality")
                if progress:
                    keep["progress"] = {
                        "completed_rows": progress.get("completed_rows"),
                        "total_rows": progress.get("total_rows"),
                        "percent": progress.get("percent"),
                    }
                if quality:
                    keep["quality"] = {
                        "passed": quality.get("passed"),
                        "issues": len(quality.get("issues") or []),
                    }
            return keep or str(data)[:300]
        return str(data)[:300]

    def make_fixtures(self) -> dict[str, Path]:
        brief = self.fixture_dir / "project-brief.md"
        brief.write_text(
            "# Hell Prison SLG\n\nEnglish localization. Keep UI short, preserve variables, and follow glossary terms.",
            encoding="utf-8",
        )
        glossary = self.fixture_dir / "glossary.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Glossary"
        ws.append(["ID", "CN", "EN", "EN2", "分类", "备注"])
        ws.append(["T1", "开始", "Start", "", "UI", "stable check"])
        ws.append(["T2", "游戏", "Game", "", "UI", "stable check"])
        ws.append(["T3", "领取", "Claim", "", "UI", "stable check"])
        ws.append(["T4", "奖励", "Reward", "Rewards", "UI", "stable check"])
        wb.save(glossary)
        wb.close()

        source = self.fixture_dir / "language_source.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["ID", "cn", "EN"])
        ws.append(["1", "开始游戏", ""])
        ws.append(["2", "领取奖励", ""])
        wb.save(source)
        wb.close()

        translated = self.fixture_dir / "translated_ready.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["ID", "cn", "EN"])
        ws.append(["1", "开始游戏", "Start Game"])
        ws.append(["2", "领取奖励", "Claim Reward"])
        wb.save(translated)
        wb.close()

        announcement = self.fixture_dir / "announcement.txt"
        announcement.write_text("开始游戏即可领取奖励。\n请在活动期间领取奖励。", encoding="utf-8")
        return {"brief": brief, "glossary": glossary, "source": source, "translated": translated, "announcement": announcement}

    def execute(self) -> int:
        fixtures = self.make_fixtures()
        try:
            self.run_step("00_backend_health", lambda: self.get("/api/health"))
            settings = self.run_step("01_provider_settings", lambda: self.get("/api/settings"))
            provider = str(settings.get("provider") or "")
            if provider in {"openai", "openai-chat", "anthropic"} and settings.get("api_key") != "configured":
                raise RuntimeError("real provider is selected but API key is not configured")

            project = self.run_step(
                "02_create_project",
                lambda: self.post(
                    "/api/projects",
                    {
                        "name": self.project_name,
                        "type": "Stable Flow",
                        "icon": "game",
                        "description": "Automated no-Codex stability check. Delete after run.",
                    },
                ),
            )
            self.project_id = project["id"]

            brief_artifact = self.run_step("03_step1_upload_project_material", lambda: self.upload(self.project_id, fixtures["brief"], "asset"))
            self.run_step(
                "04_step2_ai_analysis_via_backend_provider",
                lambda: self.post(
                    f"/api/projects/{self.project_id}/analyze",
                    {
                        "intro": "Hell prison SLG. English localization. Concise UI and strict glossary.",
                        "asset_artifact_ids": [brief_artifact["id"]],
                        "target_language": "en",
                    },
                    timeout=240,
                ),
            )

            glossary_artifact = self.run_step("05_step3_upload_glossary", lambda: self.upload(self.project_id, fixtures["glossary"], "term_base"))
            self.run_step(
                "06_step3_preview_glossary",
                lambda: self.post(f"/api/projects/{self.project_id}/glossary/import-preview", {"artifact_id": glossary_artifact["id"], "language": "en"}),
            )
            self.run_step(
                "07_step3_import_glossary",
                lambda: self.post(f"/api/projects/{self.project_id}/glossary/import", {"artifact_id": glossary_artifact["id"], "language": "en"}),
            )
            self.run_step("08_glossary_export_xlsx", lambda: self.get(f"/api/projects/{self.project_id}/glossary/export?format=xlsx&language=en"))

            source_artifact = self.run_step("09_step4_upload_language_table", lambda: self.upload(self.project_id, fixtures["source"], "language_table"))
            readiness = self.run_step(
                "10_step4_translation_readiness",
                lambda: self.get(f"/api/artifacts/{source_artifact['id']}/translation-readiness?language=en&batch_size=90"),
            )
            if int(readiness.get("source_rows") or 0) != 2:
                raise RuntimeError(f"unexpected source rows: {readiness}")

            run = self.run_step(
                "11_create_translation_run",
                lambda: self.post(
                    "/api/runs",
                    {
                        "project_id": self.project_id,
                        "kind": "translation",
                        "language": "en",
                        "input_artifact_id": source_artifact["id"],
                        "term_artifact_id": glossary_artifact["id"],
                        "batch_size": 90,
                        "task_origin": "stability_check",
                        "task_code": "T",
                    },
                ),
            )
            self.run_step("12_start_translation_background", lambda: self.post(f"/api/runs/{run['id']}/translate/start", {"batch_size": 90}, timeout=60))
            translated_run = self.run_step("13_translation_reaches_terminal_state", lambda: self.poll_run(run["id"], timeout_seconds=240))
            progress = (translated_run.get("metadata") or {}).get("translation_progress") or {}
            if progress.get("completed_rows") != 2:
                raise RuntimeError(f"translation did not complete all rows: {progress}")

            translated_artifact = self.run_step("14_upload_ready_translated_workbook", lambda: self.upload(self.project_id, fixtures["translated"], "final_workbook"))
            qa_run = self.run_step(
                "15_create_direct_qa_run",
                lambda: self.post(
                    "/api/runs",
                    {
                        "project_id": self.project_id,
                        "kind": "qa",
                        "language": "en",
                        "input_artifact_id": translated_artifact["id"],
                        "term_artifact_id": glossary_artifact["id"],
                        "task_origin": "stability_check",
                        "task_code": "QA",
                    },
                ),
            )
            qa_result = self.run_step("16_run_direct_qa", lambda: self.post(f"/api/runs/{qa_run['id']}/qa", {}, timeout=180))
            qa_status = ((qa_result.get("run") or {}).get("status") or "")
            if qa_status not in {"passed", "failed"}:
                raise RuntimeError(f"QA did not finish: {qa_status}")

            self.run_step(
                "17_import_translation_archive",
                lambda: self.post(f"/api/projects/{self.project_id}/translations/import", {"artifact_id": translated_artifact["id"], "language": "en"}),
            )
            self.run_step("18_export_translation_archive_xlsx", lambda: self.get(f"/api/projects/{self.project_id}/translations/export?format=xlsx&language=en"))
            self.run_step("19_export_translation_archive_csv", lambda: self.get(f"/api/projects/{self.project_id}/translations/export?format=csv&language=en"))

            quick_run = self.run_step(
                "20_create_quick_translation_run",
                lambda: self.post(
                    "/api/runs",
                    {
                        "project_id": self.project_id,
                        "kind": "translation",
                        "language": "en",
                        "input_artifact_id": source_artifact["id"],
                        "term_artifact_id": glossary_artifact["id"],
                        "batch_size": 90,
                        "task_origin": "quick_task",
                        "task_code": "T",
                    },
                ),
            )
            self.run_step("21_start_quick_translation_background", lambda: self.post(f"/api/runs/{quick_run['id']}/translate/start", {"batch_size": 90}, timeout=60))
            quick_terminal = self.run_step("22_quick_translation_reaches_terminal_state", lambda: self.poll_run(quick_run["id"], timeout_seconds=240))
            quick_progress = (quick_terminal.get("metadata") or {}).get("translation_progress") or {}
            if quick_progress.get("completed_rows") != 2:
                raise RuntimeError(f"quick translation did not complete all rows: {quick_progress}")

            announcement_artifact = self.run_step("23_upload_announcement_source", lambda: self.upload(self.project_id, fixtures["announcement"], "asset"))
            announcement = self.run_step(
                "24_create_announcement_task",
                lambda: self.post(
                    f"/api/projects/{self.project_id}/announcement-tasks",
                    {"source_artifact_id": announcement_artifact["id"], "title": "Stable announcement", "languages": ["en"]},
                ),
            )
            task_id = announcement["id"]
            self.run_step(
                "25_announcement_inspect_constraints",
                lambda: self.post(f"/api/announcement-tasks/{task_id}/inspect-constraints", {"languages": ["en"], "language_table_artifact_ids": [translated_artifact["id"]]}),
            )
            self.run_step(
                "26_announcement_extract_terms",
                lambda: self.post(f"/api/announcement-tasks/{task_id}/extract-terms", {"languages": ["en"], "language_table_artifact_ids": [translated_artifact["id"]], "ai_supplement": False}, timeout=180),
            )
            self.run_step(
                "27_announcement_lookup_translations",
                lambda: self.post(f"/api/announcement-tasks/{task_id}/lookup-translations", {"languages": ["en"], "language_table_artifact_ids": [translated_artifact["id"]]}, timeout=180),
            )
            self.run_step(
                "28_announcement_prepare",
                lambda: self.post(f"/api/announcement-tasks/{task_id}/prepare", {"languages": ["en"]}, timeout=180),
            )

            return 0
        finally:
            report_path = self.fixture_dir / "stability-report.json"
            report_path.write_text(json.dumps({"base_url": self.base_url, "project_id": self.project_id, "results": self.results}, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"report={report_path}")
            if self.project_id and not self.keep_project:
                try:
                    self.delete(f"/api/projects/{self.project_id}")
                    print(f"deleted_project={self.project_id}")
                except Exception as exc:
                    print(f"delete_project_failed={self.project_id}: {exc}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Run no-Codex stability checks against a running Localization Workflow Studio.")
    parser.add_argument("--base-url", default="http://127.0.0.1:5174", help="Workbench URL. Use the front-end URL so proxy failures are tested too.")
    parser.add_argument("--keep-project", action="store_true", help="Keep the temporary project for manual inspection.")
    args = parser.parse_args()
    return StabilityCheck(args.base_url, keep_project=args.keep_project).execute()


if __name__ == "__main__":
    raise SystemExit(main())
